"""SECR database — SQLite persistence for generated SECR workbooks.

This is the only module that touches the database. It is called from app.py
after `create_secr_bytes()` / `update_secr_bytes()` succeed:

    record = secr_db.record_from_workbook(secr_bytes, action="create", ...)
    secr_id = secr_db.save_secr(record)

Design doc: docs/SECR_DATABASE_DESIGN.md
"""
from __future__ import annotations

import getpass
import io
import re
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

DB_PATH = Path(__file__).resolve().parent / "data" / "secr_database.db"
SCHEMA_VERSION = 1

_SCHEMA = """
CREATE TABLE IF NOT EXISTS secr (
    id                   INTEGER PRIMARY KEY,
    secr_number          TEXT NOT NULL,
    version              TEXT NOT NULL DEFAULT 'A',
    filename             TEXT,
    action               TEXT NOT NULL CHECK (action IN ('create','update')),
    parent_secr_id       INTEGER REFERENCES secr(id),
    model_year           TEXT,
    program              TEXT,
    phase                TEXT,
    harness_family       TEXT,
    phase_implemented    TEXT,
    pull_ahead           TEXT,
    change_type          TEXT,
    subject              TEXT,
    secr_author          TEXT,
    design_release_engineer TEXT,
    change_requested_by  TEXT,
    original_issue_date  TEXT,
    reissue_date         TEXT,
    dtcr_numbers         TEXT,
    bulletin_numbers     TEXT,
    ref_secr             TEXT,
    source_def_filename  TEXT,
    enriched             INTEGER NOT NULL DEFAULT 0,
    created_at           TEXT NOT NULL DEFAULT (datetime('now','localtime')),
    created_by           TEXT,
    UNIQUE (secr_number, version)
);

CREATE TABLE IF NOT EXISTS secr_affected_item (
    id        INTEGER PRIMARY KEY,
    secr_id   INTEGER NOT NULL REFERENCES secr(id) ON DELETE CASCADE,
    category  TEXT NOT NULL CHECK (category IN ('device','circuit','part_number')),
    action    TEXT NOT NULL CHECK (action IN ('ADD','CHG','DELETE')),
    item      TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS secr_dtcr (
    id                    INTEGER PRIMARY KEY,
    secr_id               INTEGER NOT NULL REFERENCES secr(id) ON DELETE CASCADE,
    dtcr_number           TEXT NOT NULL,
    device_transmittal    TEXT,
    device_control_number TEXT,
    reason_for_change     TEXT,
    status                TEXT,
    match_method          TEXT,
    matched_dtx_value     TEXT,
    cnum                  TEXT,
    harness_family        TEXT
);

CREATE TABLE IF NOT EXISTS secr_dtcr_circuit (
    secr_dtcr_id INTEGER NOT NULL REFERENCES secr_dtcr(id) ON DELETE CASCADE,
    circuit      TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS ix_secr_lookup ON secr (model_year, program, phase);
CREATE INDEX IF NOT EXISTS ix_item_lookup ON secr_affected_item (item, category);
CREATE INDEX IF NOT EXISTS ix_dtcr_lookup ON secr_dtcr (dtcr_number);
CREATE INDEX IF NOT EXISTS ix_dtcr_ckt   ON secr_dtcr_circuit (circuit);
"""

# Summary-sheet cell map (see design doc §3)
_AFFECTED_CELLS = [
    ("device", "ADD", "C20"),
    ("device", "CHG", "C21"),
    ("device", "DELETE", "C22"),
    ("circuit", "ADD", "C25"),
    ("circuit", "CHG", "C26"),
    ("circuit", "DELETE", "C27"),
    ("part_number", "ADD", "C30"),
    ("part_number", "CHG", "C31"),
    ("part_number", "DELETE", "C32"),
]


# ---------------------------------------------------------------------------
# Connection / schema
# ---------------------------------------------------------------------------

def get_conn(db_path: Optional[Path] = None) -> sqlite3.Connection:
    path = Path(db_path) if db_path else DB_PATH
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(path), timeout=15)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA busy_timeout = 15000")
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db(db_path: Optional[Path] = None) -> None:
    """Create/migrate the schema. Safe to call repeatedly."""
    with get_conn(db_path) as conn:
        current = conn.execute("PRAGMA user_version").fetchone()[0]
        if current < 1:
            conn.executescript(_SCHEMA)
            conn.execute(f"PRAGMA user_version = {SCHEMA_VERSION}")
        # future migrations: if current < 2: ... etc.


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _date_text(value: Any) -> str:
    """Normalize openpyxl date/datetime cells to YYYY-MM-DD."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _split_list(text: Any) -> List[str]:
    """Split a comma/semicolon/newline-separated cell into clean items."""
    raw = _text(text)
    if not raw:
        return []
    return [p.strip() for p in re.split(r"[,;\n]+", raw) if p.strip()]


def read_secr_number(secr_bytes: bytes) -> Optional[str]:
    """Read the SECR # (Summary I2) from workbook bytes. Used to resolve
    the baseline record when running Update SECR."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(secr_bytes), data_only=True)
        value = _text(wb["Summary"]["I2"].value)
        wb.close()
        return value or None
    except Exception:
        return None


def record_from_workbook(
    secr_bytes: bytes,
    *,
    action: str,
    source_def_filename: str = "",
    filename: str = "",
    change_type: str = "",
    enriched: bool = False,
    dtcr_mapping_df=None,
    parent_secr_number: Optional[str] = None,
) -> Dict[str, Any]:
    """Build a plain-dict SECR record from generated workbook bytes.

    The workbook is the source of truth (it already contains form inputs and
    any DTCR enrichment). `dtcr_mapping_df` is the DataFrame produced by
    `match_dtcr_to_harness_family()`; only rows matching the SECR's harness
    family (Summary C12) are stored, mirroring the enrichment filter.
    """
    if action not in ("create", "update"):
        raise ValueError("action must be 'create' or 'update'")

    wb = openpyxl.load_workbook(io.BytesIO(secr_bytes), data_only=True)
    if "Summary" not in wb.sheetnames:
        wb.close()
        raise ValueError("SECR workbook has no 'Summary' sheet")
    ws = wb["Summary"]

    record: Dict[str, Any] = {
        "secr_number": _text(ws["I2"].value),
        "version": _text(ws["I3"].value) or "A",
        "filename": filename,
        "action": action,
        "parent_secr_number": parent_secr_number,
        "model_year": _text(ws["C10"].value),
        "program": _text(ws["C11"].value),
        "phase": _text(ws["F10"].value),
        "harness_family": _text(ws["C12"].value),
        "phase_implemented": _text(ws["F11"].value),
        "pull_ahead": _text(ws["F12"].value),
        "change_type": change_type,
        "subject": _text(ws["C7"].value),
        "secr_author": _text(ws["I10"].value),
        "design_release_engineer": _text(ws["I11"].value),
        "change_requested_by": _text(ws["I12"].value),
        "original_issue_date": _date_text(ws["I4"].value),
        "reissue_date": _date_text(ws["I5"].value),
        "dtcr_numbers": _text(ws["C14"].value),
        "bulletin_numbers": _text(ws["G14"].value),
        "ref_secr": _text(ws["C15"].value),
        "source_def_filename": source_def_filename,
        "enriched": 1 if enriched else 0,
        "created_by": _safe_username(),
        "affected_items": [],
        "dtcrs": [],
    }
    if not record["secr_number"]:
        wb.close()
        raise ValueError("Could not read SECR # from Summary!I2")

    for category, act, cell in _AFFECTED_CELLS:
        for item in _split_list(ws[cell].value):
            record["affected_items"].append(
                {"category": category, "action": act, "item": item}
            )
    wb.close()

    if dtcr_mapping_df is not None and len(dtcr_mapping_df) > 0:
        family = record["harness_family"]
        df = dtcr_mapping_df
        if family and "Harness Family" in df.columns:
            df = df[df["Harness Family"].astype(str).str.strip() == family]
        for _, row in df.iterrows():
            record["dtcrs"].append(
                {
                    "dtcr_number": _text(row.get("DTCR#")),
                    "device_transmittal": _text(row.get("Device Transmittal")),
                    "device_control_number": _text(
                        row.get("Extracted Device Control Number")
                    ),
                    "reason_for_change": _text(row.get("Reason for change")),
                    "status": _text(row.get("Status")),
                    "match_method": _text(row.get("Match Method")),
                    "matched_dtx_value": _text(row.get("Matched DTx Value")),
                    "cnum": _text(row.get("CNUM")),
                    "harness_family": _text(row.get("Harness Family")),
                }
            )
    return record


def _safe_username() -> str:
    try:
        return getpass.getuser()
    except Exception:
        return ""


# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------

def save_secr(record: Dict[str, Any], db_path: Optional[Path] = None) -> int:
    """Insert (or replace, keyed on SECR # + version) a SECR record with its
    affected items and DTCR rows. Atomic. Returns the secr row id."""
    init_db(db_path)
    with get_conn(db_path) as conn:
        parent_id = None
        if record.get("parent_secr_number"):
            row = conn.execute(
                "SELECT id FROM secr WHERE secr_number = ? ORDER BY id DESC LIMIT 1",
                (record["parent_secr_number"],),
            ).fetchone()
            parent_id = row["id"] if row else None

        # Upsert: replace any existing record for the same SECR # + version.
        conn.execute(
            "DELETE FROM secr WHERE secr_number = ? AND version = ?",
            (record["secr_number"], record["version"]),
        )
        cur = conn.execute(
            """
            INSERT INTO secr (
                secr_number, version, filename, action, parent_secr_id,
                model_year, program, phase, harness_family, phase_implemented,
                pull_ahead, change_type, subject, secr_author,
                design_release_engineer, change_requested_by,
                original_issue_date, reissue_date, dtcr_numbers,
                bulletin_numbers, ref_secr, source_def_filename, enriched,
                created_by
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                record["secr_number"], record["version"], record.get("filename"),
                record["action"], parent_id,
                record.get("model_year"), record.get("program"),
                record.get("phase"), record.get("harness_family"),
                record.get("phase_implemented"), record.get("pull_ahead"),
                record.get("change_type"), record.get("subject"),
                record.get("secr_author"), record.get("design_release_engineer"),
                record.get("change_requested_by"),
                record.get("original_issue_date"), record.get("reissue_date"),
                record.get("dtcr_numbers"), record.get("bulletin_numbers"),
                record.get("ref_secr"), record.get("source_def_filename"),
                record.get("enriched", 0), record.get("created_by"),
            ),
        )
        secr_id = cur.lastrowid

        for it in record.get("affected_items", []):
            conn.execute(
                "INSERT INTO secr_affected_item (secr_id, category, action, item)"
                " VALUES (?,?,?,?)",
                (secr_id, it["category"], it["action"], it["item"]),
            )

        for d in record.get("dtcrs", []):
            cur = conn.execute(
                """
                INSERT INTO secr_dtcr (
                    secr_id, dtcr_number, device_transmittal,
                    device_control_number, reason_for_change, status,
                    match_method, matched_dtx_value, cnum, harness_family
                ) VALUES (?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    secr_id, d["dtcr_number"], d.get("device_transmittal"),
                    d.get("device_control_number"), d.get("reason_for_change"),
                    d.get("status"), d.get("match_method"),
                    d.get("matched_dtx_value"), d.get("cnum"),
                    d.get("harness_family"),
                ),
            )
            dtcr_id = cur.lastrowid
            for circuit in _split_list(d.get("cnum")):
                if circuit.lower() != "none":
                    conn.execute(
                        "INSERT INTO secr_dtcr_circuit (secr_dtcr_id, circuit)"
                        " VALUES (?,?)",
                        (dtcr_id, circuit),
                    )
        return secr_id


# ---------------------------------------------------------------------------
# Queries
# ---------------------------------------------------------------------------

def list_secrs(
    model_year: str = "",
    program: str = "",
    phase: str = "",
    harness_family: str = "",
    author: str = "",
    db_path: Optional[Path] = None,
) -> List[Dict[str, Any]]:
    """List SECR records, newest first, with optional exact-match filters."""
    init_db(db_path)
    clauses, params = [], []
    for col, val in (
        ("model_year", model_year), ("program", program), ("phase", phase),
        ("harness_family", harness_family), ("secr_author", author),
    ):
        if val:
            clauses.append(f"{col} = ?")
            params.append(val)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    with get_conn(db_path) as conn:
        rows = conn.execute(
            f"SELECT * FROM secr {where} ORDER BY id DESC", params
        ).fetchall()
    return [dict(r) for r in rows]


def get_secr(secr_id: int, db_path: Optional[Path] = None) -> Optional[Dict[str, Any]]:
    """Full record: secr row + affected items + DTCR rows."""
    with get_conn(db_path) as conn:
        row = conn.execute("SELECT * FROM secr WHERE id = ?", (secr_id,)).fetchone()
        if row is None:
            return None
        record = dict(row)
        record["affected_items"] = [
            dict(r)
            for r in conn.execute(
                "SELECT category, action, item FROM secr_affected_item"
                " WHERE secr_id = ? ORDER BY category, action, item",
                (secr_id,),
            )
        ]
        record["dtcrs"] = [
            dict(r)
            for r in conn.execute(
                "SELECT * FROM secr_dtcr WHERE secr_id = ? ORDER BY dtcr_number",
                (secr_id,),
            )
        ]
        return record


def find_by_dtcr(dtcr_number: str, db_path: Optional[Path] = None) -> List[Dict[str, Any]]:
    """All SECRs that include a given DTCR # (via enrichment rows or the raw
    C14 text)."""
    with get_conn(db_path) as conn:
        rows = conn.execute(
            """
            SELECT DISTINCT s.* FROM secr s
            LEFT JOIN secr_dtcr d ON d.secr_id = s.id
            WHERE d.dtcr_number = ? OR s.dtcr_numbers LIKE ?
            ORDER BY s.id DESC
            """,
            (str(dtcr_number).strip(), f"%{str(dtcr_number).strip()}%"),
        ).fetchall()
    return [dict(r) for r in rows]


def find_by_item(item: str, category: str = "", db_path: Optional[Path] = None) -> List[Dict[str, Any]]:
    """All SECRs whose affected items (device / circuit / part_number) or
    per-DTCR circuits include `item`."""
    params: List[Any] = [item.strip()]
    cat_clause = ""
    if category:
        cat_clause = "AND a.category = ?"
        params.append(category)
    params.append(item.strip())
    with get_conn(db_path) as conn:
        rows = conn.execute(
            f"""
            SELECT DISTINCT s.* FROM secr s
            LEFT JOIN secr_affected_item a ON a.secr_id = s.id
            LEFT JOIN secr_dtcr d ON d.secr_id = s.id
            LEFT JOIN secr_dtcr_circuit c ON c.secr_dtcr_id = d.id
            WHERE (a.item = ? {cat_clause}) OR c.circuit = ?
            ORDER BY s.id DESC
            """,
            params,
        ).fetchall()
    return [dict(r) for r in rows]


def get_revision_chain(secr_id: int, db_path: Optional[Path] = None) -> List[Dict[str, Any]]:
    """Walk parent links from a record back to its original SECR."""
    chain: List[Dict[str, Any]] = []
    with get_conn(db_path) as conn:
        current: Optional[int] = secr_id
        seen = set()
        while current is not None and current not in seen:
            seen.add(current)
            row = conn.execute("SELECT * FROM secr WHERE id = ?", (current,)).fetchone()
            if row is None:
                break
            chain.append(dict(row))
            current = row["parent_secr_id"]
    return chain


def next_sequence(
    model_year: str,
    program: str,
    phase: str,
    change_type: str,
    default: int = 1000,
    db_path: Optional[Path] = None,
) -> int:
    """Next free sequence for SECR numbers shaped {M|D}{MY2}{PROGRAM}{PHASE}_{seq}."""
    my_two = str(model_year).strip()[-2:]
    prefix = (
        ("D" if str(change_type).strip().lower().startswith("design") else "M")
        + my_two
        + str(program).replace(" ", "").upper()
        + str(phase).replace("_", "").replace(" ", "").upper()
        + "_"
    )
    init_db(db_path)
    with get_conn(db_path) as conn:
        rows = conn.execute(
            "SELECT secr_number FROM secr WHERE secr_number LIKE ?", (prefix + "%",)
        ).fetchall()
    best = default - 1
    for r in rows:
        m = re.search(r"_(\d+)$", r["secr_number"])
        if m:
            best = max(best, int(m.group(1)))
    return best + 1
