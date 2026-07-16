from __future__ import annotations

import inspect
import io
import re
import sys
import tempfile
import time
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

from dtx_compare_engine import generate_dtx_change_report, launch_preorder_generation_tool
from secr_engine import create_secr_bytes
from secr_enrichment_engine import (
    load_dtcr_report,
    load_dtcr_matching_report,
    load_dtx_circuits_report,
    load_generated_secr_workbook,
    match_dtcr_to_harness_family,
    get_secr_harness_family_from_c12,
    find_reason_for_change_cell,
    find_dtcr_number_label_cell,
    update_secr_reason_for_change,
    update_secr_dtcr_numbers,
    build_reason_for_change_for_secr,
    build_dtcr_numbers_for_secr,
    build_enrichment_summary,
    export_dtcr_mapping_styled,
    export_secr_enriched_output,
)
from wiring_harness_processor import (
    evaluate_expression_against_all_pns,
    generate_sales_code_expression,
    generate_expression_for_selected_pns,
    get_candidate_codes_from_option_df,
    get_selected_harness_pns,
    run_analysis,
    run_analysis_from_option_df,
    simplify_expression_for_display,
    validate_generated_expression,
)
from vbom_streamlit_engine import run_vbom_workflow
from feedback_system import FeedbackStore, render_feedback_widget
from metrics.storage import build_metrics_storage
from metrics.tracker import MetricsTracker
from metrics.workflow_metrics import (
    create_secr_counts,
    dtcr_matching_counts,
    dtx_compare_counts,
    dtx_preorder_counts,
    splice_counts,
    vbom_counts,
)


st.set_page_config(page_title="System Engineer Toolkit", layout="wide")

metrics_tracker = MetricsTracker(build_metrics_storage())

APP_DIR = Path(__file__).resolve().parent
LOGO_PATH = APP_DIR / "assets" / "versigent_logo_horizontal.jpg"
SPLICE_SAMPLE_INPUT_PATH = APP_DIR / "assets" / "downloads" / "Z913_example_input.xlsx"
DTCR_EXTENSION_ZIP_PATH = APP_DIR / "assets" / "downloads" / "ispeed-dtcr-downloader.zip"
TOOL_SCROLL_IDS = {
    "Splice Generation": "splice-generation-section",
    "DTx Compare Report": "dtx-compare-section",
    "DTCR Matching Report": "dtcr-matching-section",
    "Create SECR": "create-secr-section",
    "Update SECR": "update-secr-section",
    "VBOM Risk Matrix": "vbom-risk-section",
}
if LOGO_PATH.exists():
    st.image(str(LOGO_PATH), width=300)


def open_tool(tool_name: str) -> None:
    st.session_state["selected_tool"] = tool_name
    st.session_state["scroll_to_tool"] = tool_name
    st.rerun()


def render_tool_scroll_anchor(tool_name: str) -> None:
    anchor_id = TOOL_SCROLL_IDS[tool_name]
    st.markdown(f'<div id="{anchor_id}"></div>', unsafe_allow_html=True)
    if st.session_state.pop("scroll_to_tool", None) == tool_name:
        components.html(
            f"""
            <script>
            const anchor = window.parent.document.getElementById("{anchor_id}");
            if (anchor) {{
                anchor.scrollIntoView({{ behavior: "smooth", block: "start" }});
            }}
            </script>
            """,
            height=0,
        )


def _auto_enrich_secr_if_requested(
    secr_bytes: bytes,
    dtcr_matching_bytes: bytes,
    output_filename: str,
) -> tuple[bytes, dict, pd.DataFrame, pd.DataFrame, str]:
    dtcr_mapping_df = load_dtcr_matching_report(dtcr_matching_bytes)
    dtcr_mapping_df = dtcr_mapping_df[
        dtcr_mapping_df["Status"].astype(str).str.strip().isin(["Complete", "Draft"])
    ]

    secr_wb = load_generated_secr_workbook(secr_bytes)
    secr_harness_family = get_secr_harness_family_from_c12(secr_wb)
    if dtcr_mapping_df.empty:
        raise ValueError("DTCR Matching Report is empty after filtering.")
    if not secr_harness_family:
        raise ValueError("SECR cell C12 is empty or invalid.")

    reason_text = build_reason_for_change_for_secr(secr_harness_family, dtcr_mapping_df)
    summary_df = build_enrichment_summary(dtcr_mapping_df, secr_harness_family, reason_text)

    reason_cell_info = find_reason_for_change_cell(secr_wb)
    if reason_cell_info:
        _, cell_ref = reason_cell_info
        update_secr_reason_for_change(secr_wb, cell_ref, reason_text)

    dtcr_label_info = find_dtcr_number_label_cell(secr_wb)
    if dtcr_label_info:
        _, dtcr_label_ref = dtcr_label_info
        dtcr_numbers_text = build_dtcr_numbers_for_secr(secr_harness_family, dtcr_mapping_df)
        update_secr_dtcr_numbers(secr_wb, dtcr_label_ref, dtcr_numbers_text)

    enriched_bytes, export_meta = export_secr_enriched_output(
        secr_wb,
        dtcr_mapping_df,
        dtcr_mapping_df,
        summary_df,
        output_filename=output_filename,
    )
    return enriched_bytes, export_meta, dtcr_mapping_df, summary_df, secr_harness_family


def _build_secr_number_preview(
    model_year: str,
    program: str,
    phase: str,
    secr_type_label: str,
    sequence: int = 1000,
) -> str:
    """Build SECR number preview from form values and selected change type."""
    my_clean = str(model_year or "").strip()
    program_clean = str(program or "").strip().upper().replace(" ", "")
    phase_clean = str(phase or "").strip().upper().replace(" ", "")

    if not my_clean or len(my_clean) < 2:
        return ""

    type_prefix = "D" if secr_type_label == "Design Change" else "M"
    my_two = my_clean[-2:]
    return f"{type_prefix}{my_two}{program_clean}{phase_clean}_{sequence}"


def _extract_secr_number_inputs_from_def(def_bytes: bytes) -> tuple[str, str, str]:
    """Extract MY, Program(Vehicle Line), and Phase from DEF_DEF_Summary identifier.

    Expected snippet in workbook: "DEF_New (Identifier) := 2028 RU X1_A ..."
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(def_bytes), data_only=True, read_only=True)
    except Exception:
        return "", "", ""

    try:
        if "DEF_DEF_Summary" not in wb.sheetnames:
            return "", "", ""

        ws = wb["DEF_DEF_Summary"]
        identifier_text = ""

        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 120), min_col=1, max_col=min(ws.max_column or 1, 12), values_only=True):
            for value in row:
                if not value:
                    continue
                text = str(value)
                if "DEF_New" in text and "Identifier" in text:
                    identifier_text = text
                    break
            if identifier_text:
                break

        if not identifier_text:
            return "", "", ""

        # Match: 2028 RU X1_A
        match = re.search(r"(\d{4})\s+([A-Za-z0-9]+)\s+([A-Za-z0-9]+_[A-Za-z0-9]+)", identifier_text)
        if not match:
            return "", "", ""

        my = match.group(1)
        program = match.group(2).upper()
        phase = match.group(3).replace("_", "").upper()
        return my, program, phase
    finally:
        wb.close()

st.markdown(
    """
    <style>
        .hero {
            padding: 1.25rem 1.5rem;
            border-radius: 16px;
            border: 1px solid #d9e4ee;
            background: linear-gradient(135deg, #f3f8fc 0%, #eef6f2 100%);
            margin-bottom: 1.2rem;
        }
        .tool-card {
            border: 1px solid #d6e1ea;
            border-radius: 14px;
            padding: 1rem;
            background: #ffffff;
            min-height: 220px;
            box-shadow: 0 8px 16px rgba(26, 43, 60, 0.05);
        }
        .tool-title {
            font-size: 1.2rem;
            font-weight: 700;
            margin-bottom: 0.5rem;
            color: #14324a;
        }
        .tool-desc {
            color: #35526b;
            margin-bottom: 1rem;
        }
        .tool-badge {
            display: inline-block;
            padding: 0.2rem 0.55rem;
            border-radius: 999px;
            font-size: 0.8rem;
            font-weight: 600;
            background: #e8f4ff;
            color: #0b5ea8;
            margin-right: 0.35rem;
            margin-bottom: 0.45rem;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="hero">
        <h1 style="margin-bottom: 0.35rem; color: #10273a;">System Engineer Toolkit</h1>
        <p style="margin: 0; color: #2f4b62;">
            Select a workflow below to launch wiring splice generation, DTx report comparison, or SECR creation.
        </p>
    </div>
    """,
    unsafe_allow_html=True,
)

mode = st.radio(
    "Choose Tool",
    [
        "Home",
        "Splice Generation",
        "DTx Compare Report",
        "Create SECR",
        "Update SECR",
        "DTCR Matching Report",
        "VBOM Risk Matrix",
    ],
    horizontal=True,
    label_visibility="collapsed",
)

if mode == "Home":
    row1 = st.columns(2, gap="large")

    with row1[0]:
        st.markdown(
            """
            <div class="tool-card">
                <div class="tool-title">Splice Generation</div>
                <div class="tool-desc">
                    Build harness configurations, generated connections, print matrix, and interactive sales code validation.
                </div>
                <div class="tool-desc" style="margin-top: 0.5rem;">
                    Upload one Excel workbook that contains exactly two required sheets: <strong>Complexity</strong> and <strong>OptionPerCkt</strong>.
                    Use the sample input if you need a reference layout for column order, naming, and valid sales code formatting.
                </div>
                <span class="tool-badge">Complexity</span>
                <span class="tool-badge">OptionPerCkt</span>
                <span class="tool-badge">Output Excel</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("Open Splice Generation", key="go_splice", use_container_width=True):
            open_tool("Splice Generation")
        if SPLICE_SAMPLE_INPUT_PATH.exists():
            st.download_button(
                "Download Example Splice Input",
                data=SPLICE_SAMPLE_INPUT_PATH.read_bytes(),
                file_name="Z913_example_input.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_splice_example_home",
                use_container_width=True,
            )

    with row1[1]:
        st.markdown(
            """
            <div class="tool-card">
                <div class="tool-title">DTx Compare Report</div>
                <div class="tool-desc">
                    Compare OLD vs NEW DTx reports, review added/removed/modified CNUM and circuits, and generate the compare workbook.
                </div>
                <div style="margin: 0.6rem 0 0.25rem; font-weight: 700; color: #0b5ea8;">PreOrder Generation List</div>
                <div class="tool-desc" style="margin-top: 0.15rem;">
                    Run the PreOrder workbook directly from this workflow to create the generation list output.
                </div>
                <span class="tool-badge">OLD vs NEW</span>
                <span class="tool-badge">Change Log</span>
                <span class="tool-badge">Dashboard</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("Open DTx Compare", key="go_dtx", use_container_width=True):
            open_tool("DTx Compare Report")

    row2 = st.columns(2, gap="large")

    with row2[0]:
        st.markdown(
            """
            <div class="tool-card">
                <div class="tool-title">DTCR Matching Report + Create SECR</div>
                <div class="tool-desc">
                    Use Step 1 to generate the DTCR matching workbook from iSpeed DTCR data and DTx circuits data.
                </div>
                <div class="tool-desc">
                    Use Step 2 to enrich SECR using the Step 1 DTCR_Matching_Report workbook output.
                </div>
                <div class="tool-desc" style="margin-top: 0.5rem; font-weight: 700; color: #0b5ea8;">Step 1: DTCR Matching Report</div>
                <div class="tool-desc">
                    Match DTCR records against DTx circuits and harness family mappings to produce a styled matching report workbook.
                </div>
                <span class="tool-badge">DTCR Mapping</span>
                <span class="tool-badge">Harness Family</span>
                <span class="tool-badge">Styled Output</span>
                <div class="tool-desc" style="margin-top: 0.9rem; font-weight: 700; color: #0b5ea8;">Step 2: Create SECR</div>
                <div class="tool-desc">
                    Generate a SECR workbook and enrich it with the Step 1 DTCR_Matching_Report output.
                </div>
                <div class="tool-desc" style="margin-top: 0.9rem; font-weight: 700; color: #0b5ea8;">Step 3: Update SECR</div>
                <div class="tool-desc">
                    Update a new SECR from a DEF compare and previous SECR baseline.
                </div>
                <span class="tool-badge">DEF Compare</span>
                <span class="tool-badge">SECR Template</span>
                <span class="tool-badge">Output Excel</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        step1_btn, step2_btn, step3_btn = st.columns(3)
        with step1_btn:
            if st.button("Open DTCR Matching", key="go_dtcr", use_container_width=True):
                open_tool("DTCR Matching Report")
        with step2_btn:
            if st.button("Open Create SECR", key="go_secr", use_container_width=True):
                open_tool("Create SECR")
        with step3_btn:
            if st.button("Open Update SECR", key="go_update_secr", use_container_width=True):
                open_tool("Update SECR")

    with row2[1]:
        st.markdown(
            """
            <div class="tool-card">
                <div class="tool-title">VBOM Risk Matrix</div>
                <div class="tool-desc">
                    Upload your VBOM input files and generate the same workbook bundle used by the desktop VBOM workflow, including the master complexity workbook, VIN matrix, and harness selection output.
                </div>
                <span class="tool-badge">DoAll / BuildSpec</span>
                <span class="tool-badge">Harness Complexity</span>
                <span class="tool-badge">Workbook Bundle</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("Open VBOM Risk Matrix", key="go_vbom", use_container_width=True):
            open_tool("VBOM Risk Matrix")

    row3 = st.columns(1)

    with row3[0]:
        with st.container(border=True):
            st.markdown(
                """
                <div class="tool-title">iSpeed DTCR Downloader</div>
                <div class="tool-desc">
                    Download the Chrome extension package used to capture iSpeed DTCR search results, attachments, and a DTCR summary CSV in one run.
                </div>
                <span class="tool-badge">Chrome Extension</span>
                <span class="tool-badge">DTCR Attachments</span>
                <span class="tool-badge">CSV Summary</span>
                """,
                unsafe_allow_html=True,
            )
            if DTCR_EXTENSION_ZIP_PATH.exists():
                st.download_button(
                    "Download iSpeed DTCR Downloader",
                    data=DTCR_EXTENSION_ZIP_PATH.read_bytes(),
                    file_name="ispeed-dtcr-downloader.zip",
                    mime="application/zip",
                    key="download_ispeed_dtcr_extension",
                    use_container_width=True,
                )
            else:
                st.warning("Chrome extension package not found. Expected: assets/downloads/ispeed-dtcr-downloader.zip")

            with st.expander("Install and use the extension", expanded=False):
                st.markdown(
                    """

### Install

1. Download and unzip the extension.
2. Open `chrome://extensions` in Chrome.
3. Turn on **Developer mode**.
4. Click **Load unpacked**.
5. Select the unzipped `ispeed-dtcr-downloader` folder.
6. Pin the extension from Chrome's Extensions menu.

### What it does
The extension processes the current iSpeed DTCR search results. It skips deleted or canceled DTCRs, records each Reason for Change, downloads attachments with cleaned filenames, and creates `DTCR_Summary.csv`.

### How to use it

1. Sign in to iSpeed.
2. Select a Vehicle Program and Build Phase, then click **Search**.
3. With the results visible, click the extension icon.
4. Confirm the DTCR count.
5. Click **Choose folder** and select an empty destination folder.
6. Click **Start download**.
7. Keep both tabs open until the run finishes.

iSpeed can be slow. The extension waits for each detail page and the restored search results before continuing.
                    """
                )

if mode != "Home":
    st.session_state["selected_tool"] = mode

selected_tool = st.session_state.get("selected_tool", "Home")
feedback_store = FeedbackStore()
render_feedback_widget(workflow=selected_tool, area=selected_tool, store=feedback_store, key_prefix="main_app_feedback")

if selected_tool == "Splice Generation":
    render_tool_scroll_anchor("Splice Generation")
    st.title("Wiring Harness Splice Generator")
    st.caption("Generate harness print-ready direct connections, splices, configuration groups, and validation reports.")

    with st.expander("How To Prepare The Upload File", expanded=True):
        st.markdown(
            """
            Use one Excel workbook with these two required sheets:

            1. `Complexity`: first column must be the Harness PN, and every other column must be a sales code. Mark valid harness/code combinations with `X`.
            2. `OptionPerCkt`: must include the circuit/device rows with the required columns for `CNUM`, `Pin`, `Circuit`, and `Sales Code`.
            3. Keep the sales codes in the workbook exactly as engineering defines them. Do not split the data into separate files.
            4. If you need a reference, download the bundled example workbook and match your file structure to it before uploading.
            """
        )
        if SPLICE_SAMPLE_INPUT_PATH.exists():
            st.download_button(
                "Download Example Input Workbook",
                data=SPLICE_SAMPLE_INPUT_PATH.read_bytes(),
                file_name="Z913_example_input.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_splice_example_page",
            )

    # CAN Mode Configuration (must be selected before uploading)
    st.markdown("---")
    st.subheader("Splice Configuration Options")
    can_mode = st.checkbox("Apply CAN splice rules: maximum 3 ends per splice", value=False)
    if can_mode:
        st.info("CAN mode enabled: Each splice will be limited to a maximum of 3 endpoints. Additional splices and splice-to-splice connections will be created as needed for configurations with more than 3 endpoints.")

    prev_ui_can_mode = st.session_state.get("ui_can_mode")
    if prev_ui_can_mode is not None and prev_ui_can_mode != can_mode:
        st.session_state.pop("analysis_result", None)
        st.session_state.pop("uploaded_file_name", None)
    st.session_state["ui_can_mode"] = can_mode

    uploaded_file = st.file_uploader("Upload Excel file (Complexity + OptionPerCkt)", type=["xlsx", "xls"])

    if uploaded_file is None:
        st.info("Upload Input.xlsx (or equivalent) to begin analysis.")
        st.stop()

    st.session_state["can_mode"] = can_mode

    upload_signature = f"{uploaded_file.name}:{uploaded_file.size}:{int(can_mode)}"
    previous_signature = st.session_state.get("analysis_signature")
    should_recompute = ("analysis_result" not in st.session_state) or (previous_signature != upload_signature)

    if should_recompute:
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(uploaded_file.name).suffix) as temp_file:
            temp_file.write(uploaded_file.getbuffer())
            temp_path = temp_file.name
        st.session_state["analysis_temp_path"] = temp_path

        event_key = f"splice_generation:{upload_signature}"
        try:
            with metrics_tracker.track_workflow(
                "splice_generation",
                event_key=event_key,
                input_file_count=1,
            ) as tracked_run:
                run_analysis_sig = inspect.signature(run_analysis)
                if "can_mode" in run_analysis_sig.parameters:
                    result = run_analysis(temp_path, can_mode=can_mode)
                else:
                    result = run_analysis(temp_path)
                    if can_mode:
                        st.warning("CAN mode is not available in the loaded backend yet. Please reboot/redeploy the app.")

                st.session_state["analysis_result"] = result
                counts = splice_counts(result)
                tracked_run.record_counts(
                    rows_read=counts["rows_read"],
                    rows_processed=counts["rows_processed"],
                    circuits_processed=counts["circuits_processed"],
                    harness_variants_processed=counts["harness_variants_processed"],
                    output_file_count=1,
                )
                tracked_run.record_validation_results(
                    automatic_validation_errors=counts["automatic_validation_errors"],
                    automatic_validation_warnings=counts["automatic_validation_warnings"],
                    automatic_validation_failures=counts["automatic_validation_failures"],
                )
                tracked_run.complete(output_generated=bool(result.get("output_excel_bytes")), output_file_count=1)
        except Exception as exc:
            st.error(f"Analysis failed: {exc}")
            st.stop()

        st.session_state["analysis_signature"] = upload_signature
        st.session_state["uploaded_file_name"] = uploaded_file.name
        st.session_state["analysis_can_mode"] = can_mode

    temp_path = st.session_state.get("analysis_temp_path", "")
    if not temp_path:
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(uploaded_file.name).suffix) as temp_file:
            temp_file.write(uploaded_file.getbuffer())
            temp_path = temp_file.name
        st.session_state["analysis_temp_path"] = temp_path
    result = st.session_state["analysis_result"]

    st.subheader("Input Previews")
    left, right = st.columns(2)
    with left:
        st.markdown("**Complexity Matrix (normalized)**")
        st.dataframe(result["harness_code_map_df"], use_container_width=True)
    with right:
        st.markdown("**OptionPerCircuit (normalized)**")
        st.dataframe(result["option_df"], use_container_width=True)

    st.subheader("Generated Configurations")
    st.dataframe(result["configurations_df"], use_container_width=True)

    # Display CAN validation results if CAN mode is enabled
    if result.get("can_mode", False):
        st.markdown("---")
        if result.get("can_validation_passed", True):
            st.success(f"✓ {result.get('can_validation_message', 'CAN validation passed')}")
        else:
            st.error(f"✗ {result.get('can_validation_message', 'CAN validation failed')}")

    st.subheader("Generated Connections")
    conns_df = result["generated_connections_df"]
    configs_df = result["configurations_df"]

    config_lookup = {}
    for _, cfg in configs_df.iterrows():
        key = (cfg["Circuit Name"], cfg["Configuration ID"])
        config_lookup[key] = {
            "topology_type": cfg["Topology Type"],
            "target_harness_pns": cfg["Target Harness PNs"],
        }

    for (circuit, config_id), group in conns_df.groupby(["Circuit Name", "Configuration"], sort=False):
        cfg_details = config_lookup.get((circuit, config_id), {})
        topology = cfg_details.get("topology_type", "Unknown")
        target_pns = cfg_details.get("target_harness_pns", "")

        if config_id == conns_df[conns_df["Circuit Name"] == circuit]["Configuration"].iloc[0]:
            st.markdown(f"### Circuit {circuit}")

        st.markdown(f"**Configuration {config_id} | {topology}**")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"**Connections:** {len(group)}")
        with col2:
            st.markdown(f"**Topology:** {topology}")
        with col3:
            st.markdown(f"**Target PNs:** {target_pns}")

        st.dataframe(group, use_container_width=True)
        st.markdown("---")

    st.subheader("Harness Print Matrix")
    st.markdown("Engineering applicability matrix showing which connections apply to each Harness PN:")
    st.dataframe(result["harness_print_matrix_df"], use_container_width=True)

    st.subheader("Interactive Sales Code Generator")
    st.markdown("Select circuit/row from the second sheet, edit Sales Code text, and visualize PN applicability from first-sheet rules.")

    option_df = result["option_df"].copy()
    circuits = sorted(option_df["Circuit"].dropna().astype(str).str.strip().unique().tolist())

    if not circuits:
        st.info("No circuits available in the second sheet (OptionPerCkt).")
    else:
        selected_circuit = st.selectbox("Circuit", circuits, key="interactive_circuit_selector")
        circuit_rows = option_df[option_df["Circuit"].astype(str).str.strip() == selected_circuit].copy()

        row_labels = [
            f"{idx}: {row['CNUM']} | Pin {row['Pin']} | SalesCode={row['Sales Code']}"
            for idx, row in circuit_rows.iterrows()
        ]
        selected_row_label = st.selectbox("Row", row_labels, key="interactive_row_selector")
        selected_row_idx = int(selected_row_label.split(":", 1)[0])
        selected_row = option_df.loc[selected_row_idx]

        sales_code_input = st.text_input(
            "Sales Code",
            value=str(selected_row["Sales Code"]),
            key=f"interactive_sales_code_input_{selected_row_idx}",
        )

        harness_cols = sorted({k.split("__")[0] for k in result["harness_code_map"].keys()})

        matched_pns: list[str] = []
        expression_valid = True
        validation_message = ""
        if sales_code_input.strip():
            try:
                matched_pns = evaluate_expression_against_all_pns(sales_code_input.strip(), result["harness_code_map"])
            except Exception:
                expression_valid = False
                validation_message = "Combination not valid with available salescodes"

        visualize_row = {
            "Device ID": "",
            "Connector No": str(selected_row["CNUM"]),
            "Device Name": "Interactive_Row",
            "Pin": str(selected_row["Pin"]),
            "Circuit": selected_circuit,
            "Sales Code": sales_code_input.strip(),
        }
        for pn in harness_cols:
            visualize_row[pn] = pn in matched_pns

        st.markdown("**PN Applicability Grid**")
        edited_df = st.data_editor(
            pd.DataFrame([visualize_row]),
            column_config={pn: st.column_config.CheckboxColumn(pn) for pn in harness_cols},
            use_container_width=True,
            num_rows="fixed",
            key=f"interactive_editor_{selected_row_idx}",
        )

        if not expression_valid:
            st.error(validation_message)

        col_gen, col_apply = st.columns(2)
        with col_gen:
            if st.button("Generate Sales Code", key="btn_generate_sales_code"):
                selected_by_row = get_selected_harness_pns(edited_df)
                selected_pns = selected_by_row.get(0, [])
                candidate_codes = get_candidate_codes_from_option_df(
                    result["option_df"],
                    circuit_name=selected_circuit,
                )

                selected_set = {pn.strip() for pn in selected_pns if str(pn).strip()}
                target_harness_keys = [
                    hk for hk in result["harness_code_map"].keys()
                    if hk.split("__")[0] in selected_set
                ]

                expr = ""
                if target_harness_keys and candidate_codes:
                    expr = generate_sales_code_expression(
                        target_harnesses=target_harness_keys,
                        harness_code_map=result["harness_code_map"],
                        candidate_codes=candidate_codes,
                    )
                else:
                    expr = generate_expression_for_selected_pns(selected_pns, result["harness_code_map"])

                if not expr:
                    st.session_state["interactive_generated_expr"] = None
                    st.session_state["interactive_expr_valid"] = False
                    st.error("Combination not valid with available salescodes")
                else:
                    valid = validate_generated_expression(expr, selected_pns, result["harness_code_map"])
                    if not valid:
                        st.session_state["interactive_generated_expr"] = None
                        st.session_state["interactive_expr_valid"] = False
                        st.error("Combination not valid with available salescodes")
                    else:
                        display_expr = simplify_expression_for_display(expr)
                        st.session_state["interactive_generated_expr"] = display_expr
                        st.session_state["interactive_expr_valid"] = True
                        st.session_state["interactive_target_row"] = selected_row_idx
                        st.success(f"Generated Sales Code: {display_expr}")

        with col_apply:
            can_apply = (
                st.session_state.get("interactive_expr_valid", False)
                and st.session_state.get("interactive_target_row") == selected_row_idx
            )
            if st.button("Apply Sales Code to Row", disabled=not can_apply, key="btn_apply_sales_code"):
                generated_expr = st.session_state.get("interactive_generated_expr", "")
                if not generated_expr:
                    st.error("No valid generated sales code to apply.")
                else:
                    updated_option_df = result["option_df"].copy()
                    updated_option_df.loc[selected_row_idx, "Sales Code"] = generated_expr
                    can_mode_for_refresh = st.session_state.get("can_mode", False)
                    refresh_sig = inspect.signature(run_analysis_from_option_df)
                    if "can_mode" in refresh_sig.parameters:
                        refreshed = run_analysis_from_option_df(
                            temp_path,
                            updated_option_df,
                            can_mode=can_mode_for_refresh,
                        )
                    else:
                        refreshed = run_analysis_from_option_df(temp_path, updated_option_df)
                        if can_mode_for_refresh:
                            st.warning("CAN mode is not available in the loaded backend yet. Please reboot/redeploy the app.")
                    st.session_state["analysis_result"] = refreshed
                    st.success("Sales code applied. Configurations and validation refreshed.")
                    st.rerun()

    st.download_button(
        label="Download Output Excel",
        data=result["output_excel_bytes"],
        file_name="Wiring_Harness_Output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

elif selected_tool == "DTx Compare Report":
    render_tool_scroll_anchor("DTx Compare Report")
    st.title("DTx Compare Report")
    st.caption("Upload OLD and NEW DTx files to generate an engineering change workbook.")

    col_old, col_new = st.columns(2)
    with col_old:
        old_file = st.file_uploader("OLD DTx report", type=["xlsx", "xls", "xlsm"], key="dtx_old_file")
    with col_new:
        new_file = st.file_uploader("NEW DTx report", type=["xlsx", "xls", "xlsm"], key="dtx_new_file")

    if old_file is None or new_file is None:
        st.info("Upload both OLD and NEW DTx reports to continue.")
        st.stop()

    st.subheader("PreOrder Generation List")
    st.caption("Generate the PreOrder workbook directly for the selected DTx reports.")

    if st.button("Generate PreOrder Generation List", type="secondary"):
        try:
            event_key = f"dtx_preorder_generation:{time.time_ns()}"
            with metrics_tracker.track_workflow(
                "dtx_preorder_generation",
                event_key=event_key,
                input_file_count=2,
            ) as tracked_run:
                temp_dir = Path(tempfile.mkdtemp(prefix="dtx_preorder_", dir=tempfile.gettempdir()))
                old_temp_path = temp_dir / old_file.name
                new_temp_path = temp_dir / new_file.name
                old_temp_path.write_bytes(old_file.getvalue())
                new_temp_path.write_bytes(new_file.getvalue())

                with st.spinner("Generating PreOrder workbook..."):
                    preorder_result = launch_preorder_generation_tool(
                        old_file_path=old_temp_path,
                        new_file_path=new_temp_path,
                    )

                preorder_counts = dtx_preorder_counts(preorder_result)
                tracked_run.record_counts(
                    rows_read=preorder_counts["rows_read"],
                    rows_processed=preorder_counts["rows_processed"],
                        circuits_processed=preorder_counts["circuits_processed"],
                        harness_variants_processed=preorder_counts["harness_variants_processed"],
                    output_file_count=1,
                )
                tracked_run.record_validation_results(
                    automatic_validation_errors=preorder_counts["automatic_validation_errors"],
                    automatic_validation_warnings=preorder_counts["automatic_validation_warnings"],
                    automatic_validation_failures=preorder_counts["automatic_validation_failures"],
                )
                tracked_run.complete(
                    output_generated=bool(preorder_result.get("output_excel_bytes")),
                    output_file_count=1,
                )

            st.session_state["preorder_generation_result"] = preorder_result
            st.success("PreOrder workbook generated.")
        except Exception as exc:
            st.error(f"Unable to generate the PreOrder workbook: {exc}")

    preorder_result = st.session_state.get("preorder_generation_result")
    if preorder_result is not None:
        st.dataframe(preorder_result["summary_df"], use_container_width=True)
        st.download_button(
            label="Download PreOrder Workbook",
            data=preorder_result["output_excel_bytes"],
            file_name=preorder_result["output_file_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if st.button("Generate Compare Report", type="primary"):
        try:
            event_key = f"dtx_compare_report:{time.time_ns()}"
            with metrics_tracker.track_workflow(
                "dtx_compare_report",
                event_key=event_key,
                input_file_count=2,
            ) as tracked_run:
                with st.spinner("Comparing reports and building workbook..."):
                    dtx_result = generate_dtx_change_report(
                        old_file_bytes=old_file.getvalue(),
                        new_file_bytes=new_file.getvalue(),
                        old_file_name=old_file.name,
                        new_file_name=new_file.name,
                    )

                compare_counts = dtx_compare_counts(dtx_result)
                tracked_run.record_counts(
                    rows_read=compare_counts["rows_read"],
                    rows_processed=compare_counts["rows_processed"],
                    circuits_processed=compare_counts["circuits_processed"],
                    harness_variants_processed=compare_counts["harness_variants_processed"],
                    output_file_count=1,
                )
                tracked_run.record_validation_results(
                    automatic_validation_errors=compare_counts["automatic_validation_errors"],
                    automatic_validation_warnings=compare_counts["automatic_validation_warnings"],
                    automatic_validation_failures=compare_counts["automatic_validation_failures"],
                )
                tracked_run.complete(
                    output_generated=bool(dtx_result.get("output_excel_bytes")),
                    output_file_count=1,
                )

            st.session_state["dtx_compare_result"] = dtx_result
        except Exception as exc:
            st.error(f"DTx comparison failed: {exc}")

    dtx_result = st.session_state.get("dtx_compare_result")
    if dtx_result is None:
        st.stop()

    st.success("Comparison complete.")

    metric_cols = st.columns(5)
    metric_cols[0].metric("Added CNUMs", int(dtx_result["added_cnum_count"]))
    metric_cols[1].metric("Removed CNUMs", int(dtx_result["removed_cnum_count"]))
    metric_cols[2].metric("Added Circuits", int(dtx_result["added_circuit_count"]))
    metric_cols[3].metric("Removed Circuits", int(dtx_result["removed_circuit_count"]))
    metric_cols[4].metric("Modified Circuits", int(dtx_result["modified_circuit_count"]))

    st.subheader("Detected Layout")
    layout_left, layout_right = st.columns(2)
    old_layout = dtx_result["old_layout"]
    new_layout = dtx_result["new_layout"]
    layout_left.info(f"OLD: sheet '{old_layout.sheet_name}', header row {old_layout.header_row + 1}")
    layout_right.info(f"NEW: sheet '{new_layout.sheet_name}', header row {new_layout.header_row + 1}")

    st.subheader("Preview Tables")
    with st.expander("Added Circuits", expanded=False):
        st.dataframe(dtx_result["added_circuits_df"], use_container_width=True)
    with st.expander("Removed Circuits", expanded=False):
        st.dataframe(dtx_result["removed_circuits_df"], use_container_width=True)
    with st.expander("Modified Circuits", expanded=True):
        st.dataframe(dtx_result["modified_circuits_df"], use_container_width=True)
    with st.expander("CNUM Summary", expanded=False):
        st.dataframe(dtx_result["cnum_summary_df"], use_container_width=True)
    with st.expander("Field Change Frequency", expanded=False):
        st.dataframe(dtx_result["field_change_frequency_df"], use_container_width=True)

    st.download_button(
        label="Download DTx Compare Workbook",
        data=dtx_result["output_excel_bytes"],
        file_name=dtx_result["output_file_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

elif selected_tool == "DTCR Matching Report":
    render_tool_scroll_anchor("DTCR Matching Report")
    st.title("DTCR Matching Report")
    st.caption("Match DTCR entries to harness families and export a standalone workbook using either the DTCR Excel report or the DTCR_Summary.csv from the Chrome extension, plus the DTx circuits report.")

    dtcr_match_col1, dtcr_match_col2 = st.columns(2)
    with dtcr_match_col1:
        dtcr_match_file = st.file_uploader(
            "DTCR Report or DTCR_Summary.csv",
            type=["xlsx", "xls", "xlsm", "csv"],
            key="dtcr_match_file",
        )
    with dtcr_match_col2:
        dtx_match_file = st.file_uploader(
            "DTx Circuits Report (Excel)",
            type=["xlsx", "xls", "xlsm"],
            key="dtx_match_file",
        )

    if dtcr_match_file is not None and dtx_match_file is not None:
        if st.button(
            "Generate DTCR Matching Report",
            type="primary",
            key="generate_dtcr_matching_report",
        ):
            try:
                event_key = f"dtcr_matching_report:{time.time_ns()}"
                with metrics_tracker.track_workflow(
                    "dtcr_matching_report",
                    event_key=event_key,
                    input_file_count=2,
                ) as tracked_run:
                    with st.spinner("Building DTCR matching report..."):
                        dtcr_match_df = load_dtcr_report(dtcr_match_file.getvalue(), dtcr_match_file.name)
                        dtx_match_df = load_dtx_circuits_report(dtx_match_file.getvalue())
                        dtcr_mapping_df = match_dtcr_to_harness_family(dtcr_match_df, dtx_match_df)
                        dtcr_map_bytes = export_dtcr_mapping_styled(dtcr_mapping_df)

                    counts = dtcr_matching_counts(dtcr_match_df, dtx_match_df, dtcr_mapping_df)
                    tracked_run.record_counts(
                        rows_read=counts["rows_read"],
                        rows_processed=counts["rows_processed"],
                        circuits_processed=counts["circuits_processed"],
                        harness_variants_processed=counts["harness_variants_processed"],
                        output_file_count=1,
                    )
                    tracked_run.record_validation_results(
                        automatic_validation_errors=counts["automatic_validation_errors"],
                        automatic_validation_warnings=counts["automatic_validation_warnings"],
                        automatic_validation_failures=counts["automatic_validation_failures"],
                    )
                    tracked_run.complete(
                        output_generated=bool(dtcr_map_bytes),
                        output_file_count=1,
                    )

                st.session_state["dtcr_matching_report_bytes"] = dtcr_map_bytes
                st.session_state["dtcr_matching_report_df"] = dtcr_mapping_df
                st.session_state["dtcr_matching_report_name"] = "DTCR_Matching_Report.xlsx"
                st.success("DTCR matching report generated.")
            except Exception as exc:
                st.error(f"DTCR matching report failed: {exc}")

        dtcr_matching_report_bytes = st.session_state.get("dtcr_matching_report_bytes")
        if dtcr_matching_report_bytes is not None:
            st.download_button(
                label="Download DTCR Matching Report",
                data=dtcr_matching_report_bytes,
                file_name=st.session_state.get("dtcr_matching_report_name", "DTCR_Matching_Report.xlsx"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_dtcr_matching_report",
                use_container_width=True,
            )
            st.dataframe(
                st.session_state.get("dtcr_matching_report_df", pd.DataFrame()),
                use_container_width=True,
            )

elif selected_tool == "VBOM Risk Matrix":
    render_tool_scroll_anchor("VBOM Risk Matrix")
    st.title("VBOM Risk Matrix")
    st.caption(
        "Upload a DoAll or BuildSpec file and one or more harness complexity files to generate the VBOM workbook bundle used by the desktop workflow."
    )

    with st.form("vbom_streamlit_form"):
        my = st.text_input("Model Year (MY)", value="27")
        program = st.text_input("Program", value="RU")
        source_type = st.radio("Input source", ["DoAll", "BuildSpec"], horizontal=True)
        input_upload = st.file_uploader(
            "DoAll / BuildSpec file",
            type=["xlsx", "xls", "xlsm", "csv"],
            key="vbom_input_file",
        )
        complexity_uploads = st.file_uploader(
            "Harness Complexity files",
            type=["xlsx", "xls", "xlsm"],
            accept_multiple_files=True,
            key="vbom_complexity_files",
        )
        generate_clicked = st.form_submit_button("Generate VBOM Bundle", type="primary")

    if generate_clicked:
        if input_upload is None or not complexity_uploads:
            st.error("Please upload an input file and at least one harness complexity file.")
        else:
            try:
                event_key = f"vbom_risk_matrix:{time.time_ns()}"
                with metrics_tracker.track_workflow(
                    "vbom_risk_matrix",
                    event_key=event_key,
                    input_file_count=1 + len(complexity_uploads),
                ) as tracked_run:
                    with st.spinner("Generating VBOM outputs..."):
                        result = run_vbom_workflow(
                            my=my,
                            program=program,
                            source_type=source_type,
                            input_upload=input_upload,
                            complexity_uploads=complexity_uploads,
                            output_dir=Path(tempfile.gettempdir()) / "splice_vbom_outputs",
                        )

                    counts = vbom_counts(result)
                    tracked_run.record_counts(
                        rows_read=counts["rows_read"],
                        rows_processed=counts["rows_processed"],
                        circuits_processed=counts["circuits_processed"],
                        harness_variants_processed=counts["harness_variants_processed"],
                        output_file_count=4,
                    )
                    tracked_run.record_validation_results(
                        automatic_validation_errors=counts["automatic_validation_errors"],
                        automatic_validation_warnings=counts["automatic_validation_warnings"],
                        automatic_validation_failures=counts["automatic_validation_failures"],
                    )
                    tracked_run.complete(
                        output_generated=True,
                        output_file_count=4,
                    )

                st.session_state["vbom_result"] = result
                st.success("VBOM workbook bundle generated.")
            except Exception as exc:
                st.error(f"VBOM workflow failed: {exc}")

    vbom_result = st.session_state.get("vbom_result")
    if vbom_result is not None:
        st.subheader("Generated Files")
        output_paths = [
            ("Master complexity workbook", vbom_result.get("master_path")),
            ("VIN / SalesCode matrix", vbom_result.get("vin_matrix_path")),
            ("Harness selection workbook", vbom_result.get("selections_path")),
            ("Formatted template", vbom_result.get("formatted_template_path")),
        ]
        for label, path in output_paths:
            if path is not None and Path(path).exists():
                st.write(f"- {label}: {Path(path).name}")

        archive_bytes = io.BytesIO()
        with zipfile.ZipFile(archive_bytes, "w", zipfile.ZIP_DEFLATED) as archive:
            for _, path in output_paths:
                if path is not None and Path(path).exists():
                    archive.write(path, arcname=Path(path).name)
        archive_bytes.seek(0)
        st.download_button(
            label="Download VBOM Bundle",
            data=archive_bytes.getvalue(),
            file_name="VBOM_Risk_Matrix_Bundle.zip",
            mime="application/zip",
            key="dl_vbom_bundle",
            use_container_width=True,
        )

elif selected_tool == "Create SECR":
    render_tool_scroll_anchor("Create SECR")
    st.title("Create SECR")
    st.caption(
        "Create a SECR workbook from the DEF-to-DEF compare file. If you upload a DTCR_Matching_Report workbook, the SECR is enriched automatically after generation."
    )

    def_file = st.file_uploader(
        "DEF-to-DEF Compare file",
        type=["xlsx", "xls", "xlsm"],
        key="secr_def_file",
    )

    if def_file is None:
        st.info(
            "Upload the DEF-to-DEF compare file to continue.  \n"
            "Expected filename pattern: `2027_RU_X2_A_vs_2026_RU_X2_A_IP_DEF_DEF_Compare_...xlsx`"
        )
        st.stop()

    dtcr_matching_file = st.file_uploader(
        "Optional: DTCR_Matching_Report workbook",
        type=["xlsx", "xlsm", "xls"],
        key="create_secr_dtcr_matching_file",
        help="Upload the Step 1 DTCR_Matching_Report workbook now to auto-enrich the SECR after it is created.",
    )

    with st.form("secr_details_form"):
        st.subheader("SECR Details")

        # Defaults from DEF workbook identifier; fallback to filename when needed.
        extracted_my, extracted_program, extracted_phase = _extract_secr_number_inputs_from_def(
            def_file.getvalue()
        )

        def_stem_parts = Path(def_file.name).stem.split("_")
        fallback_my = def_stem_parts[0] if len(def_stem_parts) > 0 else ""
        fallback_program = def_stem_parts[1] if len(def_stem_parts) > 1 else ""
        fallback_phase = (
            f"{def_stem_parts[2]}{def_stem_parts[3]}".replace("_", "")
            if len(def_stem_parts) > 3 else ""
        )

        default_my = extracted_my or fallback_my
        default_program = extracted_program or fallback_program
        default_phase = extracted_phase or fallback_phase

        num_col1, num_col2, num_col3 = st.columns(3)
        with num_col1:
            secr_model_year = st.text_input(
                "MY",
                value=default_my,
                key="secr_number_my",
            )
        with num_col2:
            secr_program = st.text_input(
                "Program (Vehicle Line)",
                value=default_program,
                key="secr_number_program",
            )
        with num_col3:
            secr_phase = st.text_input(
                "Phase",
                value=default_phase,
                key="secr_number_phase",
            )

        col_a, col_b = st.columns(2)
        with col_a:
            subject = st.text_area(
                "Subject", height=100, key="secr_subject"
            )
            secr_author = st.text_input("SECR Author", key="secr_author")
            design_release_engineer = st.text_input(
                "Design Release Engineer", key="secr_dre"
            )
            change_requested_by = st.text_input(
                "Change Requested By", key="secr_crb"
            )
        with col_b:
            version = st.text_input("Version", value="A", key="secr_version")
            phase_implemented = st.text_input(
                "Phase Implemented", key="secr_phase_impl"
            )
            pull_ahead = st.selectbox(
                "Pull Ahead (Y/N)", options=["", "N", "Y"], key="secr_pull_ahead"
            )
            original_issue_date = st.text_input(
                "Original Issue Date (MM/DD/YYYY)", key="secr_orig_date"
            )
            reissue_date = st.text_input(
                "ReIssue Date (MM/DD/YYYY — leave blank if N/A)",
                key="secr_reissue_date",
            )
            secr_change_type = st.selectbox(
                "SECR # Type",
                options=["Miscellaneous", "Design Change"],
                key="secr_change_type",
            )

            secr_number_preview = _build_secr_number_preview(
                secr_model_year,
                secr_program,
                secr_phase,
                secr_change_type,
                sequence=1000,
            )
            st.text_input(
                "SECR # Preview",
                value=secr_number_preview,
                disabled=True,
                key="secr_number_preview",
                help="Auto-generated as M/D + MY + PROGRAM + PHASE + _1000.",
            )

        generate_clicked = st.form_submit_button("Generate SECR", type="primary")

    if generate_clicked:
        try:
            event_key = f"create_secr:{time.time_ns()}"
            with metrics_tracker.track_workflow(
                "create_secr",
                event_key=event_key,
                input_file_count=1 + (1 if dtcr_matching_file is not None else 0),
            ) as tracked_run:
                with st.spinner("Building SECR workbook..."):
                    secr_bytes, meta = create_secr_bytes(
                        def_bytes=def_file.getvalue(),
                        def_filename=def_file.name,
                        reason_for_change=subject,
                        secr_author=secr_author,
                        design_release_engineer=design_release_engineer,
                        change_requested_by=change_requested_by,
                        original_issue_date=original_issue_date,
                        reissue_date=reissue_date,
                        version=version,
                        phase_implemented=phase_implemented,
                        pull_ahead=pull_ahead,
                        secr_change_type=secr_change_type,
                        secr_model_year=secr_model_year,
                        secr_program=secr_program,
                        secr_phase=secr_phase,
                    )
                    base_meta = dict(meta)
                    st.session_state["secr_result_bytes"] = secr_bytes
                    st.session_state["secr_result_filename"] = base_meta["filename"]
                    st.session_state["secr_result_meta"] = base_meta
                    st.session_state["secr_result_enriched"] = False
                    enrichment_summary_df = None
                    if dtcr_matching_file is not None:
                        try:
                            secr_bytes, meta, dtcr_mapping_df, summary_df, secr_harness_family = _auto_enrich_secr_if_requested(
                                secr_bytes,
                                dtcr_matching_file.getvalue(),
                                base_meta.get("filename", "SECR_output.xlsx"),
                            )
                            enrichment_summary_df = summary_df
                            st.session_state["dtcr_matching_preview_df"] = dtcr_mapping_df
                            st.session_state["dtcr_matching_preview_summary_df"] = summary_df
                            st.session_state["dtcr_matching_preview_family"] = secr_harness_family
                            st.session_state["secr_result_filename"] = base_meta["filename"]
                            st.session_state["secr_result_meta"] = base_meta
                            st.session_state["secr_result_bytes"] = secr_bytes
                            st.session_state["secr_result_enriched"] = True
                        except Exception as enrich_exc:
                            st.warning(f"DTCR matching workbook was uploaded but could not be applied: {enrich_exc}")

                counts = create_secr_counts(
                    def_file_uploaded=def_file is not None,
                    enriched=bool(st.session_state.get("secr_result_enriched")),
                    enrichment_summary_df=enrichment_summary_df,
                )
                tracked_run.record_counts(
                    rows_read=counts["rows_read"],
                    rows_processed=counts["rows_processed"],
                    circuits_processed=counts["circuits_processed"],
                    harness_variants_processed=counts["harness_variants_processed"],
                    output_file_count=1,
                )
                tracked_run.record_validation_results(
                    automatic_validation_errors=counts["automatic_validation_errors"],
                    automatic_validation_warnings=counts["automatic_validation_warnings"],
                    automatic_validation_failures=counts["automatic_validation_failures"],
                )
                tracked_run.complete(output_generated=bool(st.session_state.get("secr_result_bytes")), output_file_count=1)
        except Exception as exc:
            st.error(f"SECR creation failed: {exc}")

    secr_result = st.session_state.get("secr_result_bytes")
    if secr_result is not None:
        meta = st.session_state.get("secr_result_meta", {})
        st.success("SECR workbook created successfully.")
        col1, col2, col3 = st.columns(3)
        col1.metric("M Code", meta.get("I2", ""))
        col2.metric("Vehicle Line", meta.get("C11", ""))
        col3.metric("Phase", meta.get("F10", ""))
        st.download_button(
            label="Download SECR Excel",
            data=secr_result,
            file_name=st.session_state.get("secr_result_filename", "SECR_output.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="secr_dl_btn",
        )
        if st.session_state.get("secr_result_enriched"):
            st.success("SECR was auto-enriched from the uploaded DTCR_Matching_Report workbook.")

elif selected_tool == "Update SECR":
    render_tool_scroll_anchor("Update SECR")
    st.title("Update SECR")
    st.caption(
        "Update SECR is now available in the desktop SECR Tracker app. "
        "This web workflow entry is enabled so the action is visible in navigation."
    )
    st.info(
        "Use the desktop flow for now: select new DEF-to-DEF compare + old SECR, "
        "then apply version/reissue updates and comment carry-over."
    )


