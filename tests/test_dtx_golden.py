"""Golden guard for the V1 DTx change-report contract (ADR-0005).

The raw ``.xlsx`` bytes are not stable across runs, for two reasons: xlsxwriter stamps a
creation timestamp into the file *metadata*, and the Dashboard embeds a live "Generated: …"
*cell*. So this pins a hash of the workbook content (sheet names + cell values) with that one
volatile cell normalized — everything else is what carries engineering meaning. Any drift in
the V1 output (a changed cell, column, or sheet) fails loudly; the generation timestamp does
not. Uses committed synthetic fixtures (made-up circuits), so it runs in CI with no proprietary
data. Regenerate fixtures with ``tests/fixtures/dtx_golden/build_fixtures.py``.
"""

from __future__ import annotations

import hashlib
import io
from pathlib import Path

from openpyxl import load_workbook

from splice.dtx_compare import generate_dtx_change_report

_FIXTURES = Path(__file__).resolve().parent / "fixtures" / "dtx_golden"

# Pinned from the synthetic fixtures. If the V1 change-report content legitimately changes,
# regenerate and update this hash **as a deliberate, reviewed step** (never silently).
GOLDEN_CONTENT_SHA256 = "ff769c458679d2635feec0af64b4cc4e294f6ec10d306f51bd948142cadbfd40"


def _norm(value) -> str:
    """Cell value as text, with the single volatile 'Generated: …' cell normalized out."""
    s = "" if value is None else str(value)
    return "Generated: <normalized>" if s.startswith("Generated:") else s


def _content_hash(xlsx_bytes: bytes) -> str:
    """Hash the meaningful content of a workbook: sheet names + all (normalized) cell values."""
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    h = hashlib.sha256()
    for ws in wb.worksheets:
        h.update(f"||SHEET:{ws.title}||".encode())
        for row in ws.iter_rows(values_only=True):
            h.update(("|".join(_norm(v) for v in row) + "\n").encode())
    return h.hexdigest()


def _run():
    old = (_FIXTURES / "old.xlsx").read_bytes()
    new = (_FIXTURES / "new.xlsx").read_bytes()
    return generate_dtx_change_report(
        old_file_bytes=old, new_file_bytes=new,
        old_file_name="old.xlsx", new_file_name="new.xlsx",
    )


def test_synthetic_change_set_is_as_designed():
    r = _run()
    assert (r["added_cnum_count"], r["removed_cnum_count"], r["modified_circuit_count"]) == (1, 1, 1)


def test_v1_change_report_content_is_byte_stable():
    assert _content_hash(_run()["output_excel_bytes"]) == GOLDEN_CONTENT_SHA256
