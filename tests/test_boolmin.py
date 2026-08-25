"""Boolean minimization of sales-code windows (display only).

Field report 2026-08-25: mechanically composed windows are unreadable —
the A960 route-gap window was 207 characters for a 22-character condition.
Every reduction must be provably equivalent under the engine's evaluator;
anything doubtful returns the original string.
"""

from __future__ import annotations

from itertools import chain, combinations

from splice.inline import salescode
from splice.inline.boolmin import minimize


def assert_equivalent(a: str, b: str) -> None:
    codes = sorted(salescode.codes_in(a) | salescode.codes_in(b))
    for r in range(len(codes) + 1):
        for on in combinations(codes, r):
            assert salescode.evaluate(a, set(on)) == salescode.evaluate(b, set(on)), \
                f"{a!r} vs {b!r} differ at {set(on)}"


class TestReductions:
    def test_r732_window(self):
        assert minimize("((CG3))&-((CG3&(CYC/CYF)))") == "CG3&-CYC&-CYF"

    def test_absorption(self):
        assert minimize("(A)/(A&B)") == "A"

    def test_common_factor_grouping(self):
        # (-X&A)/(-X&B)/(-X&C)  ->  -X&(A/B/C)
        short = minimize("(-X&A)/(-X&B)/(-X&C)")
        assert short == "-X&(A/B/C)"

    def test_a960_field_window(self):
        # the 207-character route-gap window from the 2026-08-25 field report;
        # note the minimal form PROVES the window includes the XZ3-without-amp
        # branch the manual reduction (-XZ2&-XZ3&(RHH/RTC/RDU)) missed
        raw = ("(((XZ2&(RHH/RTC/RDU))/(XZ2&-RHH&-RDU))/(((RHH/RTC/RDU)&-XZ2&-XZ3)"
               "/((XZ2/XZ3)&(-RHH&-RDU))/(XZ2/XZ3&(/RHH/RTC/RDU)))/"
               "(((RHH/RTC/RDU)&-XZ2&-XZ3)/(XZ2/XZ3&(RHH/RTC/RDU))))"
               "&-((XZ2&(RHH/RTC/RDU))/(XZ2&-RHH&-RDU))")
        short = minimize(raw)
        assert short == "-XZ2&(RDU/RHH/RTC/XZ3)"
        assert_equivalent(raw, short)

    def test_tolerates_the_leading_slash_data_typo(self):
        # (/A/B) evaluates in the engine as A/B (empty operand is false)
        short = minimize("(X&(/A/B))/(X&A)")
        assert_equivalent("(X&(/A/B))/(X&A)", short)
        assert len(short) <= len("(X&(/A/B))/(X&A)")


class TestDontCares:
    """Reduction constrained to the buildable configurations of the
    complexity tables (field question 2026-08-25: XZ2 and XZ3 share the same
    applicability in Body Left, so windows must not display phantom
    XZ3-without-XZ2 branches)."""

    def _bl(self):
        from splice.inline.model import Build, Harness
        return Harness(name="BODY LEFT", def_id="1", builds=[
            Build("B1", codes=frozenset({"XZ2", "XZ3", "RHH"})),
            Build("B2", codes=frozenset({"XZ2", "XZ3", "RTC"})),
            Build("B3", codes=frozenset({"XZ2", "XZ3", "RDU"})),
            Build("B4", codes=frozenset({"XZ2", "XZ3"})),
            Build("B5", codes=frozenset({"RHH"})),
            Build("B6", codes=frozenset({"RTC"})),
            Build("B7", codes=frozenset({"RDU"})),
            Build("B8", codes=frozenset()),
        ], complexity_codes={"XZ2", "XZ3", "RHH", "RTC", "RDU"})

    def test_co_occurring_codes_collapse_the_phantom_branch(self):
        from splice.inline.boolmin import care_configurations
        raw = ("(((XZ2&(RHH/RTC/RDU))/(XZ2&-RHH&-RDU))/"
               "(((RHH/RTC/RDU)&-XZ2&-XZ3)/((XZ2/XZ3)&(-RHH&-RDU))/"
               "(XZ2/XZ3&(/RHH/RTC/RDU)))/"
               "(((RHH/RTC/RDU)&-XZ2&-XZ3)/(XZ2/XZ3&(RHH/RTC/RDU))))"
               "&-((XZ2&(RHH/RTC/RDU))/(XZ2&-RHH&-RDU))")
        short = minimize(raw, care_configurations(self._bl()))
        assert short == "-XZ2&(RDU/RHH/RTC)"

    def test_constrained_form_selects_the_same_builds(self):
        from splice.inline.boolmin import care_configurations
        from splice.inline.health import builds_where
        bl = self._bl()
        raw = "((XZ2/XZ3)&(-RHH&-RDU))&-((XZ2)&(-RHH&-RDU))"
        short = minimize(raw, care_configurations(bl))
        raw_builds = {b.part_number for b in builds_where(bl, raw)}
        short_builds = {b.part_number for b in builds_where(bl, short)}
        assert raw_builds == short_builds

    def test_constant_on_buildables_keeps_the_raw_evidence(self):
        from splice.inline.boolmin import care_configurations
        bl = self._bl()
        # XZ3&-XZ2 never builds: the window is false on every buildable
        # configuration — a constant display would hide the evidence
        raw = "XZ3&-XZ2&RHH"
        assert minimize(raw, care_configurations(bl)) == raw


class TestSafety:
    def test_result_is_always_equivalent(self):
        cases = [
            "((501))&-((LEQ/LEM))",
            "((NHS&-CAJ)/(NHS&CAJ))&-((NHS&-CAJ&(CJZ/JWG/JPZ))/(NHS&CAJ))",
            "(((XZ3))/((XZ2))/((XAC))/((-XAC&-XZ3&-XZ2)/(XAC/XZ3/XZ2/RFX)))&-((XZ2))",
            "A&-A/B",  # OR binds tighter: A&(-A|B) = A&B
        ]
        for raw in cases:
            assert_equivalent(raw, minimize(raw))

    def test_never_longer_than_the_original(self):
        for raw in ["A", "A&B", "(A/B)&-C", "-X", "((CG3))&-((CG3&(CYC/CYF)))"]:
            assert len(minimize(raw)) <= len(raw)

    def test_too_many_variables_passes_through(self):
        raw = "/".join(f"V{i}" for i in range(20))
        assert minimize(raw) == raw

    def test_tautology_and_contradiction_stay_verbose(self):
        assert minimize("A/-A") == "A/-A"
        assert minimize("A&-A") == "A&-A"

    def test_empty_passthrough(self):
        assert minimize("") == ""


class TestHealthIntegration:
    def test_findings_carry_short_form_and_stable_fingerprint(self):
        from tests.test_inline_health import _analyze, LEFT_BUILDS, RIGHT_BUILDS
        result = _analyze(
            [{"circuit": "R732", "cnum": "X10A", "cav": "8", "sc": "CG3"}],
            [{"circuit": "R732", "cnum": "Y10A", "cav": "8",
              "sc": "CG3&(CYC/CYF)"}],
            LEFT_BUILDS, RIGHT_BUILDS,
        )
        f = next(x for x in result.findings if x.kind == "one_sided_window")
        # complexity-constrained: CG3 is in every build and CYC in none, so
        # both literals are vacuous on the buildable set — only -CYF remains
        assert f.window_short == "-CYF"
        assert f.window_short in f.detail
        assert f.window == "((CG3))&-((CG3&(CYC/CYF)))"  # raw preserved
        # the fingerprint hashes the RAW window — dispositions survive
        import hashlib
        import re as _re
        basis = "|".join([f.kind, f.inline, f.cavity, f.circuit,
                          _re.sub(r"\s+", "", f.window)])
        assert f.fingerprint == hashlib.sha1(basis.encode()).hexdigest()[:16]

    def test_report_shows_the_short_form(self):
        import io
        from openpyxl import load_workbook
        from splice.inline.health import render_report
        from tests.test_inline_health import _analyze, LEFT_BUILDS, RIGHT_BUILDS
        result = _analyze(
            [{"circuit": "R732", "cnum": "X10A", "cav": "8", "sc": "CG3"}],
            [{"circuit": "R732", "cnum": "Y10A", "cav": "8",
              "sc": "CG3&(CYC/CYF)"}],
            LEFT_BUILDS, RIGHT_BUILDS,
        )
        wb = load_workbook(io.BytesIO(render_report(result, {"dispositions": {}})))
        rows = list(wb["Findings"].iter_rows(values_only=True))
        hdr = rows[0]
        windows = [str(r[hdr.index("Window (sales)")]) for r in rows[1:]
                   if r[hdr.index("Kind")] == "one_sided_window"]
        assert "-CYF" in windows
