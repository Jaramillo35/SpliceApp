"""An invented programme at the scale of a real export, generated not authored.

The chart engine's performance work was measured against a customer DTx. That
is the wrong dependency for a repository: the numbers could not be reproduced
by anyone without the file, and the file has no business being the thing CI
and a benchmark script lean on.

So this fabricates a programme (2031 ZR) big enough to exercise the same code
paths at the same order of magnitude, from a seed. Nothing here is derived
from customer data — the families, codes, part numbers and circuit names are
invented, and the *shapes* it reproduces are the ones already documented in
this repository:

* a family count and row count in the range a full vehicle export reaches,
* roughly a third of DTx rows being No Connect (``splice.dtxcircuits.chart``
  records 29% for a real export),
* inlines left blank while devices carry the condition (the same module
  records 1,924 of 2,954 blank cells sitting on inlines),
* circuits reaching three or more cavities in one harness, so splices are
  planned rather than skipped,
* circuits crossing a harness boundary through an X/Y inline pair, so the
  cross-harness mate lookup has something to find.

``programme()`` is deterministic: the same seed and size give the same
workbooks, byte for byte, so a benchmark is comparable between runs.

Sizes. ``programme()`` defaults to 12 families (~1,400 rows), which is what a
test can afford. ``FULL_SCALE`` is the 47-family size the benchmark uses, and
what it produces:

    5,450 DTx rows · 1,444 No Connect (26%) · 47 families · 376 part numbers
    47 charts · 5,790 chart rows · 533 splices · 534 mated ends

Set beside the export the original measurement used (5,412 DTx rows, 47
families, 4,214 chart rows), it is the same size and somewhat heavier work —
which is what a benchmark should be. It is deliberately NOT a reconstruction
of that export: the resemblance is in the dimensions, nowhere else.
"""

from __future__ import annotations

import io
import random
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Sequence

from openpyxl import Workbook

YEAR, VEHICLE, PHASE = "2031", "ZR", "V2_A"
PROGRAM = f"{YEAR}{VEHICLE}"
TAG = f"{YEAR[-2:]}{VEHICLE}"

#: The size of a full vehicle export, for the benchmark. A test uses less.
FULL_SCALE = 47

#: Invented sales codes. Three letters, Z-prefixed so they cannot be mistaken
#: for a real one at a glance.
CODES: List[str] = [
    f"Z{group}{n}" for group in "ABCDEF" for n in range(1, 7)
]

#: Family names built from real harness *vocabulary* (an IP is an IP) but
#: composed here, so the list matches no particular vehicle.
_ZONES = ["IP", "DASH", "BODY", "CONSOLE", "DOOR", "SEAT", "LIFTGATE",
          "HVAC", "FASCIA", "JUMPER", "HEADLINER", "POWERTRAIN",
          "BATTERY", "TRAILER", "WIRE_TRACK", "MIRROR"]
_SIDES = ["", "_LEFT", "_RIGHT", "_FRONT", "_REAR", "_2ND_ROW"]


def family_names(count: int) -> List[str]:
    """``count`` distinct family names, always in the same order."""
    out: List[str] = []
    for side in _SIDES:
        for zone in _ZONES:
            name = f"{zone}{side}"
            if name not in out:
                out.append(name)
            if len(out) == count:
                return out
    raise ValueError(f"only {len(out)} family names available, wanted {count}")


@dataclass(frozen=True)
class Build:
    pn: str
    codes: frozenset


@dataclass
class Row:
    """One DTx circuit row, in the order the report writes them."""

    family: str
    def_id: str
    cnum: str
    pin: str
    circuit: str
    function: str
    condition: str


@dataclass
class Programme:
    families: List[str]
    def_ids: Dict[str, str]
    tracked: Dict[str, List[str]]
    builds: Dict[str, List[Build]]
    rows: List[Row] = field(default_factory=list)

    # ------------------------------------------------------------ counts
    @property
    def no_connect_rows(self) -> int:
        return sum(1 for r in self.rows if r.circuit == "N0")

    @property
    def part_numbers(self) -> int:
        return sum(len(b) for b in self.builds.values())

    # ----------------------------------------------------------- writing
    def dtx_bytes(self) -> bytes:
        """A Detailed DTx Circuits Report: title block on 1-4, header on 6."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Detailed DTx Circuits Report"
        ws.append(["Detailed DTx Circuits Report"])
        ws.append([f"Vehicle Program - {PROGRAM}  "])
        ws.append([f"Build Phase - {PHASE}"])
        ws.append(["Report Date: Mar-04-2031 09:00"])
        ws.append([])
        ws.append(["Device Control Number", "Device Name", "Suffix", "CNUM",
                   "Number of Cavities", "Connector PN", "Harness Family",
                   "Pin Number", "Circuit Name", "Circuit Suffix",
                   "Circuit Function", "Color", "Terminal",
                   "Connector FCA part number", "Wire Gauge", "Wire Type",
                   "Sales Code"])
        for r in self.rows:
            ws.append(["D" + r.def_id, f"{r.family}_MODULE", "A", r.cnum, 12,
                       "8" + r.def_id[1:] + "0", r.family, r.pin, r.circuit,
                       "", r.function, "BN/GN", "T-1", "FCA" + r.def_id,
                       "0.35", "TXL", r.condition or None])
        return _to_bytes(wb)

    def complexity_bytes(self, family: str) -> bytes:
        """One individual harness complexity file (Complexity + Harness PN)."""
        codes = self.tracked[family]
        wb = Workbook()
        ws = wb.active
        ws.title = "Complexity"
        ws.append([f"ID={self.def_ids[family]}", *codes])
        for build in self.builds[family]:
            ws.append([build.pn,
                       *["X" if c in build.codes else "" for c in codes]])

        hp = wb.create_sheet("Harness PN")
        hp.append(["Previous P/N", "New P/N", "Symbol"])
        for build in self.builds[family]:
            hp.append(["", build.pn, ""])
        hp.append([])
        for label, value in (("year:", YEAR), ("vehicle:", VEHICLE),
                             ("phase:", PHASE), ("harness:", family),
                             ("id:", self.def_ids[family])):
            hp.append([label, value])
        return _to_bytes(wb)

    def complexity_name(self, family: str) -> str:
        return (f"2.- Harness_Complexity_{TAG}_{family}_"
                f"{self.def_ids[family]}_03-04-2031.xlsx")

    @property
    def dtx_name(self) -> str:
        return f"DetailedDTxCircuitsReport_{TAG}_{PHASE}.xlsx"

    def write(self, out: Path) -> Dict[str, object]:
        """Put the DTx and every complexity file on disk, as the page takes them."""
        out = Path(out)
        out.mkdir(parents=True, exist_ok=True)
        dtx = out / self.dtx_name
        dtx.write_bytes(self.dtx_bytes())
        files = []
        for family in self.families:
            path = out / self.complexity_name(family)
            path.write_bytes(self.complexity_bytes(family))
            files.append(path)
        return {"dtx": dtx, "complexities": files}


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _expression(rng: random.Random, codes: Sequence[str]) -> str:
    """A condition in the engine's grammar, in the proportions a DTx shows.

    Blank dominates, a single code is next, and the compound forms are the
    tail — including a bare universal code, which the conventions module
    reads as unconditional only when it stands alone.
    """
    roll = rng.random()
    if roll < 0.45 or not codes:
        return ""
    if roll < 0.78:
        return rng.choice(codes)
    if roll < 0.86:
        return f"{rng.choice(codes)}/{rng.choice(codes)}"
    if roll < 0.92:
        return f"{rng.choice(codes)}&{rng.choice(codes)}"
    if roll < 0.96:
        return f"-{rng.choice(codes)}"
    return "501"


def programme(families: int = 12, seed: int = 2031,
              rows_per_family: int = 115) -> Programme:
    """An invented programme of ``families`` harnesses, always the same one."""
    rng = random.Random(seed)
    names = family_names(families)
    def_ids = {f: str(72000 + i) for i, f in enumerate(names)}

    tracked: Dict[str, List[str]] = {}
    builds: Dict[str, List[Build]] = {}
    for i, family in enumerate(names):
        # Families are not the same size. A real programme has a handful of
        # big ones — an IP complexity in a 2028-era export carries 33 part
        # numbers — against a tail of two- and three-build harnesses, and the
        # planning cost is driven almost entirely by the big ones. A fixture
        # of uniformly small families would benchmark the wrong thing.
        big = i % 8 == 0
        span = (12, 24) if big else (4, 10)
        owned = sorted(rng.sample(CODES, rng.randint(*span)))
        tracked[family] = owned
        made: List[Build] = []
        for k in range(rng.randint(18, 36) if big else rng.randint(2, 8)):
            carried = frozenset(
                rng.sample(owned, rng.randint(1, max(1, len(owned) - 1))))
            made.append(Build(f"98{i:02d}{k:02d}AA", carried))
        builds[family] = made

    prog = Programme(families=names, def_ids=def_ids, tracked=tracked,
                     builds=builds)

    circuit_no = 0
    inline_no = 0
    # counted as we go: re-scanning prog.rows per iteration is quadratic, and
    # at full scale that costs more than everything the fixture is built to
    # measure
    counts: Dict[str, int] = {f: 0 for f in names}

    def emit(row: Row) -> None:
        prog.rows.append(row)
        counts[row.family] += 1

    for i, family in enumerate(names):
        owned = tracked[family]
        device_no = 0
        while counts[family] < rows_per_family:
            roll = rng.random()

            # A cavity wired to nothing. Roughly a third of a real export's
            # rows, and none of them are circuits — the chart drops them, so
            # a fixture without them would not exercise that at all.
            if roll < 0.52:
                device_no += 1
                emit(Row(family, def_ids[family], f"D{2000 + device_no}A",
                         str(rng.randint(1, 12)), "N0", "No Connect", ""))
                continue

            circuit_no += 1
            circuit = f"ZK{circuit_no:05d}"
            condition = _expression(rng, owned)

            # How many cavities this circuit reaches inside this harness.
            # Three or more is what forces a splice; most circuits are a
            # plain two-end wire, as they are in a real export.
            ends = rng.choice([2] * 8 + [3, 3, 4])
            for _ in range(ends):
                device_no += 1
                emit(Row(family, def_ids[family], f"D{2000 + device_no}A",
                         str(rng.randint(1, 12)), circuit,
                         f"{family} - LOAD {device_no}", condition))

            # Every fifth circuit leaves the harness through an inline pair.
            # The inline halves are left blank: the DTx states applicability
            # at devices, and a blank on a joint is silence, not "always".
            if names[(i + 1) % len(names)] != family and rng.random() < 0.2:
                inline_no += 1
                neighbour = names[(i + 1) % len(names)]
                stem = f"{300 + inline_no}A"
                emit(Row(family, def_ids[family], f"X{stem}", "1", circuit,
                         "INLINE", ""))
                emit(Row(neighbour, def_ids[neighbour], f"Y{stem}", "1",
                         circuit, "INLINE", ""))
                emit(Row(neighbour, def_ids[neighbour],
                         f"D{9000 + inline_no}A", "1", circuit,
                         f"{neighbour} - LOAD", condition))

    return prog
