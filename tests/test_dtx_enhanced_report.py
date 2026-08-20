"""Enhanced DTx Compare workbook — sheets, Status column, dashboard, DTCR gate."""

from __future__ import annotations

import io

import pandas as pd
import pytest
from openpyxl import load_workbook

from splice.dtx_compare.enhanced_report import (
    DTCRRequiredError,
    build_yellow_connectors_df,
    generate_enhanced_dtx_report,
)


def test_yellow_connectors_from_invalid_connector_pn():
    new_rows = pd.DataFrame({
        "CNUM": ["D1", "D1", "D2", "D3"],
        "Connector PN": ["Invalid_Contact_Connector_Enginner | 5", "805-1", "Invalid_Contact_Connector_Enginner", "770-2"],
        "Device Name": ["Mod_A", "Mod_A", "Ant_B", "Sw_C"],
        "Harness Family": ["IP", "IP", "DASH", "IP"],
    })
    df = build_yellow_connectors_df(new_rows)
    assert list(df["CNUM"]) == ["D2", "D1"]        # one row per CNUM, sorted by family then CNUM
    assert "Connector PN" in df.columns and df.iloc[0]["Harness Family"] == "DASH"


def test_dtcr_report_is_required():
    with pytest.raises(DTCRRequiredError):
        generate_enhanced_dtx_report(b"x", b"y", "o.xls", "n.xls", None)
    with pytest.raises(DTCRRequiredError):
        generate_enhanced_dtx_report(b"x", b"y", "o.xls", "n.xls", pd.DataFrame())


@pytest.fixture(scope="module")
def _real(request):
    from pathlib import Path
    d = Path("/Users/martinjaramillo/Downloads/Development/data/Validatehere")
    files = {
        "old": d / "28RU_X1_DetailedDTxCircuitsReport_4_22_26 (31) 3.xls",
        "new": d / "DetailedDTxCircuitsReport (28_RU_X2) (2) 1.xls",
        "dtcr": d / "DTCRReport (1).xls",
    }
    if not all(p.exists() for p in files.values()):
        pytest.skip("real DTx/DTCR sample files absent")
    from splice.dtx_compare.engine import load_dtcr_report
    dtcr = load_dtcr_report(files["dtcr"].read_bytes(), files["dtcr"].name)
    return generate_enhanced_dtx_report(
        files["old"].read_bytes(), files["new"].read_bytes(), "28RU_X1.xls", "28RU_X2.xls", dtcr)


def test_enhanced_workbook_sheets_status_and_dashboard(_real):
    wb = load_workbook(io.BytesIO(_real["output_excel_bytes"]))
    # the WEAVE deliverables are all present as sheets
    for sheet in ("Dashboard", "All Changes", "DTCR Matching", "Yellow Connectors", "PreOrder List"):
        assert sheet in wb.sheetnames
    # requested sheet order: Dashboard, DTCR Matching, Yellow Connectors, All Changes, …
    order = wb.sheetnames
    assert order[0] == "Dashboard"
    assert order.index("DTCR Matching") < order.index("Yellow Connectors") < order.index("All Changes")

    ac = wb["All Changes"]
    assert ac.cell(1, 1).value == "Status" and ac.cell(1, 2).value == "DTCR#"   # Status before DTCR#
    assert len(ac.data_validations.dataValidation) == 1                          # the Status dropdown
    assert len(ac.conditional_formatting._cf_rules) >= 1                         # status row-coloring
    # DTCR# is a live array formula flowing from the DTCR Matching sheet by CNUM
    dtcr_cell = ac.cell(2, 2).value
    dtcr_formula = getattr(dtcr_cell, "text", dtcr_cell)          # openpyxl wraps array formulas
    assert "DTCR Matching" in str(dtcr_formula) and "TEXTJOIN" in str(dtcr_formula)

    db = wb["Dashboard"]
    assert db["A23"].value == "My Harnesses" and db["B23"].value == "Harness Family"   # picker column
    assert db["B9"].value.startswith("=SUMPRODUCT")                               # scoped Done count
    assert db["B14"].value.startswith("=IF")                                      # % complete
    assert db["A12"].value == "Not started"                                       # Not-started row present
    assert len(db._charts) == 3                                                   # progress donut, by-family bar, DTCR pie
