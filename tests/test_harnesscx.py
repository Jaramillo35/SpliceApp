"""Tests for splice.harnesscx — the individual harness-complexity workbench engine."""

from __future__ import annotations

import io

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from splice.common.errors import SpliceError
from splice.harnesscx import adapters, checks, compare, export
from splice.harnesscx.models import ProposalClass

UNIVERSE = {"AHT", "CM5", "CVM", "RS3", "XH3", "XH4", "CJK"}


def _wb_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _master_bytes(include_aht: bool = True) -> bytes:
    """Synthetic master complexity workbook with an 'IP' family sheet.

    Row 9 layout: col3 'Made from', col4 phase 'P1', col5 'Current', then the
    sales-code columns — a single code, a separable OR-list, a combined
    expression, and an equality expression.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "IP"
    ws.cell(1, 1, "Vehicle Program: 2028 RU")
    ws.cell(2, 1, "Build Phase: X2")
    ws.cell(3, 1, "Harness: IP Harness")

    headers = {3: "Made from", 4: "P1", 5: "Current"}
    col = 6
    code_cols: dict[str, int] = {}
    if include_aht:
        headers[col] = "AHT"
        code_cols["AHT"] = col
        col += 1
    headers[col] = "CM5/CVM"          # separable OR-list
    code_cols["CM5/CVM"] = col
    col += 1
    headers[col] = "RS3+(CM5/CVM)"    # combined — SE decision
    code_cols["COMBINED"] = col
    col += 1
    headers[col] = "XH3=XH4"          # equality — auto-resolved
    code_cols["EQ"] = col
    for c, text in headers.items():
        ws.cell(9, c, text)
        ws.cell(7, c, f"feature {text}")

    # row 10: confirmed Current value, marked under AHT (when present)
    ws.cell(10, 1, "A")
    ws.cell(10, 4, "111")
    ws.cell(10, 5, "PN300")
    if include_aht:
        ws.cell(10, code_cols["AHT"], "X")
    # row 11: carryover -> most recent valid P/N from the phase column
    ws.cell(11, 1, "B")
    ws.cell(11, 4, "PN200")
    ws.cell(11, 5, "C/O")
    ws.cell(11, code_cols["CM5/CVM"], "X")
    # row 12: deleted — must be excluded everywhere
    ws.cell(12, 1, "C")
    ws.cell(12, 5, "DELETE P/N")
    # row 13: marked only under the combined and equality expressions
    ws.cell(13, 1, "D")
    ws.cell(13, 5, "PN30")            # prefix lookalike of PN300
    ws.cell(13, code_cols["COMBINED"], "X")
    ws.cell(13, code_cols["EQ"], "G")

    # partitioned family sheet: LEFT/RIGHT marker columns on row 9
    ws2 = wb.create_sheet("SEAT 2ND ROW")
    ws2.cell(9, 3, "Made from")
    ws2.cell(9, 5, "Current")
    ws2.cell(9, 6, "AHT")
    ws2.cell(9, 7, "LEFT")
    ws2.cell(9, 8, "RIGHT")
    ws2.cell(10, 1, "A")
    ws2.cell(10, 5, "PNL1")
    ws2.cell(10, 6, "X")
    ws2.cell(10, 7, "X")              # LEFT part
    ws2.cell(11, 1, "B")
    ws2.cell(11, 5, "PNR1")
    ws2.cell(11, 6, "X")
    ws2.cell(11, 8, "X")              # RIGHT part
    ws2.cell(12, 1, "C")
    ws2.cell(12, 5, "PNC1")
    ws2.cell(12, 6, "X")              # common (no marker)
    return _wb_bytes(wb)


def _crossref_bytes() -> bytes:
    df = pd.DataFrame({
        "Harness Family": ["Instrument Panel", "Seat 2nd Row"],
        "Complexity File": ["IP", "SEAT 2ND ROW"],
        "DTx Family Name": ["IP_DTX", "SEAT_2ND_ROW"],
    })
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    return buf.getvalue()


def _dtx_frames() -> list[pd.DataFrame]:
    return [pd.DataFrame({
        "Harness Family": ["IP_DTX", "IP_DTX", "SEAT_2ND_ROW"],
        "Sales Code": ["AHT", "CM5/CVM, RS3, XH3=XH4", "AHT, CJK"],
    })]


@pytest.fixture()
def crossref():
    return adapters.load_crossref(_crossref_bytes())


@pytest.fixture()
def matrix(crossref):
    return adapters.extract_family_matrix(
        _master_bytes(), "IP", UNIVERSE, "Instrument Panel",
        family_dtx_codes=adapters.family_dtx_sales_codes(_dtx_frames(), crossref, "IP"))


class TestCrossRef:
    def test_mappings(self, crossref):
        assert crossref.worksheet_for_dtx("IP_DTX") == "IP"
        assert crossref.worksheet_to_canonical["IP"] == "Instrument Panel"
        assert "SEAT 2ND ROW" in crossref.worksheets

    def test_missing_columns_rejected(self):
        df = pd.DataFrame({"Wrong": ["x"]})
        buf = io.BytesIO()
        df.to_excel(buf, index=False)
        with pytest.raises(SpliceError, match="missing column"):
            adapters.load_crossref(buf.getvalue())


class TestUniverse:
    def test_dtx_universe_and_family_codes(self, crossref):
        frames = _dtx_frames()
        universe = adapters.dtx_sales_code_universe(frames)
        assert {"AHT", "CM5", "CVM", "RS3", "XH3", "XH4", "CJK"} <= universe
        ip_codes = adapters.family_dtx_sales_codes(frames, crossref, "IP")
        assert "AHT" in ip_codes and "CJK" not in ip_codes


class TestMatrixExtraction:
    def test_sales_code_columns(self, matrix):
        codes = matrix.complexity_codes
        assert {"AHT", "CM5", "CVM"} <= codes
        # combined and equality expressions never become plain columns
        assert "RS3" not in codes and "XH3" not in codes
        cm5 = next(sc for sc in matrix.sales_codes if sc.code == "CM5")
        assert cm5.from_combined and cm5.klass is ProposalClass.UNCERTAIN

    def test_current_pn_classification(self, matrix):
        by_variant = {r.variant_id: r for r in matrix.rows}
        assert by_variant["A"].current_pn == "PN300"
        assert by_variant["A"].current_class is ProposalClass.CONFIRMED
        # C/O resolves to the most recent valid phase P/N, marked inferred
        assert by_variant["B"].current_pn == "PN200"
        assert by_variant["B"].current_class is ProposalClass.INFERRED
        # DELETE row excluded
        assert by_variant["C"].excluded
        assert matrix.excluded_count == 1

    def test_combined_expressions(self, matrix):
        exprs = {ce.original_expr: ce for ce in matrix.combined_exprs}
        assert "RS3+(CM5/CVM)" in exprs and not exprs["RS3+(CM5/CVM)"].include
        eq = exprs["XH3=XH4"]
        assert eq.is_equality and eq.include        # auto-resolved
        assert eq.output_codes == ["XH3", "XH4"]
        d = next(r for r in matrix.rows if r.variant_id == "D")
        assert d.combined_symbols[eq.key] == "G"

    def test_header_meta(self, matrix):
        assert matrix.year == "2028" and matrix.vehicle == "RU"
        assert matrix.phase == "X2" and matrix.harness_name == "IP Harness"

    def test_partition_detection(self, crossref):
        m = adapters.extract_family_matrix(
            _master_bytes(), "SEAT 2ND ROW", UNIVERSE, "Seat 2nd Row")
        assert m.partition_sides == ["LEFT", "RIGHT"]
        sides = {r.current_pn: r.partition_side for r in m.rows}
        assert sides == {"PNL1": "LEFT", "PNR1": "RIGHT", "PNC1": ""}

    def test_missing_worksheet_rejected(self):
        with pytest.raises(SpliceError, match="not found"):
            adapters.extract_family_matrix(_master_bytes(), "NOPE", UNIVERSE)


class TestCompare:
    def test_added_codes_detected(self, crossref):
        universe = adapters.dtx_sales_code_universe(_dtx_frames())
        changes = compare.compare_complexity(
            _master_bytes(include_aht=False), _master_bytes(), crossref, universe)
        ip = next(ch for ch in changes if ch.worksheet == "IP")
        assert ip.added_codes == ["AHT"] and ip.removed_codes == []

    def test_affected_families(self, crossref):
        universe = adapters.dtx_sales_code_universe(_dtx_frames())
        changes = compare.compare_complexity(
            _master_bytes(include_aht=False), _master_bytes(), crossref, universe)
        fams = compare.affected_families({"IP_DTX": 4, "GHOST": 1}, changes, crossref)
        ip = next(a for a in fams if a.worksheet == "IP")
        assert ip.by_dtx and ip.by_complexity and ip.dtx_change_count == 4
        ghost = next(a for a in fams if a.family == "GHOST")
        assert not ghost.resolved


class TestChecks:
    def test_coverage_flags_missing_dtx_code(self, crossref):
        m = adapters.extract_family_matrix(
            _master_bytes(), "SEAT 2ND ROW", UNIVERSE, "Seat 2nd Row",
            family_dtx_codes=adapters.family_dtx_sales_codes(
                _dtx_frames(), crossref, "SEAT 2ND ROW"))
        rows = {r["code"]: r for r in checks.coverage_rows(m)}
        assert rows["CJK"]["in_dtx"] and not rows["CJK"]["in_complexity"]
        assert not rows["CJK"]["ok"]
        assert rows["AHT"]["ok"]

    def test_pn_lookalikes(self, matrix):
        assert ("PN30", "PN300") in checks.pn_lookalikes(matrix)

    def test_unmarked_parts_respects_included_combined(self, matrix):
        # D is marked only under combined expressions; the equality is included,
        # so D is NOT unmarked. Turn it off and D surfaces.
        assert "PN30" not in checks.unmarked_parts(matrix)
        for ce in matrix.combined_exprs:
            ce.include = False
        assert "PN30" in checks.unmarked_parts(matrix)


class TestExport:
    def test_validation_gates(self, matrix):
        assert export.validate_before_export(matrix, "") \
            == ["Harness ID is required (enter it manually)."]
        assert export.validate_before_export(matrix, "H1") == []

    def test_generated_file_layout(self, matrix):
        files, problems = export.generate_files(matrix, "H123")
        assert problems == [] and len(files) == 1
        data, fname = files[0]
        assert fname.startswith("2.- Harness_Complexity_28RU_X2_IP_")
        wb = load_workbook(io.BytesIO(data), keep_vba=True)
        ws = wb["Complexity"]
        assert ws.cell(1, 1).value == "ID=H123"
        headers = [ws.cell(1, c).value for c in range(2, 10) if ws.cell(1, c).value]
        assert "AHT" in headers and "CM5" in headers and "CVM" in headers
        # equality auto-included as one column per side; combined expr left out
        assert "XH3" in headers and "XH4" in headers
        assert "RS3+(CM5/CVM)" not in headers
        pns = [ws.cell(r, 1).value for r in range(2, 6) if ws.cell(r, 1).value]
        assert pns == ["PN300", "PN200", "PN30"]     # DELETE row omitted
        # the equality columns carry D's mark in both
        row_d = 2 + pns.index("PN30")
        marks = {ws.cell(1, c).value: ws.cell(row_d, c).value
                 for c in range(2, 10) if ws.cell(1, c).value}
        assert marks["XH3"] == "G" and marks["XH4"] == "G"
        wb.close()

    def test_partitioned_family_yields_one_file_per_side(self):
        m = adapters.extract_family_matrix(
            _master_bytes(), "SEAT 2ND ROW", UNIVERSE, "Seat 2nd Row")
        files, problems = export.generate_files(m, "H9")
        assert problems == [] and len(files) == 2
        by_name = {}
        for data, fname in files:
            wb = load_workbook(io.BytesIO(data), keep_vba=True)
            ws = wb["Complexity"]
            by_name[fname] = [ws.cell(r, 1).value for r in range(2, 6)
                              if ws.cell(r, 1).value]
            wb.close()
        left = next(v for k, v in by_name.items() if "LEFT" in k)
        right = next(v for k, v in by_name.items() if "RIGHT" in k)
        # each variant file: its own parts plus the common (unmarked) one
        assert left == ["PNL1", "PNC1"]
        assert right == ["PNR1", "PNC1"]

    def test_template_present(self):
        assert export.TEMPLATE_PATH.exists()
        assert len(export.template_bytes()) > 10_000
