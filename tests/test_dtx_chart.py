"""The circuit chart: which part number carries which wire.

The chart is written in the Circuit Summary layout on purpose, so the strong
test is not "does it look right" but "does Circuit Health's own reader take it
back". Everything else here guards the marks themselves.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

from fixtures_dtxcircuits import circuit_rows, harnesses  # noqa: E402

from splice.dtxcircuits import analyze_harness, chart, conventions, report
from splice.dtxcircuits.models import CircuitRow, DtxMeta
from splice.inline.summary import SHEET, read_circuit_summary


META = DtxMeta(program="9000ZZ", phase="X9_A")


def _entries(rows, families=("BODY_LEFT", "IP"), with_complexity=True):
    books = harnesses()
    return [report.Entry(
        label=f"{fam} → {books[fam].name}", family=fam,
        filename=f"{fam}.xlsm",
        analysis=analyze_harness([r for r in rows if r.harness_family == fam],
                                 books[fam], harness_name=books[fam].name),
        complexity=books[fam] if with_complexity else None)
        for fam in families]


@pytest.fixture()
def charts():
    rows = circuit_rows()
    return chart.build_charts(_entries(rows), rows)


def _by_family(charts):
    return {c.family: c for c in charts}


class TestShape:
    def test_one_chart_per_pairing(self, charts):
        assert {c.family for c in charts} == {"BODY_LEFT", "IP"}

    def test_the_block_title_is_what_the_parser_keys_on(self, charts):
        assert _by_family(charts)["BODY_LEFT"].block_title == "BODY_LEFT - 90001"

    def test_part_numbers_come_from_the_complexity(self, charts):
        body = _by_family(charts)["BODY_LEFT"]
        assert body.part_numbers == ["90000001AA", "90000002AA",
                                     "90000003AA", "90000004AA"]

    def test_a_circuit_with_two_ends_is_two_rows(self, charts):
        """CKT_700 lands in two cavities; the chart is per end, not per wire."""
        body = _by_family(charts)["BODY_LEFT"]
        ends = [r for r in body.rows if r.circuit == "CKT_700"]
        assert len(ends) == 2
        assert {(r.cnum, r.cavity) for r in ends} == {("C2", "3"), ("C3", "4")}

    def test_rows_are_ordered_for_reading(self, charts):
        body = _by_family(charts)["BODY_LEFT"]
        assert body.rows == sorted(
            body.rows, key=lambda r: (r.circuit, r.cnum, chart._cavity_key(r.cavity)))

    def test_cavities_sort_numerically(self):
        assert chart._cavity_key("2") < chart._cavity_key("10")
        assert chart._cavity_key("10") < chart._cavity_key("A1")


class TestMarks:
    def test_an_unconditional_circuit_is_carried_by_every_build(self, charts):
        body = _by_family(charts)["BODY_LEFT"]
        row = next(r for r in body.rows if r.circuit == "CKT_100")
        assert row.marks(body.part_numbers) == ["X", "X", "X", "X"]

    def test_a_variant_is_marked_only_where_it_is_built(self, charts):
        body = _by_family(charts)["BODY_LEFT"]
        row = next(r for r in body.rows if r.circuit == "CKT_400")   # AAA&BBB
        assert row.marks(body.part_numbers) == ["", "", "X", ""]

    def test_a_never_built_circuit_is_an_empty_row_and_says_so(self, charts):
        body = _by_family(charts)["BODY_LEFT"]
        row = next(r for r in body.rows if r.circuit == "CKT_500")   # CCC
        assert row.marks(body.part_numbers) == ["", "", "", ""]
        assert row.is_finding, "an empty row must be flagged, not just empty"
        assert body.findings == 1

    def test_both_ends_of_one_circuit_carry_the_same_marks(self, charts):
        """Applicability is a property of the wire, not of the cavity."""
        body = _by_family(charts)["BODY_LEFT"]
        ends = [r for r in body.rows if r.circuit == "CKT_700"]
        assert ends[0].marks(body.part_numbers) == ends[1].marks(body.part_numbers)

    def test_coverage_counts_ends_per_part_number(self, charts):
        body = _by_family(charts)["BODY_LEFT"]
        for pn in body.part_numbers:
            expected = sum(1 for r in body.rows if pn in r.builds)
            assert body.coverage(pn) == expected

    def test_an_untracked_code_reads_as_carried(self, charts):
        """CKT_600 rests on ZZZ, which BODY_LEFT does not track.

        The resolver treats an unknown code as present, so the chart must show
        the circuit as carried — showing it absent would invent a decision the
        data does not support.
        """
        body = _by_family(charts)["BODY_LEFT"]
        row = next(r for r in body.rows if r.circuit == "CKT_600")
        assert row.marks(body.part_numbers) == ["X", "X", "X", "X"]


class TestRoundTrip:
    """What we write, Circuit Health reads."""

    def _reread(self, charts):
        data = chart.build_chart_workbook(charts, META.program, META.phase)
        return read_circuit_summary(data, "chart.xlsx")

    def test_harnesses_and_builds_survive(self, charts):
        harns, _ends = self._reread(charts)
        assert set(harns) == {"90001", "90002"}
        assert [b.part_number for b in harns["90001"].builds] == \
            _by_family(charts)["BODY_LEFT"].part_numbers

    def test_every_end_survives(self, charts):
        _harns, ends = self._reread(charts)
        assert len(ends) == sum(len(c.rows) for c in charts)

    def test_the_marks_survive(self, charts):
        _harns, ends = self._reread(charts)
        body = _by_family(charts)["BODY_LEFT"]
        for row in body.rows:
            end = next(e for e in ends
                       if e.circuit == row.circuit and e.connector == row.cnum
                       and e.cavity == row.cavity)
            assert set(end.builds) == set(row.builds), row.circuit

    def test_the_sales_code_survives_as_the_harness_form(self, charts):
        """The Sales Code column carries the condition restated in THIS
        harness's codes — that is the one true of this block's part numbers.
        The DTx wording rides along in Suffix so nothing is lost."""
        _harns, ends = self._reread(charts)
        body = _by_family(charts)["BODY_LEFT"]
        row = next(r for r in body.rows if r.circuit == "CKT_400")
        end = next(e for e in ends if e.circuit == "CKT_400")
        assert end.sales_code == (row.harness_expression or row.expression)
        if row.harness_expression and row.harness_expression != row.expression:
            assert end.suffix == row.expression

    def test_a_never_built_circuit_round_trips_as_carried_by_nothing(self, charts):
        _harns, ends = self._reread(charts)
        end = next(e for e in ends if e.circuit == "CKT_500")
        assert end.builds == frozenset()


class TestWorkbook:
    def test_the_sheet_is_named_for_the_parser(self, charts):
        data = chart.build_chart_workbook(charts)
        assert SHEET in load_workbook(io.BytesIO(data)).sheetnames

    def test_wire_physicals_are_left_blank_rather_than_invented(self, charts):
        """Size, material and colour are not in the DTx. A guessed gauge would
        be read as fact, so those columns stay empty."""
        _harns, ends = TestRoundTrip()._reread(charts)
        assert all(not (e.size or e.material or e.color) for e in ends)

    def test_the_review_workbook_carries_the_chart(self, charts):
        rows = circuit_rows()
        data = report.build_report(_entries(rows), {}, charts=charts)
        wb = load_workbook(io.BytesIO(data))
        assert SHEET in wb.sheetnames
        assert "Circuits" in wb.sheetnames, "the review sheets must still be there"

    def test_the_review_workbook_is_unchanged_without_a_chart(self):
        rows = circuit_rows()
        data = report.build_report(_entries(rows), {})
        assert SHEET not in load_workbook(io.BytesIO(data)).sheetnames


class TestUniversalCode:
    """``501`` alone means every harness part number (SE ruling, 2026-09-01)."""

    def test_a_bare_universal_code_is_unconditional(self):
        assert conventions.is_universal("501")
        assert conventions.effective_condition("501") is None
        assert conventions.effective_condition("  501 ") is None

    def test_inside_a_larger_expression_it_stays_an_ordinary_code(self):
        """The narrow reading the SE asked for: 501/RHV and 501&HAH are not
        rewritten, because widening the rule would silently restate them."""
        assert not conventions.is_universal("501/RHV")
        assert conventions.effective_condition("501/RHV") == "501/RHV"
        assert conventions.effective_condition("501&HAH") == "501&HAH"

    def test_a_bare_universal_circuit_is_carried_by_every_build(self):
        rows = circuit_rows() + [
            CircuitRow(harness_family="IP", circuit="CKT_970", sales_code="501",
                       cnum="C6", pin="6", connector_pn="99999999")]
        built = chart.build_charts(_entries(rows), rows)
        ip = _by_family(built)["IP"]
        row = next(r for r in ip.rows if r.circuit == "CKT_970")
        assert row.marks(ip.part_numbers) == ["X", "X"]
        assert row.expression == "", "a universal code is not a condition"

    def test_it_is_not_reported_as_a_sales_code_gap(self):
        """No complexity file lists 501, so reporting it gives the customer
        nothing they can act on."""
        from splice.dtxcircuits.analyze import code_gaps
        rows = [CircuitRow(harness_family="IP", circuit="CKT_970",
                           sales_code="501", cnum="C6", pin="6")]
        assert code_gaps(rows, harnesses()["IP"]) == []

    def test_but_a_mixed_expression_still_reports_its_real_gaps(self):
        from splice.dtxcircuits.analyze import code_gaps
        rows = [CircuitRow(harness_family="IP", circuit="CKT_971",
                           sales_code="501/ZZZ", cnum="C6", pin="6")]
        codes = {g.code for g in code_gaps(rows, harnesses()["IP"])}
        assert "ZZZ" in codes, "an untracked code in a mixed expression is real"


class TestReconciliation:
    """Applicability is per harness, decided by the devices THAT harness has.

    Drawn from real shapes in 2028RU X2_A. A937 reaches a bare-501 ground in
    the IP (always fitted) and an HBB device in the HVAC_REAR (fitted only on
    HBB); D442 reads SDE at every device and is blank only on the DASH
    inlines it passes through.
    """

    def _dev(self, family, circuit, code, cnum, pin="1"):
        return CircuitRow(harness_family=family, circuit=circuit,
                          sales_code=code, cnum=cnum, pin=pin)

    # ---------------------------------------------------------------- inlines
    def test_an_inline_is_recognised_by_its_name(self):
        assert chart.is_pass_through("X301A")
        assert chart.is_pass_through("Y301A")
        assert chart.is_pass_through("I350X")
        assert not chart.is_pass_through("D6630B")
        assert not chart.is_pass_through("G911A")
        assert not chart.is_pass_through("")

    def test_a_blank_inline_is_silence_not_unconditional(self):
        """D442: every device says SDE, only the pass-through DASH is blank.
        Reading that blank as 'always' collapsed the whole circuit."""
        rows = [
            self._dev("POWERTRAIN", "D442", "SDE", "D2798A"),
            self._dev("POWERTRAIN", "D442", "SDE", "X200A", "33"),
            self._dev("DASH", "D442", "", "X402A", "27"),
            self._dev("DASH", "D442", "SDE", "Y200A", "33"),
            self._dev("BODY_RIGHT", "D442", "", "Y402A", "27"),
            self._dev("BODY_RIGHT", "D442", "SDE", "D3872A", "9"),
        ]
        resolved = chart.harness_conditions(rows)
        assert resolved[("D442", "POWERTRAIN")] == "(SDE)"
        assert resolved[("D442", "BODY_RIGHT")] == "(SDE)"

    def test_a_pass_through_harness_inherits_rather_than_going_unconditional(self):
        """DASH holds only inlines for D442. It is SDE too, not 'always'."""
        rows = [
            self._dev("POWERTRAIN", "D442", "SDE", "D2798A"),
            self._dev("DASH", "D442", "", "X402A", "27"),
            self._dev("BODY_RIGHT", "D442", "SDE", "D3872A", "9"),
        ]
        assert chart.harness_conditions(rows)[("D442", "DASH")] == "(SDE)"

    def test_a_stated_inline_beats_inheritance(self):
        rows = [
            self._dev("POWERTRAIN", "D442", "SDE", "D2798A"),
            self._dev("DASH", "D442", "HAH", "Y200A", "33"),
        ]
        assert chart.harness_conditions(rows)[("D442", "DASH")] == "(HAH)"

    # ------------------------------------------------------- per harness, 501
    def test_a_501_ground_does_not_erase_a_condition_elsewhere(self):
        """The reported defect. A937 is always present in the IP because a
        bare-501 device is always fitted there, and present only on HBB in
        the HVAC_REAR. One circuit-wide condition cannot say both."""
        rows = [
            self._dev("IP", "A937", "501", "D3816B", "6"),
            self._dev("IP", "A937", "XC4", "D2269A", "9"),
            self._dev("HVAC_REAR", "A937", "HBB", "D3828C", "10"),
            self._dev("HVAC_FRONT", "A937", "501", "D3828A", "2"),
        ]
        resolved = chart.harness_conditions(rows)
        assert resolved[("A937", "IP")] is None, "a 501 device is always fitted"
        assert resolved[("A937", "HVAC_FRONT")] is None
        assert resolved[("A937", "HVAC_REAR")] == "(HBB)"

    def test_conditions_in_one_harness_are_or_ed(self):
        """Any fitted device pulls the wire into that harness."""
        rows = [
            self._dev("HEADLINER", "Z911", "LBR", "D6627A", "9"),
            self._dev("HEADLINER", "Z911", "GN9", "D4594A", "2"),
            self._dev("HEADLINER", "Z911", "XPR", "D3467A", "4"),
        ]
        assert chart.harness_conditions(rows)[("Z911", "HEADLINER")] == \
            "(GN9)/(LBR)/(XPR)"

    def test_a_blank_device_end_is_unconditional(self):
        """Unlike an inline, a device with no condition really is always on."""
        rows = [self._dev("IP", "CKT", "", "D100A"),
                self._dev("IP", "CKT", "HAH", "D200A")]
        assert chart.harness_conditions(rows)[("CKT", "IP")] is None

    # ------------------------------------------------------------ whole circuit
    def test_the_whole_circuit_view_ignores_inlines_too(self):
        rows = [
            self._dev("POWERTRAIN", "D442", "SDE", "D2798A"),
            self._dev("DASH", "D442", "", "X402A", "27"),
        ]
        assert chart.flowed_conditions(rows)["D442"] == "(SDE)"

    def test_each_harness_carries_its_own_condition_into_the_chart(self):
        rows = [
            self._dev("BODY_LEFT", "CKT_TIGHT", "AAA", "D1A"),
            self._dev("IP", "CKT_TIGHT", "BBB", "D2A"),
        ]
        built = {c.family: c for c in chart.build_charts(_entries(rows), rows)}
        assert {r.expression for r in built["BODY_LEFT"].rows} == {"(AAA)"}
        assert {r.expression for r in built["IP"].rows} == {"(BBB)"}


class TestHarnessExpression:
    """Restated only where the harness has no column for a code."""

    def test_a_condition_already_in_the_vocabulary_is_left_alone(self, charts):
        body = _by_family(charts)["BODY_LEFT"]
        row = next(r for r in body.rows if r.circuit == "CKT_400")   # AAA&BBB
        assert row.harness_expression == row.expression

    def test_a_condition_on_an_untracked_code_is_restated(self):
        """BODY_LEFT does not track ZZZ, so (ZZZ) cannot be stated in its
        codes; it selects every build, which is not a condition at all."""
        rows = circuit_rows()
        body = _by_family(chart.build_charts(_entries(rows), rows))["BODY_LEFT"]
        row = next(r for r in body.rows if r.circuit == "CKT_600")
        assert row.harness_expression == ""
        assert row.marks(body.part_numbers) == ["X", "X", "X", "X"]

    def test_a_restatement_selects_exactly_the_same_builds(self):
        """Whatever wording comes back must not move a single mark."""
        from splice.inline.complexity import applies_in
        rows = circuit_rows()
        books = harnesses()
        for c in chart.build_charts(_entries(rows), rows):
            complexity = books[c.family]
            for row in c.rows:
                if not row.harness_expression:
                    continue
                selected = {b.part_number for b in complexity.builds
                            if applies_in(row.harness_expression, b.codes,
                                          complexity.complexity_codes)}
                assert selected == set(row.builds), row.circuit

    def test_all_builds_and_no_builds_are_not_conditions(self, charts):
        for c in charts:
            for row in c.rows:
                if not row.builds or len(row.builds) == len(c.part_numbers):
                    assert row.harness_expression == ""

    def test_without_a_complexity_file_it_falls_back_honestly(self):
        """No complexity means no restatement — and the marks come from the
        analysis, not from an empty set that would read as never built."""
        rows = circuit_rows()
        built = chart.build_charts(_entries(rows, with_complexity=False), rows)
        body = _by_family(built)["BODY_LEFT"]
        row = next(r for r in body.rows if r.circuit == "CKT_100")
        assert row.harness_expression == ""
        assert row.marks(body.part_numbers) == ["X", "X", "X", "X"]


class TestSplices:
    """Three ends is a branch, and a branch has to join somewhere."""

    def _rows(self, ends: int, family: str = "IP"):
        return [CircuitRow(harness_family=family, circuit="CKT_BRANCH",
                           sales_code="", cnum=f"C{i}", pin=str(i + 1))
                for i in range(ends)]

    def _chart(self, ends: int):
        rows = self._rows(ends)
        return _by_family(chart.build_charts(_entries(rows, families=("IP",)),
                                             rows))["IP"]

    def test_two_ends_is_a_wire_not_a_splice(self):
        c = self._chart(2)
        assert c.splices == {}
        assert not any(r.is_splice for r in c.rows)

    def test_three_ends_gets_a_splice(self):
        c = self._chart(3)
        assert c.splices == {"CKT_BRANCH": "SCKT_BRANCHA"}

    def test_the_splice_is_named_the_way_splice_generation_names_them(self):
        assert self._chart(3).splices["CKT_BRANCH"] == "SCKT_BRANCHA"

    def test_one_splice_cavity_per_branch(self):
        c = self._chart(4)
        cavities = [r.cavity for r in c.rows if r.is_splice]
        assert cavities == ["A", "B", "C", "D"]

    def test_the_splice_wires_carry_what_their_branch_carries(self):
        rows = self._rows(3)
        rows[0] = CircuitRow(harness_family="IP", circuit="CKT_BRANCH",
                             sales_code="AAA", cnum="C0", pin="1")
        c = _by_family(chart.build_charts(_entries(rows, families=("IP",)),
                                          rows))["IP"]
        device_ends = [r for r in c.rows if not r.is_splice]
        splice_ends = [r for r in c.rows if r.is_splice]
        assert len(splice_ends) == len(device_ends)
        assert {tuple(r.builds) for r in splice_ends} == \
               {tuple(r.builds) for r in device_ends}

    def test_cavities_continue_past_Z_rather_than_truncating(self):
        """A real ground net in 2028RU X2_A has 269 ends. Stopping at Z would
        silently drop wires."""
        c = self._chart(30)
        cavities = [r.cavity for r in c.rows if r.is_splice]
        assert cavities[:2] == ["A", "B"]
        assert cavities[25:28] == ["Z", "AA", "AB"]
        assert len(cavities) == 30

    def test_cavities_sort_in_allocation_order(self):
        assert chart._cavity_key("Z") < chart._cavity_key("AA")
        assert chart._cavity_key("AA") < chart._cavity_key("AB")
        assert chart._cavity_key("2") < chart._cavity_key("10")

    def test_splices_can_be_turned_off(self):
        rows = self._rows(5)
        c = _by_family(chart.build_charts(_entries(rows, families=("IP",)), rows,
                                          splice_min_ends=0))["IP"]
        assert c.splices == {}

    def test_a_spliced_chart_still_round_trips(self):
        rows = self._rows(4)
        built = chart.build_charts(_entries(rows, families=("IP",)), rows)
        data = chart.build_chart_workbook(built)
        _harns, ends = read_circuit_summary(data, "spliced.xlsx")
        assert len(ends) == 8, "four device ends and four splice ends"
        assert sum(1 for e in ends if e.connector == "SCKT_BRANCHA") == 4


class TestOtherEnd:
    """A wire has two ends; the chart writes one per row, so each says where
    it goes. Shapes taken from 2028RU X2_A."""

    def _dev(self, family, circuit, cnum, pin="1", code=""):
        return CircuitRow(harness_family=family, circuit=circuit,
                          sales_code=code, cnum=cnum, pin=pin)

    def _rows_for(self, charts, circuit):
        return {(c.family, r.cnum): r for c in charts for r in c.rows
                if r.circuit == circuit}

    def test_an_inline_mates_with_its_other_half_next_door(self):
        """Crossing a harness boundary is a mating joint, not a wire."""
        rows = [self._dev("IP", "D442", "X402A", "27"),
                self._dev("BODY_LEFT", "D442", "Y402A", "27")]
        built = chart.build_charts(_entries(rows), rows)
        ends = self._rows_for(built, "D442")
        near = ends[("IP", "X402A")]
        assert near.mate_family == "BODY_LEFT"
        assert near.mate_cnum == "Y402A"
        assert ends[("BODY_LEFT", "Y402A")].mate_cnum == "X402A"

    def test_a_lone_end_gets_no_wire_of_its_own(self):
        """One end in a harness has nothing inside it to wire to. Saying so
        beats inventing a partner."""
        rows = [self._dev("IP", "D442", "X402A", "27"),
                self._dev("BODY_LEFT", "D442", "Y402A", "27")]
        built = chart.build_charts(_entries(rows), rows)
        assert all(not r.other_cnum for c in built for r in c.rows)

    def test_two_device_ends_in_ONE_harness_join_each_other(self):
        rows = [self._dev("IP", "F946", "D2798A", "5"),
                self._dev("IP", "F946", "D3872A", "13")]
        built = chart.build_charts(_entries(rows, families=("IP",)), rows)
        ends = self._rows_for(built, "F946")
        assert ends[("IP", "D2798A")].other_cnum == "D3872A"
        assert ends[("IP", "D3872A")].other_cnum == "D2798A"

    def test_devices_in_DIFFERENT_harnesses_are_not_wired_together(self):
        """A wire cannot leave its harness except through a connector, so
        two devices either side of a boundary are not one wire — drawing
        one ran it straight through whatever lay in between."""
        rows = [self._dev("IP", "F946", "D2798A", "5"),
                self._dev("BODY_LEFT", "F946", "D3872A", "13")]
        built = chart.build_charts(_entries(rows), rows)
        assert all(not r.other_cnum for c in built for r in c.rows)

    def test_every_branch_of_a_splice_points_at_the_splice(self):
        rows = [self._dev("IP", "A910", f"D{i}A", str(i)) for i in range(3)]
        built = chart.build_charts(_entries(rows, families=("IP",)), rows)
        ip = _by_family(built)["IP"]
        for row in ip.rows:
            if row.is_splice:
                assert not row.other_cnum.startswith("S")
            else:
                assert row.other_cnum == "SA910A", row.cnum

    def test_the_splice_end_points_back_at_its_branch(self):
        rows = [self._dev("IP", "A910", f"D{i}A", str(i)) for i in range(3)]
        built = chart.build_charts(_entries(rows, families=("IP",)), rows)
        ip = _by_family(built)["IP"]
        pairs = {(r.cnum, r.cavity): (r.other_cnum, r.other_cavity)
                 for r in ip.rows}
        for (cnum, cav), (other_cnum, other_cav) in pairs.items():
            assert pairs[(other_cnum, other_cav)] == (cnum, cav), \
                "the far end must point back"

    def test_an_ambiguous_circuit_is_left_blank_rather_than_guessed(self):
        """Three unpaired ends spread over harnesses have no single far end.
        Inventing one would put a wire in the chart nobody drew."""
        rows = [self._dev("IP", "CKT_X", "D1A"),
                self._dev("BODY_LEFT", "CKT_X", "D2A"),
                self._dev("DASH", "CKT_X", "D3A")]
        built = chart.build_charts(_entries(rows, families=("IP", "BODY_LEFT")),
                                   rows)
        assert all(not r.other_cnum for c in built for r in c.rows)

    def test_end_type_names_what_the_row_is(self):
        rows = [self._dev("IP", "CKT_Y", "D1A"),
                self._dev("IP", "CKT_Y", "X301A", "2")]
        built = chart.build_charts(_entries(rows, families=("IP",)), rows)
        kinds = {r.cnum: r.end_type for r in _by_family(built)["IP"].rows}
        assert kinds["D1A"] == "Device"
        assert kinds["X301A"] == "Inline"


class TestOneFamilyManyHarnesses:
    """The reported bug. A DTx family may be mapped to several complexity
    files, and then the same DTx rows appear once per chart. Pairing used to
    count a circuit's ends across every chart, so the count never matched the
    export and nothing was joined — on 2028RU X1, circuit A0's two devices
    inside BATTERY POSITIVE were unconnected in both of its charts.
    """

    def _rows(self):
        return [CircuitRow(harness_family="BATTERY_POSITIVE", circuit="A0",
                           sales_code="XHZ", cnum=cnum, pin="1",
                           connector_pn="99999999",
                           function="PWR - BATT TO PWR DISTRIBUTION POS")
                for cnum in ("D6630A", "D7402E")]

    def _two_charts(self, rows):
        """One family, two harnesses — as the SE maps ESS1 and ESS2."""
        from splice.dtxcircuits import analyze_harness
        return chart.build_charts([
            report.Entry(label=name, family="BATTERY_POSITIVE", filename=name,
                         analysis=analyze_harness(rows, None, harness_name=name))
            for name in ("BATTERY POSITIVE ESS1", "BATTERY POSITIVE ESS2")], rows)

    def test_each_chart_wires_its_own_copy(self):
        built = self._two_charts(self._rows())
        assert len(built) == 2
        for c in built:
            pairs = {r.cnum: r.other_cnum for r in c.rows}
            assert pairs == {"D6630A": "D7402E", "D7402E": "D6630A"}, c.harness

    def test_the_wire_stays_inside_the_harness_it_belongs_to(self):
        for c in self._two_charts(self._rows()):
            for row in c.rows:
                assert row.other_family == c.family

    def test_a_third_harness_does_not_disturb_them(self):
        rows = self._rows() + [
            CircuitRow(harness_family="IP", circuit="A0", sales_code="XHZ",
                       cnum="D100A", pin="1", connector_pn="99999999")]
        from splice.dtxcircuits import analyze_harness
        built = chart.build_charts([
            report.Entry(label=n, family=f, filename=n,
                         analysis=analyze_harness(
                             [r for r in rows if r.harness_family == f], None,
                             harness_name=n))
            for f, n in (("BATTERY_POSITIVE", "ESS1"),
                         ("BATTERY_POSITIVE", "ESS2"), ("IP", "IP"))], rows)
        battery = [c for c in built if c.family == "BATTERY_POSITIVE"]
        assert all(r.other_cnum for c in battery for r in c.rows)
        # the lone IP end has nothing in its harness to wire to
        ip = next(c for c in built if c.family == "IP")
        assert all(not r.other_cnum for r in ip.rows)


class TestInlinePath:
    """Both connections, which is what the SE asked for: the wire from the
    device to the inline inside one harness, and the inline-to-inline joint
    that carries the circuit into the next.
    """

    def _dev(self, family, cnum, pin="1"):
        return CircuitRow(harness_family=family, circuit="D442", sales_code="",
                          cnum=cnum, pin=pin, connector_pn="99999999")

    def _built(self):
        rows = [self._dev("POWERTRAIN", "D2798A", "1"),
                self._dev("POWERTRAIN", "X200A", "33"),
                self._dev("DASH", "Y200A", "33"),
                self._dev("DASH", "X402A", "27"),
                self._dev("BODY_RIGHT", "Y402A", "27"),
                self._dev("BODY_RIGHT", "D3872A", "9")]
        from splice.dtxcircuits import analyze_harness
        return {c.family: c for c in chart.build_charts([
            report.Entry(label=f, family=f, filename=f,
                         analysis=analyze_harness(
                             [r for r in rows if r.harness_family == f], None,
                             harness_name=f))
            for f in ("POWERTRAIN", "DASH", "BODY_RIGHT")], rows)}

    def test_the_device_is_wired_to_the_inline_in_its_own_harness(self):
        pt = self._built()["POWERTRAIN"]
        wires = {r.cnum: r.other_cnum for r in pt.rows}
        assert wires == {"D2798A": "X200A", "X200A": "D2798A"}

    def test_the_inline_mates_across_the_boundary(self):
        built = self._built()
        pt = next(r for r in built["POWERTRAIN"].rows if r.cnum == "X200A")
        assert (pt.mate_family, pt.mate_cnum) == ("DASH", "Y200A")

    def test_a_pass_through_harness_wires_its_two_inlines_together(self):
        """DASH only passes the circuit along: its wire runs inline to inline."""
        dash = self._built()["DASH"]
        wires = {r.cnum: r.other_cnum for r in dash.rows}
        assert wires == {"Y200A": "X402A", "X402A": "Y200A"}

    def test_the_whole_path_is_walkable_end_to_end(self):
        """Device to device across three harnesses, alternating wire and
        mate, with nothing missing in the middle."""
        built = self._built()
        index = {(c.family, r.cnum): (c, r) for c in built.values() for r in c.rows}
        family, cnum, walked = "POWERTRAIN", "D2798A", ["D2798A"]
        for _ in range(6):
            _c, row = index[(family, cnum)]
            if row.other_cnum and row.other_cnum not in walked:
                cnum = row.other_cnum
            elif row.mate_cnum and row.mate_cnum not in walked:
                family, cnum = row.mate_family, row.mate_cnum
            else:
                break
            walked.append(cnum)
        assert walked == ["D2798A", "X200A", "Y200A", "X402A", "Y402A", "D3872A"]


class TestFlatSheet:
    """One table, one header, no column that is empty for every row."""

    def _sheet(self, charts, name=chart.FLAT_SHEET):
        data = chart.build_chart_workbook(charts, "9000ZZ", "X9_A")
        return load_workbook(io.BytesIO(data))[name]

    def test_the_only_header_is_row_two(self, charts):
        ws = self._sheet(charts)
        assert ws.cell(1, 1).value == "Circuit Chart"
        assert [c.value for c in ws[2]][:4] == \
            ["Harness Family", "Harness", "Def Id", "Circuit"]
        # no repeated banner anywhere below it
        first = [c.value for c in ws[2]]
        for row in ws.iter_rows(min_row=3, values_only=True):
            assert list(row) != first

    def test_every_harness_is_in_the_one_table(self, charts):
        ws = self._sheet(charts)
        families = {row[0] for row in ws.iter_rows(min_row=3, values_only=True)}
        assert families == {c.family for c in charts}

    def test_no_column_is_empty_for_every_row(self):
        """Built with an inline pair, so the Mates With columns are exercised
        too — a column nothing ever fills has no business on the sheet."""
        rows = circuit_rows() + [
            CircuitRow(harness_family="BODY_LEFT", circuit="CKT_SPAN",
                       sales_code="", cnum="X301A", pin="1",
                       connector_pn="99999999", function="TEST"),
            CircuitRow(harness_family="BODY_LEFT", circuit="CKT_SPAN",
                       sales_code="", cnum="D900A", pin="2",
                       connector_pn="99999999", function="TEST"),
            CircuitRow(harness_family="IP", circuit="CKT_SPAN",
                       sales_code="", cnum="Y301A", pin="1",
                       connector_pn="99999999", function="TEST"),
            CircuitRow(harness_family="IP", circuit="CKT_SPAN",
                       sales_code="", cnum="D901A", pin="2",
                       connector_pn="99999999", function="TEST"),
        ]
        charts = chart.build_charts(_entries(rows), rows)
        ws = self._sheet(charts)
        headers = [c.value for c in ws[2]]
        body = list(ws.iter_rows(min_row=3, values_only=True))
        for index, name in enumerate(headers):
            assert any(row[index] not in (None, "") for row in body), \
                f"column {name!r} is empty everywhere"

    def test_the_dropped_columns_are_the_ones_the_dtx_cannot_fill(self, charts):
        headers = {c.value for c in self._sheet(charts)[2]}
        assert not {"Size", "Material", "Color", "Suffix"} & headers

    def test_the_other_end_columns_are_present(self, charts):
        headers = [c.value for c in self._sheet(charts)[2]]
        for name in ("Other End Harness", "Other End CNUM",
                     "Other End Cavity", "Other End Device"):
            assert name in headers

    def test_a_part_number_column_marks_only_its_own_harness(self):
        rows = circuit_rows()
        charts = chart.build_charts(_entries(rows), rows)
        ws = self._sheet(charts)
        headers = [c.value for c in ws[2]]
        column = headers.index("90000001AA")          # a BODY_LEFT build
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] != "BODY_LEFT":
                assert row[column] in (None, ""), \
                    "a build must not be marked outside its own harness"

    def test_part_number_columns_stay_in_harness_order(self, charts):
        columns = chart.part_number_columns(charts)
        assert columns == sorted(
            columns, key=lambda pair: [c.harness for c in charts].index(pair[0]))

    def test_the_blocked_sheet_is_still_written_beside_it(self, charts):
        """Flattening the blocks must not cost the Circuit Health round trip."""
        data = chart.build_chart_workbook(charts)
        wb = load_workbook(io.BytesIO(data))
        assert wb.sheetnames == [chart.FLAT_SHEET, SHEET]
        harns, ends = read_circuit_summary(data, "both.xlsx")
        assert len(ends) == sum(len(c.rows) for c in charts)

    def test_the_review_workbook_carries_both(self, charts):
        rows = circuit_rows()
        wb = load_workbook(io.BytesIO(
            report.build_report(_entries(rows), {}, charts=charts)))
        assert chart.FLAT_SHEET in wb.sheetnames and SHEET in wb.sheetnames


class TestNoConnect:
    """``N0`` is the DTx's marker for a cavity wired to nothing.

    It is 1,570 of 5,412 rows in 2028RU X2_A. Treated as a circuit it became
    3,120 chart rows, one fabricated 269-cavity splice, and 3,106 far-end
    links to wires nobody drew. The chart's job is to say where wires go, so
    a No Connect is not one of its rows.

    Note the scope: the *chart* excludes them. The applicability analysis
    still sees them, because that was the ask and because a No Connect can
    still carry a sales code worth reviewing.
    """

    def _row(self, family, circuit, cnum, pin="1", function="TEST", code=""):
        return CircuitRow(harness_family=family, circuit=circuit,
                          sales_code=code, cnum=cnum, pin=pin,
                          connector_pn="99999999", function=function)

    # ------------------------------------------------------- the predicate
    def test_it_knows_a_no_connect_by_name(self):
        assert conventions.is_no_connect("N0")
        assert conventions.is_no_connect(" n0 "), "case and padding are noise"

    def test_it_knows_one_by_its_function_too(self):
        """Both columns agreed on all 1,570 rows of the reference export, so
        either alone identifies the row and a rename of one is still caught."""
        assert conventions.is_no_connect("SOMETHING", "No Connect")
        assert conventions.is_no_connect("SOMETHING", "  no connect  ")

    def test_a_real_circuit_is_not_one(self):
        assert not conventions.is_no_connect("QK101", "IP - POWER FEED")
        assert not conventions.is_no_connect("", "")

    def test_a_name_that_merely_starts_with_n0_is_not_one(self):
        """N01 is a circuit. The rule matches the whole name, not a prefix."""
        assert not conventions.is_no_connect("N01")
        assert not conventions.is_no_connect("N0A")

    # ----------------------------------------------------------- the chart
    def _built(self, rows, families=("IP",)):
        return _by_family(chart.build_charts(_entries(rows, families=families), rows))

    def test_no_connect_rows_never_become_chart_rows(self):
        rows = [self._row("IP", "QK900", "C1"),
                self._row("IP", "N0", "C2", "2", function="No Connect"),
                self._row("IP", "N0", "C3", "3", function="No Connect")]
        ip = self._built(rows)["IP"]
        assert [r.circuit for r in ip.rows] == ["QK900"]

    def test_what_was_dropped_is_counted_not_hidden(self):
        rows = [self._row("IP", "QK900", "C1"),
                self._row("IP", "N0", "C2", "2", function="No Connect"),
                self._row("IP", "N0", "C3", "3", function="No Connect")]
        assert self._built(rows)["IP"].no_connect_rows == 2

    def test_nothing_is_counted_when_there_is_nothing_to_drop(self):
        rows = [self._row("IP", "QK900", "C1")]
        assert self._built(rows)["IP"].no_connect_rows == 0

    def test_no_connects_are_never_spliced_together(self):
        """The 269-cavity SN0A splice: three or more ends of the same circuit
        normally get one, and N0 had hundreds. It must not."""
        rows = [self._row("IP", "N0", f"C{i}", str(i), function="No Connect")
                for i in range(5)]
        ip = self._built(rows)["IP"]
        assert ip.splices == {}
        assert ip.rows == []
        assert ip.no_connect_rows == 5

    def test_a_real_circuit_still_splices_beside_them(self):
        rows = [self._row("IP", "QK900", f"C{i}", str(i)) for i in range(3)]
        rows += [self._row("IP", "N0", f"D{i}", str(i), function="No Connect")
                 for i in range(4)]
        ip = self._built(rows)["IP"]
        assert ip.splices == {"QK900": "SQK900A"}
        assert all(r.circuit == "QK900" for r in ip.rows)

    def test_no_far_end_is_invented_between_two_no_connects(self):
        rows = [self._row("IP", "N0", "X301A", "1", function="No Connect"),
                self._row("BODY_LEFT", "N0", "Y301A", "1", function="No Connect")]
        built = chart.build_charts(
            _entries(rows, families=("IP", "BODY_LEFT")), rows)
        assert all(not r.other_cnum for c in built for r in c.rows)
        assert sum(c.no_connect_rows for c in built) == 2

    def test_a_real_pair_still_finds_each_other_past_the_dropped_rows(self):
        """Removing rows must not strand the circuits that remain."""
        rows = [self._row("IP", "N0", "C9", "9", function="No Connect"),
                self._row("IP", "D442", "X301A", "1"),
                self._row("BODY_LEFT", "D442", "Y301A", "1")]
        built = {c.family: c for c in chart.build_charts(
            _entries(rows, families=("IP", "BODY_LEFT")), rows)}
        near = next(r for r in built["IP"].rows if r.circuit == "D442")
        assert near.mate_cnum == "Y301A"
        assert near.mate_family == "BODY_LEFT"

    def test_a_family_of_nothing_but_no_connects_is_empty_not_broken(self):
        rows = [self._row("IP", "N0", "C1", "1", function="No Connect")]
        ip = self._built(rows)["IP"]
        assert ip.rows == [] and ip.no_connect_rows == 1
        assert ip.circuits == 0 and ip.findings == 0

    # -------------------------------------------------------- the workbook
    def test_the_workbook_says_what_it_left_out(self):
        rows = [self._row("IP", "QK900", "C1"),
                self._row("IP", "N0", "C2", "2", function="No Connect")]
        built = chart.build_charts(_entries(rows, families=("IP",)), rows)
        data = chart.build_chart_workbook(built, "9000ZZ", "X9_A")
        wb = load_workbook(io.BytesIO(data))
        for sheet in (chart.FLAT_SHEET, SHEET):
            note = str(wb[sheet].cell(1, 3).value or "")
            assert "1 No Connect row(s) excluded" in note, sheet

    def test_the_note_is_absent_when_nothing_was_excluded(self):
        rows = [self._row("IP", "QK900", "C1")]
        built = chart.build_charts(_entries(rows, families=("IP",)), rows)
        wb = load_workbook(io.BytesIO(
            chart.build_chart_workbook(built, "9000ZZ", "X9_A")))
        assert not wb[chart.FLAT_SHEET].cell(1, 3).value

    def test_the_workbook_still_round_trips_without_them(self):
        rows = [self._row("IP", "QK900", "C1"),
                self._row("IP", "QK901", "C2", "2"),
                self._row("IP", "N0", "C3", "3", function="No Connect")]
        built = chart.build_charts(_entries(rows, families=("IP",)), rows)
        data = chart.build_chart_workbook(built)
        _harns, ends = read_circuit_summary(data, "nc.xlsx")
        assert {e.circuit for e in ends} == {"QK900", "QK901"}


class TestScale:
    """A real programme is ~5,400 circuit ends against ~20 part numbers.

    The first version styled each row via ``ws[ws.max_row]``. ``max_row``
    rescans every cell written so far, so the write was quadratic: 33 seconds
    for one export, run on the event loop, which dropped the browser's
    connection and read to the user as the app restarting. These tests hold
    the shape of the fix rather than a stopwatch — a timing assertion would
    be flaky on a loaded machine, but a quadratic write cannot pass a ratio
    check.
    """

    def _charts(self, families: int, ends: int, parts: int = 12):
        built = []
        for f in range(families):
            pns = [f"9900{f:02d}{k:02d}AA" for k in range(parts)]
            rows = [chart.ChartRow(
                circuit=f"QK{i:05d}", cnum=f"C{i % 40:03d}",
                cavity=str(i % 20 + 1), expression="(QA1/QB2)",
                classification="variant",
                builds=pns[: (i % parts) + 1]) for i in range(ends)]
            built.append(chart.Chart(family=f"FAM_{f}", harness=f"FAM_{f}",
                                     def_id=str(70000 + f),
                                     part_numbers=pns, rows=rows))
        return built

    def test_the_sheet_is_never_rescanned_while_writing(self):
        """The defect itself, pinned.

        ``max_row``/``max_column`` walk every cell written so far. Consulting
        either one per row is what made the export quadratic, so the writer
        tracks its own row index and must not read them back at all. A timing
        assertion would be flaky on a loaded machine; this cannot be.
        """
        from openpyxl.worksheet.worksheet import Worksheet

        reads = []
        for name in ("max_row", "max_column"):
            original = getattr(Worksheet, name)

            def probe(self, _name=name, _original=original):
                reads.append(_name)
                return _original.fget(self)

            monkey = property(probe)
            setattr(Worksheet, name, monkey)
        try:
            chart.build_chart_workbook(self._charts(3, 50))
        finally:
            # restore the real descriptors
            import importlib
            importlib.reload(importlib.import_module(
                "openpyxl.worksheet.worksheet"))

        assert not reads, (
            f"the writer rescanned the sheet {len(reads)} time(s): "
            f"{sorted(set(reads))}")

    def test_a_large_chart_still_round_trips(self):
        charts = self._charts(6, 400)
        data = chart.build_chart_workbook(charts, "9000ZZ", "X9_A")
        harns, ends = read_circuit_summary(data, "big.xlsx")
        assert len(harns) == 6
        assert len(ends) == sum(len(c.rows) for c in charts)


class TestEdges:
    def test_a_family_with_no_rows_yields_an_empty_chart(self):
        rows = circuit_rows()
        entries = _entries(rows, families=("IP",))
        built = chart.build_charts(entries, [r for r in rows
                                             if r.harness_family == "BODY_LEFT"])
        assert built and built[0].rows == []

    def test_a_duplicate_dtx_row_is_written_once(self):
        """The DTx repeats a row per configuration; the chart is per end."""
        rows = circuit_rows()
        doubled = rows + [r for r in rows if r.circuit == "CKT_100"]
        built = chart.build_charts(_entries(doubled), doubled)
        body = _by_family(built)["BODY_LEFT"]
        assert sum(1 for r in body.rows if r.circuit == "CKT_100") == 1

    def test_a_row_with_no_circuit_is_skipped(self):
        rows = circuit_rows() + [CircuitRow(harness_family="IP", circuit="",
                                            sales_code="AAA", cnum="C9", pin="9")]
        built = chart.build_charts(_entries(rows), rows)
        assert all(r.circuit for c in built for r in c.rows)
