from __future__ import annotations

import io

import openpyxl
import pandas as pd

from splice.secr import db as secr_db
from splice.secr.enrich import (
    build_bulletin_numbers_for_secr,
    build_dtcr_numbers_for_secr,
    build_enrichment_summary,
    build_reason_for_change_for_secr,
    harness_family_matches,
)
from splice.secr.generate import (
    _process_circuit_sheet,
    _process_connector_sheet,
    _process_def_def_summary,
)


def _sheet_with_rows(title: str, headers: list[str], rows: list[list[str]]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    for column, header in enumerate(headers, 1):
        ws.cell(row=3, column=column, value=header)
    for row_number, values in enumerate(rows, 4):
        for column, value in enumerate(values, 1):
            ws.cell(row=row_number, column=column, value=value)
    wb.create_sheet("Summary")
    return wb


def test_affected_category_counts_and_values_are_kept_in_sync() -> None:
    connector_wb = _sheet_with_rows(
        "Connector",
        ["Action", "Connector"],
        [
            ["ADD", "A"],
            ["ADD", "B"],
            ["ADD", "A"],
            ["DELETE", "X"],
            ["ADD", "X"],
            ["CHG", "C"],
        ],
    )
    connector_summary = connector_wb["Summary"]
    _process_connector_sheet(connector_wb, connector_summary)
    assert connector_summary["B20"].value == "Added: (2)"
    assert connector_summary["C20"].value == "A, B"
    assert connector_summary["B21"].value == "Changed: (2)"
    assert connector_summary["C21"].value == "C, X"
    assert connector_summary["B22"].value == "Removed: (0)"
    assert connector_summary["C22"].value == ""

    circuit_wb = _sheet_with_rows(
        "Circuit",
        ["Action", "Circuit", "Suffix"],
        [["ADD", "100", "A"], ["CHG", "200", "B"], ["DELETE", "300", "C"]],
    )
    circuit_summary = circuit_wb["Summary"]
    _process_circuit_sheet(circuit_wb, circuit_summary)
    assert circuit_summary["B25"].value == "Added: (1)"
    assert circuit_summary["B26"].value == "Changed: (1)"
    assert circuit_summary["B27"].value == "Removed: (1)"

    part_wb = _sheet_with_rows(
        "DEF_DEF_Summary",
        ["Action", "Unused 1", "Unused 2", "Part Number"],
        [["ADD", "", "", "P1"], ["ADD", "", "", "P2"], ["DELETE", "", "", "P3"]],
    )
    part_summary = part_wb["Summary"]
    _process_def_def_summary(part_wb, part_summary)
    assert part_summary["B30"].value == "Added: (2)"
    assert part_summary["B31"].value == "Changed: (0)"
    assert part_summary["B32"].value == "Removed: (1)"


def test_multi_family_dtcr_rows_enrich_each_matching_family() -> None:
    mapping = pd.DataFrame(
        [
            {
                "DTCR#": "1001",
                "Reason for change": "Routing update Bulletin 555-01",
                "Harness Family": "IP, Body",
                "Match Method": "Device Control Number",
            },
            {
                "DTCR#": "1002",
                "Reason for change": "Jumper-only update",
                "Harness Family": "IP Jumper",
                "Match Method": "Device Control Number",
            },
        ]
    )

    assert harness_family_matches("IP, Body", "IP")
    assert harness_family_matches("ip / body", "BODY")
    assert not harness_family_matches("IP Jumper", "IP")
    assert build_dtcr_numbers_for_secr("IP", mapping) == "1001"
    assert build_reason_for_change_for_secr("IP", mapping) == (
        "1001: Routing update Bulletin 555-01"
    )
    assert build_bulletin_numbers_for_secr("IP", mapping) == "555-01"

    summary = build_enrichment_summary(mapping, "IP", "applied")
    matching_count = summary.loc[
        summary["Metric"] == "DTCRs Matching This SECR", "Value"
    ].iloc[0]
    assert matching_count == 1


def test_database_record_uses_multi_family_matching() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["I2"] = "M28RUX1A_1000"
    ws["C12"] = "IP"
    workbook_bytes = io.BytesIO()
    wb.save(workbook_bytes)
    wb.close()

    mapping = pd.DataFrame(
        [
            {"DTCR#": "1001", "Harness Family": "IP, Body"},
            {"DTCR#": "1002", "Harness Family": "Body"},
        ]
    )
    record = secr_db.record_from_workbook(
        workbook_bytes.getvalue(),
        action="create",
        dtcr_mapping_df=mapping,
    )
    assert [row["dtcr_number"] for row in record["dtcrs"]] == ["1001"]
