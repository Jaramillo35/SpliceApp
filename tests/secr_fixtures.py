"""Shared test fixtures: a structurally faithful SECR workbook.

Lifted from the Splice test-suite fixture that is validated against the real
27-file corpus, so the assistant is exercised against the same shape of data
the parser was proven on: a ``Summary`` sheet, an
``Add_Remove_Report_Summary`` holding *two* tables (connectors then circuits),
a ``Connector`` sheet, a ``Circuit`` sheet whose old values come from ``(Old)``
columns and ``"Old DEF :"`` cell comments, and a ``DEF_DEF_Summary``.
"""

from __future__ import annotations

import io
from typing import List, Optional

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

YELLOW = PatternFill("solid", fgColor="FFFFFF00")


def _write_row(ws, row: int, values: List[Optional[str]]) -> None:
    for column, value in enumerate(values, start=1):
        ws.cell(row=row, column=column, value=value)


def build_secr_workbook(
    *,
    secr_number: str = "D50319A",
    version: str = "1",
    harness_family: str = "IP",
    program: str = "RU",
    model_year: str = "2028",
) -> bytes:
    """A minimal but structurally faithful SECR workbook."""
    wb = openpyxl.Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary["I2"] = secr_number
    summary["I3"] = version
    summary["I4"] = "2026-05-07"
    summary["C7"] = "28 X1 RELEASE"
    summary["C8"] = f"SECR_28RU_X1_IP_{secr_number}_V{version}"
    summary["C10"] = model_year
    summary["C11"] = program
    summary["F10"] = "X1"
    summary["C12"] = harness_family
    summary["F11"] = "X1"
    summary["F12"] = "N"
    summary["I10"] = "M. Aguilar"
    summary["I11"] = "Ken Kopf"
    summary["I12"] = "STELLANTIS"
    summary["C14"] = "49754, 50319"
    summary["G14"] = "320767, 321299"
    summary["C20"] = "D2784J"          # devices added
    summary["C22"] = "X350A"           # devices removed
    summary["C26"] = "A937F"           # circuits changed
    summary["C27"] = "F982A"           # circuits removed

    # Two tables on one sheet, exactly as the real workbooks lay it out.
    add_remove = wb.create_sheet("Add_Remove_Report_Summary")
    _write_row(add_remove, 1, ["Connector Add/Remove Report"])
    _write_row(
        add_remove,
        2,
        [
            "SE Comment", "Action", "FCA-CNUM", "Suffix(New)", "Suffix(Old)",
            "DEF_Connector_PN", "DEF_Connector_PN(Old)", "DEF_Connector_Supplier",
        ],
    )
    _write_row(
        add_remove, 3,
        [None, "ADD", "D2784J", "MOD_CMCM", "", "D4Z080-000-B", "", "G3"],
    )
    _write_row(
        add_remove, 4,
        [None, "DELETE", "D2784J", "", "MOD_CMCM", "", "D4K10A-1D5A5-B", ""],
    )
    _write_row(
        add_remove, 5,
        ["DTCR 49919", "DELETE", "X350A", "", "INLINE_X350", "", "6098-8716", ""],
    )
    _write_row(add_remove, 6, ["Circuit Add/Remove Report"])
    _write_row(
        add_remove,
        7,
        [
            "SE Comment", "Action", "CKT NBR", "CKT Suffix", "DEF CKT COLOR",
            "DEF CKT COLOR(Old)", "DEF GAUGE", "DEF GAUGE(Old)",
        ],
    )
    _write_row(add_remove, 8, [None, "DELETE", "F982", "A", "", "PK/YE", "", "0.35"])

    connector = wb.create_sheet("Connector")
    _write_row(
        connector,
        3,
        [
            "SE Comment", "Action", "FCA-CNUM", "Suffix(New)", "Suffix(Old)",
            "DEF_Connector_PN", "DEF_Connector_PN(Old)", "DEF_Connector_Supplier",
        ],
    )
    _write_row(
        connector, 4,
        ["DTCR 50319", "COMP CHG", "SD401", "NEW_SFX", "OLD_SFX", "PN-NEW", "PN-OLD", "DZ"],
    )

    circuit = wb.create_sheet("Circuit")
    _write_row(
        circuit,
        3,
        [
            "SE Comments", "Action", "CKT NBR", "CKT Suffix", "DEF CKT COLOR",
            "DEF CKT COLOR(Old)", "DEF GAUGE", "DEF GAUGE(Old)",
            "CKT FROM (DNUM | CAV)", "Sales_Code", "Sales_Code (Old)",
            # Spelled "To", not "TO", exactly as the real workbooks write it.
            "CKT To (DNUM | CAV)",
        ],
    )
    # Old value in an explicit "(Old)" column.
    _write_row(
        circuit, 4,
        ["DTCR 49919", "CHG", "A937", "F", "GN/RD", "GN/RD", "0.50", "0.35",
         "I350X|4", "XZ4", "XZ4", "D2784J|7"],
    )
    # Old value only recoverable from the cell comment, cell marked yellow.
    # The TO cavity is a letter here — a quarter of real cavities are.
    _write_row(
        circuit, 5,
        ["DTCR's 50315, 50317", "CHG", "C205", "", "VT/BU", "VT/BU", "0.35",
         "0.35", "I350X|17", "XZ4", "XZ4", "D3840A|C"],
    )
    from_cell = circuit.cell(row=5, column=9)
    from_cell.fill = YELLOW
    from_cell.comment = Comment("Old DEF :X350A|17", "compare")

    def_def = wb.create_sheet("DEF_DEF_Summary")
    _write_row(
        def_def,
        3,
        [
            "SE Comment", "Action", "Harness PN", "DEF Symbol",
            "Harness PN (Old)", "DEF Symbol (Old)",
        ],
    )
    _write_row(def_def, 4, [None, "CHG", "68774881AB", "18", "68774881AA", "18"])

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()
