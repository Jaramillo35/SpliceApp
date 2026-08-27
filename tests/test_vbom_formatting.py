"""Workbook formatting: correctness and the scaling guarantee.

Formatting a 44-harness master workbook stalled the VBOM run at 88% because
``_style_worksheet`` walked rows as ``worksheet[n]``, which re-resolves the
sheet bounds on every access and turned the pass superlinear. These pin both
the styling result and the linear scaling, so the regression cannot return.
"""

from __future__ import annotations

import io
import time

import pytest
from openpyxl import Workbook, load_workbook

from splice.vbom.workflow import _load_vbom_module

vbom = _load_vbom_module()


def _sheet(rows: int, cols: int, filler: str = "value"):
    wb = Workbook()
    ws = wb.active
    ws.append([f"Header {c}" for c in range(cols)])
    for r in range(rows):
        ws.append([f"{filler}_{r}_{c}" for c in range(cols)])
    return wb, ws


class TestStyling:
    def test_header_and_body_are_styled(self):
        _, ws = _sheet(5, 4)
        vbom._style_worksheet(ws)
        header = ws.cell(row=1, column=1)
        assert header.font.bold and header.alignment.horizontal == "center"
        assert header.fill.fgColor.rgb.endswith("D9EAF7")
        body = ws.cell(row=3, column=2)
        assert body.alignment.vertical == "center" and body.alignment.wrap_text
        assert body.border.left.style == "thin"

    def test_freeze_panes_and_autofilter(self):
        _, ws = _sheet(5, 3)
        vbom._style_worksheet(ws)
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref == ws.dimensions

    def test_single_row_sheet_is_not_frozen_below_itself(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["only", "header", "row"])
        vbom._style_worksheet(ws)
        # openpyxl normalises a freeze at "A1" to None — nothing is frozen
        assert ws.freeze_panes is None

    def test_column_width_floor_and_cap(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["tiny", "x" * 100])          # under the floor, over the cap
        ws.append(["ab", "y" * 100])
        vbom._style_worksheet(ws)
        assert ws.column_dimensions["A"].width == 12     # 12-char floor
        assert ws.column_dimensions["B"].width == 60     # 60-char cap

    def test_width_tracks_the_longest_value(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["h"])
        ws.append(["x" * 30])
        vbom._style_worksheet(ws)
        assert ws.column_dimensions["A"].width == 32     # longest + 2

    def test_empty_cells_do_not_break_the_pass(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["a", "b", "c"])
        ws.append([None, "value", None])
        vbom._style_worksheet(ws)
        assert ws.column_dimensions["A"].width == 12


class TestScaling:
    """The bug was superlinear cost, so scaling is the thing worth pinning."""

    def test_doubling_the_rows_does_not_triple_the_time(self):
        def timed(rows: int) -> float:
            _, ws = _sheet(rows, 12)
            start = time.perf_counter()
            vbom._style_worksheet(ws)
            return time.perf_counter() - start

        small = timed(1500)
        large = timed(3000)
        if small < 0.02:                       # too fast to measure reliably
            pytest.skip("styling too fast on this machine to compare ratios")
        # Linear would be ~2x. The old implementation was ~3-4x and worsening;
        # 2.6x leaves headroom for noise while still catching a regression.
        assert large / small < 2.6, (
            f"formatting is scaling superlinearly: {small:.3f}s -> {large:.3f}s")


class TestFormatWorkbookProgress:
    def _saved(self, tmp_path, sheets: int):
        wb = Workbook()
        wb.remove(wb.active)
        for i in range(sheets):
            ws = wb.create_sheet(f"Sheet{i}")
            ws.append(["a", "b"])
            ws.append(["1", "2"])
        path = tmp_path / "book.xlsx"
        wb.save(path)
        return path

    def test_reports_once_per_sheet(self, tmp_path):
        path = self._saved(tmp_path, 5)
        seen = []
        vbom.format_workbook_output(path, lambda i, t, title: seen.append((i, t, title)))
        assert [i for i, _, _ in seen] == [1, 2, 3, 4, 5]
        assert {t for _, t, _ in seen} == {5}
        assert seen[0][2] == "Sheet0"

    def test_callback_is_optional(self, tmp_path):
        path = self._saved(tmp_path, 2)
        vbom.format_workbook_output(path)        # must not raise
        assert load_workbook(path)["Sheet0"].freeze_panes == "A2"

    def test_callback_errors_never_lose_the_workbook(self, tmp_path):
        # A dead browser session must not cost the user the finished workbook.
        path = self._saved(tmp_path, 2)

        def explode(index, total, title):
            raise RuntimeError("client gone")

        vbom.format_workbook_output(path, explode)
        assert load_workbook(path)["Sheet1"].cell(row=1, column=1).font.bold


class TestSelectionsWorkbookSheets:
    """Global_Code_Overview was dropped from the selections workbook."""

    def test_global_code_overview_is_not_written(self):
        import inspect
        from splice.vbom.workflow import run_vbom_workflow
        source = inspect.getsource(run_vbom_workflow)
        assert 'sheet_name="Global_Code_Overview"' not in source
        # the per-family stats sheet it summarised is still emitted
        assert 'sheet_name="Family_Code_Stats"' in source


class TestGuide:
    """The instructions ship with the bundle, so they must stay accurate."""

    def test_guide_names_every_output_file(self):
        from splice.vbom.guide import GUIDE_MD
        for stem in ("Harness_Selection_Review", "VIN_to_Harness_Selection",
                     "VIN_Salescode_matrix", "Master_Combined_Harness_Complexity",
                     "VBOM_Template_for_DEFE"):
            assert stem in GUIDE_MD, f"{stem} is not documented"

    def test_guide_names_every_review_reason(self):
        # The reasons the engine can raise must all be explained.
        from splice.vbom.guide import GUIDE_MD
        for reason in ("No complete PN covers every required sales code",
                       "Multiple PNs share the best score",
                       "N/A conflicts with an available base/default PN"):
            assert reason in GUIDE_MD, f"unexplained review reason: {reason}"

    def test_guide_documents_the_sheets_that_are_written(self):
        from splice.vbom.guide import GUIDE_MD
        for sheet in ("Selections", "AllCandidates", "Excluded_SalesCodes",
                      "Final_BOM_By_VIN", "Family_Code_Stats", "SalesCode_Diff"):
            assert sheet in GUIDE_MD
        # ...and not the one that was removed
        assert "Global_Code_Overview" not in GUIDE_MD

    def test_bundle_readme_is_stamped_with_the_run(self):
        from splice.vbom.guide import bundle_readme
        text = bundle_readme("28_DT", "28_DT_VBOM_Template_for_DEFE.xlsx", 7)
        assert "28_DT" in text and "7 selection(s)" in text
        assert "28_DT_VBOM_Template_for_DEFE.xlsx" in text

    def test_bundle_readme_handles_a_clean_run(self):
        from splice.vbom.guide import bundle_readme
        text = bundle_readme("28_DT", "x.xlsx", 0)
        assert "flagged **no** selections" in text
