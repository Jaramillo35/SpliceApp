"""The applicability export, and the cleanup notes it carries.

A ticked row must reach the workbook as a written instruction — the point of
the Complexity Cleanup Notes column is that someone can act on it without the
workbench open.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

from fixtures_dtxcircuits import circuit_rows, harnesses  # noqa: E402
from splice.dtxcircuits import analyze, report  # noqa: E402


@pytest.fixture()
def entries():
    return [report.Entry(label=f"{a.harness} → {a.harness}", family=a.harness,
                         filename=f"{a.harness}.xlsm", analysis=a)
            for a in analyze(circuit_rows(), harnesses())]


@pytest.fixture()
def body_left(entries):
    return next(e for e in entries if e.family == "BODY_LEFT")


class TestItemKey:
    def test_key_is_stable_and_distinguishes_kinds(self):
        a = report.item_key("IP", "IP", report.KIND_CIRCUIT, "CKT_1")
        b = report.item_key("IP", "IP", report.KIND_CONNECTOR, "CKT_1")
        assert a == report.item_key("IP", "IP", report.KIND_CIRCUIT, "CKT_1")
        assert a != b

    def test_same_ident_on_two_harnesses_is_two_keys(self):
        # one DTx family may map to several harnesses; a circuit selected on
        # one must not silently select it on the other
        assert report.item_key("IP", "LEFT", "circuit", "CKT_1") != \
            report.item_key("IP", "RIGHT", "circuit", "CKT_1")


class TestNotes:
    def test_never_built_note_states_the_verdict_and_the_condition(self, body_left):
        s = report.selection_for(body_left, report.KIND_CIRCUIT, "CKT_500")
        assert "Never built" in s.note
        assert "(CCC)" in s.note
        assert s.verdict == "never built"

    def test_untracked_note_names_the_code_and_the_fix(self, body_left):
        s = report.selection_for(body_left, report.KIND_CIRCUIT, "CKT_600")
        assert "ZZZ" in s.note
        assert "Remove ZZZ" in s.note, "the note must say what to do about it"

    def test_gap_note_is_one_sentence_about_the_code(self, body_left):
        """The circuits resting on the code live in the sheet's own column;
        repeating them in the note is what made the old one unreadable."""
        s = report.selection_for(body_left, report.KIND_GAP, "ZZZ")
        assert s.note == ("Sales code ZZZ is in the DTx report but not in the "
                          "complexity file.")
        assert s.verdict == "sales-code gap"

    def test_the_gaps_sheet_still_carries_the_circuits(self, entries):
        """Dropping them from the note must not drop them from the report."""
        data = report.build_report(entries, {})
        ws = load_workbook(io.BytesIO(data))["Sales-code gaps"]
        headers = [c.value for c in ws[1]]
        col = headers.index("Circuits")
        code = headers.index("Sales code")
        found = {row[code].value: row[col].value for row in ws.iter_rows(min_row=2)}
        assert "CKT_600" in (found.get("ZZZ") or "")

    def test_variant_note_reports_the_split(self, body_left):
        s = report.selection_for(body_left, report.KIND_CIRCUIT, "CKT_200")
        assert "2 of 4 builds" in s.note

    def test_unknown_item_returns_none(self, body_left):
        assert report.selection_for(body_left, report.KIND_CIRCUIT, "NOPE") is None
        assert report.selection_for(body_left, "nonsense", "CKT_500") is None


class TestWorkbook:
    def _book(self, entries, cleanup):
        data = report.build_report(entries, cleanup, dtx_program="2030QX",
                                   dtx_phase="V1_A")
        return load_workbook(io.BytesIO(data))

    def test_every_sheet_is_present(self, entries):
        wb = self._book(entries, {})
        assert wb.sheetnames == ["Read Me", "Circuits", "Connectors",
                                 "Sales-code gaps", "Sales-code repairs",
                                 "Complexity Cleanup", "Customer email"]

    def test_cleanup_column_is_named_exactly(self, entries):
        wb = self._book(entries, {})
        for sheet in ("Circuits", "Connectors", "Sales-code gaps"):
            headers = [c.value for c in wb[sheet][1]]
            assert report.CLEANUP_COLUMN in headers, sheet
        assert report.CLEANUP_COLUMN == "Complexity Cleanup Notes"

    def test_unselected_rows_leave_the_column_empty(self, entries):
        wb = self._book(entries, {})
        ws = wb["Circuits"]
        col = [c.value for c in ws[1]].index(report.CLEANUP_COLUMN)
        assert all(row[col].value in (None, "") for row in ws.iter_rows(min_row=2))

    def test_a_selected_row_carries_its_note(self, entries, body_left):
        s = report.selection_for(body_left, report.KIND_CIRCUIT, "CKT_500")
        wb = self._book(entries, {s.key: s})
        ws = wb["Circuits"]
        headers = [c.value for c in ws[1]]
        col = headers.index(report.CLEANUP_COLUMN)
        ident = headers.index("Circuit")
        notes = {row[ident].value: row[col].value
                 for row in ws.iter_rows(min_row=2)}
        assert notes["CKT_500"] and "Never built" in notes["CKT_500"]
        assert not notes["CKT_100"], "an unticked row must stay empty"

    def test_cleanup_sheet_lists_only_the_selection(self, entries, body_left):
        picks = {}
        for kind, ident in ((report.KIND_CIRCUIT, "CKT_500"),
                            (report.KIND_GAP, "ZZZ")):
            s = report.selection_for(body_left, kind, ident)
            picks[s.key] = s
        ws = self._book(entries, picks)["Complexity Cleanup"]
        head = report.CLEANUP_HEADERS
        rows = [r for r in ws.iter_rows(min_row=7, values_only=True) if r[0]]
        assert len(rows) == 2
        assert {r[head.index("Item")] for r in rows} == {"CKT_500", "ZZZ"}
        types = {r[head.index("Type")] for r in rows}
        assert types == {"circuit", "sales code"}, "no engine words in the sheet"

    def test_empty_selection_still_produces_a_valid_workbook(self, entries):
        ws = self._book(entries, {})["Complexity Cleanup"]
        assert ws.max_row == 6          # the band and the header, no rows
        assert [c.value for c in ws[6]] == report.CLEANUP_HEADERS
        assert "nothing outstanding" in ws["A3"].value

    def test_read_me_states_the_programme(self, entries):
        ws = self._book(entries, {})["Read Me"]
        text = " ".join(str(r[0]) for r in ws.iter_rows(values_only=True) if r[0])
        assert "2030QX" in text and "V1_A" in text
        assert report.CLEANUP_COLUMN in text

    def test_family_and_harness_are_both_reported(self, entries):
        # a family mapped to several harnesses needs both to tell rows apart
        ws = self._book(entries, {})["Circuits"]
        headers = [c.value for c in ws[1]]
        assert headers[0] == "DTx family" and headers[1] == "Harness"


class TestBothExpressions:
    """The export must show what the DTx said AND what the SE decided.

    Without the original, a repaired circuit reads as though the DTx always
    said the corrected thing, and there is no record of the decision that
    moved its verdict.
    """

    def _entry_with_repair(self):
        analysis = analyze(circuit_rows(), harnesses())[0]
        return report.Entry(
            label="BODY_LEFT", family="BODY_LEFT", filename="bl.xlsm",
            analysis=analysis,
            original_circuit_conditions={"CKT_200": "(AAA-BBB)"},
            original_cnum_conditions={"C2": "(AAA-BBB)"})

    def _book(self, entries, repairs=None, context=None):
        data = report.build_report(entries, {}, dtx_program="2030QX",
                                   dtx_phase="V1_A", repairs=repairs,
                                   repair_context=context)
        return load_workbook(io.BytesIO(data))

    def test_circuits_carry_both_conditions(self):
        ws = self._book([self._entry_with_repair()])["Circuits"]
        headers = [c.value for c in ws[1]]
        assert "Condition as in DTx" in headers
        assert "Condition as decided" in headers
        assert "Repaired" in headers

    def test_a_repaired_row_is_flagged_and_shows_both(self):
        ws = self._book([self._entry_with_repair()])["Circuits"]
        headers = [c.value for c in ws[1]]
        i_ckt = headers.index("Circuit")
        i_was = headers.index("Condition as in DTx")
        i_now = headers.index("Condition as decided")
        i_flag = headers.index("Repaired")
        row = next(r for r in ws.iter_rows(min_row=2, values_only=True)
                   if r[i_ckt] == "CKT_200")
        assert row[i_was] == "(AAA-BBB)"
        assert row[i_now] == "(AAA)"
        assert row[i_flag] == "yes"

    def test_an_untouched_row_is_not_flagged(self):
        ws = self._book([self._entry_with_repair()])["Circuits"]
        headers = [c.value for c in ws[1]]
        i_ckt, i_flag = headers.index("Circuit"), headers.index("Repaired")
        row = next(r for r in ws.iter_rows(min_row=2, values_only=True)
                   if r[i_ckt] == "CKT_300")
        assert not row[i_flag]

    def test_connectors_carry_both_conditions_too(self):
        ws = self._book([self._entry_with_repair()])["Connectors"]
        headers = [c.value for c in ws[1]]
        assert "Condition as in DTx" in headers and "Condition as decided" in headers

    def test_the_repairs_sheet_lists_each_decision_once(self):
        wb = self._book([self._entry_with_repair()],
                        repairs={"AAA-BBB": "AAA&-BBB"},
                        context={"AAA-BBB": {"kind": "missing operator", "rows": 3,
                                             "families": ["BODY_LEFT"],
                                             "circuits": ["CKT_200"]}})
        ws = wb["Sales-code repairs"]
        assert [c.value for c in ws[1]][:3] == ["Expression as in DTx",
                                                "Expression as decided", "Problem"]
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(rows) == 1
        assert rows[0][0] == "AAA-BBB" and rows[0][1] == "AAA&-BBB"
        assert rows[0][2] == "missing operator" and rows[0][3] == 3
        assert "CKT_200" in rows[0][5]

    def test_no_repairs_still_produces_the_sheet(self):
        ws = self._book([self._entry_with_repair()])["Sales-code repairs"]
        assert ws.max_row == 1          # header only

    def test_read_me_explains_the_two_columns(self):
        ws = self._book([self._entry_with_repair()])["Read Me"]
        text = " ".join(str(r[0]) for r in ws.iter_rows(values_only=True) if r[0])
        assert "as in DTx" in text and "as decided" in text


class TestWorkList:
    """The cleanup sheet is a work list: every row says what to change and
    which document to change it in, because that is the SE's first question
    and the thing an email to the customer has to state."""

    def _rows(self, entries, cleanup, **kwargs):
        work = report.cleanup_rows(cleanup, entries, kwargs.get("repairs"),
                                   kwargs.get("context"))
        return {r.ident: r for r in work}, work

    def test_a_contradiction_the_complexity_can_see_is_a_dtx_fix(self, entries, body_left):
        """CKT_500 is conditioned on CCC, which BODY_LEFT's complexity tracks;
        no build satisfies it, so the DTx wrote a condition that cannot hold."""
        s = report.selection_for(body_left, report.KIND_CIRCUIT, "CKT_500",
                                 tracked_elsewhere=report.tracked_codes(entries))
        assert s.priority == report.P_BLOCKER
        assert s.fix_in == report.FIX_DTX
        assert "Correct" in s.action and "CCC" in s.action
        assert "no build satisfies it" in s.action

    def test_a_code_missing_from_the_complexity_is_a_question_for_the_customer(
            self, entries, body_left):
        """ZZZ is in no complexity file, so the data cannot say whether the
        DTx invented it or every matrix is missing it — the action asks."""
        s = report.selection_for(body_left, report.KIND_GAP, "ZZZ",
                                 tracked_elsewhere=report.tracked_codes(entries))
        assert s.fix_in == report.FIX_EITHER
        assert "Confirm whether" in s.action and "ZZZ" in s.action

    def test_a_code_another_harness_tracks_is_a_complexity_fix(self, entries, body_left):
        """The same gap becomes unambiguous once a sibling file tracks the
        code: the code is real, so this file is the one missing it."""
        s = report.selection_for(body_left, report.KIND_GAP, "ZZZ",
                                 tracked_elsewhere={"ZZZ"})
        assert s.fix_in == report.FIX_COMPLEXITY
        assert "already tracks it" in s.action

    def test_a_confirmed_repair_is_exported_as_a_dtx_correction(self, entries):
        """The workbench applied the repair locally; unless the customer fixes
        the source it comes back with the next export."""
        _, work = self._rows(
            entries, {}, repairs={"QB1-QA1": "QB1&-QA1"},
            context={"QB1-QA1": {"kind": "missing operator", "rows": 1,
                                 "families": ["IP"], "circuits": ["QK109"]}})
        repair = next(r for r in work if r.kind == report.KIND_EXPRESSION)
        assert repair.fix_in == report.FIX_DTX
        assert repair.priority == report.P_BLOCKER
        assert repair.action == 'Correct "QB1-QA1" to "QB1&-QA1" in the DTx'
        assert "QK109" in repair.evidence

    def test_the_worst_item_is_the_first_row(self, entries, body_left):
        picks = {}
        for kind, ident in ((report.KIND_GAP, "ZZZ"),
                            (report.KIND_CIRCUIT, "CKT_500")):
            s = report.selection_for(body_left, kind, ident)
            picks[s.key] = s
        _, work = self._rows(entries, picks)
        assert work[0].priority == report.P_BLOCKER
        assert [r.priority for r in work] == sorted(
            (r.priority for r in work),
            key=lambda p: report.PRIORITY_ORDER[p])

    def test_every_row_carries_an_action_and_a_place_to_fix_it(self, entries, body_left):
        picks = {}
        for kind, ident in ((report.KIND_CIRCUIT, "CKT_500"),
                            (report.KIND_CIRCUIT, "CKT_600"),
                            (report.KIND_CIRCUIT, "CKT_200"),
                            (report.KIND_GAP, "ZZZ")):
            s = report.selection_for(body_left, kind, ident)
            picks[s.key] = s
        _, work = self._rows(entries, picks)
        for row in work:
            assert row.action, row.ident
            assert row.fix_in in report.FIX_ORDER, row.ident
            assert row.priority in report.PRIORITY_ORDER, row.ident

    def test_a_tick_from_another_run_keeps_its_instruction(self, body_left):
        """Restored ticks are re-prescribed when the item is in this run and
        keep their stored instruction when it is not."""
        from splice.dtxcircuits import store
        s = report.selection_for(body_left, report.KIND_CIRCUIT, "CKT_500")
        restored = store.restore_cleanup(store.remember_cleanup({s.key: s}))[s.key]
        assert restored.action == s.action and restored.fix_in == s.fix_in
        _, work = self._rows([], {s.key: restored})
        assert work[0].action == s.action, "no entry this run, so keep the note"


class TestCustomerEmail:
    def test_the_draft_groups_by_where_the_change_belongs(self, entries, body_left):
        picks = {}
        for kind, ident in ((report.KIND_CIRCUIT, "CKT_500"),
                            (report.KIND_GAP, "ZZZ")):
            s = report.selection_for(body_left, kind, ident)
            picks[s.key] = s
        work = report.cleanup_rows(picks, entries)
        lines = report.email_lines(work, dtx_program="2030QX", dtx_phase="V1_A",
                                   prepared_by="M. Jaramillo")
        text = "\n".join(lines)
        assert lines[0].startswith("Subject: 2030QX V1_A —")
        assert "correction(s) requested" in lines[0]
        assert "IN THE DTx REPORT (1)" in text
        assert "EITHER DOCUMENT" in text
        assert "1. [Blocker]" in text and "2. [High]" in text
        assert text.rstrip().endswith("M. Jaramillo")
        assert "CKT_500" in text and "ZZZ" in text

    def test_a_clean_review_says_so_rather_than_sending_an_empty_list(self, entries):
        lines = report.email_lines([], dtx_program="2030QX", dtx_phase="V1_A")
        text = "\n".join(lines)
        assert "Nothing outstanding" in text
        assert "<your name>" in text

    def test_the_sheet_carries_the_draft(self, entries, body_left):
        s = report.selection_for(body_left, report.KIND_CIRCUIT, "CKT_500")
        data = report.build_report(entries, {s.key: s}, dtx_program="2030QX",
                                   dtx_phase="V1_A", prepared_by="M. Jaramillo")
        ws = load_workbook(io.BytesIO(data))["Customer email"]
        text = "\n".join(str(r[0]) for r in ws.iter_rows(values_only=True) if r[0])
        assert "Subject: 2030QX V1_A" in text
        assert "M. Jaramillo" in text


def test_no_internal_kind_word_reaches_the_reader():
    """'gap' is what the engine calls it; the customer reads 'sales code'."""
    from splice.dtxcircuits import report as r
    assert r.kind_label(r.KIND_GAP) == "sales code"
    assert set(r.KIND_LABEL) == {r.KIND_CIRCUIT, r.KIND_CONNECTOR, r.KIND_GAP,
                                 r.KIND_EXPRESSION}
