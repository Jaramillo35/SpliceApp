"""Holistic health check: window coverage, route gaps, gates, baseline.

Every scenario is a miniature of a defect found (or ruled out) in the 28RU X1
Body Left / Body Right validation on 2026-08-22:

* R732 — a one-sided option window with real builds on the wireless side.
* F984 — a superset OR on one side that provably covers both variants
  (must auto-clear, with the proof kept).
* A960 — a circuit live on a harness in a window at some inlines but absent
  at another crossing in that same window (route gap).
"""

from __future__ import annotations

from splice.inline.complexity import read_complexity, satisfying_builds
from splice.inline.pairing import resolve
from splice.inline.summary import read_circuit_summary
from splice.inline import health
from splice.inline.health import (
    SEV_BLOCKER, SEV_HIGH, SEV_REVIEW,
    analyze, builds_where, disposition, load_baseline, render_report,
    save_baseline, sign_off, union_expression, window_minus,
)

from tests.test_inline_continuity import build_complexity, build_summary


def _harness(def_id: str, codes, builds, filename: str):
    payload = build_complexity(def_id, list(codes), builds)
    harness = read_complexity(payload, filename)
    harness.complexity_file = filename
    return harness


def _analyze(left_ends, right_ends, left_builds, right_builds,
             codes=("CG3", "CYF", "CY3", "CYC"),
             left_file="cx_left_01-10-2026.xlsm",
             right_file="cx_right_01-10-2026.xlsm"):
    complexity = {
        "100": _harness("100", codes, left_builds, left_file),
        "200": _harness("200", codes, right_builds, right_file),
    }

    def stamp(ends, def_id):
        harness = complexity[def_id]
        for end in ends:
            end.setdefault("builds", [
                b.part_number
                for b in satisfying_builds(harness, end.get("sc", ""))
            ])
        return [b.part_number for b in harness.builds]

    lp = stamp(left_ends, "100")
    rp = stamp(right_ends, "200")
    payload = build_summary([("Left", "100", lp, left_ends),
                             ("Right", "200", rp, right_ends)])
    harnesses, ends = read_circuit_summary(payload)
    pairs, unmated = resolve(ends, set(harnesses))
    return analyze(harnesses, ends, complexity, pairs, unmated)


LEFT_BUILDS = [("L-CYF", {"CG3", "CYF"}), ("L-CY3", {"CG3", "CY3"})]
RIGHT_BUILDS = [("R-CYF", {"CG3", "CYF"}), ("R-CY3", {"CG3", "CY3"})]


class TestWindowCoverage:
    def test_one_sided_window_with_builds_is_a_blocker(self):
        # R732: left covers all of CG3, right omits the CY3 case
        result = _analyze(
            [{"circuit": "R732", "cnum": "X10A", "cav": "8", "sc": "CG3"}],
            [{"circuit": "R732", "cnum": "Y10A", "cav": "8", "sc": "CG3&(CYC/CYF)"}],
            LEFT_BUILDS, RIGHT_BUILDS,
        )
        blockers = [f for f in result.findings
                    if f.severity == SEV_BLOCKER and f.kind == "one_sided_window"]
        assert len(blockers) == 1
        f = blockers[0]
        assert f.cavity == "8" and "R732" in f.circuit
        assert f.harness_without == "Right"
        assert "R-CY3" in f.builds_without  # the concrete affected build

    def test_superset_or_auto_clears_with_proof(self):
        # F984: right's single OR wire covers both left variants
        result = _analyze(
            [{"circuit": "F984", "suffix": "A", "cnum": "X10A", "cav": "3", "sc": "CG3&CYF"},
             {"circuit": "F984", "suffix": "G", "cnum": "X10A", "cav": "3", "sc": "CG3&-CYF"}],
            [{"circuit": "F984", "suffix": "AB", "cnum": "Y10A", "cav": "3", "sc": "CG3"}],
            LEFT_BUILDS, RIGHT_BUILDS,
        )
        assert not [f for f in result.findings if f.kind == "one_sided_window"]
        assert any(p.cavity == "3" for p in result.cleared)

    def test_unconditional_wire_covers_everything(self):
        result = _analyze(
            [{"circuit": "G74", "cnum": "X10A", "cav": "1", "sc": "CYF"}],
            [{"circuit": "G74", "cnum": "Y10A", "cav": "1", "sc": ""}],  # 100% wire
            LEFT_BUILDS, RIGHT_BUILDS,
        )
        one_sided = [f for f in result.findings if f.kind == "one_sided_window"]
        # right's unconditional wire covers left; the reverse window (right has
        # a wire where left has none) correctly flags left's CY3 builds
        assert all(f.harness_without == "Left" for f in one_sided)

    def test_missing_continuation_is_a_blocker(self):
        result = _analyze(
            [{"circuit": "A900", "cnum": "X10A", "cav": "5", "sc": "CG3"}],
            [{"circuit": "Z1", "cnum": "Y10A", "cav": "6", "sc": "CG3"}],
            LEFT_BUILDS, RIGHT_BUILDS,
        )
        cavity_blockers = [f for f in result.findings
                           if f.kind == "cavity" and f.severity == SEV_BLOCKER]
        assert cavity_blockers  # occupied on one side, empty on the other


class TestRouteCompleteness:
    def test_a960_signature_is_a_high_finding(self):
        # A960 crosses two inlines on Left: broad window at X20A, narrow at X21A
        result = _analyze(
            [{"circuit": "A960", "cnum": "X20A", "cav": "1", "sc": "CYF/CY3"},
             {"circuit": "A960", "cnum": "X21A", "cav": "1", "sc": "CYF"}],
            [{"circuit": "A960", "cnum": "Y20A", "cav": "1", "sc": "CYF/CY3"},
             {"circuit": "A960", "cnum": "Y21A", "cav": "1", "sc": "CYF"}],
            LEFT_BUILDS, RIGHT_BUILDS,
        )
        gaps = [f for f in result.findings if f.kind == "route_window_gap"]
        assert any(f.inline == "X21A" and f.circuit == "A960"
                   and f.severity == SEV_HIGH for f in gaps)
        gap = next(f for f in gaps if f.inline == "X21A")
        assert "L-CY3" in f"{gap.builds_with}"  # the affected build is named


class TestExpressionAlgebra:
    def test_union_expression(self):
        class E:
            def __init__(self, sc): self.sales_code = sc
        assert union_expression([E("A"), E("B")]) == "(A)/(B)"
        assert union_expression([E("A"), E("")]) is None   # unconditional
        assert union_expression([]) is None

    def test_window_minus(self):
        assert window_minus("A", None) is None
        assert window_minus(None, "B") == "-(B)"
        assert window_minus("A", "B") == "(A)&-(B)"

    def test_unknown_codes_are_treated_as_present(self):
        h = _harness("300", ["CG3"], [("P1", {"CG3"})], "cx.xlsm")
        # ZZZ is not in the vocabulary: silence must not manufacture findings
        assert len(builds_where(h, "CG3&ZZZ")) == 1
        assert len(builds_where(h, None)) == 1


class TestGates:
    def _one_blocker(self):
        return _analyze(
            [{"circuit": "R732", "cnum": "X10A", "cav": "8", "sc": "CG3"}],
            [{"circuit": "R732", "cnum": "Y10A", "cav": "8", "sc": "CG3&(CYC/CYF)"}],
            LEFT_BUILDS, RIGHT_BUILDS,
        )

    def test_fingerprints_are_stable_across_runs(self):
        a = self._one_blocker().findings[0].fingerprint
        b = self._one_blocker().findings[0].fingerprint
        assert a == b and len(a) == 16

    def test_disposition_closes_a_finding(self, tmp_path):
        result = self._one_blocker()
        finding = result.findings[0]
        baseline = load_baseline(tmp_path / "baseline.json")
        assert result.blocking_open(baseline)  # open before disposition
        disposition(baseline, finding, "Accepted variant", "covered by spare", "SE")
        save_baseline(tmp_path / "baseline.json", baseline)
        reloaded = load_baseline(tmp_path / "baseline.json")
        assert finding.fingerprint in reloaded["dispositions"]
        assert not [f for f in result.open_findings(reloaded)
                    if f.fingerprint == finding.fingerprint]

    def test_sign_off_is_recorded(self, tmp_path):
        baseline = load_baseline(tmp_path / "b.json")
        sign_off(baseline, "Martin", "X1 release")
        assert baseline["signoffs"][0]["by"] == "Martin"

    def test_input_report_flags_skew_and_gaps(self):
        result = _analyze(
            [{"circuit": "G74", "cnum": "X10A", "cav": "1", "sc": ""}],
            [{"circuit": "G74", "cnum": "Y10A", "cav": "1", "sc": ""}],
            LEFT_BUILDS, RIGHT_BUILDS,
            left_file="cx_left_01-10-2026.xlsm",
            right_file="cx_right_03-15-2026.xlsm",
        )
        assert result.inputs.skew_days > 30
        assert result.inputs.has_issues
        assert result.inputs.missing_complexity == []


class TestReport:
    def test_report_workbook_has_all_sheets(self, tmp_path):
        import io
        import openpyxl
        result = _analyze(
            [{"circuit": "R732", "cnum": "X10A", "cav": "8", "sc": "CG3"}],
            [{"circuit": "R732", "cnum": "Y10A", "cav": "8", "sc": "CG3&(CYC/CYF)"}],
            LEFT_BUILDS, RIGHT_BUILDS,
        )
        baseline = load_baseline(tmp_path / "b.json")
        disposition(baseline, result.findings[0], "Defect", "file SECR", "SE")
        data = render_report(result, baseline)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert set(wb.sheetnames) == {"Read Me", "Findings", "Scorecard",
                                      "Inputs", "Cleared proofs", "Sign-off"}
        # the reviewer guide ships inside the report itself
        readme = " ".join(str(c.value) for row in wb["Read Me"].iter_rows()
                          for c in row if c.value)
        assert "Technical considerations" in readme
        assert "minimized against the BUILDABLE" in readme
        rows = list(wb["Findings"].iter_rows(values_only=True))
        assert rows[1][11] == "Defect"  # disposition column round-trips


class TestDeduplication:
    def test_identical_fingerprints_collapse_to_one_finding(self):
        # Streamlit renders one widget set per finding; duplicate fingerprints
        # crashed the page with StreamlitDuplicateElementKey (field report).
        result = _analyze(
            [{"circuit": "R732", "cnum": "X10A", "cav": "8", "sc": "CG3"}],
            [{"circuit": "R732", "cnum": "Y10A", "cav": "8", "sc": "CG3&(CYC/CYF)"}],
            LEFT_BUILDS, RIGHT_BUILDS,
        )
        prints = [f.fingerprint for f in result.findings]
        assert len(prints) == len(set(prints))


class TestPresentation:
    """Field reports 2026-08-24: names and within-harness rendering."""

    def test_findings_wear_summary_names_not_filenames(self):
        # complexity harnesses are named after their uploaded FILE; every
        # finding (all layers) and the study must show the summary's names
        result = _analyze(
            [{"circuit": "R732", "cnum": "X10A", "cav": "8", "sc": "CG3"}],
            [{"circuit": "R732", "cnum": "Y10A", "cav": "8", "sc": "CG3&(CYC/CYF)"}],
            LEFT_BUILDS, RIGHT_BUILDS,
        )
        names = {f.harness_with for f in result.findings} \
            | {f.harness_without for f in result.findings} \
            | {f.harness_a for f in result.study.findings} \
            | {f.harness_b for f in result.study.findings}
        leaked = {n for n in names if "cx_" in str(n) or ".xlsm" in str(n)}
        assert not leaked, f"complexity filenames leaked into findings: {leaked}"

    def test_route_gap_reports_as_within_one_harness(self):
        # a route gap lives inside ONE harness; the report must not print the
        # harness on both sides ("BODY LEFT ↔ BODY LEFT" misread as an inline)
        import io
        from openpyxl import load_workbook
        result = _analyze(
            [{"circuit": "A960", "cnum": "X20A", "cav": "1", "sc": "CYF/CY3"},
             {"circuit": "A960", "cnum": "X21A", "cav": "1", "sc": "CYF"}],
            [{"circuit": "A960", "cnum": "Y20A", "cav": "1", "sc": "CYF/CY3"},
             {"circuit": "A960", "cnum": "Y21A", "cav": "1", "sc": "CYF"}],
            LEFT_BUILDS, RIGHT_BUILDS,
        )
        gap = next(f for f in result.findings if f.kind == "route_window_gap")
        assert gap.within_harness
        assert gap.crossings and gap.inline not in gap.crossings

        wb = load_workbook(io.BytesIO(render_report(result, {"dispositions": {}})))
        rows = list(wb["Findings"].iter_rows(values_only=True))
        hdr = rows[0]
        for r in rows[1:]:
            if r[hdr.index("Kind")] == "route_window_gap":
                missing_on = str(r[hdr.index("Missing on")])
                has_wire = str(r[hdr.index("Has wire")])
                assert missing_on != has_wire
                assert "within" in missing_on


class TestConflictedCavityMerge:
    def test_inconsistent_cavity_yields_one_complete_finding(self):
        """Field reports (L206/N0 @ X103A cav 14): first double-reported,
        then suppressed the window — the fix is ONE finding carrying the
        identity conflict AND the merged coverage picture."""
        result = _analyze(
            [{"circuit": "N0", "cnum": "X10A", "cav": "14", "sc": "CG3"}],
            [{"circuit": "L206", "cnum": "Y10A", "cav": "14",
              "sc": "CG3&(CYC/CYF)"}],
            LEFT_BUILDS, RIGHT_BUILDS,
        )
        at_cavity = [f for f in result.findings if f.cavity == "14"]
        assert len(at_cavity) == 1
        f = at_cavity[0]
        assert f.kind == "cavity" and f.severity == SEV_BLOCKER
        assert "Coverage also differs" in f.detail
        assert f.builds_without and "R-CY3" in f.builds_without
        assert f.window_short  # merged window shown, raw window stays ""
        assert f.window == ""  # fingerprint basis unchanged


class TestIntegrityGrouping:
    def test_rows_sharing_a_root_cause_collapse_to_one_finding(self):
        """Field report (R5/A942B/P968): one truncated part number in the
        summary columns flagged every row marking that build — group them."""
        result = _analyze(
            [{"circuit": "R5", "cnum": "X10A", "cav": "1", "sc": "",
              "builds": ["L-CYF"]},
             {"circuit": "A942", "cnum": "X10A", "cav": "2", "sc": "",
              "builds": ["L-CYF"]}],
            [{"circuit": "R5", "cnum": "Y10A", "cav": "1", "sc": ""},
             {"circuit": "A942", "cnum": "Y10A", "cav": "2", "sc": ""}],
            LEFT_BUILDS, RIGHT_BUILDS,
        )
        groups = [f for f in result.findings if f.kind == "integrity"]
        assert len(groups) == 1
        g = groups[0]
        assert "2 row(s)" in g.circuit
        assert "A942" in g.circuit and "R5" in g.circuit
        assert "L-CY3" in g.detail  # the expression-only build is named

    def test_truncation_hint_names_the_twin_pns(self):
        from splice.inline.health import _truncation_hint
        hint = _truncation_hint(("687894643A",), ("687894643AA",))
        assert "687894643A" in hint and "687894643AA" in hint
        assert "truncated" in hint
        assert _truncation_hint(("AAA",), ("BBB",)) == ""
