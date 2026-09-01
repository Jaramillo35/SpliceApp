"""DTx data quality — the numbers the customer is shown, on invented data.

The point of the dashboard is that a finding must be *defensible*: the SE
builds the complexity files from what the customer supplied, so a mismatch is
the customer's own data disagreeing with itself. These tests hold the line on
which things count as findings and which are only context.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent))

from fixtures_dtxcircuits import circuit_rows, harnesses  # noqa: E402

from splice.dtxcircuits import analyze_harness, integrity, quality, report
from splice.dtxcircuits.models import CircuitRow, DtxMeta


META = DtxMeta(program="9000ZZ", phase="X9_A", report_date="Jan-01-9000")


def _entries(rows=None, families=("BODY_LEFT", "IP")):
    rows = circuit_rows() if rows is None else rows
    books = harnesses()
    out = []
    for family in families:
        harness = books[family]
        family_rows = [r for r in rows if r.harness_family == family]
        out.append(report.Entry(
            label=f"{family} → {harness.name}", family=family,
            filename=f"{family}.xlsm",
            analysis=analyze_harness(family_rows, harness,
                                     harness_name=harness.name)))
    return out


class TestStructure:
    def test_counts_what_the_dtx_contains(self):
        rows = circuit_rows()
        q = quality.assess(rows, META)
        assert q.program == "9000ZZ" and q.phase == "X9_A"
        assert q.rows == len(rows)
        # CKT_700 and CKT_800 each occupy two rows
        assert q.circuits == 11
        assert q.families == 3                      # BODY_LEFT, IP, DASH
        assert q.conditioned_rows + q.unconditional_rows == q.rows
        assert 0.0 < q.conditioned_share < 1.0

    def test_structure_alone_is_not_a_finding(self):
        """A DTx with blanks but nothing contradictory is reported clean."""
        rows = [CircuitRow(harness_family="IP", circuit="CKT_900",
                           sales_code="AAA", cnum="", pin="1",
                           connector_pn="", function="TEST")]
        q = quality.assess(rows, META)
        assert q.blank_cnum == 1 and q.blank_connector_pn == 1
        assert q.finding_total == 0 and q.clean


class TestFindings:
    def test_never_built_circuits_are_counted(self):
        q = quality.assess(circuit_rows(), META, entries=_entries())
        # CKT_500 asks for CCC, which no BODY_LEFT build carries
        assert q.never_built_circuits == 1
        assert not q.clean

    def test_unmapped_family_is_named_not_blamed(self):
        q = quality.assess(circuit_rows(), META, entries=_entries())
        assert q.families_unmapped == ["DASH"]
        assert q.families_mapped == 2
        # DASH being unassessed does not, by itself, add a finding
        assert "DASH" not in str(q.codes_not_tracked_anywhere)

    def test_malformed_expressions_count_and_clear_when_repaired(self):
        rows = circuit_rows() + [
            CircuitRow(harness_family="BODY_LEFT", circuit="CKT_810",
                       sales_code="AAA-BBB", cnum="C9", pin="9",
                       connector_pn="99999999", function="TEST")]
        issues = integrity.scan(rows)
        assert issues, "the fixture must actually be malformed"

        before = quality.assess(rows, META, issues)
        assert before.malformed_expressions == len(issues)
        assert before.malformed_rows >= 1
        assert before.repaired_expressions == 0

        fixes = {i.expression: i.expression.replace("-", "&-") for i in issues}
        after = quality.assess(rows, META, issues, fixes=fixes)
        assert after.repaired_expressions == len(issues)
        # the expression is still a finding — the *customer* must fix it at
        # source; the repair only stops it from faking never-built circuits
        assert after.malformed_expressions == len(issues)


class TestCoverage:
    def test_says_where_a_code_exists_and_where_it_does_not(self):
        q = quality.assess(circuit_rows(), META, entries=_entries())
        by_code = {c.code: c for c in q.coverage}

        aaa = by_code["AAA"]
        assert aaa.status == quality.TRACKED
        assert aaa.tracked_by and not aaa.missing_from

        # ZZZ is used by BODY_LEFT and listed by no complexity file
        zzz = by_code["ZZZ"]
        assert zzz.status == quality.UNTRACKED
        assert zzz.missing_from == ["BODY_LEFT → BODY_LEFT"]
        assert zzz.is_gap

    def test_a_code_only_used_by_an_unmapped_family_is_unassessed(self):
        rows = circuit_rows() + [
            CircuitRow(harness_family="DASH", circuit="CKT_998",
                       sales_code="QQQ", cnum="C8", pin="8",
                       connector_pn="99999999", function="TEST")]
        q = quality.assess(rows, META, entries=_entries())
        qqq = {c.code: c for c in q.coverage}["QQQ"]
        assert qqq.status == quality.UNASSESSED
        assert not qqq.is_gap, "we cannot call it a gap without a complexity file"

    def test_partial_tracking_is_distinguished_from_absent(self):
        """AAA known to both harnesses; a code known to one only is PARTIAL."""
        rows = circuit_rows() + [
            CircuitRow(harness_family="IP", circuit="CKT_960",
                       sales_code="BBB", cnum="C7", pin="7",
                       connector_pn="99999999", function="TEST")]
        q = quality.assess(rows, META, entries=_entries(rows))
        bbb = {c.code: c for c in q.coverage}["BBB"]
        # BODY_LEFT lists BBB, IP does not
        assert bbb.status == quality.PARTIAL
        assert bbb.tracked_by == ["BODY_LEFT → BODY_LEFT"]
        assert bbb.missing_from == ["IP → IP"]


class TestAutoSelect:
    def test_picks_never_built_and_every_gap(self):
        entries = _entries()
        picked = report.auto_select(entries)
        kinds = {(s.kind, s.ident) for s in picked.values()}
        assert ("circuit", "CKT_500") in kinds, "never-built circuit"
        assert ("gap", "ZZZ") in kinds, "sales code no complexity tracks"
        # a healthy circuit is never volunteered to the customer
        assert not any(ident == "CKT_200" for _kind, ident in kinds)

    def test_an_explicit_untick_is_not_undone(self):
        entries = _entries()
        picked = report.auto_select(entries)
        victim = next(k for k, s in picked.items() if s.ident == "CKT_500")
        again = report.auto_select(entries, {victim})
        assert victim not in again
        assert len(again) == len(picked) - 1

    def test_notes_say_why_each_row_was_picked(self):
        for selection in report.auto_select(_entries()).values():
            assert selection.note.strip(), "a picked row must explain itself"

    def test_notes_stay_short(self):
        """The cleanup list is read row by row; a paragraph in a cell is not."""
        for selection in report.auto_select(_entries()).values():
            assert len(selection.note) <= 140, selection.note

    def test_gap_note_names_the_code_and_nothing_else(self):
        gap = next(s for s in report.auto_select(_entries()).values()
                   if s.kind == "gap" and s.ident == "ZZZ")
        assert gap.note == ("Sales code ZZZ is in the DTx report but not in "
                            "the complexity file.")

    def test_an_inert_condition_is_picked_up_with_a_removal(self):
        """CKT_600 is 'ZZZ' — untracked, so it reads as built on every build.

        The condition is not a condition at all; the fix is to drop the code.
        """
        picked = report.auto_select(_entries())
        item = next((s for s in picked.values() if s.ident == "CKT_600"), None)
        assert item is not None, "an all-builds circuit resting on a gap code " \
                                 "must reach the cleanup list"
        assert "Remove ZZZ" in item.note

    def test_a_healthy_all_builds_circuit_is_not_picked(self):
        """CKT_700 is unconditional on tracked codes — nothing to remove."""
        picked = report.auto_select(_entries())
        assert not any(s.ident == "CKT_700" for s in picked.values())


class TestRemovalHint:
    def test_one_phrasing_for_one_fix(self):
        """A circuit and a connector resting on the same untracked code must
        say the same thing — two wordings read as two different problems."""
        entries = _entries()
        notes = {s.ident: s.note for s in report.auto_select(entries).values()}
        inert = [n for n in notes.values() if "reads as always true" in n]
        assert inert, "the fixture must contain an inert condition"
        for note in inert:
            assert note.endswith("or add it."), note

    def test_only_gap_codes_are_suggested_for_removal(self):
        assert report.removal_hint("AAA&ZZZ", {"ZZZ"}) == \
            " Remove ZZZ, or add it to the complexity file."
        assert report.removal_hint("AAA&BBB", {"ZZZ"}) == ""
        assert report.removal_hint("", {"ZZZ"}) == ""
        assert report.removal_hint("AAA", set()) == ""

    def test_several_gap_codes_are_listed_once_and_sorted(self):
        assert report.removal_hint("QZ9/QY8&QZ9", {"QZ9", "QY8"}) == \
            " Remove QY8, QZ9, or add it to the complexity file."

    def test_a_negated_gap_code_is_still_a_gap_code(self):
        assert "Remove ZZZ" in report.removal_hint("AAA&-ZZZ", {"ZZZ"})


class TestExport:
    def test_quality_sheets_are_written(self):
        import io
        from openpyxl import load_workbook

        entries = _entries()
        q = quality.assess(circuit_rows(), META, entries=entries)
        data = report.build_report(entries, {}, dtx_program=q.program,
                                   dtx_phase=q.phase, quality=q)
        wb = load_workbook(io.BytesIO(data))
        assert "Data quality" in wb.sheetnames
        assert "Sales-code coverage" in wb.sheetnames

        text = "\n".join(str(c.value) for row in wb["Data quality"].iter_rows()
                         for c in row)
        assert "9000ZZ" in text and "Never-built circuits" in text

        codes = [r[0].value for r in wb["Sales-code coverage"].iter_rows(min_row=2)]
        assert "ZZZ" in codes

    def test_the_report_still_builds_without_a_measurement(self):
        """Quality is optional — an older workbench state must still export."""
        data = report.build_report(_entries(), {})
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(data))
        assert "Data quality" not in wb.sheetnames
        assert "Circuits" in wb.sheetnames
