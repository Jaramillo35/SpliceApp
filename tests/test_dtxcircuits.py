"""Per-harness circuit applicability, on invented data.

No real programme data: fixtures_dtxcircuits builds a synthetic programme
("9000ZZ" / "X9_A") shaped to hit every classification and every failure mode.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent))

from fixtures_dtxcircuits import circuit_rows, harnesses, body_left, ip  # noqa: E402

from splice.common.errors import SpliceError
from splice.dtxcircuits import (
    ALL_BUILDS,
    NEVER,
    NO_COMPLEXITY,
    UNCONDITIONAL,
    VARIANT,
    analyze,
    analyze_harness,
    read_dtx_circuits,
    union_condition,
)
from splice.dtxcircuits.models import CircuitRow


def _by_circuit(analysis):
    return {c.circuit: c for c in analysis.circuits}


@pytest.fixture()
def results():
    return {a.harness: a for a in analyze(circuit_rows(), harnesses())}


class TestUnionCondition:
    def test_single_condition_is_kept(self):
        rows = [CircuitRow("H", "C", "AAA")]
        assert union_condition(rows) == "(AAA)"

    def test_several_conditions_are_ored(self):
        rows = [CircuitRow("H", "C", "AAA"), CircuitRow("H", "C", "BBB")]
        assert union_condition(rows) == "(AAA)/(BBB)"

    def test_one_unconditional_occurrence_wins(self):
        # a circuit reachable unconditionally at any pin is unconditional
        rows = [CircuitRow("H", "C", "AAA"), CircuitRow("H", "C", "")]
        assert union_condition(rows) is None

    def test_duplicate_conditions_collapse(self):
        rows = [CircuitRow("H", "C", "AAA"), CircuitRow("H", "C", "AAA")]
        assert union_condition(rows) == "(AAA)"


class TestClassification:
    def test_unconditional_circuit_is_on_every_build(self, results):
        c = _by_circuit(results["BODY_LEFT"])["CKT_100"]
        assert c.classification == UNCONDITIONAL
        assert c.expression is None
        assert len(c.builds_with) == 4 and not c.builds_without

    def test_variant_circuit_splits_the_builds(self, results):
        c = _by_circuit(results["BODY_LEFT"])["CKT_200"]
        assert c.classification == VARIANT
        assert c.builds_with == ["90000001AA", "90000003AA"]
        assert c.builds_without == ["90000002AA", "90000004AA"]

    def test_or_widens_and_and_narrows(self, results):
        circuits = _by_circuit(results["BODY_LEFT"])
        assert len(circuits["CKT_300"].builds_with) == 3   # AAA/BBB
        assert len(circuits["CKT_400"].builds_with) == 1   # AAA&BBB

    def test_condition_no_build_satisfies_is_a_finding(self, results):
        # CCC is tracked by the complexity but carried by no build
        c = _by_circuit(results["BODY_LEFT"])["CKT_500"]
        assert c.classification == NEVER
        assert c.is_finding and not c.builds_with
        assert results["BODY_LEFT"].findings == [c]

    def test_negation_is_evaluated(self, results):
        c = _by_circuit(results["IP"])["CKT_950"]
        assert c.classification == VARIANT
        assert c.builds_with == ["90000011AA"]   # the build WITHOUT AAA


class TestUntrackedCodes:
    """A code the complexity does not track is unknown, not absent."""

    def test_untracked_code_is_treated_as_present(self, results):
        c = _by_circuit(results["BODY_LEFT"])["CKT_600"]
        assert c.classification == ALL_BUILDS      # never NEVER
        assert c.untracked_codes == ["ZZZ"]
        assert c.relies_on_untracked

    def test_silence_can_never_manufacture_a_finding(self, results):
        # every circuit resting on an untracked code must be wide, not empty
        for a in results.values():
            for c in a.circuits:
                if c.relies_on_untracked and c.build_count:
                    assert c.builds_with, (
                        f"{c.circuit}: untracked codes must widen, not remove")

    def test_harness_reports_its_complexity_gap(self, results):
        assert results["BODY_LEFT"].untracked_codes == ["ZZZ"]
        assert results["IP"].untracked_codes == []

    def test_tracked_code_still_narrows(self, results):
        # CCC IS tracked, so it is absent (not unknown) -> never built
        c = _by_circuit(results["BODY_LEFT"])["CKT_500"]
        assert c.untracked_codes == [] and c.classification == NEVER


class TestMultiPinCircuits:
    def test_pins_are_collected(self, results):
        c = _by_circuit(results["BODY_LEFT"])["CKT_800"]
        assert c.pins == ["C4/5", "C5/6"]
        assert c.raw_expressions == ["AAA", "BBB"]

    def test_union_of_two_pins_widens_applicability(self, results):
        c = _by_circuit(results["BODY_LEFT"])["CKT_800"]
        assert c.classification == VARIANT
        assert len(c.builds_with) == 3      # AAA or BBB

    def test_unconditional_pin_makes_the_circuit_unconditional(self, results):
        c = _by_circuit(results["BODY_LEFT"])["CKT_700"]
        assert c.classification == UNCONDITIONAL
        assert len(c.builds_with) == 4


class TestMissingComplexity:
    def test_family_without_a_complexity_file_is_marked_not_guessed(self, results):
        a = results["DASH"]
        assert a.builds == 0 and a.def_id == ""
        assert _by_circuit(a)["CKT_999"].classification == NO_COMPLEXITY

    def test_missing_complexity_is_never_a_finding(self, results):
        # not knowing is not the same as knowing it is wrong
        assert results["DASH"].findings == []

    def test_harness_with_no_builds_behaves_like_no_complexity(self):
        from splice.inline.model import Harness
        empty = Harness(name="X", def_id="1", builds=[], complexity_codes={"AAA"})
        a = analyze_harness([CircuitRow("X", "C", "AAA")], empty)
        assert a.circuits[0].classification == NO_COMPLEXITY


class TestScoping:
    def test_a_condition_is_only_evaluated_against_its_own_harness(self):
        """IP tracks only AAA; BODY_LEFT's BBB must not leak into IP."""
        rows = [CircuitRow("IP", "CKT_X", "BBB")]
        a = analyze_harness(rows, ip())
        c = a.circuits[0]
        # BBB is untracked ON IP -> unknown -> present, NOT read from BODY_LEFT
        assert c.untracked_codes == ["BBB"]
        assert len(c.builds_with) == 2

    def test_families_are_analyzed_independently(self, results):
        assert set(results) == {"BODY_LEFT", "DASH", "IP"}
        assert results["IP"].builds == 2 and results["BODY_LEFT"].builds == 4

    def test_counts_add_up_to_the_circuit_total(self, results):
        for a in results.values():
            assert sum(a.counts.values()) == len(a.circuits)


def _dtx_workbook(**kw) -> bytes:
    """A synthetic DTx export, title block and all."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed DTx Circuits Report"
    ws.append(["Detailed DTx Circuits Report"])
    ws.append([kw.get("program_line", "Vehicle Program - 9000ZZ  ")])
    ws.append([kw.get("phase_line", "Build Phase - X9_A")])
    ws.append(["Report Date: Jan-01-2099 00:00"])
    ws.append([])
    ws.append(["Device Control Number", "Device Name", "CNUM", "Connector PN",
               "Harness Family", "Pin Number", "Circuit Name",
               "Circuit Function", "Sales Code"])
    for r in circuit_rows():
        ws.append(["76151", "Dev", r.cnum, r.connector_pn, r.harness_family,
                   r.pin, r.circuit, r.function, r.sales_code or None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestDtxReader:
    def test_reads_rows_past_the_title_block(self):
        rows, meta = read_dtx_circuits(_dtx_workbook(), "invented.xlsx")
        assert len(rows) == len(circuit_rows())
        assert {r.harness_family for r in rows} == {"BODY_LEFT", "IP", "DASH"}

    def test_program_and_phase_come_from_the_file_not_the_name(self):
        _rows, meta = read_dtx_circuits(_dtx_workbook(), "WRONG_NAME_1234.xlsx")
        assert meta.program == "9000ZZ"
        assert meta.phase == "X9_A"
        assert meta.rows == len(circuit_rows())
        assert meta.families == 3

    def test_blank_sales_code_becomes_unconditional(self):
        rows, _ = read_dtx_circuits(_dtx_workbook(), "invented.xlsx")
        ckt_100 = [r for r in rows if r.circuit == "CKT_100"]
        assert ckt_100 and ckt_100[0].sales_code == ""

    def test_round_trip_matches_the_in_memory_fixture(self):
        rows, _ = read_dtx_circuits(_dtx_workbook(), "invented.xlsx")
        parsed = {a.harness: a.counts for a in analyze(rows, harnesses())}
        direct = {a.harness: a.counts for a in analyze(circuit_rows(), harnesses())}
        assert parsed == direct

    def test_a_file_without_the_expected_columns_is_rejected(self):
        wb = Workbook()
        wb.active.append(["Something", "Else"])
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(SpliceError, match="Detailed DTx Circuits Report"):
            read_dtx_circuits(buf.getvalue(), "not-a-dtx.xlsx")

    def test_missing_metadata_does_not_break_the_read(self):
        data = _dtx_workbook(program_line="", phase_line="")
        rows, meta = read_dtx_circuits(data, "invented.xlsx")
        assert len(rows) == len(circuit_rows())
        assert meta.program == "" and meta.phase == ""


def _complexity_workbook(year="9000", vehicle="ZZ", phase="X9_A",
                         harness="BODY_LEFT", def_id="90001",
                         info_sheet=True) -> bytes:
    """An invented individual complexity file: build table + Harness PN info."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Complexity"
    ws.append([f"ID={def_id}", "AAA", "BBB", "CCC"])
    ws.append(["90000001AA", "X", "", ""])
    ws.append(["90000002AA", "", "X", ""])
    if info_sheet:
        hp = wb.create_sheet("Harness PN")
        hp.append(["Previous P/N", "New P/N", "Symbol"])
        hp.append([None, None, None])
        for label, value in (("year:", year), ("vehicle:", vehicle),
                             ("phase:", phase), ("harness:", harness),
                             ("id:", def_id)):
            hp.append([label, value])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestComplexityIdentity:
    """Identity comes from the Harness PN sheet, never the filename."""

    def test_reads_the_info_table(self):
        from splice.dtxcircuits import read_info
        meta = read_info(_complexity_workbook(), "WRONG-NAME-1234.xlsm")
        assert (meta.year, meta.vehicle, meta.phase) == ("9000", "ZZ", "X9_A")
        assert meta.harness == "BODY_LEFT" and meta.def_id == "90001"
        assert meta.program == "9000ZZ" and meta.complete

    def test_filename_is_ignored_for_identity(self):
        from splice.dtxcircuits import read_info
        meta = read_info(_complexity_workbook(phase="X9_A"),
                         "2.- Harness_Complexity_28RU_X2_IP_99999_01-01-2026.xlsm")
        assert meta.phase == "X9_A" and meta.harness == "BODY_LEFT"

    def test_missing_info_sheet_is_reported_not_guessed(self):
        from splice.dtxcircuits import read_info
        meta = read_info(_complexity_workbook(info_sheet=False), "x.xlsm")
        assert not meta.complete and meta.program == ""

    def test_read_harness_file_returns_builds_and_identity(self):
        from splice.dtxcircuits import read_harness_file
        harness, meta = read_harness_file(_complexity_workbook(), "x.xlsm")
        assert harness.def_id == "90001"
        assert {b.part_number for b in harness.builds} == {"90000001AA", "90000002AA"}
        assert harness.complexity_codes == {"AAA", "BBB", "CCC"}
        # the harness takes the name the file states, not the filename stem
        assert harness.name == "BODY_LEFT" and meta.harness == "BODY_LEFT"


class TestNormalisation:
    def test_program_forms_compare_equal(self):
        from splice.dtxcircuits import normalize_program as n
        assert n("2028RU") == n("28RU") == n("2028 RU") == "28RU"

    def test_phase_punctuation_is_absorbed(self):
        from splice.dtxcircuits import normalize_phase as n
        assert n("X2_A") == n("X2A") == n("x2 a") == "X2A"

    def test_a_shorter_phase_is_not_the_same_phase(self):
        from splice.dtxcircuits import normalize_phase as n
        assert n("X2") != n("X2_A")


class TestCorrespondence:
    def _dtx(self):
        from splice.dtxcircuits.models import DtxMeta
        return DtxMeta(program="9000ZZ", phase="X9_A")

    def _meta(self, **kw):
        from splice.dtxcircuits import ComplexityMeta
        base = dict(year="9000", vehicle="ZZ", phase="X9_A", harness="H",
                    filename="f.xlsm")
        base.update(kw)
        return ComplexityMeta(**base)

    def test_exact_match_passes(self):
        from splice.dtxcircuits import correspond
        result = correspond.check(self._dtx(), [self._meta()])
        assert result.all_match and len(result.matched) == 1

    def test_spelling_difference_warns_but_does_not_block(self):
        from splice.dtxcircuits import correspond
        result = correspond.check(self._dtx(), [self._meta(phase="X9A")])
        assert result.all_match                      # not blocking
        assert result.files[0].status == correspond.PHASE_SPELLING

    def test_a_different_phase_blocks(self):
        from splice.dtxcircuits import correspond
        result = correspond.check(self._dtx(), [self._meta(phase="X9")])
        assert not result.all_match
        assert result.files[0].status == correspond.PHASE_MISMATCH

    def test_a_different_program_blocks(self):
        from splice.dtxcircuits import correspond
        result = correspond.check(self._dtx(), [self._meta(vehicle="DT")])
        assert not result.all_match
        assert result.files[0].status == correspond.PROGRAM_MISMATCH

    def test_unstated_identity_warns_without_blocking(self):
        from splice.dtxcircuits import correspond
        result = correspond.check(self._dtx(), [self._meta(year="", vehicle="")])
        assert result.all_match
        assert result.files[0].status == correspond.UNKNOWN


class TestFirstSheetOnly:
    def test_only_the_first_sheet_is_read(self):
        """The Duplicate DTx Report repeats rows under the same headers;
        reading it too would double every circuit's occurrences."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Detailed DTx Circuits Report"
        ws.append(["Detailed DTx Circuits Report"])
        ws.append(["Vehicle Program - 9000ZZ"])
        ws.append(["Build Phase - X9_A"])
        ws.append([])
        ws.append(["Harness Family", "Circuit Name", "Sales Code"])
        ws.append(["BODY_LEFT", "CKT_100", "AAA"])
        dup = wb.create_sheet("Duplicate DTx Report")
        dup.append(["Detailed DTx Circuits Report"])
        dup.append(["Harness Family", "Circuit Name", "Sales Code"])
        dup.append(["BODY_LEFT", "CKT_100", "AAA"])
        buf = io.BytesIO()
        wb.save(buf)
        rows, meta = read_dtx_circuits(buf.getvalue(), "invented.xlsx")
        assert len(rows) == 1, "the duplicate sheet must not be counted"
        assert meta.program == "9000ZZ" and meta.phase == "X9_A"
