"""Tests for generated-SECR identity: numbering, naming, versioning, validation.

Covers the enumerated acceptance cases: sequences are scoped to model year +
phase and start at 1000, numbers are never reused, filenames are built from
structured metadata, updates increment the version without touching the
previous one, and a scope change blocks the update instead of silently
renumbering.

Everything runs against a temporary database.
"""

from __future__ import annotations

import io
import sqlite3
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from secrdb.core.common.errors import SpliceInputError
from secrdb.core.secr import db as secr_db, generation, identity
from secrdb.core.secr.identity import (
    CHANGE_TYPE_DESIGN,
    CHANGE_TYPE_MISCELLANEOUS,
    SecrMetadata,
    build_filename,
    build_secr_number,
    changed_fields,
    compare_metadata,
    extract_metadata_from_def,
    format_generation_date,
    validate_metadata,
)

IP = SecrMetadata(harness_family="IP", model_year="2028", phase="X1", program="RU")


@pytest.fixture()
def db_path(tmp_path: Path) -> Path:
    return tmp_path / "identity_test.db"


# ---------------------------------------------------------------------------
# DEF compare fixture
# ---------------------------------------------------------------------------

def build_def_compare(
    *,
    harness: str = "IP",
    model_year: str = "2028",
    program: str = "RU",
    new_phase: str = "X1_A",
    old_phase: str = "X0_A",
    old_program: str = "",
    old_model_year: str = "",
    include_identifiers: bool = True,
    connector_se_comment: str = "DTCR 50319",
) -> bytes:
    """A DEF→DEF compare workbook with the sheets the SECR engine expects."""
    wb = openpyxl.Workbook()

    summary = wb.active
    summary.title = "DEF_DEF_Summary"
    if include_identifiers:
        summary["B1"] = (
            f"DEF_New (Identifier) := {model_year} {program} {new_phase} "
            f"05_07_26_09_25_34 {harness}  ID: 11430"
        )
        summary["B2"] = (
            f"DEF_Old (Identifier) := {old_model_year or model_year} "
            f"{old_program or program} {old_phase} 05_06_26_08_43_41 "
            f"{harness}  ID: 11184"
        )
    headers = [
        "SE Comment", "Action", "Harness PN", "DEF Symbol",
        "Harness PN (Old)", "DEF Symbol (Old)",
    ]
    for column, header in enumerate(headers, start=1):
        summary.cell(row=3, column=column, value=header)
    for column, value in enumerate(
        [None, "CHG", "68774881AB", "18", "68774881AA", "18"], start=1
    ):
        summary.cell(row=4, column=column, value=value)

    connector = wb.create_sheet("Connector")
    for column, header in enumerate(
        [
            "SE Comment", "Action", "FCA-CNUM", "Suffix(New)", "Suffix(Old)",
            "DEF_Connector_PN", "DEF_Connector_PN(Old)",
        ],
        start=1,
    ):
        connector.cell(row=3, column=column, value=header)
    for column, value in enumerate(
        [connector_se_comment, "COMP CHG", "SD401", "NEW", "OLD", "PN-NEW", "PN-OLD"],
        start=1,
    ):
        connector.cell(row=4, column=column, value=value)

    circuit = wb.create_sheet("Circuit")
    for column, header in enumerate(
        [
            "SE Comments", "Action", "CKT NBR", "CKT Suffix", "DEF CKT COLOR",
            "DEF CKT COLOR(Old)", "DEF GAUGE", "DEF GAUGE(Old)",
        ],
        start=1,
    ):
        circuit.cell(row=3, column=column, value=header)
    for column, value in enumerate(
        ["DTCR 49919", "CHG", "A937", "F", "GN/RD", "GN/RD", "0.50", "0.35"],
        start=1,
    ):
        circuit.cell(row=4, column=column, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()


#: The engine parses the harness family out of the DEF filename, so the fixture
#: filename must follow the real convention.
def def_filename(harness: str = "IP", model_year: str = "2028", program: str = "RU",
                 phase: str = "X1") -> str:
    return (
        f"{model_year}_{program}_{phase}_A_vs_{model_year}_{program}_X0_A_"
        f"{harness}_DEF_DEF_Compare_20260507.xlsx"
    )


# ---------------------------------------------------------------------------
# Naming
# ---------------------------------------------------------------------------

def test_filename_matches_the_specified_format() -> None:
    assert (
        build_filename(IP, CHANGE_TYPE_DESIGN, 1000, 1, date(2026, 5, 7))
        == "SECR_IP_D28X1RU_1000_V1_05072026.xlsx"
    )


def test_design_change_produces_d_and_miscellaneous_produces_m() -> None:
    assert build_secr_number(IP, CHANGE_TYPE_DESIGN, 1000) == "D28X1RU_1000"
    assert build_secr_number(IP, CHANGE_TYPE_MISCELLANEOUS, 1000) == "M28X1RU_1000"


def test_generation_date_uses_mmddyyyy() -> None:
    assert format_generation_date(date(2026, 5, 7)) == "05072026"
    assert format_generation_date(date(2026, 12, 31)) == "12312026"


def test_filename_embeds_model_year_phase_and_version() -> None:
    x2 = SecrMetadata("IP", "2028", "X2", "RU")
    assert (
        build_filename(x2, CHANGE_TYPE_DESIGN, 1003, 3, date(2026, 5, 15))
        == "SECR_IP_D28X2RU_1003_V3_05152026.xlsx"
    )


def test_harness_family_with_spaces_is_filename_safe() -> None:
    console = SecrMetadata("CENTER CONSOLE_CUP", "2028", "X1", "RU")
    name = build_filename(console, CHANGE_TYPE_DESIGN, 1000, 1, date(2026, 5, 7))
    assert name == "SECR_CENTER_CONSOLE_CUP_D28X1RU_1000_V1_05072026.xlsx"
    assert " " not in name


def test_unknown_change_type_is_rejected() -> None:
    with pytest.raises(ValueError, match="Unknown change type"):
        build_secr_number(IP, "Something Else", 1000)


# ---------------------------------------------------------------------------
# Metadata extraction and validation
# ---------------------------------------------------------------------------

def test_metadata_is_extracted_from_the_def_identifier() -> None:
    extracted = extract_metadata_from_def(build_def_compare(), def_filename())

    assert extracted.metadata == SecrMetadata("IP", "2028", "X1", "RU")
    assert extracted.warnings == []
    assert validate_metadata(extracted.metadata, CHANGE_TYPE_DESIGN) == []


def test_scope_comes_from_the_new_def_and_a_phase_step_is_only_a_note() -> None:
    """A compare across phases (X0 → X1) is the normal case."""
    extracted = extract_metadata_from_def(
        build_def_compare(new_phase="X1_A", old_phase="X0_A"), def_filename()
    )

    assert extracted.metadata.phase == "X1"
    assert extracted.warnings == []
    assert any("across phases" in note for note in extracted.notes)


def test_program_mismatch_between_the_two_defs_is_a_warning() -> None:
    extracted = extract_metadata_from_def(
        build_def_compare(program="RU", old_program="WS"), def_filename()
    )
    assert any("Program differs" in warning for warning in extracted.warnings)


def test_harness_ambiguity_between_filename_and_identifier_is_a_warning() -> None:
    extracted = extract_metadata_from_def(
        build_def_compare(harness="IP"), def_filename(harness="BODY")
    )
    assert any("ambiguous" in warning for warning in extracted.warnings)
    # The identifier wins — a filename is typed by hand, an identifier is not.
    assert extracted.metadata.harness_family == "IP"


# ---------------------------------------------------------------------------
# Warnings inform; they do not block
# ---------------------------------------------------------------------------

def test_a_program_mismatch_does_not_block_generation(db_path: Path) -> None:
    """Removed by request: an unusual DEF pairing is the engineer's call."""
    plan = generation.plan_new_secr(
        build_def_compare(program="RU", old_program="WS"),
        def_filename(),
        CHANGE_TYPE_DESIGN,
        db_path=db_path,
    )

    assert plan.can_generate
    assert plan.problems == []
    assert any("Program differs" in warning for warning in plan.warnings)
    # The scope is the NEW DEF, and a real number is previewed for it.
    assert plan.metadata.program == "RU"
    assert plan.sequence_number == identity.FIRST_SEQUENCE_NUMBER


def test_an_ambiguous_harness_does_not_block_generation(db_path: Path) -> None:
    plan = generation.plan_new_secr(
        build_def_compare(harness="IP"),
        def_filename(harness="BODY"),
        CHANGE_TYPE_DESIGN,
        db_path=db_path,
    )

    assert plan.can_generate
    assert any("ambiguous" in warning for warning in plan.warnings)
    assert plan.metadata.harness_family == "IP"


def test_a_warned_secr_generates_and_is_stored(db_path: Path) -> None:
    """The whole point of the change: the workbook actually comes out."""
    payload = build_def_compare(program="RU", old_program="WS")
    plan = generation.plan_new_secr(
        payload, def_filename(), CHANGE_TYPE_DESIGN, db_path=db_path
    )

    result = generation.generate_new_secr(
        payload, def_filename(), plan, db_path=db_path
    )

    assert result.secr_bytes
    assert result.secr_id > 0
    assert result.metadata.program == "RU"


@pytest.mark.parametrize(
    "missing, expected",
    [
        ("harness_family", "Harness Family"),
        ("model_year", "Model Year"),
        ("phase", "Phase"),
        ("program", "Program"),
    ],
)
def test_missing_metadata_blocks_generation(missing: str, expected: str) -> None:
    values = IP.as_dict()
    values[missing] = ""
    problems = validate_metadata(SecrMetadata(**values), CHANGE_TYPE_DESIGN)
    assert any(expected in problem for problem in problems)


def test_unreadable_identifier_blocks_generation(db_path: Path) -> None:
    plan = generation.plan_new_secr(
        build_def_compare(include_identifiers=False),
        # A filename that also yields no harness family.
        "mystery.xlsx",
        CHANGE_TYPE_DESIGN,
        db_path=db_path,
    )
    assert not plan.can_generate
    assert plan.problems
    assert plan.sequence_number == 0  # nothing was reserved


# ---------------------------------------------------------------------------
# Numbering
# ---------------------------------------------------------------------------

def test_first_secr_in_a_scope_gets_1000(db_path: Path) -> None:
    assert secr_db.reserve_next_secr_number("2028", "X1", db_path=db_path) == 1000


def test_second_secr_in_the_same_scope_gets_1001(db_path: Path) -> None:
    secr_db.reserve_next_secr_number("2028", "X1", db_path=db_path)
    assert secr_db.reserve_next_secr_number("2028", "X1", db_path=db_path) == 1001


def test_each_phase_has_its_own_sequence(db_path: Path) -> None:
    assert secr_db.reserve_next_secr_number("2028", "X1", db_path=db_path) == 1000
    assert secr_db.reserve_next_secr_number("2028", "X1", db_path=db_path) == 1001
    # A different phase restarts.
    assert secr_db.reserve_next_secr_number("2028", "X2", db_path=db_path) == 1000


def test_each_model_year_has_its_own_sequence(db_path: Path) -> None:
    secr_db.reserve_next_secr_number("2028", "X1", db_path=db_path)
    secr_db.reserve_next_secr_number("2028", "X1", db_path=db_path)
    assert secr_db.reserve_next_secr_number("2029", "X1", db_path=db_path) == 1000


def test_scope_is_normalized_across_two_and_four_digit_years(db_path: Path) -> None:
    assert secr_db.reserve_next_secr_number("2028", "X1", db_path=db_path) == 1000
    assert secr_db.reserve_next_secr_number("28", "x1", db_path=db_path) == 1001


def test_peeking_does_not_consume_a_number(db_path: Path) -> None:
    assert secr_db.peek_next_secr_number("2028", "X1", db_path=db_path) == 1000
    assert secr_db.peek_next_secr_number("2028", "X1", db_path=db_path) == 1000
    assert secr_db.reserve_next_secr_number("2028", "X1", db_path=db_path) == 1000


def _generate(db_path: Path, harness: str = "IP"):
    payload = build_def_compare(harness=harness)
    name = def_filename(harness=harness)
    return generation.generate_new_secr(
        payload,
        name,
        generation.plan_new_secr(
            payload, name, CHANGE_TYPE_DESIGN, db_path=db_path
        ),
        db_path=db_path,
    )


def test_deleting_the_most_recent_secr_gives_its_number_back(
    db_path: Path,
) -> None:
    """Reported from the field: deleted 1006 and 1007, next SECR got 1008.

    Deleting is how a mistaken generation is undone, so skipping the numbers
    of SECRs that no longer exist helps nobody.
    """
    first = _generate(db_path)
    second = _generate(db_path, harness="BODY_LEFT")
    third = _generate(db_path, harness="DASH")
    assert [s.identity.sequence_number for s in (first, second, third)] == [
        1000, 1001, 1002,
    ]

    assert secr_db.delete_secr(third.secr_id, db_path=db_path) is True
    assert secr_db.delete_secr(second.secr_id, db_path=db_path) is True

    assert secr_db.peek_next_secr_number("2028", "X1", db_path=db_path) == 1001
    assert _generate(db_path, harness="DASH").identity.sequence_number == 1001


def test_a_number_in_the_middle_of_the_range_stays_spent(db_path: Path) -> None:
    """Reissuing 1000 while 1001 is in circulation would clone an identity."""
    first = _generate(db_path)
    second = _generate(db_path, harness="BODY_LEFT")

    assert secr_db.delete_secr(first.secr_id, db_path=db_path) is True

    assert secr_db.peek_next_secr_number("2028", "X1", db_path=db_path) == 1002
    assert second.identity.sequence_number == 1001


def test_emptying_a_scope_restarts_it_at_1000(db_path: Path) -> None:
    result = _generate(db_path)
    assert secr_db.delete_secr(result.secr_id, db_path=db_path) is True

    assert secr_db.reserve_next_secr_number("2028", "X1", db_path=db_path) == 1000


def test_reclaiming_is_scoped_to_one_model_year_and_phase(
    db_path: Path,
) -> None:
    """A delete in one scope must not disturb another scope's sequence."""
    other = build_def_compare(new_phase="X2_A", old_phase="X1_A")
    other_name = def_filename(phase="X2")
    x2 = generation.generate_new_secr(
        other,
        other_name,
        generation.plan_new_secr(
            other, other_name, CHANGE_TYPE_DESIGN, db_path=db_path
        ),
        db_path=db_path,
    )
    x1 = _generate(db_path)

    assert secr_db.delete_secr(x1.secr_id, db_path=db_path) is True

    assert secr_db.peek_next_secr_number("2028", "X1", db_path=db_path) == 1000
    assert secr_db.peek_next_secr_number("2028", "X2", db_path=db_path) == 1001
    assert x2.identity.sequence_number == 1000


def test_deleting_an_imported_secr_touches_no_sequence(db_path: Path) -> None:
    """Imported SECRs carry no sequence number, so there is nothing to reclaim."""
    from secrdb.core.secr.importer import import_secr_files
    from tests.secr_fixtures import build_secr_workbook

    generated = _generate(db_path)
    imported = import_secr_files(
        [("outside.xlsx", build_secr_workbook(secr_number="D99999Z"))],
        db_path=db_path,
    )
    imported_id = imported.imported[0].secr_id

    assert secr_db.delete_secr(imported_id, db_path=db_path) is True

    assert secr_db.peek_next_secr_number("2028", "X1", db_path=db_path) == 1001
    assert generated.identity.sequence_number == 1000


def test_reserving_without_a_scope_is_rejected(db_path: Path) -> None:
    with pytest.raises(ValueError, match="Model Year and Phase"):
        secr_db.reserve_next_secr_number("", "X1", db_path=db_path)
    with pytest.raises(ValueError, match="Model Year and Phase"):
        secr_db.reserve_next_secr_number("2028", "", db_path=db_path)


def test_concurrent_allocation_never_hands_out_the_same_number(
    db_path: Path,
) -> None:
    """Two requests arriving together must get two different numbers."""
    secr_db.init_db(db_path)

    def reserve(_: int) -> int:
        return secr_db.reserve_next_secr_number("2028", "X1", db_path=db_path)

    with ThreadPoolExecutor(max_workers=8) as pool:
        allocated = list(pool.map(reserve, range(8)))

    assert sorted(allocated) == list(range(1000, 1008))
    assert len(set(allocated)) == 8


def test_duplicate_generated_identity_cannot_be_inserted(db_path: Path) -> None:
    """MY + Phase + number + version is unique for generated SECRs."""
    plan = generation.plan_new_secr(
        build_def_compare(), def_filename(), CHANGE_TYPE_DESIGN, db_path=db_path
    )
    generation.generate_new_secr(
        build_def_compare(), def_filename(), plan, db_path=db_path
    )

    with secr_db.get_conn(db_path) as conn:
        row = conn.execute(
            "SELECT * FROM secr WHERE secr_sequence_number = 1000"
        ).fetchone()
        with pytest.raises(sqlite3.IntegrityError):
            conn.execute(
                """
                INSERT INTO secr (secr_number, version, action,
                                  scope_model_year, scope_phase,
                                  secr_sequence_number, version_number)
                VALUES (?,?,?,?,?,?,?)
                """,
                (
                    "DIFFERENT_CODE", row["version"], "create",
                    row["scope_model_year"], row["scope_phase"],
                    row["secr_sequence_number"], row["version_number"],
                ),
            )


def test_same_number_in_two_scopes_is_allowed(db_path: Path) -> None:
    """MY28/X1/1000 and MY28/X2/1000 are both valid, separate SECRs."""
    for phase in ("X1", "X2"):
        plan = generation.plan_new_secr(
            build_def_compare(new_phase=f"{phase}_A"),
            def_filename(phase=phase),
            CHANGE_TYPE_DESIGN,
            db_path=db_path,
        )
        result = generation.generate_new_secr(
            build_def_compare(new_phase=f"{phase}_A"),
            def_filename(phase=phase),
            plan,
            db_path=db_path,
        )
        assert result.identity.sequence_number == 1000

    numbers = {r["secr_number"] for r in secr_db.list_generated_secrs(db_path=db_path)}
    assert numbers == {"D28X1RU_1000", "D28X2RU_1000"}


# ---------------------------------------------------------------------------
# Create workflow
# ---------------------------------------------------------------------------

def test_new_secr_is_generated_stored_and_named(db_path: Path) -> None:
    payload, name = build_def_compare(), def_filename()
    plan = generation.plan_new_secr(
        payload, name, CHANGE_TYPE_DESIGN, when=date(2026, 5, 7), db_path=db_path
    )
    assert plan.can_generate
    assert plan.scope_is_new
    assert plan.filename == "SECR_IP_D28X1RU_1000_V1_05072026.xlsx"

    result = generation.generate_new_secr(payload, name, plan, db_path=db_path)

    assert result.filename == "SECR_IP_D28X1RU_1000_V1_05072026.xlsx"
    assert result.secr_number == "D28X1RU_1000"
    assert result.version_number == 1
    assert result.change_count > 0

    stored = secr_db.get_secr(result.secr_id, db_path=db_path)
    assert stored["import_origin"] == secr_db.ORIGIN_GENERATED
    assert stored["secr_sequence_number"] == 1000
    assert stored["scope_model_year"] == "28"
    assert stored["scope_phase"] == "X1"
    assert stored["version_number"] == 1
    assert stored["generation_date"] == "05072026"
    assert stored["source_file"] is not None  # the workbook is kept


def test_generated_workbook_records_the_resolved_metadata(db_path: Path) -> None:
    """The Summary carries the validated scope, not whatever the filename said."""
    payload = build_def_compare()
    # A filename with a leading token, which the engine's positional parse
    # would otherwise misread as the model year.
    name = "7_" + def_filename()
    plan = generation.plan_new_secr(
        payload, name, CHANGE_TYPE_DESIGN, db_path=db_path
    )
    result = generation.generate_new_secr(payload, name, plan, db_path=db_path)

    workbook = openpyxl.load_workbook(io.BytesIO(result.secr_bytes), data_only=True)
    summary = workbook["Summary"]
    assert str(summary["C10"].value) == "2028"
    assert summary["C11"].value == "RU"
    assert summary["C12"].value == "IP"
    assert summary["I2"].value == "D28X1RU_1000"
    workbook.close()


def test_generation_is_refused_when_metadata_is_invalid(db_path: Path) -> None:
    plan = generation.plan_new_secr(
        build_def_compare(include_identifiers=False),
        "mystery.xlsx",
        CHANGE_TYPE_DESIGN,
        db_path=db_path,
    )
    with pytest.raises(SpliceInputError, match="must be resolved"):
        generation.generate_new_secr(
            build_def_compare(include_identifiers=False), "mystery.xlsx", plan,
            db_path=db_path,
        )
    assert secr_db.peek_next_secr_number("2028", "X1", db_path=db_path) == 1000


# ---------------------------------------------------------------------------
# Update workflow
# ---------------------------------------------------------------------------

def _create(db_path: Path, **kwargs) -> generation.GeneratedSecr:
    payload = build_def_compare(**kwargs)
    name = def_filename(
        harness=kwargs.get("harness", "IP"),
        model_year=kwargs.get("model_year", "2028"),
        program=kwargs.get("program", "RU"),
        phase=kwargs.get("new_phase", "X1_A").split("_")[0],
    )
    plan = generation.plan_new_secr(
        payload, name, CHANGE_TYPE_DESIGN, when=date(2026, 5, 7), db_path=db_path
    )
    return generation.generate_new_secr(payload, name, plan, db_path=db_path)


def test_matching_metadata_increments_the_version(db_path: Path) -> None:
    first = _create(db_path)
    payload, name = build_def_compare(), def_filename()

    plan = generation.plan_secr_update(
        first.secr_id, payload, name, when=date(2026, 5, 9), db_path=db_path
    )
    assert not plan.scope_changed
    assert plan.can_generate
    assert plan.current_version == 1
    assert plan.next_version == 2
    assert plan.filename == "SECR_IP_D28X1RU_1000_V2_05092026.xlsx"

    second = generation.generate_secr_update(
        payload, name, first.secr_bytes, plan, db_path=db_path
    )
    assert second.version_number == 2
    assert second.secr_number == first.secr_number  # same number


def test_update_keeps_the_same_identity(db_path: Path) -> None:
    first = _create(db_path)
    payload, name = build_def_compare(), def_filename()
    plan = generation.plan_secr_update(first.secr_id, payload, name, db_path=db_path)
    second = generation.generate_secr_update(
        payload, name, first.secr_bytes, plan, db_path=db_path
    )

    assert second.identity == first.identity
    assert secr_db.peek_next_secr_number("2028", "X1", db_path=db_path) == 1001


def test_previous_version_is_never_overwritten(db_path: Path) -> None:
    first = _create(db_path)
    payload, name = build_def_compare(), def_filename()

    plan = generation.plan_secr_update(
        first.secr_id, payload, name, when=date(2026, 5, 9), db_path=db_path
    )
    generation.generate_secr_update(
        payload, name, first.secr_bytes, plan, db_path=db_path
    )

    versions = secr_db.get_versions("28", "X1", 1000, db_path=db_path)
    assert [v["version_number"] for v in versions] == [1, 2]
    assert secr_db.get_secr(first.secr_id, db_path=db_path) is not None
    assert versions[0]["filename"] == "SECR_IP_D28X1RU_1000_V1_05072026.xlsx"
    assert versions[1]["filename"] == "SECR_IP_D28X1RU_1000_V2_05092026.xlsx"


def test_three_versions_advance_v1_v2_v3(db_path: Path) -> None:
    current = _create(db_path)
    payload, name = build_def_compare(), def_filename()
    for expected in (2, 3):
        plan = generation.plan_secr_update(
            current.secr_id, payload, name, db_path=db_path
        )
        current = generation.generate_secr_update(
            payload, name, current.secr_bytes, plan, db_path=db_path
        )
        assert current.version_number == expected
    assert [
        v["version_number"] for v in secr_db.get_versions("28", "X1", 1000, db_path=db_path)
    ] == [1, 2, 3]


def test_phase_change_blocks_the_update(db_path: Path) -> None:
    first = _create(db_path)
    payload = build_def_compare(new_phase="X2_A", old_phase="X1_A")
    name = def_filename(phase="X2")

    plan = generation.plan_secr_update(first.secr_id, payload, name, db_path=db_path)

    assert plan.scope_changed
    assert not plan.can_generate
    changed = {d.field for d in plan.changed}
    assert changed == {"phase"}

    with pytest.raises(generation.SecrScopeChanged, match="Phase"):
        generation.generate_secr_update(
            payload, name, first.secr_bytes, plan, db_path=db_path
        )


def test_model_year_change_blocks_the_update(db_path: Path) -> None:
    first = _create(db_path)
    payload = build_def_compare(model_year="2029")
    name = def_filename(model_year="2029")

    plan = generation.plan_secr_update(first.secr_id, payload, name, db_path=db_path)
    assert {d.field for d in plan.changed} == {"model_year"}
    assert not plan.can_generate


def test_harness_change_blocks_the_update(db_path: Path) -> None:
    first = _create(db_path)
    payload = build_def_compare(harness="BODY")
    name = def_filename(harness="BODY")

    plan = generation.plan_secr_update(first.secr_id, payload, name, db_path=db_path)
    assert {d.field for d in plan.changed} == {"harness_family"}
    assert not plan.can_generate


def test_the_comparison_names_every_scope_field(db_path: Path) -> None:
    first = _create(db_path)
    plan = generation.plan_secr_update(
        first.secr_id,
        build_def_compare(new_phase="X2_A"),
        def_filename(phase="X2"),
        db_path=db_path,
    )
    assert [d.label for d in plan.differences] == [
        "Harness Family", "Model Year", "Phase", "Program"
    ]
    phase = next(d for d in plan.differences if d.field == "phase")
    assert (phase.existing, phase.new) == ("X1", "X2")


def test_new_secr_after_a_scope_change_uses_the_new_scope_sequence(
    db_path: Path,
) -> None:
    """Blocked update → create a new SECR → it draws from MY28/X2, not X1."""
    _create(db_path)  # MY28/X1/1000

    payload = build_def_compare(new_phase="X2_A", old_phase="X1_A")
    name = def_filename(phase="X2")
    plan = generation.plan_new_secr(
        payload, name, CHANGE_TYPE_DESIGN, when=date(2026, 5, 15), db_path=db_path
    )
    result = generation.generate_new_secr(payload, name, plan, db_path=db_path)

    assert result.identity.sequence_number == 1000  # X2 has never issued a number
    assert result.filename == "SECR_IP_D28X2RU_1000_V1_05152026.xlsx"
    assert result.version_number == 1


def test_new_secr_in_a_used_scope_continues_that_sequence(db_path: Path) -> None:
    for _ in range(3):
        payload = build_def_compare(new_phase="X2_A")
        name = def_filename(phase="X2")
        plan = generation.plan_new_secr(
            payload, name, CHANGE_TYPE_DESIGN, db_path=db_path
        )
        result = generation.generate_new_secr(payload, name, plan, db_path=db_path)
    assert result.identity.sequence_number == 1002

    payload = build_def_compare(new_phase="X2_A")
    name = def_filename(phase="X2")
    plan = generation.plan_new_secr(
        payload, name, CHANGE_TYPE_DESIGN, when=date(2026, 5, 15), db_path=db_path
    )
    assert plan.sequence_number == 1003
    assert plan.filename == "SECR_IP_D28X2RU_1003_V1_05152026.xlsx"


def test_imported_secrs_are_not_offered_for_versioned_update(db_path: Path) -> None:
    """Automatic numbering applies only to generated SECRs."""
    from tests.secr_fixtures import build_secr_workbook
    from secrdb.core.secr.importer import import_secr_bytes

    imported = import_secr_bytes(
        build_secr_workbook(), "historical.xlsx", db_path=db_path
    )
    assert imported.secr_id is not None
    assert secr_db.list_generated_secrs(db_path=db_path) == []

    with pytest.raises(SpliceInputError, match="imported, not generated"):
        generation.plan_secr_update(
            imported.secr_id, build_def_compare(), def_filename(), db_path=db_path
        )


def test_import_does_not_rename_or_renumber_a_historical_secr(db_path: Path) -> None:
    from tests.secr_fixtures import build_secr_workbook
    from secrdb.core.secr.importer import import_secr_bytes

    result = import_secr_bytes(
        build_secr_workbook(secr_number="D50319A", version="1"),
        "SECR_28RU_X1_IP_D50319A_V1_05072026.xlsx",
        db_path=db_path,
    )
    stored = secr_db.get_secr(result.secr_id, db_path=db_path)

    assert stored["secr_number"] == "D50319A"
    assert stored["filename"] == "SECR_28RU_X1_IP_D50319A_V1_05072026.xlsx"
    assert stored["secr_sequence_number"] is None
    assert stored["import_origin"] == secr_db.ORIGIN_IMPORTED
