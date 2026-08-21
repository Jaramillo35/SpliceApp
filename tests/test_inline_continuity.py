"""Tests for inline continuity validation.

The grammar tests are the important ones: the sales-code language was recovered
from data rather than documented, and two of its properties are counter-intuitive
enough that a plausible-looking change would silently break them.
"""

from __future__ import annotations

import io
from typing import List, Optional

import openpyxl
import pytest

from splice.common.errors import SpliceInputError
from splice.inline import salescode
from splice.inline.complexity import read_complexity, satisfying_builds
from splice.inline.model import (
    CONDITIONS_EXCLUSIVE,
    CONTINUOUS,
    INCONSISTENT,
    MARK_SUFFIX,
    MISSING_CONTINUATION,
    NOT_IN_SUMMARY,
)
from splice.inline.pairing import mate_name, resolve
from splice.inline.readiness import BLOCKING, assess, duplicate_complexity
from splice.inline.summary import normalise_cavity, read_circuit_summary
from splice.inline.validate import run_study


# ---------------------------------------------------------------------------
# The grammar
# ---------------------------------------------------------------------------

def test_or_binds_tighter_than_and() -> None:
    """The reading that matches the data, and not the programming default.

    Reading `/` with the looser precedence it has in most languages mispredicts
    20 of the 700 expression-bearing rows in the reference export.
    """
    assert salescode.evaluate("XZ2/XZ3/XAC&-RFX", {"XZ2"}) is True
    assert salescode.evaluate("XZ2/XZ3/XAC&-RFX", {"XZ2", "RFX"}) is False
    # Under the other precedence this would be true on LBH alone.
    assert salescode.evaluate("LBH/LBR&LBB", {"LBH"}) is False
    assert salescode.evaluate("LBH/LBR&LBB", {"LBH", "LBB"}) is True


def test_a_sales_code_may_be_numeric() -> None:
    """`501` is a sales code. Requiring a leading letter turned 24 sound
    cavities into false failures during validation."""
    assert salescode.codes_in("501") == {"501"}
    assert salescode.evaluate("501", {"501"}) is True
    assert salescode.evaluate("XC4/HAH/HBB/501", {"501"}) is True


def test_an_empty_expression_is_unconditional() -> None:
    assert salescode.evaluate("", set()) is True
    assert salescode.evaluate(None, set()) is True


def test_negation_applies_to_atoms_and_groups() -> None:
    assert salescode.evaluate("-XPR", set()) is True
    assert salescode.evaluate("-XPR", {"XPR"}) is False
    assert salescode.evaluate("RSY&(CVM/CM5)&-RTC", {"RSY", "CVM"}) is True
    assert salescode.evaluate("RSY&(CVM/CM5)&-RTC", {"RSY", "CVM", "RTC"}) is False


def test_unbalanced_parentheses_are_rejected() -> None:
    assert salescode.is_valid("RSY&(CVM/CM5)")
    assert not salescode.is_valid("RSY&(CVM")


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def build_complexity(def_id: str, codes: List[str], builds) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Complexity"
    sheet.cell(row=1, column=1, value=f"ID={def_id}")
    for index, code in enumerate(codes, start=2):
        sheet.cell(row=1, column=index, value=code)
    for row_number, (part, carried) in enumerate(builds, start=2):
        sheet.cell(row=row_number, column=1, value=part)
        for index, code in enumerate(codes, start=2):
            if code in carried:
                sheet.cell(row=row_number, column=index, value="X")
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


HEAD = ["Harness Family", "Circuit", "Suffix", "Size", "Material", "Color",
        "Stripe", "CNumber", "Cav", "Twist Ckt", "Device Name", "Part Nbr",
        "Conn End", "Term Mat", "Term PN", "Seal PN", "Sales Code"]


def build_summary(blocks) -> bytes:
    """blocks: [(name, def_id, [part numbers], [end dicts])]"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Circuit Summary"
    for index, label in enumerate(HEAD, start=1):
        sheet.cell(row=1, column=index, value=label)
    row = 2
    for name, def_id, parts, ends in blocks:
        sheet.cell(row=row, column=1, value=f"{name} - {def_id}")
        for index, label in enumerate(HEAD[1:], start=2):
            sheet.cell(row=row, column=index, value=label)
        for offset, part in enumerate(parts):
            sheet.cell(row=row, column=18 + offset, value=f"{offset+1}~{part}")
        row += 1
        for end in ends:
            sheet.cell(row=row, column=1, value=f"{name} - {def_id}")
            sheet.cell(row=row, column=2, value=end["circuit"])
            sheet.cell(row=row, column=3, value=end.get("suffix", ""))
            sheet.cell(row=row, column=4, value=end.get("size", "0.35"))
            sheet.cell(row=row, column=5, value=end.get("material", "CABT"))
            sheet.cell(row=row, column=8, value=end["cnum"])
            sheet.cell(row=row, column=9, value=end["cav"])
            sheet.cell(row=row, column=11, value=end.get("device", "Inline_X1 A_B"))
            sheet.cell(row=row, column=17, value=end.get("sc", ""))
            for offset, part in enumerate(parts):
                if part in end.get("builds", parts):
                    sheet.cell(row=row, column=18 + offset, value="X")
            row += 1
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Readers
# ---------------------------------------------------------------------------

def test_a_cavity_occurrence_index_is_normalised() -> None:
    """`8:1` is a second wire in cavity 8, not a cavity called "8:1"."""
    assert normalise_cavity("8") == "8"
    assert normalise_cavity("8:1") == "8"
    assert normalise_cavity(" 24:2 ") == "24"


def test_the_def_id_comes_from_inside_the_file() -> None:
    """The reference IP complexity is named 11645 and declares 11661."""
    payload = build_complexity("11661", ["501"], [("68774874AC", {"501"})])
    harness = read_complexity(payload, "Harness_Complexity_IP_11645.xlsm")
    assert harness.def_id == "11661"


def test_a_file_without_a_complexity_sheet_is_refused() -> None:
    workbook = openpyxl.Workbook()
    buffer = io.BytesIO()
    workbook.save(buffer)
    with pytest.raises(SpliceInputError, match="Complexity"):
        read_complexity(buffer.getvalue(), "wrong.xlsx")


def test_each_harness_block_carries_its_own_part_numbers() -> None:
    payload = build_summary([
        ("IP", "11661", ["A1", "A2"],
         [{"circuit": "C1", "cnum": "X1A", "cav": "1", "builds": ["A1"]}]),
        ("Dash", "11671", ["B1"],
         [{"circuit": "C1", "cnum": "Y1A", "cav": "1"}]),
    ])
    harnesses, ends = read_circuit_summary(payload)
    assert set(harnesses) == {"11661", "11671"}
    ip = [e for e in ends if e.harness_id == "11661"][0]
    dash = [e for e in ends if e.harness_id == "11671"][0]
    assert ip.builds == frozenset({"A1"})
    assert dash.builds == frozenset({"B1"})


# ---------------------------------------------------------------------------
# Applicability
# ---------------------------------------------------------------------------

def test_a_condition_is_evaluated_against_its_own_harness_builds() -> None:
    """`RSY & RTC` names two real codes that no build carries.

    Free Boolean enumeration invents that vehicle and reports a gap in it.
    """
    payload = build_complexity(
        "11661", ["RSY", "RTC", "501"],
        [("P1", {"RSY", "501"}), ("P2", {"RTC", "501"})],
    )
    harness = read_complexity(payload, "ip.xlsm")

    assert len(satisfying_builds(harness, "RSY&RTC")) == 0
    assert len(satisfying_builds(harness, "RSY&-RTC")) == 1
    # 501 is standard: on every build.
    assert len(satisfying_builds(harness, "501")) == 2


# ---------------------------------------------------------------------------
# Pairing
# ---------------------------------------------------------------------------

def test_mate_names_follow_both_conventions() -> None:
    assert mate_name("X301A") == "Y301A"
    assert mate_name("Y313A") == "X313A"
    assert mate_name("I350X") == "I350Y"
    assert mate_name("I41A") is None      # no convention -> never guessed


def test_a_connector_with_no_mate_is_reported_not_guessed() -> None:
    payload = build_summary([
        ("IP", "11661", ["A1"],
         [{"circuit": "C1", "cnum": "X301A", "cav": "1"},
          {"circuit": "C2", "cnum": "Y316A", "cav": "1"}]),
        ("Dash", "11671", ["B1"],
         [{"circuit": "C1", "cnum": "Y301A", "cav": "1"}]),
    ])
    harnesses, ends = read_circuit_summary(payload)
    pairs, unmated = resolve(ends, set(harnesses))
    assert len(pairs) == 1
    assert ("Y316A", "11661") in unmated


def test_an_excluded_harness_neither_pairs_nor_absorbs_a_mate() -> None:
    payload = build_summary([
        ("IP", "11661", ["A1"], [{"circuit": "C1", "cnum": "X301A", "cav": "1"}]),
        ("Test", "11670", ["T1"], [{"circuit": "C1", "cnum": "Y301A", "cav": "1"}]),
    ])
    harnesses, ends = read_circuit_summary(payload)
    pairs, unmated = resolve(ends, {"11661"})
    assert pairs == []
    assert ("X301A", "11661") in unmated


# ---------------------------------------------------------------------------
# The rules
# ---------------------------------------------------------------------------

def _matrix(def_id: str, codes) -> bytes:
    """A small but realistic complexity: one build per code, plus one with all.

    A single build carrying every code would make a condition like ``RSY&-CVM``
    unsatisfiable, which is a property of the fixture rather than of the data.
    """
    codes = list(codes or ["501"])
    builds = [(f"{def_id}-{i}", {code}) for i, code in enumerate(codes, start=1)]
    builds.append((f"{def_id}-ALL", set(codes)))
    return build_complexity(def_id, codes, builds)


def _study(ip_ends, dash_ends, ip_codes=None, dash_codes=None):
    """Build a two-harness study whose two applicability sources agree.

    The Circuit Summary states applicability twice — expression and part-number
    columns — and in real exports they agree on every row. The fixture derives
    the columns from the expression so it has that property too; otherwise every
    test would trip the integrity check.
    """
    complexity = {
        "11661": read_complexity(_matrix("11661", ip_codes), "ip.xlsm"),
        "11671": read_complexity(_matrix("11671", dash_codes), "dash.xlsm"),
    }

    def stamp(ends, def_id):
        harness = complexity[def_id]
        parts = [b.part_number for b in harness.builds]
        for end in ends:
            end.setdefault("builds", [
                b.part_number
                for b in satisfying_builds(harness, end.get("sc", ""))
            ])
        return parts

    ip_parts = stamp(ip_ends, "11661")
    dash_parts = stamp(dash_ends, "11671")
    payload = build_summary([
        ("IP", "11661", ip_parts, ip_ends),
        ("Dash", "11671", dash_parts, dash_ends),
    ])
    harnesses, ends = read_circuit_summary(payload)
    pairs, unmated = resolve(ends, set(harnesses))
    return run_study(harnesses, ends, complexity, pairs, unmated)


def test_any_option_that_continues_clears_the_cavity() -> None:
    """The adopted rule: several options per cavity, one pairing is enough."""
    result = _study(
        [{"circuit": "A934", "suffix": "A", "cnum": "X301A", "cav": "8", "sc": "RSY"},
         {"circuit": "A934", "suffix": "B", "cnum": "X301A", "cav": "8", "sc": "CVM"}],
        [{"circuit": "A934", "cnum": "Y301A", "cav": "8", "sc": "RSY"}],
        ip_codes=["RSY", "CVM"], dash_codes=["RSY"],
    )
    cavity = [f for f in result.findings if f.cavity == "8"][0]
    assert cavity.verdict == CONTINUOUS


def test_a_cavity_empty_on_one_side_is_a_missing_continuation() -> None:
    result = _study(
        [{"circuit": "A900", "cnum": "X301A", "cav": "3"}],
        [{"circuit": "A900", "cnum": "Y301A", "cav": "4"}],
    )
    verdicts = {f.cavity: f.verdict for f in result.findings if f.cavity}
    assert verdicts["3"] == MISSING_CONTINUATION
    assert verdicts["4"] == MISSING_CONTINUATION


def test_different_circuits_at_one_cavity_are_inconsistent() -> None:
    result = _study(
        [{"circuit": "T108", "cnum": "X301A", "cav": "1"}],
        [{"circuit": "Z909", "cnum": "Y301A", "cav": "1"}],
    )
    finding = [f for f in result.findings if f.cavity == "1"][0]
    assert finding.verdict == INCONSISTENT
    assert "T108" in finding.reason and "Z909" in finding.reason


def test_a_condition_that_is_never_built_is_reported() -> None:
    result = _study(
        [{"circuit": "A900", "cnum": "X301A", "cav": "1", "sc": "NOPE"}],
        [{"circuit": "A900", "cnum": "Y301A", "cav": "1", "sc": "501"}],
        ip_codes=["501"], dash_codes=["501"],
    )
    finding = [f for f in result.findings if f.cavity == "1"][0]
    assert finding.verdict == CONDITIONS_EXCLUSIVE
    assert "never built" in finding.reason


def test_attribute_differences_are_marked_not_failed() -> None:
    result = _study(
        [{"circuit": "F944", "suffix": "AB", "cnum": "X301A", "cav": "4"}],
        [{"circuit": "F944", "suffix": "", "cnum": "Y301A", "cav": "4"}],
    )
    finding = [f for f in result.findings if f.cavity == "4"][0]
    assert finding.verdict == CONTINUOUS
    assert MARK_SUFFIX in finding.marks


def test_an_unmated_connector_becomes_a_finding() -> None:
    payload = build_summary([
        ("IP", "11661", ["A1"], [{"circuit": "C1", "cnum": "Y316A", "cav": "1"}]),
    ])
    harnesses, ends = read_circuit_summary(payload)
    pairs, unmated = resolve(ends, set(harnesses))
    result = run_study(harnesses, ends, {}, pairs, unmated)
    assert [f.verdict for f in result.findings] == [NOT_IN_SUMMARY]
    assert result.review[0].connector_a == "Y316A"


# ---------------------------------------------------------------------------
# Readiness
# ---------------------------------------------------------------------------

def test_a_harness_without_complexity_is_reported_as_blocking() -> None:
    harnesses, _ = read_circuit_summary(build_summary([
        ("IP", "11661", ["A1"], [{"circuit": "C1", "cnum": "X1A", "cav": "1"}]),
    ]))
    gaps = assess(harnesses, {}, set())
    assert any(g.severity == BLOCKING and "11661" in g.what for g in gaps)


def test_duplicate_complexity_files_are_reported() -> None:
    one = read_complexity(build_complexity("11667", ["A"], [("P", {"A"})]), "a.xlsm")
    two = read_complexity(
        build_complexity("11667", ["A", "JJB"], [("P", {"A", "JJB"})]), "b.xlsm"
    )
    gaps = duplicate_complexity([one, two])
    assert gaps and gaps[0].severity == BLOCKING     # they disagree
    assert "a.xlsm" in gaps[0].affects and "b.xlsm" in gaps[0].affects


# ---------------------------------------------------------------------------
# Reporting: one circuit per row
# ---------------------------------------------------------------------------

def test_each_circuit_gets_its_own_row() -> None:
    """A934A and A934B are separate wires and must not share a row.

    Both continue through the single A934 on the mating side: the IP splits the
    cavity by sales code, the Dash does not, and one Dash wire is the
    counterpart of both variants.
    """
    from splice.inline import report

    result = _study(
        [{"circuit": "A934", "suffix": "A", "cnum": "X301A", "cav": "8",
          "sc": "RSY&-CVM", "material": "5ABT"},
         {"circuit": "A934", "suffix": "B", "cnum": "X301A", "cav": "8",
          "sc": "RSY&CVM", "material": "5ABT"}],
        [{"circuit": "A934", "cnum": "Y301A", "cav": "8", "sc": "RSY",
          "material": "CABT"}],
        ip_codes=["RSY", "CVM"], dash_codes=["RSY"],
    )
    frame = report.marked_frame(result)

    circuits = list(frame["Circuit A"])
    assert "A934A" in circuits and "A934B" in circuits
    assert len(frame) == 2, "the two wires were collapsed into one row"

    # The Dash wire is not consumed by the first variant.
    for label in ("A934A", "A934B"):
        row = frame[frame["Circuit A"] == label].iloc[0]
        assert row["Circuit B"] == "A934", f"{label} lost its counterpart"
        assert "no counterpart" not in row["Marked"]
        assert "material differs" in row["Marked"]


def test_one_wire_serves_every_variant_that_it_applies_to() -> None:
    """The mechanism behind the row above, at the engine level."""
    result = _study(
        [{"circuit": "A934", "suffix": "A", "cnum": "X301A", "cav": "8",
          "sc": "RSY&-CVM"},
         {"circuit": "A934", "suffix": "B", "cnum": "X301A", "cav": "8",
          "sc": "RSY&CVM"}],
        [{"circuit": "A934", "cnum": "Y301A", "cav": "8", "sc": "RSY"}],
        ip_codes=["RSY", "CVM"], dash_codes=["RSY"],
    )
    finding = [f for f in result.findings if f.cavity == "8"][0]
    matched = [o for o in finding.options if o.matched]
    assert len(matched) == 2, "the single far-side wire was consumed by one variant"
    assert {o.suffix_a for o in matched} == {"A", "B"}


def test_a_counterpart_that_never_applies_is_still_unpaired() -> None:
    """One-to-many must not become "anything matches"."""
    result = _study(
        [{"circuit": "A934", "cnum": "X301A", "cav": "8", "sc": "RSY"}],
        [{"circuit": "A934", "cnum": "Y301A", "cav": "8", "sc": "NOPE"}],
        ip_codes=["RSY"], dash_codes=["RSY"],
    )
    finding = [f for f in result.findings if f.cavity == "8"][0]
    assert not any(o.matched for o in finding.options)
    assert finding.verdict == CONDITIONS_EXCLUSIVE


def test_a_wire_only_on_the_mating_side_is_shown_too() -> None:
    from splice.inline import report

    result = _study(
        [{"circuit": "F948", "cnum": "X301A", "cav": "9"}],
        [{"circuit": "F948", "cnum": "Y301A", "cav": "9"},
         {"circuit": "F948", "suffix": "AB", "cnum": "Y301A", "cav": "9"}],
    )
    frame = report.marked_frame(result)
    extra = frame[frame["Circuit A"] == "—"]
    assert len(extra) == 1
    assert extra.iloc[0]["Circuit B"] == "F948AB"
    assert "no counterpart" in extra.iloc[0]["Marked"]


def test_suffix_matching_prefers_the_same_suffix() -> None:
    """A934A pairs with A934A, not with a bare A934 that appears first."""
    result = _study(
        [{"circuit": "A934", "suffix": "A", "cnum": "X301A", "cav": "8"}],
        [{"circuit": "A934", "cnum": "Y301A", "cav": "8"},
         {"circuit": "A934", "suffix": "A", "cnum": "Y301A", "cav": "8"}],
    )
    finding = [f for f in result.findings if f.cavity == "8"][0]
    matched = [o for o in finding.options if o.circuit_a and o.circuit_b][0]
    assert matched.suffix_a == "A" and matched.suffix_b == "A"


# ---------------------------------------------------------------------------
# The second applicability source
# ---------------------------------------------------------------------------

def test_the_two_applicability_sources_are_cross_checked() -> None:
    """The summary states applicability twice; a disagreement is escalated."""
    from splice.inline.model import INTEGRITY
    from splice.inline.validate import check_integrity

    complexity = {
        "11661": read_complexity(
            build_complexity("11661", ["RSY", "CVM"],
                             [("A1", {"RSY"}), ("A2", {"CVM"})]), "ip.xlsm")
    }
    payload = build_summary([
        ("IP", "11661", ["A1", "A2"],
         [{"circuit": "C1", "cnum": "X1A", "cav": "1", "sc": "RSY",
           "builds": ["A1"]}]),
    ])
    _, ends = read_circuit_summary(payload)
    assert check_integrity(ends, complexity) == []   # expression matches columns

    mismatched = build_summary([
        ("IP", "11661", ["A1", "A2"],
         [{"circuit": "C1", "cnum": "X1A", "cav": "1", "sc": "RSY",
           "builds": ["A1", "A2"]}]),
    ])
    _, bad_ends = read_circuit_summary(mismatched)
    findings = check_integrity(bad_ends, complexity)
    assert len(findings) == 1
    assert findings[0].verdict == INTEGRITY
    assert findings[0].needs_review


def test_a_code_the_other_harness_does_not_track_is_unknown_not_absent() -> None:
    """From the reference release, and it cost a false finding.

    ``G74`` is conditioned ``JPB/JRK`` on the Body_Right and ``JPB`` on the
    Body_Left. ``JPB`` is on all 19 Body_Left builds but appears nowhere in the
    Body_Right complexity — not even in its header. Reading that silence as
    "this vehicle has no JPB" reports a continuity break on a sound joint.
    """
    from splice.inline.complexity import counterpart_applies

    body_right = read_complexity(
        build_complexity("11680", ["JRK", "501"],
                         [("R1", {"JRK", "501"}), ("R2", {"JRK", "501"})]),
        "body_right.xlsm",
    )
    body_left = read_complexity(
        build_complexity("11667", ["JPB", "JRK", "501"],
                         [("L1", {"JPB", "JRK", "501"})]),
        "body_left.xlsm",
    )

    assert "JPB" not in body_right.complexity_codes
    assert counterpart_applies(body_right, "JPB/JRK", body_left, "JPB") is True


def test_an_unknown_code_does_not_make_everything_match() -> None:
    """The permissive reading is still bounded by the far side being built."""
    from splice.inline.complexity import counterpart_applies

    left = read_complexity(
        build_complexity("11680", ["JRK"], [("R1", {"JRK"})]), "a.xlsm"
    )
    right = read_complexity(
        build_complexity("11667", ["JPB"], [("L1", {"JPB"})]), "b.xlsm"
    )
    # ZZZ is built on neither, so no counterpart exists whatever the vocabulary.
    assert counterpart_applies(left, "JRK", right, "ZZZ") is False
