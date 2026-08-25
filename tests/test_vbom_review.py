"""In-app VBOM review gate: resolutions, reasons, and DEFE generation."""

from __future__ import annotations

import io

import pandas as pd
import pytest
from openpyxl import load_workbook

from splice.vbom import review


REVIEW_DF = pd.DataFrame([{
    "ReviewID": "VIN001|IP", "VIN": "VIN001", "HarnessFamily": "IP",
    "ReviewReason": "No complete PN covers every required sales code; "
                    "Multiple PNs share the best score",
    "EngineRecommendation": "PN-A", "AllowedPNs": "PN-A,PN-B,N/A",
    "RequiredSalesCodes": "XZ2", "MissingSalesCodes": "", "ExtraSalesCodes": "",
    "Giveaway": "", "CandidateDetails": "PN-A | CANDIDATE",
    "SelectedPN": "", "ReviewerNotes": "",
}])

SELECTIONS = pd.DataFrame([
    {"VIN": "VIN001", "HarnessFamily": "IP", "SelectedHarnessPN": "PN-A",
     "MatchStatus": "INCOMPLETE"},
    {"VIN": "VIN001", "HarnessFamily": "BODY", "SelectedHarnessPN": "PN-C",
     "MatchStatus": "COMPLETE"},
    {"VIN": "VIN002", "HarnessFamily": "IP", "SelectedHarnessPN": "PN-B",
     "MatchStatus": "COMPLETE"},
])

VINS = pd.DataFrame({"VIN": ["VIN001", "VIN002"]})


class TestHelpers:
    def test_allowed_pns_split(self):
        assert review.allowed_pns(REVIEW_DF.iloc[0]) == ["PN-A", "PN-B", "N/A"]

    def test_reason_counts_split_compound_reasons(self):
        counts = review.reason_counts(REVIEW_DF)
        assert counts == {
            "No complete PN covers every required sales code": 1,
            "Multiple PNs share the best score": 1,
        }

    def test_apply_resolutions_targets_only_the_flagged_row(self):
        resolved = review.apply_resolutions(SELECTIONS, {"VIN001|IP": "PN-B"})
        row = resolved[(resolved["VIN"] == "VIN001")
                       & (resolved["HarnessFamily"] == "IP")].iloc[0]
        assert row["SelectedHarnessPN"] == "PN-B"
        assert row["MatchStatus"] == "RESOLVED (SE)"
        untouched = resolved[(resolved["VIN"] == "VIN002")].iloc[0]
        assert untouched["SelectedHarnessPN"] == "PN-B"
        assert untouched["MatchStatus"] == "COMPLETE"

    def test_resolving_to_na_marks_not_applicable(self):
        resolved = review.apply_resolutions(SELECTIONS, {"VIN001|IP": "N/A"})
        row = resolved[(resolved["VIN"] == "VIN001")
                       & (resolved["HarnessFamily"] == "IP")].iloc[0]
        assert row["MatchStatus"] == "NOT_APPLICABLE"


class TestDefeGeneration:
    def test_defe_reflects_resolved_selection(self):
        try:
            from splice.vbom.workflow import _load_vbom_module
            _load_vbom_module()
        except Exception as exc:  # pragma: no cover - environment-specific
            pytest.skip(f"legacy VBOM module unavailable: {exc}")

        resolved = review.apply_resolutions(SELECTIONS, {"VIN001|IP": "PN-B"})
        name, payload = review.generate_defe("2027", "RU", resolved, VINS)
        assert name == "27_RU_VBOM_Template_for_DEFE.xlsx"

        ws = load_workbook(io.BytesIO(payload)).active
        rows = {str(ws.cell(row=r, column=1).value): r
                for r in range(3, ws.max_row + 1)}
        # the SE's choice (PN-B) covers both VINs' IP rows now
        assert "PN-B" in rows
        pnb = rows["PN-B"]
        assert ws.cell(row=pnb, column=3).value == "x"   # VIN001
        assert ws.cell(row=pnb, column=4).value == "x"   # VIN002
        # the engine's overridden pick must NOT appear for VIN001 anymore
        assert "PN-A" not in rows
