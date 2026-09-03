"""Dressing a workbook changes how it looks, never what it says."""

from __future__ import annotations

import hashlib
import io

from openpyxl import Workbook, load_workbook

from splice.common import workbook


def _values(data: bytes, skip_readme: bool = True) -> str:
    wb = load_workbook(io.BytesIO(data))
    h = hashlib.sha256()
    for ws in wb.worksheets:
        if skip_readme and ws.title == workbook.README_TITLE:
            continue
        h.update(ws.title.encode())
        for row in ws.iter_rows(values_only=True):
            h.update(repr(row).encode())
    return h.hexdigest()


def _sample() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"
    ws.append(["Circuit", "Verdict", "Detail"])
    for i in range(40):
        ws.append([f"QK{i:03d}", "variant" if i % 2 else "never built",
                   "a longer explanation that should wrap in a notes column " * 2])
    form = wb.create_sheet("Summary")
    form["A1"] = "Systems Engineering Change Request"
    form.merge_cells("A1:D1")
    form["A3"] = "SECR #"
    form["B3"] = "D28X1RU_1001"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def test_values_are_untouched_and_a_readme_is_added_last():
    before = _sample()
    after = workbook.dress(before, "Findings.xlsx", tool="Circuit Health",
                           version="0.1.0 (abc1234)", by="MJ", context="2030QX · V1_A",
                           purpose="Missing circuits across inlines.",
                           inputs=["Circuit_Summary.xlsx"])
    assert _values(before) == _values(after)
    wb = load_workbook(io.BytesIO(after))
    assert wb.sheetnames[-1] == workbook.README_TITLE
    assert wb.active.title == "Findings", "readers that use wb.active keep working"
    readme = wb[workbook.README_TITLE]
    text = " ".join(str(c.value) for row in readme.iter_rows() for c in row if c.value)
    assert "Circuit Health" in text and "2030QX" in text and "MJ" in text
    assert "0.1.0 (abc1234)" in text and "Circuit_Summary.xlsx" in text
    assert "Findings" in text and "40 row(s)" in text
    assert readme._images, "the Versigent mark is embedded"


def test_tables_get_the_anatomy_and_forms_keep_theirs():
    after = workbook.dress(_sample(), "Findings.xlsx")
    wb = load_workbook(io.BytesIO(after))
    ws = wb["Findings"]
    head = ws["A1"]
    assert head.font.bold and head.font.color.rgb.endswith(workbook.HEADER_TEXT)
    assert head.fill.fgColor.rgb.endswith(workbook.HEADER_FILL)
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:C41"
    assert ws.column_dimensions["C"].width == workbook.NOTES_WIDTH
    assert ws.page_setup.orientation == "landscape"
    assert ws.print_title_rows in ("1:1", "$1:$1")
    form = wb["Summary"]
    assert not form["A1"].fill.fill_type, "a form with merged cells is left alone"
    assert form.freeze_panes is None


def test_macro_files_and_chart_files_pass_through():
    sample = _sample()
    assert workbook.dress(sample, "Template.xlsm") is sample
    assert not workbook.can_dress("x.xlsm", sample)
    assert workbook.dress(b"not a workbook", "x.xlsx") == b"not a workbook"


def test_customer_formats_pass_through_untouched():
    sample = _sample()
    for name in ("SECR_IP_D28X1RU_1001_V1.xlsx", "30_QX_VBOM_Template_for_DEFE.xlsx",
                 "2.- Harness_Complexity_30QX_IP.xlsx"):
        assert workbook.dress(sample, name) is sample, name


def test_dressing_is_idempotent():
    once = workbook.dress(_sample(), "F.xlsx")
    twice = workbook.dress(once, "F.xlsx")
    wb = load_workbook(io.BytesIO(twice))
    assert wb.sheetnames.count(workbook.README_TITLE) == 1
