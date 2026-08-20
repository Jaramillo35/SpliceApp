"""Reproducibly build the synthetic DTx old/new pair used by the golden-content test.

The data is entirely made up (no proprietary circuits) but matches the real DTx layout the
engine expects: sheet "Detailed DTx Circuits Report", 5 preamble rows, a header on Excel row 6,
then data. The pair is engineered to yield a known change set:

    added   = D4A   (present in NEW only)
    removed = D3A   (present in OLD only)
    modified= D2A   (Wire Gauge differs OLD→NEW)
    unchanged = D1A

Run:  python tests/fixtures/dtx_golden/build_fixtures.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

SHEET = "Detailed DTx Circuits Report"
COLUMNS = [
    "CNUM", "Pin Number", "Device Control Number", "Device Name", "Suffix",
    "Number of Cavities", "Connector PN", "Harness Family", "Circuit Name",
    "Circuit Suffix", "Circuit Function", "Color", "Terminal",
    "Connector FCA part number", "Wire Gauge", "Wire Type", "Sales Code",
]


def _row(cnum, dcn, name, family, gauge):
    """One fully-specified circuit row keyed by the fields the change engine compares."""
    return {
        "CNUM": cnum, "Pin Number": "1", "Device Control Number": dcn, "Device Name": name,
        "Suffix": "", "Number of Cavities": "6", "Connector PN": "1355082-1",
        "Harness Family": family, "Circuit Name": "C34", "Circuit Suffix": "",
        "Circuit Function": "HVC - ACTUATOR COM", "Color": "DB/BU", "Terminal": "Tin",
        "Connector FCA part number": "AHY06440", "Wire Gauge": gauge, "Wire Type": "TXL",
        "Sales Code": "501",
    }


OLD_ROWS = [
    _row("D1A", "50063", "Act_HVAC_Recirc", "HVAC_FRONT", "20"),   # unchanged
    _row("D2A", "50064", "Sensor_Temp", "HVAC_FRONT", "20"),       # modified (gauge 20→18)
    _row("D3A", "50065", "Relay_Fan", "COOLING", "16"),            # removed
]
NEW_ROWS = [
    _row("D1A", "50063", "Act_HVAC_Recirc", "HVAC_FRONT", "20"),   # unchanged
    _row("D2A", "50064", "Sensor_Temp", "HVAC_FRONT", "18"),       # modified
    _row("D4A", "50066", "Module_Body", "BODY", "18"),             # added
]


def _write(path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for i in range(1, 6):                       # 5 preamble rows (header lands on Excel row 6)
        ws.cell(row=i, column=1, value=f"(preamble line {i})" if i == 1 else "")
    for c, name in enumerate(COLUMNS, start=1):
        ws.cell(row=6, column=c, value=name)
    for r, row in enumerate(rows, start=7):
        for c, name in enumerate(COLUMNS, start=1):
            ws.cell(row=r, column=c, value=row[name])
    wb.save(path)


def main() -> None:
    here = Path(__file__).resolve().parent
    _write(here / "old.xlsx", OLD_ROWS)
    _write(here / "new.xlsx", NEW_ROWS)
    print(f"wrote {here/'old.xlsx'} and {here/'new.xlsx'}")


if __name__ == "__main__":
    main()
