"""Tests for the SECR database: parser, persistence, import and query API.

Everything runs against a temporary database file, so the real
``data/secr_database.db`` is never touched.

The workbook fixtures are built in-memory to mirror the structure of a real
generated SECR: a ``Summary`` sheet, an ``Add_Remove_Report_Summary`` holding
*two* tables (connectors then circuits), a ``Connector`` sheet, a ``Circuit``
sheet whose old values come from ``(Old)`` columns and ``"Old DEF :"`` cell
comments, and a ``DEF_DEF_Summary``.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import List, Optional

import openpyxl
import pytest
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from secrdb.core.common.errors import SpliceSchemaError
from secrdb.core.secr import api, db as secr_db
from secrdb.core.secr.importer import (
    STATUS_DUPLICATE,
    STATUS_FAILED,
    STATUS_IMPORTED,
    import_secr_bytes,
    import_secr_files,
)
from secrdb.core.secr.parse import (
    extract_dtcr_numbers,
    parse_secr_bytes,
    split_endpoint,
)

YELLOW = PatternFill("solid", fgColor="FFFFFF00")


@pytest.fixture()
def db_path(tmp_path: Path) -> Path:
    """A throwaway database; never the real one."""
    return tmp_path / "secr_test.db"


# ---------------------------------------------------------------------------
# Workbook fixture builder
# ---------------------------------------------------------------------------

def _write_row(ws, row: int, values: List[Optional[str]]) -> None:
    for column, value in enumerate(values, start=1):
        ws.cell(row=row, column=column, value=value)


def build_secr_workbook(
    *,
    secr_number: str = "D50319A",
    version: str = "1",
    harness_family: str = "IP",
    program: str = "RU",
    model_year: str = "2028",
) -> bytes:
    """A minimal but structurally faithful SECR workbook."""
    wb = openpyxl.Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary["I2"] = secr_number
    summary["I3"] = version
    summary["I4"] = "2026-05-07"
    summary["C7"] = "28 X1 RELEASE"
    summary["C8"] = f"SECR_28RU_X1_IP_{secr_number}_V{version}"
    summary["C10"] = model_year
    summary["C11"] = program
    summary["F10"] = "X1"
    summary["C12"] = harness_family
    summary["F11"] = "X1"
    summary["F12"] = "N"
    summary["I10"] = "M. Aguilar"
    summary["I11"] = "Ken Kopf"
    summary["I12"] = "STELLANTIS"
    summary["C14"] = "49754, 50319"
    summary["G14"] = "320767, 321299"
    summary["C20"] = "D2784J"          # devices added
    summary["C22"] = "X350A"           # devices removed
    summary["C26"] = "A937F"           # circuits changed
    summary["C27"] = "F982A"           # circuits removed

    # Two tables on one sheet, exactly as the real workbooks lay it out.
    add_remove = wb.create_sheet("Add_Remove_Report_Summary")
    _write_row(add_remove, 1, ["Connector Add/Remove Report"])
    _write_row(
        add_remove,
        2,
        [
            "SE Comment", "Action", "FCA-CNUM", "Suffix(New)", "Suffix(Old)",
            "DEF_Connector_PN", "DEF_Connector_PN(Old)", "DEF_Connector_Supplier",
        ],
    )
    _write_row(
        add_remove, 3,
        [None, "ADD", "D2784J", "MOD_CMCM", "", "D4Z080-000-B", "", "G3"],
    )
    _write_row(
        add_remove, 4,
        [None, "DELETE", "D2784J", "", "MOD_CMCM", "", "D4K10A-1D5A5-B", ""],
    )
    _write_row(
        add_remove, 5,
        ["DTCR 49919", "DELETE", "X350A", "", "INLINE_X350", "", "6098-8716", ""],
    )
    _write_row(add_remove, 6, ["Circuit Add/Remove Report"])
    _write_row(
        add_remove,
        7,
        [
            "SE Comment", "Action", "CKT NBR", "CKT Suffix", "DEF CKT COLOR",
            "DEF CKT COLOR(Old)", "DEF GAUGE", "DEF GAUGE(Old)",
        ],
    )
    _write_row(add_remove, 8, [None, "DELETE", "F982", "A", "", "PK/YE", "", "0.35"])

    connector = wb.create_sheet("Connector")
    _write_row(
        connector,
        3,
        [
            "SE Comment", "Action", "FCA-CNUM", "Suffix(New)", "Suffix(Old)",
            "DEF_Connector_PN", "DEF_Connector_PN(Old)", "DEF_Connector_Supplier",
        ],
    )
    _write_row(
        connector, 4,
        ["DTCR 50319", "COMP CHG", "SD401", "NEW_SFX", "OLD_SFX", "PN-NEW", "PN-OLD", "DZ"],
    )

    circuit = wb.create_sheet("Circuit")
    _write_row(
        circuit,
        3,
        [
            "SE Comments", "Action", "CKT NBR", "CKT Suffix", "DEF CKT COLOR",
            "DEF CKT COLOR(Old)", "DEF GAUGE", "DEF GAUGE(Old)",
            "CKT FROM (DNUM | CAV)", "Sales_Code", "Sales_Code (Old)",
            # Spelled "To", not "TO", exactly as the real workbooks write it.
            "CKT To (DNUM | CAV)",
        ],
    )
    # Old value in an explicit "(Old)" column.
    _write_row(
        circuit, 4,
        # Two fields changed on one row (colour and gauge), as real rows do.
        ["DTCR 49919", "CHG", "A937", "F", "GN/RD", "BU/WT", "0.50", "0.35",
         "I350X|4", "XZ4", "XZ4", "D2784J|7"],
    )
    # Old value only recoverable from the cell comment, cell marked yellow.
    # The TO cavity is a letter here — a quarter of real cavities are.
    _write_row(
        circuit, 5,
        ["DTCR's 50315, 50317", "CHG", "C205", "", "VT/BU", "VT/BU", "0.35",
         "0.35", "I350X|17", "XZ4", "XZ4", "D3840A|C"],
    )
    from_cell = circuit.cell(row=5, column=9)
    from_cell.fill = YELLOW
    from_cell.comment = Comment("Old DEF :X350A|17", "compare")

    def_def = wb.create_sheet("DEF_DEF_Summary")
    _write_row(
        def_def,
        3,
        [
            "SE Comment", "Action", "Harness PN", "DEF Symbol",
            "Harness PN (Old)", "DEF Symbol (Old)",
        ],
    )
    _write_row(def_def, 4, [None, "CHG", "68774881AB", "18", "68774881AA", "18"])

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

def test_parses_summary_metadata() -> None:
    parsed = parse_secr_bytes(build_secr_workbook(), filename="secr.xlsx")

    assert parsed.secr_number == "D50319A"
    assert parsed.version == "1"
    assert parsed.metadata["harness_family"] == "IP"
    assert parsed.metadata["program"] == "RU"
    assert parsed.metadata["model_year"] == "2028"
    assert parsed.bulletin_numbers == ["320767", "321299"]
    assert "49754" in parsed.dtcr_numbers and "50319" in parsed.dtcr_numbers
    assert parsed.source_sha256  # provenance hash is always computed


def test_delete_and_add_of_one_cnum_becomes_a_pn_change() -> None:
    parsed = parse_secr_bytes(build_secr_workbook())

    pn_changes = [
        c
        for c in parsed.changes
        if c.object_id == "D2784J" and c.field == "DEF_Connector_PN"
    ]
    assert len(pn_changes) == 1
    change = pn_changes[0]
    assert change.action == "PN CHANGE"
    assert change.old_value == "D4K10A-1D5A5-B"
    assert change.new_value == "D4Z080-000-B"


def test_unpaired_connector_delete_is_kept_as_a_delete() -> None:
    parsed = parse_secr_bytes(build_secr_workbook())

    deletes = [c for c in parsed.changes if c.object_id == "X350A"]
    assert len(deletes) == 1
    assert deletes[0].action == "DELETE"
    assert deletes[0].old_value == "6098-8716"
    assert deletes[0].new_value is None
    assert deletes[0].dtcr_number == "49919"


def test_circuit_id_combines_number_and_suffix() -> None:
    parsed = parse_secr_bytes(build_secr_workbook())

    circuit_ids = {c.object_id for c in parsed.changes if c.object_type == "circuit"}
    assert "A937F" in circuit_ids  # CKT NBR "A937" + CKT Suffix "F"
    assert "F982A" in circuit_ids


def test_circuit_old_value_from_paired_column() -> None:
    parsed = parse_secr_bytes(build_secr_workbook())

    gauge = [
        c
        for c in parsed.changes
        if c.object_id == "A937F" and c.field == "DEF GAUGE"
    ]
    assert len(gauge) == 1
    assert gauge[0].old_value == "0.35"
    assert gauge[0].new_value == "0.50"


def test_circuit_old_value_from_cell_comment() -> None:
    parsed = parse_secr_bytes(build_secr_workbook())

    moved = [
        c
        for c in parsed.changes
        if c.object_id == "C205" and c.field == "CKT FROM (DNUM | CAV)"
    ]
    assert len(moved) == 1
    assert moved[0].old_value == "X350A|17"
    assert moved[0].new_value == "I350X|17"
    # The comment names two DTCRs and the row keeps both, so the one written
    # second is still queryable.
    assert moved[0].dtcr_number == "50315, 50317"


def test_add_remove_sheet_circuit_block_is_not_read_as_a_connector() -> None:
    """The sheet holds a connector table and a circuit table; F982A is a circuit."""
    parsed = parse_secr_bytes(build_secr_workbook())

    connectors = {c.object_id for c in parsed.changes if c.object_type == "connector"}
    assert "F982A" not in connectors
    assert "F982" not in connectors


def test_harness_pn_change_is_recorded() -> None:
    parsed = parse_secr_bytes(build_secr_workbook())

    harness = [c for c in parsed.changes if c.object_type == "harness"]
    assert len(harness) == 1
    assert harness[0].old_value == "68774881AA"
    assert harness[0].new_value == "68774881AB"


def test_summary_items_without_a_change_row_are_flagged() -> None:
    """A Summary roll-up entry with no backing change row must not look complete."""
    wb = openpyxl.load_workbook(io.BytesIO(build_secr_workbook()))
    wb["Summary"]["C26"] = "A937F, R372"  # R372 exists nowhere else
    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()

    parsed = parse_secr_bytes(buffer.getvalue())
    assert any("R372" in warning for warning in parsed.warnings)


def test_extract_dtcr_numbers_handles_real_comment_forms() -> None:
    assert extract_dtcr_numbers("DTCR 49793") == ["49793"]
    assert extract_dtcr_numbers("DTCR's 50315, 50317") == ["50315", "50317"]
    # Forms the keyword-anchored pattern used to miss, all taken from the
    # comments in the 27 validation workbooks.
    assert extract_dtcr_numbers("DTCR-50326") == ["50326"]
    assert extract_dtcr_numbers("318725-02_Bulletin DTCR_50073") == ["50073"]
    assert extract_dtcr_numbers("DTCR D50351 & D50298") == ["50351", "50298"]
    assert extract_dtcr_numbers("DTCR 50298 AND 50318") == ["50298", "50318"]
    assert extract_dtcr_numbers(
        "Bulletin 320767: Delete PHEV:49900/49951/49745"
    ) == ["49900", "49951", "49745"]


def test_only_five_digit_numbers_are_dtcrs() -> None:
    """The width is what separates a DTCR from a bulletin or a part number."""
    assert extract_dtcr_numbers("Bulletin 320767") == []      # 6 digits
    assert extract_dtcr_numbers("PN 318725-02") == []         # 6 digits, no cut
    assert extract_dtcr_numbers("issued 06092026") == []      # 8-digit date
    assert extract_dtcr_numbers("CAV 12 gauge 0.35") == []    # too short
    assert extract_dtcr_numbers("X501A|7") == []
    assert extract_dtcr_numbers(
        "XZ4 deleted from PNs per complexity. DTCR 50277"
    ) == ["50277"]
    assert extract_dtcr_numbers("Remove Suffix") == []


def test_non_secr_workbook_is_rejected_with_a_clear_message() -> None:
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()

    with pytest.raises(SpliceSchemaError, match="Summary"):
        parse_secr_bytes(buffer.getvalue(), filename="not_a_secr.xlsx")


def test_empty_and_corrupt_files_are_rejected() -> None:
    with pytest.raises(SpliceSchemaError):
        parse_secr_bytes(b"")
    with pytest.raises(SpliceSchemaError):
        parse_secr_bytes(b"this is not a workbook", filename="junk.xlsx")


# ---------------------------------------------------------------------------
# Persistence
# ---------------------------------------------------------------------------

def test_save_get_and_delete_round_trip(db_path: Path) -> None:
    payload = build_secr_workbook()
    parsed = parse_secr_bytes(payload, filename="secr.xlsx")
    record = secr_db.record_from_parsed(parsed, filename="secr.xlsx")

    secr_id = secr_db.save_secr(record, db_path, source_bytes=payload)
    stored = secr_db.get_secr(secr_id, db_path=db_path)

    assert stored is not None
    assert stored["secr_number"] == "D50319A"
    assert len(stored["changes"]) == parsed.change_count
    assert stored["source_file"]["sha256"] == parsed.source_sha256

    assert secr_db.delete_secr(secr_id, db_path=db_path) is True
    assert secr_db.get_secr(secr_id, db_path=db_path) is None
    assert secr_db.delete_secr(secr_id, db_path=db_path) is False


def test_delete_cascades_to_changes(db_path: Path) -> None:
    payload = build_secr_workbook()
    record = secr_db.record_from_parsed(
        parse_secr_bytes(payload), filename="secr.xlsx"
    )
    secr_id = secr_db.save_secr(record, db_path, source_bytes=payload)
    assert secr_db.get_changes(secr_id, db_path=db_path)

    secr_db.delete_secr(secr_id, db_path=db_path)
    assert secr_db.get_changes(secr_id, db_path=db_path) == []


def test_deletion_is_written_to_the_audit_log(db_path: Path) -> None:
    record = secr_db.record_from_parsed(
        parse_secr_bytes(build_secr_workbook()), filename="secr.xlsx"
    )
    secr_id = secr_db.save_secr(record, db_path)
    secr_db.delete_secr(secr_id, db_path=db_path)

    events = [e["event"] for e in secr_db.get_audit_log(db_path=db_path)]
    assert "deleted" in events
    assert "saved" in events


def test_skip_policy_never_overwrites_a_stored_secr(db_path: Path) -> None:
    first = secr_db.record_from_parsed(
        parse_secr_bytes(build_secr_workbook()), filename="first.xlsx"
    )
    original_id = secr_db.save_secr(first, db_path)

    second = secr_db.record_from_parsed(
        parse_secr_bytes(build_secr_workbook()), filename="second.xlsx"
    )
    returned_id = secr_db.save_secr(
        second, db_path, on_conflict=secr_db.CONFLICT_SKIP
    )

    assert returned_id == original_id
    stored = secr_db.get_secr(original_id, db_path=db_path)
    assert stored is not None and stored["filename"] == "first.xlsx"
    assert len(secr_db.list_secrs(db_path=db_path)) == 1


def test_replace_policy_overwrites_and_is_audited(db_path: Path) -> None:
    secr_db.save_secr(
        secr_db.record_from_parsed(
            parse_secr_bytes(build_secr_workbook()), filename="first.xlsx"
        ),
        db_path,
    )
    secr_db.save_secr(
        secr_db.record_from_parsed(
            parse_secr_bytes(build_secr_workbook()), filename="second.xlsx"
        ),
        db_path,
        on_conflict=secr_db.CONFLICT_REPLACE,
    )

    records = secr_db.list_secrs(db_path=db_path)
    assert len(records) == 1
    assert records[0]["filename"] == "second.xlsx"
    assert "replaced" in [e["event"] for e in secr_db.get_audit_log(db_path=db_path)]


def test_error_policy_raises_on_duplicate(db_path: Path) -> None:
    record = secr_db.record_from_parsed(
        parse_secr_bytes(build_secr_workbook()), filename="secr.xlsx"
    )
    secr_db.save_secr(record, db_path)

    with pytest.raises(secr_db.DuplicateSecrError):
        secr_db.save_secr(record, db_path, on_conflict=secr_db.CONFLICT_ERROR)


def test_different_versions_of_one_secr_coexist(db_path: Path) -> None:
    secr_db.save_secr(
        secr_db.record_from_parsed(
            parse_secr_bytes(build_secr_workbook(version="1")), filename="v1.xlsx"
        ),
        db_path,
    )
    secr_db.save_secr(
        secr_db.record_from_parsed(
            parse_secr_bytes(build_secr_workbook(version="2")), filename="v2.xlsx"
        ),
        db_path,
    )

    assert len(secr_db.list_secrs(db_path=db_path)) == 2


def test_schema_migrates_a_v1_database_without_losing_rows(db_path: Path) -> None:
    """A database created before change tracking keeps its rows and gains v2."""
    with secr_db.get_conn(db_path) as conn:
        conn.executescript(secr_db._SCHEMA)
        conn.execute("PRAGMA user_version = 1")
        conn.execute(
            "INSERT INTO secr (secr_number, version, action) VALUES (?,?,?)",
            ("LEGACY1", "A", "create"),
        )

    secr_db.init_db(db_path)

    with secr_db.get_conn(db_path) as conn:
        assert conn.execute("PRAGMA user_version").fetchone()[0] == secr_db.SCHEMA_VERSION
        assert conn.execute("SELECT COUNT(*) FROM secr").fetchone()[0] == 1
        columns = {row["name"] for row in conn.execute("PRAGMA table_info(secr)")}
    assert {
        "import_origin", "source_sha256", "parse_warnings",
        "secr_sequence_number", "scope_model_year", "scope_phase",
        "version_number", "generation_date",
    } <= columns

    # And the migrated database still accepts new records with changes.
    secr_id = secr_db.save_secr(
        secr_db.record_from_parsed(
            parse_secr_bytes(build_secr_workbook()), filename="secr.xlsx"
        ),
        db_path,
    )
    assert secr_db.get_changes(secr_id, db_path=db_path)


def test_data_survives_reopening_the_database(db_path: Path) -> None:
    """Persistence across an application restart — a new connection sees it."""
    secr_db.save_secr(
        secr_db.record_from_parsed(
            parse_secr_bytes(build_secr_workbook()), filename="secr.xlsx"
        ),
        db_path,
    )
    del_conn = secr_db.get_conn(db_path)
    del_conn.close()

    assert len(secr_db.list_secrs(db_path=db_path)) == 1
    assert api.get_secr_summary("D50319A", db_path=db_path) is not None


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

def test_bulk_import_reports_every_file(db_path: Path) -> None:
    files = [
        ("a.xlsx", build_secr_workbook(secr_number="D1001")),
        ("b.xlsx", build_secr_workbook(secr_number="D1002")),
        ("duplicate.xlsx", build_secr_workbook(secr_number="D1001")),
        ("broken.xlsx", b"not a workbook"),
    ]

    summary = import_secr_files(files, db_path=db_path)

    assert summary.processed == 4
    assert len(summary.imported) == 2
    assert len(summary.duplicates) == 1
    assert len(summary.failed) == 1
    assert summary.failed[0].filename == "broken.xlsx"
    assert summary.failed[0].message  # the reason is always reported
    assert summary.headline() == "4 SECR files processed"


def test_import_stores_change_records_and_source_file(db_path: Path) -> None:
    payload = build_secr_workbook()
    result = import_secr_bytes(payload, "secr.xlsx", db_path=db_path)

    assert result.status == STATUS_IMPORTED
    assert result.change_count > 0
    stored = secr_db.get_secr(result.secr_id, db_path=db_path)
    assert stored["import_origin"] == "imported"
    source = secr_db.get_source_file(result.secr_id, db_path=db_path)
    assert source["content"] == payload


def test_import_can_skip_storing_the_source_file(db_path: Path) -> None:
    result = import_secr_bytes(
        build_secr_workbook(), "secr.xlsx", store_source=False, db_path=db_path
    )
    assert secr_db.get_source_file(result.secr_id, db_path=db_path) is None


def test_reimporting_the_same_file_does_not_duplicate_changes(db_path: Path) -> None:
    payload = build_secr_workbook()
    first = import_secr_bytes(payload, "secr.xlsx", db_path=db_path)
    second = import_secr_bytes(payload, "secr.xlsx", db_path=db_path)

    assert second.status == STATUS_DUPLICATE
    assert second.secr_id == first.secr_id
    assert len(secr_db.get_changes(first.secr_id, db_path=db_path)) == first.change_count


def test_a_corrupt_file_does_not_abort_the_run(db_path: Path) -> None:
    summary = import_secr_files(
        [
            ("broken.xlsx", b"garbage"),
            ("good.xlsx", build_secr_workbook()),
        ],
        db_path=db_path,
    )

    assert [r.status for r in summary.results] == [STATUS_FAILED, STATUS_IMPORTED]
    assert len(secr_db.list_secrs(db_path=db_path)) == 1


def test_import_progress_is_reported(db_path: Path) -> None:
    seen: list[tuple[int, int]] = []
    import_secr_files(
        [
            ("a.xlsx", build_secr_workbook(secr_number="D1")),
            ("b.xlsx", build_secr_workbook(secr_number="D2")),
        ],
        progress=lambda done, total, name: seen.append((done, total)),
        db_path=db_path,
    )
    assert seen == [(1, 2), (2, 2)]


# ---------------------------------------------------------------------------
# Query API
# ---------------------------------------------------------------------------

@pytest.fixture()
def populated_db(db_path: Path) -> Path:
    import_secr_files(
        [
            ("ip.xlsx", build_secr_workbook(secr_number="D50319A")),
            (
                "body.xlsx",
                build_secr_workbook(
                    secr_number="D49957A", harness_family="BODY", model_year="2027"
                ),
            ),
        ],
        db_path=db_path,
    )
    return db_path


def test_search_finds_a_secr_by_connector_number(populated_db: Path) -> None:
    hits = api.search_secrs("D2784J", db_path=populated_db)

    assert {h["secr_number"] for h in hits} == {"D50319A", "D49957A"}
    assert "connector D2784J" in hits[0]["match_reason"]


def test_search_by_circuit_prefix_matches_the_suffixed_circuit(
    populated_db: Path,
) -> None:
    """Searching A937 must find circuit A937F."""
    hits = api.search_secrs("A937", db_path=populated_db)
    assert hits
    assert "circuit A937F" in hits[0]["match_reason"]


def test_filters_combine(populated_db: Path) -> None:
    hits = api.search_secrs(
        harness_family="BODY", model_year="2027", db_path=populated_db
    )
    assert [h["secr_number"] for h in hits] == ["D49957A"]

    assert api.search_secrs(
        harness_family="BODY", model_year="2028", db_path=populated_db
    ) == []


def test_search_by_dtcr(populated_db: Path) -> None:
    hits = api.search_secrs(dtcr="49919", db_path=populated_db)
    assert {h["secr_number"] for h in hits} == {"D50319A", "D49957A"}


def test_change_lookups_by_object(populated_db: Path) -> None:
    by_cnum = api.get_changes_by_cnum("D2784J", db_path=populated_db)
    # Both roles come back; the connector's own change is still there.
    own = [c for c in by_cnum if c["connector_role"] == "connector"]
    assert own and all(c["object_type"] == "connector" for c in own)
    assert own[0]["old_value"] == "D4K10A-1D5A5-B"

    by_circuit = api.get_changes_by_circuit("A937F", db_path=populated_db)
    assert by_circuit and all(c["object_type"] == "circuit" for c in by_circuit)

    by_dtcr = api.get_changes_by_dtcr("49919", db_path=populated_db)
    assert by_dtcr and all(c["dtcr_number"] == "49919" for c in by_dtcr)

    by_harness = api.get_changes_by_harness("BODY", db_path=populated_db)
    assert by_harness and all(c["harness_family"] == "BODY" for c in by_harness)


def test_connector_pn_lookup_matches_old_and_new_values(populated_db: Path) -> None:
    old_side = api.get_connector_changes("D4K10A-1D5A5-B", db_path=populated_db)
    new_side = api.get_connector_changes("D4Z080-000-B", db_path=populated_db)
    assert old_side and new_side


def test_secr_summary_reports_change_counts(populated_db: Path) -> None:
    summary = api.get_secr_summary("D50319A", db_path=populated_db)

    assert summary is not None
    assert summary["harness_family"] == "IP"
    assert summary["change_count"] > 0
    assert "PN CHANGE" in summary["changes_by_action"]
    assert api.get_secr_summary("NOT-A-SECR", db_path=populated_db) is None


def test_program_and_model_year_summaries(populated_db: Path) -> None:
    program = api.get_program_summary("RU", db_path=populated_db)
    assert program["secr_count"] == 2
    assert program["change_count"] > 0

    year = api.get_model_year_summary("2027", db_path=populated_db)
    assert year["secr_numbers"] == ["D49957A"]


def test_database_summary_breakdowns(populated_db: Path) -> None:
    summary = api.get_database_summary(db_path=populated_db)

    assert summary["totals"]["secrs"] == 2
    assert summary["totals"]["changes"] > 0
    families = {row["name"] for row in summary["by_harness_family"]}
    assert families == {"IP", "BODY"}


def test_read_only_tool_surface_exposes_no_writes() -> None:
    """The assistant-facing surface must not expose a way to modify data.

    No tool may write, and none may take raw SQL — a future local assistant
    calls these and nothing else.
    """
    write_functions = {
        "save_secr", "delete_secr", "init_db", "get_conn", "record_from_parsed",
        "record_from_workbook", "import_secr_bytes", "import_secr_files",
        "import_folder",
    }
    for name, function in api.READ_ONLY_TOOLS.items():
        assert name not in write_functions, name
        assert getattr(api, name, None) is function, name
        assert "sql" not in function.__code__.co_varnames, name

    for name in write_functions:
        assert not hasattr(api, name), f"api re-exports the write function {name}"


def test_query_api_does_not_modify_the_database(populated_db: Path) -> None:
    before = [dict(r) for r in secr_db.list_secrs(db_path=populated_db)]
    change_count = secr_db.database_summary(db_path=populated_db)["totals"]["changes"]

    api.search_secrs("D2784J", db_path=populated_db)
    api.get_changes_by_dtcr("49919", db_path=populated_db)
    api.get_program_summary("RU", db_path=populated_db)
    api.get_database_summary(db_path=populated_db)

    assert [dict(r) for r in secr_db.list_secrs(db_path=populated_db)] == before
    assert (
        secr_db.database_summary(db_path=populated_db)["totals"]["changes"]
        == change_count
    )


# ---------------------------------------------------------------------------
# File handles
# ---------------------------------------------------------------------------

def test_operations_leave_no_open_file_handles(tmp_path: Path) -> None:
    """Every database call must close its connection.

    ``with sqlite3.connect(...) as conn`` manages the *transaction* and leaves
    the handle open. On macOS/Linux that is invisible — an open file can still
    be deleted — but on Windows it locks the database and its WAL sidecars, so
    a backup, a replace, or a temp-directory cleanup fails outright. The WAL
    and shared-memory files are removed by SQLite when the last connection
    closes, which makes their absence a portable check for "handle released".
    """
    db_path = tmp_path / "handles.db"
    payload = build_secr_workbook()

    import_secr_bytes(payload, "secr.xlsx", db_path=db_path)
    secr_db.list_secrs(db_path=db_path)
    secr_db.search_secrs("D2784J", db_path=db_path)
    secr_db.change_facets(db_path=db_path)
    secr_db.reserve_next_secr_number("2028", "X1", db_path=db_path)

    leftovers = [
        path.name
        for path in tmp_path.iterdir()
        if path.name.endswith(("-wal", "-shm"))
    ]
    assert not leftovers, f"database handles still open: {leftovers}"


def test_a_directory_holding_the_database_can_be_deleted(tmp_path: Path) -> None:
    """The exact failure the Windows build hit: temp cleanup after a db call."""
    import shutil
    import tempfile

    with tempfile.TemporaryDirectory(dir=tmp_path) as temp_dir:
        db_path = Path(temp_dir) / "probe.db"
        secr_db.init_db(db_path)
        secr_db.list_secrs(db_path=db_path)
        assert db_path.is_file()
    # Exiting the context deletes the directory; on Windows that raises
    # PermissionError if any connection is still open.

    workspace = tmp_path / "workspace"
    workspace.mkdir()
    db_path = workspace / "probe.db"
    import_secr_bytes(build_secr_workbook(), "secr.xlsx", db_path=db_path)
    shutil.rmtree(workspace)
    assert not workspace.exists()


# ---------------------------------------------------------------------------
# Circuit endpoints (schema v4)
# ---------------------------------------------------------------------------

def test_split_endpoint_reads_dnum_and_cavity() -> None:
    assert split_endpoint("X501A|7") == ("X501A", "7")
    assert split_endpoint(" x501a | 7 ") == ("X501A", "7")
    assert split_endpoint("D3840A|C") == ("D3840A", "C")   # lettered cavity
    assert split_endpoint("no pipe here") == (None, None)
    assert split_endpoint("") == (None, None)
    assert split_endpoint(None) == (None, None)


def test_circuit_changes_carry_their_endpoints() -> None:
    """A DTCR reaches a circuit *and* the connectors and cavities it lands on."""
    parsed = parse_secr_bytes(build_secr_workbook())

    a937f = [c for c in parsed.changes if c.object_id == "A937F"]
    assert a937f, "the circuit was not parsed"
    for change in a937f:
        assert change.from_dnum == "I350X" and change.from_cav == "4"
        assert change.to_dnum == "D2784J" and change.to_cav == "7"
        assert change.dtcr_number == "49919"


def test_every_changed_field_of_a_circuit_row_shares_its_endpoints() -> None:
    """The endpoints belong to the row, so each field change carries them."""
    parsed = parse_secr_bytes(build_secr_workbook())

    a937f = [c for c in parsed.changes if c.object_id == "A937F"]
    assert {c.field for c in a937f} == {"DEF CKT COLOR", "DEF GAUGE"}
    assert {(c.from_dnum, c.from_cav) for c in a937f} == {("I350X", "4")}
    assert {(c.to_dnum, c.to_cav) for c in a937f} == {("D2784J", "7")}


def test_connectors_do_not_get_endpoints() -> None:
    """Only circuits terminate; a connector change has none."""
    parsed = parse_secr_bytes(build_secr_workbook())

    for change in parsed.changes:
        if change.object_type != "circuit":
            assert change.from_dnum is None and change.to_dnum is None


def test_endpoints_are_stored_and_queryable(db_path: Path) -> None:
    import_secr_files([("secr.xlsx", build_secr_workbook())], db_path=db_path)

    with secr_db.connect(db_path) as conn:
        assert conn.execute("PRAGMA user_version").fetchone()[0] >= 4
        row = conn.execute(
            "SELECT from_dnum, from_cav, to_dnum, to_cav FROM secr_change"
            " WHERE object_id = 'A937F' LIMIT 1"
        ).fetchone()
    assert (row["from_dnum"], row["from_cav"]) == ("I350X", "4")
    assert (row["to_dnum"], row["to_cav"]) == ("D2784J", "7")

    # The question that used to be unanswerable: D2784J never appears on the
    # Connector sheet here, only as the far end of a circuit.
    landed = secr_db.find_changes(endpoint="D2784J", db_path=db_path)
    assert landed and all(r["object_type"] == "circuit" for r in landed)
    assert secr_db.find_changes(endpoint="d2784j", db_path=db_path) == landed

    at_cavity = secr_db.find_changes(
        endpoint="D3840A", cavity="C", db_path=db_path
    )
    assert at_cavity and all(r["object_id"] == "C205" for r in at_cavity)
    assert secr_db.find_changes(
        endpoint="D3840A", cavity="99", db_path=db_path
    ) == []


def test_a_connector_query_returns_both_roles_labelled(db_path: Path) -> None:
    """"Anything involving D2784J" spans the connector and the circuits on it."""
    import_secr_files([("secr.xlsx", build_secr_workbook())], db_path=db_path)

    rows = api.get_changes_by_cnum("D2784J", db_path=db_path)
    roles = {r["connector_role"] for r in rows}
    assert "endpoint" in roles, "circuits landing on the connector were missed"

    endpoints = [r for r in rows if r["connector_role"] == "endpoint"]
    assert all(r["object_type"] == "circuit" for r in endpoints)
    assert {r["endpoint_side"] for r in endpoints} == {"TO"}
    assert {r["endpoint_cavity"] for r in endpoints} == {"7"}


def test_the_cnum_filter_still_means_the_object_only(db_path: Path) -> None:
    """Browse and the Dashboard count connector *objects* — unchanged by v4."""
    import_secr_files([("secr.xlsx", build_secr_workbook())], db_path=db_path)

    only_objects = secr_db.find_changes(cnum="D2784J", db_path=db_path)
    assert all(r["object_type"] == "connector" for r in only_objects)


# ---------------------------------------------------------------------------
# Pin history: one connector + one cavity, over time (schema v5)
# ---------------------------------------------------------------------------

def test_a_pin_query_pairs_the_connector_with_its_own_cavity(
    db_path: Path,
) -> None:
    """The two ends have different cavities; a pin is one specific pair.

    A937F runs I350X|4 -> D2784J|7. Asking for "I350X cavity 7" must find
    nothing: cavity 7 belongs to the other end. Testing
    (from_dnum OR to_dnum) AND (from_cav OR to_cav) would wrongly match.
    """
    import_secr_files([("secr.xlsx", build_secr_workbook())], db_path=db_path)

    assert api.get_changes_by_endpoint("I350X", "4", db_path=db_path)
    assert api.get_changes_by_endpoint("D2784J", "7", db_path=db_path)
    assert api.get_changes_by_endpoint("I350X", "7", db_path=db_path) == []
    assert api.get_changes_by_endpoint("D2784J", "4", db_path=db_path) == []


def test_a_circuit_leaving_a_pin_is_a_change_to_that_pin(db_path: Path) -> None:
    """C205 moved off X350A|17 onto I350X|17.

    X350A is not where the circuit ended up, so v4 lost it entirely. But
    "has X350A cavity 17 changed?" is exactly the question an engineer asks
    when something disappeared from that pin.
    """
    import_secr_files([("secr.xlsx", build_secr_workbook())], db_path=db_path)

    left = api.get_changes_by_endpoint("X350A", "17", db_path=db_path)
    assert left, "the circuit that moved off this pin was not found"
    assert {r["object_id"] for r in left} == {"C205"}
    assert all(r["endpoint_state"] == "left" for r in left)

    arrived = api.get_changes_by_endpoint("I350X", "17", db_path=db_path)
    assert {r["object_id"] for r in arrived} == {"C205"}
    assert all(r["endpoint_state"] == "arrived" for r in arrived)


def test_a_pin_that_did_not_move_reads_as_present(db_path: Path) -> None:
    """A937F's endpoints are unchanged; its colour and gauge are what moved."""
    import_secr_files([("secr.xlsx", build_secr_workbook())], db_path=db_path)

    rows = api.get_changes_by_endpoint("D2784J", "7", db_path=db_path)
    assert rows and all(r["endpoint_state"] == "present" for r in rows)
    assert all(r["endpoint_side"] == "TO" for r in rows)


def test_pin_history_carries_the_phase_of_each_change(db_path: Path) -> None:
    """The same pin can hold different circuits in different phases.

    Two SECRs in different phases both touch D2784J|7; the answer must keep
    them apart rather than merge them into one contradictory statement.
    """
    import_secr_files([("x1.xlsx", build_secr_workbook())], db_path=db_path)
    later = build_secr_workbook(secr_number="D50320A")
    workbook = openpyxl.load_workbook(io.BytesIO(later))
    workbook["Summary"]["F10"] = "X2"
    workbook["Circuit"].cell(row=4, column=3, value="B111")   # a different CKT
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    import_secr_files([("x2.xlsx", buffer.getvalue())], db_path=db_path)

    rows = api.get_changes_by_endpoint("D2784J", "7", db_path=db_path)
    by_phase = {(r["phase"], r["object_id"]) for r in rows}
    assert ("X1", "A937F") in by_phase
    assert ("X2", "B111F") in by_phase


def test_the_cavity_must_match_exactly(db_path: Path) -> None:
    import_secr_files([("secr.xlsx", build_secr_workbook())], db_path=db_path)

    assert api.get_changes_by_endpoint("D3840A", "C", db_path=db_path)
    assert api.get_changes_by_endpoint("D3840A", "c", db_path=db_path)
    assert api.get_changes_by_endpoint("D3840A", "1", db_path=db_path) == []
    # Without a cavity, every pin of the connector comes back.
    assert api.get_changes_by_endpoint("D3840A", db_path=db_path)


# ---------------------------------------------------------------------------
# DTCRs listed on a SECR but on no change row (field report, 2026-08-12)
# ---------------------------------------------------------------------------

def test_a_dtcr_listed_only_on_the_secr_still_answers(db_path: Path) -> None:
    """The fixture's Summary lists 49754; no SE comment mentions it.

    About half the DTCRs in a real database are like this — written on the
    Summary sheet and never into an individual row's comment. Returning
    nothing for them made the assistant silent about DTCRs the engineer could
    see on screen, twice reported from the field.
    """
    import_secr_files([("secr.xlsx", build_secr_workbook())], db_path=db_path)

    rows = api.get_changes_by_dtcr("49754", db_path=db_path)

    assert rows, "a DTCR listed on the SECR came back empty"
    assert all(row["dtcr_attribution"] == "secr" for row in rows)
    assert rows[0]["secr_number"] == "D50319A"
    assert rows[0]["harness_family"] == "IP"
    assert rows[0]["dtcr_number"] == "49754"


def test_a_dtcr_on_change_rows_still_returns_those_rows(db_path: Path) -> None:
    """The precise answer wins whenever it exists — the fallback is a fallback."""
    import_secr_files([("secr.xlsx", build_secr_workbook())], db_path=db_path)

    rows = api.get_changes_by_dtcr("49919", db_path=db_path)

    assert rows
    assert all(row["dtcr_attribution"] == "change" for row in rows)
    assert all(row["dtcr_number"] == "49919" for row in rows)
    assert {row["object_type"] for row in rows} <= {"connector", "circuit"}


def test_an_unknown_dtcr_returns_nothing(db_path: Path) -> None:
    """The fallback must not turn "no record" into a false positive."""
    import_secr_files([("secr.xlsx", build_secr_workbook())], db_path=db_path)

    assert api.get_changes_by_dtcr("99999", db_path=db_path) == []
    assert api.get_changes_by_dtcr("", db_path=db_path) == []
