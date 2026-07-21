from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dtx_compare_engine import generate_preorder_generation_workbook


REQUIRED_COLUMNS = [
    "Device Control Number",
    "Device Name",
    "Suffix",
    "CNUM",
    "Number of Cavities",
    "Connector PN",
    "Harness Family",
    "Pin Number",
    "Circuit Name",
    "Circuit Suffix",
    "Circuit Function",
    "Color",
    "Terminal",
    "Connector FCA part number",
    "Wire Gauge",
    "Wire Type",
    "Sales Code",
]


def _write_dtx_excel(path: Path, rows: list[dict[str, object]]) -> None:
    frame = pd.DataFrame(rows, columns=REQUIRED_COLUMNS)
    with pd.ExcelWriter(path) as writer:
        frame.to_excel(writer, index=False)


def test_generate_preorder_generation_workbook_creates_excel_bytes(tmp_path: Path) -> None:
    old_path = tmp_path / "old.xlsx"
    new_path = tmp_path / "new.xlsx"
    _write_dtx_excel(
        old_path,
        [
            {
                "Device Control Number": "DCN1",
                "Device Name": "Device A",
                "Suffix": "A",
                "CNUM": "C1",
                "Number of Cavities": "2",
                "Connector PN": "PN-OLD",
                "Harness Family": "HF-1",
                "Pin Number": "1",
                "Circuit Name": "CIRCUIT1",
                "Circuit Suffix": "S1",
                "Circuit Function": "FUNC",
                "Color": "RED",
                "Terminal": "T1",
                "Connector FCA part number": "FCA1",
                "Wire Gauge": "18",
                "Wire Type": "W1",
                "Sales Code": "S1",
            }
        ],
    )
    _write_dtx_excel(
        new_path,
        [
            {
                "Device Control Number": "DCN1",
                "Device Name": "Device A",
                "Suffix": "A",
                "CNUM": "C1",
                "Number of Cavities": "2",
                "Connector PN": "PN-NEW",
                "Harness Family": "HF-1",
                "Pin Number": "1",
                "Circuit Name": "CIRCUIT1",
                "Circuit Suffix": "S1",
                "Circuit Function": "FUNC",
                "Color": "RED",
                "Terminal": "T1",
                "Connector FCA part number": "FCA1",
                "Wire Gauge": "18",
                "Wire Type": "W1",
                "Sales Code": "S1",
            },
            {
                "Device Control Number": "DCN2",
                "Device Name": "Device B",
                "Suffix": "B",
                "CNUM": "C2",
                "Number of Cavities": "4",
                "Connector PN": "PN-NEW2",
                "Harness Family": "HF-2",
                "Pin Number": "2",
                "Circuit Name": "CIRCUIT2",
                "Circuit Suffix": "S2",
                "Circuit Function": "FUNC2",
                "Color": "BLUE",
                "Terminal": "T2",
                "Connector FCA part number": "FCA2",
                "Wire Gauge": "20",
                "Wire Type": "W2",
                "Sales Code": "S2",
            },
        ],
    )

    result = generate_preorder_generation_workbook(
        old_file_bytes=old_path.read_bytes(),
        new_file_bytes=new_path.read_bytes(),
        old_file_name=old_path.name,
        new_file_name=new_path.name,
    )

    assert result["output_excel_bytes"]
    assert "summary_df" in result
    assert "connector_changes_df" in result
    assert result["summary_df"].shape[0] >= 1
    assert any(result["summary_df"]["Change Type"] == "Connector PN Change")


def test_generate_preorder_generation_workbook_matches_sample_layout(tmp_path: Path) -> None:
    old_path = tmp_path / "old.xlsx"
    new_path = tmp_path / "new.xlsx"
    _write_dtx_excel(
        old_path,
        [
            {
                "Device Control Number": "DCN1",
                "Device Name": "Device A",
                "Suffix": "A",
                "CNUM": "C1",
                "Number of Cavities": "2",
                "Connector PN": "PN-OLD",
                "Harness Family": "HF-1",
                "Pin Number": "1",
                "Circuit Name": "CIRCUIT1",
                "Circuit Suffix": "S1",
                "Circuit Function": "FUNC",
                "Color": "RED",
                "Terminal": "T1",
                "Connector FCA part number": "FCA1",
                "Wire Gauge": "18",
                "Wire Type": "W1",
                "Sales Code": "S1",
            }
        ],
    )
    _write_dtx_excel(
        new_path,
        [
            {
                "Device Control Number": "DCN1",
                "Device Name": "Device A",
                "Suffix": "A",
                "CNUM": "C1",
                "Number of Cavities": "2",
                "Connector PN": "PN-NEW",
                "Harness Family": "HF-1",
                "Pin Number": "1",
                "Circuit Name": "CIRCUIT1",
                "Circuit Suffix": "S1",
                "Circuit Function": "FUNC",
                "Color": "RED",
                "Terminal": "T1",
                "Connector FCA part number": "FCA1",
                "Wire Gauge": "18",
                "Wire Type": "W1",
                "Sales Code": "S1",
            }
        ],
    )

    result = generate_preorder_generation_workbook(
        old_file_bytes=old_path.read_bytes(),
        new_file_bytes=new_path.read_bytes(),
        old_file_name=old_path.name,
        new_file_name=new_path.name,
    )

    workbook = pd.ExcelFile(BytesIO(result["output_excel_bytes"]))
    assert workbook.sheet_names == ["Connector Changes", "Summary"]

    connector_changes = pd.read_excel(BytesIO(result["output_excel_bytes"]), sheet_name="Connector Changes", header=None)
    assert connector_changes.iloc[5, 0] == "CNUM_Device Name-Suffix (Device Control Number)"
    assert connector_changes.iloc[5, 4] == "Connector PN Change"
    assert connector_changes.iloc[5, 6] == "Change Type"

    workbook = load_workbook(BytesIO(result["output_excel_bytes"]))
    connector_ws = workbook["Connector Changes"]
    assert connector_ws["A6"].fill.fgColor.rgb == "004472C4"
    assert connector_ws["A6"].font.bold is True
    assert connector_ws["A6"].border.top.style == "thick"
    assert connector_ws["A6"].border.left.style == "thick"
    assert connector_ws["E7"].fill.fgColor.rgb == "00FFFF00"

    summary = pd.read_excel(BytesIO(result["output_excel_bytes"]), sheet_name="Summary", header=None)
    assert summary.iloc[5, 0] == "CNUM"
    assert summary.iloc[5, 1] == "Connector PN Change"
    assert summary.iloc[5, 2] == "Harness Family"
    assert summary.iloc[5, 3] == "Change Type"
    assert workbook["Summary"]["B7"].fill.fgColor.rgb == "00FFFF00"


def test_generate_preorder_generation_workbook_matches_real_dtx_reports() -> None:
    dtx_dir = Path(__file__).resolve().parents[1].parent / "DTx"
    old_path = dtx_dir / "1 2027 KM AWDV1_FWDX2 DetailedDTxCircuitsReport_revA_EC.xls"
    new_path = dtx_dir / "2 2028 KM X2_49-X1_74 DetailedDTxCircuitsReport_revA_ec.xls"

    if not old_path.exists() or not new_path.exists():
        pytest.skip("Real DTx fixture files are not available in this environment.")

    result = generate_preorder_generation_workbook(
        old_file_bytes=old_path.read_bytes(),
        new_file_bytes=new_path.read_bytes(),
        old_file_name=old_path.name,
        new_file_name=new_path.name,
    )

    assert len(result["summary_df"]) >= 10
    assert result["summary_df"].iloc[0, 0] == "D1606B"
    assert result["output_file_name"].startswith("PreOrderList_")
    assert "2028KM" in result["output_file_name"]
    assert "X2_49" in result["output_file_name"]
    assert result["output_file_name"].endswith(".xlsx")

    workbook = load_workbook(BytesIO(result["output_excel_bytes"]))
    assert workbook["Connector Changes"]["A2"].value == "Vehicle Program - 2027KM"
    assert workbook["Connector Changes"]["H2"].value == "Vehicle Program - 2028KM"


def test_generate_preorder_generation_workbook_reports_deleted_and_added_connectors(tmp_path: Path) -> None:
    old_path = tmp_path / "old.xlsx"
    new_path = tmp_path / "new.xlsx"
    _write_dtx_excel(
        old_path,
        [
            {
                "Device Control Number": "DCN1",
                "Device Name": "Device A",
                "Suffix": "A",
                "CNUM": "C1",
                "Number of Cavities": "2",
                "Connector PN": "PN-OLD",
                "Harness Family": "HF-1",
                "Pin Number": "1",
                "Circuit Name": "CIRCUIT1",
                "Circuit Suffix": "S1",
                "Circuit Function": "FUNC",
                "Color": "RED",
                "Terminal": "T1",
                "Connector FCA part number": "FCA1",
                "Wire Gauge": "18",
                "Wire Type": "W1",
                "Sales Code": "S1",
            }
        ],
    )
    _write_dtx_excel(
        new_path,
        [
            {
                "Device Control Number": "DCN2",
                "Device Name": "Device B",
                "Suffix": "B",
                "CNUM": "C2",
                "Number of Cavities": "4",
                "Connector PN": "PN-NEW",
                "Harness Family": "HF-2",
                "Pin Number": "2",
                "Circuit Name": "CIRCUIT2",
                "Circuit Suffix": "S2",
                "Circuit Function": "FUNC2",
                "Color": "BLUE",
                "Terminal": "T2",
                "Connector FCA part number": "FCA2",
                "Wire Gauge": "20",
                "Wire Type": "W2",
                "Sales Code": "S2",
            }
        ],
    )

    result = generate_preorder_generation_workbook(
        old_file_bytes=old_path.read_bytes(),
        new_file_bytes=new_path.read_bytes(),
        old_file_name=old_path.name,
        new_file_name=new_path.name,
    )

    assert any(result["connector_changes_df"]["Change Type"] == "Deleted")
    assert any(result["connector_changes_df"]["Change Type"] == "Added")
    assert any(result["summary_df"]["Change Type"] == "Deleted")
    assert any(result["summary_df"]["Change Type"] == "Added")


def test_generate_preorder_generation_workbook_requires_valid_input_files(tmp_path: Path) -> None:
    old_path = tmp_path / "old.xlsx"
    new_path = tmp_path / "new.xlsx"
    old_path.write_bytes(b"not an excel file")
    new_path.write_bytes(b"not an excel file")

    with pytest.raises(ValueError):
        generate_preorder_generation_workbook(
            old_file_bytes=old_path.read_bytes(),
            new_file_bytes=new_path.read_bytes(),
            old_file_name=old_path.name,
            new_file_name=new_path.name,
        )


def test_preorder_uses_cnum_within_harness_as_connector_identity(tmp_path: Path) -> None:
    old_path = tmp_path / "old.xlsx"
    new_path = tmp_path / "new.xlsx"

    def row(cnum: str, suffix: str, connector_pn: str) -> dict[str, object]:
        values = {column: "" for column in REQUIRED_COLUMNS}
        values.update({
            "Device Control Number": f"DCN-{cnum}",
            "Device Name": "Old Label" if suffix == "OLD" else "New Label",
            "Suffix": suffix,
            "CNUM": cnum,
            "Number of Cavities": "4",
            "Connector PN": connector_pn,
            "Harness Family": "DASH",
            "Pin Number": "1",
        })
        return values

    _write_dtx_excel(old_path, [row("D3821C", "OLD", "PN-SAME"), row("Y200A", "OLD", "PN-OLD")])
    _write_dtx_excel(new_path, [row("D3821C", "NEW", "PN-SAME"), row("Y200A", "NEW", "PN-NEW")])

    result = generate_preorder_generation_workbook(
        old_file_bytes=old_path.read_bytes(),
        new_file_bytes=new_path.read_bytes(),
        old_file_name=old_path.name,
        new_file_name=new_path.name,
    )

    assert "D3821C" not in set(result["summary_df"]["CNUM"])
    y200 = result["summary_df"].loc[result["summary_df"]["CNUM"] == "Y200A"]
    assert len(y200) == 1
    assert y200.iloc[0]["Change Type"] == "Connector PN Change"
    assert y200.iloc[0]["Connector PN Change"] == "PN-OLD >> PN-NEW"
