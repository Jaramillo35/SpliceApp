"""Tests for the DTCR Matching Report library.

The library exists so a report is uploaded once per program + model year +
phase instead of attached to every SECR. What has to hold: a report is found
again from any spelling of its scope, re-uploading replaces rather than
duplicates, multi-family cells are counted as several families, and "unmatched"
means what an engineer thinks it means.

Report fixtures are built in memory, so nothing here needs the real file.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import pytest

from secrdb.core.common.errors import SpliceInputError
from secrdb.core.dtcr import library

_HEADERS = [
    "DTCR#",
    "Device Transmittal",
    "Extracted Device Control Number",
    "Reason for change",
    "Status",
    "Bulletin",
    "Match Method",
    "Matched DTx Value",
    "CNUM",
    "Harness Family",
]


def build_report(rows: Optional[List[Dict[str, Any]]] = None) -> bytes:
    """A DTCR Matching Report shaped like the real export."""
    if rows is None:
        rows = [
            {
                "DTCR#": "50868", "Device Transmittal": "115899-Sw_Seat",
                "Extracted Device Control Number": "115899",
                "Reason for change": "Seat belt warning", "Status": "Complete",
                "Bulletin": "320880-02", "Match Method": "Device Control Number",
                "Matched DTx Value": "115899", "CNUM": "D2360A",
                "Harness Family": "SEAT_3RD_ROW_LEFT",
            },
            {
                # One DTCR, several harness families in one cell.
                "DTCR#": "50888", "Device Transmittal": "71929-System",
                "Extracted Device Control Number": "71929",
                "Reason for change": "Complexity", "Status": "Complete",
                "Bulletin": "", "Match Method": "Python Script",
                "Matched DTx Value": "", "CNUM": "",
                "Harness Family": "BODY_RIGHT, BODY_LEFT",
            },
            {
                # Mixed case: must land on the same bar as BODY_LEFT above.
                "DTCR#": "50752", "Device Transmittal": "74904-Mod",
                "Extracted Device Control Number": "74904",
                "Reason for change": "", "Status": "Complete",
                "Bulletin": "320880-02", "Match Method": "Device Control Number",
                "Matched DTx Value": "74904", "CNUM": "D3920A",
                "Harness Family": "Body_Left",
            },
            {
                # No family at all — the real gap.
                "DTCR#": "50672", "Device Transmittal": "99999-Unknown",
                "Extracted Device Control Number": "99999",
                "Reason for change": "", "Status": "Rejected",
                "Bulletin": "", "Match Method": "No Match",
                "Matched DTx Value": "", "CNUM": "", "Harness Family": "",
            },
            {
                # "No Match" as a method, but the families were resolved anyway.
                "DTCR#": "50889", "Device Transmittal": "88888-Global",
                "Extracted Device Control Number": "88888",
                "Reason for change": "", "Status": "Complete",
                "Bulletin": "", "Match Method": "No Match",
                "Matched DTx Value": "", "CNUM": "D5900A,D5900B",
                "Harness Family": "IP, DASH",
            },
        ]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "DTCR_Harness_Family_Mapping"
    for column, header in enumerate(_HEADERS, start=1):
        sheet.cell(row=1, column=column, value=header)
    for index, row in enumerate(rows, start=2):
        for column, header in enumerate(_HEADERS, start=1):
            sheet.cell(row=index, column=column, value=row.get(header, ""))
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


@pytest.fixture()
def db_path(tmp_path: Path) -> Path:
    return tmp_path / "library.db"


@pytest.fixture()
def filed(db_path: Path) -> int:
    return library.save_report(
        build_report(), "DTCR_Matching_Report_28RU_X1_vs_X2.xlsx",
        program="RU", model_year="2028", phase="X2", db_path=db_path,
    )


# ---------------------------------------------------------------------------
# Scope
# ---------------------------------------------------------------------------

def test_scope_is_read_from_the_filename() -> None:
    scope = library.parse_scope_from_filename(
        "DTCR_Matching_Report_28RU_X1_vs_X2_H 2.xlsx"
    )
    # A phase compare is filed under the later phase, as a SECR is scoped to
    # its NEW DEF.
    assert (scope.program, scope.model_year, scope.phase) == ("RU", "28", "X2")


def test_a_single_phase_filename_keeps_that_phase() -> None:
    scope = library.parse_scope_from_filename("DTCR_Matching_Report_28RU_X1.xlsx")
    assert scope.phase == "X1"


def test_an_unreadable_filename_yields_an_empty_scope() -> None:
    scope = library.parse_scope_from_filename("report final v2.xlsx")
    assert not scope.is_complete


def test_the_model_year_is_stored_two_digits(db_path: Path) -> None:
    """2028 and 28 are the same year; two spellings must not be two scopes."""
    library.save_report(
        build_report(), "r.xlsx", program="ru", model_year="2028", phase="x2",
        db_path=db_path,
    )
    for year in ("28", "2028"):
        found = library.find_report_for_scope("RU", year, "X2", db_path=db_path)
        assert found is not None, f"not findable by {year}"
    assert library.find_report_for_scope(
        "ru", "28", "x2", db_path=db_path
    ) is not None


def test_an_incomplete_scope_is_refused(db_path: Path) -> None:
    with pytest.raises(SpliceInputError, match="Program, Model Year and Phase"):
        library.save_report(
            build_report(), "r.xlsx", program="RU", model_year="", phase="X2",
            db_path=db_path,
        )


# ---------------------------------------------------------------------------
# Storing
# ---------------------------------------------------------------------------

def test_re_uploading_a_scope_replaces_it(db_path: Path, filed: int) -> None:
    """Two reports disagreeing about one scope is a contradiction, not history."""
    library.save_report(
        build_report(), "corrected.xlsx", program="RU", model_year="28",
        phase="X2", db_path=db_path,
    )
    reports = library.list_reports(db_path=db_path)
    assert len(reports) == 1
    assert reports[0]["filename"] == "corrected.xlsx"


def test_different_scopes_coexist(db_path: Path, filed: int) -> None:
    library.save_report(
        build_report(), "x1.xlsx", program="RU", model_year="28", phase="X1",
        db_path=db_path,
    )
    library.save_report(
        build_report(), "ws.xlsx", program="WS", model_year="28", phase="X2",
        db_path=db_path,
    )
    assert len(library.list_reports(db_path=db_path)) == 3


def test_the_original_workbook_is_kept_for_enrichment(
    db_path: Path, filed: int
) -> None:
    payload = library.report_bytes(filed, db_path=db_path)
    assert payload and payload[:2] == b"PK"      # a real xlsx, not a summary
    assert library.read_report(payload).shape[0] == 5


def test_a_file_that_is_not_a_report_is_refused(db_path: Path) -> None:
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "not a DTCR report"
    buffer = io.BytesIO()
    workbook.save(buffer)
    with pytest.raises(SpliceInputError, match="DTCR#"):
        library.save_report(
            buffer.getvalue(), "wrong.xlsx", program="RU", model_year="28",
            phase="X2", db_path=db_path,
        )


def test_deleting_a_report_removes_its_rows(db_path: Path, filed: int) -> None:
    assert library.delete_report(filed, db_path=db_path)
    assert library.list_reports(db_path=db_path) == []
    assert library.report_rows(filed, db_path=db_path) == []


# ---------------------------------------------------------------------------
# Harness families
# ---------------------------------------------------------------------------

def test_split_families_handles_multi_value_cells() -> None:
    assert library.split_families("BODY_RIGHT, BODY_LEFT") == [
        "BODY_RIGHT", "BODY_LEFT",
    ]
    assert library.split_families("Body_Left") == ["BODY_LEFT"]
    assert library.split_families("IP") == ["IP"]
    assert library.split_families("") == []
    assert library.split_families(None) == []
    assert library.split_families("IP, IP") == ["IP"]     # deduped


def test_a_dtcr_naming_several_families_counts_in_each(
    db_path: Path, filed: int
) -> None:
    stats = library.report_statistics(filed, db_path=db_path)
    counts = {row["harness_family"]: row["count"] for row in stats.by_harness_family}

    # 50888 names BODY_RIGHT and BODY_LEFT; 50752 names Body_Left.
    assert counts["BODY_LEFT"] == 2
    assert counts["BODY_RIGHT"] == 1
    assert stats.multi_family == 2          # 50888 and 50889


def test_case_variants_are_one_family(db_path: Path) -> None:
    """"Body_Left" and "BODY_LEFT" must not become two bars."""
    stats_source = build_report([
        {"DTCR#": "1", "Status": "Complete", "Match Method": "Manual",
         "Harness Family": "BODY_LEFT"},
        {"DTCR#": "2", "Status": "Complete", "Match Method": "Manual",
         "Harness Family": "Body_Left"},
    ])
    report_id = library.save_report(
        stats_source, "r.xlsx", program="RU", model_year="28", phase="X2",
        db_path=db_path,
    )
    stats = library.report_statistics(report_id, db_path=db_path)
    assert stats.by_harness_family == [{"harness_family": "BODY_LEFT", "count": 2}]


# ---------------------------------------------------------------------------
# Statistics
# ---------------------------------------------------------------------------

def test_unmatched_means_no_harness_family_not_no_match_method(
    db_path: Path, filed: int
) -> None:
    """50889 is reported "No Match" but has families — it is not a gap.

    Counting it as unmatched sends an engineer looking for a problem that is
    already solved.
    """
    stats = library.report_statistics(filed, db_path=db_path)

    assert stats.unmatched == 1
    assert [row["dtcr_number"] for row in stats.unmatched_rows] == ["50672"]
    # ...while the method breakdown still reports both as "No Match".
    methods = {row["match_method"]: row["count"] for row in stats.by_match_method}
    assert methods["No Match"] == 2


def test_the_headline_numbers(db_path: Path, filed: int) -> None:
    stats = library.report_statistics(filed, db_path=db_path)

    assert stats.total == 5
    assert stats.with_cnum == 3
    assert stats.with_harness == 4
    assert stats.with_bulletin == 2
    assert stats.matched == 4
    assert round(stats.match_rate) == 80


def test_status_and_bulletin_breakdowns(db_path: Path, filed: int) -> None:
    stats = library.report_statistics(filed, db_path=db_path)

    assert {row["status"]: row["count"] for row in stats.by_status} == {
        "Complete": 4, "Rejected": 1,
    }
    assert stats.by_bulletin == [{"bulletin": "320880-02", "count": 2}]


def test_statistics_for_a_missing_report_are_empty(db_path: Path) -> None:
    stats = library.report_statistics(999, db_path=db_path)
    assert stats.total == 0 and stats.by_harness_family == []
    assert stats.match_rate == 0.0          # no division by zero
