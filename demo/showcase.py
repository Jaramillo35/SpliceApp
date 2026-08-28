"""Generate the showcase dataset: one invented programme, every section covered.

Everything here is fabricated. The programme (2030 QX), the sales codes (QA1,
QB2, ...), the part numbers (99xxxxxx) and the circuit names (QK1xx) are chosen
so they cannot collide with a real Stellantis programme, and no customer data
of any kind is used or derived from.

The files cross-reference each other on purpose: the same eight harness
families, the same sales codes and the same part numbers run through the DTx,
the complexity files, the build spec and the circuit summary. That coherence is
what makes a demo believable — and it is what lets one section's output be
shown as another's input.

Defects are planted deliberately, so each section has something to *find*
rather than a clean sheet that proves nothing. They are listed in DEMO_NOTES
and in demo/README.md.

Run:  python -m demo.showcase [--out demo/showcase]
"""

from __future__ import annotations

import argparse
import io
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

from openpyxl import Workbook

# --------------------------------------------------------------------------
# The invented programme
# --------------------------------------------------------------------------

YEAR, VEHICLE, PHASE = "2030", "QX", "V1_A"
PROGRAM = f"{YEAR}{VEHICLE}"          # 2030QX — the DTx title block wording
TAG = f"{YEAR[-2:]}{VEHICLE}"         # 30QX — the filename/tag wording

#: Invented sales codes with a plain-language meaning, so the demo can be
#: narrated ("this circuit only exists on cars with the panoramic roof").
SALES_CODES: Dict[str, str] = {
    "QA1": "Panoramic roof",
    "QA2": "Fixed roof",
    "QB1": "Premium audio",
    "QB2": "Base audio",
    "QC1": "Heated seats",
    "QD1": "Trailer tow",
    "QE1": "360 camera",
    "QF1": "Power liftgate",
    "QZ9": "Night vision (in the DTx only — never tracked by a complexity file)",
}

FAMILIES = [
    "IP", "DASH", "BODY_LEFT", "BODY_RIGHT",
    "CONSOLE_CENTER", "DOOR_FRONT_LEFT", "LIFTGATE", "SEAT_2ND_ROW_LEFT",
]

#: def id per family — the id that lives in the complexity file and is what
#: every engine matches on (never the filename).
DEF_IDS = {
    "IP": "70101", "DASH": "70102", "BODY_LEFT": "70103", "BODY_RIGHT": "70104",
    "CONSOLE_CENTER": "70105", "DOOR_FRONT_LEFT": "70106",
    "LIFTGATE": "70107", "SEAT_2ND_ROW_LEFT": "70108",
}


@dataclass(frozen=True)
class Build:
    pn: str
    codes: frozenset


@dataclass(frozen=True)
class Circuit:
    circuit: str
    cnum: str
    pin: str
    condition: str          # "" = unconditional
    function: str
    gauge: str = "0.35"
    color: str = "BN/GN"


#: Builds per family. Codes are what that part number carries.
BUILDS: Dict[str, List[Build]] = {
    "IP": [
        Build("99000101AA", frozenset({"QA1", "QB1", "QE1"})),
        Build("99000102AA", frozenset({"QA1", "QB2"})),
        Build("99000103AA", frozenset({"QA2", "QB1", "QE1"})),
        Build("99000104AA", frozenset({"QA2", "QB2"})),
    ],
    "DASH": [
        Build("99000201AA", frozenset({"QB1", "QD1"})),
        Build("99000202AA", frozenset({"QB1"})),
        Build("99000203AA", frozenset({"QB2"})),
    ],
    "BODY_LEFT": [
        Build("99000301AA", frozenset({"QA1", "QD1", "QF1"})),
        Build("99000302AA", frozenset({"QA1", "QF1"})),
        Build("99000303AA", frozenset({"QA2", "QD1"})),
        Build("99000304AA", frozenset({"QA2"})),
    ],
    "BODY_RIGHT": [
        Build("99000401AA", frozenset({"QA1", "QF1"})),
        Build("99000402AA", frozenset({"QA2"})),
    ],
    "CONSOLE_CENTER": [
        Build("99000501AA", frozenset({"QB1", "QC1"})),
        Build("99000502AA", frozenset({"QB2"})),
    ],
    "DOOR_FRONT_LEFT": [
        Build("99000601AA", frozenset({"QB1", "QE1"})),
        Build("99000602AA", frozenset({"QB2"})),
    ],
    "LIFTGATE": [
        Build("99000701AA", frozenset({"QF1"})),
        Build("99000702AA", frozenset()),
    ],
    "SEAT_2ND_ROW_LEFT": [
        Build("99000801AA", frozenset({"QC1"})),
        Build("99000802AA", frozenset()),
    ],
}

#: Codes each family's complexity file TRACKS. QZ9 is deliberately absent
#: everywhere, so the DTx conditions on a code no file knows — the code-gap
#: demo. DOOR_FRONT_LEFT also omits QE1 for a second, smaller gap.
TRACKED: Dict[str, List[str]] = {
    "IP": ["QA1", "QA2", "QB1", "QB2", "QE1"],
    "DASH": ["QB1", "QB2", "QD1"],
    "BODY_LEFT": ["QA1", "QA2", "QD1", "QF1"],
    "BODY_RIGHT": ["QA1", "QA2", "QF1"],
    "CONSOLE_CENTER": ["QB1", "QB2", "QC1"],
    "DOOR_FRONT_LEFT": ["QB1", "QB2"],
    "LIFTGATE": ["QF1"],
    "SEAT_2ND_ROW_LEFT": ["QC1"],
}

#: The DTx circuit table. Conditions use the engine's grammar: / = OR,
#: & = AND, - = NOT.
CIRCUITS: Dict[str, List[Circuit]] = {
    "IP": [
        Circuit("QK101", "CQ101", "1", "", "IP - POWER FEED"),
        Circuit("QK102", "CQ101", "2", "", "IP - GROUND"),
        Circuit("QK103", "CQ101", "3", "QA1", "ROOF - SUNSHADE MOTOR"),
        Circuit("QK104", "CQ102", "1", "QB1", "AUDIO - AMPLIFIER FEED"),
        Circuit("QK105", "CQ102", "2", "QB1/QB2", "AUDIO - HEAD UNIT"),
        Circuit("QK106", "CQ103", "1", "QE1", "CAMERA - FRONT VIDEO"),
        # planted: mutually exclusive roof codes -> no IP build satisfies it
        Circuit("QK107", "CQ103", "2", "QA1&QA2", "ROOF - DUAL PANEL SENSE"),
        # planted: conditions on a code no complexity file tracks
        Circuit("QK108", "CQ104", "1", "QZ9", "NIGHT VISION - CAMERA FEED"),
        # planted: MALFORMED — a NOT with no connector before it. Read as
        # written it is false for every vehicle, so the circuit would look
        # "never built" until the integrity check repairs it to QB1&-QA1.
        Circuit("QK109", "CQ104", "2", "QB1-QA1", "AUDIO - ROOFLESS TRIM"),
    ],
    "DASH": [
        Circuit("QK201", "CQ201", "1", "", "DASH - IGNITION FEED"),
        Circuit("QK202", "CQ201", "2", "QD1", "TRAILER - BRAKE CONTROL"),
        Circuit("QK203", "CQ202", "1", "QB1", "AUDIO - SUBWOOFER"),
        Circuit("QK204", "CQ202", "2", "-QD1", "DASH - BLANK PLUG SENSE"),
    ],
    "BODY_LEFT": [
        Circuit("QK301", "CQ301", "1", "", "BODY - LEFT POWER"),
        Circuit("QK302", "CQ301", "2", "QF1", "LIFTGATE - RELEASE SIGNAL"),
        Circuit("QK303", "CQ302", "1", "QD1", "TRAILER - LEFT LAMP"),
        Circuit("QK304", "CQ302", "2", "QA1/QA2", "ROOF - DRAIN HEATER"),
    ],
    "BODY_RIGHT": [
        Circuit("QK401", "CQ401", "1", "", "BODY - RIGHT POWER"),
        Circuit("QK402", "CQ401", "2", "QF1", "LIFTGATE - RIGHT LATCH"),
        Circuit("QK403", "CQ402", "1", "QD1", "TRAILER - RIGHT LAMP"),
    ],
    "CONSOLE_CENTER": [
        Circuit("QK501", "CQ501", "1", "", "CONSOLE - ACCESSORY FEED"),
        Circuit("QK502", "CQ501", "2", "QC1", "SEAT - HEATER SWITCH"),
        Circuit("QK503", "CQ502", "1", "QB1", "AUDIO - CONSOLE SPEAKER"),
    ],
    "DOOR_FRONT_LEFT": [
        Circuit("QK601", "CQ601", "1", "", "DOOR - WINDOW MOTOR"),
        Circuit("QK602", "CQ601", "2", "QB1", "AUDIO - DOOR SPEAKER"),
        # planted: QE1 is real but this family's complexity does not track it
        Circuit("QK603", "CQ602", "1", "QE1", "CAMERA - MIRROR VIDEO"),
    ],
    "LIFTGATE": [
        Circuit("QK701", "CQ701", "1", "", "LIFTGATE - LAMP FEED"),
        Circuit("QK702", "CQ701", "2", "QF1", "LIFTGATE - POWER STRUT"),
    ],
    "SEAT_2ND_ROW_LEFT": [
        Circuit("QK801", "CQ801", "1", "", "SEAT - POSITION SENSE"),
        Circuit("QK802", "CQ801", "2", "QC1", "SEAT - HEATER ELEMENT"),
    ],
    # planted: a DTx family with NO complexity file at all, so the mapping
    # workbench has an unconnected row to demonstrate
    "HEADLINER": [
        Circuit("QK901", "CQ901", "1", "", "HEADLINER - DOME LAMP"),
        Circuit("QK902", "CQ901", "2", "QA1", "ROOF - SHADE SWITCH"),
    ],
}

#: VINs for the build spec, each with the codes that vehicle was ordered with.
VINS: List[tuple] = [
    ("QX30000001", {"QA1", "QB1", "QE1", "QF1"}),
    ("QX30000002", {"QA2", "QB2"}),
    ("QX30000003", {"QA1", "QB2", "QF1"}),
    ("QX30000004", {"QA2", "QB1", "QD1", "QE1"}),
    ("QX30000005", {"QA1", "QB1", "QC1", "QD1", "QE1", "QF1"}),
    ("QX30000006", {"QA2", "QB2", "QC1"}),
    # planted: needs a roof code AND trailer tow on a family whose parts
    # never combine them -> a VBOM review case
    ("QX30000007", {"QA1", "QB1", "QD1", "QE1"}),
    ("QX30000008", {"QA2", "QB1", "QF1"}),
]

DEMO_NOTES = [
    ("Circuit Applicability", "QK107 (IP) is conditioned on QA1&QA2 — the two "
     "roof codes are mutually exclusive, so NO build carries it: a 'never "
     "built' finding."),
    ("Circuit Applicability", "QK108 (IP) and QK603 (DOOR_FRONT_LEFT) depend "
     "on codes their complexity file does not track (QZ9, QE1) — the "
     "Sales-code gaps tab."),
    ("Circuit Applicability", "QK109 (IP) carries the malformed expression "
     "'QB1-QA1' — the Sales-code integrity step catches it before analysis and "
     "suggests 'QB1&-QA1'."),
    ("Circuit Applicability", "HEADLINER appears in the DTx with no complexity "
     "file — a red dotted row for the mapping demo."),
    ("Circuit Applicability", "The DOOR_FRONT_LEFT file is named "
     "DOOR_FRONT_LEFT_MAIN so it does NOT auto-match — it appears as a "
     "candidate to drag or click."),
    ("VBOM Risk Matrix", "8 VINs produce 13 review cases in two flavours — "
     "'no complete PN covers every required sales code' and 'N/A conflicts "
     "with an available base/default PN'."),
    ("Harness Complexity", "The master carries a C/O carryover row, a DELETE "
     "row, a combined expression (QB1+(QA1/QA2)) and an equality (QA1=QA2)."),
    ("HRN Chart Builder", "One circuit uses supplier prefix 'ZQ', which is not "
     "in the shipped supplier list — it triggers the update ticket."),
    ("Circuit Health", "Inline X350 <-> Y350 cavity 2 has a wire on BODY_LEFT "
     "and nothing opposite on LIFTGATE — a Blocker, plus 2 auto-cleared."),
    ("DTx Compare", "The OLD export omits QK106 and QK702, so the comparison "
     "reports them as added."),
]


# --------------------------------------------------------------------------
# Writers
# --------------------------------------------------------------------------

def _save(wb: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def write_dtx(path: Path, *, families: Sequence[str] | None = None,
              drop_circuits: Iterable[str] = ()) -> Path:
    """A Detailed DTx Circuits Report: title block on rows 1-4, header on 6."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed DTx Circuits Report"
    ws.append(["Detailed DTx Circuits Report"])
    ws.append([f"Vehicle Program - {PROGRAM}  "])
    ws.append([f"Build Phase - {PHASE}"])
    ws.append(["Report Date: Jan-15-2030 08:00"])
    ws.append([])
    ws.append(["Device Control Number", "Device Name", "CNUM", "Connector PN",
               "Harness Family", "Pin Number", "Circuit Name",
               "Circuit Function", "Wire Gauge", "Color", "Sales Code"])
    skip = set(drop_circuits)
    for family in (families or CIRCUITS):
        for c in CIRCUITS[family]:
            if c.circuit in skip:
                continue
            ws.append(["D" + DEF_IDS.get(family, "70999"), f"{family}_MODULE",
                       c.cnum, "8" + DEF_IDS.get(family, "70999")[1:] + "0",
                       family, c.pin, c.circuit, c.function, c.gauge, c.color,
                       c.condition or None])
    return _save(wb, path)


def write_dtx_flat(path: Path) -> Path:
    """A flat DTx sales-code export: header on row 1.

    The harness-complexity workbench reads DTx exports as ordinary tables to
    learn which row-9 tokens are real sales codes, so it needs this shape
    rather than the title-block report the circuit tools consume.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "DTx"
    ws.append(["Harness Family", "Circuit Name", "CNUM", "Sales Code"])
    for family, circuits in CIRCUITS.items():
        for c in circuits:
            ws.append([family, c.circuit, c.cnum, c.condition or None])
    return _save(wb, path)


def write_individual_complexity(path: Path, family: str, *,
                                harness_name: str | None = None) -> Path:
    """One individual harness complexity file (Complexity + Harness PN)."""
    codes = TRACKED[family]
    wb = Workbook()
    ws = wb.active
    ws.title = "Complexity"
    ws.append([f"ID={DEF_IDS[family]}", *codes])
    for build in BUILDS[family]:
        ws.append([build.pn, *["X" if c in build.codes else "" for c in codes]])

    hp = wb.create_sheet("Harness PN")
    hp.append(["Previous P/N", "New P/N", "Symbol"])
    for build in BUILDS[family]:
        hp.append(["", build.pn, ""])
    hp.append([])
    for label, value in (("year:", YEAR), ("vehicle:", VEHICLE),
                         ("phase:", PHASE),
                         ("harness:", harness_name or family),
                         ("id:", DEF_IDS[family])):
        hp.append([label, value])
    return _save(wb, path)


def write_build_spec(path: Path) -> Path:
    """A BuildSpec in the layout the VBOM engine reads.

    Row 6 carries MVON, row 7 the VIN, column A the sales codes, and an ``X``
    marks a code ordered on that vehicle. (One-based rows: the parser indexes
    rows 5 and 6 zero-based and starts data at row 8.)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "BuildSpec"
    ws.cell(1, 1, f"Build Spec {YEAR} {VEHICLE} — phase {PHASE}")
    ws.cell(6, 1, "MVON")
    ws.cell(7, 1, "VIN")
    for column, (vin, _codes) in enumerate(VINS, start=2):
        ws.cell(6, column, f"M{vin[-6:]}")
        ws.cell(7, column, vin)
    for offset, code in enumerate(c for c in SALES_CODES if c != "QZ9"):
        row = 8 + offset
        ws.cell(row, 1, code)
        for column, (_vin, codes) in enumerate(VINS, start=2):
            if code in codes:
                ws.cell(row, column, "X")
    return _save(wb, path)


def write_master_complexity(path: Path, *, include_new_code: bool = True) -> Path:
    """A Master Complexity workbook: one sheet per family, sales codes on row 9.

    Row 7 carries the feature description, row 9 the codes, and the part rows
    start at row 10 under a 'Current' column — the layout the workbench reads.
    """
    wb = Workbook()
    first = True
    for family in FAMILIES:
        ws = wb.active if first else wb.create_sheet()
        ws.title = family[:31]
        first = False
        ws.cell(1, 1, f"Vehicle Program: {YEAR} {VEHICLE}")
        ws.cell(2, 1, f"Build Phase: {PHASE}")
        ws.cell(3, 1, f"Harness: {family}")

        codes = list(TRACKED[family])
        if not include_new_code and "QF1" in codes:
            codes.remove("QF1")          # the OLD master lacks QF1 -> a delta

        ws.cell(9, 3, "Made from")
        ws.cell(9, 4, "P1")
        ws.cell(9, 5, "Current")
        col = 6
        for code in codes:
            ws.cell(7, col, SALES_CODES[code])
            ws.cell(9, col, code)
            col += 1
        # a combined expression the SE must decide on, and an equality
        if family == "IP":
            ws.cell(7, col, "Audio with either roof")
            ws.cell(9, col, "QB1+(QA1/QA2)")
            combined_col = col
            col += 1
            ws.cell(7, col, "Roof equivalence")
            ws.cell(9, col, "QA1=QA2")
            equality_col = col
            col += 1
        else:
            combined_col = equality_col = None

        row = 10
        for index, build in enumerate(BUILDS[family]):
            ws.cell(row, 1, chr(65 + index))            # variant symbol
            ws.cell(row, 4, build.pn)                    # phase P1 value
            ws.cell(row, 5, build.pn)                    # Current
            for offset, code in enumerate(codes):
                if code in build.codes:
                    ws.cell(row, 6 + offset, "X")
            if combined_col and index == 0:
                ws.cell(row, combined_col, "X")
            if equality_col and index == 1:
                ws.cell(row, equality_col, "G")
            row += 1
        if family == "IP":
            # a carryover row and a deleted row, so the workbench shows an
            # Inferred value and an excluded one
            ws.cell(row, 1, "E")
            ws.cell(row, 4, "99000105AA")
            ws.cell(row, 5, "C/O")
            row += 1
            ws.cell(row, 1, "F")
            ws.cell(row, 5, "DELETE P/N")
    return _save(wb, path)


def write_crossref(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "CrossRef"
    ws.append(["Harness Family", "Complexity File", "DTx Family Name"])
    for family in FAMILIES:
        ws.append([family.replace("_", " ").title(), family, family])
    return _save(wb, path)


def write_splice_input(path: Path) -> Path:
    """The Splice Generation workbook: Complexity + OptionPerCkt sheets."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Complexity"
    codes = [c for c in SALES_CODES if c != "QZ9"]
    ws.append(["Harness PN", *codes])
    for family in ("IP", "DASH", "BODY_LEFT"):
        for build in BUILDS[family]:
            ws.append([build.pn,
                       *["X" if c in build.codes else "" for c in codes]])

    op = wb.create_sheet("OptionPerCkt")
    op.append(["CNUM", "Pin", "Circuit", "Sales Code"])
    for family in ("IP", "DASH", "BODY_LEFT"):
        for c in CIRCUITS[family]:
            op.append([c.cnum, c.pin, c.circuit, c.condition or None])
    return _save(wb, path)


def write_circuit_summary(path: Path) -> Path:
    """A Circuit Summary: a block per harness, then its circuit ends.

    Column positions match splice.inline.summary (family 0, circuit 1,
    suffix 2, size 3, material 4, colour 5, CNUM 7, cavity 8, device 10,
    sales code 16, builds from 17).

    Two harnesses share an inline: BODY_LEFT carries X350 and LIFTGATE the
    mating Y350. An end only counts as an inline when its DEVICE says so, so
    those rows are named "INLINE 350". Cavity 2 is planted as a defect — the
    wire exists on BODY_LEFT and nothing sits opposite it on LIFTGATE.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Circuit Summary"
    ws.append(["Circuit Summary", f"{PROGRAM} {PHASE}"])

    def end_row(family: str, circuit: str, cnum: str, cavity: str,
                device: str, condition: str, carried) -> List[str]:
        row = [""] * 17
        row[0] = family
        row[1] = circuit
        row[3] = "0.35"
        row[4] = "TXL"
        row[5] = "BN/GN"
        row[7] = cnum
        row[8] = cavity
        row[10] = device
        row[16] = condition
        row.extend("X" if pn in carried else "" for pn in
                   (b.pn for b in BUILDS[family]))
        return row

    def block(family: str, rows: List[List[str]]) -> None:
        header = [""] * 17
        header[0] = f"{family} - {DEF_IDS[family]}"
        header[1] = "Circuit"
        for build in BUILDS[family]:
            header.append(f"X~{build.pn}")
        ws.append(header)
        for row in rows:
            ws.append(row)

    def carriers(family: str, condition: str) -> set:
        if not condition:
            return {b.pn for b in BUILDS[family]}
        wanted = {t for t in condition.replace("-", " ").replace("&", " ")
                  .replace("/", " ").replace("(", " ").replace(")", " ").split()}
        return {b.pn for b in BUILDS[family] if wanted & b.codes}

    # BODY_LEFT: normal circuits, then its side of inline 350
    bl = [end_row("BODY_LEFT", c.circuit, c.cnum, c.pin, f"{'BODY_LEFT'}_MODULE",
                  c.condition, carriers("BODY_LEFT", c.condition))
          for c in CIRCUITS["BODY_LEFT"]]
    bl.append(end_row("BODY_LEFT", "QK302", "X350", "1", "INLINE 350",
                      "QF1", carriers("BODY_LEFT", "QF1")))
    # planted defect: this wire has no counterpart on the LIFTGATE side
    bl.append(end_row("BODY_LEFT", "QK303", "X350", "2", "INLINE 350",
                      "QD1", carriers("BODY_LEFT", "QD1")))
    block("BODY_LEFT", bl)

    lg = [end_row("LIFTGATE", c.circuit, c.cnum, c.pin, "LIFTGATE_MODULE",
                  c.condition, carriers("LIFTGATE", c.condition))
          for c in CIRCUITS["LIFTGATE"]]
    lg.append(end_row("LIFTGATE", "QK302", "Y350", "1", "INLINE 350",
                      "QF1", carriers("LIFTGATE", "QF1")))
    block("LIFTGATE", lg)

    for family in ("IP", "DASH"):
        block(family, [
            end_row(family, c.circuit, c.cnum, c.pin, f"{family}_MODULE",
                    c.condition, carriers(family, c.condition))
            for c in CIRCUITS[family]])
    return _save(wb, path)


def write_hrn_triple(out_dir: Path) -> List[Path]:
    """An HRN + CSV (+ CMP) triple for the chart builder.

    The stem must read <family>_<MY><program>_..._<date> for the engine to
    pull the family, model year and programme out of it.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = f"{YEAR}{VEHICLE}1P_1234_IP_01-15-2030"
    rows = []
    for c in CIRCUITS["IP"]:
        # 'ZQ' is deliberately not in the shipped supplier list
        prefix = "ZQ" if c.circuit == "QK106" else "AA"
        rows.append([c.circuit, c.cnum, f"{prefix}-{c.cnum}", c.pin,
                     "CQ199", "AA-CQ199", "9", c.gauge, "TXL", "1.8",
                     "", "", "IP", c.color, c.condition, "", "", "", "",
                     "", "", f"{c.gauge} TXL", "", "", c.function])
    hrn = out_dir / f"{stem}.hrn"
    hrn.write_text("\n".join("|".join(str(v) for v in r) for r in rows),
                   encoding="utf-8")

    codes = [c for c in SALES_CODES if c != "QZ9"]
    csv_lines = ["CKT," + ",".join(codes)]
    for c in CIRCUITS["IP"]:
        marks = ["X" if (not c.condition or code in c.condition) else ""
                 for code in codes]
        csv_lines.append(c.circuit + "," + ",".join(marks))
    csv = out_dir / f"{stem}.csv"
    csv.write_text("\n".join(csv_lines), encoding="utf-8")

    cmp_file = out_dir / f"{stem}.cmp"
    cmp_file.write_text("\n".join(f"{c.cnum},AA,CONNECTOR {c.cnum}"
                                  for c in CIRCUITS["IP"]), encoding="utf-8")
    return [hrn, csv, cmp_file]


# --------------------------------------------------------------------------
# Assembly
# --------------------------------------------------------------------------

def build(out: Path) -> Dict[str, List[Path]]:
    """Write every showcase file. Returns {section: [paths]}.

    The output directory is cleared first: a renamed file would otherwise
    linger and the pack would ship two versions of the same input, which is
    exactly the confusion a demo cannot afford.
    """
    if out.exists():
        for path in sorted(out.rglob("*"), reverse=True):
            path.unlink() if path.is_file() else path.rmdir()
    made: Dict[str, List[Path]] = {}

    ca = out / "1_circuit_applicability"
    files = [write_dtx(ca / f"DetailedDTxCircuitsReport_{TAG}_{PHASE}.xlsx")]
    for family in FAMILIES:
        # DOOR_FRONT_LEFT is named so it will NOT auto-match, to demo the
        # candidate column and drag-and-drop
        name = "DOOR_FRONT_LEFT_MAIN" if family == "DOOR_FRONT_LEFT" else family
        files.append(write_individual_complexity(
            ca / f"2.- Harness_Complexity_{TAG}_{PHASE}_{name}_01-15-2030.xlsx",
            family, harness_name=name))
    made["Circuit Applicability"] = files

    vb = out / "2_vbom_risk_matrix"
    files = [write_build_spec(vb / f"{YEAR}_{VEHICLE}_BuildSpec_V1A.xlsx")]
    for family in FAMILIES:
        files.append(write_individual_complexity(
            vb / f"2.- Harness_Complexity_{TAG}_{PHASE}_{family}_01-15-2030.xlsx",
            family))
    made["VBOM Risk Matrix"] = files

    hc = out / "3_harness_complexity"
    made["Harness Complexity"] = [
        write_crossref(hc / "Harness_Family_CrossRef.xlsx"),
        write_master_complexity(hc / f"Master_Complexity_{TAG}_{PHASE}_NEW.xlsx"),
        write_master_complexity(hc / f"Master_Complexity_{TAG}_{PHASE}_OLD.xlsx",
                                include_new_code=False),
        write_dtx_flat(hc / f"DTx_SalesCodes_{TAG}_{PHASE}.xlsx"),
    ]

    dx = out / "4_dtx_compare"
    made["DTx Compare"] = [
        write_dtx(dx / f"DetailedDTxCircuitsReport_{TAG}_V1_OLD.xlsx",
                  drop_circuits={"QK106", "QK702"}),
        write_dtx(dx / f"DetailedDTxCircuitsReport_{TAG}_V1_NEW.xlsx"),
    ]

    sg = out / "5_splice_generation"
    made["Splice Generation"] = [
        write_splice_input(sg / f"Splice_Input_{TAG}_{PHASE}.xlsx")]

    ch = out / "6_circuit_health"
    files = [write_circuit_summary(ch / f"Circuit_Summary_{TAG}_{PHASE}.xlsx")]
    for family in ("BODY_LEFT", "LIFTGATE", "IP", "DASH"):
        files.append(write_individual_complexity(
            ch / f"2.- Harness_Complexity_{TAG}_{PHASE}_{family}_01-15-2030.xlsx",
            family))
    made["Circuit Health"] = files

    made["HRN Chart Builder"] = write_hrn_triple(out / "7_hrn_chart_builder")
    return made


def bundle(out: Path, made: Dict[str, List[Path]]) -> Path:
    """Zip the whole showcase for handing to a laptop that has no repo."""
    archive = out / f"Splice_Showcase_{TAG}.zip"
    with zipfile.ZipFile(archive, "w", zipfile.ZIP_DEFLATED) as zf:
        for paths in made.values():
            for path in paths:
                zf.write(path, arcname=str(path.relative_to(out)))
        zf.writestr("README.md", (Path(__file__).parent / "README.md").read_text()
                    if (Path(__file__).parent / "README.md").exists()
                    else "See demo/README.md in the repository.")
    return archive


def main(argv: Sequence[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", default="demo/showcase", type=Path)
    args = parser.parse_args(argv)
    made = build(args.out)
    archive = bundle(args.out, made)
    total = sum(len(v) for v in made.values())
    for section, paths in made.items():
        print(f"{section:24} {len(paths)} file(s)")
    print(f"\n{total} files under {args.out}")
    print(f"bundle: {archive}")


if __name__ == "__main__":
    main()
