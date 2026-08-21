"""SECR numbering and auto-enrichment orchestration.

These functions were previously defined inside ``app.py`` (the Streamlit layer),
even though they are pure business logic with no UI. They live here so the SECR
Management page — and any scripted/agent caller — can drive the same flow.
"""

from __future__ import annotations

import io
import re

import openpyxl
import pandas as pd

from secrdb.core.secr.enrich import (
    build_bulletin_numbers_for_secr,
    build_dtcr_numbers_for_secr,
    build_enrichment_summary,
    build_reason_for_change_for_secr,
    export_secr_enriched_output,
    find_dtcr_number_label_cell,
    find_reason_for_change_cell,
    get_secr_harness_family_from_c12,
    load_dtcr_matching_report,
    load_generated_secr_workbook,
    update_secr_bulletin_numbers,
    update_secr_dtcr_numbers,
    update_secr_reason_for_change,
)

__all__ = [
    "build_secr_number_preview",
    "extract_secr_number_inputs_from_def",
    "auto_enrich_secr",
]


def build_secr_number_preview(
    model_year: str,
    program: str,
    phase: str,
    secr_type_label: str,
    sequence: int = 1000,
) -> str:
    """Build the SECR-number preview from form values and the change type.

    Format: ``{D|M}{MY2}{PROGRAM}{PHASE}_{sequence}`` where the prefix is ``D``
    for a Design Change and ``M`` otherwise, and ``MY2`` is the last two digits
    of the model year. Returns ``""`` if the model year is too short to be valid.
    """
    my_clean = str(model_year or "").strip()
    program_clean = str(program or "").strip().upper().replace(" ", "")
    phase_clean = str(phase or "").strip().upper().replace(" ", "")

    if not my_clean or len(my_clean) < 2:
        return ""

    type_prefix = "D" if secr_type_label == "Design Change" else "M"
    my_two = my_clean[-2:]
    return f"{type_prefix}{my_two}{program_clean}{phase_clean}_{sequence}"


def extract_secr_number_inputs_from_def(def_bytes: bytes) -> tuple[str, str, str]:
    """Extract (MY, Program/Vehicle Line, Phase) from a DEF workbook identifier.

    Reads the ``DEF_DEF_Summary`` sheet and parses an identifier snippet such as
    ``DEF_New (Identifier) := 2028 RU X1_A ...``. Returns ``("", "", "")`` if the
    workbook can't be read or the identifier isn't found.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(def_bytes), data_only=True, read_only=True)
    except Exception:
        return "", "", ""

    try:
        if "DEF_DEF_Summary" not in wb.sheetnames:
            return "", "", ""

        ws = wb["DEF_DEF_Summary"]
        identifier_text = ""

        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row or 1, 120),
            min_col=1,
            max_col=min(ws.max_column or 1, 12),
            values_only=True,
        ):
            for value in row:
                if not value:
                    continue
                text = str(value)
                if "DEF_New" in text and "Identifier" in text:
                    identifier_text = text
                    break
            if identifier_text:
                break

        if not identifier_text:
            return "", "", ""

        # Match: 2028 RU X1_A
        match = re.search(r"(\d{4})\s+([A-Za-z0-9]+)\s+([A-Za-z0-9]+_[A-Za-z0-9]+)", identifier_text)
        if not match:
            return "", "", ""

        my = match.group(1)
        program = match.group(2).upper()
        phase = match.group(3).replace("_", "").upper()
        return my, program, phase
    finally:
        wb.close()


def auto_enrich_secr(
    secr_bytes: bytes,
    dtcr_matching_bytes: bytes,
    output_filename: str,
) -> tuple[bytes, dict, pd.DataFrame, pd.DataFrame, str]:
    """Enrich a freshly generated SECR from a DTCR Matching Report workbook.

    Filters the mapping to Complete/Draft rows, reads the SECR harness family
    from cell C12, fills Reason-for-Change / DTCR numbers / bulletin numbers, and
    exports the enriched workbook.

    Returns ``(enriched_bytes, export_meta, dtcr_mapping_df, summary_df,
    secr_harness_family)``.

    Raises
    ------
    ValueError
        If the mapping is empty after filtering, or the SECR C12 harness family
        cell is empty/invalid.
    """
    dtcr_mapping_df = load_dtcr_matching_report(dtcr_matching_bytes)
    dtcr_mapping_df = dtcr_mapping_df[
        dtcr_mapping_df["Status"].astype(str).str.strip().isin(["Complete", "Draft"])
    ]

    secr_wb = load_generated_secr_workbook(secr_bytes)
    secr_harness_family = get_secr_harness_family_from_c12(secr_wb)
    if dtcr_mapping_df.empty:
        raise ValueError("DTCR Matching Report is empty after filtering.")
    if not secr_harness_family:
        raise ValueError("SECR cell C12 is empty or invalid.")

    reason_text = build_reason_for_change_for_secr(secr_harness_family, dtcr_mapping_df)
    summary_df = build_enrichment_summary(dtcr_mapping_df, secr_harness_family, reason_text)

    reason_cell_info = find_reason_for_change_cell(secr_wb)
    if reason_cell_info:
        _, cell_ref = reason_cell_info
        update_secr_reason_for_change(secr_wb, cell_ref, reason_text)

    dtcr_label_info = find_dtcr_number_label_cell(secr_wb)
    if dtcr_label_info:
        _, dtcr_label_ref = dtcr_label_info
        dtcr_numbers_text = build_dtcr_numbers_for_secr(secr_harness_family, dtcr_mapping_df)
        update_secr_dtcr_numbers(secr_wb, dtcr_label_ref, dtcr_numbers_text)

    bulletin_numbers_text = build_bulletin_numbers_for_secr(secr_harness_family, dtcr_mapping_df)
    if bulletin_numbers_text:
        update_secr_bulletin_numbers(secr_wb, bulletin_numbers_text)

    enriched_bytes, export_meta = export_secr_enriched_output(
        secr_wb,
        dtcr_mapping_df,
        dtcr_mapping_df,
        summary_df,
        output_filename=output_filename,
    )
    return enriched_bytes, export_meta, dtcr_mapping_df, summary_df, secr_harness_family
