"""Identity, numbering, naming and versioning for **generated** SECRs.

These rules apply only to SECRs the application creates from a DEF→DEF compare.
Imported historical SECRs keep the identifiers and filenames they arrived with.

Four concepts are kept separate on purpose:

``SECR identity``
    ``model_year + phase + sequence number`` — e.g. ``28 / X1 / 1000``.
    ``MY28/X1/1000`` and ``MY28/X2/1000`` are two different SECRs.
``SECR number``
    The human-facing code written into the workbook: ``D28X1RU_1000``.
``SECR version``
    ``V1 → V2 → V3`` within one identity. Versions never overwrite each other.
``Generated filename``
    Derived *from* the structured fields, never parsed back out of them:
    ``SECR_IP_D28X1RU_1000_V1_05072026.xlsx``

The filename is an output, not a source of truth.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

CHANGE_TYPE_DESIGN = "Design Change"
CHANGE_TYPE_MISCELLANEOUS = "Miscellaneous Change"
CHANGE_TYPES = (CHANGE_TYPE_DESIGN, CHANGE_TYPE_MISCELLANEOUS)

#: Change type → the letter that opens the SECR number.
CHANGE_TYPE_CODES = {CHANGE_TYPE_DESIGN: "D", CHANGE_TYPE_MISCELLANEOUS: "M"}

#: The first number issued in any previously unused model-year + phase scope.
FIRST_SEQUENCE_NUMBER = 1000

#: Where a metadata value came from — recorded so a reviewer can tell an
#: extracted value from one a person chose.
PROVENANCE_EXTRACTED = "EXTRACTED"
PROVENANCE_USER_SELECTED = "USER_SELECTED"
PROVENANCE_USER_RESOLVED = "USER_RESOLVED"
PROVENANCE_GENERATED = "GENERATED"

#: Metadata fields that define a SECR's scope. A change in any of them during an
#: update normally means a new SECR is required.
SCOPE_FIELDS = ("harness_family", "model_year", "phase", "program")

_IDENTIFIER_RE = re.compile(
    r"(?P<my>\d{4})\s+(?P<program>[A-Za-z0-9]+)\s+(?P<phase>[A-Za-z0-9]+(?:_[A-Za-z0-9]+)?)"
    r"(?:\s+[\d_]+)?\s*(?P<harness>[A-Za-z0-9 _\-]*?)\s*(?:ID\s*:|$)"
)

_ILLEGAL_FILENAME_CHARS = re.compile(r"[^A-Za-z0-9_\-]+")


# ---------------------------------------------------------------------------
# Metadata
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class SecrMetadata:
    """The four values that scope a generated SECR."""

    harness_family: str = ""
    model_year: str = ""
    phase: str = ""
    program: str = ""

    @property
    def model_year_2(self) -> str:
        """Last two digits of the model year, as used in the SECR number."""
        digits = re.sub(r"\D", "", self.model_year)
        return digits[-2:] if len(digits) >= 2 else digits

    def as_dict(self) -> Dict[str, str]:
        return {
            "harness_family": self.harness_family,
            "model_year": self.model_year,
            "phase": self.phase,
            "program": self.program,
        }

    def is_complete(self) -> bool:
        return all(self.as_dict().values())


@dataclass
class ExtractedMetadata:
    """What the DEF compare workbook says about the SECR being generated.

    ``metadata`` is the scope of the SECR, taken from the **NEW** DEF — the new
    DEF is the state the SECR describes. ``old`` is kept for reporting only.
    """

    metadata: SecrMetadata
    old: SecrMetadata = field(default_factory=SecrMetadata)
    new: SecrMetadata = field(default_factory=SecrMetadata)
    sources: Dict[str, str] = field(default_factory=dict)
    old_def_source: str = ""
    new_def_source: str = ""
    notes: List[str] = field(default_factory=list)
    #: Disagreements between the two DEFs, or between the filename and the
    #: identifier. **Advisory only** — they are shown to the engineer and never
    #: block generation. See :func:`extract_metadata_from_def`.
    warnings: List[str] = field(default_factory=list)
    #: Set when the workbook itself could not be opened. This *does* block, but
    #: it is a read failure, not a metadata disagreement.
    read_error: str = ""


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _normalize_phase(value: str) -> str:
    """``X1_A`` and ``x1 a`` both normalize to ``X1A``; ``X1`` stays ``X1``."""
    return re.sub(r"[\s_]+", "", _clean(value)).upper()


def _scope_phase(value: str) -> str:
    """The phase used for numbering: the leading phase token (``X1_A`` → ``X1``).

    The DEF identifier carries a revision letter after the phase (``X1_A``); the
    SECR number and the numbering scope use the phase itself.
    """
    cleaned = _clean(value).upper()
    head = re.split(r"[\s_]+", cleaned)[0] if cleaned else ""
    return head


def _parse_identifier(text: str) -> Tuple[SecrMetadata, str]:
    """Parse ``2028 RU X1_A 05_07_26_09_25_34 IP  ID: 11430``."""
    match = _IDENTIFIER_RE.search(text)
    if not match:
        return SecrMetadata(), ""
    harness = _clean(match.group("harness"))
    return (
        SecrMetadata(
            harness_family=harness.upper(),
            model_year=match.group("my"),
            phase=_scope_phase(match.group("phase")),
            program=match.group("program").upper(),
        ),
        _clean(text),
    )


def _harness_from_def_filename(def_filename: str) -> str:
    """Harness family from the DEF filename: the part right before ``DEF``.

    Mirrors how :func:`secrdb.core.secr.generate.create_secr_bytes` reads it, so the
    two never disagree about where the value comes from.
    """
    parts = Path(def_filename).stem.split("_")
    try:
        index = parts.index("DEF")
    except ValueError:
        return ""
    return parts[index - 1].upper() if index > 0 else ""


def extract_metadata_from_def(
    def_bytes: bytes, def_filename: str = ""
) -> ExtractedMetadata:
    """Read the SECR scope out of a DEF→DEF compare workbook.

    The workbook's ``DEF_DEF_Summary`` sheet carries a ``DEF_New (Identifier)``
    and a ``DEF_Old (Identifier)`` line holding model year, program, phase and
    harness family. The SECR's scope is the **new** DEF; the old one is read so
    that a suspicious pairing can be reported.

    A phase difference between OLD and NEW is normal — a compare is usually run
    across phases (``X0_A`` vs ``X1_A``) — and is reported as a note. A
    difference in program, model year or harness family is unusual and is
    reported as a **warning**: it is surfaced to the engineer, who decides,
    because only they can tell an unusual-but-intended pairing (a carryover, a
    rebadge, a harness renamed between phases) from a genuine mistake.

    Nothing here blocks generation. The only blocking outcome is
    ``read_error`` — a workbook that cannot be opened yields no metadata at
    all, which :func:`validate_metadata` then rejects on its own terms.
    """
    old_text = new_text = ""
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(def_bytes), data_only=True, read_only=True
        )
    except Exception as exc:  # noqa: BLE001 - reported, not raised
        return ExtractedMetadata(
            metadata=SecrMetadata(),
            read_error=f"The DEF compare file could not be read: {exc}",
        )

    try:
        if "DEF_DEF_Summary" in workbook.sheetnames:
            worksheet = workbook["DEF_DEF_Summary"]
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=min(worksheet.max_row or 1, 120),
                max_col=min(worksheet.max_column or 1, 12),
                values_only=True,
            ):
                for value in row:
                    text = str(value or "")
                    if "Identifier" not in text:
                        continue
                    if not new_text and "DEF_New" in text:
                        new_text = text
                    elif not old_text and "DEF_Old" in text:
                        old_text = text
                if new_text and old_text:
                    break
    finally:
        workbook.close()

    new_meta, new_source = _parse_identifier(new_text)
    old_meta, old_source = _parse_identifier(old_text)

    sources: Dict[str, str] = {}
    notes: List[str] = []
    warnings: List[str] = []

    harness = new_meta.harness_family
    if harness:
        sources["harness_family"] = "DEF_New identifier"
    filename_harness = _harness_from_def_filename(def_filename)
    if filename_harness:
        if not harness:
            harness = filename_harness
            sources["harness_family"] = "DEF filename"
        elif filename_harness != harness:
            # The identifier wins: it is written by the compare tool, whereas a
            # filename is typed and renamed by hand.
            warnings.append(
                f"Harness family is ambiguous: the DEF filename says "
                f"'{filename_harness}' but the DEF_New identifier says "
                f"'{harness}'. The identifier is used."
            )

    metadata = SecrMetadata(
        harness_family=harness,
        model_year=new_meta.model_year,
        phase=new_meta.phase,
        program=new_meta.program,
    )
    for name in ("model_year", "phase", "program"):
        if getattr(metadata, name):
            sources.setdefault(name, "DEF_New identifier")

    if old_meta.is_complete() and new_meta.model_year:
        if old_meta.phase and old_meta.phase != new_meta.phase:
            notes.append(
                f"The compare runs across phases: OLD DEF is {old_meta.phase}, "
                f"NEW DEF is {new_meta.phase}. The SECR is scoped to the NEW "
                f"DEF phase ({new_meta.phase})."
            )
        for label, name in (
            ("Program", "program"),
            ("Model year", "model_year"),
            ("Harness family", "harness_family"),
        ):
            old_value = getattr(old_meta, name)
            new_value = getattr(new_meta, name)
            if old_value and new_value and old_value != new_value:
                warnings.append(
                    f"{label} differs between the two DEFs: OLD is '{old_value}', "
                    f"NEW is '{new_value}'. Check this is the pair you meant to "
                    f"compare — the SECR is scoped to the NEW DEF ('{new_value}')."
                )

    return ExtractedMetadata(
        metadata=metadata,
        old=old_meta,
        new=new_meta,
        sources=sources,
        old_def_source=old_source,
        new_def_source=new_source,
        notes=notes,
        warnings=warnings,
    )


def validate_metadata(
    metadata: SecrMetadata, change_type: str = ""
) -> List[str]:
    """Return the reasons this metadata cannot produce a SECR. Empty means valid.

    Nothing is guessed or defaulted: a missing value is a blocking problem the
    engineer has to resolve.
    """
    labels = {
        "harness_family": "Harness Family",
        "model_year": "Model Year",
        "phase": "Phase",
        "program": "Program",
    }
    problems = [
        f"{label} could not be determined from the DEF compare file."
        for name, label in labels.items()
        if not getattr(metadata, name)
    ]
    if metadata.model_year and not re.fullmatch(r"\d{2}|\d{4}", metadata.model_year):
        problems.append(
            f"Model Year '{metadata.model_year}' is not a 2- or 4-digit year."
        )
    if change_type and change_type not in CHANGE_TYPE_CODES:
        problems.append(
            f"Change Type must be one of {', '.join(CHANGE_TYPES)}."
        )
    return problems


@dataclass(frozen=True)
class MetadataDifference:
    """One scope field that changed between an existing SECR and new input."""

    field: str
    label: str
    existing: str
    new: str

    @property
    def changed(self) -> bool:
        return self.existing != self.new


def compare_metadata(
    existing: SecrMetadata, new: SecrMetadata
) -> List[MetadataDifference]:
    """Compare the scope fields of an existing SECR against new DEF input.

    Returns every scope field with both values, changed or not, so the UI can
    show the full table the engineer needs to make the call.
    """
    labels = {
        "harness_family": "Harness Family",
        "model_year": "Model Year",
        "phase": "Phase",
        "program": "Program",
    }
    return [
        MetadataDifference(
            field=name,
            label=label,
            existing=getattr(existing, name) or "",
            new=getattr(new, name) or "",
        )
        for name, label in labels.items()
    ]


def changed_fields(
    differences: List[MetadataDifference],
) -> List[MetadataDifference]:
    return [difference for difference in differences if difference.changed]


# ---------------------------------------------------------------------------
# Numbering and naming
# ---------------------------------------------------------------------------

def change_type_code(change_type: str) -> str:
    """``Design Change`` → ``D``, ``Miscellaneous Change`` → ``M``."""
    try:
        return CHANGE_TYPE_CODES[change_type]
    except KeyError:
        raise ValueError(
            f"Unknown change type {change_type!r}; expected one of "
            f"{', '.join(CHANGE_TYPES)}."
        ) from None


def build_secr_number(
    metadata: SecrMetadata, change_type: str, sequence_number: int
) -> str:
    """``D28X1RU_1000`` — ``{D|M}{MY2}{PHASE}{PROGRAM}_{sequence}``.

    The model year and phase are embedded, so the code stays unique even though
    the sequence restarts at 1000 in every model-year + phase scope.
    """
    if not metadata.model_year_2:
        raise ValueError("Model Year is required to build a SECR number.")
    if not metadata.phase:
        raise ValueError("Phase is required to build a SECR number.")
    if not metadata.program:
        raise ValueError("Program is required to build a SECR number.")
    return (
        f"{change_type_code(change_type)}{metadata.model_year_2}"
        f"{_normalize_phase(metadata.phase)}"
        f"{re.sub(r'[^A-Za-z0-9]', '', metadata.program).upper()}"
        f"_{int(sequence_number)}"
    )


def _filename_token(value: str) -> str:
    """Make a metadata value safe for a filename without losing information."""
    token = re.sub(r"\s+", "_", _clean(value)).upper()
    token = _ILLEGAL_FILENAME_CHARS.sub("_", token)
    return token.strip("_")


def format_generation_date(when: Optional[date] = None) -> str:
    """``MMDDYYYY`` — 7 May 2026 → ``05072026``."""
    return (when or date.today()).strftime("%m%d%Y")


def build_filename(
    metadata: SecrMetadata,
    change_type: str,
    sequence_number: int,
    version_number: int,
    when: Optional[date] = None,
    extension: str = ".xlsx",
) -> str:
    """``SECR_IP_D28X1RU_1000_V1_05072026.xlsx``.

    Built from the structured fields every time — the filename is never parsed
    back to recover them.
    """
    if version_number < 1:
        raise ValueError("Version number starts at 1.")
    harness = _filename_token(metadata.harness_family)
    if not harness:
        raise ValueError("Harness Family is required to build a SECR filename.")
    number = build_secr_number(metadata, change_type, sequence_number)
    return (
        f"SECR_{harness}_{number}_V{int(version_number)}"
        f"_{format_generation_date(when)}{extension}"
    )


@dataclass(frozen=True)
class SecrIdentity:
    """``model_year + phase + sequence number`` — what makes a SECR *that* SECR."""

    model_year: str
    phase: str
    sequence_number: int

    def __str__(self) -> str:
        return f"MY{self.model_year} / {self.phase} / {self.sequence_number}"

    @property
    def scope(self) -> Tuple[str, str]:
        """The numbering scope this identity draws its number from."""
        return (self.model_year, self.phase)


def identity_from_metadata(
    metadata: SecrMetadata, sequence_number: int
) -> SecrIdentity:
    return SecrIdentity(
        model_year=metadata.model_year_2,
        phase=_normalize_phase(metadata.phase),
        sequence_number=int(sequence_number),
    )


def scope_key(model_year: str, phase: str) -> Tuple[str, str]:
    """Normalize a numbering scope so ``2028``/``28`` and ``x1``/``X1`` agree."""
    digits = re.sub(r"\D", "", str(model_year or ""))
    return (digits[-2:] if len(digits) >= 2 else digits, _normalize_phase(phase))
