"""SECR creation engine — Streamlit-compatible (BytesIO, no file paths)."""
from __future__ import annotations

import copy
import datetime
import io
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.cell import MergedCell

TEMPLATE_PATH = Path(__file__).resolve().parent / "assets" / "SECR_TEMPLATE.xlsx"

SUMMARY_SHEET = "Summary"
DEF_DEF_SUMMARY_SHEET = "DEF_DEF_Summary"
CONNECTOR_SHEET = "Connector"
CIRCUIT_SHEET = "Circuit"


# ---------------------------------------------------------------------------
# Sheet copying helpers
# ---------------------------------------------------------------------------

def _copy_cell_style(source_cell, target_cell) -> None:
    try:
        if source_cell.has_style:
            target_cell.font = copy.copy(source_cell.font)
            target_cell.fill = copy.copy(source_cell.fill)
            target_cell.border = copy.copy(source_cell.border)
            target_cell.alignment = copy.copy(source_cell.alignment)
            target_cell.number_format = copy.copy(source_cell.number_format)
            target_cell.protection = copy.copy(source_cell.protection)
        if source_cell.hyperlink:
            target_cell._hyperlink = copy.copy(source_cell.hyperlink)
        if source_cell.comment:
            target_cell.comment = copy.copy(source_cell.comment)
    except Exception:
        pass


def _copy_sheet(source_ws, target_wb) -> None:
    target_ws = target_wb.create_sheet(title=source_ws.title)
    max_row = source_ws.max_row or 0
    max_col = source_ws.max_column or 0
    if max_row > 0 and max_col > 0:
        for row in source_ws.iter_rows():
            if row is None:
                continue
            for cell in row:
                if cell is None or isinstance(cell, MergedCell):
                    continue
                new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                _copy_cell_style(cell, new_cell)
    if hasattr(source_ws, "column_dimensions"):
        for col_letter, dim in source_ws.column_dimensions.items():
            target_ws.column_dimensions[col_letter].width = dim.width
    if hasattr(source_ws, "row_dimensions"):
        for row_idx, dim in source_ws.row_dimensions.items():
            target_ws.row_dimensions[row_idx].height = dim.height
    if hasattr(source_ws, "merged_cells") and source_ws.merged_cells:
        for merged_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_range))
    target_ws.sheet_format.defaultColWidth = source_ws.sheet_format.defaultColWidth
    target_ws.sheet_format.defaultRowHeight = source_ws.sheet_format.defaultRowHeight
    target_ws.freeze_panes = source_ws.freeze_panes


# ---------------------------------------------------------------------------
# Action-column lookup (find by header name, not by position)
# ---------------------------------------------------------------------------

def _find_action_col(ws, header_row: int = 3) -> int:
    """Return the 1-based column index whose header equals 'Action' (case-insensitive).
    Falls back to the first non-empty column in that row if not found."""
    for c in range(1, (ws.max_column or 1) + 1):
        val = ws.cell(row=header_row, column=c).value
        if val and str(val).strip().lower() == "action":
            return c
    return 1  # fallback


# ---------------------------------------------------------------------------
# DEF sheet processors — update Summary from copied DEF sheets
# ---------------------------------------------------------------------------

def _process_def_def_summary(wb_secr, ws_summary) -> None:
    if DEF_DEF_SUMMARY_SHEET not in wb_secr.sheetnames:
        return
    ws = wb_secr[DEF_DEF_SUMMARY_SHEET]

    action_col = _find_action_col(ws, header_row=3)
    # The value column was originally 3 positions after Action (col D when Action=col A)
    value_col = action_col + 3

    delete_values, chg_values, add_values = [], [], []
    for row_idx in range(4, (ws.max_row or 3) + 1):
        action = ws.cell(row=row_idx, column=action_col).value
        value = ws.cell(row=row_idx, column=value_col).value
        if value is not None:
            s = str(value)
            if action == "DELETE":
                delete_values.append(s)
            elif action == "CHG":
                chg_values.append(s)
            elif action == "ADD":
                add_values.append(s)

    ws_summary["C32"] = ", ".join(delete_values) if delete_values else ""
    ws_summary["C31"] = ", ".join(chg_values) if chg_values else ""
    ws_summary["C30"] = ", ".join(add_values) if add_values else ""


def _process_connector_sheet(wb_secr, ws_summary) -> None:
    if CONNECTOR_SHEET not in wb_secr.sheetnames:
        return
    ws = wb_secr[CONNECTOR_SHEET]

    action_col = _find_action_col(ws, header_row=3)
    connector_col = action_col + 1

    delete_vals, chg_vals, add_vals = [], [], []
    for row_idx in range(4, (ws.max_row or 3) + 1):
        action = ws.cell(row=row_idx, column=action_col).value
        connector = ws.cell(row=row_idx, column=connector_col).value
        if connector is not None:
            s = str(connector)
            if action == "DELETE":
                delete_vals.append(s)
            elif action in ("COMP CHG", "CHG"):
                chg_vals.append(s)
            elif action == "ADD":
                add_vals.append(s)

    delete_set, add_set = set(delete_vals), set(add_vals)
    common = delete_set & add_set
    if common:
        chg_vals.extend(common)
        delete_vals = list(delete_set - common)
        add_vals = list(add_set - common)

    ws_summary["C22"] = ", ".join(sorted(set(delete_vals))) if delete_vals else ""
    combined_chg_add = sorted(set(chg_vals + add_vals))
    ws_summary["C21"] = ", ".join(combined_chg_add) if combined_chg_add else ""


def _process_circuit_sheet(wb_secr, ws_summary) -> None:
    if CIRCUIT_SHEET not in wb_secr.sheetnames:
        return
    ws = wb_secr[CIRCUIT_SHEET]

    action_col = _find_action_col(ws, header_row=3)
    val_b_col = action_col + 1
    val_c_col = action_col + 2

    delete_vals, chg_vals, add_vals = [], [], []
    for row_idx in range(4, (ws.max_row or 3) + 1):
        action = ws.cell(row=row_idx, column=action_col).value
        b = ws.cell(row=row_idx, column=val_b_col).value
        c = ws.cell(row=row_idx, column=val_c_col).value
        combined = (
            (str(b) if b is not None else "") + (str(c) if c is not None else "")
        ).strip()
        if combined:
            if action == "DELETE":
                delete_vals.append(combined)
            elif action in ("CHG", "COMP CHG", "COMP CHG "):
                chg_vals.append(combined)
            elif action == "ADD":
                add_vals.append(combined)

    delete_set, add_set = set(delete_vals), set(add_vals)
    common = delete_set & add_set
    if common:
        chg_vals.extend(common)
        delete_vals = list(delete_set - common)
        add_vals = list(add_set - common)

    ws_summary["C27"] = ", ".join(sorted(set(delete_vals))) if delete_vals else ""
    ws_summary["C25"] = ", ".join(sorted(set(add_vals))) if add_vals else ""
    ws_summary["C26"] = ", ".join(sorted(set(chg_vals))) if chg_vals else ""


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_header_map(ws, header_row: int = 3) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for col_idx in range(1, (ws.max_column or 0) + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        key = _normalize_text(val).upper()
        if key:
            header_map[key] = col_idx
    return header_map


def _find_first_col(header_map: Dict[str, int], names: List[str]) -> Optional[int]:
    for name in names:
        idx = header_map.get(name.upper())
        if idx:
            return idx
    return None


def _row_key(ws, row_idx: int, cols: List[int]) -> Tuple[str, ...]:
    return tuple(_normalize_text(ws.cell(row=row_idx, column=c).value) for c in cols)


def _merge_comments_from_old(new_ws, old_ws) -> None:
    new_header = _find_header_map(new_ws)
    old_header = _find_header_map(old_ws)

    new_comment_col = _find_first_col(new_header, ["SE COMMENTS", "COMMENTS"])
    old_comment_col = _find_first_col(old_header, ["SE COMMENTS", "COMMENTS"])
    if not new_comment_col or not old_comment_col:
        return

    excluded = {"SE COMMENTS", "COMMENTS"}
    shared = [h for h in new_header if h in old_header and h not in excluded]
    if not shared:
        return

    old_key_cols = [old_header[h] for h in shared]
    new_key_cols = [new_header[h] for h in shared]

    old_comments: Dict[Tuple[str, ...], str] = {}
    for row_idx in range(4, (old_ws.max_row or 3) + 1):
        comment = _normalize_text(old_ws.cell(row=row_idx, column=old_comment_col).value)
        if not comment:
            continue
        key = _row_key(old_ws, row_idx, old_key_cols)
        if key and key not in old_comments:
            old_comments[key] = comment

    for row_idx in range(4, (new_ws.max_row or 3) + 1):
        existing = _normalize_text(new_ws.cell(row=row_idx, column=new_comment_col).value)
        if existing:
            continue
        key = _row_key(new_ws, row_idx, new_key_cols)
        old_comment = old_comments.get(key)
        if old_comment:
            new_ws.cell(row=row_idx, column=new_comment_col, value=old_comment)


def _copy_summary_values(new_summary, old_summary, excluded_cells: set[str]) -> None:
    for r in range(1, (old_summary.max_row or 0) + 1):
        for c in range(1, (old_summary.max_column or 0) + 1):
            coord = old_summary.cell(row=r, column=c).coordinate
            if coord in excluded_cells:
                continue
            new_summary.cell(row=r, column=c, value=old_summary.cell(row=r, column=c).value)


def _populate_update_table_row(ws_summary, version: str, reissue_date: str, subject: str) -> None:
    row_idx = 34
    max_row = ws_summary.max_row or 34
    while row_idx <= max_row and ws_summary.cell(row=row_idx, column=2).value:
        row_idx += 1

    ws_summary.cell(row=row_idx, column=2, value=reissue_date)
    ws_summary.cell(row=row_idx, column=3, value=version)
    ws_summary.cell(row=row_idx, column=4, value=subject or "SECR update")


def _build_secr_code(
    c10_value: str,
    c11_value: str,
    code1: str,
    code2: str,
    secr_change_type: str,
    secr_sequence: int,
    secr_model_year: str = "",
    secr_program: str = "",
    secr_phase: str = "",
) -> str:
    my_for_code = str(secr_model_year).strip() or c10_value
    program_for_code = str(secr_program).strip() or c11_value
    phase_for_code = str(secr_phase).strip() or f"{code1}{code2}"

    my_two = my_for_code[-2:] if len(my_for_code) >= 2 else my_for_code
    type_prefix = "D" if str(secr_change_type).strip().lower().startswith("design") else "M"
    phase = phase_for_code.replace("_", "").replace(" ", "").upper()
    program_clean = program_for_code.replace(" ", "").upper()
    return f"{type_prefix}{my_two}{program_clean}{phase}_{int(secr_sequence)}"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def create_secr_bytes(
    def_bytes: bytes,
    def_filename: str,
    reason_for_change: str,
    secr_author: str,
    design_release_engineer: str,
    change_requested_by: str,
    original_issue_date: str,
    reissue_date: str,
    version: str,
    phase_implemented: str,
    pull_ahead: str,
    secr_change_type: str = "Miscellaneous",
    secr_sequence: int = 1000,
    secr_model_year: str = "",
    secr_program: str = "",
    secr_phase: str = "",
) -> Tuple[bytes, Dict[str, Any]]:
    """Create a SECR workbook from DEF compare bytes.

    Parameters mirror the original GUI dialog fields. Returns (excel_bytes, metadata).
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"SECR template not found: {TEMPLATE_PATH}")

    # Parse metadata from DEF filename
    stem = Path(def_filename).stem
    parts = stem.split("_")
    if len(parts) < 5:
        raise ValueError(
            "DEF filename must have at least 5 underscore-separated parts "
            "(e.g. 2027_RU_X2_A_vs_2026_RU_X2_A_IP_DEF_DEF_Compare_...)."
        )

    my_full = parts[0]
    vehicle_line = parts[1]
    code1 = parts[2]
    code2 = parts[3]

    idx_def = next((i for i, p in enumerate(parts) if p == "DEF"), None)
    if idx_def is None or idx_def == 0:
        raise ValueError("Could not find 'DEF' segment in the filename.")

    pre_def_string = parts[idx_def - 1]
    c10_value = my_full
    c11_value = vehicle_line
    f10_value = f"{code1}_{code2}"
    c12_value = pre_def_string

    m_code = _build_secr_code(
        c10_value,
        c11_value,
        code1,
        code2,
        secr_change_type,
        secr_sequence,
        secr_model_year=secr_model_year,
        secr_program=secr_program,
        secr_phase=secr_phase,
    )

    wb_template = openpyxl.load_workbook(str(TEMPLATE_PATH))
    wb_def = openpyxl.load_workbook(io.BytesIO(def_bytes), data_only=False)

    if SUMMARY_SHEET not in wb_template.sheetnames:
        raise ValueError("SECR template has no 'Summary' sheet.")

    ws_summary = wb_template[SUMMARY_SHEET]

    # Fill Summary from filename metadata
    ws_summary["C10"] = int(c10_value) if c10_value.isdigit() else c10_value
    ws_summary["C11"] = c11_value
    ws_summary["F10"] = f10_value
    ws_summary["C12"] = c12_value
    ws_summary["I2"] = m_code
    ws_summary["C8"] = m_code

    # Copy all DEF sheets into the template workbook
    for ws in wb_def.worksheets:
        _copy_sheet(ws, wb_template)
    wb_def.close()

    # Reorder: Summary first
    summary_ws = next(
        (ws for ws in wb_template.worksheets if ws.title == SUMMARY_SHEET), None
    )
    if summary_ws:
        others = [ws for ws in wb_template.worksheets if ws is not summary_ws]
        wb_template._sheets = [summary_ws] + others

    # Fill user-provided details into Summary
    ws_summary["C7"] = reason_for_change
    ws_summary["I10"] = secr_author
    ws_summary["I11"] = design_release_engineer
    ws_summary["I12"] = change_requested_by
    ws_summary["I3"] = version
    ws_summary["F11"] = phase_implemented
    ws_summary["F12"] = pull_ahead
    if original_issue_date:
        ws_summary["I4"] = original_issue_date
    if reissue_date:
        ws_summary["I5"] = reissue_date

    # Process copied DEF sheets → populate Summary circuit/connector blocks
    _process_def_def_summary(wb_template, ws_summary)
    _process_connector_sheet(wb_template, ws_summary)
    _process_circuit_sheet(wb_template, ws_summary)

    # Unprotect all sheets so the user can edit the output
    for ws in wb_template.worksheets:
        ws.protection.disable()
    if hasattr(wb_template, "security") and wb_template.security is not None:
        wb_template.security.lockStructure = False
        wb_template.security.lockWindows = False
        wb_template.security.lockRevision = False

    secr_filename = f"{m_code}.xlsx"

    buf = io.BytesIO()
    wb_template.save(buf)
    wb_template.close()
    buf.seek(0)

    return buf.read(), {
        "C10": c10_value,
        "C11": c11_value,
        "F10": f10_value,
        "C12": c12_value,
        "I2": m_code,
        "filename": secr_filename,
    }


def update_secr_bytes(
    def_bytes: bytes,
    def_filename: str,
    old_secr_bytes: bytes,
    subject: str,
    secr_author: str,
    design_release_engineer: str,
    change_requested_by: str,
    reissue_date: str,
    version: str,
    phase_implemented: str,
    pull_ahead: str,
    secr_change_type: str = "Miscellaneous",
    secr_sequence: int = 1000,
    secr_model_year: str = "",
    secr_program: str = "",
    secr_phase: str = "",
) -> Tuple[bytes, Dict[str, Any]]:
    """Create updated SECR workbook from new DEF and old SECR baseline."""
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"SECR template not found: {TEMPLATE_PATH}")

    stem = Path(def_filename).stem
    parts = stem.split("_")
    if len(parts) < 5:
        raise ValueError(
            "DEF filename must have at least 5 underscore-separated parts "
            "(e.g. 2027_RU_X2_A_vs_2026_RU_X2_A_IP_DEF_DEF_Compare_...)."
        )

    my_full = parts[0]
    vehicle_line = parts[1]
    code1 = parts[2]
    code2 = parts[3]
    idx_def = next((i for i, p in enumerate(parts) if p == "DEF"), None)
    if idx_def is None or idx_def == 0:
        raise ValueError("Could not find 'DEF' segment in the filename.")

    pre_def_string = parts[idx_def - 1]
    c10_value = my_full
    c11_value = vehicle_line
    f10_value = f"{code1}_{code2}"
    c12_value = pre_def_string

    m_code = _build_secr_code(
        c10_value,
        c11_value,
        code1,
        code2,
        secr_change_type,
        secr_sequence,
        secr_model_year=secr_model_year,
        secr_program=secr_program,
        secr_phase=secr_phase,
    )

    wb_template = openpyxl.load_workbook(str(TEMPLATE_PATH))
    wb_def = openpyxl.load_workbook(io.BytesIO(def_bytes), data_only=False)
    wb_old = openpyxl.load_workbook(io.BytesIO(old_secr_bytes), data_only=False)

    if SUMMARY_SHEET not in wb_template.sheetnames:
        raise ValueError("SECR template has no 'Summary' sheet.")
    if SUMMARY_SHEET not in wb_old.sheetnames:
        raise ValueError("Old SECR workbook has no 'Summary' sheet.")

    ws_summary = wb_template[SUMMARY_SHEET]

    ws_summary["C10"] = int(c10_value) if c10_value.isdigit() else c10_value
    ws_summary["C11"] = c11_value
    ws_summary["F10"] = f10_value
    ws_summary["C12"] = c12_value
    ws_summary["I2"] = m_code
    ws_summary["C8"] = m_code

    for ws in wb_def.worksheets:
        _copy_sheet(ws, wb_template)
    wb_def.close()

    summary_ws = next((ws for ws in wb_template.worksheets if ws.title == SUMMARY_SHEET), None)
    if summary_ws:
        others = [ws for ws in wb_template.worksheets if ws is not summary_ws]
        wb_template._sheets = [summary_ws] + others

    _copy_summary_values(
        ws_summary,
        wb_old[SUMMARY_SHEET],
        excluded_cells={"C8", "I2", "I3", "I4"},
    )
    ws_summary["C8"] = m_code
    ws_summary["I2"] = m_code

    if CONNECTOR_SHEET in wb_template.sheetnames and CONNECTOR_SHEET in wb_old.sheetnames:
        _merge_comments_from_old(wb_template[CONNECTOR_SHEET], wb_old[CONNECTOR_SHEET])
    if CIRCUIT_SHEET in wb_template.sheetnames and CIRCUIT_SHEET in wb_old.sheetnames:
        _merge_comments_from_old(wb_template[CIRCUIT_SHEET], wb_old[CIRCUIT_SHEET])

    ws_summary["C7"] = subject
    ws_summary["I10"] = secr_author
    ws_summary["I11"] = design_release_engineer
    ws_summary["I12"] = change_requested_by
    ws_summary["I3"] = version
    ws_summary["F11"] = phase_implemented
    ws_summary["F12"] = pull_ahead
    if reissue_date:
        ws_summary["I4"] = reissue_date

    _populate_update_table_row(ws_summary, version, reissue_date, subject)

    _process_def_def_summary(wb_template, ws_summary)
    _process_connector_sheet(wb_template, ws_summary)
    _process_circuit_sheet(wb_template, ws_summary)

    for ws in wb_template.worksheets:
        ws.protection.disable()
    if hasattr(wb_template, "security") and wb_template.security is not None:
        wb_template.security.lockStructure = False
        wb_template.security.lockWindows = False
        wb_template.security.lockRevision = False

    secr_filename = f"{m_code}.xlsx"

    buf = io.BytesIO()
    wb_template.save(buf)
    wb_template.close()
    wb_old.close()
    buf.seek(0)

    return buf.read(), {
        "C10": c10_value,
        "C11": c11_value,
        "F10": f10_value,
        "C12": c12_value,
        "I2": m_code,
        "filename": secr_filename,
    }
