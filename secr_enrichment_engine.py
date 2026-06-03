"""SECR Reason for Change Enrichment — match DTCRs to harness families and update SECR."""
from __future__ import annotations

import io
import re
from typing import Any, Dict, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# File Loaders
# ---------------------------------------------------------------------------

def _find_header_row(file_bytes: bytes, keyword_sets: list, max_scan: int = 30) -> int:
    """Scan the first `max_scan` rows of an Excel file to find the header row.

    `keyword_sets` is a list of lists; each inner list contains lowercase
    substrings that are expected to appear in column headers.  The first row
    that matches the most keywords wins.  Returns the 0-based row index.
    """
    raw = pd.read_excel(io.BytesIO(file_bytes), header=None, nrows=max_scan, dtype=str)
    best_row = 0
    best_score = -1
    all_keywords = [kw for group in keyword_sets for kw in group]
    for row_idx in range(len(raw)):
        row_vals = " ".join(str(v).lower() for v in raw.iloc[row_idx] if pd.notna(v))
        score = sum(1 for kw in all_keywords if kw in row_vals)
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


def _map_columns(df: pd.DataFrame, mappings: dict) -> pd.DataFrame:
    """Rename df columns to canonical names using case-insensitive substring matching.

    `mappings` maps canonical_name -> list of lowercase substrings to look for.
    Exact-name columns take precedence over keyword-matched ones.
    Only ONE column is mapped per canonical name.
    Returns a DataFrame with columns renamed where matches were found.
    """
    col_map = {}
    for canonical, keywords in mappings.items():
        # First pass: exact case-insensitive match
        for col in df.columns:
            if str(col).strip().lower() == canonical.lower():
                col_map[col] = canonical
                break
        if canonical in col_map.values():
            continue
        # Second pass: substring keyword match
        for col in df.columns:
            if col in col_map:          # already mapped to something
                continue
            col_lower = str(col).strip().lower()
            for kw in keywords:
                if kw in col_lower:
                    col_map[col] = canonical
                    break
            if canonical in col_map.values():
                break
    return df.rename(columns=col_map)


def load_dtcr_report(file_bytes: bytes) -> pd.DataFrame:
    """Load DTCR Report from Excel bytes with auto header-row detection."""
    keyword_sets = [["dtcr"], ["transmittal"], ["reason"], ["status"]]
    header_row = _find_header_row(file_bytes, keyword_sets)
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_row, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    mappings = {
        "DTCR#":               ["dtcr#", "dtcr #", "dtcr number", "dtcr no", "dtcr"],
        "Device Transmittal":  ["device transmittal", "transmittal", "device trans"],
        "Reason for change":   ["reason for change", "reason for", "reason"],
        "Status":              ["status", "request action", "action"],
    }
    df = _map_columns(df, mappings)

    required = ["DTCR#", "Device Transmittal", "Reason for change"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"DTCR Report missing columns: {missing}. "
            f"Detected columns: {list(df.columns[:20])}"
        )
    # Status is optional — fall back to empty string if not present
    if "Status" not in df.columns:
        df["Status"] = ""
    # Deduplicate: if there are multiple "Status" columns, keep the first
    if isinstance(df.columns, pd.Index) and df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated()]
    keep_cols = required + ["Status"]
    df = df[keep_cols].copy()
    # Drop rows where DTCR# is blank (wrapped/continuation rows from Excel)
    df = df.dropna(subset=["DTCR#"]).reset_index(drop=True)
    df["DTCR#"] = df["DTCR#"].astype(str).str.strip()
    return df


def load_dtx_circuits_report(file_bytes: bytes) -> pd.DataFrame:
    """Load DTx Circuits Report from Excel bytes with auto header-row detection."""
    keyword_sets = [["control number", "device control", "dcn"],
                    ["device name", "name"],
                    ["suffix"],
                    ["harness family", "harness", "family"]]
    header_row = _find_header_row(file_bytes, keyword_sets)
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_row, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    mappings = {
        "Device Control Number": ["device control number", "control number", "dcn", "device control"],
        "Device Name":           ["device name", "device nm"],
        "Suffix":                ["suffix"],
        "Harness Family":        ["harness family", "harnfamily", "harness fam", "family"],
    }
    df = _map_columns(df, mappings)

    required = ["Device Control Number", "Device Name", "Suffix", "Harness Family"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"DTx Circuits Report missing columns: {missing}. "
            f"Detected columns: {list(df.columns[:20])}"
        )
    df = df[required].dropna(how="all").reset_index(drop=True)
    return df


def load_generated_secr_workbook(file_bytes: bytes) -> openpyxl.Workbook:
    """Load the generated SECR workbook from bytes."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        return wb
    except Exception as e:
        raise ValueError(f"Failed to load SECR workbook: {e}")


# ---------------------------------------------------------------------------
# Text Processing
# ---------------------------------------------------------------------------

def normalize_text(value: Optional[str]) -> str:
    """Normalize text: uppercase, remove special chars, collapse spaces."""
    if not value:
        return ""
    s = str(value).upper().strip()
    s = re.sub(r"[^\w\s]", " ", s)  # Replace special chars with space
    s = re.sub(r"\s+", " ", s)  # Collapse multiple spaces
    return s


def extract_device_control_number(device_transmittal: str) -> Optional[str]:
    """Extract Device Control Number from Device Transmittal string.
    
    Example: "123456 - SWITCH BANK LEFT" -> "123456"
    """
    if not device_transmittal:
        return None
    # Look for leading digits before dash or space
    match = re.match(r"^(\d+)", str(device_transmittal).strip())
    if match:
        return match.group(1)
    return None


# ---------------------------------------------------------------------------
# Matching Logic
# ---------------------------------------------------------------------------

def match_dtcr_to_harness_family(dtcr_df: pd.DataFrame, dtx_df: pd.DataFrame) -> pd.DataFrame:
    """Match each DTCR to a Harness Family using strict priority logic.

    Priority:
    1) Device Control Number
    2) Device Name
    No suffix matching is performed.
    
    Returns a DataFrame with columns:
    - DTCR#
    - Device Transmittal
    - Extracted Device Control Number
    - Reason for change
    - Status
    - Match Method
    - Matched DTx Value
    - Harness Family
    """
    results = []

    for _, row in dtcr_df.iterrows():
        dtcr_num = row["DTCR#"]
        device_transmittal = row["Device Transmittal"]
        reason = row["Reason for change"]
        status = row["Status"]

        extracted_dcn = extract_device_control_number(device_transmittal)
        match_method = "No Match"
        matched_dtx_value = None
        harness_family = None

        # Step 1: Match by Device Control Number
        if extracted_dcn:
            matching_rows = dtx_df[dtx_df["Device Control Number"].astype(str).str.strip() == extracted_dcn.strip()]
            if not matching_rows.empty:
                harness_family = matching_rows.iloc[0]["Harness Family"]
                matched_dtx_value = extracted_dcn
                match_method = "Device Control Number"

        # Step 2: Match by Device Name (if no DCN match)
        if match_method == "No Match" and device_transmittal:
            normalized_transmittal = normalize_text(device_transmittal)

            # Try Device Name matching
            for _, dtx_row in dtx_df.iterrows():
                device_name = dtx_row.get("Device Name")
                if device_name:
                    normalized_name = normalize_text(device_name)
                    if normalized_name and normalized_name in normalized_transmittal:
                        harness_family = dtx_row["Harness Family"]
                        matched_dtx_value = device_name
                        match_method = "Device Name"
                        break

        results.append({
            "DTCR#": dtcr_num,
            "Device Transmittal": device_transmittal,
            "Extracted Device Control Number": extracted_dcn or "",
            "Reason for change": reason,
            "Status": status,
            "Match Method": match_method,
            "Matched DTx Value": matched_dtx_value or "",
            "Harness Family": harness_family or "",
        })

    return pd.DataFrame(results)


# ---------------------------------------------------------------------------
# SECR Reading & Writing
# ---------------------------------------------------------------------------

def get_secr_harness_family_from_c12(secr_workbook: openpyxl.Workbook) -> Optional[str]:
    """Read the SECR Harness Family from cell C12 of the Summary sheet."""
    try:
        ws = secr_workbook["Summary"]
        value = ws["C12"].value
        return str(value).strip() if value else None
    except Exception:
        return None


def find_reason_for_change_cell(secr_workbook: openpyxl.Workbook) -> Optional[Tuple[str, str]]:
    """Find the Reason for Change field in the SECR workbook.
    
    Returns tuple of (sheet_name, cell_reference) or None if not found.
    """
    try:
        ws = secr_workbook["Summary"]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "Reason for Change" in str(cell.value):
                    return ("Summary", cell.coordinate)
    except Exception:
        pass
    return None


def find_dtcr_number_label_cell(secr_workbook: openpyxl.Workbook) -> Optional[Tuple[str, str]]:
    """Find the cell containing DTCR # label in Summary sheet."""
    try:
        ws = secr_workbook["Summary"]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "DTCR #" in str(cell.value):
                    return ("Summary", cell.coordinate)
    except Exception:
        pass
    return None


def _autosize_cell_for_text(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row_idx: int,
    col_idx: int,
    text: str,
) -> None:
    """Autosize a cell's row/column for multiline text readability."""
    safe_text = text or ""
    lines = safe_text.split("\n") if safe_text else [""]
    longest_line = max((len(line) for line in lines), default=0)
    line_count = max(1, len(lines))

    # Width/height bounds keep layout usable while showing full content.
    target_width = min(max(longest_line + 2, 18), 120)
    target_height = min(max(line_count * 15, 18), 409)

    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = max(
        ws.column_dimensions[col_letter].width or 0,
        target_width,
    )
    ws.row_dimensions[row_idx].height = max(
        ws.row_dimensions[row_idx].height or 0,
        target_height,
    )


def update_secr_reason_for_change(
    secr_workbook: openpyxl.Workbook,
    cell_ref: str,
    reason_for_change_text: str,
) -> None:
    """Write Reason for Change text in the cell directly below the label cell."""
    try:
        ws = secr_workbook["Summary"]
        label_cell = ws[cell_ref]
        target_row = label_cell.row + 1
        target_col = label_cell.column
        target_cell = ws.cell(row=target_row, column=target_col)
        target_cell.value = reason_for_change_text
        target_cell.alignment = Alignment(wrap_text=True, vertical="top")
        _autosize_cell_for_text(ws, target_row, target_col, reason_for_change_text)
    except Exception as e:
        raise RuntimeError(f"Failed to update SECR cell {cell_ref}: {e}")


def update_secr_dtcr_numbers(
    secr_workbook: openpyxl.Workbook,
    dtcr_label_cell_ref: str,
    dtcr_numbers_text: str,
) -> None:
    """Write DTCR numbers in the cell to the right of DTCR # label."""
    try:
        ws = secr_workbook["Summary"]
        label_cell = ws[dtcr_label_cell_ref]
        target_row = label_cell.row
        target_col = label_cell.column + 1
        target_cell = ws.cell(row=target_row, column=target_col)
        target_cell.value = dtcr_numbers_text
        target_cell.alignment = Alignment(wrap_text=True, vertical="top")
        _autosize_cell_for_text(ws, target_row, target_col, dtcr_numbers_text)
    except Exception as e:
        raise RuntimeError(f"Failed to update DTCR numbers near {dtcr_label_cell_ref}: {e}")


# ---------------------------------------------------------------------------
# SECR Enrichment Logic
# ---------------------------------------------------------------------------

def build_reason_for_change_for_secr(
    secr_harness_family: str,
    dtcr_mapping_df: pd.DataFrame,
) -> str:
    """Build the Reason for Change text for a specific SECR Harness Family.
    
    Returns string in format:
    DTCR#: reason
    DTCR#: reason
    """
    # Filter DTCRs matching the SECR Harness Family
    matching = dtcr_mapping_df[dtcr_mapping_df["Harness Family"] == secr_harness_family].copy()

    if matching.empty:
        return ""

    # Build text
    entries = []
    for _, row in matching.iterrows():
        dtcr_num = str(row["DTCR#"]).strip()
        reason = str(row["Reason for change"]).strip()
        entries.append(f"{dtcr_num}: {reason}")

    return "\n".join(entries)


def build_dtcr_numbers_for_secr(
    secr_harness_family: str,
    dtcr_mapping_df: pd.DataFrame,
) -> str:
    """Build newline-separated DTCR numbers matching the SECR Harness Family."""
    matching = dtcr_mapping_df[dtcr_mapping_df["Harness Family"] == secr_harness_family].copy()
    if matching.empty:
        return ""
    dtcr_values = (
        matching["DTCR#"].astype(str).str.strip().dropna().drop_duplicates().tolist()
    )
    return ", ".join(dtcr_values)


def _style_dtcr_mapping_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Apply readable formatting to the DTCR_Harness_Mapping sheet."""
    if ws.max_row < 1 or ws.max_column < 1:
        return

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Header styling
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data alignment and zebra striping
    stripe_fill = PatternFill(start_color="EAF2FA", end_color="EAF2FA", fill_type="solid")
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row % 2 == 0:
                cell.fill = stripe_fill

    # Auto filter + freeze header
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Column widths based on content (bounded for readability)
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 120)

    # Give multiline cells enough height
    for row_idx in range(2, ws.max_row + 1):
        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            max_lines = max(max_lines, str(val).count("\n") + 1)
        ws.row_dimensions[row_idx].height = min(max(max_lines * 15, 18), 120)

def export_dtcr_mapping_styled(dtcr_mapping_df: pd.DataFrame) -> bytes:
    """Export DTCR mapping as a styled standalone workbook (table + autofit)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DTCR_Harness_Family_Mapping"

    for r_idx, row in enumerate(
        [dtcr_mapping_df.columns] + dtcr_mapping_df.values.tolist(), 1
    ):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    _style_dtcr_mapping_sheet(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_secr_enriched_output(
    secr_workbook: openpyxl.Workbook,
    dtcr_extracted_df: pd.DataFrame,
    dtcr_mapping_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    output_filename: Optional[str] = None,
) -> Tuple[bytes, Dict[str, Any]]:
    """Export the enriched SECR as Excel bytes with additional sheets.
    
    Adds sheets: DTCR_Extracted, DTCR_Harness_Mapping, SECR_Enrichment_Summary, Updated_SECR.
    Returns (excel_bytes, metadata_dict).
    """
    # Create a new workbook starting from the updated SECR
    wb_out = secr_workbook

    # Add DTCR_Extracted sheet
    if "DTCR_Extracted" in wb_out.sheetnames:
        del wb_out["DTCR_Extracted"]
    ws_dtcr = wb_out.create_sheet("DTCR_Extracted")
    for r_idx, row in enumerate(
        [dtcr_extracted_df.columns] + dtcr_extracted_df.values.tolist(), 1
    ):
        for c_idx, val in enumerate(row, 1):
            ws_dtcr.cell(row=r_idx, column=c_idx, value=val)

    # Add DTCR_Harness_Mapping sheet
    if "DTCR_Harness_Mapping" in wb_out.sheetnames:
        del wb_out["DTCR_Harness_Mapping"]
    ws_mapping = wb_out.create_sheet("DTCR_Harness_Mapping")
    for r_idx, row in enumerate(
        [dtcr_mapping_df.columns] + dtcr_mapping_df.values.tolist(), 1
    ):
        for c_idx, val in enumerate(row, 1):
            ws_mapping.cell(row=r_idx, column=c_idx, value=val)
    _style_dtcr_mapping_sheet(ws_mapping)

    # Add SECR_Enrichment_Summary sheet
    if "SECR_Enrichment_Summary" in wb_out.sheetnames:
        del wb_out["SECR_Enrichment_Summary"]
    ws_summary = wb_out.create_sheet("SECR_Enrichment_Summary")
    for r_idx, row in enumerate(
        [summary_df.columns] + summary_df.values.tolist(), 1
    ):
        for c_idx, val in enumerate(row, 1):
            ws_summary.cell(row=r_idx, column=c_idx, value=val)

    # Rename the "Updated_SECR" sheet pointer (Note: we modify the output workbook in place,
    # so the original summary sheet is already updated)
    # If needed, create a copy of summary as Updated_SECR for clarity
    # Actually, leave the existing Summary sheet as-is; the enrichment updates it in place.

    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)

    return buf.read(), {
        "filename": output_filename or "SECR_Enriched.xlsx",
        "sheets_added": ["DTCR_Extracted", "DTCR_Harness_Mapping", "SECR_Enrichment_Summary"],
    }


# ---------------------------------------------------------------------------
# Validation & Summary
# ---------------------------------------------------------------------------

def build_enrichment_summary(
    dtcr_mapping_df: pd.DataFrame,
    secr_harness_family: str,
    reason_for_change_applied: str,
) -> pd.DataFrame:
    """Build a summary DataFrame of the enrichment process."""
    total_processed = len(dtcr_mapping_df)
    total_matched = len(dtcr_mapping_df[dtcr_mapping_df["Match Method"] != "No Match"])
    matched_to_secr = len(dtcr_mapping_df[dtcr_mapping_df["Harness Family"] == secr_harness_family])
    matched_by_dcn = len(dtcr_mapping_df[dtcr_mapping_df["Match Method"] == "Device Control Number"])
    matched_by_name = len(dtcr_mapping_df[dtcr_mapping_df["Match Method"] == "Device Name"])
    matched_by_suffix = len(dtcr_mapping_df[dtcr_mapping_df["Match Method"] == "Suffix"])
    not_matched = len(dtcr_mapping_df[dtcr_mapping_df["Match Method"] == "No Match"])

    summary = pd.DataFrame(
        [
            {"Metric": "SECR Harness Family (from C12)", "Value": secr_harness_family},
            {"Metric": "Total DTCRs Processed", "Value": total_processed},
            {"Metric": "Total DTCRs Matched to Any Harness", "Value": total_matched},
            {"Metric": "DTCRs Matched by Device Control Number", "Value": matched_by_dcn},
            {"Metric": "DTCRs Matched by Device Name", "Value": matched_by_name},
            {"Metric": "DTCRs Matched by Suffix", "Value": matched_by_suffix},
            {"Metric": "DTCRs Not Matched", "Value": not_matched},
            {"Metric": "DTCRs Matching This SECR", "Value": matched_to_secr},
            {"Metric": "Reason for Change Applied", "Value": "Yes" if reason_for_change_applied else "No"},
        ]
    )
    return summary


def validate_enrichment_inputs(
    dtcr_df: pd.DataFrame,
    dtx_df: pd.DataFrame,
    secr_harness_family: Optional[str],
) -> Tuple[bool, list]:
    """Validate all inputs for SECR enrichment.
    
    Returns (is_valid, list_of_warnings).
    """
    warnings = []

    if dtcr_df.empty:
        warnings.append("DTCR Report is empty.")
    if dtx_df.empty:
        warnings.append("DTx Circuits Report is empty.")
    if not secr_harness_family:
        warnings.append("SECR cell C12 is empty or invalid.")

    is_valid = len(warnings) == 0
    return is_valid, warnings
