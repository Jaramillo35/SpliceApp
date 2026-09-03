"""Dress an exported workbook so every export from the toolkit reads the same.

The interface schema study, part two (2026-09-03), found eleven writers
producing five visual dialects. Rather than teach eleven writers one style,
the app dresses the bytes on the way out: the engine's values are untouched
(the golden guard and the cell-diff proofs see exactly what they saw), and
the workbook the engineer opens has one header style, frozen panes, filters,
column widths, a print setup, and a Read Me sheet with the run's envelope
and the Versigent mark.

What is deliberately NOT done here:

* ``.xlsm`` files are customer templates with macros — never re-saved.
* Workbooks that carry charts or drawings are left alone; an openpyxl
  round-trip would drop them (the DTx change report's dashboard).
* Customer-format files (the SECR, the DEFE template, the complexity
  workbooks) pass through whatever their extension.
* Sheets that are forms rather than tables (merged cells, a title block
  above the header) keep their layout — the SECR form, the PreOrder title
  block, the DTx dashboard.
* The active sheet is never changed: Circuit Health and the DTx engine read
  ``wb.active`` of the workbooks that feed them, so the Read Me goes last.
"""

from __future__ import annotations

import io
import logging
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

logger = logging.getLogger(__name__)

# ----------------------------------------------------------------- tokens
HEADER_FILL = "1F3B57"      # the applicability review's navy — the most used
HEADER_TEXT = "FFFFFF"
SUBHEADER_FILL = "DDEBF7"
RULE = "D5D9E0"
MUTED = "6B7280"
README_TAB = "6B7280"
ANSWER_TAB = "1F3B57"
STATUS = {                  # fill, text — always beside the word, never alone
    "blocker": ("F8D7DA", "9C1C1C"),
    "high": ("FCE4C6", "8A4B00"),
    "review": ("FFF3CD", "7A5A00"),
    "ok": ("D6EBD6", "14532D"),
    "info": ("DDEBF7", "1F3B57"),
}
MIN_WIDTH, MAX_WIDTH, NOTES_WIDTH = 8, 60, 70
_ASSETS = Path(__file__).resolve().parents[2] / "assets"
#: the ink wordmark on a transparent ground — the black JPG plate would
#: print as a rectangle on a white sheet (scripts/make_logo_variants.py)
LOGO = _ASSETS / "versigent_logo_light.png"
LOGO_FALLBACK = _ASSETS / "versigent_logo_horizontal.jpg"
README_TITLE = "Read Me"
#: customer-format files pass through untouched, whatever their extension
PASS_THROUGH = (r"^SECR_", r"template", r"defe", r"Harness_Complexity_")


def can_dress(filename: str, data: bytes) -> bool:
    """Only plain .xlsx without charts or drawings survive a round-trip intact."""
    if not filename.lower().endswith(".xlsx"):
        return False
    if any(re.search(pat, filename, re.I) for pat in PASS_THROUGH):
        return False
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            names = z.namelist()
    except zipfile.BadZipFile:
        return False
    return not any(n.startswith(("xl/charts/", "xl/drawings/")) for n in names)


def dress(data: bytes, filename: str, *, tool: str = "", version: str = "",
          by: str = "", context: str = "", purpose: str = "",
          inputs: Iterable[str] = (), at: Optional[str] = None) -> bytes:
    """Return the dressed workbook, or the original bytes when it cannot be
    dressed safely. Never raises: a styling failure must not lose a download."""
    if not can_dress(filename, data):
        return data
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data))
        active = wb.active
        for ws in wb.worksheets:
            if ws.sheet_state == "visible" and ws.title != README_TITLE:
                _dress_sheet(ws)
        _readme(wb, filename=filename, tool=tool, version=version, by=by,
                context=context, purpose=purpose, inputs=list(inputs), at=at)
        wb.active = wb.worksheets.index(active)
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()
    except Exception as exc:  # noqa: BLE001 — the download must still happen
        logger.warning("Could not dress %s: %s", filename, exc)
        return data


# --------------------------------------------------------------- one sheet
def header_row_of(ws) -> Optional[int]:
    """The header row: the widest of the top few rows.

    Taking the first row with two filled cells reads a title band as the
    header — and then the Read Me reports a row count six too high. The
    widest row wins instead, earliest on a tie, which lands on row 1 for a
    plain table and on the real header under a title block.
    """
    best, best_filled = None, 1
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8)):
        filled = sum(1 for c in row if c.value not in (None, ""))
        if filled > best_filled:
            best, best_filled = row[0].row, filled
    return best


def is_table(ws) -> bool:
    """A table has a header in its top three rows, rows under it, and no
    merged cells (a form or a title block has merges)."""
    if ws.merged_cells.ranges:
        return False
    header = header_row_of(ws)
    return header is not None and ws.max_row > header and ws.max_column >= 2


def _dress_sheet(ws) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if not is_table(ws):
        return
    header = header_row_of(ws)
    last_col = ws.max_column
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    font = Font(bold=True, color=HEADER_TEXT)
    border = Border(bottom=Side(style="thin", color=RULE))
    for cell in ws[header]:
        if cell.column > last_col:
            break
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center",
                                   horizontal=cell.alignment.horizontal or "left")
    # freeze below the header unless the writer chose a deeper pane already
    if not ws.freeze_panes:
        ws.freeze_panes = f"A{header + 1}"
    if not ws.auto_filter.ref:
        ws.auto_filter.ref = f"A{header}:{get_column_letter(last_col)}{ws.max_row}"
    _widths(ws, header)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header}:{header}"


def _widths(ws, header: int) -> None:
    """Widths from content, for columns the writer left unset."""
    from openpyxl.utils import get_column_letter
    sample_rows = min(ws.max_row, 300)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        # openpyxl invents a 13-wide dimension on access; only one that came
        # from the file counts as "the writer chose a width"
        if letter in ws.column_dimensions and ws.column_dimensions[letter].width:
            continue
        longest = 0
        for row in ws.iter_rows(min_row=header, max_row=sample_rows,
                                min_col=col, max_col=col):
            v = row[0].value
            if v is None:
                continue
            longest = max(longest, max(len(part) for part in str(v).split("\n")))
        head = str(ws.cell(row=header, column=col).value or "")
        wide = re.search(r"note|detail|reason|what it means|description|proof",
                         head, re.I)
        width = min(NOTES_WIDTH if wide else MAX_WIDTH, max(MIN_WIDTH, longest + 2))
        ws.column_dimensions[letter].width = width
        if wide:
            for row in ws.iter_rows(min_row=header + 1, max_row=sample_rows,
                                    min_col=col, max_col=col):
                row[0].alignment = row[0].alignment.copy(wrap_text=True, vertical="top")


# ---------------------------------------------------------------- read me
def _readme(wb, *, filename: str, tool: str, version: str, by: str,
            context: str, purpose: str, inputs: list, at: Optional[str]) -> None:
    from openpyxl.styles import Alignment, Font

    if README_TITLE in wb.sheetnames:
        return
    ws = wb.create_sheet(README_TITLE)
    ws.sheet_properties.tabColor = README_TAB
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 96
    row = 1
    if _logo(ws, "A1"):
        row = 5
    ws.cell(row=row, column=1, value=(tool or "System Engineer Toolkit")
            + (f" — {context}" if context else "")).font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value=filename).font = Font(color=MUTED)
    row += 2
    stamp = at or datetime.now().strftime("%Y-%m-%d %H:%M")
    facts = [("Generated", stamp + (f" · by {by}" if by else "")),
             ("Tool", " · ".join(x for x in ("System Engineer Toolkit " + version
                                              if version else "System Engineer Toolkit",
                                              tool) if x)),
             ("Programme", context or "—")]
    if purpose:
        facts.append(("What this is", purpose))
    if inputs:
        facts.append(("Inputs", "\n".join(inputs)))
    for key, value in facts:
        ws.cell(row=row, column=1, value=key).font = Font(bold=True, color=MUTED)
        c = ws.cell(row=row, column=2, value=value)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Sheets").font = Font(bold=True, color=MUTED)
    ws.cell(row=row, column=2, value="Row counts exclude the header. A sheet with 0 rows "
            "found nothing — it is not broken.").font = Font(color=MUTED, italic=True)
    row += 1
    for sheet in wb.worksheets:
        if sheet.title == README_TITLE:
            continue
        header = header_row_of(sheet)
        n = max(sheet.max_row - header, 0) if header else sheet.max_row
        state = "" if sheet.sheet_state == "visible" else f" ({sheet.sheet_state})"
        ws.cell(row=row, column=1, value=sheet.title + state)
        ws.cell(row=row, column=2, value=f"{n:,} row(s)")
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Legend").font = Font(bold=True, color=MUTED)
    ws.cell(row=row, column=2, value="Status is always written as a word; the tint beside it "
            "is a reminder, not the message.").font = Font(color=MUTED, italic=True)
    row += 1
    for kind, (fill, text) in STATUS.items():
        from openpyxl.styles import PatternFill
        c = ws.cell(row=row, column=1, value=kind)
        c.fill = PatternFill("solid", fgColor=fill)
        c.font = Font(color=text, bold=True)
        ws.cell(row=row, column=2, value={
            "blocker": "a vehicle would be built wrong — must be resolved",
            "high": "usually real, needs a decision",
            "review": "worth a look; bookkeeping or attribute differences",
            "ok": "checked and clean",
            "info": "context, not a finding",
        }[kind]).font = Font(color=MUTED)
        row += 1


def _logo(ws, anchor: str) -> bool:
    try:
        from openpyxl.drawing.image import Image
        source = LOGO if LOGO.exists() else LOGO_FALLBACK
        if not source.exists():
            return False
        img = Image(str(source))
        img.width, img.height = 214, 60   # the mark at print size
        ws.add_image(img, anchor)
        ws.row_dimensions[1].height = 48
        return True
    except Exception as exc:  # noqa: BLE001 — a missing logo is not a failure
        logger.debug("logo not embedded: %s", exc)
        return False


def export_name(stem: str, *, context: str = "", at: Optional[datetime] = None,
                ext: str = ".xlsx") -> str:
    """``Circuit_Chart_2030QX_V1_A_2026-09-03_0754.xlsx``.

    A constant filename means yesterday's export and today's are the same
    file in a Downloads folder. The programme, the phase and the moment the
    run happened make each one identifiable; ``context`` is the page's
    programme chip, in whatever punctuation it uses.
    """
    parts = [stem.strip("_")]
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", context or "").strip("_")
    if cleaned:
        parts.append(cleaned)
    parts.append((at or datetime.now()).strftime("%Y-%m-%d_%H%M"))
    return "_".join(parts) + ext


def status_style(kind: str):
    """Fill and font for a status cell, for writers that adopt the tokens."""
    from openpyxl.styles import Font, PatternFill
    fill, text = STATUS.get(kind, STATUS["info"])
    return PatternFill("solid", fgColor=fill), Font(color=text, bold=True)
