"""HRN + CMP chart engine — deterministic, file-in/file-out, no UI imports.

This is the single home of the converter (the standalone desktop app it was
ported from has been retired); the workbook output was verified cell-for-cell
against it. Everything here works on bytes so the engine runs identically
under Streamlit, splice-api (Docker/k8s), and tests.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from splice.config import ASSETS_DIR

#: Supplier list shipped with the app (SUPPLIER_PREFIX / SUPPLIER_NAME columns).
DEFAULT_SUPPLIER_PATH = ASSETS_DIR / "DEF Supplier Codes.xlsx"

# Fallback supplier prefixes used when no supplier file is available.
EMBEDDED_SUPPLIER_MAP = {
    'TE CONNECTIVITY': 'DZ',
    'ROSENBERGER': 'G3',
    'YAZAKI': 'YZ',
    'BENTLEY-HARRIS INC': 'BH',
    'BENTLEY-HARRIS': 'BH',
    'AMP': 'H3',
    'MOLEX': 'CM',
    'JAEIL ENGINEERING': 'JA',
}

# 0-based HRN field index -> column name. Unmapped indices become C<n>.
HRN_COLUMN_NAMES = {
    0: 'CKT_ID', 1: 'FROM_CNUM', 2: 'FROM', 3: 'FROM_CAV', 4: 'TO_CNUM',
    5: 'TO', 6: 'TO_CAV', 7: 'SIZE', 8: 'MSPEC', 9: 'OD',
    12: 'HARNESS_FAMILY', 13: 'Ckt_Color', 14: 'SALESCODE',
    21: 'SIZE_MSPEC', 24: 'CKT FUNCTION',
}

HRN_ASSEMBLY_FIELD = 12
HRN_SALESCODE_FIELD = 14


# ---------------------------------------------------------------------------
# Output naming
# ---------------------------------------------------------------------------

@dataclass
class HrnNameInfo:
    """Fields extracted from an HRN file name.

    Example: 68605261AA_2028DJ2P_X1_A_07_07_26_14_14_45_EC_MIRROR_08-07-2026
      model_year='2028', program='DJ', family='EC_MIRROR'
    """
    family: Optional[str] = None
    model_year: Optional[str] = None
    program: Optional[str] = None

    @property
    def complete(self) -> bool:
        return bool(self.family and self.model_year and self.program)


def parse_hrn_filename(stem: str) -> HrnNameInfo:
    """Extract harness family, model year, and program from an HRN stem.

    - model year + program come from the token that starts '20YY' followed by
      two letters (e.g. '2028DJ2P' -> 2028, DJ)
    - the harness family is everything after the last all-digit token and
      before the trailing MM-DD-YYYY / MMDDYYYY date token (families may
      contain underscores, e.g. EC_MIRROR)
    """
    tokens = stem.split('_')
    info = HrnNameInfo()

    date_idx = None
    for i, t in enumerate(tokens):
        if re.fullmatch(r'\d{2}-\d{2}-\d{4}', t) or re.fullmatch(r'\d{8}', t):
            date_idx = i

    for t in tokens:
        m = re.match(r'^(20\d{2})([A-Za-z]{2})', t)
        if m:
            info.model_year, info.program = m.group(1), m.group(2).upper()
            break

    end = date_idx if date_idx is not None else len(tokens)
    last_numeric = None
    for i in range(end):
        if re.fullmatch(r'\d+', tokens[i]):
            last_numeric = i
    if last_numeric is not None and last_numeric + 1 < end:
        family = '_'.join(tokens[last_numeric + 1:end]).strip('_')
        info.family = family or None

    return info


def output_basename(hrn_filename: str) -> str:
    """{Family}_{ModelYear}{Program}_Chart_{MMDDYYYY}, dated the day the
    conversion runs. Falls back to the HRN's own stem when the name doesn't
    follow the pattern."""
    stem = Path(hrn_filename).stem
    info = parse_hrn_filename(stem)
    if not info.complete:
        return stem
    return (f'{info.family}_{info.model_year}{info.program}_Chart_'
            f'{datetime.now().strftime("%m%d%Y")}')


# ---------------------------------------------------------------------------
# Supplier list
# ---------------------------------------------------------------------------

def _norm_alnum(s) -> str:
    return re.sub(r'[^A-Z0-9]', '', str(s).upper())


def load_supplier_map(source) -> Optional[dict]:
    """Load supplier name -> prefix mapping from a path or bytes (Excel/CSV).

    Accepts SUPPLIER_NAME / SUPPLIER_PREFIX columns in either order; for
    headerless two-column files the prefix column is detected as the one with
    the short codes so the mapping is never built backwards.
    """
    buf = io.BytesIO(source) if isinstance(source, (bytes, bytearray)) else source
    try:
        sdf = pd.read_excel(buf, dtype=str)
    except Exception:
        try:
            if hasattr(buf, 'seek'):
                buf.seek(0)
            sdf = pd.read_csv(buf, dtype=str)
        except Exception:
            return None

    sdf = sdf.fillna('')
    cols_upper = [str(c).upper() for c in sdf.columns]
    if 'SUPPLIER_NAME' in cols_upper and 'SUPPLIER_PREFIX' in cols_upper:
        name_col = sdf.columns[cols_upper.index('SUPPLIER_NAME')]
        pref_col = sdf.columns[cols_upper.index('SUPPLIER_PREFIX')]
    elif sdf.shape[1] >= 2:
        col_a, col_b = sdf.columns[0], sdf.columns[1]
        avg_a = sdf[col_a].astype(str).str.strip().str.len().mean()
        avg_b = sdf[col_b].astype(str).str.strip().str.len().mean()
        pref_col, name_col = (col_a, col_b) if avg_a <= avg_b else (col_b, col_a)
    else:
        return None

    supplier_map = {}
    for _, row in sdf.iterrows():
        name = str(row[name_col]).strip()
        prefix = str(row[pref_col]).strip()
        if name and prefix:
            supplier_map[name.upper()] = prefix
    return supplier_map or None


def default_supplier_map() -> dict:
    """The shipped supplier list, or the embedded fallback subset."""
    if DEFAULT_SUPPLIER_PATH.exists():
        sm = load_supplier_map(DEFAULT_SUPPLIER_PATH)
        if sm:
            return sm
    return dict(EMBEDDED_SUPPLIER_MAP)


def find_supplier_prefix(token, supplier_map) -> Optional[str]:
    """Exact key, normalized key, whole-word, then substring matching."""
    if not token or not supplier_map:
        return None
    tu = str(token).strip().upper()
    if not tu:
        return None
    if tu in supplier_map:
        return supplier_map[tu]
    nt = _norm_alnum(tu)
    if not nt:
        return None
    for k, v in supplier_map.items():
        if _norm_alnum(k) == nt:
            return v
    for k, v in supplier_map.items():
        words = [w for w in re.split(r'[^A-Z0-9]+', k.upper()) if w]
        if tu in words:
            return v
    best, best_score = None, 0
    for k, v in supplier_map.items():
        nk = _norm_alnum(k)
        if nt in nk and len(nt) > best_score:
            best_score = len(nt)
            best = v
    return best


# ---------------------------------------------------------------------------
# Input parsing
# ---------------------------------------------------------------------------

@dataclass
class HrnData:
    rows: list
    assembly: str = ''
    tokens: list = field(default_factory=list)


def parse_hrn(hrn_bytes: bytes) -> HrnData:
    """Parse comma-delimited .hrn content into rows plus assembly/salescode."""
    rows, tokens, assemblies = [], [], []
    text = hrn_bytes.decode('utf-8', errors='replace')
    for line in text.splitlines():
        if not line.strip():
            continue
        parts = [p.strip() for p in line.split(',')]
        rows.append(parts)
        if len(parts) > HRN_ASSEMBLY_FIELD and parts[HRN_ASSEMBLY_FIELD]:
            assemblies.append(parts[HRN_ASSEMBLY_FIELD])
        if len(parts) > HRN_SALESCODE_FIELD and parts[HRN_SALESCODE_FIELD]:
            t = parts[HRN_SALESCODE_FIELD]
            if t not in tokens:
                tokens.append(t)
    if not rows:
        raise ValueError("HRN file contains no data rows")
    return HrnData(rows=rows, assembly=assemblies[0] if assemblies else '', tokens=tokens)


def read_matrix_csv(csv_bytes: bytes) -> pd.DataFrame:
    """Harness matrix CSV (semicolon-delimited, with comma fallback)."""
    try:
        df = pd.read_csv(io.BytesIO(csv_bytes), sep=';', engine='python', dtype=str)
    except Exception:
        df = pd.read_csv(io.BytesIO(csv_bytes), dtype=str)
    if df.shape[1] < 1:
        raise ValueError("Matrix CSV has no columns")
    return df.fillna('')


def _cmp_entry_from_tokens(tokens, supplier_map):
    cleaned = ['' if t is None else str(t).strip() for t in tokens]
    nonempty = [t for t in cleaned if t and t.upper() != 'NAN']
    if not nonempty:
        return None
    key, last = nonempty[0], nonempty[-1]
    sup_code = None
    for t in nonempty:
        sp = find_supplier_prefix(t, supplier_map)
        if sp:
            sup_code = sp
            break
    return key, (f"{last}~{sup_code}" if sup_code else last)


def parse_cmp(cmp_bytes: bytes, supplier_map) -> Optional[dict]:
    """Parse .cmp content into {connector name: CNUM[~supplier prefix]}."""
    cmp_map = {}
    try:
        cmp_df = pd.read_csv(io.BytesIO(cmp_bytes), sep=None, engine='python',
                             header=None, dtype=str)
        for _, row in cmp_df.iterrows():
            entry = _cmp_entry_from_tokens(row.tolist(), supplier_map)
            if entry:
                cmp_map[entry[0]] = entry[1]
    except Exception:
        text = cmp_bytes.decode('utf-8', errors='ignore')
        for line in text.splitlines():
            s = line.strip()
            if not s:
                continue
            for tokens in (s.split(','), s.split(), re.findall(r'[A-Za-z0-9_.-]+', s)):
                entry = _cmp_entry_from_tokens(tokens, supplier_map)
                if entry:
                    cmp_map[entry[0]] = entry[1]
                    break
    return cmp_map or None


# ---------------------------------------------------------------------------
# Connector matching
# ---------------------------------------------------------------------------

def normalize_connector_key(k) -> str:
    if k is None:
        return ''
    s = str(k).strip().upper()
    for sep in ('.', '/', '-'):
        if sep in s:
            s = s.split(sep)[0]
    return _norm_alnum(s)


class ConnectorMatcher:
    """Match HRN FROM/TO connector names against a CMP map (memoized)."""

    def __init__(self, cmp_map: dict):
        self.cmp_map = cmp_map
        self.norm_map = {}
        for k, v in cmp_map.items():
            nk = normalize_connector_key(k)
            if nk and nk not in self.norm_map:
                self.norm_map[nk] = v
        self._cache = {}

    def match(self, value: str) -> str:
        value = str(value).strip()
        if not value:
            return ''
        if value not in self._cache:
            self._cache[value] = self._match_uncached(value)
        return self._cache[value]

    def _match_uncached(self, value: str) -> str:
        for candidate in (value, value.upper()):
            if candidate in self.cmp_map:
                return self.cmp_map[candidate]
        nc = normalize_connector_key(value)
        if nc and nc in self.norm_map:
            return self.norm_map[nc]
        if nc:
            for nk, v in self.norm_map.items():
                if nk.startswith(nc) or nc.startswith(nk) or nc in nk or nk in nc:
                    return v
        return ''


# ---------------------------------------------------------------------------
# Workbook build
# ---------------------------------------------------------------------------

@dataclass
class ChartResult:
    workbook: bytes
    filename: str            # e.g. EC_MIRROR_2028DJ_Chart_08212026.xlsx
    unmatched: list          # rows where FROM/TO found no CMP match
    invalid_prefixes: list   # CNUM supplier suffixes not in the supplier list


def _hrn_dataframe(hrn: HrnData) -> pd.DataFrame:
    max_cols = max(len(r) for r in hrn.rows)
    padded = [r + [''] * (max_cols - len(r)) for r in hrn.rows]
    df = pd.DataFrame(padded)
    df.columns = [HRN_COLUMN_NAMES.get(i, f'C{i + 1}') for i in range(max_cols)]
    return df


def _apply_cmp(hrn_df: pd.DataFrame, cmp_map: dict) -> None:
    matcher = ConnectorMatcher(cmp_map)
    for target_col, source_col in (('FROM_CNUM', 'FROM'), ('TO_CNUM', 'TO')):
        if target_col not in hrn_df.columns:
            hrn_df[target_col] = ''
        if source_col not in hrn_df.columns:
            continue
        hrn_df[target_col] = [
            matcher.match(v) or existing
            for v, existing in zip(hrn_df[source_col], hrn_df[target_col])
        ]


def _harness_applicability(hrn_df: pd.DataFrame, csv_df: pd.DataFrame):
    harness_ids = [str(x).strip() for x in csv_df.iloc[:, 0].astype(str)]
    sales_to_rows = {}
    for j, col in enumerate(list(csv_df.columns[1:]), start=1):
        key = str(col).strip()
        sales_to_rows[key] = {
            i for i in range(len(csv_df))
            if str(csv_df.iat[i, j]).strip().upper() == 'X'
        }

    matrix = [[''] * len(harness_ids) for _ in range(len(hrn_df))]
    if 'SALESCODE' in hrn_df.columns:
        compact = {k.replace(' ', ''): k for k in sales_to_rows}
        for ridx, sale in enumerate(hrn_df['SALESCODE']):
            sale = str(sale).strip()
            if not sale:
                continue
            key = sale if sale in sales_to_rows else compact.get(sale.replace(' ', ''))
            if not key:
                continue
            for h_idx in sales_to_rows.get(key, ()):
                if 0 <= h_idx < len(harness_ids):
                    matrix[ridx][h_idx] = 'X'
    return harness_ids, matrix


def _collect_unmatched(hrn_df: pd.DataFrame) -> list:
    unmatched = []
    for ridx, row in hrn_df.iterrows():
        from_val = str(row.get('FROM', '')).strip()
        to_val = str(row.get('TO', '')).strip()
        from_cnum = str(row.get('FROM_CNUM', '')).strip()
        to_cnum = str(row.get('TO_CNUM', '')).strip()
        if (from_val and not from_cnum) or (to_val and not to_cnum):
            unmatched.append({
                'row': ridx, 'CKT_ID': row.get('CKT_ID', ''),
                'FROM': from_val, 'FROM_CNUM': from_cnum,
                'TO': to_val, 'TO_CNUM': to_cnum,
            })
    return unmatched


def _collect_invalid_prefixes(hrn_df: pd.DataFrame, supplier_map: dict) -> list:
    valid = {str(v).strip().upper() for v in supplier_map.values() if v}
    invalids = []
    for ridx, row in hrn_df.iterrows():
        for col in ('FROM_CNUM', 'TO_CNUM'):
            val = str(row.get(col, '')).strip()
            if '~' in val:
                suffix = val.split('~')[-1].strip().upper()
                if suffix and suffix not in valid:
                    invalids.append({
                        'row': ridx + 1, 'CKT_ID': row.get('CKT_ID', ''),
                        'column': col, 'value': val, 'prefix': suffix,
                    })
    return invalids


def _style_workbook(buf: io.BytesIO) -> io.BytesIO:
    """Auto-size columns everywhere; style HRN_Raw (borders, rotated harness
    headers, frozen header row, sanitized unique headers, autofilter)."""
    wb = load_workbook(buf)

    for ws in wb.worksheets:
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=0,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    if 'HRN_Raw' in wb.sheetnames:
        ws = wb['HRN_Raw']
        thin = Side(border_style='thin', color='FF000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        ckt_idx = headers.index('CKT FUNCTION') if 'CKT FUNCTION' in headers else None

        harness_cols = []
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            if ckt_idx is not None and col_idx - 1 > ckt_idx:
                harness_cols.append(col_idx)

        for col_idx in harness_cols:
            cell = ws.cell(row=1, column=col_idx)
            cell.alignment = Alignment(textRotation=90, horizontal='center',
                                       vertical='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = 4

        ws.row_dimensions[1].height = 80

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                if cell.column in harness_cols:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.freeze_panes = ws['A2']

        seen = set()
        for col_idx, cell in enumerate(ws[1], start=1):
            val = str(cell.value).strip() if cell.value is not None else ''
            safe = re.sub(r'_+', '_', re.sub(r'[^A-Za-z0-9_]', '_', val)).strip('_')
            safe = (safe or f'Column{col_idx}')[:31]
            base, suffix = safe, 1
            while safe.upper() in seen:
                safe = f"{base}_{suffix}"[:31]
                suffix += 1
            cell.value = safe
            seen.add(safe.upper())

        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"

    out = io.BytesIO()
    wb.save(out)
    return out


def build_chart(hrn_filename: str, hrn_bytes: bytes, csv_bytes: bytes,
                cmp_bytes: Optional[bytes] = None,
                supplier_map: Optional[dict] = None) -> ChartResult:
    """Build the chart workbook for one HRN/CSV(/CMP) set, fully in memory."""
    if supplier_map is None:
        supplier_map = default_supplier_map()

    hrn = parse_hrn(hrn_bytes)
    csv_df = read_matrix_csv(csv_bytes)
    cmp_map = parse_cmp(cmp_bytes, supplier_map) if cmp_bytes else None

    hrn_df = _hrn_dataframe(hrn)
    if cmp_map:
        _apply_cmp(hrn_df, cmp_map)

    harness_ids, matrix = _harness_applicability(hrn_df, csv_df)

    cols = list(hrn_df.columns)
    insert_pos = cols.index('CKT FUNCTION') + 1 if 'CKT FUNCTION' in cols else len(cols)
    for i, hname in enumerate(harness_ids):
        cols.insert(insert_pos + i, hname)
    for hidx, hname in enumerate(harness_ids):
        hrn_df[hname] = [matrix[r][hidx] for r in range(len(hrn_df))]
    hrn_df = hrn_df[cols]

    raw = io.BytesIO()
    with pd.ExcelWriter(raw, engine='openpyxl') as writer:
        csv_df.to_excel(writer, sheet_name='Matrix', index=False)
        hrn_df.to_excel(writer, sheet_name='HRN_Raw', index=False)
    raw.seek(0)
    styled = _style_workbook(raw)

    unmatched = _collect_unmatched(hrn_df) if cmp_map else []
    invalids = _collect_invalid_prefixes(hrn_df, supplier_map) if cmp_map else []

    return ChartResult(
        workbook=styled.getvalue(),
        filename=f'{output_basename(hrn_filename)}.xlsx',
        unmatched=unmatched,
        invalid_prefixes=invalids,
    )
