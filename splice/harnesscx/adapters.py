"""Master-complexity and harness-family cross-reference adapters.

Ported from WEAVE (harness-suite-v2) into Splice. The rules:

* the cross-reference workbook maps DTx family names <-> master worksheet
  names <-> canonical harness-family names;
* sales-code expressions live in **row 9** of each per-family master
  worksheet, mixed with phases / PC / market codes — so a row-9 token counts
  as a sales code only if it also appears in the DTx ``Sales Code`` data;
* sources are never modified; always parsed from bytes.
"""

from __future__ import annotations

import io
import logging
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from splice.common.errors import SpliceError
from splice.harnesscx.models import (
    CombinedExpr,
    FamilyMatrix,
    MatrixRow,
    ProposalClass,
    SalesCodeColumn,
)
from splice.harnesscx.partition import ordered_sides, partition_columns, row_side

logger = logging.getLogger(__name__)

SALES_CODE_ROW = 9                       # 1-indexed Excel row holding the sales-code headers
FEATURE_ROW = 7                          # feature descriptions live on row 7 of every sheet
FIRST_DATA_ROW = 10                      # part-number rows begin here
_CODE_TOKEN = re.compile(r"\b[A-Z0-9]{3}\b")  # 3-char sales-code shape (e.g. AHT, CJK, 501)
_CARRYOVER = "C/O"
_NA = "N/A"

# Current-column sentinels that are NOT real part numbers — a part row carrying one of
# these must never appear in the workbench PN table or a generated file.
_NON_PART_PREFIXES = ("NO HARNESS", "DELETE", "CANCEL", "N/A")


def is_non_part_pn(pn: object) -> bool:
    """True when a 'part number' is really a placeholder/sentinel (NO HARNESS, DELETE, …)."""
    u = str(pn or "").strip().upper()
    return not u or any(u.startswith(p) for p in _NON_PART_PREFIXES)


def extract_code_tokens(text: object) -> list[str]:
    """Pull candidate 3-character sales-code tokens from a (possibly combined) value.

    ``"CJK/LEQ"`` -> ``["CJK", "LEQ"]``; ``"AHT (XEY)"`` -> ``["AHT", "XEY"]``.
    """
    if text is None:
        return []
    return _CODE_TOKEN.findall(str(text).upper())


def _separable_or_list(expr: str, tokens: list[str]) -> bool:
    """True when ``expr`` is a plain OR-list of ``tokens`` — only ``/``, ``,`` and
    whitespace between them (e.g. ``"CM5/CVM"``). Safe to split into independent
    columns. AND (``+``/``&``), negation (``-``), grouping (``()``) or equality
    (``=``) make it a :class:`CombinedExpr` the SE reviews (a pure equality is
    then auto-resolved: included with its two codes pre-filled).
    """
    residue = expr.upper()
    for t in set(tokens):
        residue = residue.replace(t, " ")
    residue = re.sub(r"[\s/,]+", "", residue)
    return residue == ""


class CrossRef:
    """Harness-family name mappings from the cross-reference workbook."""

    def __init__(self) -> None:
        self.dtx_to_worksheet: dict[str, str] = {}
        self.dtx_to_canonical: dict[str, str] = {}
        self.worksheet_to_canonical: dict[str, str] = {}
        self.family_map: dict[str, list[str]] = {}   # alias -> DTx name(s)
        self.dtx_families: set[str] = set()

    @property
    def worksheets(self) -> set[str]:
        return set(self.worksheet_to_canonical)

    def worksheet_for_dtx(self, dtx_family: str) -> str | None:
        return self.dtx_to_worksheet.get(str(dtx_family).strip())

    def _add_alias(self, alias: str, dtx: str) -> None:
        alias = str(alias).strip()
        if alias and dtx:
            lst = self.family_map.setdefault(alias, [])
            if dtx not in lst:
                lst.append(dtx)


def load_crossref(file_bytes: bytes) -> CrossRef:
    """Load the family cross-reference (columns: Harness Family / Complexity File /
    DTx Family Name)."""
    if not file_bytes:
        raise SpliceError("Cross-reference workbook is empty or was not provided.")
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    except Exception as exc:
        raise SpliceError(f"Could not read the cross-reference workbook: {exc}") from exc
    df.columns = [str(c).strip() for c in df.columns]
    need = {"Harness Family", "Complexity File", "DTx Family Name"}
    missing = need - set(df.columns)
    if missing:
        raise SpliceError(f"Cross-reference is missing column(s): {sorted(missing)}. "
                          f"Found: {list(df.columns)}")

    cr = CrossRef()
    for _, row in df.iterrows():
        canonical = str(row.get("Harness Family") or "").strip()
        worksheet = str(row.get("Complexity File") or "").strip()
        dtx = str(row.get("DTx Family Name") or "").strip()
        if dtx in ("", "Column3", "nan"):   # skip the spurious 'Column1/2/3' header row
            continue
        if worksheet:
            cr.dtx_to_worksheet[dtx] = worksheet
            if canonical:
                cr.worksheet_to_canonical.setdefault(worksheet, canonical)
        if canonical:
            cr.dtx_to_canonical[dtx] = canonical
        cr.dtx_families.add(dtx)
        cr._add_alias(dtx, dtx)
        cr._add_alias(worksheet, dtx)
        cr._add_alias(canonical, dtx)
    logger.info("Cross-reference: %d DTx-family maps, %d worksheets.",
                len(cr.dtx_to_worksheet), len(cr.worksheets))
    return cr


def extract_family_sales_codes(
    master_bytes: bytes,
    worksheets: set[str],
    universe: set[str],
) -> dict[str, set[str]]:
    """Per family worksheet present in the master, its row-9 sales codes.

    A row-9 token is kept only if it is in ``universe`` (the DTx sales-code set).
    """
    if not master_bytes:
        raise SpliceError("Master complexity workbook is empty or was not provided.")
    try:
        wb = load_workbook(io.BytesIO(master_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise SpliceError(f"Could not read the master complexity workbook: {exc}") from exc

    result: dict[str, set[str]] = {}
    try:
        for sheet in wb.sheetnames:
            if sheet not in worksheets:
                continue
            ws = wb[sheet]
            codes: set[str] = set()
            for cell in next(ws.iter_rows(min_row=SALES_CODE_ROW, max_row=SALES_CODE_ROW), []):
                for token in extract_code_tokens(cell.value):
                    if token in universe:
                        codes.add(token)
            result[sheet] = codes
    finally:
        wb.close()
    return result


def _cell_text(cell) -> str:
    return "" if cell is None or cell.value is None else str(cell.value).strip()


def master_worksheets(master_bytes: bytes) -> list[str]:
    """The sheet names of a master complexity workbook (for unmapped fallback)."""
    if not master_bytes:
        return []
    try:
        wb = load_workbook(io.BytesIO(master_bytes), read_only=True)
    except Exception as exc:
        raise SpliceError(f"Could not read the master complexity workbook: {exc}") from exc
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def extract_family_matrix(
    master_bytes: bytes,
    worksheet: str,
    universe: set[str],
    canonical_family: str = "",
    *,
    family_dtx_codes: set[str] | None = None,
) -> FamilyMatrix:
    """Build the proposed individual-complexity matrix for one harness family.

    Sales codes are row-9 tokens present in ``universe`` (separable OR-lists split,
    originals retained); the Current P/N is under the ``Current`` column with
    carryover (``C/O``) resolution and DELETE/``N/A``/Cancel exclusion; the
    Previous P/N is the most recent valid PN before ``Current``. Every proposed
    value is classified and carries its source. A pure equality expression
    (``XH3=XH4``) is auto-resolved: included with one column per side.
    """
    if not master_bytes:
        raise SpliceError("Master complexity workbook is empty or was not provided.")
    try:
        # data_only=True resolves the Current-column formulas to their cached values.
        wb = load_workbook(io.BytesIO(master_bytes), data_only=True)
    except Exception as exc:
        raise SpliceError(f"Could not read the master complexity workbook: {exc}") from exc
    if worksheet not in wb.sheetnames:
        wb.close()
        raise SpliceError(f"Worksheet '{worksheet}' not found in the master complexity workbook.")
    ws = wb[worksheet]
    try:
        return _build_matrix(ws, worksheet, universe, canonical_family, family_dtx_codes)
    finally:
        wb.close()


def _build_matrix(ws, worksheet: str, universe: set[str], canonical_family: str,
                  family_dtx_codes: set[str] | None = None) -> FamilyMatrix:
    # Locate the 'Made from' and 'Current' anchor columns on row 9.
    made_from_col = current_col = None
    for c in range(1, ws.max_column + 1):
        txt = _cell_text(ws.cell(SALES_CODE_ROW, c)).lower()
        if txt.startswith("made from"):
            made_from_col = c
        elif txt == "current":
            current_col = c
    if current_col is None:
        raise SpliceError(f"'Current' column not found on row {SALES_CODE_ROW} of "
                          f"'{worksheet}'.")
    made_from_col = made_from_col or 3

    phase_cols = [c for c in range(made_from_col + 1, current_col)
                  if _cell_text(ws.cell(SALES_CODE_ROW, c))]

    # Sales-code columns (right of Current). A cell with one token, or a pure OR-list
    # (only '/'/','), becomes independent sales-code column(s); a cell that mixes AND /
    # negation / grouping cannot be separated safely and becomes a CombinedExpr the SE
    # reviews. First occurrence of a token wins.
    sales_codes: list[SalesCodeColumn] = []
    combined_exprs: list[CombinedExpr] = []
    code_to_col: dict[str, int] = {}
    combined_cols: list[tuple[CombinedExpr, int]] = []
    for c in range(current_col + 1, ws.max_column + 1):
        txt = _cell_text(ws.cell(SALES_CODE_ROW, c))
        tokens = [t for t in extract_code_tokens(txt) if t in universe]
        if not tokens:
            continue
        feature = _cell_text(ws.cell(FEATURE_ROW, c))
        if len(tokens) == 1 or _separable_or_list(txt, tokens):
            combined = len(tokens) > 1
            for token in tokens:
                if token in code_to_col:
                    continue
                code_to_col[token] = c
                sales_codes.append(SalesCodeColumn(
                    code=token, feature=feature, original_expr=txt, from_combined=combined,
                    source_col=get_column_letter(c),
                    klass=ProposalClass.UNCERTAIN if combined else ProposalClass.CONFIRMED,
                ))
        else:
            ce = CombinedExpr(
                original_expr=txt, source_col=get_column_letter(c), tokens=tokens,
                feature=feature)
            if ce.is_equality:
                ce.include = True   # equivalent codes — unambiguous, pre-approved
            combined_exprs.append(ce)
            combined_cols.append((ce, c))

    # Variant partitioning: marker columns on row 9 (LEFT/RIGHT, DRIVER/PASSENGER,
    # CUP vs CM5/CVM). Each part row is X/G-marked under exactly one -> its side.
    part_cols = partition_columns(ws, SALES_CODE_ROW)

    rows: list[MatrixRow] = []
    seq = phase_cols + [current_col]
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        current_text = _cell_text(ws.cell(r, current_col))
        if not current_text:
            continue  # a variant row must have a Current value
        variant_id = _cell_text(ws.cell(r, 1))
        row = _build_row(ws, r, variant_id, seq, current_col, sales_codes, code_to_col,
                         combined_cols, worksheet)
        if part_cols:
            row.partition_side = row_side(ws, r, part_cols)
        rows.append(row)
        if len(rows) >= 2000:  # safety cap
            break

    # only keep sides that actually have a non-excluded part; a real partition needs
    # at least two variants (a lone marker column is not a split).
    present = {x.partition_side for x in rows if x.partition_side and not x.excluded}
    partition_sides = [s for s in ordered_sides(part_cols) if s in present]
    if len(partition_sides) < 2:
        partition_sides = []
        for x in rows:
            x.partition_side = ""

    meta = _extract_header_meta(ws)
    logger.info("Family matrix '%s': %d rows, %d sales codes, %d combined, sides=%s.",
                worksheet, len(rows), len(sales_codes), len(combined_exprs), partition_sides)
    return FamilyMatrix(
        worksheet=worksheet, canonical_family=canonical_family or worksheet,
        sales_codes=sales_codes, rows=rows, combined_exprs=combined_exprs,
        dtx_codes=sorted(family_dtx_codes or set()), partition_sides=partition_sides,
        year=meta.get("year", ""), vehicle=meta.get("vehicle", ""),
        phase=meta.get("phase", ""), harness_name=meta.get("harness", "") or worksheet,
    )


def _extract_header_meta(ws) -> dict[str, str]:
    """Parse the family-sheet header (rows 1-5): Vehicle Program (year+vehicle),
    Build Phase, and Harness name — for the individual file's info table."""
    meta: dict[str, str] = {}
    for r in range(1, 6):
        for c in range(1, 4):
            text = _cell_text(ws.cell(r, c))
            low = text.lower()
            if "vehicle program" in low and ":" in text:
                parts = text.split(":", 1)[1].split()
                if parts and parts[0].isdigit():
                    meta["year"] = parts[0]
                if len(parts) > 1:
                    meta["vehicle"] = parts[1]
            elif "build phase" in low and ":" in text:
                meta["phase"] = text.split(":", 1)[1].strip()
            elif low.startswith("harness:"):
                meta["harness"] = text.split(":", 1)[1].strip()
    return meta


def _build_row(ws, r, variant_id, seq, current_col, sales_codes, code_to_col,
               combined_cols, worksheet) -> MatrixRow:
    # Deletion is signalled by a 'DELETE P/N' / 'Cancel' sentinel — a cell that STARTS
    # WITH one of those — either the Current cell or the note columns beside it.
    # (Bulletin prose mentions 'delete' mid-text; that must NOT exclude.)
    def _is_delete_marker(text: str) -> bool:
        u = text.strip().upper()
        return u.startswith("DELETE P") or u.startswith("CANCEL")

    note_removed = any(
        _is_delete_marker(_cell_text(ws.cell(r, c)))
        for c in range(current_col + 1, min(current_col + 4, ws.max_column + 1))
    )

    def _valid(val: str) -> bool:
        u = val.upper()
        return bool(val) and u not in (_CARRYOVER, _NA) and not _is_delete_marker(val)

    last_valid = ""
    prev_before_current = ""
    for c in seq:
        if c == current_col:
            prev_before_current = last_valid
        val = _cell_text(ws.cell(r, c))
        if _valid(val):
            last_valid = val

    current_text = _cell_text(ws.cell(r, current_col))
    source = f"{worksheet}!{get_column_letter(current_col)}{r}"
    excluded = False
    if _is_delete_marker(current_text) or note_removed or current_text.upper() == _NA:
        current_pn, klass, reason, excluded = \
            "", ProposalClass.EXCLUDED, "deleted / cancelled / N/A", True
    elif current_text.upper() == _CARRYOVER:
        current_pn, klass, reason = \
            prev_before_current, ProposalClass.INFERRED, "carryover (C/O) → most recent valid P/N"
    elif current_text:
        current_pn, klass, reason = current_text, ProposalClass.CONFIRMED, "Current column value"
    else:
        current_pn, klass, reason = \
            prev_before_current, ProposalClass.INFERRED, "no Current value → most recent valid P/N"

    symbols: dict[str, str] = {}
    symbol_class: dict[str, ProposalClass] = {}
    for sc in sales_codes:
        sym = _cell_text(ws.cell(r, code_to_col[sc.code])).upper()
        if sym in ("X", "G"):
            symbols[sc.code] = sym
            symbol_class[sc.code] = \
                ProposalClass.UNCERTAIN if sc.from_combined else ProposalClass.CONFIRMED
    combined_symbols: dict[str, str] = {}
    for ce, c in combined_cols:
        sym = _cell_text(ws.cell(r, c)).upper()
        if sym in ("X", "G"):
            combined_symbols[ce.key] = sym
    return MatrixRow(
        variant_id=variant_id, current_pn=current_pn, previous_pn=prev_before_current,
        current_class=klass, current_reason=reason, current_source=source,
        excluded=excluded, symbols=symbols, symbol_class=symbol_class,
        combined_symbols=combined_symbols,
    )


def dtx_sales_code_universe(frames: list[pd.DataFrame]) -> set[str]:
    """The set of 3-char sales codes present in the DTx ``Sales Code`` column(s)."""
    universe: set[str] = set()
    for df in frames:
        if df is None or "Sales Code" not in df.columns:
            continue
        for value in df["Sales Code"].dropna().unique():
            universe.update(extract_code_tokens(value))
    return universe


def family_dtx_sales_codes(
    frames: list[pd.DataFrame],
    crossref: CrossRef | None,
    worksheet: str,
) -> set[str]:
    """Sales codes used in the DTx for one harness family (its master worksheet).

    Resolves which DTx ``Harness Family`` names map to ``worksheet`` (via the
    cross-reference, plus the worksheet name itself as a fallback), then collects
    the 3-char tokens from those rows' ``Sales Code`` values.
    """
    families = {worksheet}
    if crossref is not None:
        families |= {d for d, w in crossref.dtx_to_worksheet.items() if w == worksheet}
    codes: set[str] = set()
    for df in frames:
        if df is None or "Harness Family" not in df.columns or "Sales Code" not in df.columns:
            continue
        sub = df[df["Harness Family"].astype(str).isin(families)]
        for value in sub["Sales Code"].dropna():
            codes.update(extract_code_tokens(value))
    return codes


def read_dtx_frames(files: list[tuple[str, bytes]]) -> list[pd.DataFrame]:
    """Read DTx export workbook(s)/CSVs into dataframes (all sheets, as strings)."""
    frames: list[pd.DataFrame] = []
    for name, data in files:
        try:
            if name.lower().endswith(".csv"):
                frames.append(pd.read_csv(io.BytesIO(data), dtype=str))
            else:
                book = pd.read_excel(io.BytesIO(data), dtype=str, sheet_name=None)
                frames.extend(book.values())
        except Exception as exc:
            raise SpliceError(f"Could not read DTx file '{name}': {exc}") from exc
    return frames
