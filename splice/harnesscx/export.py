"""Generate an individual harness-complexity .xlsm from the bundled template.

Writes only the ``Complexity`` matrix into a **copy** of the template
(``assets/Harness_Complexity.xlsm``), preserving its VBA macros, the
``Harness PN`` sheet formulas (which reference ``Complexity!``), the
``Information`` sheet, and all styling. Sources are never modified.

Layout (from the filled examples): A1 = ``ID=<harness id>``; row 1 columns 2..
= the sales codes; column A rows 2.. = the (non-excluded) part numbers; the
``X``/``G`` symbols sit at the part-number × sales-code intersections.
"""

from __future__ import annotations

import copy
import io
import logging
from datetime import date
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

from splice.common.errors import SpliceError
from splice.harnesscx.adapters import is_non_part_pn
from splice.harnesscx.models import FamilyMatrix

logger = logging.getLogger(__name__)

TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "assets" / "Harness_Complexity.xlsm"


def template_bytes() -> bytes:
    """The bundled individual-complexity template."""
    if not TEMPLATE_PATH.exists():
        raise SpliceError(f"Individual-complexity template missing: {TEMPLATE_PATH}")
    return TEMPLATE_PATH.read_bytes()


def validate_before_export(matrix: FamilyMatrix, harness_id: str) -> list[str]:
    """Return blocking problems (empty = ok to export). Warnings are separate."""
    problems: list[str] = []
    if not str(harness_id).strip():
        problems.append("Harness ID is required (enter it manually).")
    if not any(not r.excluded and r.current_pn for r in matrix.rows):
        problems.append("No part numbers to write (every row is excluded or empty).")
    if not matrix.sales_codes:
        problems.append("No sales-code columns were identified.")
    return problems


def unresolved_warnings(matrix: FamilyMatrix) -> list[str]:
    """Non-blocking items the SE should confirm before relying on the export."""
    warnings: list[str] = []
    n_uncertain = matrix.unresolved_count
    if n_uncertain:
        warnings.append(f"{n_uncertain} row(s) still have Uncertain values "
                        "(e.g. split combined codes).")
    n_combined = sum(1 for s in matrix.sales_codes if s.from_combined)
    if n_combined:
        warnings.append(f"{n_combined} sales code(s) came from combined expressions "
                        "— verify applicability.")
    return warnings


def generate_files(
    matrix: FamilyMatrix,
    harness_id: str,
    template: bytes | None = None,
) -> tuple[list[tuple[bytes, str]], list[str]]:
    """Generate the individual complexity file(s) for a family matrix.

    Returns ``(files, blocking_problems)`` where ``files`` is a list of
    ``(bytes, filename)``. A worksheet the master partitions (LEFT/RIGHT,
    DRIVER/PASSENGER, CUP vs CM5/CVM) yields ONE file per variant — each with
    only that variant's part numbers plus the common (unmarked) ones.
    """
    matrix.harness_id = harness_id
    problems = validate_before_export(matrix, harness_id)
    if problems:
        return [], problems
    template = template or template_bytes()

    if not matrix.partition_sides:
        return [build_individual_complexity(matrix, template, harness_id=harness_id)], []

    files: list[tuple[bytes, str]] = []
    for side in matrix.partition_sides:
        sub = copy.deepcopy(matrix)
        # keep this variant's part numbers (+ any common part with no side marker)
        sub.rows = [r for r in matrix.rows if r.partition_side == side or not r.partition_side]
        sub.partition_sides = []
        sub.worksheet = f"{matrix.worksheet} {side.replace('/', '')}"
        files.append(build_individual_complexity(sub, template, harness_id=harness_id))
    return files, []


def build_individual_complexity(
    matrix: FamilyMatrix,
    template: bytes,
    *,
    harness_id: str = "",
) -> tuple[bytes, str]:
    """Return (xlsm bytes, filename) for the individual harness-complexity file."""
    if not template:
        raise SpliceError("Individual harness-complexity template was not provided.")
    try:
        wb = load_workbook(io.BytesIO(template), keep_vba=True)  # preserve macros
    except Exception as exc:
        raise SpliceError(f"Could not read the individual-complexity template: {exc}") from exc
    if "Complexity" not in wb.sheetnames:
        raise SpliceError("Template has no 'Complexity' sheet.")

    ws = wb["Complexity"]
    # The 'Harness PN' sheet holds the manually-entered columns the template's
    # formulas build on: Previous P/N, New P/N, and Symbol (master column A).
    hp_ws = wb["Harness PN"] if "Harness PN" in wb.sheetnames else None

    def _hp_col(name: str) -> int | None:
        if hp_ws is None:
            return None
        for c in range(1, hp_ws.max_column + 1):
            if str(hp_ws.cell(1, c).value or "").strip().lower() == name.lower():
                return c
        return None

    prev_col, new_col, symbol_col = _hp_col("Previous P/N"), _hp_col("New P/N"), _hp_col("Symbol")

    ws.cell(1, 1, f"ID={str(harness_id).strip()}")
    # Output columns: the separated sales codes, then any SE-approved combined
    # expressions (each a header + a getter for that column's per-row mark).
    out_cols: list[tuple[str, Callable]] = [
        (sc.code, (lambda row, code=sc.code: row.symbols.get(code, "")))
        for sc in matrix.sales_codes
    ]
    for ce in matrix.combined_exprs:
        if ce.include:
            # An equality ('XH3=XH4') or a comma-separated definition writes one
            # column per code, each sharing this expression's per-row content.
            get = (lambda row, key=ce.key: row.combined_symbols.get(key, ""))
            for code in ce.output_codes:
                if code:
                    out_cols.append((code, get))
    for col, (header, _get) in enumerate(out_cols, start=2):
        ws.cell(1, col, header)

    r = 2
    written = 0
    for row in matrix.rows:
        if row.excluded or is_non_part_pn(row.current_pn):
            continue  # excluded / NO HARNESS / DELETE / cancel / N/A must not appear
        ws.cell(r, 1, row.current_pn)
        if hp_ws is not None:
            if new_col:
                hp_ws.cell(r, new_col, row.current_pn)
            if prev_col and row.previous_pn:
                hp_ws.cell(r, prev_col, row.previous_pn)
            if symbol_col and str(row.variant_id).strip() not in ("", "(added)"):
                hp_ws.cell(r, symbol_col, row.variant_id)
        for col, (_header, get) in enumerate(out_cols, start=2):
            mark = get(row)
            if mark:
                ws.cell(r, col, mark)
        r += 1
        written += 1

    _apply_matrix_borders(ws, last_row=written + 1, last_col=1 + len(out_cols))

    if hp_ws is not None:
        _write_info_table(hp_ws, matrix, harness_id)
        # The template only carries its per-row formulas down a fixed number of
        # rows; extend them so every part number written above is covered.
        _extend_row_formulas(hp_ws, last_row=written + 1)

    # Force Excel to recompute every formula when the file opens — openpyxl keeps
    # the template's *cached* formula results otherwise, which show as stale cells
    # until the user manually recalculates.
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:  # noqa: BLE001 — never fail generation over a calc hint
        pass

    buf = io.BytesIO()
    wb.save(buf)
    filename = _build_filename(matrix)
    logger.info("Generated '%s': %d parts x %d cols.", filename, written, len(out_cols))
    return buf.getvalue(), filename


_THIN = Side(style="thin", color="FF000000")
_MATRIX_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _apply_matrix_borders(ws, *, last_row: int, last_col: int) -> None:
    """Draw thin borders on every cell of the populated matrix rectangle."""
    if last_row < 1 or last_col < 1:
        return
    for r in range(1, last_row + 1):
        for c in range(1, last_col + 1):
            ws.cell(r, c).border = _MATRIX_BORDER


def _build_filename(matrix: FamilyMatrix) -> str:
    """``2.- Harness_Complexity_28RU_X2_IP_MM-DD-YYYY.xlsm`` (date = generated today)."""
    yy = (matrix.year or "").strip()[-2:]
    program = f"{yy}{matrix.vehicle}".strip()
    harness = (matrix.worksheet or matrix.harness_name or matrix.canonical_family or "Harness")
    tokens = [t.replace(" ", "").replace("/", "-")
              for t in (program, matrix.phase, harness) if t]
    stem = "_".join(tokens)
    return f"2.- Harness_Complexity_{stem}_{date.today().strftime('%m-%d-%Y')}.xlsm"


def _extend_row_formulas(hp_ws, *, last_row: int) -> None:
    """Fill each per-row formula column of the 'Harness PN' sheet down to ``last_row``.

    A per-row formula column is one whose row 2 AND row 3 hold formulas; single-cell
    formulas like ``=COUNTA(H:H)`` are left untouched. Existing formula cells are
    preserved; only empty rows below the template's range are filled, with relative
    references translated. Columns intersecting a merged range are skipped."""
    merged_cols = {
        c for mr in hp_ws.merged_cells.ranges for c in range(mr.min_col, mr.max_col + 1)
    }
    for c in range(1, hp_ws.max_column + 1):
        if c in merged_cols:
            continue
        f2, f3 = hp_ws.cell(2, c).value, hp_ws.cell(3, c).value
        if not (isinstance(f2, str) and f2.startswith("=")
                and isinstance(f3, str) and f3.startswith("=")):
            continue
        letter = get_column_letter(c)
        for r in range(3, last_row + 1):
            existing = hp_ws.cell(r, c).value
            if isinstance(existing, str) and existing.startswith("="):
                continue  # keep the template's own formula rows
            hp_ws.cell(r, c, Translator(f2, origin=f"{letter}2").translate_formula(f"{letter}{r}"))


def _write_info_table(hp_ws, matrix: FamilyMatrix, harness_id: str) -> None:
    """Fill the Year / Vehicle / Phase / Harness / ID values beside their labels
    (the labels already exist in the template's info table)."""
    values = {
        "year:": matrix.year,
        "vehicle:": matrix.vehicle,
        "phase:": matrix.phase,
        "harness:": matrix.harness_name or matrix.canonical_family or matrix.worksheet,
        "id:": str(harness_id).strip(),
    }
    for r in range(1, 41):
        for c in range(1, hp_ws.max_column + 1):
            label = str(hp_ws.cell(r, c).value or "").strip().lower()
            if label in values and values[label]:
                hp_ws.cell(r, c + 1, values[label])
                break
