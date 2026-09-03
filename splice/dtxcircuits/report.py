"""Export the applicability review, carrying the SE's cleanup selections.

The workbook is the hand-off: it repeats what the workbench showed, and adds
a **Complexity Cleanup Notes** column. A row the Systems Engineer ticked in
the workbench carries a written note saying what has to be fixed in the
complexity file; everything else leaves that column empty, so the column is a
work list rather than a wall of commentary.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from splice.dtxcircuits.models import (
    ALL_BUILDS,
    NEVER,
    NO_COMPLEXITY,
    UNCONDITIONAL,
    VARIANT,
    HarnessAnalysis,
)
from splice.inline import salescode

#: the column (and sheet) the cleanup notes land in — named by the request
CLEANUP_COLUMN = "Complexity Cleanup Notes"

KIND_CIRCUIT, KIND_CONNECTOR, KIND_GAP = "circuit", "connector", "gap"
KIND_EXPRESSION = "expression"

#: Where the correction belongs. The SE's first question about any finding
#: is which document is wrong, so it is a column rather than a sentence.
#: "Either" is not a shrug — it is the honest answer when the data cannot
#: settle it, and the action then names the question to put to the customer.
FIX_DTX = "DTx"
FIX_COMPLEXITY = "Complexity matrix"
FIX_EITHER = "DTx or complexity"
FIX_CONFIRM = "Confirm only"
FIX_MAPPING = "Your mapping"
FIX_ORDER = {FIX_DTX: 0, FIX_COMPLEXITY: 1, FIX_EITHER: 2, FIX_CONFIRM: 3,
             FIX_MAPPING: 4}

P_BLOCKER, P_HIGH, P_REVIEW = "Blocker", "High", "Review"
PRIORITY_ORDER = {P_BLOCKER: 0, P_HIGH: 1, P_REVIEW: 2}

#: "gap" is what the engine calls it; this is what the customer reads
KIND_LABEL = {KIND_CIRCUIT: "circuit", KIND_CONNECTOR: "connector",
              KIND_GAP: "sales code", KIND_EXPRESSION: "expression"}


def kind_label(kind: str) -> str:
    return KIND_LABEL.get(kind, kind)


def item_key(family: str, harness: str, kind: str, ident: str) -> str:
    """A stable handle for one selectable row, across refreshes and reruns."""
    return "|".join((family, harness, kind, ident))


@dataclass
class Entry:
    """One analysed pairing: a DTx family resolved against one harness file.

    ``original_*`` hold the conditions as the DTx actually stated them, before
    any sales-code repair was applied. The analysis runs on repaired rows, so
    without these the export would only ever show the corrected wording and
    there would be no record of what was changed or why a verdict moved.
    """

    label: str
    family: str
    filename: str
    analysis: HarnessAnalysis
    original_circuit_conditions: Dict[str, str] = field(default_factory=dict)
    original_cnum_conditions: Dict[str, str] = field(default_factory=dict)
    #: the complexity file itself. The analysis is a summary of it; the chart
    #: needs the builds and their codes to re-express a condition in the terms
    #: this harness actually tracks.
    complexity: object = None


@dataclass
class CleanupSelection:
    """What the SE ticked, and the note that will be exported for it."""

    key: str
    family: str
    harness: str
    kind: str
    ident: str
    verdict: str = ""
    condition: str = ""
    note: str = ""
    #: the work-list fields: what to do, where, and how urgent
    priority: str = ""
    fix_in: str = ""
    action: str = ""
    def_id: str = ""
    builds: str = ""
    evidence: str = ""


def gap_codes(analysis: HarnessAnalysis) -> set:
    """Codes the DTx conditions on that this harness's complexity never lists."""
    return {gap.code for gap in analysis.code_gaps}


def inert_codes(expression: str, gaps: set) -> List[str]:
    """The codes in ``expression`` this harness's complexity cannot see.

    A code the complexity does not track is treated as present, so it can
    never narrow anything: leaving it in the condition only makes the item
    look conditional when it is not.
    """
    if not expression or not gaps:
        return []
    return sorted({code for code in salescode.codes_in(expression)
                   if code in gaps})


def removal_hint(expression: str, gaps: set) -> str:
    """The one fix, phrased the same way everywhere it appears."""
    hits = inert_codes(expression, gaps)
    if not hits:
        return ""
    codes = ", ".join(hits)
    return f" Remove {codes}, or add it to the complexity file."


def _inert_note(condition: str, hits: List[str]) -> str:
    codes = ", ".join(hits)
    return (f"{condition} reads as always true — {codes} not in the complexity "
            f"file. Remove {codes}, or add it.")


def circuit_note(analysis: HarnessAnalysis, circuit) -> str:
    """Why this circuit needs attention. One line, no restated columns."""
    condition = circuit.expression or "(no sales code)"
    gaps = gap_codes(analysis)
    hits = inert_codes(circuit.expression, gaps)
    if circuit.classification == NEVER:
        return (f"Never built — no build satisfies {condition}."
                + removal_hint(circuit.expression, gaps))
    if hits or circuit.untracked_codes:
        return _inert_note(condition,
                           hits or sorted(circuit.untracked_codes))
    if circuit.classification == NO_COMPLEXITY:
        return f"No complexity file mapped, so {condition} was not resolved."
    if circuit.classification == VARIANT:
        return (f"{len(circuit.builds_with)} of {circuit.build_count} builds "
                f"under {condition} — confirm the split is intended.")
    if circuit.classification in (ALL_BUILDS, UNCONDITIONAL):
        return (f"Every build carries it under {condition} — confirm the "
                f"condition is still needed.")
    return f"Review under {condition}."


def connector_note(analysis: HarnessAnalysis, cnum) -> str:
    condition = cnum.expression or "(no sales code)"
    gaps = gap_codes(analysis)
    hits = inert_codes(cnum.expression, gaps)
    if cnum.classification == NEVER:
        return (f"Never populated — no build satisfies {condition}."
                + removal_hint(cnum.expression, gaps))
    if hits:
        return _inert_note(condition, hits)
    if cnum.classification in (ALL_BUILDS, UNCONDITIONAL):
        return (f"Every build populates it under {condition} — confirm the "
                f"condition is still needed.")
    return f"{len(cnum.circuits)} circuit(s) under {condition}."


def gap_note(analysis: HarnessAnalysis, gap) -> str:
    return (f"Sales code {gap.code} is in the DTx report but not in the "
            f"complexity file.")



def _codes(item, gaps: set) -> List[str]:
    """The codes in this item's condition the complexity cannot see."""
    return inert_codes(getattr(item, "expression", ""), gaps) \
        or sorted(getattr(item, "untracked_codes", []) or [])


def _code_fix(codes: str, harness: str, tracked_elsewhere: set,
              proven: Optional[bool] = None) -> tuple:
    """``(fix_in, action)`` for a code the complexity does not track.

    Phrased identically wherever the same code turns up, so the email can
    collapse a circuit, its connector and the gap itself into one request.
    """
    if proven is None:
        proven = all(code.strip() in tracked_elsewhere
                     for code in codes.split(",") if code.strip())
    where = f"the {harness} complexity file"
    if proven:
        return (FIX_COMPLEXITY,
                f"Add {codes} to {where} — another harness already tracks it")
    return (FIX_EITHER,
            f"Confirm whether {harness} varies by {codes}, then add it to "
            f"{where} or remove it from the DTx condition")


def prescribe(analysis: HarnessAnalysis, kind: str, item,
              tracked_elsewhere: Optional[set] = None) -> tuple:
    """What to do about one finding: ``(priority, fix_in, action)``.

    The DTx and the complexity matrix are two descriptions of the same
    vehicle, and a finding is a place they disagree. Which one to change is
    decided from the data wherever the data can decide it:

    * a condition no build can satisfy **whose codes the complexity does
      track** is a contradiction the DTx wrote — the DTx is wrong;
    * a code this complexity does not track **but another loaded complexity
      does** is real, so this file is the one missing it;
    * a code no loaded complexity tracks could be a DTx typo or a matrix-wide
      omission, and nothing in the data separates those — so the action asks
      the customer the question instead of pretending to know.
    """
    tracked_elsewhere = tracked_elsewhere or set()
    gaps = gap_codes(analysis)
    harness = analysis.harness

    if kind == KIND_EXPRESSION:
        return (P_BLOCKER, FIX_DTX,
                f'Correct "{item}" — it is not a valid sales-code expression')

    if kind == KIND_GAP:
        return (P_HIGH,) + _code_fix(item.code, harness, tracked_elsewhere)

    condition = getattr(item, "expression", "") or "(no sales code)"
    hits = _codes(item, gaps)
    codes = ", ".join(hits)
    verb = "carries" if kind == KIND_CIRCUIT else "populates"

    # a code a sibling complexity file already tracks is real, so the same
    # code gets the same verdict wherever it turns up — not "either" here and
    # "complexity" one row above
    proven = bool(hits) and all(code in tracked_elsewhere for code in hits)

    if item.classification == NEVER:
        if hits:
            return (P_BLOCKER,) + _code_fix(codes, harness, tracked_elsewhere,
                                            proven=proven)
        return (P_BLOCKER, FIX_DTX,
                f"Correct {condition} in the DTx, or move this "
                f"{kind_label(kind)} off {harness} — no build satisfies it")
    if hits:
        return (P_HIGH,) + _code_fix(codes, harness, tracked_elsewhere,
                                     proven=proven)
    if item.classification == NO_COMPLEXITY:
        return (P_HIGH, FIX_MAPPING,
                f"Map a complexity file for {analysis.harness} — "
                f"{condition} could not be resolved")
    if item.classification in (ALL_BUILDS, UNCONDITIONAL):
        return (P_REVIEW, FIX_CONFIRM,
                f"Every build {verb} it under {condition} — confirm the "
                f"condition is still needed")
    if item.classification == VARIANT:
        return (P_REVIEW, FIX_CONFIRM,
                f"{len(item.builds_with)} of {item.build_count} builds "
                f"{verb} it under {condition} — confirm the split is intended")
    return (P_REVIEW, FIX_CONFIRM, f"Review under {condition}")


def _evidence(item) -> tuple:
    """``(builds, part numbers)`` — the proof an SE forwards to the customer."""
    with_it = list(getattr(item, "builds_with", []) or [])
    without = list(getattr(item, "builds_without", []) or [])
    total = len(with_it) + len(without)
    builds = f"{len(with_it)} of {total}" if total else "—"
    shown = without[:6] if not with_it else with_it[:6]
    label = "without it: " if not with_it else "with it: "
    more = f" (+{len(without if not with_it else with_it) - 6})" \
        if len(without if not with_it else with_it) > 6 else ""
    return builds, (label + ", ".join(shown) + more if shown else "—")


def selection_for(entry: Entry, kind: str, ident: str,
                  tracked_elsewhere: Optional[set] = None) -> Optional[CleanupSelection]:
    """Build the selection record (and its note) for one row of one entry."""
    analysis = entry.analysis
    if kind == KIND_CIRCUIT:
        item = next((c for c in analysis.circuits if c.circuit == ident), None)
        if item is None:
            return None
        priority, fix_in, action = prescribe(analysis, kind, item,
                                            tracked_elsewhere)
        builds, evidence = _evidence(item)
        return CleanupSelection(
            key=item_key(entry.family, analysis.harness, kind, ident),
            family=entry.family, harness=analysis.harness, kind=kind,
            ident=ident, verdict=item.classification,
            condition=item.expression or "", note=circuit_note(analysis, item),
            priority=priority, fix_in=fix_in, action=action,
            def_id=analysis.def_id, builds=builds, evidence=evidence)
    if kind == KIND_CONNECTOR:
        item = next((c for c in analysis.cnums if c.cnum == ident), None)
        if item is None:
            return None
        priority, fix_in, action = prescribe(analysis, kind, item,
                                            tracked_elsewhere)
        builds, evidence = _evidence(item)
        return CleanupSelection(
            key=item_key(entry.family, analysis.harness, kind, ident),
            family=entry.family, harness=analysis.harness, kind=kind,
            ident=ident, verdict=item.classification,
            condition=item.expression or "", note=connector_note(analysis, item),
            priority=priority, fix_in=fix_in, action=action,
            def_id=analysis.def_id, builds=builds, evidence=evidence)
    if kind == KIND_GAP:
        item = next((g for g in analysis.code_gaps if g.code == ident), None)
        if item is None:
            return None
        priority, fix_in, action = prescribe(analysis, kind, item,
                                            tracked_elsewhere)
        return CleanupSelection(
            key=item_key(entry.family, analysis.harness, kind, ident),
            family=entry.family, harness=analysis.harness, kind=kind,
            ident=ident, verdict="sales-code gap", condition="",
            note=gap_note(analysis, item), priority=priority, fix_in=fix_in,
            action=action, def_id=analysis.def_id,
            builds=f"{item.occurrences} DTx row(s)",
            evidence=("circuits: " + ", ".join(item.circuits[:6])
                      if item.circuits else "—"))
    return None


def auto_select(entries: Iterable[Entry],
                dismissed: Optional[set] = None) -> Dict[str, CleanupSelection]:
    """The findings worth putting in front of the customer, pre-ticked.

    Never-built circuits and connectors, and every sales-code gap: each is a
    place the DTx cannot be reconciled with complexity built from the
    customer's own information.

    Also taken: anything built on *every* build whose condition contains one
    of those gap codes. An untracked code is read as present, so it cannot
    narrow anything — the condition only looks like a condition. That is the
    other half of the same defect, and it carries a concrete fix (drop the
    code), so it belongs in the cleanup list rather than being left for the
    SE to find by hand.

    Anything the SE has explicitly unticked is left alone — a proposal that
    reappears every run is not a proposal.
    """
    dismissed = dismissed or set()
    picked: Dict[str, CleanupSelection] = {}

    def take(entry: Entry, kind: str, ident: str) -> None:
        key = item_key(entry.family, entry.analysis.harness, kind, ident)
        if key in dismissed or key in picked:
            return
        selection = selection_for(entry, kind, ident)
        if selection:
            picked[key] = selection

    for entry in entries:
        analysis = entry.analysis
        gaps = gap_codes(analysis)
        for circuit in analysis.circuits:
            if circuit.classification == NEVER or (
                    circuit.classification in (ALL_BUILDS, UNCONDITIONAL)
                    and removal_hint(circuit.expression, gaps)):
                take(entry, KIND_CIRCUIT, circuit.circuit)
        for connector in analysis.cnums:
            if connector.classification == NEVER or (
                    connector.classification in (ALL_BUILDS, UNCONDITIONAL)
                    and removal_hint(connector.expression, gaps)):
                take(entry, KIND_CONNECTOR, connector.cnum)
        for gap in analysis.code_gaps:
            take(entry, KIND_GAP, gap.code)
    return picked


# --------------------------------------------------------------------------
# workbook
# --------------------------------------------------------------------------

_HEAD_FILL = PatternFill("solid", fgColor="1F3B57")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_NOTE_FILL = PatternFill("solid", fgColor="FFF3CD")
_MUTED_INK = "6B7280"


def _sheet(wb: Workbook, title: str, headers: List[str], rows: List[list],
           first: bool = False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title[:31]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill, cell.font = _HEAD_FILL, _HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    note_col = headers.index(CLEANUP_COLUMN) + 1 if CLEANUP_COLUMN in headers else None
    # The row index is tracked, not read back: ws.max_row rescans every cell
    # written so far, which turns a long sheet quadratic.
    line_no = 1
    for row in rows:
        ws.append(row)
        line_no += 1
        if note_col and ws.cell(line_no, note_col).value:
            ws.cell(line_no, note_col).fill = _NOTE_FILL
    widths = [12] * len(headers)
    for index, header in enumerate(headers):
        longest = max([len(str(header))] + [len(str(r[index])) for r in rows[:400]]
                      or [10])
        widths[index] = min(max(12, longest + 2), 70)
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions
    return ws


def _quality_sheets(wb: Workbook, q) -> None:
    """Two sheets written for the customer: what the export got right, and
    where each sales code is and is not known."""
    _sheet(wb, "Data quality", ["Measure", "Value", "What it means"], [
        ["Programme", q.program or "?", "Read from the DTx title block"],
        ["Phase", q.phase or "?", ""],
        ["Report date", q.report_date or "?", ""],
        ["", "", ""],
        ["Rows", q.rows, "Circuit/connector rows read"],
        ["Circuits", q.circuits, ""],
        ["Connectors", q.connectors, ""],
        ["Harness families", q.families, ""],
        ["Conditioned rows", f"{q.conditioned_rows} ({q.conditioned_share:.0%})",
         "Rows carrying a sales-code condition"],
        ["Unconditional rows", q.unconditional_rows, "Built on every "
         "configuration"],
        ["Distinct expressions", q.distinct_expressions, ""],
        ["Distinct sales codes", q.distinct_codes, ""],
        ["", "", ""],
        ["FINDINGS", "", "Each one is something to correct in the next export"],
        ["Malformed expressions", q.malformed_expressions,
         "No boolean operator between codes — the expression is false for "
         "every configuration, so its circuits read as never built"],
        ["  rows affected", q.malformed_rows, ""],
        ["  repaired in review", q.repaired_expressions,
         "The Systems Engineer supplied the intended expression; see the "
         "Sales-code repairs sheet"],
        ["Never-built circuits", q.never_built_circuits,
         "No configuration of the mapped harness satisfies the condition"],
        ["Never-built connectors", q.never_built_connectors, ""],
        ["Codes tracked nowhere", len(q.codes_not_tracked_anywhere),
         "Used by the DTx, absent from every complexity file it was checked "
         "against"],
        ["Codes partly tracked", len(q.codes_partially_tracked),
         "Known to some harnesses and not others"],
        ["", "", ""],
        ["Families assessed", q.families_mapped, ""],
        ["Families not assessed", len(q.families_unmapped),
         ", ".join(q.families_unmapped)],
    ] + [[k, v, ""] for k, v in sorted(q.malformed_by_kind.items())])

    _sheet(wb, "Sales-code coverage",
           ["Sales code", "Status", "DTx rows", "Used by families",
            "Tracked by", "Missing from", "Circuits"],
           [[x.code, x.status, x.dtx_rows, ", ".join(x.families),
             ", ".join(x.tracked_by) or "—", ", ".join(x.missing_from) or "—",
             ", ".join(x.circuits[:20])] for x in q.coverage])



# --------------------------------------------------------------------------
# the cleanup work list, and the email that goes with it
# --------------------------------------------------------------------------
CLEANUP_HEADERS = [
    "#", "Priority", "Fix in", "Action for the customer", "Why it is wrong",
    "Harness family", "Harness (DEF id)", "Type", "Item",
    "Condition (as in DTx)", "Builds carrying it", "Evidence — part numbers",
    "Status (yours)",
]
_PRIORITY_FILL = {
    P_BLOCKER: PatternFill("solid", fgColor="F8D7DA"),
    P_HIGH: PatternFill("solid", fgColor="FCE4C6"),
    P_REVIEW: PatternFill("solid", fgColor="FFF3CD"),
}
_PRIORITY_FONT = {P_BLOCKER: Font(bold=True, color="9C1C1C"),
                  P_HIGH: Font(bold=True, color="8A4B00"),
                  P_REVIEW: Font(bold=True, color="7A5A00")}
_BAND_FILL = PatternFill("solid", fgColor="DDEBF7")


def tracked_codes(entries: Iterable[Entry]) -> set:
    """Every sales code any loaded complexity file tracks.

    This is what separates "the DTx invented a code" from "this one file is
    missing a code its neighbours have".
    """
    out: set = set()
    for entry in entries:
        codes = getattr(entry.complexity, "complexity_codes", None) or []
        out.update(str(code).strip().upper() for code in codes)
    return out


def _repair_rows(repairs: dict, context: dict) -> List[CleanupSelection]:
    """A repair the SE confirmed is a DTx defect the customer must mirror.

    The workbench applied it locally so the analysis could run; unless the
    customer fixes the source, the next export brings it back.
    """
    out = []
    for original, decided in sorted((repairs or {}).items()):
        info = (context or {}).get(original, {})
        families = ", ".join(info.get("families", []) or [])
        circuits = list(info.get("circuits", []) or [])
        priority, fix_in, _ = prescribe(HarnessAnalysis(harness=""),
                                        KIND_EXPRESSION, original)
        out.append(CleanupSelection(
            key=f"expression|{original}", family=families or "—",
            harness="—", kind=KIND_EXPRESSION, ident=original,
            verdict=info.get("kind", "malformed expression"), condition=original,
            note=(f"{info.get('kind', 'Malformed')} — the expression is false "
                  f"for every configuration, so its circuits read as never "
                  f"built. Reviewed as \u201c{decided}\u201d."),
            priority=priority, fix_in=fix_in,
            action=f'Correct "{original}" to "{decided}" in the DTx',
            builds=f"{info.get('rows', '')} DTx row(s)".strip(),
            evidence=("circuits: " + ", ".join(circuits[:6]) if circuits else "—")))
    return out


def cleanup_rows(cleanup: Dict[str, CleanupSelection], entries: List[Entry],
                 repairs: Optional[dict] = None,
                 context: Optional[dict] = None) -> List[CleanupSelection]:
    """The work list: every ticked finding plus every confirmed repair,
    re-prescribed against this run and ordered so the top row is the first
    thing to raise with the customer."""
    tracked = tracked_codes(entries)
    by_pair = {(e.family, e.analysis.harness): e for e in entries}
    out: List[CleanupSelection] = []
    for selection in cleanup.values():
        entry = by_pair.get((selection.family, selection.harness))
        fresh = None
        if entry is not None:
            fresh = selection_for(entry, selection.kind, selection.ident,
                                  tracked_elsewhere=tracked)
        out.append(fresh or selection)
    out.extend(_repair_rows(repairs or {}, context or {}))
    out.sort(key=lambda s: (PRIORITY_ORDER.get(s.priority, 9),
                            FIX_ORDER.get(s.fix_in, 9), s.family, s.kind,
                            s.ident))
    return out


def _cleanup_sheet(wb: Workbook, rows: List[CleanupSelection], *,
                   dtx_program: str, dtx_phase: str) -> None:
    ws = wb.create_sheet("Complexity Cleanup")
    counts = {}
    for row in rows:
        counts[row.fix_in] = counts.get(row.fix_in, 0) + 1
    priorities = {}
    for row in rows:
        priorities[row.priority] = priorities.get(row.priority, 0) + 1

    def tally(mapping, order):
        parts = [f"{mapping[k]} {k}" for k in order if mapping.get(k)]
        return " · ".join(parts) or "nothing outstanding"

    ws["A1"] = "Complexity cleanup — what the customer has to correct"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")
    ws["A2"] = (f"DTx {dtx_program or '?'} · phase {dtx_phase or '?'} · "
                f"reviewed {date.today():%Y-%m-%d}")
    ws.merge_cells("A2:F2")
    ws["A3"] = f"{len(rows)} item(s) — where the change belongs: " \
               + tally(counts, [FIX_DTX, FIX_COMPLEXITY, FIX_EITHER,
                                FIX_CONFIRM, FIX_MAPPING])
    ws.merge_cells("A3:F3")
    ws["A4"] = "By priority: " + tally(priorities, [P_BLOCKER, P_HIGH, P_REVIEW]) \
               + ". Sorted so the top row is the first thing to raise; the " \
                 "'Customer email' sheet is the same list as sendable text."
    ws.merge_cells("A4:F4")
    for ref in ("A2", "A3", "A4"):
        ws[ref].font = Font(color="1F3B57")
        ws[ref].alignment = Alignment(vertical="center")
    ws["A3"].font = Font(bold=True, color="1F3B57")

    header_row = 6
    for column, title in enumerate(CLEANUP_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=column, value=title)
        cell.fill, cell.font = _HEAD_FILL, _HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for index, row in enumerate(rows, start=1):
        line = header_row + index
        harness = f"{row.harness} ({row.def_id})" if row.def_id else row.harness
        values = [index, row.priority, row.fix_in, row.action, row.note,
                  row.family, harness, kind_label(row.kind), row.ident,
                  row.condition,
                  row.builds, row.evidence, ""]
        for column, value in enumerate(values, start=1):
            ws.cell(row=line, column=column, value=value)
        ws.cell(row=line, column=2).fill = _PRIORITY_FILL.get(
            row.priority, _PRIORITY_FILL[P_REVIEW])
        ws.cell(row=line, column=2).font = _PRIORITY_FONT.get(
            row.priority, _PRIORITY_FONT[P_REVIEW])
        ws.cell(row=line, column=3).font = Font(bold=True)
        ws.cell(row=line, column=4).font = Font(bold=True)
        for column in (4, 5, 12):
            ws.cell(row=line, column=column).alignment = Alignment(
                wrap_text=True, vertical="top")

    widths = [5, 10, 18, 52, 58, 18, 22, 11, 14, 22, 15, 34, 18]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.freeze_panes = f"A{header_row + 1}"
    if rows:
        ws.auto_filter.ref = (f"A{header_row}:"
                              f"{get_column_letter(len(CLEANUP_HEADERS))}"
                              f"{header_row + len(rows)}")
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"


_EMAIL_GROUPS = [
    (FIX_DTX, "IN THE DTx REPORT", "Please correct these in the export."),
    (FIX_COMPLEXITY, "IN THE HARNESS COMPLEXITY MATRIX",
     "Another harness already tracks each code below, so the code is real "
     "and these files are the ones missing it."),
    (FIX_EITHER, "EITHER DOCUMENT — PLEASE CONFIRM WHICH IS RIGHT",
     "The two documents disagree and neither is provably wrong; each line "
     "names the question."),
    (FIX_CONFIRM, "FOR CONFIRMATION ONLY", "No change requested."),
]


def _where(row: CleanupSelection) -> str:
    """The place, without saying the same name twice."""
    parts = [x for x in (row.family, row.harness) if x and x != "—"]
    if len(parts) == 2 and parts[0] == parts[1]:
        parts = parts[:1]
    return " · ".join(parts) or "—"


def email_lines(rows: List[CleanupSelection], *, dtx_program: str = "",
                dtx_phase: str = "", prepared_by: str = "") -> List[str]:
    """The same work list as text an SE can paste into a mail client."""
    programme = f"{dtx_program or '?'} {dtx_phase or '?'}".strip()
    actionable = [r for r in rows if r.fix_in != FIX_MAPPING]
    requested = len({(r.action, r.priority, _where(r)) for r in actionable})
    lines = [
        f"Subject: {programme} — {requested} correction(s) requested "
        f"({len(actionable)} finding(s))",
        "",
        "Hello,",
        "",
        f"We resolved the Detailed DTx Circuits Report for {programme} against "
        "the harness complexity files. The items below are places the two "
        "documents disagree; each one needs a correction at source, because a "
        "fix made on our side comes back with the next export.",
        "",
    ]
    number = 0
    for fix_in, title, blurb in _EMAIL_GROUPS:
        group = [r for r in actionable if r.fix_in == fix_in]
        if not group:
            continue
        requests: List[tuple] = []
        seen: Dict[tuple, list] = {}
        for row in group:
            handle = (row.action, row.priority, _where(row))
            if handle not in seen:
                seen[handle] = []
                requests.append(handle)
            seen[handle].append(f"{kind_label(row.kind)} {row.ident}")
        lines.append(f"{title} ({len(requests)})")
        lines.append(f"  {blurb}")
        for handle in requests:
            number += 1
            action, priority, where = handle
            affected = seen[handle]
            lines.append(f"  {number}. [{priority}] {where} — {action}.")
            lines.append(f"     Affects: {', '.join(affected)}")
        lines.append("")
    if not number:
        lines.append("Nothing outstanding — the export reconciles against "
                     "every complexity file we checked it against.")
        lines.append("")
    lines += [
        "The attached workbook has the full detail on the 'Complexity Cleanup' "
        "sheet, including the part numbers behind each finding.",
        "",
        "Thank you,",
        prepared_by or "<your name>",
    ]
    return lines


def _email_sheet(wb: Workbook, rows: List[CleanupSelection], *,
                 dtx_program: str, dtx_phase: str, prepared_by: str) -> None:
    ws = wb.create_sheet("Customer email")
    ws["A1"] = "Draft — read it, then copy column A into your mail client"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = ("Every line comes from the Complexity Cleanup sheet; nothing "
                "here is sent for you.")
    ws["A2"].font = Font(color=_MUTED_INK, italic=True)
    for index, line in enumerate(email_lines(rows, dtx_program=dtx_program,
                                             dtx_phase=dtx_phase,
                                             prepared_by=prepared_by), start=4):
        cell = ws.cell(row=index, column=1, value=line)
        cell.alignment = Alignment(vertical="top", wrap_text=False)
        if line.startswith("Subject:") or line[:3].isupper() and line.strip():
            cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 110
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


def build_report(entries: Iterable[Entry],
                 cleanup: Dict[str, CleanupSelection],
                 *, dtx_program: str = "", dtx_phase: str = "",
                 repairs: Optional[Dict[str, str]] = None,
                 repair_context: Optional[Dict[str, dict]] = None,
                 quality=None, charts=None, prepared_by: str = "") -> bytes:
    """The review workbook, with the cleanup notes folded into every sheet."""
    entries = list(entries)
    wb = Workbook()

    _sheet(wb, "Read Me", ["Circuit applicability review"], [
        [f"DTx programme {dtx_program or '?'} · phase {dtx_phase or '?'}"],
        [f"Generated {date.today():%Y-%m-%d}"],
        [""],
        ["Each DTx harness family is resolved against the complexity file(s) "
         "mapped to it. A family may be mapped to several harnesses; each "
         "pairing is reported separately."],
        [""],
        [f"'{CLEANUP_COLUMN}' carries a note only for the rows the Systems "
         "Engineer selected in the workbench."],
        [""],
        ["START ON THE 'Complexity Cleanup' SHEET. It is the work list: one "
         "row per item, sorted worst first, and the third column says whether "
         "the correction belongs in the DTx, in the complexity matrix, or "
         "whether the two documents disagree and the customer has to say "
         "which is right. Every sales-code repair confirmed during the review "
         "is listed there too — the workbench applied it locally, but unless "
         "the customer fixes the source it returns with the next export."],
        [""],
        ["'Customer email' is that same list as text to paste into a mail "
         "client, grouped by where the change belongs. Read it before you "
         "send it; nothing is sent for you."],
        [""],
        ["Every circuit and connector shows its condition twice: 'as in DTx' is "
         "what the export stated, 'as decided' is what was analysed after any "
         "sales-code repair. 'Repaired' marks the rows where they differ, and "
         "the Sales-code repairs sheet lists each decision once with what "
         "depended on it."],
        [""],
        ["Verdicts: unconditional (no sales code, every build); all builds "
         "(conditioned but true for all); variant (some part numbers); never "
         "built (no build satisfies the condition — a defect or a missing "
         "code); no complexity (nothing mapped)."],
        ["A sales code the DTx uses that the complexity file does not track is "
         "treated as PRESENT, so circuits resting on it read wider than the "
         "data can justify."],
    ], first=True)

    circuit_rows = []
    for entry in entries:
        a = entry.analysis
        for c in a.circuits:
            key = item_key(entry.family, a.harness, KIND_CIRCUIT, c.circuit)
            decided = c.expression or ""
            # an absent original means "not recorded", never "changed" — the
            # Repaired flag must only fire on a difference we actually saw
            original = entry.original_circuit_conditions.get(c.circuit, decided)
            circuit_rows.append([
                entry.family, a.harness, a.def_id, c.circuit, c.classification,
                original, decided, "yes" if original != decided else "",
                len(c.builds_with), c.build_count,
                ", ".join(c.builds_with), ", ".join(c.untracked_codes),
                ", ".join(c.pins),
                cleanup[key].note if key in cleanup else "",
            ])
    if quality is not None:
        _quality_sheets(wb, quality)

    _sheet(wb, "Circuits", ["DTx family", "Harness", "Def id", "Circuit",
                            "Verdict", "Condition as in DTx",
                            "Condition as decided", "Repaired",
                            "Builds with", "Builds",
                            "Carried by", "Untracked codes", "Pins",
                            CLEANUP_COLUMN], circuit_rows)

    cnum_rows = []
    for entry in entries:
        a = entry.analysis
        for c in a.cnums:
            key = item_key(entry.family, a.harness, KIND_CONNECTOR, c.cnum)
            decided = c.expression or ""
            original = entry.original_cnum_conditions.get(c.cnum, decided)
            cnum_rows.append([
                entry.family, a.harness, a.def_id, c.cnum, c.connector_pn,
                c.classification, original, decided,
                "yes" if original != decided else "", len(c.builds_with),
                c.build_count, len(c.circuits), ", ".join(c.circuits),
                ", ".join(c.untracked_codes),
                cleanup[key].note if key in cleanup else "",
            ])
    _sheet(wb, "Connectors", ["DTx family", "Harness", "Def id", "CNUM",
                              "Connector PN", "Verdict", "Condition as in DTx",
                              "Condition as decided", "Repaired",
                              "Builds with", "Builds", "# circuits", "Circuits",
                              "Untracked codes", CLEANUP_COLUMN], cnum_rows)

    gap_rows = []
    for entry in entries:
        a = entry.analysis
        for g in a.code_gaps:
            key = item_key(entry.family, a.harness, KIND_GAP, g.code)
            gap_rows.append([
                entry.family, a.harness, a.def_id, g.code, g.occurrences,
                ", ".join(g.circuits), ", ".join(g.cnums),
                cleanup[key].note if key in cleanup else "",
            ])
    _sheet(wb, "Sales-code gaps", ["DTx family", "Harness", "Def id",
                                   "Sales code", "DTx rows", "Circuits",
                                   "Connectors", CLEANUP_COLUMN], gap_rows)

    repair_rows = []
    for original in sorted(repairs or {}):
        context = (repair_context or {}).get(original, {})
        repair_rows.append([
            original, (repairs or {})[original],
            context.get("kind", ""), context.get("rows", ""),
            ", ".join(context.get("families", [])),
            ", ".join(context.get("circuits", [])),
        ])
    _sheet(wb, "Sales-code repairs",
           ["Expression as in DTx", "Expression as decided", "Problem",
            "DTx rows", "Families", "Circuits"], repair_rows)

    if charts:
        # Written last so it sits at the end of the book, but in the exact
        # layout splice.inline.summary parses: this sheet is also an input.
        from splice.dtxcircuits.chart import write_chart_sheet, write_flat_sheet
        write_flat_sheet(wb, charts, dtx_program, dtx_phase)
        write_chart_sheet(wb, charts, dtx_program, dtx_phase)

    work = cleanup_rows(cleanup, entries, repairs, repair_context)
    _cleanup_sheet(wb, work, dtx_program=dtx_program, dtx_phase=dtx_phase)
    _email_sheet(wb, work, dtx_program=dtx_program, dtx_phase=dtx_phase,
                 prepared_by=prepared_by)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
