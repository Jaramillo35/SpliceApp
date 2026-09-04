"""A circuit chart per harness family: which part number carries which wire.

This is the thing the SE has been reconstructing by hand. Everything it needs
is already resolved by the time the review is done — the DTx gives the circuit
ends (circuit, connector, cavity, device, condition) and the complexity gives
the part numbers and the codes each one carries — so the chart is not new
analysis, it is the analysis written out in the shape people read.

The layout deliberately matches the **Circuit Summary** that Circuit Health
takes as input: a block per harness, opened by a ``<Family> - <def id>`` /
``Circuit`` header row carrying ``X~<part number>`` columns, then one row per
circuit end marked ``X`` under every build that carries it. Two reasons for
copying an existing format rather than inventing one:

* SEs already read it. A chart in a familiar shape needs no explanation.
* It round-trips. What this module writes, ``splice.inline.summary`` reads
  back, so a DTx-derived chart can be fed straight into Circuit Health and
  compared against the CAD-derived one. A test holds that both ways.

The one thing it cannot invent is wire physicals — size, material, colour are
not in the DTx, so those columns are left empty rather than guessed. A blank
cell says "not stated"; a made-up gauge would be read as fact.

The one thing it leaves out is the **No Connect**. ``N0`` is the DTx's marker
for a cavity wired to nothing — 29% of a real export's rows. It is not a
circuit, so it is not a row here; see ``splice.dtxcircuits.conventions``. The
count of what was dropped is kept on each chart and printed on the sheet,
because a chart much shorter than its export should say why.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from typing import Callable, Dict, Iterable, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from splice.dtxcircuits import conventions
from splice.dtxcircuits.models import (
    ALL_BUILDS, NEVER, NO_COMPLEXITY, VARIANT,
)
from splice.inline import salescode
from splice.inline.complexity import applies_in
from splice.inline.pairing import mate_name
from splice.inline.summary import (
    COL_CAV,
    COL_CIRCUIT,
    COL_CNUM,
    COL_COLOR,
    COL_DEVICE,
    COL_FAMILY,
    COL_MATERIAL,
    COL_SALES_CODE,
    COL_SIZE,
    COL_SUFFIX,
    FIRST_BUILD_COL,
    SHEET,
)

logger = logging.getLogger(__name__)

MARK = "X"

#: The readable sheet. ``SHEET`` (imported from splice.inline.summary) is the
#: blocked one Circuit Health parses; both are written, because flattening the
#: blocks would have quietly cost the round trip.
FLAT_SHEET = "Circuit Chart"

#: Column headers for the human-facing sheet. Positions are fixed by the
#: Circuit Summary contract, so the labels are placed, not appended.
_LABELS = {
    COL_FAMILY: "Harness Family",
    COL_CIRCUIT: "Circuit",
    COL_SUFFIX: "Suffix",
    COL_SIZE: "Size",
    COL_MATERIAL: "Material",
    COL_COLOR: "Color",
    COL_CNUM: "CNUM",
    COL_CAV: "Cavity",
    COL_DEVICE: "Device",
    COL_SALES_CODE: "Sales Code",
}


@dataclass
class ChartRow:
    """One circuit end, and the part numbers that carry it."""

    circuit: str
    cnum: str = ""
    cavity: str = ""
    device: str = ""
    connector_pn: str = ""
    #: the condition as the DTx states it for this circuit, flowed across
    #: every harness the circuit reaches
    expression: str = ""
    #: the same condition re-expressed in the codes THIS harness tracks
    harness_expression: str = ""
    classification: str = ""
    builds: List[str] = field(default_factory=list)
    #: a wire running to a generated splice rather than to a device
    is_splice: bool = False
    #: For a splice leg: the expression under which THIS leg exists — the
    #: end's own condition and the configuration's, together. A circuit can
    #: be spliced on one part number and a plain wire on another, so a
    #: single expression per circuit cannot describe its legs.
    leg_expression: str = ""
    #: which configuration of the circuit this row belongs to, where the
    #: complexity supports more than one
    configuration: str = ""
    #: The far end of the WIRE, which never leaves this harness. A wire has
    #: two ends; the chart is written one end per row, so without these the
    #: far end has to be found by eye.
    other_family: str = ""
    other_cnum: str = ""
    other_cavity: str = ""
    other_device: str = ""
    #: The connector this one mates with in the NEXT harness. An inline is a
    #: joint, not a wire: the circuit runs device-to-inline inside this
    #: harness, the two inline halves mate, and another wire continues on the
    #: far side. Those are two different connections and the chart states
    #: both — collapsing them into one drew wires through harnesses the
    #: circuit only passes through.
    mate_family: str = ""
    mate_cnum: str = ""
    mate_cavity: str = ""

    @property
    def end_type(self) -> str:
        if self.is_splice:
            return "Splice"
        return "Inline" if is_pass_through(self.cnum) else "Device"

    def carried_by(self, part_number: str) -> bool:
        return part_number in self.builds

    @property
    def is_finding(self) -> bool:
        return self.classification == NEVER

    def marks(self, part_numbers: Sequence[str]) -> List[str]:
        return [MARK if self.carried_by(pn) else "" for pn in part_numbers]


@dataclass
class Chart:
    """One harness family resolved against one complexity file."""

    family: str
    harness: str
    def_id: str = ""
    part_numbers: List[str] = field(default_factory=list)
    rows: List[ChartRow] = field(default_factory=list)
    #: circuit -> generated splice name, for the circuits that needed one
    splices: Dict[str, str] = field(default_factory=dict)
    #: DTx rows dropped as No Connect. Reported, never silent: the chart is
    #: markedly shorter than the export it came from, and an SE comparing
    #: the two deserves to be told why rather than left to wonder.
    no_connect_rows: int = 0

    @property
    def block_title(self) -> str:
        """``BODY_LEFT - 7010`` — the header the Circuit Summary parser keys on."""
        return f"{self.harness} - {self.def_id}" if self.def_id else self.harness

    @property
    def circuits(self) -> int:
        return len({row.circuit for row in self.rows})

    @property
    def findings(self) -> int:
        return sum(1 for row in self.rows if row.is_finding)

    def coverage(self, part_number: str) -> int:
        """How many circuit ends this part number carries."""
        return sum(1 for row in self.rows if row.carried_by(part_number))


#: A circuit reaching this many cavities inside one harness cannot be run as
#: a single wire — the branches have to meet somewhere, and that somewhere is
#: a splice.
SPLICE_MIN_ENDS = 3


#: Re-exported: the rule now lives with the other DTx reading conventions,
#: because the analysis needs it too — a blank on a joint must not make a
#: circuit read as unconditional there either.
is_pass_through = conventions.is_pass_through


def _union(parts: Sequence[str]) -> Optional[str]:
    unique = sorted({f"({p})" for p in parts if p})
    return "/".join(unique) if unique else None


def flowed_conditions(rows: Sequence) -> Dict[str, Optional[str]]:
    """One condition per circuit, unioned over its device ends everywhere.

    The whole-circuit view, used where a harness states nothing of its own.
    ``None`` means unconditional, which no expression says.
    """
    stated: Dict[str, List[str]] = {}
    unconditional: set = set()
    for row in rows:
        circuit = getattr(row, "circuit", "")
        if not circuit:
            continue
        stated.setdefault(circuit, [])
        if is_pass_through(getattr(row, "cnum", "")):
            continue
        condition = conventions.effective_condition(getattr(row, "sales_code", ""))
        if condition is None:
            unconditional.add(circuit)
        else:
            stated[circuit].append(condition)
    return {circuit: (None if circuit in unconditional else _union(parts))
            for circuit, parts in stated.items()}


def harness_conditions(rows: Sequence) -> Dict[tuple, Optional[str]]:
    """``(circuit, family) -> condition``, reconciled per harness.

    A circuit does not have one applicability. It has one per harness, because
    what makes it present in a harness is the devices *that harness* connects
    it to. ``501`` on a ground stud in the IP is ground truth that the segment
    is always there; the same circuit reaching a HAH device in the HVAC is
    present only on HAH builds. Collapsing those to a single circuit-wide
    condition loses whichever one is narrower — 60 circuits in 2028RU X2_A
    have exactly that shape.

    So each harness is resolved from its own ends, in order of authority:

    1. **Its device ends.** One unconditional device end (blank, or a bare
       ``501``) makes the segment unconditional; otherwise the stated
       conditions are OR-ed, since any fitted device pulls the wire in.
    2. **Its inline ends that state something**, for a harness the circuit
       only passes through. Rare (194 rows) but decisive when present.
    3. **The circuit's conditions elsewhere.** A pure pass-through states
       nothing, and inheriting is better than calling it unconditional: in
       ``D442`` every device end reads ``SDE`` and only the DASH inlines are
       blank, so the DASH segment is ``SDE`` too, not "always".

    A blank on an inline never reaches any of these. It is silence, and
    reading it as "unconditional" is what made a whole circuit collapse.
    """
    device: Dict[tuple, List[str]] = {}
    inline: Dict[tuple, List[str]] = {}
    unconditional: set = set()
    keys: set = set()

    for row in rows:
        circuit = getattr(row, "circuit", "")
        family = getattr(row, "harness_family", "")
        if not circuit:
            continue
        key = (circuit, family)
        keys.add(key)
        condition = conventions.effective_condition(getattr(row, "sales_code", ""))
        if is_pass_through(getattr(row, "cnum", "")):
            if condition is not None:
                inline.setdefault(key, []).append(condition)
            continue
        if condition is None:
            unconditional.add(key)
        else:
            device.setdefault(key, []).append(condition)

    whole = flowed_conditions(rows)
    resolved: Dict[tuple, Optional[str]] = {}
    for key in keys:
        circuit, _family = key
        if key in unconditional:
            resolved[key] = None
        elif key in device:
            resolved[key] = _union(device[key])
        elif key in inline:
            resolved[key] = _union(inline[key])
        else:
            resolved[key] = whole.get(circuit)
    return resolved


def carrying_builds(condition: Optional[str], complexity) -> List[str]:
    """The part numbers of ``complexity`` that satisfy ``condition``."""
    if complexity is None or not getattr(complexity, "builds", None):
        return []
    if condition is None:
        return [b.part_number for b in complexity.builds]
    vocabulary = getattr(complexity, "complexity_codes", set())
    return [b.part_number for b in complexity.builds
            if applies_in(condition, b.codes, vocabulary)]


def harness_expression(condition: Optional[str], builds: Sequence[str],
                       complexity) -> str:
    """Re-express a condition in the codes THIS harness actually tracks.

    The DTx writes one condition for the whole circuit, in whatever codes the
    programme uses. A harness's complexity only tracks some of them, so the
    same condition has to be restated per harness — otherwise the chart cites
    codes that harness has no column for.

    A condition already written entirely in codes this harness tracks is kept
    as it is. It selects the right builds by construction, and rewriting it
    would hand the SE a different-looking expression for no gain — the DTx
    ``(QA1)`` came back as ``-QA2`` before this rule, which is equally true
    and needlessly unfamiliar.

    Only where the condition leans on a code this harness has no column for is
    it restated, and then from the outcome rather than the text: take the part
    numbers it selects, and ask the Splice Generation engine for an expression
    selecting exactly those. Returns ``""`` for "every build" and for "no
    build", since neither is a condition, and falls back to ``""`` rather than
    guessing when the generated expression does not verify.
    """
    if complexity is None or not getattr(complexity, "builds", None):
        return ""
    everything = [b.part_number for b in complexity.builds]
    carried = [pn for pn in everything if pn in set(builds)]
    if not carried or len(carried) == len(everything):
        return ""

    vocabulary = getattr(complexity, "complexity_codes", set())
    if condition and all(code in vocabulary
                         for code in salescode.codes_in(condition)):
        return condition

    code_map = {b.part_number: set(b.codes) for b in complexity.builds}
    try:
        from splice.splice_gen import (generate_expression_for_selected_pns,
                                       validate_generated_expression)
        expression = generate_expression_for_selected_pns(carried, code_map)
        if expression and validate_generated_expression(expression, carried, code_map):
            return expression
        if expression:
            logger.info("Generated expression for %s did not verify; dropped",
                        getattr(complexity, "name", "?"))
    except Exception as exc:  # noqa: BLE001 — a chart must never fail to build
        logger.info("Could not restate a condition for %s: %s",
                    getattr(complexity, "name", "?"), exc)
    return ""


def build_charts(entries: Iterable, rows: Sequence,
                 splice_min_ends: int = SPLICE_MIN_ENDS,
                 progress: Optional[Callable[[float, str], None]] = None,
                 ) -> List[Chart]:
    """One chart per family × harness pairing, from the resolved analysis.

    ``rows`` are the DTx circuit rows the analysis was run on — the repaired
    ones, so the chart shows what was actually resolved rather than what the
    export said before the SE fixed it.

    ``progress`` is called once per family with ``(fraction, message)``. This
    is the longest step in the page — planning every circuit's wiring against
    every part number — so it reports rather than going quiet, and the caller
    runs it off the event loop.
    """
    by_family: Dict[str, List] = {}
    for row in rows:
        by_family.setdefault(getattr(row, "harness_family", ""), []).append(row)
    # Reconciled per harness, not once per circuit: a 501 ground in one
    # harness must not erase a HAH device in another.
    conditions = harness_conditions(rows)

    charts: List[Chart] = []
    entries = list(entries)
    for index, entry in enumerate(entries, start=1):
        if progress is not None:
            progress((index - 1) / max(len(entries), 1),
                     f"Charting {entry.family} ({index} of {len(entries)})…")
        analysis = entry.analysis
        complexity = getattr(entry, "complexity", None)
        family_rows = by_family.get(entry.family, [])

        applicability = {c.circuit: c for c in analysis.circuits}
        connector_pn = {}
        for row in family_rows:
            if row.cnum and row.cnum not in connector_pn:
                connector_pn[row.cnum] = row.connector_pn or ""

        chart = Chart(family=entry.family, harness=analysis.harness,
                      def_id=analysis.def_id,
                      part_numbers=_part_numbers(analysis, complexity))

        # resolved once per circuit, not once per end: the condition is a
        # property of the wire, so every end of it is carried identically
        resolved: Dict[str, tuple] = {}

        # the DTx condition stated at each individual end, which is what
        # decides that end's presence on a given part number
        conditions_by_end: Dict[tuple, str] = {}

        seen = set()
        for row in family_rows:
            key = (row.circuit, row.cnum, row.pin)
            if not row.circuit or key in seen:
                continue
            # A No Connect is a cavity wired to nothing. It has no far end and
            # nothing to splice to, so it never becomes a chart row — leaving
            # it in fabricated both.
            if conventions.is_no_connect(row.circuit, getattr(row, "function", "")):
                chart.no_connect_rows += 1
                continue
            seen.add(key)
            if row.circuit not in resolved:
                condition = conditions.get((row.circuit, entry.family))
                if complexity is not None:
                    builds = carrying_builds(condition, complexity)
                    restated = harness_expression(condition, builds, complexity)
                else:
                    # No complexity file to resolve against. Fall back to what
                    # the analysis already worked out rather than reporting
                    # nothing carried — an empty chart would read as every
                    # circuit being never built, which is a lie, not a gap.
                    item = applicability.get(row.circuit)
                    builds = list(item.builds_with) if item else []
                    restated = ""
                resolved[row.circuit] = (
                    condition or "", builds, restated,
                    _classify(builds, chart.part_numbers))
            conditions_by_end[(row.circuit, row.cnum, row.pin)] = \
                conventions.effective_condition(row.sales_code) or ""
            expression, builds, restated, classification = resolved[row.circuit]
            chart.rows.append(ChartRow(
                circuit=row.circuit, cnum=row.cnum, cavity=row.pin,
                device=row.function or "",
                connector_pn=connector_pn.get(row.cnum, ""),
                expression=expression, harness_expression=restated,
                classification=classification, builds=list(builds)))

        _add_splices(chart, splice_min_ends, complexity, conditions_by_end)
        chart.rows.sort(key=lambda r: (r.circuit, r.cnum, _cavity_key(r.cavity)))
        charts.append(chart)

    if progress is not None:
        progress(0.98, "Linking circuit ends across harnesses…")
    link_ends(charts)
    logger.info("Built %d circuit chart(s), %d row(s), %d splice(s)",
                len(charts), sum(len(c.rows) for c in charts),
                sum(len(c.splices) for c in charts))
    return charts


def _classify(builds: Sequence[str], part_numbers: Sequence[str]) -> str:
    if not part_numbers:
        return NO_COMPLEXITY
    if not builds:
        return NEVER
    return ALL_BUILDS if len(builds) == len(part_numbers) else VARIANT


def _add_splices(chart: Chart, minimum: int, complexity=None,
                 conditions_by_end: Optional[Dict[tuple, str]] = None) -> None:
    """Wire each circuit the way this harness's own options require.

    A circuit does not have one topology. M34 in Door_Driver_2 reaches an LCF
    device, a LEQ/LEM device and an inline: the part number carrying LCF has
    all three ends and needs a splice, the one without has two and is a plain
    wire. So the ends are grouped by which part numbers actually carry them
    and each group is planned separately, by the same planner Splice
    Generation uses — the two surfaces cannot disagree about what a splice is.

    Every leg carries its own expression: its end's condition together with
    its configuration's. An end that connects somewhere different in another
    configuration gets a second row, because one row cannot say two things;
    an end that connects to the same place in every configuration keeps one
    row and the expressions are merged by the planner.

    Without a complexity file there are no part numbers to group by, so the
    simple rule stands: three or more ends share one splice.
    """
    if minimum <= 0:
        return
    by_circuit: Dict[str, List[ChartRow]] = {}
    for row in chart.rows:
        by_circuit.setdefault(row.circuit, []).append(row)

    builds = list(getattr(complexity, "builds", []) or [])
    conditions_by_end = conditions_by_end or {}

    for circuit, ends in sorted(by_circuit.items()):
        if not builds:
            if len(ends) >= minimum:
                _simple_splice(chart, circuit, ends)
            continue
        connections = _plan_circuit(circuit, ends, builds, conditions_by_end)
        if connections is None:
            if len(ends) >= minimum:
                _simple_splice(chart, circuit, ends)
            continue
        _wire_from_plan(chart, circuit, ends, connections, minimum)


def _simple_splice(chart: Chart, circuit: str, ends: List[ChartRow]) -> None:
    """One splice for every end — what the chart does with no complexity."""
    name = f"S{circuit}{_int_to_alpha_suffix(0)}"
    chart.splices[circuit] = name
    for index, end in enumerate(list(ends)):
        branch = _branch_row(end, circuit, name, _int_to_alpha_suffix(index),
                             end.harness_expression or end.expression, "")
        _join(end, chart.family, branch)
        _join(branch, chart.family, end)
        chart.rows.append(branch)


def _leg_text(expression: str) -> str:
    """The engine writes TRUE for an unconditional leg; the chart leaves it
    blank, as it does for every other unconditional condition."""
    text = (expression or "").strip()
    return "" if text.upper() == "TRUE" else text


def _branch_row(end: ChartRow, circuit: str, name: str, cavity: str,
                expression: str, configuration: str) -> ChartRow:
    return ChartRow(
        circuit=circuit, cnum=name, cavity=cavity, device=f"SPLICE {name}",
        expression=end.expression, harness_expression=end.harness_expression,
        classification=end.classification, builds=list(end.builds),
        is_splice=True, leg_expression=_leg_text(expression),
        configuration=configuration)


def _leg_condition(circuit: str, end: ChartRow,
                   conditions_by_end: Dict[tuple, str],
                   circuit_condition: str) -> str:
    """The condition under which THIS end of the circuit exists.

    A blank cell on an inline is silence, not "unconditional". The DTx states
    applicability at devices and leaves the joints empty — 1,924 of the 2,954
    blank cells in a real export sit on inlines — so an inline carries
    whatever the circuit carries.

    Reading that blank as TRUE was a Boolean mistake with a visible cost: a
    circuit conditioned ``-AAA`` produced a splice whose leg to the inline
    had no expression at all, which says the leg is on every part number
    while the circuit it belongs to is on five of twenty. A leg cannot be
    present where its circuit is absent.
    """
    stated = conditions_by_end.get((circuit, end.cnum, end.cavity), "")
    if stated or not is_pass_through(end.cnum):
        return stated
    return circuit_condition


def _plan_circuit(circuit: str, ends: List[ChartRow], builds,
                  conditions_by_end: Dict[tuple, str]):
    """Ask Splice Generation how this circuit is wired on this harness."""
    try:
        import pandas as pd

        from splice.splice_gen import plan_connections
        # every end of a circuit carries the same resolved condition
        circuit_condition = next(
            (e.expression for e in ends if e.expression), "")
        option_rows = [{
            "CNUM": end.cnum, "Pin": end.cavity, "Circuit": circuit,
            "Sales Code": _leg_condition(circuit, end, conditions_by_end,
                                         circuit_condition),
        } for end in ends]
        code_map = {build.part_number: set(build.codes) for build in builds}
        _configurations, connections = plan_connections(pd.DataFrame(option_rows), code_map)
        return [row for row in connections if row.get("Circuit Name") == circuit]
    except Exception as exc:  # noqa: BLE001 — a chart must still be drawn
        logger.info("Could not plan %s: %s", circuit, exc)
        return None


def _wire_from_plan(chart: Chart, circuit: str, ends: List[ChartRow],
                    connections: List[dict], minimum: int) -> None:
    by_key = {(end.cnum, end.cavity): end for end in ends}
    used: set = set()
    splice_cavity = 0

    for row in connections:
        name = str(row.get("Splice Name", ""))
        expression = str(row.get("Sales Code", ""))
        configuration = str(row.get("Configuration", ""))
        left = (str(row.get("From CNUM", "")), str(row.get("From Pin", "")))
        right = (str(row.get("To CNUM", "")), str(row.get("To Pin", "")))

        if name and name in (left[0], right[0]):
            chart.splices.setdefault(circuit, name)
            device_key = right if left[0] == name else left
            end = _end_for(chart, circuit, by_key, used, device_key,
                           expression, configuration)
            if end is None:
                continue
            branch = _branch_row(end, circuit, name,
                                 _int_to_alpha_suffix(splice_cavity),
                                 expression, configuration)
            splice_cavity += 1
            _join(end, chart.family, branch)
            _join(branch, chart.family, end)
            chart.rows.append(branch)
            continue

        # a direct connection: two device ends, one wire
        first = _end_for(chart, circuit, by_key, used, left, expression, configuration)
        second = _end_for(chart, circuit, by_key, used, right, expression, configuration)
        if first is not None and second is not None:
            _join(first, chart.family, second)
            _join(second, chart.family, first)


def _end_for(chart: Chart, circuit: str, by_key: Dict[tuple, ChartRow],
             used: set, key: tuple, expression: str, configuration: str):
    """The chart row for one end of a planned wire.

    An end already wired by another configuration gets a duplicate row: it
    connects somewhere different there, and one row cannot carry two far
    ends. Where the planner merged two configurations into one connection
    there is only ever one row to make.
    """
    base = by_key.get(key)
    if base is None:
        return None
    if key not in used:
        used.add(key)
        base.leg_expression = _leg_text(expression)
        base.configuration = configuration
        return base
    duplicate = ChartRow(
        circuit=base.circuit, cnum=base.cnum, cavity=base.cavity,
        device=base.device, connector_pn=base.connector_pn,
        expression=base.expression, harness_expression=base.harness_expression,
        classification=base.classification, builds=list(base.builds),
        leg_expression=_leg_text(expression), configuration=configuration)
    chart.rows.append(duplicate)
    return duplicate


def _join(row: ChartRow, family: str, other: ChartRow) -> None:
    row.other_family = family
    row.other_cnum = other.cnum
    row.other_cavity = other.cavity
    row.other_device = other.device


def link_ends(charts: Sequence[Chart]) -> None:
    """Fill in each wire's far end, and each inline's mate.

    Two different relationships, and the chart needs both:

    * **The wire.** It runs between two ends *inside one harness* and never
      leaves it — a wire cannot cross a harness boundary except through a
      connector. So the pairing is done per chart: two open ends of a circuit
      in the same harness are the two ends of one wire. Three or more were
      already joined to a splice when it was made.
    * **The mate.** An inline's other half sits in the next harness —
      ``X301A`` at ``Y301A``. That is how the circuit continues, and it is a
      joint rather than a wire.

    Pairing used to be global: "the circuit has exactly two unpaired ends
    anywhere, join them". That drew a wire straight from a device in one
    harness to a device in another, and it failed outright whenever one DTx
    family was mapped to several complexity files — the same two DTx rows
    then appear once per chart, so the count never matched and nothing was
    joined at all. On 2028RU X1 that left circuit A0, two devices inside
    BATTERY POSITIVE, unconnected in both of its charts.

    A single end alone in a harness gets no wire. That is honest: the DTx
    does not say what it joins to.
    """
    for chart in charts:
        by_circuit: Dict[str, List[ChartRow]] = {}
        for row in chart.rows:
            by_circuit.setdefault(row.circuit, []).append(row)
        for ends in by_circuit.values():
            open_ends = [row for row in ends if not row.other_cnum]
            if len(open_ends) == 2:
                first, second = open_ends
                _join(first, chart.family, second)
                _join(second, chart.family, first)

    # the mate is looked up across every chart, since it is by definition in
    # another harness; a row is indexed per (circuit, connector)
    elsewhere: Dict[tuple, List[tuple]] = {}
    for chart in charts:
        for row in chart.rows:
            elsewhere.setdefault((row.circuit, row.cnum.upper()), []).append(
                (chart, row))

    for chart in charts:
        for row in chart.rows:
            mate = mate_name(row.cnum)
            if not mate:
                continue
            candidates = elsewhere.get((row.circuit, mate), [])
            # prefer a half in a different harness — that is what a mate is
            match = next((pair for pair in candidates
                          if pair[0].harness != chart.harness), None)
            if match is None:
                continue
            other_chart, other_row = match
            row.mate_family = other_chart.family
            row.mate_cnum = other_row.cnum
            row.mate_cavity = other_row.cavity


def _int_to_alpha_suffix(index: int) -> str:
    """0 -> A, 25 -> Z, 26 -> AA. The scheme Splice Generation already uses."""
    result, value = "", index
    while True:
        value, remainder = divmod(value, 26)
        result = chr(ord("A") + remainder) + result
        if value == 0:
            break
        value -= 1
    return result


def _part_numbers(analysis, complexity=None) -> List[str]:
    """Every part number of the harness, in a stable order.

    The complexity file is the authority when it is available: it lists every
    build, including one that carries nothing, and an empty column is itself
    the finding. Without it, fall back to whatever the analysis mentions.
    """
    if complexity is not None and getattr(complexity, "builds", None):
        return [b.part_number for b in complexity.builds]
    ordered: List[str] = []
    for circuit in analysis.circuits:
        for part in list(circuit.builds_with) + list(circuit.builds_without):
            if part not in ordered:
                ordered.append(part)
    return sorted(ordered)


def _cavity_key(value: str):
    """Order cavities the way they are allocated, not the way they spell.

    Numbers sort numerically (2 before 10). Letters sort by length first, so a
    splice runs A, B … Z, AA, AB — plain alphabetical order would file AA
    between A and B and scatter a 269-way ground net.
    """
    text = (value or "").strip()
    head = text.split(":")[0]
    if head.isdigit():
        return (0, len(head), int(head), text)
    return (1, len(head), 0, text)


# --------------------------------------------------------------------------
# workbook
# --------------------------------------------------------------------------

_TITLE_FONT = Font(bold=True, size=12)
_BLOCK_FILL = PatternFill("solid", fgColor="1F3B57")
_BLOCK_FONT = Font(bold=True, color="FFFFFF")
_MARK_FONT = Font(bold=True)
_NEVER_FILL = PatternFill("solid", fgColor="F8D7DA")

# Hoisted, not built per cell. A real programme is ~5,400 circuit ends against
# ~20 part numbers — 115,000 mark cells — and constructing a style object for
# each one cost more than everything else in this module put together.
_CENTER = Alignment(horizontal="center")
_BLOCK_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_chart_sheet(wb: Workbook, charts: Sequence[Chart],
                      program: str = "", phase: str = "") -> None:
    """Write the blocked Circuit Summary sheet into ``wb``.

    The sheet name and geometry are the ones ``splice.inline.summary`` parses,
    so Circuit Health can take this workbook as an input unchanged.

    The row index is tracked here rather than read back from the sheet.
    ``ws.max_row`` rescans every cell written so far, so styling each row via
    ``ws[ws.max_row]`` is quadratic: on a real programme (~5,400 circuit ends
    against ~20 part numbers) that alone took 33 seconds and, run on the event
    loop, dropped the browser's connection.
    """
    ws = wb.create_sheet(SHEET)
    ws.append(["Circuit Summary", " ".join(p for p in (program, phase) if p),
               _no_connect_note(charts)])
    ws.cell(1, 1).font = _TITLE_FONT
    line_no = 1

    widest = 0
    for chart in charts:
        span = len(chart.part_numbers)
        widest = max(widest, span)

        header = [""] * (FIRST_BUILD_COL + span)
        header[COL_FAMILY] = chart.block_title
        header[COL_CIRCUIT] = "Circuit"
        for offset, part in enumerate(chart.part_numbers):
            header[FIRST_BUILD_COL + offset] = f"{MARK}~{part}"
        ws.append(header)
        line_no += 1
        for column in (COL_FAMILY + 1, COL_CIRCUIT + 1):
            cell = ws.cell(line_no, column)
            cell.fill, cell.font, cell.alignment = _BLOCK_FILL, _BLOCK_FONT, _BLOCK_ALIGN
        for offset in range(span):
            cell = ws.cell(line_no, FIRST_BUILD_COL + offset + 1)
            cell.fill, cell.font, cell.alignment = _BLOCK_FILL, _BLOCK_FONT, _BLOCK_ALIGN

        for row in chart.rows:
            line = [""] * (FIRST_BUILD_COL + span)
            line[COL_FAMILY] = chart.family
            line[COL_CIRCUIT] = row.circuit
            line[COL_CNUM] = row.cnum
            line[COL_CAV] = row.cavity
            line[COL_DEVICE] = row.device
            # The Circuit Summary's own Sales Code column carries the harness
            # form, because that is the one true of this block's part numbers.
            # The DTx form rides in the Suffix column, which the parser reads
            # but nothing downstream conditions on.
            line[COL_SALES_CODE] = row.harness_expression or row.expression
            if row.harness_expression and row.harness_expression != row.expression:
                line[COL_SUFFIX] = row.expression
            marks = row.marks(chart.part_numbers)
            for offset, mark in enumerate(marks):
                line[FIRST_BUILD_COL + offset] = mark
            ws.append(line)
            line_no += 1

            # Only the cells that carry something are styled: an empty cell
            # gains nothing from being centred.
            for offset, mark in enumerate(marks):
                if mark:
                    cell = ws.cell(line_no, FIRST_BUILD_COL + offset + 1)
                    cell.alignment, cell.font = _CENTER, _MARK_FONT
            # A circuit no build carries is the whole point of the chart, and
            # an entirely empty row is easy to scroll past — so it is coloured.
            if row.is_finding:
                for column in range(1, FIRST_BUILD_COL + span + 1):
                    ws.cell(line_no, column).fill = _NEVER_FILL

        ws.append([])
        line_no += 1

    for index in range(1, FIRST_BUILD_COL + widest + 1):
        letter = get_column_letter(index)
        ws.column_dimensions[letter].width = 16 if index < FIRST_BUILD_COL else 14
    ws.freeze_panes = ws.cell(3, FIRST_BUILD_COL + 1)


#: The flat sheet's fixed columns, in reading order. Everything the DTx does
#: not state (wire size, material, colour) is left out rather than written as
#: an empty column — the blocked sheet has to carry them because the Circuit
#: Summary contract fixes their positions, but nothing here does.
FLAT_COLUMNS = [
    "Harness Family", "Harness", "Def Id", "Circuit", "End", "CNUM",
    "Cavity", "Connector PN", "Device",
    "Sales Code (DTx)", "Sales Code (this harness)", "Verdict",
    "Other End Harness", "Other End CNUM", "Other End Cavity",
    "Other End Device",
    # the inline's other half, in the next harness — a joint, not a wire
    "Mates With Harness", "Mates With CNUM", "Mates With Cavity",
    # what has to be true for THIS wire to exist, and which of the circuit's
    # configurations it belongs to
    "Leg Sales Code", "Configuration",
    "Builds Carrying",
]


def _no_connect_note(charts: Sequence[Chart]) -> str:
    """What was left out, said on the sheet rather than left to be noticed."""
    dropped = sum(c.no_connect_rows for c in charts)
    if not dropped:
        return ""
    return (f"{dropped} No Connect row(s) excluded — a cavity wired to "
            f"nothing is not a circuit")


def part_number_columns(charts: Sequence[Chart]) -> List[tuple]:
    """Every part number in the study, kept in harness order.

    One column each, grouped by the harness they belong to, so the matrix
    reads left to right the way the harnesses are listed. Returned as
    ``(harness, part number)`` because two harnesses can ship the same number
    and the header has to say which column is which.
    """
    columns: List[tuple] = []
    seen = set()
    for chart in charts:
        for part in chart.part_numbers:
            key = (chart.harness, part)
            if key not in seen:
                seen.add(key)
                columns.append(key)
    return columns


def write_flat_sheet(wb: Workbook, charts: Sequence[Chart],
                     program: str = "", phase: str = "") -> None:
    """One table, one header row, one row per circuit end.

    The blocked sheet repeats a header for every harness because that is what
    the Circuit Summary parser keys on. It is the wrong shape to read or to
    filter: the repeated banners break sorting, and the fixed column positions
    force four columns the DTx cannot fill. This sheet is the same data as a
    plain table — title on row 1, the only header on row 2, and no column that
    is empty for every row.
    """
    ws = wb.create_sheet(FLAT_SHEET)
    ws.append(["Circuit Chart", " ".join(p for p in (program, phase) if p),
               _no_connect_note(charts)])
    ws.cell(1, 1).font = _TITLE_FONT

    parts = part_number_columns(charts)
    header = list(FLAT_COLUMNS) + [part for _harness, part in parts]
    ws.append(header)
    for index in range(1, len(header) + 1):
        cell = ws.cell(2, index)
        cell.fill, cell.font = _BLOCK_FILL, _BLOCK_FONT
        cell.alignment = _BLOCK_ALIGN

    index_of = {key: len(FLAT_COLUMNS) + offset
                for offset, key in enumerate(parts)}
    line_no = 2
    for chart in charts:
        own = [(index_of[(chart.harness, part)], part)
               for part in chart.part_numbers]
        for row in chart.rows:
            line = [""] * len(header)
            line[0] = chart.family
            line[1] = chart.harness
            line[2] = chart.def_id
            line[3] = row.circuit
            line[4] = row.end_type
            line[5] = row.cnum
            line[6] = row.cavity
            line[7] = row.connector_pn
            line[8] = row.device
            line[9] = row.expression
            line[10] = row.harness_expression
            line[11] = row.classification
            line[12] = row.other_family
            line[13] = row.other_cnum
            line[14] = row.other_cavity
            line[15] = row.other_device
            line[16] = row.mate_family
            line[17] = row.mate_cnum
            line[18] = row.mate_cavity
            line[19] = row.leg_expression
            line[20] = row.configuration
            line[21] = len(row.builds)
            carried = set(row.builds)
            for column, part in own:
                if part in carried:
                    line[column] = MARK
            ws.append(line)
            line_no += 1
            for column, part in own:
                if part in carried:
                    cell = ws.cell(line_no, column + 1)
                    cell.alignment, cell.font = _CENTER, _MARK_FONT
            if row.is_finding:
                for column in range(1, len(header) + 1):
                    ws.cell(line_no, column).fill = _NEVER_FILL

    widths = [18, 18, 9, 12, 9, 12, 8, 15, 22, 24, 24, 14, 18, 15, 9, 22,
              18, 15, 9, 24, 12, 10]
    for offset, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(offset)].width = width
    for offset in range(len(FLAT_COLUMNS) + 1, len(header) + 1):
        ws.column_dimensions[get_column_letter(offset)].width = 14
    ws.freeze_panes = ws.cell(3, 5)
    if line_no > 2:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(header))}{line_no}"


def build_chart_workbook(charts: Sequence[Chart], program: str = "",
                         phase: str = "") -> bytes:
    """The chart on its own, ready to hand to Circuit Health."""
    wb = Workbook()
    wb.remove(wb.active)
    write_flat_sheet(wb, charts, program, phase)
    write_chart_sheet(wb, charts, program, phase)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
