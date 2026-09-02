"""Read an individual harness complexity file: its build table AND its identity.

The build table lives on the ``Complexity`` sheet (``ID=`` in A1, sales codes
across row 1, part numbers down column A) and is parsed by the existing
:func:`splice.inline.complexity.read_complexity`.

The programme identity — year, vehicle, build phase, harness name — lives in
the info table on the **Harness PN** sheet, beside its labels. That is the only
trustworthy source: filenames carry ids and phases that disagree with the file
contents often enough that trusting them attributes applicability to the wrong
harness.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass
from typing import Tuple

from openpyxl import load_workbook

from splice.common.errors import SpliceError
from splice.inline.complexity import read_complexity
from splice.inline.model import Harness

logger = logging.getLogger(__name__)

INFO_SHEET = "Harness PN"
#: label in the info table -> field it fills
_LABELS = {"year:": "year", "vehicle:": "vehicle", "phase:": "phase",
           "harness:": "harness", "id:": "def_id"}
MAX_INFO_ROWS = 60
MAX_INFO_COLS = 30


@dataclass
class ComplexityMeta:
    """What a complexity file says about itself, read from ``Harness PN``."""

    year: str = ""
    vehicle: str = ""
    phase: str = ""
    harness: str = ""
    def_id: str = ""
    filename: str = ""

    @property
    def program(self) -> str:
        """``2028`` + ``RU`` -> ``2028RU``, matching the DTx's own wording."""
        return f"{self.year}{self.vehicle}".strip()

    @property
    def complete(self) -> bool:
        return bool(self.year and self.vehicle and self.phase)


def normalize_program(value: str) -> str:
    """``2028RU`` / ``28RU`` / ``2028 RU`` -> ``28RU`` for comparison."""
    text = re.sub(r"[^A-Z0-9]", "", str(value or "").upper())
    match = re.match(r"^(?:20)?(\d{2})([A-Z].*)$", text)
    return f"{match.group(1)}{match.group(2)}" if match else text


def normalize_phase(value: str) -> str:
    """``X2_A`` / ``X2A`` -> ``X2A``; plain ``X2`` stays ``X2`` (a real
    difference, not a spelling one, and the caller must be told)."""
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def read_info(payload, filename: str = "") -> ComplexityMeta:
    """The identity block from the ``Harness PN`` sheet."""
    source = io.BytesIO(payload) if isinstance(payload, (bytes, bytearray)) else payload
    try:
        workbook = load_workbook(source, data_only=True, read_only=True)
    except Exception as exc:
        raise SpliceError(
            f"Could not read {filename or 'the complexity file'}: {exc}") from exc
    meta = ComplexityMeta(filename=filename)
    try:
        if INFO_SHEET not in workbook.sheetnames:
            logger.info("%s has no '%s' sheet; identity unknown", filename, INFO_SHEET)
            return meta
        sheet = workbook[INFO_SHEET]
        for row in sheet.iter_rows(min_row=1, max_row=MAX_INFO_ROWS,
                                   max_col=MAX_INFO_COLS, values_only=True):
            for index, cell in enumerate(row):
                field = _LABELS.get(str(cell or "").strip().lower())
                if not field:
                    continue
                value = row[index + 1] if index + 1 < len(row) else None
                if value not in (None, "") and not getattr(meta, field):
                    setattr(meta, field, str(value).strip())
    finally:
        workbook.close()
    return meta


def read_harness_file(payload: bytes,
                      filename: str = "") -> Tuple[Harness, ComplexityMeta]:
    """Build table plus identity, from one individual complexity file."""
    harness = read_complexity(payload, filename)
    meta = read_info(payload, filename)
    if meta.def_id and harness.def_id and \
            re.sub(r"\D", "", meta.def_id) != re.sub(r"\D", "", harness.def_id):
        # The two ids come from different cells; disagreement means the file was
        # hand-edited. The Complexity sheet's id is the one every engine matches
        # on, so it wins — but the discrepancy is worth logging.
        logger.warning("%s: Harness PN id %s disagrees with Complexity id %s",
                       filename, meta.def_id, harness.def_id)
    if meta.harness:
        harness.name = meta.harness   # the file's own name, never the filename
    return harness, meta
