from __future__ import annotations

import os
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from splice.common.errors import SpliceInputError
from splice.common.logging import get_logger
from splice.common.text import (
    normalize_value,
    normalize_cell,
    normalize_match_text as _normalize_match_text,
    split_delimited_values as _split_delimited_values,
    extract_transmittal_number as _extract_transmittal_number,
    extract_bulletin_number as _extract_bulletin_number,
)
from splice.common.validation import ensure_non_empty_upload, require_columns

logger = get_logger(__name__)
from splice.dtcr.matching import (
    DTCR_MATCHING_COLUMNS,
    prepare_dtcr_for_matching as _prepare_dtcr_for_matching,
    match_dtcr_to_harness_family,
)


REQUIRED_COLUMNS = [
    "Device Control Number",
    "Device Name",
    "Suffix",
    "CNUM",
    "Number of Cavities",
    "Connector PN",
    "Harness Family",
    "Pin Number",
    "Circuit Name",
    "Circuit Suffix",
    "Circuit Function",
    "Color",
    "Terminal",
    "Connector FCA part number",
    "Wire Gauge",
    "Wire Type",
    "Sales Code",
]

KEY_COLUMNS = ["CNUM", "Pin Number"]
COMPARISON_COLUMNS = [column for column in REQUIRED_COLUMNS if column not in KEY_COLUMNS]
MAX_HEADER_SCAN_ROWS = 30
STATUS_COLORS = {
    "Added": "#C6EFCE",
    "Removed": "#FFC7CE",
    "Modified": "#FFEB9C",
    "Unchanged": "#D9D9D9",
    "New CNUM": "#A9D08E",
    "Removed CNUM": "#F1948A",
    "Added Circuit": "#C6EFCE",
    "Removed Circuit": "#FFC7CE",
}

UNASSIGNED_FAMILY = "(No Harness Family)"

ALL_CHANGES_COLUMNS = [
    "DTCR#",
    "Harness Family",
    "Change Type",
    "Device Control Number",
    "Connector PN",
    "Device Name",
    "CNUM",
    "Pin Number",
    "Circuit Name",
    "Circuit Function",
    "Wire Gauge",
    "Wire Type",
    "Sales Code",
    "Changed Fields",
    "Change Detail",
]

# Comparison columns shown directly in All Changes: a modification in one of these
# is rendered inline in its own cell as "old --> new". Anything else goes to Change Detail.
ALL_CHANGES_INLINE_COLUMNS = [
    "Harness Family",
    "Device Control Number",
    "Connector PN",
    "Device Name",
    "Circuit Name",
    "Circuit Function",
    "Wire Gauge",
    "Wire Type",
    "Sales Code",
]

CHANGE_TYPE_ORDER = {
    "New CNUM": 0,
    "Removed CNUM": 1,
    "Added Circuit": 2,
    "Removed Circuit": 3,
    "Modified": 4,
}


def _build_dtcr_lookup_by_cnum(
    old_df: pd.DataFrame,
    new_df: pd.DataFrame,
    dtcr_df: pd.DataFrame,
) -> dict[str, list[str]]:
    required_columns = {"DTCR#", "Device Transmittal"}
    if not required_columns.issubset(set(dtcr_df.columns)):
        raise ValueError("DTCR report must include DTCR# and Device Transmittal columns.")

    dcn_to_dtcr: dict[str, list[str]] = {}
    for _, row in dtcr_df.iterrows():
        dtcr_number = normalize_value(row.get("DTCR#"))
        dcn = _extract_transmittal_number(row.get("Device Transmittal", ""))
        if not dtcr_number or not dcn:
            continue
        if dcn not in dcn_to_dtcr:
            dcn_to_dtcr[dcn] = []
        if dtcr_number not in dcn_to_dtcr[dcn]:
            dcn_to_dtcr[dcn].append(dtcr_number)

    cnum_to_dtcr: dict[str, list[str]] = {}
    combined = pd.concat([old_df, new_df], ignore_index=True)
    for _, row in combined.iterrows():
        cnum = normalize_value(row.get("CNUM", ""))
        if not cnum:
            continue
        dcn_values = _split_delimited_values(row.get("Device Control Number", ""))
        if cnum not in cnum_to_dtcr:
            cnum_to_dtcr[cnum] = []
        for dcn in dcn_values:
            for dtcr_number in dcn_to_dtcr.get(dcn, []):
                if dtcr_number not in cnum_to_dtcr[cnum]:
                    cnum_to_dtcr[cnum].append(dtcr_number)

    return cnum_to_dtcr


def _insert_dtcr_column(data_frame: pd.DataFrame, values: list[str]) -> pd.DataFrame:
    annotated = data_frame.copy()
    if "DTCR#" in annotated.columns:
        annotated = annotated.drop(columns=["DTCR#"])

    insert_at = 1 if "CNUM" in annotated.columns else len(annotated.columns)
    annotated.insert(insert_at, "DTCR#", values)
    return annotated


def _match_dtcr_numbers_for_row(row: pd.Series, dtcr_by_cnum: dict[str, list[str]]) -> str:
    matched: list[str] = []

    cnum_values = _split_delimited_values(row.get("CNUM", ""))
    for cnum in cnum_values:
        for dtcr_number in dtcr_by_cnum.get(cnum, []):
            if dtcr_number not in matched:
                matched.append(dtcr_number)

    return ", ".join(matched)


def _annotate_results_with_dtcr(
    results: dict[str, object],
    dtcr_by_cnum: dict[str, list[str]],
) -> dict[str, object]:
    tables_to_annotate = [
        "added_cnums_df",
        "removed_cnums_df",
        "added_circuits_df",
        "removed_circuits_df",
        "modified_circuits_df",
        "change_log_df",
        "cnum_summary_df",
    ]

    for table_name in tables_to_annotate:
        data_frame = results.get(table_name)
        if not isinstance(data_frame, pd.DataFrame):
            continue
        if data_frame.empty:
            results[table_name] = _insert_dtcr_column(data_frame, [])
            continue

        dtcr_values = [
            _match_dtcr_numbers_for_row(row, dtcr_by_cnum)
            for _, row in data_frame.iterrows()
        ]
        results[table_name] = _insert_dtcr_column(data_frame, dtcr_values)

    if isinstance(results.get("top_20_cnums_df"), pd.DataFrame):
        top_df = results["top_20_cnums_df"]
        if top_df.empty:
            results["top_20_cnums_df"] = _insert_dtcr_column(top_df, [])
        else:
            top_dtcr_values = [
                _match_dtcr_numbers_for_row(row, dtcr_by_cnum)
                for _, row in top_df.iterrows()
            ]
            results["top_20_cnums_df"] = _insert_dtcr_column(top_df, top_dtcr_values)

    return results


def get_preorder_generation_exe_path() -> Path:
    """Resolve the PreOrder tool path (env override or bundled assets fallback)."""
    from splice.config import PREORDER_EXE_PATH

    return PREORDER_EXE_PATH


@dataclass(frozen=True)
class WorkbookLayout:
    sheet_name: str
    header_row: int


def collapse_values(values: pd.Series) -> str:
    unique_values = []
    seen = set()
    for value in values.tolist():
        normalized = normalize_value(value)
        if normalized and normalized not in seen:
            seen.add(normalized)
            unique_values.append(normalized)
    return " | ".join(unique_values)


def excel_column_name(column_index: int) -> str:
    name = ""
    current = column_index + 1
    while current:
        current, remainder = divmod(current - 1, 26)
        name = chr(65 + remainder) + name
    return name


def detect_layout(file_bytes: bytes, file_name: str) -> WorkbookLayout:
    excel_file = pd.ExcelFile(BytesIO(file_bytes))
    required = set(REQUIRED_COLUMNS)

    for sheet_name in excel_file.sheet_names:
        preview = pd.read_excel(
            BytesIO(file_bytes),
            sheet_name=sheet_name,
            header=None,
            nrows=MAX_HEADER_SCAN_ROWS,
        )
        for header_row, row in preview.iterrows():
            row_values = {normalize_value(value) for value in row.tolist() if normalize_value(value)}
            if required.issubset(row_values):
                return WorkbookLayout(sheet_name=sheet_name, header_row=header_row)

    raise ValueError(
        f"Could not find the DTx header row in {file_name}. Expected columns: {', '.join(REQUIRED_COLUMNS)}"
    )


def load_dtx_report(file_bytes: bytes, file_name: str) -> tuple[pd.DataFrame, WorkbookLayout]:
    ensure_non_empty_upload(file_bytes, name=f"DTx report '{file_name}'")
    layout = detect_layout(file_bytes, file_name)
    try:
        data_frame = pd.read_excel(
            BytesIO(file_bytes),
            sheet_name=layout.sheet_name,
            header=layout.header_row,
            dtype=object,
        )
    except Exception as exc:  # pragma: no cover - pandas/openpyxl parse errors vary
        logger.warning("Failed to parse DTx report '%s': %s", file_name, exc)
        raise SpliceInputError(f"Could not read the DTx report '{file_name}': {exc}") from exc
    data_frame.columns = [normalize_value(column) for column in data_frame.columns]

    require_columns(data_frame, REQUIRED_COLUMNS, context=f"DTx report '{file_name}'")

    data_frame = data_frame[REQUIRED_COLUMNS].copy()
    data_frame = data_frame.map(normalize_cell)
    data_frame = data_frame.loc[
        ~((data_frame["CNUM"] == "") & (data_frame["Pin Number"] == ""))
    ].reset_index(drop=True)

    aggregation = {column: collapse_values for column in COMPARISON_COLUMNS}
    data_frame = (
        data_frame.groupby(KEY_COLUMNS, dropna=False, as_index=False)
        .agg(aggregation)
        .sort_values(KEY_COLUMNS)
        .reset_index(drop=True)
    )

    return data_frame, layout


def load_dtcr_report(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    """Load a DTCR search report, auto-detecting the header row."""
    ensure_non_empty_upload(file_bytes, name=f"DTCR report '{file_name}'")
    required = {"DTCR#", "Device Transmittal"}
    try:
        excel_file = pd.ExcelFile(BytesIO(file_bytes))
    except Exception as exc:  # pragma: no cover - pandas/openpyxl parse errors vary
        logger.warning("Failed to open DTCR report '%s': %s", file_name, exc)
        raise SpliceInputError(f"Could not read the DTCR report '{file_name}': {exc}") from exc

    for sheet_name in excel_file.sheet_names:
        preview = pd.read_excel(
            BytesIO(file_bytes),
            sheet_name=sheet_name,
            header=None,
            nrows=MAX_HEADER_SCAN_ROWS,
        )
        for header_row, row in preview.iterrows():
            row_values = {normalize_value(value) for value in row.tolist() if normalize_value(value)}
            if required.issubset(row_values):
                data_frame = pd.read_excel(
                    BytesIO(file_bytes),
                    sheet_name=sheet_name,
                    header=header_row,
                    dtype=object,
                )
                data_frame.columns = [normalize_value(column) for column in data_frame.columns]
                return data_frame

    raise ValueError(
        f"Could not find the DTCR header row in {file_name}. Expected columns: DTCR#, Device Transmittal."
    )


# ---------------------------------------------------------------------------
# DTCR Matching Report (same format as SECR Step 1, built from BOTH DTx files)
# ---------------------------------------------------------------------------

def _build_combined_dtx_frame(old_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    """Combine OLD + NEW DTx rows (NEW first) into one dedup'd matching frame."""
    columns = ["Device Control Number", "Device Name", "Suffix", "Harness Family", "CNUM"]
    combined = pd.concat([new_df[columns], old_df[columns]], ignore_index=True)
    combined = combined.map(normalize_cell).astype(str)
    return combined.drop_duplicates().reset_index(drop=True)


def export_dtcr_matching_report(mapping_df: pd.DataFrame) -> bytes:
    """Export the DTCR mapping as a styled standalone workbook (same format as SECR Step 1)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "DTCR_Harness_Family_Mapping"

    for row_index, row in enumerate([mapping_df.columns.tolist()] + mapping_df.values.tolist(), 1):
        for column_index, value in enumerate(row, 1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    stripe_fill = PatternFill(start_color="EAF2FA", end_color="EAF2FA", fill_type="solid")
    for row_index in range(2, worksheet.max_row + 1):
        for column_index in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index % 2 == 0:
                cell.fill = stripe_fill

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for row_index in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 120)

    for row_index in range(2, worksheet.max_row + 1):
        max_lines = 1
        for column_index in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value is not None:
                max_lines = max(max_lines, str(value).count("\n") + 1)
        worksheet.row_dimensions[row_index].height = min(max(max_lines * 15, 18), 120)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_dtcr_matching_report(
    old_file_bytes: bytes,
    new_file_bytes: bytes,
    old_file_name: str,
    new_file_name: str,
    dtcr_df: pd.DataFrame,
) -> dict[str, object]:
    """Build the DTCR_Matching_Report using BOTH DTx reports (old + new)."""
    old_rows = _read_dtx_report_rows(old_file_bytes, old_file_name)
    new_rows = _read_dtx_report_rows(new_file_bytes, new_file_name)
    combined = _build_combined_dtx_frame(old_rows, new_rows)
    mapping_df = match_dtcr_to_harness_family(dtcr_df, combined)
    return {
        "dtcr_matching_df": mapping_df,
        "dtcr_matching_bytes": export_dtcr_matching_report(mapping_df),
        "dtcr_matching_file_name": "DTCR_Matching_Report.xlsx",
    }


def _all_changes_record(row: pd.Series, change_type: str) -> dict[str, object]:
    return {
        "DTCR#": normalize_value(row.get("DTCR#", "")),
        "Harness Family": normalize_value(row.get("Harness Family", "")) or UNASSIGNED_FAMILY,
        "Change Type": change_type,
        "Device Control Number": row.get("Device Control Number", ""),
        "Connector PN": row.get("Connector PN", ""),
        "Device Name": row.get("Device Name", ""),
        "CNUM": row.get("CNUM", ""),
        "Pin Number": row.get("Pin Number", ""),
        "Circuit Name": row.get("Circuit Name", ""),
        "Circuit Function": row.get("Circuit Function", ""),
        "Wire Gauge": row.get("Wire Gauge", ""),
        "Wire Type": row.get("Wire Type", ""),
        "Sales Code": row.get("Sales Code", ""),
        "Changed Fields": "",
        "Change Detail": "",
    }


def _pick_new_or_old(row: pd.Series, column: str) -> str:
    return normalize_value(row.get(f"{column}_new", "")) or normalize_value(row.get(f"{column}_old", ""))


def build_all_changes_df(results: dict[str, object]) -> pd.DataFrame:
    """Combine every change (CNUM-level, circuit-level, modifications) into one table."""
    records: list[dict[str, object]] = []

    for table_name, change_type in (
        ("added_cnums_df", "New CNUM"),
        ("removed_cnums_df", "Removed CNUM"),
        ("added_circuits_df", "Added Circuit"),
        ("removed_circuits_df", "Removed Circuit"),
    ):
        data_frame = results.get(table_name)
        if isinstance(data_frame, pd.DataFrame) and not data_frame.empty:
            for _, row in data_frame.iterrows():
                records.append(_all_changes_record(row, change_type))

    modified_df = results.get("modified_circuits_df")
    if isinstance(modified_df, pd.DataFrame) and not modified_df.empty:
        for _, row in modified_df.iterrows():
            changed_fields = normalize_value(row.get("Changed Fields", ""))
            changed_set = {field.strip() for field in changed_fields.split(",") if field.strip()}

            record: dict[str, object] = {
                "DTCR#": normalize_value(row.get("DTCR#", "")),
                "Change Type": "Modified",
                "CNUM": row.get("CNUM", ""),
                "Pin Number": row.get("Pin Number", ""),
                "Changed Fields": changed_fields,
            }

            # Displayed comparison columns: show the change inline as "old --> new".
            for column in ALL_CHANGES_INLINE_COLUMNS:
                old_value = normalize_value(row.get(f"{column}_old", ""))
                new_value = normalize_value(row.get(f"{column}_new", ""))
                if column in changed_set:
                    record[column] = f"{old_value or '(blank)'} --> {new_value or '(blank)'}"
                else:
                    record[column] = new_value or old_value
            if not record["Harness Family"]:
                record["Harness Family"] = UNASSIGNED_FAMILY

            # Everything else that changed goes to Change Detail.
            details: list[str] = []
            for column in sorted(changed_set - set(ALL_CHANGES_INLINE_COLUMNS)):
                old_value = normalize_value(row.get(f"{column}_old", ""))
                new_value = normalize_value(row.get(f"{column}_new", ""))
                details.append(f"{column}: {old_value or '(blank)'} --> {new_value or '(blank)'}")
            record["Change Detail"] = "; ".join(details)

            records.append(record)

    all_changes = pd.DataFrame(records, columns=ALL_CHANGES_COLUMNS)
    if all_changes.empty:
        return all_changes

    all_changes["_type_order"] = all_changes["Change Type"].map(CHANGE_TYPE_ORDER).fillna(9)
    all_changes = (
        all_changes.sort_values(["Harness Family", "_type_order", "CNUM", "Pin Number"])
        .drop(columns=["_type_order"])
        .reset_index(drop=True)
    )
    return all_changes


FAMILY_SUMMARY_COLUMNS = [
    "Harness Family",
    "New CNUM Circuits",
    "Removed CNUM Circuits",
    "Added Circuits",
    "Removed Circuits",
    "Modified Circuits",
    "Total Changes",
    "Affected CNUMs",
    "DTCR#s",
]


def build_family_summary_df(all_changes_df: pd.DataFrame) -> pd.DataFrame:
    """Roll up the All Changes table into one row per Harness Family."""
    if all_changes_df.empty:
        return pd.DataFrame(columns=FAMILY_SUMMARY_COLUMNS)

    label_map = {
        "New CNUM": "New CNUM Circuits",
        "Removed CNUM": "Removed CNUM Circuits",
        "Added Circuit": "Added Circuits",
        "Removed Circuit": "Removed Circuits",
        "Modified": "Modified Circuits",
    }

    rows: list[dict[str, object]] = []
    for family, group in all_changes_df.groupby("Harness Family", dropna=False):
        counts = group["Change Type"].value_counts()
        record: dict[str, object] = {"Harness Family": family}
        for change_type, column_name in label_map.items():
            record[column_name] = int(counts.get(change_type, 0))
        record["Total Changes"] = int(len(group))
        record["Affected CNUMs"] = int(group["CNUM"].nunique())

        dtcr_numbers: list[str] = []
        for value in group["DTCR#"].tolist():
            for token in _split_delimited_values(value):
                if token not in dtcr_numbers:
                    dtcr_numbers.append(token)
        record["DTCR#s"] = ", ".join(dtcr_numbers)
        rows.append(record)

    summary = pd.DataFrame(rows, columns=FAMILY_SUMMARY_COLUMNS)
    return summary.sort_values(
        ["Total Changes", "Harness Family"], ascending=[False, True]
    ).reset_index(drop=True)


def _collapse_to_unique_connector_values(values: pd.Series) -> str:
    unique_values: list[str] = []
    seen: set[str] = set()

    for value in values.tolist():
        normalized = normalize_value(value)
        if normalized and normalized not in seen:
            seen.add(normalized)
            unique_values.append(normalized)

    return " | ".join(unique_values)


def _first_non_empty(values: pd.Series) -> str:
    for value in values.tolist():
        normalized = normalize_value(value)
        if normalized:
            return normalized
    return ""


def _read_dtx_report_rows(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    layout = detect_layout(file_bytes, file_name)
    data_frame = pd.read_excel(
        BytesIO(file_bytes),
        sheet_name=layout.sheet_name,
        header=layout.header_row,
        dtype=object,
    )
    data_frame.columns = [normalize_value(column) for column in data_frame.columns]

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in data_frame.columns]
    if missing_columns:
        raise ValueError(f"{file_name} is missing required columns: {', '.join(missing_columns)}")

    data_frame = data_frame[REQUIRED_COLUMNS].copy()
    data_frame = data_frame.map(normalize_cell)
    data_frame = data_frame.loc[
        ~((data_frame["CNUM"] == "") & (data_frame["Pin Number"] == ""))
    ].reset_index(drop=True)
    return data_frame


def _extract_report_metadata(file_bytes: bytes, file_name: str) -> dict[str, str]:
    excel_file = pd.ExcelFile(BytesIO(file_bytes))
    sheet_name = next(
        (sheet for sheet in excel_file.sheet_names if "Detailed DTx Circuits Report" in sheet),
        excel_file.sheet_names[0],
    )
    preview = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None, nrows=12)

    metadata = {"Vehicle Program": "", "Build Phase": "", "Report Date": ""}
    for _, row in preview.iterrows():
        for value in row.tolist():
            normalized = normalize_value(value)
            if normalized.startswith("Vehicle Program -"):
                metadata["Vehicle Program"] = normalized
            elif normalized.startswith("Build Phase -"):
                metadata["Build Phase"] = normalized
            elif normalized.startswith("Report Date:"):
                metadata["Report Date"] = normalized

    if not metadata["Vehicle Program"]:
        metadata["Vehicle Program"] = Path(file_name).stem
    if not metadata["Build Phase"]:
        metadata["Build Phase"] = "Build Phase - Unknown"
    if not metadata["Report Date"]:
        metadata["Report Date"] = datetime.now().strftime("%b-%d-%Y %I:%M %p")

    return metadata


def _build_output_file_name(new_metadata: dict[str, str]) -> str:
    vehicle_program = new_metadata.get("Vehicle Program", "")
    build_phase = new_metadata.get("Build Phase", "")

    def _file_component(raw_value: str, prefix: str) -> str:
        cleaned = normalize_value(raw_value)
        if cleaned.startswith(prefix):
            cleaned = cleaned[len(prefix):].strip()
        cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", cleaned).strip("._-")
        return cleaned or "Unknown"

    vehicle_component = _file_component(vehicle_program, "Vehicle Program -")
    build_component = _file_component(build_phase, "Build Phase -")
    date_component = datetime.now().strftime("%m%d%y")
    return f"PreOrderList_{vehicle_component}_{build_component}_{date_component}.xlsx"


def _build_connector_grouped_frame(data_frame: pd.DataFrame) -> pd.DataFrame:
    connector_columns = [
        "CNUM",
        "Device Control Number",
        "Device Name",
        "Suffix",
        "Number of Cavities",
        "Connector PN",
        "Harness Family",
    ]
    grouped_frame = data_frame[connector_columns].copy()
    grouped_frame = grouped_frame.map(normalize_cell)

    aggregation = {
        "Device Control Number": _first_non_empty,
        "Device Name": _collapse_to_unique_connector_values,
        "Suffix": _collapse_to_unique_connector_values,
        "Number of Cavities": _first_non_empty,
        "Connector PN": _collapse_to_unique_connector_values,
    }
    grouped_frame = (
        grouped_frame.groupby(["CNUM", "Harness Family"], dropna=False, as_index=False)
        .agg(aggregation)
        .sort_values(["Harness Family", "CNUM"])
        .reset_index(drop=True)
    )
    return grouped_frame


def _apply_preorder_workbook_styles(workbook_bytes: bytes) -> bytes:
    workbook = load_workbook(BytesIO(workbook_bytes))
    header_fill = PatternFill(fill_type="solid", fgColor="004472C4")
    yellow_fill = PatternFill(fill_type="solid", fgColor="00FFFF00")
    deleted_fill = PatternFill(fill_type="solid", fgColor="00E6E6FA")
    added_fill = PatternFill(fill_type="solid", fgColor="00FFD580")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    def build_border(column: int, is_header: bool) -> Border:
        thin_side = Side(style="thin", color="000000")
        thick_side = Side(style="thick", color="000000")
        if worksheet.max_column == 4:
            block_starts = {1}
            block_ends = {4}
        elif worksheet.max_column == 11:
            block_starts = {1, 5, 8}
            block_ends = {4, 7, 11}
        else:
            block_starts = {1}
            block_ends = {worksheet.max_column}

        left = thick_side if column in block_starts else thin_side
        right = thick_side if column in block_ends else thin_side
        top = thick_side if is_header else thin_side
        bottom = thin_side if is_header else thin_side
        return Border(left=left, right=right, top=top, bottom=bottom)

    for worksheet in workbook.worksheets:
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        for row in range(6, max_row + 1):
            for column in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=column)
                if row == 6:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = build_border(column, is_header=True)
                else:
                    cell.border = build_border(column, is_header=False)
                    if worksheet.title == "Connector Changes" and column == 5:
                        change_type = worksheet.cell(row=row, column=7).value
                        if change_type == "Deleted":
                            cell.fill = deleted_fill
                        elif change_type == "Added":
                            cell.fill = added_fill
                        elif isinstance(cell.value, str) and ">>" in cell.value:
                            cell.fill = yellow_fill
                    elif worksheet.title == "Summary" and column == 2:
                        change_type = worksheet.cell(row=row, column=4).value
                        if change_type == "Deleted":
                            cell.fill = deleted_fill
                        elif change_type == "Added":
                            cell.fill = added_fill
                        elif isinstance(cell.value, str) and ">>" in cell.value:
                            cell.fill = yellow_fill

        worksheet.freeze_panes = "A7"
        if worksheet.title == "Connector Changes":
            widths = {"A": 42, "B": 14, "C": 18, "D": 30, "E": 42, "F": 24, "G": 22,
                      "H": 42, "I": 14, "J": 18, "K": 30}
        else:
            widths = {"A": 14, "B": 42, "C": 24, "D": 22}
        for column_letter, width in widths.items():
            worksheet.column_dimensions[column_letter].width = width
        worksheet.row_dimensions[6].height = 34
        for cell in worksheet[6]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    output_buffer = BytesIO()
    workbook.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer.getvalue()


def generate_preorder_generation_workbook(
    old_file_bytes: bytes,
    new_file_bytes: bytes,
    old_file_name: str,
    new_file_name: str,
) -> dict[str, object]:
    old_df = _read_dtx_report_rows(old_file_bytes, old_file_name)
    new_df = _read_dtx_report_rows(new_file_bytes, new_file_name)

    old_grouped = _build_connector_grouped_frame(old_df)
    new_grouped = _build_connector_grouped_frame(new_df)

    # A physical connector is identified by CNUM within its harness family.
    # Device Name and Suffix are descriptive and may legitimately change between
    # releases; using them as keys creates false Deleted + Added pairs.
    key_columns = ["CNUM", "Harness Family"]
    old_grouped_keys = old_grouped[key_columns].drop_duplicates()
    new_grouped_keys = new_grouped[key_columns].drop_duplicates()

    merged = old_grouped.merge(
        new_grouped,
        on=key_columns,
        how="inner",
        suffixes=("_old", "_new"),
    )
    connector_pn_mask = (
        merged["Connector PN_old"].fillna("").astype(str).str.strip()
        != merged["Connector PN_new"].fillna("").astype(str).str.strip()
    )
    changed_connector_pn_df = merged.loc[connector_pn_mask, [
        "CNUM",
        "Device Control Number_old",
        "Device Name_old",
        "Suffix_old",
        "Number of Cavities_old",
        "Harness Family",
        "Connector PN_old",
        "Device Control Number_new",
        "Device Name_new",
        "Suffix_new",
        "Number of Cavities_new",
        "Connector PN_new",
    ]].copy()
    changed_connector_pn_df.insert(0, "Change Type", "Connector PN Change")

    deleted_df = (
        old_grouped.merge(
            new_grouped_keys,
            on=key_columns,
            how="left",
            indicator=True,
        )
        .query("_merge == 'left_only'")
        .reset_index(drop=True)
    )
    deleted_df.insert(0, "Change Type", "Deleted")

    added_df = (
        new_grouped.merge(
            old_grouped_keys,
            on=key_columns,
            how="left",
            indicator=True,
        )
        .query("_merge == 'left_only'")
        .reset_index(drop=True)
    )
    added_df.insert(0, "Change Type", "Added")

    old_metadata = _extract_report_metadata(old_file_bytes, old_file_name)
    new_metadata = _extract_report_metadata(new_file_bytes, new_file_name)

    connector_changes_columns = [
        "CNUM_Device Name-Suffix (Device Control Number)",
        "CNUM",
        "Number of Cavities",
        "Connector PN",
        "Connector PN Change",
        "Harness Family",
        "Change Type",
        "CNUM_Device Name-Suffix (Device Control Number)",
        "CNUM",
        "Number of Cavities",
        "Connector PN",
    ]
    summary_columns = ["CNUM", "Connector PN Change", "Harness Family", "Change Type"]

    connector_changes_rows: list[list[object]] = []
    summary_rows: list[list[object]] = []

    for _, row in changed_connector_pn_df.iterrows():
        old_identifier = (
            f"{row['CNUM']}_{row['Device Name_old']}-{row['Suffix_old']} ({row['Device Control Number_old']})"
        )
        new_identifier = (
            f"{row['CNUM']}_{row['Device Name_new']}-{row['Suffix_new']} ({row['Device Control Number_new']})"
        )
        old_connector_pn = normalize_value(row["Connector PN_old"])
        new_connector_pn = normalize_value(row["Connector PN_new"])
        connector_changes_rows.append(
            [
                old_identifier,
                row["CNUM"],
                row["Number of Cavities_old"],
                old_connector_pn,
                f"{old_connector_pn} >> {new_connector_pn}",
                row["Harness Family"],
                "Connector PN Change",
                new_identifier,
                row["CNUM"],
                row["Number of Cavities_new"],
                new_connector_pn,
            ]
        )
        summary_rows.append(
            [
                row["CNUM"],
                f"{old_connector_pn} >> {new_connector_pn}",
                row["Harness Family"],
                "Connector PN Change",
            ]
        )

    for _, row in deleted_df.iterrows():
        old_identifier = (
            f"{row['CNUM']}_{row['Device Name']}-{row['Suffix']} ({row['Device Control Number']})"
        )
        old_connector_pn = normalize_value(row["Connector PN"])
        connector_changes_rows.append(
            [
                old_identifier,
                row["CNUM"],
                row["Number of Cavities"],
                old_connector_pn,
                f"{old_connector_pn} >>",
                row["Harness Family"],
                "Deleted",
                None,
                None,
                None,
                None,
            ]
        )
        summary_rows.append(
            [
                row["CNUM"],
                f"{old_connector_pn} >>",
                row["Harness Family"],
                "Deleted",
            ]
        )

    for _, row in added_df.iterrows():
        new_identifier = (
            f"{row['CNUM']}_{row['Device Name']}-{row['Suffix']} ({row['Device Control Number']})"
        )
        new_connector_pn = normalize_value(row["Connector PN"])
        connector_changes_rows.append(
            [
                None,
                None,
                None,
                None,
                f">> {new_connector_pn}",
                row["Harness Family"],
                "Added",
                new_identifier,
                row["CNUM"],
                row["Number of Cavities"],
                new_connector_pn,
            ]
        )
        summary_rows.append(
            [
                row["CNUM"],
                f">> {new_connector_pn}",
                row["Harness Family"],
                "Added",
            ]
        )

    connector_changes_df = pd.DataFrame(connector_changes_rows, columns=connector_changes_columns)
    summary_df = pd.DataFrame(summary_rows, columns=summary_columns)

    connector_metadata_rows = [
        [
            "Detailed DTx Circuits Report",
            None,
            None,
            None,
            f"{old_metadata['Vehicle Program']} >> {new_metadata['Vehicle Program']}",
            None,
            None,
            "Detailed DTx Circuits Report",
            None,
            None,
            None,
        ],
        [
            old_metadata["Vehicle Program"],
            None,
            None,
            None,
            f"{old_metadata['Build Phase']} >> {new_metadata['Build Phase']}",
            None,
            None,
            new_metadata["Vehicle Program"],
            None,
            None,
            None,
        ],
        [
            old_metadata["Build Phase"],
            None,
            None,
            None,
            f"{old_metadata['Report Date']} >> {new_metadata['Report Date']}",
            None,
            None,
            new_metadata["Build Phase"],
            None,
            None,
            None,
        ],
        [
            old_metadata["Report Date"],
            None,
            None,
            None,
            "Old >> New",
            None,
            None,
            new_metadata["Report Date"],
            None,
            None,
            None,
        ],
        [None, None, None, None, None, None, None, None, None, None, None],
    ]
    summary_metadata_rows = [
        [
            f"{old_metadata['Vehicle Program']} >> {new_metadata['Vehicle Program']}",
            None,
            None,
            None,
        ],
        [
            f"{old_metadata['Build Phase']} >> {new_metadata['Build Phase']}",
            None,
            None,
            None,
        ],
        [
            f"{old_metadata['Report Date']} >> {new_metadata['Report Date']}",
            None,
            None,
            None,
        ],
        ["Old >> New", None, None, None],
        [None, None, None, None],
    ]

    output_buffer = BytesIO()
    with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
        pd.DataFrame(connector_metadata_rows + [connector_changes_columns] + connector_changes_df.values.tolist()).to_excel(
            writer,
            sheet_name="Connector Changes",
            header=False,
            index=False,
        )
        pd.DataFrame(summary_metadata_rows + [summary_columns] + summary_df.values.tolist()).to_excel(
            writer,
            sheet_name="Summary",
            header=False,
            index=False,
        )

    output_buffer.seek(0)
    styled_bytes = _apply_preorder_workbook_styles(output_buffer.getvalue())
    return {
        "output_excel_bytes": styled_bytes,
        "output_file_name": _build_output_file_name(new_metadata),
        "summary_df": summary_df,
        "connector_changes_df": connector_changes_df,
        "suffix_changes_df": pd.DataFrame(),
        "deleted_df": deleted_df,
        "added_df": added_df,
        "changed_connector_pn_df": changed_connector_pn_df,
    }


def launch_preorder_generation_tool(
    old_file_path: Path | str | None = None,
    new_file_path: Path | str | None = None,
) -> dict[str, object]:
    if old_file_path is None or new_file_path is None:
        raise ValueError("Both old and new DTx files are required to generate the PreOrder workbook.")

    old_path = Path(old_file_path)
    new_path = Path(new_file_path)
    return generate_preorder_generation_workbook(
        old_file_bytes=old_path.read_bytes(),
        new_file_bytes=new_path.read_bytes(),
        old_file_name=old_path.name,
        new_file_name=new_path.name,
    )


def build_modified_views(
    old_existing: pd.DataFrame,
    new_existing: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, Counter[str]]:
    if old_existing.empty or new_existing.empty:
        return pd.DataFrame(), pd.DataFrame(), Counter()

    old_indexed = old_existing.set_index(KEY_COLUMNS)
    new_indexed = new_existing.set_index(KEY_COLUMNS)
    shared_index = old_indexed.index.intersection(new_indexed.index)

    modified_rows: list[dict[str, object]] = []
    change_log_rows: list[dict[str, object]] = []
    field_counter: Counter[str] = Counter()

    for key in shared_index:
        old_row = old_indexed.loc[key]
        new_row = new_indexed.loc[key]
        changed_fields: list[str] = []

        old_harness_family = normalize_cell(old_row.get("Harness Family", ""))
        new_harness_family = normalize_cell(new_row.get("Harness Family", ""))
        harness_family_value = (
            old_harness_family
            if old_harness_family == new_harness_family
            else f"{old_harness_family} >> {new_harness_family}"
        )

        pair_values: dict[str, object] = {}
        for column in COMPARISON_COLUMNS:
            old_value = normalize_cell(old_row[column])
            new_value = normalize_cell(new_row[column])
            pair_values[f"{column}_old"] = old_value
            pair_values[f"{column}_new"] = new_value

            if old_value != new_value:
                changed_fields.append(column)
                field_counter[column] += 1
                change_log_rows.append(
                    {
                        "CNUM": key[0],
                        "Pin Number": key[1],
                        "Change Type": "Modified",
                        "Harness Family": harness_family_value,
                        "Field": column,
                        "Old Value": old_value,
                        "New Value": new_value,
                    }
                )

        if changed_fields:
            row_record: dict[str, object] = {
                "CNUM": key[0],
                "Pin Number": key[1],
                "Change Type": "Modified",
                "Harness Family": harness_family_value,
                "Changed Fields": ", ".join(changed_fields),
                "Change Count": len(changed_fields),
                **pair_values,
            }
            modified_rows.append(row_record)

    modified_df = pd.DataFrame(modified_rows)
    change_log_df = pd.DataFrame(change_log_rows)

    if not modified_df.empty:
        modified_df = modified_df.sort_values(["CNUM", "Pin Number"]).reset_index(drop=True)
    if not change_log_df.empty:
        change_log_df = change_log_df.sort_values(["CNUM", "Pin Number", "Field"]).reset_index(drop=True)

    return modified_df, change_log_df, field_counter


def compare_reports(old_df: pd.DataFrame, new_df: pd.DataFrame) -> dict[str, object]:
    old_cnums = set(old_df["CNUM"])
    new_cnums = set(new_df["CNUM"])
    shared_cnums = old_cnums & new_cnums

    added_cnums = sorted(new_cnums - old_cnums)
    removed_cnums = sorted(old_cnums - new_cnums)

    added_cnums_df = new_df[new_df["CNUM"].isin(added_cnums)].copy()
    removed_cnums_df = old_df[old_df["CNUM"].isin(removed_cnums)].copy()
    if not added_cnums_df.empty:
        added_cnums_df.insert(0, "Change Type", "Added")
    if not removed_cnums_df.empty:
        removed_cnums_df.insert(0, "Change Type", "Removed")

    old_existing = old_df[old_df["CNUM"].isin(shared_cnums)].copy()
    new_existing = new_df[new_df["CNUM"].isin(shared_cnums)].copy()

    old_existing_indexed = old_existing.set_index(KEY_COLUMNS)
    new_existing_indexed = new_existing.set_index(KEY_COLUMNS)

    added_circuit_index = new_existing_indexed.index.difference(old_existing_indexed.index)
    removed_circuit_index = old_existing_indexed.index.difference(new_existing_indexed.index)
    unchanged_index = []

    for key in old_existing_indexed.index.intersection(new_existing_indexed.index):
        if all(
            normalize_cell(old_existing_indexed.at[key, column])
            == normalize_cell(new_existing_indexed.at[key, column])
            for column in COMPARISON_COLUMNS
        ):
            unchanged_index.append(key)

    added_circuits_df = (
        new_existing_indexed.loc[added_circuit_index].reset_index().sort_values(KEY_COLUMNS)
        if len(added_circuit_index)
        else pd.DataFrame(columns=REQUIRED_COLUMNS)
    )
    removed_circuits_df = (
        old_existing_indexed.loc[removed_circuit_index].reset_index().sort_values(KEY_COLUMNS)
        if len(removed_circuit_index)
        else pd.DataFrame(columns=REQUIRED_COLUMNS)
    )

    if not added_circuits_df.empty:
        added_circuits_df.insert(0, "Change Type", "Added")
    if not removed_circuits_df.empty:
        removed_circuits_df.insert(0, "Change Type", "Removed")

    modified_circuits_df, change_log_df, field_counter = build_modified_views(old_existing, new_existing)

    cnum_summary = pd.DataFrame({"CNUM": sorted(shared_cnums)})

    def count_by_cnum(data_frame: pd.DataFrame, column_name: str) -> pd.DataFrame:
        if data_frame.empty:
            return pd.DataFrame(columns=["CNUM", column_name])
        return data_frame.groupby("CNUM").size().rename(column_name).reset_index()

    cnum_summary = cnum_summary.merge(
        count_by_cnum(added_circuits_df, "Added Circuits"),
        on="CNUM",
        how="left",
    )
    cnum_summary = cnum_summary.merge(
        count_by_cnum(removed_circuits_df, "Removed Circuits"),
        on="CNUM",
        how="left",
    )
    cnum_summary = cnum_summary.merge(
        count_by_cnum(modified_circuits_df, "Modified Circuits"),
        on="CNUM",
        how="left",
    )
    cnum_summary = cnum_summary.fillna(0)
    for column in ["Added Circuits", "Removed Circuits", "Modified Circuits"]:
        cnum_summary[column] = cnum_summary[column].astype(int)

    cnum_to_family: dict[str, str] = {}
    for source_df in (new_df, old_df):
        pairs = source_df[["CNUM", "Harness Family"]].drop_duplicates()
        for cnum, family in pairs.itertuples(index=False):
            cnum_key = normalize_value(cnum)
            family_value = normalize_value(family)
            if cnum_key and family_value and cnum_key not in cnum_to_family:
                cnum_to_family[cnum_key] = family_value
    cnum_summary.insert(
        1,
        "Harness Family",
        [cnum_to_family.get(normalize_value(cnum), "") for cnum in cnum_summary["CNUM"]],
    )

    cnum_summary["Total Changes"] = (
        cnum_summary["Added Circuits"]
        + cnum_summary["Removed Circuits"]
        + cnum_summary["Modified Circuits"]
    )
    cnum_summary = cnum_summary.sort_values(
        ["Total Changes", "Modified Circuits", "Added Circuits", "Removed Circuits", "CNUM"],
        ascending=[False, False, False, False, True],
    ).reset_index(drop=True)

    field_change_frequency = pd.DataFrame(
        sorted(field_counter.items(), key=lambda item: (-item[1], item[0])),
        columns=["Field Name", "Number of Changes"],
    )

    return {
        "old_total_cnums": len(old_cnums),
        "new_total_cnums": len(new_cnums),
        "added_cnum_count": len(added_cnums),
        "removed_cnum_count": len(removed_cnums),
        "old_total_circuits": len(old_df),
        "new_total_circuits": len(new_df),
        "added_circuit_count": len(added_circuits_df),
        "removed_circuit_count": len(removed_circuits_df),
        "modified_circuit_count": len(modified_circuits_df),
        "unchanged_circuit_count": len(unchanged_index),
        "added_cnums_df": added_cnums_df.sort_values(KEY_COLUMNS).reset_index(drop=True),
        "removed_cnums_df": removed_cnums_df.sort_values(KEY_COLUMNS).reset_index(drop=True),
        "added_circuits_df": added_circuits_df.reset_index(drop=True),
        "removed_circuits_df": removed_circuits_df.reset_index(drop=True),
        "modified_circuits_df": modified_circuits_df,
        "change_log_df": change_log_df,
        "cnum_summary_df": cnum_summary,
        "field_change_frequency_df": field_change_frequency,
        "top_20_cnums_df": cnum_summary.head(20).copy(),
    }


def write_table(
    writer: pd.ExcelWriter,
    sheet_name: str,
    data_frame: pd.DataFrame,
    workbook,
    formats: dict[str, object],
    compare_pairs: Iterable[tuple[int, int]] | None = None,
) -> None:
    data_frame.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    row_count, column_count = data_frame.shape

    worksheet.freeze_panes(1, 0)
    if column_count:
        worksheet.autofilter(0, 0, max(row_count, 1), column_count - 1)
    worksheet.set_row(0, None, formats["header"])

    for column_index, column_name in enumerate(data_frame.columns):
        values = [column_name] + [normalize_value(value) for value in data_frame[column_name].tolist()]
        width = min(max(len(value) for value in values) + 2, 40)
        worksheet.set_column(column_index, column_index, width)

    if row_count == 0 or column_count == 0:
        return

    if "Change Type" in data_frame.columns:
        change_type_index = data_frame.columns.get_loc("Change Type")
        change_type_column = excel_column_name(change_type_index)
        for label, color in STATUS_COLORS.items():
            worksheet.conditional_format(
                1,
                0,
                row_count,
                column_count - 1,
                {
                    "type": "formula",
                    "criteria": f'=${change_type_column}2="{label}"',
                    "format": workbook.add_format({"bg_color": color}),
                },
            )

    if compare_pairs is not None:
        for old_index, new_index in compare_pairs:
            old_column = excel_column_name(old_index)
            new_column = excel_column_name(new_index)
            worksheet.conditional_format(
                1,
                old_index,
                row_count,
                old_index,
                {
                    "type": "formula",
                    "criteria": f'=${old_column}2<>${new_column}2',
                    "format": formats["modified"],
                },
            )
            worksheet.conditional_format(
                1,
                new_index,
                row_count,
                new_index,
                {
                    "type": "formula",
                    "criteria": f'=${old_column}2<>${new_column}2',
                    "format": formats["modified"],
                },
            )


def build_dashboard_sheet(
    writer: pd.ExcelWriter,
    results: dict[str, object],
    workbook,
    formats: dict[str, object],
    old_name: str = "",
    new_name: str = "",
) -> None:
    dashboard = workbook.add_worksheet("Dashboard")
    writer.sheets["Dashboard"] = dashboard

    dashboard.hide_gridlines(2)
    dashboard.set_column("A:A", 28)
    dashboard.set_column("B:B", 12)
    dashboard.set_column("C:C", 3)
    dashboard.set_column("D:D", 22)
    dashboard.set_column("E:E", 10)
    dashboard.set_column("F:F", 3)
    dashboard.set_column("G:G", 26)
    dashboard.set_column("H:M", 12)

    # --- Title block ---
    dashboard.merge_range("A1:M1", "DTx Engineering Change Report", formats["title"])
    dashboard.write("A2", f"OLD: {old_name}", formats["meta"])
    dashboard.write("A3", f"NEW: {new_name}", formats["meta"])
    dashboard.write("A4", f"Generated: {datetime.now().strftime('%b-%d-%Y %I:%M %p')}", formats["meta"])

    # --- Key metrics (A6:B16) ---
    dashboard.write("A6", "Metric", formats["header"])
    dashboard.write("B6", "Value", formats["header"])
    metrics = [
        ("Total CNUMs (old)", results["old_total_cnums"], formats["default"]),
        ("Total CNUMs (new)", results["new_total_cnums"], formats["default"]),
        ("Added CNUMs", results["added_cnum_count"], formats["added"]),
        ("Removed CNUMs", results["removed_cnum_count"], formats["removed"]),
        ("Total circuits (old)", results["old_total_circuits"], formats["default"]),
        ("Total circuits (new)", results["new_total_circuits"], formats["default"]),
        ("Added circuits", results["added_circuit_count"], formats["added"]),
        ("Removed circuits", results["removed_circuit_count"], formats["removed"]),
        ("Modified circuits", results["modified_circuit_count"], formats["modified"]),
        ("Unchanged circuits", results["unchanged_circuit_count"], formats["unchanged"]),
    ]
    for row_index, (label, value, cell_format) in enumerate(metrics, start=6):
        dashboard.write(row_index, 0, label, cell_format)
        dashboard.write_number(row_index, 1, int(value), cell_format)

    # --- Circuit change type counts (D6:E10) ---
    dashboard.write("D6", "Circuit Change Type", formats["header"])
    dashboard.write("E6", "Count", formats["header"])
    circuit_rows = [
        ("Added", results["added_circuit_count"], formats["added"]),
        ("Removed", results["removed_circuit_count"], formats["removed"]),
        ("Modified", results["modified_circuit_count"], formats["modified"]),
        ("Unchanged", results["unchanged_circuit_count"], formats["unchanged"]),
    ]
    for row_index, (label, value, cell_format) in enumerate(circuit_rows, start=6):
        dashboard.write(row_index, 3, label, cell_format)
        dashboard.write_number(row_index, 4, int(value), cell_format)

    # --- CNUM change type counts (D13:E15) ---
    dashboard.write("D13", "CNUM Change Type", formats["header"])
    dashboard.write("E13", "Count", formats["header"])
    dashboard.write("D14", "Added CNUMs", formats["added"])
    dashboard.write_number("E14", int(results["added_cnum_count"]), formats["added"])
    dashboard.write("D15", "Removed CNUMs", formats["removed"])
    dashboard.write_number("E15", int(results["removed_cnum_count"]), formats["removed"])

    # --- Harness Family impact table (G6:M...) ---
    family_summary = results.get("harness_family_summary_df")
    family_rows_written = 0
    family_data_start = 8  # zero-based row index of first family data row
    if isinstance(family_summary, pd.DataFrame) and not family_summary.empty:
        top_families = family_summary.head(15)
        dashboard.merge_range("G6:M6", "Harness Family Impact (Top 15 by Total Changes)", formats["subheader"])
        family_headers = [
            "Harness Family",
            "New CNUM",
            "Removed CNUM",
            "Added",
            "Removed",
            "Modified",
            "Total",
        ]
        for col_offset, header in enumerate(family_headers):
            dashboard.write(6, 6 + col_offset, header, formats["header"])
        for row_offset, (_, row) in enumerate(top_families.iterrows()):
            row_index = family_data_start + row_offset - 1
            dashboard.write(row_index, 6, row["Harness Family"], formats["default"])
            dashboard.write_number(row_index, 7, int(row["New CNUM Circuits"]), formats["added"])
            dashboard.write_number(row_index, 8, int(row["Removed CNUM Circuits"]), formats["removed"])
            dashboard.write_number(row_index, 9, int(row["Added Circuits"]), formats["added"])
            dashboard.write_number(row_index, 10, int(row["Removed Circuits"]), formats["removed"])
            dashboard.write_number(row_index, 11, int(row["Modified Circuits"]), formats["modified"])
            dashboard.write_number(row_index, 12, int(row["Total Changes"]), formats["default"])
            family_rows_written += 1

    # --- Charts ---
    circuit_chart = workbook.add_chart({"type": "column"})
    circuit_chart.add_series(
        {
            "name": "Circuit Changes",
            "categories": ["Dashboard", 6, 3, 9, 3],
            "values": ["Dashboard", 6, 4, 9, 4],
            "points": [
                {"fill": {"color": STATUS_COLORS["Added"]}},
                {"fill": {"color": STATUS_COLORS["Removed"]}},
                {"fill": {"color": STATUS_COLORS["Modified"]}},
                {"fill": {"color": STATUS_COLORS["Unchanged"]}},
            ],
        }
    )
    circuit_chart.set_title({"name": "Circuit Changes", "name_font": {"size": 12}})
    circuit_chart.set_legend({"none": True})
    dashboard.insert_chart("A19", circuit_chart, {"x_scale": 1.0, "y_scale": 1.0})

    if family_rows_written:
        first_data_row = family_data_start - 1
        last_data_row = first_data_row + family_rows_written - 1
        family_chart = workbook.add_chart({"type": "bar", "subtype": "stacked"})
        series_specs = [
            ("New CNUM", 7, "#70AD47"),
            ("Removed CNUM", 8, "#C00000"),
            ("Added", 9, STATUS_COLORS["Added"]),
            ("Removed", 10, STATUS_COLORS["Removed"]),
            ("Modified", 11, STATUS_COLORS["Modified"]),
        ]
        for series_name, column_index, color in series_specs:
            family_chart.add_series(
                {
                    "name": series_name,
                    "categories": ["Dashboard", first_data_row, 6, last_data_row, 6],
                    "values": ["Dashboard", first_data_row, column_index, last_data_row, column_index],
                    "fill": {"color": color},
                }
            )
        family_chart.set_title({"name": "Changes by Harness Family", "name_font": {"size": 12}})
        family_chart.set_legend({"position": "bottom"})
        family_chart.set_y_axis({"reverse": True})
        dashboard.insert_chart("G25", family_chart, {"x_scale": 1.9, "y_scale": 1.6})

    # --- Top 20 CNUMs (A36) ---
    top_table_header_row = 35
    dashboard.write(top_table_header_row, 0, "CNUM", formats["header"])
    dashboard.write(top_table_header_row, 1, "Total Changes", formats["header"])
    dashboard.merge_range("A35:B35", "Top 20 CNUMs by Changes", formats["subheader"])
    top_20 = results["top_20_cnums_df"]
    for row_offset, (_, row) in enumerate(top_20.iterrows(), start=1):
        dashboard.write(top_table_header_row + row_offset, 0, row["CNUM"], formats["default"])
        dashboard.write_number(
            top_table_header_row + row_offset, 1, int(row["Total Changes"]), formats["default"]
        )

    # --- Field change frequency (D36) ---
    dashboard.merge_range("D35:E35", "Field Change Frequency", formats["subheader"])
    dashboard.write(top_table_header_row, 3, "Field Name", formats["header"])
    dashboard.write(top_table_header_row, 4, "Changes", formats["header"])
    field_frequency = results["field_change_frequency_df"]
    for row_offset, (_, row) in enumerate(field_frequency.iterrows(), start=1):
        dashboard.write(top_table_header_row + row_offset, 3, row["Field Name"], formats["default"])
        dashboard.write_number(
            top_table_header_row + row_offset, 4, int(row["Number of Changes"]), formats["default"]
        )

    top_count = min(10, len(field_frequency))
    if top_count:
        field_chart = workbook.add_chart({"type": "bar"})
        field_chart.add_series(
            {
                "name": "Top 10 Most Changed Fields",
                "categories": ["Dashboard", top_table_header_row + 1, 3, top_table_header_row + top_count, 3],
                "values": ["Dashboard", top_table_header_row + 1, 4, top_table_header_row + top_count, 4],
                "fill": {"color": STATUS_COLORS["Modified"]},
            }
        )
        field_chart.set_title({"name": "Top 10 Most Changed Fields", "name_font": {"size": 12}})
        field_chart.set_legend({"none": True})
        field_chart.set_y_axis({"reverse": True})
        dashboard.insert_chart("G60", field_chart, {"x_scale": 1.4, "y_scale": 1.2})


def write_report_to_bytes(
    old_name: str,
    new_name: str,
    old_layout: WorkbookLayout,
    new_layout: WorkbookLayout,
    results: dict[str, object],
) -> bytes:
    output_buffer = BytesIO()

    with pd.ExcelWriter(output_buffer, engine="xlsxwriter") as writer:
        workbook = writer.book
        workbook.set_properties(
            {
                "title": "DTx Engineering Change Report",
                "subject": f"{old_name} vs {new_name}",
                "comments": (
                    f"OLD sheet '{old_layout.sheet_name}' row {old_layout.header_row + 1}; "
                    f"NEW sheet '{new_layout.sheet_name}' row {new_layout.header_row + 1}"
                ),
            }
        )

        formats = {
            "header": workbook.add_format(
                {"bold": True, "bg_color": "#1F4E78", "font_color": "#FFFFFF", "border": 1}
            ),
            "default": workbook.add_format({"border": 1}),
            "added": workbook.add_format({"border": 1, "bg_color": STATUS_COLORS["Added"]}),
            "removed": workbook.add_format({"border": 1, "bg_color": STATUS_COLORS["Removed"]}),
            "modified": workbook.add_format({"border": 1, "bg_color": STATUS_COLORS["Modified"]}),
            "unchanged": workbook.add_format({"border": 1, "bg_color": STATUS_COLORS["Unchanged"]}),
            "title": workbook.add_format({"bold": True, "font_size": 16, "font_color": "#1F4E78"}),
            "meta": workbook.add_format({"italic": True, "font_color": "#595959"}),
            "subheader": workbook.add_format(
                {"bold": True, "bg_color": "#DDEBF7", "font_color": "#1F4E78", "border": 1}
            ),
            "family_banner": workbook.add_format(
                {
                    "bold": True,
                    "bg_color": "#1F4E78",
                    "font_color": "#FFFFFF",
                    "border": 1,
                    "font_size": 11,
                }
            ),
        }

        build_dashboard_sheet(writer, results, workbook, formats, old_name=old_name, new_name=new_name)

        all_changes_df = results.get("all_changes_df")
        family_summary_df = results.get("harness_family_summary_df")
        if isinstance(family_summary_df, pd.DataFrame):
            write_table(writer, "Harness Family Summary", family_summary_df, workbook, formats)
        if isinstance(all_changes_df, pd.DataFrame):
            write_table(writer, "All Changes", all_changes_df, workbook, formats)

        write_table(writer, "Added CNUMs", results["added_cnums_df"], workbook, formats)
        write_table(writer, "Removed CNUMs", results["removed_cnums_df"], workbook, formats)
        write_table(writer, "Added Circuits", results["added_circuits_df"], workbook, formats)
        write_table(writer, "Removed Circuits", results["removed_circuits_df"], workbook, formats)

        modified_df = results["modified_circuits_df"]
        compare_pairs: list[tuple[int, int]] = []
        if not modified_df.empty:
            for column in COMPARISON_COLUMNS:
                old_col = modified_df.columns.get_loc(f"{column}_old")
                new_col = modified_df.columns.get_loc(f"{column}_new")
                compare_pairs.append((old_col, new_col))

        write_table(
            writer,
            "Modified Circuits",
            modified_df,
            workbook,
            formats,
            compare_pairs=compare_pairs,
        )
        write_table(writer, "Change Log", results["change_log_df"], workbook, formats)
        write_table(writer, "CNUM Summary", results["cnum_summary_df"], workbook, formats)

        workbook.worksheets_objs[0].activate()

    output_buffer.seek(0)
    return output_buffer.getvalue()


def build_output_filename(old_name: str, new_name: str) -> str:
    old_stem = Path(old_name).stem
    new_stem = Path(new_name).stem
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"DTx_Change_Report_{old_stem}_vs_{new_stem}_{timestamp}.xlsx"


def generate_dtx_change_report(
    old_file_bytes: bytes,
    new_file_bytes: bytes,
    old_file_name: str,
    new_file_name: str,
    dtcr_df: pd.DataFrame | None = None,
) -> dict[str, object]:
    old_df, old_layout = load_dtx_report(old_file_bytes, old_file_name)
    new_df, new_layout = load_dtx_report(new_file_bytes, new_file_name)

    results = compare_reports(old_df, new_df)
    if dtcr_df is not None:
        dtcr_by_cnum = _build_dtcr_lookup_by_cnum(old_df, new_df, dtcr_df)
        results = _annotate_results_with_dtcr(results, dtcr_by_cnum)
        # Second output: DTCR_Matching_Report built from BOTH DTx reports.
        results.update(
            generate_dtcr_matching_report(
                old_file_bytes=old_file_bytes,
                new_file_bytes=new_file_bytes,
                old_file_name=old_file_name,
                new_file_name=new_file_name,
                dtcr_df=dtcr_df,
            )
        )

    all_changes_df = build_all_changes_df(results)
    results["all_changes_df"] = all_changes_df
    results["harness_family_summary_df"] = build_family_summary_df(all_changes_df)

    output_bytes = write_report_to_bytes(old_file_name, new_file_name, old_layout, new_layout, results)

    return {
        **results,
        "output_excel_bytes": output_bytes,
        "output_file_name": build_output_filename(old_file_name, new_file_name),
        "old_layout": old_layout,
        "new_layout": new_layout,
    }
