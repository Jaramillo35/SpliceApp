"""Reading a Harness Complexity workbook.

The ``Complexity`` sheet is a part-number × sales-code matrix: the first row
holds ``ID=<def id>`` then the sales codes, and every row after it is one
harness part number with an ``X`` under each code it carries. Fourteen builds
and sixty-four codes for the sample IP harness.

Those rows are the only vehicle configurations that exist. That matters more
than it sounds: ``RSY`` and ``RTC`` both appear on the IP and the Dash, but no
build carries both, so any reasoning that treats sales codes as independent
invents vehicles that are never made.

The DEF id is read from **inside** the file, never from its name. In the sample
release the IP complexity is called ``…_IP_11645_…`` and declares
``ID=11661`` — matching on the filename binds it to the wrong harness.
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from splice.common.errors import SpliceInputError
from splice.common.logging import get_logger
from splice.inline.model import Build, Harness

logger = get_logger(__name__)

SHEET = "Complexity"


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def read_complexity(payload: bytes, filename: str = "") -> Harness:
    """Parse one complexity workbook into a :class:`Harness`.

    Raises :class:`SpliceInputError` when the file is not a complexity export;
    the caller turns that into a reported gap rather than a crash.
    """
    import io

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(payload), data_only=True, read_only=True
        )
    except Exception as exc:  # noqa: BLE001 - reported, not raised onward
        raise SpliceInputError(
            f"{filename or 'The file'} could not be opened as a workbook: {exc}"
        ) from exc

    try:
        if SHEET not in workbook.sheetnames:
            raise SpliceInputError(
                f"{filename or 'The file'} has no '{SHEET}' sheet, so it is not a "
                "Harness Complexity export."
            )
        rows = [row for row in workbook[SHEET].iter_rows(values_only=True)]
    finally:
        workbook.close()

    if not rows:
        raise SpliceInputError(f"{filename or 'The file'} has an empty {SHEET} sheet.")

    header = rows[0]
    def_id = re.sub(r"\D", "", _text(header[0]))
    if not def_id:
        raise SpliceInputError(
            f"{filename or 'The file'} does not declare a DEF id in the first cell "
            f"of the {SHEET} sheet (expected something like 'ID=11661')."
        )

    codes = [_text(c) for c in header[1:] if _text(c)]
    builds = []
    for row in rows[1:]:
        part = _text(row[0])
        if not part:
            continue
        carried = frozenset(
            codes[i]
            for i in range(len(codes))
            if i + 1 < len(row) and _text(row[i + 1])
        )
        builds.append(Build(part_number=part, codes=carried))

    harness = Harness(
        name=Path(filename).stem if filename else def_id,
        def_id=def_id,
        builds=builds,
        complexity_codes=set(codes),
        complexity_file=filename,
    )
    logger.info(
        "Complexity %s: id=%s builds=%d codes=%d",
        filename, def_id, len(builds), len(codes),
    )
    return harness


def satisfying_builds(harness: Harness, expression: str) -> list:
    """The builds of THIS harness in which a condition holds.

    Applicability is only ever evaluated against the harness that owns the
    condition, using the builds that harness actually ships as. That keeps the
    engine away from a question it cannot answer — how a part number on one
    harness corresponds to a part number on another — and it keeps impossible
    vehicles out of the reasoning for free: ``RSY`` and ``RTC`` both exist on
    the IP and the Dash, but no build carries both, so no evaluation here can
    ever land in an ``RSY & RTC`` vehicle.

    An empty result means the condition is never built, which is a defect worth
    reporting on its own.
    """
    from splice.inline import salescode

    return [
        build
        for build in harness.builds
        if salescode.evaluate(expression, build.codes)
    ]


def applies_in(expression: str, config, vocabulary) -> bool:
    """Does a condition hold in a vehicle described by ``config``?

    ``config`` is the sales-code set of a build belonging to the *other*
    harness, and ``vocabulary`` is that harness's complexity header — every code
    it tracks. A code the expression mentions but the vocabulary does not track
    is **unknown**, not absent, and is treated as present.

    The distinction is the difference between a finding and a false alarm. In
    the reference release ``G74`` is conditioned ``JPB/JRK`` on the Body_Right
    and ``JPB`` on the Body_Left; ``JPB`` appears on all 19 Body_Left builds but
    is not listed anywhere in the Body_Right complexity. Reading that silence as
    "the vehicle does not have JPB" reports a continuity break on a joint that
    is fine. The Body_Right simply does not record that code.
    """
    from splice.inline import salescode

    unknown = {
        code for code in salescode.codes_in(expression) if code not in vocabulary
    }
    return salescode.evaluate(expression, set(config) | unknown)


def counterpart_applies(
    harness_a: Harness | None,
    expression_a: str,
    harness_b: Harness | None,
    expression_b: str,
) -> bool:
    """Is the far-side circuit present in the vehicles where the near-side one is?

    This is the case the engineering team called out. On the IP a cavity may
    split by sales code — ``A934A`` on some part numbers, ``A934B`` on others —
    while the Dash carries a single ``A934`` across every one of its part
    numbers. The Dash wire is the counterpart of *both* IP variants, because its
    condition holds in the vehicles where each of them is built.

    Pairing one-to-one instead reports the second variant as having no
    counterpart, which is wrong: if it applies, it applies.
    """
    if harness_a is None or not harness_a.builds:
        return True
    if harness_b is not None and harness_b.builds:
        if not satisfying_builds(harness_b, expression_b):
            return False

    where_a_applies = satisfying_builds(harness_a, expression_a)
    if not where_a_applies:
        return False
    if harness_b is None or not harness_b.builds:
        return True
    return any(
        applies_in(expression_b, build.codes, harness_a.complexity_codes)
        for build in where_a_applies
    )
