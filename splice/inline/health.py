"""Holistic missing-circuit health check over the inline study.

Layers on top of :mod:`splice.inline.validate`, which decides cavities from
the wires that exist. The defects that motivated this module (circuits R732
and A960 in the 28RU X1 validation) were invisible at that level: every wire
that existed paired cleanly, and the defect was a *window* of the option
space — real builds on both harnesses — where no wire existed at all.

Three layers:

1. Cavity continuity — delegated to ``run_study`` (missing continuation,
   inconsistent circuits, conditions never built together).
2. Option-window coverage — per cavity, union each side's sales expressions,
   difference them, and evaluate every one-sided window against the build
   tables. A window with builds on the side that has no wire is a Blocker;
   a window that provably never builds is auto-cleared with the proof kept.
3. Circuit-route completeness — a circuit that is live on a harness in some
   window (it crosses other inlines there) but absent at one crossing in that
   same window is flagged for review; that is the A960 signature.

Unknown-code semantics follow :func:`splice.inline.complexity.applies_in`: a
code a harness does not track is unknown, not absent, and is treated as
present — silence in a complexity header must not manufacture findings.

Findings carry a stable fingerprint so Systems-Engineer dispositions persist
across runs (the review-gate baseline).
"""

from __future__ import annotations

import hashlib
import io
import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from splice.inline import boolmin, salescode
from splice.inline.model import CircuitEnd, Harness, InlinePair
from splice.inline.validate import run_study

SEV_BLOCKER = "Blocker"
SEV_HIGH = "High"
SEV_REVIEW = "Review"
SEV_INFO = "Info"

#: verdicts from the cavity layer mapped to health severities
_VERDICT_SEVERITY = {
    "Missing continuation": SEV_BLOCKER,
    "Inconsistent definition": SEV_BLOCKER,
    "Conditions exclusive": SEV_HIGH,
    "Applicability sources disagree": SEV_HIGH,
    "Undetermined": SEV_REVIEW,
    "Not in Ckt Summary": SEV_INFO,
}

_DATE_IN_NAME = re.compile(r"(\d{2})-(\d{2})-(\d{4})")

#: dispositions a Systems Engineer can record on a finding
DISPOSITIONS = ("Accepted variant", "Defect", "By design")


# ---------------------------------------------------------------------------
# Expression algebra
# ---------------------------------------------------------------------------

def builds_where(harness: Harness, expression: Optional[str]) -> list:
    """Builds of ``harness`` in which ``expression`` holds.

    ``None`` means unconditional (an empty sales code — a 100% wire) and
    matches every build. Codes outside the harness's tracked vocabulary are
    treated as present (see module docstring).
    """
    if expression is None:
        return list(harness.builds)
    vocabulary = harness.complexity_codes
    unknown = {c for c in salescode.codes_in(expression) if c not in vocabulary}
    out = []
    for build in harness.builds:
        if salescode.evaluate(expression, set(build.codes) | unknown):
            out.append(build)
    return out


def union_expression(ends: List[CircuitEnd]) -> Optional[str]:
    """One expression covering every wire in ``ends``; ``None`` = always."""
    exprs = []
    for end in ends:
        sc = (end.sales_code or "").strip()
        if not sc:
            return None  # an unconditional wire covers the whole cavity
        exprs.append(f"({sc})")
    if not exprs:
        return None
    return "/".join(sorted(set(exprs)))


def window_minus(cover: Optional[str], other: Optional[str]) -> Optional[str]:
    """Expression for 'covered by ``cover`` but not by ``other``'.

    Returns ``None`` when the window is provably empty without evaluation
    (``other`` unconditional).
    """
    if other is None:
        return None
    if cover is None:
        return f"-({other})"
    return f"({cover})&-({other})"


# ---------------------------------------------------------------------------
# Findings
# ---------------------------------------------------------------------------

@dataclass
class HealthFinding:
    severity: str
    kind: str            # cavity | one_sided_window | route_window_gap | config_skew
    inline: str
    cavity: str = ""
    circuit: str = ""
    harness_with: str = ""     # side that has the wire / coverage
    harness_without: str = ""  # side that lacks it
    window: str = ""           # sales-expression window of the gap
    builds_with: List[str] = field(default_factory=list)
    builds_without: List[str] = field(default_factory=list)
    detail: str = ""
    #: route gaps only: the crossings where the circuit IS covered — the
    #: finding lives within one harness, and presentation must say so
    crossings: List[str] = field(default_factory=list)
    #: minimal boolean form of ``window`` for display; the raw window stays
    #: on ``window`` because fingerprints (and stored dispositions) hash it
    window_short: str = ""

    @property
    def window_display(self) -> str:
        return self.window_short or self.window

    @property
    def within_harness(self) -> bool:
        return self.kind == "route_window_gap"

    @property
    def fingerprint(self) -> str:
        basis = "|".join([
            self.kind, self.inline, self.cavity, self.circuit,
            re.sub(r"\s+", "", self.window),
        ])
        return hashlib.sha1(basis.encode("utf-8")).hexdigest()[:16]


@dataclass
class ClearedProof:
    inline: str
    cavity: str
    window: str
    detail: str


@dataclass
class InputRow:
    harness: str
    def_id: str
    file: str
    file_date: str
    builds: int


@dataclass
class InputReport:
    rows: List[InputRow] = field(default_factory=list)
    missing_complexity: List[str] = field(default_factory=list)
    skew_days: int = 0
    skew_pair: str = ""

    @property
    def has_issues(self) -> bool:
        return bool(self.missing_complexity) or self.skew_days > 30


@dataclass
class HealthResult:
    inputs: InputReport
    findings: List[HealthFinding]
    cleared: List[ClearedProof]
    pairs: List[InlinePair] = field(default_factory=list)
    cavities_checked: int = 0
    #: the full cavity study (layer 1) — kept so the continuity audit views
    #: (every cavity, marked differences, findings workbook) need no rerun
    study: object = None
    #: readiness gaps from splice.inline.readiness (input pre-checks)
    gaps: list = field(default_factory=list)

    def open_findings(self, baseline: dict) -> List[HealthFinding]:
        return [f for f in self.findings
                if f.fingerprint not in baseline.get("dispositions", {})]

    def blocking_open(self, baseline: dict) -> List[HealthFinding]:
        return [f for f in self.open_findings(baseline)
                if f.severity in (SEV_BLOCKER, SEV_HIGH)]


# ---------------------------------------------------------------------------
# Gate 0 — inputs
# ---------------------------------------------------------------------------

_INTEGRITY_LISTS = re.compile(
    r"Only in columns: (\[.*?\]|—); only in the expression: (\[.*?\]|—)")


def _integrity_sets(reason: str) -> tuple:
    """The two part-number sets out of an integrity finding's reason text."""
    import ast as _ast
    m = _INTEGRITY_LISTS.search(reason or "")
    if not m:
        return (), ()

    def parse(chunk: str) -> tuple:
        if not chunk.startswith("["):
            return ()
        try:
            return tuple(sorted(str(x) for x in _ast.literal_eval(chunk)))
        except Exception:
            return (chunk,)

    return parse(m.group(1)), parse(m.group(2))


def _truncation_hint(cols: tuple, expr: tuple) -> str:
    """When a column PN and an expression PN differ only by a trailing
    fragment, the likely root cause is a truncated part number in one
    source — say so (field report: '687894643A' vs '687894643AA')."""
    for a in cols:
        for b in expr:
            if a != b and (a.startswith(b) or b.startswith(a)):
                return (f" The pair '{a}' / '{b}' differs only by a trailing "
                        "fragment — likely a truncated part number in the "
                        "Circuit Summary column header or the complexity "
                        "table; fix that one cell and every row here clears.")
    return ""


def file_date(filename: str) -> Optional[datetime]:
    m = _DATE_IN_NAME.search(filename or "")
    if not m:
        return None
    try:
        return datetime(int(m.group(3)), int(m.group(1)), int(m.group(2)))
    except ValueError:
        return None


def build_input_report(summary: Dict[str, Harness],
                       complexity: Dict[str, Harness]) -> InputReport:
    report = InputReport()
    dated: List[Tuple[str, datetime]] = []
    for hid, sh in sorted(summary.items()):
        cx = complexity.get(hid)
        if cx is None:
            report.missing_complexity.append(f"{sh.name} ({hid})")
            continue
        d = file_date(cx.complexity_file or cx.name)
        report.rows.append(InputRow(
            harness=sh.name, def_id=hid,
            file=cx.complexity_file or cx.name,
            file_date=d.strftime("%Y-%m-%d") if d else "unknown",
            builds=len(cx.builds),
        ))
        if d:
            dated.append((sh.name, d))
    if len(dated) >= 2:
        dated.sort(key=lambda t: t[1])
        oldest, newest = dated[0], dated[-1]
        report.skew_days = (newest[1] - oldest[1]).days
        report.skew_pair = f"{oldest[0]} ({oldest[1]:%Y-%m-%d}) vs {newest[0]} ({newest[1]:%Y-%m-%d})"
    return report


# ---------------------------------------------------------------------------
# The analysis
# ---------------------------------------------------------------------------

def _parts(builds, cap: int = 8) -> List[str]:
    return [b.part_number for b in builds[:cap]]


def analyze(summary: Dict[str, Harness], ends: List[CircuitEnd],
            complexity: Dict[str, Harness], pairs: List[InlinePair],
            unmated: List) -> HealthResult:
    """Run all three layers and return the consolidated result."""
    # Complexity harnesses are named after their FILE (read_complexity gets the
    # filename); findings must wear the summary's harness names, or the pair
    # matrix and reports show "2_Harness_Complexity_..._DASH_04-19-2026"
    # instead of "DASH" (field report, 2026-08-24).
    for hid, cx in complexity.items():
        if hid in summary and summary[hid].name:
            cx.name = summary[hid].name

    result = HealthResult(inputs=build_input_report(summary, complexity),
                          findings=[], cleared=[], pairs=pairs)

    display = {hid: h.name for hid, h in summary.items()}

    # ---- Gate 0 extras: readiness pre-checks ------------------------------
    # (duplicate-upload detection needs the pre-dedup file list, which this
    # API does not receive — callers dedupe by DEF id before analyze)
    from splice.inline.readiness import assess
    result.gaps = assess(summary, complexity)

    # ---- Layer 1: cavity continuity, via the existing engine -------------
    study = run_study(summary, ends, complexity, pairs, unmated)
    result.study = study
    # Cavities whose circuits disagree (or are never built together) get ONE
    # finding carrying the complete picture: the identity conflict AND the
    # coverage windows, merged (field reports: L206/N0 at X103A cavity 14
    # first double-reported, then lost the window when suppressed).
    conflicted: Dict[tuple, HealthFinding] = {}
    integrity_groups: Dict[tuple, list] = {}
    for f in study.findings:
        sev = _VERDICT_SEVERITY.get(f.verdict)
        if sev is None or f.verdict == "Continuous":
            continue
        if sev == SEV_INFO:
            continue  # unmated-connector notes belong to Gate 0 context
        if f.verdict == "Applicability sources disagree":
            # hundreds of rows can share ONE root cause (e.g. a truncated
            # part number in a summary column header) — group them
            cols, expr = _integrity_sets(f.reason)
            integrity_groups.setdefault(
                (f.harness_a, cols, expr), []).append(f)
            continue
        finding = HealthFinding(
            severity=sev, kind="cavity", inline=f.inline, cavity=f.cavity,
            circuit=", ".join(sorted(set(f.circuits_a + f.circuits_b))),
            harness_with=f.harness_a, harness_without=f.harness_b,
            window="", detail=f.reason,
        )
        if f.verdict in ("Inconsistent definition", "Conditions exclusive"):
            conflicted[(f.connector_a, f.connector_b, f.cavity)] = finding
        result.findings.append(finding)

    for (h_name, cols, expr), group in sorted(integrity_groups.items(),
                                              key=lambda kv: str(kv[0])):
        circuits = sorted({c for f in group for c in f.circuits_a})
        shown = ", ".join(circuits[:6]) + ("…" if len(circuits) > 6 else "")
        hint = _truncation_hint(cols, expr)
        result.findings.append(HealthFinding(
            severity=SEV_HIGH, kind="integrity", inline=h_name, cavity="",
            circuit=f"{len(group)} row(s): {shown}",
            harness_with=h_name, harness_without=h_name,
            window=f"columns:{'/'.join(cols) or '—'} vs "
                   f"expression:{'/'.join(expr) or '—'}",
            detail=(f"The Circuit Summary's part-number columns and the "
                    f"complexity table disagree on {len(group)} row(s) of "
                    f"{h_name} (circuits {shown}). Marked only in the "
                    f"columns: {list(cols) or '—'}; resolved only by the "
                    f"expression: {list(expr) or '—'}.{hint}"),
        ))

    # ---- Layer 2: option-window coverage ---------------------------------
    by_harness: Dict[str, List[CircuitEnd]] = {}
    for end in ends:
        by_harness.setdefault(end.harness_id, []).append(end)

    def at(harness_id: str, connector: str) -> Dict[str, List[CircuitEnd]]:
        grouped: Dict[str, List[CircuitEnd]] = {}
        for e in by_harness.get(harness_id, []):
            if e.connector == connector:
                grouped.setdefault(e.cavity, []).append(e)
        return grouped

    for pair in pairs:
        ha, hb = complexity.get(pair.harness_a), complexity.get(pair.harness_b)
        if ha is None or hb is None or not ha.builds or not hb.builds:
            continue  # reported by Gate 0 / layer 1 instead
        name_a = display.get(pair.harness_a, pair.harness_a)
        name_b = display.get(pair.harness_b, pair.harness_b)
        ends_a, ends_b = at(pair.harness_a, pair.connector_a), at(pair.harness_b, pair.connector_b)
        for cavity in sorted(set(ends_a) | set(ends_b)):
            side_a, side_b = ends_a.get(cavity, []), ends_b.get(cavity, [])
            if not side_a or not side_b:
                continue  # layer 1 already decided fully empty sides
            merge_into = conflicted.get((pair.connector_a, pair.connector_b,
                                         cavity))
            result.cavities_checked += 1
            u_a, u_b = union_expression(side_a), union_expression(side_b)
            circuits = ", ".join(sorted({e.circuit for e in side_a + side_b}))
            for (u_have, u_lack, h_have, h_lack, n_have, n_lack, c_have, c_lack) in (
                (u_a, u_b, ha, hb, name_a, name_b, pair.connector_a, pair.connector_b),
                (u_b, u_a, hb, ha, name_b, name_a, pair.connector_b, pair.connector_a),
            ):
                window = window_minus(u_have, u_lack)
                if window is None:
                    continue  # lacking side unconditional: covered by definition
                short = boolmin.minimize(
                    window, boolmin.care_configurations(h_have, h_lack))
                lacking = builds_where(h_lack, window)
                having = builds_where(h_have, window)
                if merge_into is not None:
                    # conflicted cavity: fold the coverage picture into the
                    # ONE cavity finding instead of emitting a second row
                    if lacking:
                        merge_into.severity = SEV_BLOCKER
                        merge_into.window_short = merge_into.window_short or short
                        merge_into.builds_with = list(dict.fromkeys(
                            merge_into.builds_with
                            + [b.part_number for b in having]))[:8]
                        merge_into.builds_without = list(dict.fromkeys(
                            merge_into.builds_without
                            + [b.part_number for b in lacking]))[:8]
                        merge_into.detail += (
                            f" Coverage also differs: in window {short}, "
                            f"{n_have} has a wire ({len(having)} build(s)) "
                            f"while {n_lack} has none ({len(lacking)} "
                            f"build(s)).")
                    continue
                if lacking:
                    result.findings.append(HealthFinding(
                        severity=SEV_BLOCKER, kind="one_sided_window",
                        inline=f"{pair.connector_a} ↔ {pair.connector_b}",
                        cavity=cavity, circuit=circuits,
                        harness_with=n_have, harness_without=n_lack,
                        window=window, window_short=short,
                        builds_with=_parts(having), builds_without=_parts(lacking),
                        detail=(f"In window {short}, {n_have} has a wire at "
                                f"{c_have} cavity {cavity} but {n_lack} builds "
                                f"{len(lacking)} part number(s) with no wire at "
                                f"{c_lack}."),
                    ))
                elif having:
                    result.findings.append(HealthFinding(
                        severity=SEV_REVIEW, kind="one_sided_window",
                        inline=f"{pair.connector_a} ↔ {pair.connector_b}",
                        cavity=cavity, circuit=circuits,
                        harness_with=n_have, harness_without=n_lack,
                        window=window, window_short=short,
                        builds_with=_parts(having),
                        detail=(f"Window {short} builds only on {n_have} "
                                f"({len(having)} build(s)); {n_lack} has no "
                                "build there — possible revision/config skew."),
                    ))
                else:
                    result.cleared.append(ClearedProof(
                        inline=f"{pair.connector_a} ↔ {pair.connector_b}",
                        cavity=cavity, window=short,
                        detail=(f"Window never builds on either harness — "
                                f"{n_have}'s extra coverage is provably unreachable."),
                    ))

    # ---- Layer 3: circuit-route completeness ------------------------------
    inline_connectors: Dict[str, List[Tuple[InlinePair, str]]] = {}
    for pair in pairs:
        inline_connectors.setdefault(pair.harness_a, []).append((pair, pair.connector_a))
        inline_connectors.setdefault(pair.harness_b, []).append((pair, pair.connector_b))

    for hid, crossings in inline_connectors.items():
        harness = complexity.get(hid)
        if harness is None or not harness.builds:
            continue
        h_name = display.get(hid, hid)
        per_circuit: Dict[str, Dict[str, List[CircuitEnd]]] = {}
        for e in by_harness.get(hid, []):
            for pair, connector in crossings:
                if e.connector == connector:
                    per_circuit.setdefault(e.circuit, {}).setdefault(connector, []).append(e)
        for circuit, at_connector in per_circuit.items():
            if len(at_connector) < 2:
                continue
            unions = {c: union_expression(ends_) for c, ends_ in at_connector.items()}
            if any(u is None for u in unions.values()):
                overall = None
            else:
                overall = "/".join(f"({u})" for u in unions.values())
            for connector, u_here in unions.items():
                gap = window_minus(overall, u_here)
                if gap is None:
                    continue
                gap_builds = builds_where(harness, gap)
                if not gap_builds:
                    continue
                short = boolmin.minimize(
                    gap, boolmin.care_configurations(harness))
                elsewhere = sorted(set(at_connector) - {connector})
                result.findings.append(HealthFinding(
                    severity=SEV_HIGH, kind="route_window_gap",
                    inline=connector, cavity="", circuit=circuit,
                    harness_with=h_name, harness_without=h_name,
                    window=gap, window_short=short,
                    builds_with=_parts(gap_builds),
                    crossings=elsewhere,
                    detail=(f"{circuit} is live on {h_name} in window {short} "
                            f"(it crosses {', '.join(elsewhere)}) but has no "
                            f"variant at {connector} there — "
                            f"{len(gap_builds)} build(s) affected."),
                ))

    # Two findings with the same fingerprint are the same logical finding
    # (one disposition must clear both); keep the more severe occurrence.
    order = {SEV_BLOCKER: 0, SEV_HIGH: 1, SEV_REVIEW: 2, SEV_INFO: 3}
    result.findings.sort(key=lambda f: (order.get(f.severity, 9), f.inline, f.cavity))
    seen: Set[str] = set()
    deduped = []
    for f in result.findings:
        if f.fingerprint in seen:
            continue
        seen.add(f.fingerprint)
        deduped.append(f)
    result.findings = deduped
    return result


# ---------------------------------------------------------------------------
# Disposition baseline (the SE review-gate memory)
# ---------------------------------------------------------------------------

def load_baseline(path: Path) -> dict:
    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception:
        return {"dispositions": {}, "signoffs": []}


def save_baseline(path: Path, baseline: dict) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")
    tmp.write_text(json.dumps(baseline, indent=2, sort_keys=True), encoding="utf-8")
    tmp.replace(path)


def disposition(baseline: dict, finding: HealthFinding, verdict: str,
                reason: str, by: str) -> dict:
    if verdict not in DISPOSITIONS:
        raise ValueError(f"Unknown disposition {verdict!r}")
    baseline.setdefault("dispositions", {})[finding.fingerprint] = {
        "verdict": verdict, "reason": reason.strip(), "by": by.strip(),
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "summary": f"[{finding.severity}] {finding.kind} {finding.inline} "
                   f"cav {finding.cavity} {finding.circuit} {finding.window}",
    }
    return baseline


def sign_off(baseline: dict, by: str, note: str = "") -> dict:
    baseline.setdefault("signoffs", []).append({
        "by": by.strip(), "note": note.strip(),
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
    })
    return baseline


# ---------------------------------------------------------------------------
# Report workbook
# ---------------------------------------------------------------------------

#: The reviewer guide embedded in every report (and shown in the app).
#: (title, lines) sections; keep in sync with the workbench guide.
REVIEWER_GUIDE = [
    ("What this report is", [
        "Every inline crossing was checked three ways: cavity continuity "
        "(wire on one side, nothing on the other), option-window coverage "
        "(both sides wired, but their sales conditions cover different "
        "vehicles), and route completeness (a circuit live on a harness in a "
        "window at some crossings but absent at another).",
        "A finding exists ONLY if real build part numbers satisfy its window "
        "— everything is validated against the Harness Complexity tables.",
    ]),
    ("Severities", [
        "Blocker: builds exist with a wire on one side and nothing opposite "
        "— a vehicle will be built with a dead-ended circuit. Review first.",
        "High: route gaps (within ONE harness) and configuration skew — "
        "usually real, but routing can legitimately differ by option; these "
        "need engineering judgment.",
        "Review: attribute or bookkeeping differences worth a look.",
        "Auto-cleared: the algebra PROVED the window never builds on either "
        "harness — no action needed; kept in its own sheet for audit.",
    ]),
    ("Technical considerations for the review", [
        "Files are matched by the DEF id INSIDE each complexity file, never "
        "by filename; unusable files are listed on the Inputs sheet.",
        "Check the Inputs sheet dates first: findings that involve a stale "
        "complexity revision may reflect old data, not defects. A skew above "
        "30 days is flagged.",
        "Window expressions are minimized against the BUILDABLE "
        "configurations of the complexity files loaded for THAT finding: "
        "codes carried by every build (or by none) vanish as vacuous, and "
        "two codes that no loaded build ever separates collapse into one. "
        "That is read from the data per finding, never assumed programme-"
        "wide — if any harness in the window can build one code without the "
        "other, both stay in the expression. The raw expression is preserved "
        "internally and the minimal form is machine-verified equivalent on "
        "every buildable configuration.",
        "Codes a harness's complexity does not track are treated as PRESENT "
        "(unknown, not absent) — silence never manufactures a finding, but "
        "it can hide one: if a code matters to a harness, its complexity "
        "must track it.",
        "The build lists are the evidence: 'Builds w/o wire' names the part "
        "numbers that would carry a dead-ended circuit (list capped; the "
        "Detail column carries the full count).",
        "A route gap lives WITHIN one harness — 'Missing on' says so "
        "explicitly. The question to answer: should this circuit cross this "
        "inline in this window, or does it legitimately route elsewhere?",
        "Dispositions are remembered by fingerprint: re-running with the "
        "same inputs re-surfaces only new findings. If engineering edits a "
        "sales expression, its finding returns as new — by design, because "
        "the evidence changed.",
        "Sign-off is only possible with zero open Blockers/Highs; this "
        "report's Findings sheet carries each disposition, reason, and "
        "reviewer as the audit trail.",
        "Attribute marks (suffix/size/sales differences inside continuous "
        "cavities) are recorded in the Continuity audit, pending the "
        "wire-attribute equivalence table — they are notes, not verdicts.",
    ]),
    ("Suggested review order", [
        "1. Inputs sheet: completeness and revision skew.",
        "2. Blockers, largest affected-build counts first.",
        "3. Route gaps, grouped by harness.",
        "4. Review items, then spot-check a few auto-cleared proofs.",
    ]),
]


def render_report(result: HealthResult, baseline: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    ws0 = wb.active
    ws0.title = "Read Me"
    row = 1
    for title, lines in REVIEWER_GUIDE:
        cell = ws0.cell(row=row, column=1, value=title)
        cell.font = bold
        row += 1
        for line in lines:
            ws0.cell(row=row, column=1, value=line)
            row += 1
        row += 1
    ws0.column_dimensions["A"].width = 110
    for r in ws0.iter_rows(min_col=1, max_col=1):
        for cell in r:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("Findings")
    headers = ["Severity", "Kind", "Inline", "Cavity", "Circuit(s)",
               "Window (sales)", "Has wire", "Missing on",
               "Builds w/ wire", "Builds w/o wire", "Detail",
               "Disposition", "Reason", "By", "Fingerprint"]
    ws.append(headers)
    for c in ws[1]:
        c.font = bold
    dispositions = baseline.get("dispositions", {})
    for f in result.findings:
        d = dispositions.get(f.fingerprint, {})
        # A route gap lives within ONE harness; printing the harness in both
        # columns read as "BODY LEFT ↔ BODY LEFT" (field report, 2026-08-24).
        missing_on = (f"(within {f.harness_with} — no variant at {f.inline})"
                      if f.within_harness else f.harness_without)
        ws.append([f.severity, f.kind, f.inline, f.cavity, f.circuit,
                   f.window_display,
                   f.harness_with, missing_on,
                   ", ".join(f.builds_with), ", ".join(f.builds_without),
                   f.detail, d.get("verdict", "OPEN"), d.get("reason", ""),
                   d.get("by", ""), f.fingerprint])

    ws2 = wb.create_sheet("Scorecard")
    ws2.append(["Inline", "Blockers", "High", "Review", "Auto-cleared"])
    for c in ws2[1]:
        c.font = bold
    per_inline: Dict[str, List[int]] = {}
    for f in result.findings:
        row = per_inline.setdefault(f.inline, [0, 0, 0, 0])
        idx = {SEV_BLOCKER: 0, SEV_HIGH: 1, SEV_REVIEW: 2}.get(f.severity)
        if idx is not None:
            row[idx] += 1
    for proof in result.cleared:
        per_inline.setdefault(proof.inline, [0, 0, 0, 0])[3] += 1
    for inline, counts in sorted(per_inline.items()):
        ws2.append([inline, *counts])

    ws3 = wb.create_sheet("Inputs")
    ws3.append(["Harness", "DEF id", "Complexity file", "File date", "Builds"])
    for c in ws3[1]:
        c.font = bold
    for row in result.inputs.rows:
        ws3.append([row.harness, row.def_id, row.file, row.file_date, row.builds])
    if result.inputs.missing_complexity:
        ws3.append([])
        ws3.append(["Missing complexity:", ", ".join(result.inputs.missing_complexity)])
    if result.inputs.skew_days:
        ws3.append(["Revision skew (days):", result.inputs.skew_days, result.inputs.skew_pair])

    ws4 = wb.create_sheet("Cleared proofs")
    ws4.append(["Inline", "Cavity", "Window", "Proof"])
    for c in ws4[1]:
        c.font = bold
    for proof in result.cleared:
        ws4.append([proof.inline, proof.cavity, proof.window, proof.detail])

    ws5 = wb.create_sheet("Sign-off")
    ws5.append(["Date", "By", "Note"])
    for c in ws5[1]:
        c.font = bold
    for s in baseline.get("signoffs", []):
        ws5.append([s.get("date", ""), s.get("by", ""), s.get("note", "")])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
