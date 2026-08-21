"""Reading a Circuit Summary export.

The export is blocked by harness. Each block opens with a header row whose
first cell is ``<Family> - <def id>`` and whose second cell is the literal
``Circuit``; that row also carries the harness part numbers, one per column,
written ``<symbol>~<part number>``. Every row after it, until the next header,
is one circuit end: a wire terminating in a cavity, marked ``X`` under each
build it appears in.

Two details the data forced:

* **A cavity can be written ``8`` or ``8:1``.** The suffix is a second wire
  occupying the same physical cavity. Before normalising, this export and the
  manual Inline Report disagreed about 10 of 64 cavities on ``X301A``; after,
  they agree on all 64.
* **The part-number columns are per block.** A multi-harness export repeats the
  header for every harness, so the columns cannot be read once and reused.
"""

from __future__ import annotations

import io
import re
from typing import Dict, List, Tuple

import openpyxl

from splice.common.errors import SpliceInputError
from splice.common.logging import get_logger
from splice.inline.model import Build, CircuitEnd, Harness

logger = get_logger(__name__)

SHEET = "Circuit Summary"

#: Column positions are fixed in the export; the part numbers start after them.
COL_FAMILY, COL_CIRCUIT, COL_SUFFIX, COL_SIZE = 0, 1, 2, 3
COL_MATERIAL, COL_COLOR, COL_CNUM, COL_CAV = 4, 5, 7, 8
COL_DEVICE, COL_SALES_CODE = 10, 16
FIRST_BUILD_COL = 17

_BLOCK = re.compile(r"^(?P<name>.*?)\s*-\s*(?P<id>\d+)$")


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def normalise_cavity(value) -> str:
    """``"8:1"`` -> ``"8"`` — a second wire in the same physical cavity."""
    return _text(value).split(":")[0].strip()


def read_circuit_summary(
    payload: bytes, filename: str = ""
) -> Tuple[Dict[str, Harness], List[CircuitEnd]]:
    """Parse the export into harnesses and the circuit ends they contain."""
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(payload), data_only=True, read_only=True
        )
    except Exception as exc:  # noqa: BLE001
        raise SpliceInputError(
            f"{filename or 'The Circuit Summary'} could not be opened: {exc}"
        ) from exc

    try:
        sheet = workbook[SHEET] if SHEET in workbook.sheetnames else workbook.active
        rows = [row for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    if not rows:
        raise SpliceInputError("The Circuit Summary is empty.")

    harnesses: Dict[str, Harness] = {}
    ends: List[CircuitEnd] = []
    current: Harness | None = None
    build_columns: Dict[int, str] = {}

    for number, row in enumerate(rows[1:], start=2):
        family = _text(row[COL_FAMILY]) if len(row) > COL_FAMILY else ""
        if not family:
            continue

        if len(row) > COL_CIRCUIT and _text(row[COL_CIRCUIT]) == "Circuit":
            match = _BLOCK.match(family)
            name, def_id = (
                (match.group("name"), match.group("id")) if match else (family, "")
            )
            build_columns = {}
            builds = []
            for index in range(FIRST_BUILD_COL, len(row)):
                cell = _text(row[index])
                if not cell:
                    continue
                symbol, _, part = cell.partition("~")
                part = part or symbol
                build_columns[index] = part
                builds.append(Build(part_number=part, symbol=symbol))
            current = Harness(name=name, def_id=def_id, builds=builds)
            harnesses[def_id or name] = current
            continue

        if current is None or len(row) <= COL_CIRCUIT or not _text(row[COL_CIRCUIT]):
            continue

        present = frozenset(
            part
            for index, part in build_columns.items()
            if index < len(row) and _text(row[index])
        )
        ends.append(
            CircuitEnd(
                harness_id=current.def_id or current.name,
                circuit=_text(row[COL_CIRCUIT]),
                suffix=_text(row[COL_SUFFIX]) if len(row) > COL_SUFFIX else "",
                connector=_text(row[COL_CNUM]) if len(row) > COL_CNUM else "",
                cavity=normalise_cavity(row[COL_CAV]) if len(row) > COL_CAV else "",
                device=_text(row[COL_DEVICE]) if len(row) > COL_DEVICE else "",
                sales_code=(
                    _text(row[COL_SALES_CODE]) if len(row) > COL_SALES_CODE else ""
                ),
                size=_text(row[COL_SIZE]) if len(row) > COL_SIZE else "",
                material=_text(row[COL_MATERIAL]) if len(row) > COL_MATERIAL else "",
                color=_text(row[COL_COLOR]) if len(row) > COL_COLOR else "",
                builds=present,
                source_row=number,
            )
        )

    if not harnesses:
        raise SpliceInputError(
            "No harness blocks were found. A Circuit Summary block starts with a row "
            "whose first cell reads '<Harness> - <id>' and whose second cell is "
            "'Circuit'."
        )
    logger.info(
        "Circuit Summary %s: %d harnesses, %d circuit ends",
        filename, len(harnesses), len(ends),
    )
    return harnesses, ends


def is_inline(end: CircuitEnd) -> bool:
    """Inline ends are named by their device, not by their connector.

    Connector names alone would miss ``I350X`` and ``I41A``, which do not follow
    the X/Y convention.
    """
    return "inline" in (end.device or "").lower()
