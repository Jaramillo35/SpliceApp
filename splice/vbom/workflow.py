from __future__ import annotations

import importlib.util
import os
import re
import tempfile
from pathlib import Path
from typing import Callable, Iterable

import pandas as pd

#: Progress callback: ``(fraction 0..1, human-readable stage)``.
ProgressFn = Callable[[float, str], None]


class _Progress:
    """Reports workflow progress, and never lets a UI callback break the run.

    Stage fractions are calibrated to observed cost, not spaced evenly: reading
    the complexity files and the VIN/harness matching dominate a real run, so
    they own most of the bar.
    """

    def __init__(self, callback: ProgressFn | None) -> None:
        self._callback = callback

    def __call__(self, fraction: float, message: str) -> None:
        if self._callback is None:
            return
        try:
            self._callback(max(0.0, min(1.0, float(fraction))), str(message))
        except Exception:  # noqa: BLE001 - progress must never fail the workflow
            pass


def _build_short_sheet_name(base_name: str) -> str:
    stem = Path(base_name).stem
    tokens = [token for token in re.split(r"[^A-Za-z0-9]+", stem) if token]
    cleaned: list[str] = []
    for token in tokens:
        token_up = token.upper()
        if token_up in {"HARNESS", "COMPLEXITY", "COMPLEX", "INPUT", "OUTPUT", "SHEET", "FILE"}:
            continue
        if re.fullmatch(r"2[678][A-Z0-9]{2}", token_up):
            continue
        if re.fullmatch(r"[VX][0-9]", token_up):
            continue
        cleaned.append(token)

    name = "_".join(cleaned[:3]).strip("_") if cleaned else stem
    name = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")
    return (name[:31] if len(name) > 31 else name) or "Harness"


def _style_worksheet(worksheet):
    """Apply the standard header/body styling and column widths.

    Performance note: every loop here walks the sheet with ``iter_rows`` and
    reuses one style object per role. Indexing a row as ``worksheet[n]``
    re-resolves the sheet bounds on every call, which made this function
    superlinear — a 120k-cell sheet took ~32s that way versus ~1.6s here, and a
    44-harness master workbook could stall for hours. The styles produced are
    byte-for-byte identical either way.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="0B3D66")
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="center", wrap_text=True)

    max_row, max_col = worksheet.max_row, worksheet.max_column

    for row in worksheet.iter_rows(min_row=1, max_row=min(max_row, 1)):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

    # Column widths from the longest rendered value, measured in one pass.
    widths = [10] * max_col
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1,
                                   max_col=max_col, values_only=True):
        for idx, value in enumerate(row):
            if value is not None:
                length = len(str(value))
                if length > widths[idx]:
                    widths[idx] = length
    for col_idx in range(1, max_col + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = \
            min(max(12, widths[col_idx - 1] + 2), 60)

    worksheet.freeze_panes = "A2" if max_row > 1 else "A1"
    if max_row > 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for row in worksheet.iter_rows(min_row=2, max_row=max_row, min_col=1,
                                   max_col=max_col):
        for cell in row:
            cell.alignment = body_alignment
            cell.border = border


def format_workbook_output(path: str | os.PathLike[str], on_sheet=None) -> str:
    """Style every sheet of a workbook in place.

    ``on_sheet(index, total, title)`` is called before each sheet is styled so
    a caller can report progress. Mirrors the legacy engine's entry point.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    workbook = load_workbook(path)
    total = len(workbook.worksheets)
    for index, worksheet in enumerate(workbook.worksheets, start=1):
        if on_sheet is not None:
            try:
                on_sheet(index, total, worksheet.title)
            except Exception:  # noqa: BLE001 - progress must never break saving
                pass
        _style_worksheet(worksheet)

    if "AllCandidates" in workbook.sheetnames:
        ws = workbook["AllCandidates"]
        header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        isbest_col = header_map.get("IsBest")
        if isbest_col is not None:
            highlight_fill = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")
            for row_num in range(2, ws.max_row + 1):
                val = ws.cell(row=row_num, column=isbest_col).value
                is_true = (
                    (isinstance(val, bool) and val)
                    or (isinstance(val, (int, float)) and val == 1)
                    or (isinstance(val, str) and val.strip().upper() == "TRUE")
                )
                if is_true:
                    for col_num in range(1, ws.max_column + 1):
                        ws.cell(row=row_num, column=col_num).fill = highlight_fill

    workbook.save(path)
    return str(path)


def _resolve_vbom_root() -> Path:
    from splice.config import VBOM_ROOT_CANDIDATES

    fallback = None
    for candidate in VBOM_ROOT_CANDIDATES:
        if (candidate / "main_app.py").exists():
            if fallback is None:
                fallback = candidate
            # Prefer an engine that also ships the review VBA project so the
            # web flow can build the same macro-enabled review workbook the
            # desktop app does (older roots lack review_vbaProject.bin).
            if (candidate / "review_vbaProject.bin").exists():
                return candidate
    return fallback or VBOM_ROOT_CANDIDATES[0]


def _load_vbom_module():
    vbom_root = _resolve_vbom_root()
    module_path = vbom_root / "main_app.py"
    if not module_path.exists():
        raise FileNotFoundError(f"Could not find VBOM legacy module at: {module_path}")

    spec = importlib.util.spec_from_file_location("vbom_legacy_main", module_path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Unable to load VBOM legacy module from: {module_path}")

    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _write_uploaded_file(uploaded_file, destination_path: Path) -> Path:
    destination_path.write_bytes(uploaded_file.getvalue())
    return destination_path


def run_vbom_workflow(
    *,
    my: str,
    program: str,
    source_type: str,
    input_upload,
    complexity_uploads: Iterable,
    output_dir: Path | None = None,
    progress: ProgressFn | None = None,
) -> dict:
    report = _Progress(progress)
    report(0.0, "Loading the VBOM engine…")
    vbom_main_app = _load_vbom_module()

    target_dir = output_dir or Path(tempfile.mkdtemp(prefix="vbom_streamlit_", dir=str(Path.cwd())))
    target_dir.mkdir(parents=True, exist_ok=True)

    input_name = getattr(input_upload, "name", "input_file") or "input_file"
    input_path = target_dir / Path(input_name).name
    report(0.03, f"Saving {Path(input_name).name}…")
    _write_uploaded_file(input_upload, input_path)

    complexity_paths = []
    for uploaded_file in complexity_uploads:
        file_name = getattr(uploaded_file, "name", "harness_complexity.xlsx") or "harness_complexity.xlsx"
        destination_path = target_dir / Path(file_name).name
        _write_uploaded_file(uploaded_file, destination_path)
        complexity_paths.append(destination_path)

    if not complexity_paths:
        raise ValueError("At least one harness complexity file is required.")

    source_mode = "buildspec" if source_type.lower() == "buildspec" else "doall"
    report(0.06, f"Building the VIN / sales-code matrix from {Path(input_name).name}…")
    vin_matrix_df, vin_codes_sorted = vbom_main_app.build_vin_matrix(str(input_path), source_type=source_mode)

    per_file_master = []
    per_file_families = []
    all_complexity_codes = set()

    total_complexity = len(complexity_paths)
    for index, complexity_path in enumerate(complexity_paths, start=1):
        report(0.15 + 0.30 * (index - 1) / total_complexity,
               f"Reading harness complexity {index} of {total_complexity} — "
               f"{Path(complexity_path).name}")
        df_comp, header_codes, pn_rows = vbom_main_app.read_complexity_sheet(str(complexity_path))
        family = vbom_main_app.try_get_harness_family(str(complexity_path))
        per_file_master.append((str(complexity_path), df_comp))
        per_file_families.append(
            {
                "family": family,
                "header_codes": set(header_codes),
                "pns": pn_rows,
            }
        )
        all_complexity_codes.update(header_codes)

    report(0.45, "Comparing sales codes across the VIN matrix and the harnesses…")
    vin_code_set = set(vin_codes_sorted)
    complexity_code_set = set(all_complexity_codes)
    diff_df = vbom_main_app.build_salescode_diff(vin_code_set, complexity_code_set)
    excluded_codes = sorted(list(vin_code_set - complexity_code_set))
    excluded_df = pd.DataFrame({"SalesCode_Not_In_Any_Harness": excluded_codes})

    family_stats_df, _global_code_df = vbom_main_app.build_salescode_statistics(
        per_file_families)
    selected_codes_by_family = {
        fam["family"]: set(fam["header_codes"]) for fam in per_file_families
    }
    families_for_matching = vbom_main_app.filter_per_file_families(per_file_families, selected_codes_by_family)

    report(0.52, f"Matching {len(vin_matrix_df)} VIN(s) against "
                 f"{len(families_for_matching)} harness family(ies)…")
    selections_df, all_candidates_df, final_bom_df = vbom_main_app.build_outputs(vin_matrix_df, families_for_matching)

    # Program-qualified naming: every output carries the {MY_last2}_{Program}
    # tag (e.g. 27_RU), matching the DEFE template create_formatted_output emits.
    my_short = my[-2:] if len(my) >= 2 else my
    tag = f"{my_short}_{program}"
    master_path = target_dir / f"Master_Combined_Harness_Complexity_{tag}.xlsx"
    vin_matrix_path = target_dir / f"VIN_Salescode_matrix_{tag}.xlsx"
    selections_path = target_dir / f"VIN_to_Harness_Selection_{tag}.xlsx"

    report(0.72, "Writing the combined master complexity workbook…")
    used_sheetnames = set()
    with pd.ExcelWriter(master_path, engine="openpyxl") as writer:
        for complexity_path, df_comp in per_file_master:
            sheet_name = vbom_main_app.safe_sheetname(Path(complexity_path).name, used_sheetnames)
            df_comp.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    report(0.78, "Writing the VIN sales-code matrix…")
    vin_matrix_df.to_excel(vin_matrix_path, index=False, engine="openpyxl")
    vbom_main_app.write_df_to_excel_append(str(vin_matrix_path), "SalesCode_Diff", diff_df)

    report(0.83, "Writing the harness selections workbook…")
    with pd.ExcelWriter(selections_path, engine="openpyxl") as writer:
        selections_df.to_excel(writer, sheet_name="Selections", index=False)
        all_candidates_df.to_excel(writer, sheet_name="AllCandidates", index=False)
        excluded_df.to_excel(writer, sheet_name="Excluded_SalesCodes", index=False)
        final_bom_df.to_excel(writer, sheet_name="Final_BOM_By_VIN", index=False)
        if not family_stats_df.empty:
            family_stats_df.to_excel(writer, sheet_name="Family_Code_Stats", index=False)
        # Global_Code_Overview is deliberately not emitted: it restated the
        # per-family stats programme-wide and no one worked from it.

    # Formatting walks every cell of every sheet, and the master workbook has
    # one sheet per harness — so this reports per sheet rather than sitting on
    # a single fraction for the whole step.
    def _format_progress(low: float, high: float, book: str):
        def on_sheet(index: int, total: int, title: str) -> None:
            report(low + (high - low) * (index - 1) / max(total, 1),
                   f"Formatting {book} — sheet {index} of {total} ({title})")
        return on_sheet

    vbom_main_app.format_workbook_output(
        str(master_path), _format_progress(0.86, 0.90, "the master workbook"))
    vbom_main_app.format_workbook_output(
        str(selections_path), _format_progress(0.90, 0.93, "the selections workbook"))

    # Match the desktop workflow: emit the macro-enabled SE review workbook and
    # WITHHOLD the DEFE template. The reviewer resolves every flagged selection,
    # then generates {tag}_VBOM_Template_for_DEFE.xlsx from inside the workbook
    # via its Generate DEFE Template macro button (Config!B5 carries the name).
    vbom_root = _resolve_vbom_root()
    template_path = vbom_root / vbom_main_app.TEMPLATE_SOURCE_FILE
    vba_project_path = vbom_root / vbom_main_app.REVIEW_VBA_PROJECT_FILE
    defe_output_name = f"{tag}_VBOM_Template_for_DEFE.xlsx"
    review_path = target_dir / f"Harness_Selection_Review_{tag}.xlsm"
    report(0.93, "Building the SE review cases…")
    review_df = vbom_main_app.build_selection_review_cases(
        selections_df, all_candidates_df, families_for_matching
    )
    report(0.96, "Writing the macro-enabled review workbook…")
    vbom_main_app.create_selection_review_workbook(
        str(review_path),
        review_df,
        selections_df,
        str(template_path),
        str(vba_project_path),
        defe_output_name=defe_output_name,
    )

    report(1.0, "Done")
    metrics_stats = {
        # rows_read combines primary input matrix rows and all harness complexity sheet rows.
        "rows_read": int(len(vin_matrix_df) + sum(len(df_comp) for _, df_comp in per_file_master)),
        # rows_processed reflects generated engineering output row totals.
        "rows_processed": int(len(selections_df) + len(all_candidates_df) + len(final_bom_df)),
        "circuits_processed": None,
        "harness_variants_processed": int(len(per_file_families)),
        "validation_warnings": int(len(excluded_df)),
    }

    return {
        "output_dir": target_dir,
        "master_path": master_path,
        "vin_matrix_path": vin_matrix_path,
        "selections_path": selections_path,
        "review_path": review_path,
        "review_case_count": int(len(review_df)),
        "defe_output_name": defe_output_name,
        "metrics_stats": metrics_stats,
        # In-memory frames for the in-app review gate (splice.vbom.review):
        # the same data the review workbook carries, without a file round-trip.
        "review_df": review_df,
        "selections_df": selections_df,
        "vin_matrix_df": vin_matrix_df,
        "my": my,
        "program": program,
    }
