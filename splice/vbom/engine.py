"""The VBOM Risk Matrix engine — every function the server needs, and none
of the desktop window.

Lifted verbatim from ``vbom_legacy/main_app.py`` on 2026-09-02. That file
was a tkinter desktop app with the engine inside it, and the web workflow
loaded the whole thing by file path with importlib to reach thirteen of
these functions — a server-side engine that depended on a desktop GUI
module. The desktop app still exists and imports this module, so there is
one copy of the logic and the two surfaces cannot drift.

Nothing here was edited in the move. The showcase workflow's fourteen
output workbooks hash identically before and after (cell values and
formula text), and the VBOM tests pass unchanged.
"""

import os
import re
import pandas as pd

VIN_COL = "VIN"
SALES_CODES_COL = "Sales Code ( 3 Char)"   # matches your file
INCLUDE_NUMERIC_3CHAR = False   # if False, drop '191', '590', etc
ONLY_ALPHA_3CHAR = False         # if True, keep only A–Z triplets
SPLIT_SUFFIX_X3 = False         # True => RLX3 -> RLX
SAVE_OUTPUT = True
MASTER_FILE_NAME = "Master_Combined_Harness_Complexity.xlsx"
VIN_MATRIX_FILE = "VIN_Salescode_matrix_.xlsx"
SELECTIONS_FILE = "VIN_to_Harness_Selection.xlsx"  # final multi-tab output
REVIEW_FILE = "Harness_Selection_Review.xlsm"
TEMPLATE_SOURCE_FILE = "Template.xlsx"
REVIEW_VBA_PROJECT_FILE = "review_vbaProject.bin"
NOT_APPLICABLE_PN = "N/A"
NOT_APPLICABLE_STATUS = "NOT_APPLICABLE"
COMPLEXITY_ALLOW_ALPHANUMERIC = True   # regex [A-Z0-9]{3} then filters out 3-digit codes
def _parse_drop_files(raw: str):
    """Parse tkdnd payload into local file paths (supports macOS file:// URIs)."""
    from urllib.parse import unquote, urlparse

    if not raw:
        return []

    tokens = re.findall(r'\{[^}]+\}|[^\s]+', raw)
    out = []
    for token in tokens:
        path = token[1:-1] if token.startswith("{") and token.endswith("}") else token
        path = path.strip().strip('"').strip("'")
        if not path:
            continue

        # Finder drops may come as file:// URLs; decode and normalize them.
        if path.startswith("file://"):
            parsed = urlparse(path)
            path = unquote(parsed.path)
            if os.name == "nt" and path.startswith("/") and len(path) > 2 and path[2] == ":":
                path = path[1:]

        path = os.path.normpath(path)
        out.append(path)

    return out
def load_dataframe(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return pd.read_csv(path)
    elif ext in (".xlsx", ".xls", ".xlsm"):
        engine = "openpyxl" if ext in (".xlsx", ".xlsm") else "xlrd"
        return pd.read_excel(path, engine=engine)
    else:
        raise ValueError(f"Unsupported file type: {ext}")
def write_df_to_excel_append(path: str, sheet_name: str, df: pd.DataFrame):
    """Append/overwrite a sheet in an xlsx file (engine=openpyxl)."""
    from openpyxl import load_workbook
    if not os.path.exists(path):
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
        wb.save(path)
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
def _build_short_sheet_name(base_name: str) -> str:
    stem = os.path.splitext(os.path.basename(base_name))[0]
    tokens = [token for token in re.split(r"[^A-Za-z0-9]+", stem) if token]
    cleaned: list[str] = []
    for token in tokens:
        token_up = token.upper()
        if token_up in {"HARNESS", "COMPLEXITY", "COMPLEX", "INPUT", "OUTPUT", "SHEET", "FILE"}:
            continue
        if re.fullmatch(r"2[678][A-Z0-9]{2}", token_up):
            continue
        if re.fullmatch(r"[VX][0-9]", token_up):
            continue
        cleaned.append(token)

    name = "_".join(cleaned[:3]).strip("_") if cleaned else stem
    name = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")
    return (name[:31] if len(name) > 31 else name) or "Harness"
def _style_worksheet(worksheet):
    """Apply the standard header/body styling and column widths.

    Performance note: every loop here walks the sheet with ``iter_rows`` and
    reuses one style object per role. Indexing a row as ``worksheet[n]``
    re-resolves the sheet bounds on every call, which made this function
    superlinear — a 120k-cell sheet took ~32s that way versus ~1.6s here, and a
    44-harness master workbook could stall for hours. The styles produced are
    byte-for-byte identical either way.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="0B3D66")
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="center", wrap_text=True)

    max_row, max_col = worksheet.max_row, worksheet.max_column

    for row in worksheet.iter_rows(min_row=1, max_row=min(max_row, 1)):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

    # Column widths from the longest rendered value, measured in one pass.
    widths = [10] * max_col
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1,
                                   max_col=max_col, values_only=True):
        for idx, value in enumerate(row):
            if value is not None:
                length = len(str(value))
                if length > widths[idx]:
                    widths[idx] = length
    for col_idx in range(1, max_col + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = \
            min(max(12, widths[col_idx - 1] + 2), 60)

    worksheet.freeze_panes = "A2" if max_row > 1 else "A1"
    if max_row > 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for row in worksheet.iter_rows(min_row=2, max_row=max_row, min_col=1,
                                   max_col=max_col):
        for cell in row:
            cell.alignment = body_alignment
            cell.border = border
def format_workbook_output(path: str | os.PathLike[str], on_sheet=None) -> str:
    """Style every sheet of a workbook in place.

    ``on_sheet(index, total, title)`` is called before each sheet is styled, so
    a caller can report progress — a 44-harness master workbook is many sheets
    and the step must not look frozen.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    workbook = load_workbook(path)
    total = len(workbook.worksheets)
    for index, worksheet in enumerate(workbook.worksheets, start=1):
        if on_sheet is not None:
            try:
                on_sheet(index, total, worksheet.title)
            except Exception:  # noqa: BLE001 - progress must never break saving
                pass
        _style_worksheet(worksheet)

    if "AllCandidates" in workbook.sheetnames:
        ws = workbook["AllCandidates"]
        header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        isbest_col = header_map.get("IsBest")
        if isbest_col is not None:
            highlight_fill = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")
            for row_num in range(2, ws.max_row + 1):
                val = ws.cell(row=row_num, column=isbest_col).value
                is_true = (
                    (isinstance(val, bool) and val)
                    or (isinstance(val, (int, float)) and val == 1)
                    or (isinstance(val, str) and val.strip().upper() == "TRUE")
                )
                if is_true:
                    for col_num in range(1, ws.max_column + 1):
                        ws.cell(row=row_num, column=col_num).fill = highlight_fill

    workbook.save(path)
    return str(path)
def safe_sheetname(base: str, used: set):
    """Return a safe/unique Excel sheet name (<=31 chars)."""
    name = _build_short_sheet_name(base)
    candidate = name; i = 1
    while candidate in used:
        suffix = f"_{i}"
        candidate = (name[: (31 - len(suffix))] + suffix) if len(name) + len(suffix) > 31 else name + suffix
        i += 1
    used.add(candidate)
    return candidate
def _is_nullish(x):
    if x is None:
        return True
    try:
        return pd.isna(x)
    except Exception:
        return False
def parse_sales_codes(cell):
    """
    Normalize one cell from 'Sales Code ( 3 Char)' into a list of 3-char codes.
    Respects INCLUDE_NUMERIC_3CHAR / ONLY_ALPHA_3CHAR.
    """
    if isinstance(cell, (list, tuple, set)):
        items = []
        for it in cell:
            items.extend(parse_sales_codes(str(it)))
        # de-dup in-order
        seen, dedup = set(), []
        for c in items:
            if c not in seen:
                seen.add(c); dedup.append(c)
        return dedup

    if _is_nullish(cell):
        return []

    text = str(cell)
    tokens = []
    for raw in text.split():
        t = re.sub(r'^[^A-Za-z0-9]+|[^A-Za-z0-9]+$', '', raw)
        if not t:
            continue
        if SPLIT_SUFFIX_X3 and len(t) == 4 and t.upper().endswith("X3"):
            t = t[:3]
        t_up = t.upper()
        if len(t_up) != 3:
            continue
        if ONLY_ALPHA_3CHAR and not re.fullmatch(r'[A-Z]{3}', t_up):
            continue
        if not INCLUDE_NUMERIC_3CHAR and t_up.isdigit():
            continue
        if not ONLY_ALPHA_3CHAR and not re.fullmatch(r'[A-Z0-9]{3}', t_up):
            continue
        tokens.append(t_up)

    seen, dedup = set(), []
    for c in tokens:
        if c not in seen:
            seen.add(c); dedup.append(c)
    return dedup
def derive_sorted_unique_codes(series: pd.Series):
    code_set = set()
    for cell in series:
        items = cell if isinstance(cell, (list, tuple, set)) else parse_sales_codes(cell)
        for c in items:
            code_set.add(c)
    return sorted(code_set)
APPLICABLE_MARKS = frozenset({"X", "O"})
GIVEAWAY_MARK = "G"
def read_complexity_sheet(path: str):
    """
    Returns:
      df_complexity: raw Complexity sheet (header=None)
      header_codes: sorted unique 3-char codes found in the FIRST ROW
            rows: list of (pn, set_of_codes, set_of_giveaway_codes)
                        where:
                            set_of_codes          => cells marked 'X'
                            set_of_giveaway_codes => cells marked 'G'
    """
    ext = os.path.splitext(path)[1].lower()
    engine = "openpyxl" if ext in (".xlsx", ".xlsm") else "xlrd"
    df = pd.read_excel(path, sheet_name="Complexity", header=None, engine=engine)

    # find header codes from row 0
    header_values = df.iloc[0].tolist()
    code_regex = r'[A-Z0-9]{3}' if COMPLEXITY_ALLOW_ALPHANUMERIC else r'[A-Z]{3}'
    colidx_to_code = {}
    for j, v in enumerate(header_values):
        if _is_nullish(v):
            continue
        txt = str(v).upper()
        m = re.search(code_regex, txt)
        if m:
            code = m.group(0)
            # Drop numeric-only 3-char codes like 108, 110.
            if code.isdigit():
                continue
            colidx_to_code[j] = code
    header_codes = sorted(set(colidx_to_code.values()))

    # pick PNs and their 'X' / 'G' codes
    pn_pattern = re.compile(r'^\d{8}[A-Z]{2}$')  # e.g., 68720520AA
    rows = []
    for i in range(1, len(df)):
        a0 = df.iat[i, 0]
        if _is_nullish(a0):
            continue
        pn = str(a0).strip()
        if not pn_pattern.match(pn):
            continue
        pn_codes = set()
        pn_giveaway_codes = set()
        row_vals = df.iloc[i].tolist()
        for j, val in enumerate(row_vals):
            if j in colidx_to_code:
                val_up = (val if isinstance(val, str) else ("" if _is_nullish(val) else str(val))).strip().upper()
                # "O" is a hand-entered stand-in for "X" that turns up in real
                # complexity files. It means the part carries the code, so it
                # is normalised here rather than being silently dropped, which
                # used to make a harness vanish from VINs that needed it.
                if val_up in APPLICABLE_MARKS:
                    pn_codes.add(colidx_to_code[j])
                elif val_up == GIVEAWAY_MARK:
                    pn_giveaway_codes.add(colidx_to_code[j])
        rows.append((pn, pn_codes, pn_giveaway_codes))

    return df, header_codes, rows
def try_get_harness_family(path: str):
    """
    Try to get HarnessFamily from 'Harness PN' tab near the 'Harness:' label,
    fallback to the filename segment.
    """
    base = os.path.basename(path)
    m = re.match(r'^Harness_Complexity_[^_]+_[^_]+_([^_]+)_', base)
    fallback = m.group(1) if m else os.path.splitext(base)[0]
    try:
        ext = os.path.splitext(path)[1].lower()
        engine = "openpyxl" if ext in (".xlsx", ".xlsm") else "xlrd"
        df_hp = pd.read_excel(path, sheet_name="Harness PN", header=None, engine=engine)
        for i in range(min(200, df_hp.shape[0])):
            for j in range(min(20, df_hp.shape[1])):
                val = df_hp.iat[i, j]
                if isinstance(val, str) and val.strip().lower() == "harness:":
                    if j + 1 < df_hp.shape[1]:
                        v = df_hp.iat[i, j+1]
                        if isinstance(v, str) and v.strip():
                            return v.strip().upper()
                    if i + 1 < df_hp.shape[0]:
                        v = df_hp.iat[i+1, j]
                        if isinstance(v, str) and v.strip():
                            return v.strip().upper()
        return fallback.upper()
    except Exception:
        return fallback.upper()
def _build_vin_matrix_from_doall(vin_file_path: str):
    df = load_dataframe(vin_file_path)
    missing = [c for c in (VIN_COL, SALES_CODES_COL) if c not in df.columns]
    if missing:
        raise KeyError(f"Missing required column(s): {missing}. Columns found: {list(df.columns)}")
    df = df.copy()
    df["_parsed_codes"] = df[SALES_CODES_COL].apply(parse_sales_codes)
    sorted_unique_codes = derive_sorted_unique_codes(df["_parsed_codes"])

    rows = []
    for _, row in df.iterrows():
        row_data = {VIN_COL: row[VIN_COL]}
        present = set(row["_parsed_codes"])
        for code in sorted_unique_codes:
            row_data[code] = '|' if code in present else ''
        rows.append(row_data)
    matrix_df = pd.DataFrame(rows)
    return matrix_df, sorted_unique_codes
def _is_buildspec_marked(val) -> bool:
    if _is_nullish(val):
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val != 0
    s = str(val).strip().upper()
    if s in ("", "0", "N", "NO", "FALSE", "-"):
        return False
    return True
def _build_vin_matrix_from_buildspec(path: str):
    ext = os.path.splitext(path)[1].lower()
    engine = "openpyxl" if ext in (".xlsx", ".xlsm") else "xlrd"
    df = pd.read_excel(path, header=None, engine=engine)

    # User-defined BuildSpec layout:
    # - Row 7 contains VIN values
    # - Row 6 contains MVON values
    # - Column A contains SalesCodes
    # Prefer VIN when present, else use MVON.
    vin_row_idx = 6
    mvon_row_idx = 5
    code_col_idx = 0
    data_start_row = 7

    if df.shape[0] <= vin_row_idx:
        raise ValueError("BuildSpec does not contain required VIN/MVON rows (rows 6 and 7).")
    if df.shape[1] <= 1:
        raise ValueError("BuildSpec does not contain VIN/MVON data columns.")

    vin_values = df.iloc[vin_row_idx].tolist() if df.shape[0] > vin_row_idx else []
    mvon_values = df.iloc[mvon_row_idx].tolist() if df.shape[0] > mvon_row_idx else []

    vin_to_codes = {}

    for col in range(1, df.shape[1]):
        vin_raw = vin_values[col] if col < len(vin_values) else None
        mvon_raw = mvon_values[col] if col < len(mvon_values) else None

        vin_text = "" if _is_nullish(vin_raw) else str(vin_raw).strip()
        mvon_text = "" if _is_nullish(mvon_raw) else str(mvon_raw).strip()
        vin_id = vin_text if vin_text else mvon_text
        if not vin_id:
            continue
        vin_id_upper = vin_id.strip().upper()
        if vin_id_upper in {"VIN", "MVON", "SALES CODE", "SALESCODE", "SALESCODES"}:
            continue

        code_set = vin_to_codes.setdefault(vin_id, set())

        for r in range(data_start_row, df.shape[0]):
            code_cell = df.iat[r, code_col_idx] if code_col_idx < df.shape[1] else None
            parsed_codes = parse_sales_codes(code_cell)
            if not parsed_codes:
                continue
            marker = df.iat[r, col]
            if _is_buildspec_marked(marker):
                for c in parsed_codes:
                    code_set.add(c)

    if not vin_to_codes:
        raise ValueError("No VIN/MVON and SalesCode assignments were found in BuildSpec.")

    all_codes = sorted({code for codes in vin_to_codes.values() for code in codes})
    rows = []
    for vin_id in sorted(vin_to_codes.keys()):
        present = vin_to_codes[vin_id]
        row_data = {VIN_COL: vin_id}
        for code in all_codes:
            row_data[code] = "|" if code in present else ""
        rows.append(row_data)

    return pd.DataFrame(rows), all_codes
def build_vin_matrix(vin_file_path: str, source_type: str = "doall"):
    if source_type == "buildspec":
        return _build_vin_matrix_from_buildspec(vin_file_path)
    return _build_vin_matrix_from_doall(vin_file_path)
def build_salescode_diff(vin_codes_set, complexity_codes_set):
    vin_not_in_complex = sorted(vin_codes_set - complexity_codes_set)
    complex_not_in_vin = sorted(complexity_codes_set - vin_codes_set)
    max_len = max(len(vin_not_in_complex), len(complex_not_in_vin))
    vin_col = vin_not_in_complex + [""] * (max_len - len(vin_not_in_complex))
    comp_col = complex_not_in_vin + [""] * (max_len - len(complex_not_in_vin))
    return pd.DataFrame({
        "Salescode_not_in_Complexity": vin_col,
        "Salescode_not_in_SPEC": comp_col
    })
def build_salescode_statistics(per_file_complexity: list):
    """
    For each harness family compute, per sales code:
      - how many PNs contain it
      - coverage % (PNsWithCode / TotalPNs * 100)
    - status: STANDARD (100%) | OPTIONAL (>0 and <100%) | UNUSED (0%)

        Also produces a global overview across all families, including a single
        classification bucket per SalesCode:
            - STANDARD_IN_ALL_HARNESSES
            - STANDARD_IN_SOME_HARNESSES
            - OPTIONAL_IN_ALL_HARNESSES
            - OTHER

    Returns:
        family_stats_df  – one row per (HarnessFamily, SalesCode)
        global_df        – one row per SalesCode summarising cross-family presence
                   and final classification bucket
        (no pivot returned)
    """
    family_stats_rows = []
    for fam in per_file_complexity:
        family_name = fam["family"]
        pns = fam["pns"]          # list of (pn, set_of_codes, set_of_giveaway_codes)
        total_pns = len(pns)
        if total_pns == 0:
            continue

        # count how many PNs include each code.
        # G (giveaway) means the code is present on that PN too.
        code_counts = {}
        for _pn, codes, _giveaway_codes in pns:
            effective_codes = set(codes) | set(_giveaway_codes)
            for code in effective_codes:
                code_counts[code] = code_counts.get(code, 0) + 1

        # also register header codes that appear in 0 PNs
        for code in fam["header_codes"]:
            if code not in code_counts:
                code_counts[code] = 0

        for code in sorted(code_counts):
            count = code_counts[code]
            pct = round(count / total_pns * 100, 1)
            if pct == 100.0:
                status = "STANDARD"
            elif pct == 0.0:
                status = "UNUSED"
            else:
                status = "OPTIONAL"
            family_stats_rows.append({
                "HarnessFamily": family_name,
                "SalesCode": code,
                "PNsWithCode": count,
                "TotalPNs": total_pns,
                "Coverage_Pct": pct,
                "Status": status,
            })

    family_stats_df = pd.DataFrame(family_stats_rows, columns=[
        "HarnessFamily", "SalesCode", "PNsWithCode", "TotalPNs", "Coverage_Pct", "Status"
    ])

    if family_stats_df.empty:
        return family_stats_df, pd.DataFrame()

    all_families = [fam["family"] for fam in per_file_complexity]
    total_families = len(all_families)
    all_codes = sorted(family_stats_df["SalesCode"].unique())

    global_rows = []
    for code in all_codes:
        sub = family_stats_df[family_stats_df["SalesCode"] == code].copy()
        status_map = dict(zip(sub["HarnessFamily"], sub["Status"]))
        statuses = [status_map.get(fam, "UNUSED") for fam in all_families]

        standard_cnt = sum(1 for s in statuses if s == "STANDARD")
        optional_cnt = sum(1 for s in statuses if s == "OPTIONAL")
        unused_cnt = sum(1 for s in statuses if s == "UNUSED")
        fams_present = [fam for fam in all_families if status_map.get(fam, "UNUSED") != "UNUSED"]

        if standard_cnt == total_families:
            classification = "STANDARD_IN_ALL_HARNESSES"
        elif 0 < standard_cnt < total_families:
            classification = "STANDARD_IN_SOME_HARNESSES"
        elif optional_cnt == total_families:
            classification = "OPTIONAL_IN_ALL_HARNESSES"
        else:
            classification = "OTHER"

        global_rows.append({
            "SalesCode": code,
            "FamiliesPresent": len(fams_present),
            "TotalFamilies": total_families,
            "FamilyCoverage_Pct": round(len(fams_present) / total_families * 100, 1),
            "StandardInFamilies": standard_cnt,
            "OptionalInFamilies": optional_cnt,
            "UnusedInFamilies": unused_cnt,
            "Classification": classification,
            "FamilyList": ", ".join(sorted(fams_present)),
        })

    global_df = pd.DataFrame(global_rows, columns=[
        "SalesCode", "FamiliesPresent", "TotalFamilies",
        "FamilyCoverage_Pct", "StandardInFamilies", "OptionalInFamilies",
        "UnusedInFamilies", "Classification", "FamilyList"
    ])

    return family_stats_df, global_df
def _format_stats_sheets(wb_path: str):
    """Apply conditional color formatting to the three SalesCode statistics sheets."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    STATUS_FILLS = {
        "STANDARD": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "OPTIONAL": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "UNUSED":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    wb = load_workbook(wb_path)

    # Family_Code_Stats: color every row by Status
    if "Family_Code_Stats" in wb.sheetnames:
        ws = wb["Family_Code_Stats"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        hdr = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        status_col = hdr.get("Status")
        if status_col:
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=status_col).value
                fill = STATUS_FILLS.get(str(val).upper() if val else "", None)
                if fill:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill

    # Global_Code_Overview: freeze + autofilter only
    if "Global_Code_Overview" in wb.sheetnames:
        ws = wb["Global_Code_Overview"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Highlight the requested single-view categories.
        from openpyxl.styles import PatternFill
        cls_fills = {
            "STANDARD_IN_ALL_HARNESSES": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "STANDARD_IN_SOME_HARNESSES": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "OPTIONAL_IN_ALL_HARNESSES": PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid"),
        }
        hdr = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        cls_col = hdr.get("Classification")
        if cls_col:
            for r in range(2, ws.max_row + 1):
                cls = ws.cell(row=r, column=cls_col).value
                fill = cls_fills.get(str(cls).upper() if cls else "")
                if fill:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill

    wb.save(wb_path)
def compute_status_score(req_cnt, matched, extra, missing):
    if missing == 0 and extra == 0:
        status = "EXACT"
    elif missing == 0 and extra > 0:
        status = "OVERBUILT"
    else:
        status = "INCOMPLETE"
    score = matched - extra - 100 * missing
    return status, score
def _variant_base_family(name: str):
    """Normalize family name to group Pacifica/Voyager variants."""
    if not isinstance(name, str):
        return str(name)
    s = name.upper()
    # Remove PACIFICA and VOYAGER variants (handles both word boundaries and underscores)
    s = re.sub(r'[_\s]*(PACIFICA|VOYAGER)[_\s]*', ' ', s)
    # Normalize whitespace
    s = re.sub(r'[\s_]+', '_', s).strip().rstrip('_')
    return s
def _family_matching_codes(fam: dict) -> tuple[set, set]:
    """
    Return (matching_codes, standard_codes) for one harness family.

    Codes present on every PN are standard/common and do not distinguish
    applicability or PN choice when the family also has optional codes.
    If the family has no optional codes (for example a one-PN Trailer Tow
    family), fall back to its standard codes so the family can still be
    recognized as applicable. Header codes present on no PN are ignored.
    """
    header_codes = set(fam["header_codes"])
    pns = fam["pns"]
    if not pns:
        return set(), set()

    code_counts = {code: 0 for code in header_codes}
    for _pn, pn_codes, giveaway_codes in pns:
        effective_codes = (set(pn_codes) | set(giveaway_codes)) & header_codes
        for code in effective_codes:
            code_counts[code] += 1

    total_pns = len(pns)
    standard_codes = {code for code, count in code_counts.items() if count == total_pns}
    optional_codes = {code for code, count in code_counts.items() if 0 < count < total_pns}
    matching_codes = optional_codes if optional_codes else standard_codes
    return matching_codes, standard_codes
def build_outputs(vin_matrix_df: pd.DataFrame,
                  per_file_complexity: list):
    """
    per_file_complexity = list of dicts: {
       'family': str,
       'header_codes': set(...),
         'pns': list of (pn, set_of_codes, set_of_giveaway_codes)
    }
    Returns: selections_df, all_candidates_df, final_bom_df
    """
    families = per_file_complexity
    selections_rows = []
    all_candidates_rows = []

    code_columns = [c for c in vin_matrix_df.columns if c != VIN_COL]

    for _, vrow in vin_matrix_df.iterrows():
        vin = vrow[VIN_COL]
        vin_codes = {c for c in code_columns if str(vrow[c]) == '|'}

        # group families by Pacifica/Voyager base to prevent double selection
        grouped = {}
        for fam in families:
            base = _variant_base_family(fam["family"])
            grouped.setdefault(base, []).append(fam)

        for _base, fam_list in grouped.items():
            group_best = None  # (key, family_name, best_pn, best_info)

            applicable_families = []
            for fam in fam_list:
                matching_codes, _standard_codes = _family_matching_codes(fam)
                vin_required_for_family = vin_codes & matching_codes
                if vin_required_for_family:
                    applicable_families.append((fam, vin_required_for_family, matching_codes))

            # A harness is not needed when the VIN has none of the selected
            # sales codes represented by that harness family's complexity.
            # Record the result explicitly so the final BOM is auditable, but
            # do not create PN candidates for a non-applicable family.
            if not applicable_families:
                for fam in fam_list:
                    selections_rows.append({
                        "VIN": vin,
                        "HarnessFamily": fam["family"],
                        "SelectedHarnessPN": NOT_APPLICABLE_PN,
                        "MatchStatus": NOT_APPLICABLE_STATUS,
                        "RequiredCount": 0,
                        "MatchedCount": 0,
                        "MissingCount": 0,
                        "ExtraCount": 0,
                        "RequiredSalesCodes": None,
                        "MatchedSalesCodes": None,
                        "Giveaway": None,
                        "MissingSalesCodes": None,
                        "ExtraSalesCodes": None,
                        "Score": None,
                    })
                continue

            for fam, vin_required_for_family, matching_codes in applicable_families:
                family_name = fam["family"]

                # evaluate all PNs for this family
                best_key = None
                best_pn = None
                best_info = None  # selection details for the winning PN

                pn_evals = []
                for pn, pn_codes, pn_giveaway_codes in fam["pns"]:
                    # Treat both X and G as present for matching semantics.
                    effective_pn_codes = (set(pn_codes) | set(pn_giveaway_codes)) & matching_codes
                    matching_giveaway_codes = set(pn_giveaway_codes) & matching_codes

                    matched_codes = vin_required_for_family & effective_pn_codes
                    missing_codes = vin_required_for_family - effective_pn_codes
                    extra_codes = effective_pn_codes - vin_required_for_family

                    matched = len(matched_codes)
                    missing = len(missing_codes)
                    extra = len(extra_codes)
                    status, score = compute_status_score(len(vin_required_for_family), matched, extra, missing)
                    required_list = ",".join(sorted(vin_required_for_family))
                    matched_list = ",".join(sorted(matched_codes))

                    # sorting key for "best"
                    key = (score, matched, -missing, -extra, -len(extra_codes), pn)
                    pn_evals.append((pn, score, status, matched, missing, extra,
                                     required_list,
                                     matched_list,
                                     ",".join(sorted(missing_codes)),
                                     ",".join(sorted(extra_codes)),
                                     len(vin_required_for_family),
                                     ",".join(sorted(matching_giveaway_codes))))

                    if (best_key is None) or (key > best_key):
                        best_key = key
                        best_pn = pn
                        best_info = (status, len(vin_required_for_family), matched, missing, extra,
                                     required_list,
                                     matched_list,
                                     ",".join(sorted(missing_codes)),
                                     ",".join(sorted(extra_codes)),
                                     score,
                                     ",".join(sorted(matching_giveaway_codes)))

                # add candidates (and mark IsBest later if they win the group)
                for (pn, score, candidate_status, matched_count, missing_count, extra_count,
                     required_list, matched_list, miss_list, extra_list, required_count,
                     giveaway_list) in pn_evals:
                    all_candidates_rows.append({
                        "VIN": vin,
                        "HarnessFamily": family_name,
                        "PN": pn,
                        "Score": score,
                        "IsBest": False,
                        "MatchStatus": candidate_status,
                        "RequiredCount": required_count,
                        "MatchedCount": matched_count,
                        "MissingCount": missing_count,
                        "ExtraCount": extra_count,
                        "RequiredSalesCodes": required_list if required_list else None,
                        "MatchedSalesCodes": matched_list if matched_list else None,
                        "Giveaway": giveaway_list if giveaway_list else None,
                        "MissingSalesCodes": miss_list if miss_list else None,
                        "ExtraSalesCodes": extra_list if extra_list else None
                    })

                if best_info is not None:
                    if (group_best is None) or (best_key > group_best[0]):
                        group_best = (best_key, family_name, best_pn, best_info)

            if group_best is None:
                continue

            # mark best in AllCandidates for the selected family/pn
            _, family_name, best_pn, best_info = group_best
            for row in all_candidates_rows:
                if row["VIN"] == vin and row["HarnessFamily"] == family_name and row["PN"] == best_pn:
                    row["IsBest"] = True

            # add selection (only one per group)
            (status, req_cnt, matched_cnt, missing_cnt, extra_cnt,
             required_list, matched_list, miss_list, extra_list,
             score, giveaway_list) = best_info
            selections_rows.append({
                "VIN": vin,
                "HarnessFamily": family_name,
                "SelectedHarnessPN": best_pn,
                "MatchStatus": status,
                "RequiredCount": req_cnt,
                "MatchedCount": matched_cnt,
                "MissingCount": missing_cnt,
                "ExtraCount": extra_cnt,
                "RequiredSalesCodes": required_list if required_list else None,
                "MatchedSalesCodes": matched_list if matched_list else None,
                "Giveaway": giveaway_list if giveaway_list else None,
                "MissingSalesCodes": miss_list if miss_list else None,
                "ExtraSalesCodes": extra_list if extra_list else None,
                "Score": score
            })

    selections_df = pd.DataFrame(selections_rows, columns=[
        "VIN","HarnessFamily","SelectedHarnessPN","MatchStatus",
        "RequiredCount","MatchedCount","MissingCount","ExtraCount",
        "RequiredSalesCodes","MatchedSalesCodes","Giveaway",
        "MissingSalesCodes","ExtraSalesCodes","Score"
    ])

    all_candidates_df = pd.DataFrame(all_candidates_rows, columns=[
        "VIN","HarnessFamily","PN","Score","IsBest","MatchStatus",
        "RequiredCount","MatchedCount","MissingCount","ExtraCount",
        "RequiredSalesCodes","MatchedSalesCodes","Giveaway",
        "MissingSalesCodes","ExtraSalesCodes"
    ])

    # Final_BOM_By_VIN: pivot selections to VIN x family => PN
    bom = selections_df.pivot(index="VIN", columns="HarnessFamily", values="SelectedHarnessPN")
    bom = bom.reset_index()
    # stable column order: VIN first, then families sorted
    ordered_cols = ["VIN"] + sorted([c for c in bom.columns if c != "VIN"])
    bom = bom[ordered_cols]
    final_bom_df = bom

    return selections_df, all_candidates_df, final_bom_df
def find_same_score_ties(all_candidates_df: pd.DataFrame):
    if all_candidates_df.empty:
        return []
    ties = []
    grouped = all_candidates_df.groupby(["VIN", "HarnessFamily"], dropna=False)
    for (vin, family), grp in grouped:
        if grp.empty:
            continue
        max_score = grp["Score"].max()
        top = grp[grp["Score"] == max_score].copy()
        if len(top) > 1:
            ties.append((vin, family, top.reset_index(drop=True)))
    return ties
def _split_codes_cell(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return []
    s = str(val).strip()
    return [x.strip() for x in s.split(",") if x.strip()]
def apply_tie_break_overrides(selections_df: pd.DataFrame, all_candidates_df: pd.DataFrame, overrides: dict):
    if not overrides:
        return selections_df, all_candidates_df

    selections = selections_df.copy()
    candidates = all_candidates_df.copy()

    for (vin, family), chosen_pn in overrides.items():
        mask_group = (candidates["VIN"] == vin) & (candidates["HarnessFamily"] == family)
        if not mask_group.any():
            continue
        candidates.loc[mask_group, "IsBest"] = False
        mask_chosen = mask_group & (candidates["PN"] == chosen_pn)
        if not mask_chosen.any():
            continue
        candidates.loc[mask_chosen, "IsBest"] = True

        chosen_row = candidates.loc[mask_chosen].iloc[0]
        mask_sel = (selections["VIN"] == vin) & (selections["HarnessFamily"] == family)
        if not mask_sel.any():
            continue

        required = int(selections.loc[mask_sel, "RequiredCount"].iloc[0])
        missing_codes = _split_codes_cell(chosen_row.get("MissingSalesCodes"))
        extra_codes = _split_codes_cell(chosen_row.get("ExtraSalesCodes"))
        missing_cnt = len(missing_codes)
        extra_cnt = len(extra_codes)
        matched_cnt = max(required - missing_cnt, 0)
        if missing_cnt == 0 and extra_cnt == 0:
            status = "EXACT"
        elif missing_cnt == 0:
            status = "OVERBUILT"
        else:
            status = "INCOMPLETE"

        selections.loc[mask_sel, "SelectedHarnessPN"] = chosen_pn
        selections.loc[mask_sel, "MatchStatus"] = status
        selections.loc[mask_sel, "MatchedCount"] = matched_cnt
        selections.loc[mask_sel, "MissingCount"] = missing_cnt
        selections.loc[mask_sel, "ExtraCount"] = extra_cnt
        selections.loc[mask_sel, "RequiredSalesCodes"] = chosen_row.get("RequiredSalesCodes")
        selections.loc[mask_sel, "MatchedSalesCodes"] = chosen_row.get("MatchedSalesCodes")
        selections.loc[mask_sel, "Giveaway"] = chosen_row.get("Giveaway")
        selections.loc[mask_sel, "MissingSalesCodes"] = chosen_row.get("MissingSalesCodes")
        selections.loc[mask_sel, "ExtraSalesCodes"] = chosen_row.get("ExtraSalesCodes")
        selections.loc[mask_sel, "Score"] = chosen_row.get("Score")

    return selections, candidates
def filter_per_file_families(per_file_families: list, selected_codes_by_family: dict) -> list:
    """
    Return a copy of per_file_families where each family's header_codes and
    per-PN code sets are restricted to that family's selected SalesCodes.
    """
    filtered = []
    for fam in per_file_families:
        family_name = fam["family"]
        chosen = selected_codes_by_family.get(family_name)
        if chosen is None:
            chosen = set(fam["header_codes"])
        else:
            chosen = set(chosen)
        filtered.append({
            "family": family_name,
            "header_codes": fam["header_codes"] & chosen,
            "pns": [(pn, codes & chosen, giveaway_codes & chosen) for pn, codes, giveaway_codes in fam["pns"]],
        })
    return filtered
def _review_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()
def _ordered_unique(values):
    seen = set()
    result = []
    for value in values:
        text = _review_text(value)
        if text and text not in seen:
            seen.add(text)
            result.append(text)
    return result
def build_selection_review_cases(selections_df: pd.DataFrame,
                                 all_candidates_df: pd.DataFrame,
                                 per_file_families: list) -> pd.DataFrame:
    """Build the small SE review queue from uncertain VIN/family selections.

    A row requires review when any of these conditions is true:
      - the engine's best candidate is INCOMPLETE;
      - two or more candidates share the best score;
      - N/A was selected although the family has a zero-option/base candidate.
    """
    columns = [
        "ReviewID", "VIN", "HarnessFamily", "ReviewReason",
        "EngineRecommendation", "RequiredSalesCodes", "MissingSalesCodes",
        "ExtraSalesCodes", "Giveaway", "CandidateDetails",
        "AllowedPNs", "SelectedPN", "ReviewerNotes",
    ]
    if selections_df.empty:
        return pd.DataFrame(columns=columns)

    family_defs = {}
    for family in per_file_families:
        family_defs.setdefault(family["family"], []).append(family)

    candidate_groups = {}
    if not all_candidates_df.empty:
        for key, group in all_candidates_df.groupby(["VIN", "HarnessFamily"], dropna=False):
            candidate_groups[(str(key[0]), str(key[1]))] = group.copy()

    review_rows = []
    for _, selection in selections_df.iterrows():
        vin = _review_text(selection.get("VIN"))
        family_name = _review_text(selection.get("HarnessFamily"))
        engine_pn = _review_text(selection.get("SelectedHarnessPN"))
        match_status = _review_text(selection.get("MatchStatus")).upper()
        key = (vin, family_name)
        candidate_group = candidate_groups.get(key, pd.DataFrame())

        reasons = []
        candidate_pns = []

        if match_status == "INCOMPLETE":
            reasons.append("No complete PN covers every required sales code")
            if not candidate_group.empty:
                ordered = candidate_group.sort_values(
                    by=["Score", "MissingCount", "ExtraCount", "PN"],
                    ascending=[False, True, True, True],
                )
                candidate_pns.extend(ordered["PN"].tolist())

        if not candidate_group.empty:
            max_score = candidate_group["Score"].max()
            top = candidate_group[candidate_group["Score"] == max_score]
            if len(top) > 1:
                reasons.append("Multiple PNs share the best score")
                candidate_pns.extend(top.sort_values("PN")["PN"].tolist())

        if match_status == NOT_APPLICABLE_STATUS:
            base_pns = []
            for family_def in family_defs.get(family_name, []):
                matching_codes, standard_codes = _family_matching_codes(family_def)
                has_optional_codes = bool(matching_codes) and matching_codes != standard_codes
                if not has_optional_codes:
                    continue
                for pn, pn_codes, giveaway_codes in family_def["pns"]:
                    effective_matching_codes = (
                        set(pn_codes) | set(giveaway_codes)
                    ) & matching_codes
                    if not effective_matching_codes:
                        base_pns.append(pn)
            if base_pns:
                reasons.append("N/A conflicts with an available base/default PN")
                candidate_pns.extend(sorted(base_pns))

        if not reasons:
            continue

        candidate_pns.append(engine_pn)
        candidate_pns.append(NOT_APPLICABLE_PN)
        candidate_pns = _ordered_unique(candidate_pns)

        detail_lines = []
        for pn in candidate_pns:
            if pn.upper() == NOT_APPLICABLE_PN:
                detail_lines.append(f"{NOT_APPLICABLE_PN} | Harness not required")
                continue
            candidate_rows = (
                candidate_group[candidate_group["PN"].astype(str) == pn]
                if not candidate_group.empty else pd.DataFrame()
            )
            if candidate_rows.empty:
                detail_lines.append(f"{pn} | Base/default candidate")
                continue
            candidate = candidate_rows.iloc[0]
            detail_lines.append(
                f"{pn} | { _review_text(candidate.get('MatchStatus')) or 'CANDIDATE' }"
                f" | Score {_review_text(candidate.get('Score')) or '-'}"
                f" | Missing {_review_text(candidate.get('MissingSalesCodes')) or '-'}"
                f" | Extra {_review_text(candidate.get('ExtraSalesCodes')) or '-'}"
                f" | Giveaway {_review_text(candidate.get('Giveaway')) or '-'}"
            )

        review_rows.append({
            "ReviewID": f"{vin}|{family_name}",
            "VIN": vin,
            "HarnessFamily": family_name,
            "ReviewReason": "; ".join(_ordered_unique(reasons)),
            "EngineRecommendation": engine_pn,
            "RequiredSalesCodes": _review_text(selection.get("RequiredSalesCodes")),
            "MissingSalesCodes": _review_text(selection.get("MissingSalesCodes")),
            "ExtraSalesCodes": _review_text(selection.get("ExtraSalesCodes")),
            "Giveaway": _review_text(selection.get("Giveaway")),
            "CandidateDetails": "\n".join(detail_lines),
            "AllowedPNs": ",".join(candidate_pns),
            "SelectedPN": "",
            "ReviewerNotes": "",
        })

    return pd.DataFrame(review_rows, columns=columns)
def _openpyxl_color_hex(color):
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and len(rgb) >= 6:
        return f"#{rgb[-6:]}"
    return None
def _copy_template_sheet_to_xlsxwriter(workbook, template_path: str):
    """Embed the existing DEFE template as a hidden sheet in the review file."""
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter

    source_workbook = load_workbook(template_path, data_only=False)
    source_sheet = source_workbook.active
    target_sheet = workbook.add_worksheet("DEFE_Template")
    format_cache = {}

    border_style_map = {
        "thin": 1, "medium": 2, "dashed": 3, "dotted": 4,
        "thick": 5, "double": 6, "hair": 7,
        "mediumDashed": 8, "dashDot": 9, "mediumDashDot": 10,
        "dashDotDot": 11, "mediumDashDotDot": 12, "slantDashDot": 13,
    }

    def cell_format(cell):
        style_id = cell.style_id
        if style_id in format_cache:
            return format_cache[style_id]
        properties = {}
        font = cell.font
        if font:
            properties.update({
                "bold": bool(font.bold),
                "italic": bool(font.italic),
                "font_name": font.name or "Calibri",
                "font_size": font.sz or 11,
            })
            font_color = _openpyxl_color_hex(font.color)
            if font_color:
                properties["font_color"] = font_color
        fill_color = _openpyxl_color_hex(cell.fill.fgColor)
        if cell.fill and cell.fill.fill_type == "solid" and fill_color:
            properties["bg_color"] = fill_color
            properties["pattern"] = 1
        alignment = cell.alignment
        if alignment:
            if alignment.horizontal:
                properties["align"] = alignment.horizontal
            if alignment.vertical:
                properties["valign"] = alignment.vertical
            if alignment.wrap_text:
                properties["text_wrap"] = True
            if alignment.text_rotation:
                properties["rotation"] = alignment.text_rotation
        if cell.number_format and cell.number_format != "General":
            properties["num_format"] = cell.number_format
        for side_name in ("left", "right", "top", "bottom"):
            side = getattr(cell.border, side_name)
            if side and side.style:
                properties[side_name] = border_style_map.get(side.style, 1)
                side_color = _openpyxl_color_hex(side.color)
                if side_color:
                    properties[f"{side_name}_color"] = side_color
        result = workbook.add_format(properties)
        format_cache[style_id] = result
        return result

    merged_cells = set()
    for merged_range in source_sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left = source_sheet.cell(min_row, min_col)
        target_sheet.merge_range(
            min_row - 1, min_col - 1, max_row - 1, max_col - 1,
            top_left.value if top_left.value is not None else "",
            cell_format(top_left),
        )
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_cells.add((row, col))

    for row in source_sheet.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or (cell.row, cell.column) in merged_cells:
                continue
            value = cell.value
            fmt = cell_format(cell)
            if isinstance(value, str) and value.startswith("="):
                target_sheet.write_formula(cell.row - 1, cell.column - 1, value, fmt)
            elif value is not None:
                target_sheet.write(cell.row - 1, cell.column - 1, value, fmt)
            elif cell.has_style:
                target_sheet.write_blank(cell.row - 1, cell.column - 1, None, fmt)

    for col_idx in range(1, source_sheet.max_column + 1):
        dimension = source_sheet.column_dimensions.get(get_column_letter(col_idx))
        if dimension and dimension.width:
            target_sheet.set_column(col_idx - 1, col_idx - 1, dimension.width)
    for row_idx, dimension in source_sheet.row_dimensions.items():
        if dimension.height:
            target_sheet.set_row(row_idx - 1, dimension.height)

    target_sheet.very_hidden()
    return target_sheet
def create_selection_review_workbook(output_path: str,
                                     review_df: pd.DataFrame,
                                     selections_df: pd.DataFrame,
                                     template_path: str,
                                     vba_project_path: str,
                                     defe_output_name: str = "Template_for_DEFE.xlsx") -> str:
    """Create the self-contained macro-enabled SE review workbook.

    defe_output_name is written to Config!B5 and is the single source of truth
    for the file name the GenerateTemplateForDEFE macro emits, e.g.
    27_KX_VBOM_Template_for_DEFE.xlsx.
    """
    import xlsxwriter
    from xlsxwriter.utility import xl_col_to_name

    if not os.path.isfile(vba_project_path):
        raise FileNotFoundError(f"Review VBA project not found: {vba_project_path}")
    if not os.path.isfile(template_path):
        raise FileNotFoundError(f"DEFE template not found: {template_path}")

    workbook = xlsxwriter.Workbook(output_path)
    workbook.add_vba_project(vba_project_path)
    workbook.set_calc_mode("auto")
    workbook.set_properties({
        "title": "Harness Selection Review",
        "subject": "SE review of uncertain VIN-to-harness selections",
        "author": "VBOM Generator",
        "comments": "Resolve every flagged selection, then use the macro button to generate Template_for_DEFE.xlsx.",
    })

    navy = "#11314F"
    blue = "#D9EAF7"
    pale_blue = "#EAF3F8"
    amber = "#FFF2CC"
    red = "#FCE4D6"
    green = "#E2F0D9"
    gray = "#E7E6E6"
    white = "#FFFFFF"

    title_fmt = workbook.add_format({
        "bold": True, "font_size": 18, "font_color": white,
        "bg_color": navy, "align": "left", "valign": "vcenter",
    })
    instruction_fmt = workbook.add_format({
        "font_color": "#334E68", "bg_color": pale_blue,
        "text_wrap": True, "valign": "vcenter",
    })
    label_fmt = workbook.add_format({
        "bold": True, "font_color": navy, "bg_color": blue,
        "border": 1, "align": "center", "valign": "vcenter",
    })
    value_fmt = workbook.add_format({
        "bold": True, "font_size": 12, "border": 1,
        "align": "center", "valign": "vcenter",
    })
    header_fmt = workbook.add_format({
        "bold": True, "font_color": white, "bg_color": navy,
        "border": 1, "align": "center", "valign": "vcenter",
        "text_wrap": True,
    })
    text_fmt = workbook.add_format({"border": 1, "valign": "top", "text_wrap": True})
    center_fmt = workbook.add_format({"border": 1, "align": "center", "valign": "top", "text_wrap": True})
    choice_fmt = workbook.add_format({
        "border": 1, "bg_color": amber, "bold": True,
        "align": "center", "valign": "vcenter",
    })
    pending_fmt = workbook.add_format({"bg_color": red, "font_color": "#9C0006", "bold": True})
    resolved_fmt = workbook.add_format({"bg_color": green, "font_color": "#276221", "bold": True})

    review_sheet = workbook.add_worksheet("Review")
    review_sheet.hide_gridlines(2)
    review_sheet.freeze_panes(6, 4)
    review_sheet.set_zoom(85)
    review_sheet.set_row(0, 30)
    review_sheet.merge_range("A1:F1", "Harness Selection Review", title_fmt)
    review_sheet.merge_range(
        "A2:J3",
        "Review only the rows below. Choose one Final PN from each yellow dropdown. "
        "When Pending Reviews reaches 0, select Generate DEFE Template. "
        "N/A is available when the SE confirms the harness is not required.",
        instruction_fmt,
    )
    review_sheet.insert_button("G1", {
        "macro": "GenerateTemplateForDEFE",
        "caption": f"Generate DEFE Template ({defe_output_name})",
        "width": 255,
        "height": 48,
        "font": {"bold": True, "color": white},
        "fill": {"color": "#2E7D32"},
    })

    first_data_row = 6  # zero-based row 7
    last_data_row = first_data_row + max(len(review_df) - 1, 0)
    excel_first_row = first_data_row + 1
    excel_last_row = max(last_data_row + 1, excel_first_row)

    review_sheet.write("A5", "Total Reviews", label_fmt)
    review_sheet.write("B5", len(review_df), value_fmt)
    review_sheet.write("C5", "Pending Reviews", label_fmt)
    review_sheet.write_formula(
        "D5", f'=COUNTIF(G{excel_first_row}:G{excel_last_row},"PENDING")',
        value_fmt, len(review_df),
    )
    review_sheet.write("E5", "Workbook Status", label_fmt)
    review_sheet.write_formula(
        "F5", '=IF(D5=0,"READY TO GENERATE","REVIEW REQUIRED")',
        value_fmt, "REVIEW REQUIRED" if len(review_df) else "READY TO GENERATE",
    )
    review_sheet.merge_range("G5:J5", "Yellow cells are the only required user inputs.", instruction_fmt)

    headers = [
        "Review ID", "VIN", "Harness Family", "Why Review Is Needed",
        "Engine Recommendation", "Final PN (SE Selection)", "Review Status",
        "Required Codes", "Missing Codes", "Extra Codes", "Giveaway Codes",
        "Candidate Details", "SE Notes", "Allowed PNs",
    ]
    for col, header in enumerate(headers):
        review_sheet.write(5, col, header, header_fmt)

    widths = [4, 19, 22, 27, 17, 20, 14, 22, 17, 17, 17, 66, 30, 4]
    for col, width in enumerate(widths):
        review_sheet.set_column(col, col, width)
    review_sheet.set_column(0, 0, None, None, {"hidden": True})
    review_sheet.set_column(13, 13, None, None, {"hidden": True})

    candidate_sheet = workbook.add_worksheet("Candidate_Options")
    candidate_sheet.hide()
    for row_offset, (_, review_row) in enumerate(review_df.iterrows()):
        target_row = first_data_row + row_offset
        values = [
            review_row["ReviewID"], review_row["VIN"], review_row["HarnessFamily"],
            review_row["ReviewReason"], review_row["EngineRecommendation"],
            "", "", review_row["RequiredSalesCodes"],
            review_row["MissingSalesCodes"], review_row["ExtraSalesCodes"],
            review_row["Giveaway"], review_row["CandidateDetails"],
            "", review_row["AllowedPNs"],
        ]
        for col, value in enumerate(values):
            fmt = choice_fmt if col == 5 else (center_fmt if col in (1, 2, 4, 6, 7, 8, 9, 10) else text_fmt)
            review_sheet.write(target_row, col, value, fmt)
        excel_row = target_row + 1
        review_sheet.write_formula(
            target_row, 6, f'=IF(F{excel_row}="","PENDING","RESOLVED")',
            center_fmt, "PENDING",
        )
        review_sheet.set_row(target_row, 62)

        options = [value for value in str(review_row["AllowedPNs"]).split(",") if value]
        option_col = row_offset
        candidate_sheet.write(0, option_col, review_row["ReviewID"])
        for option_row, option in enumerate(options, start=1):
            candidate_sheet.write(option_row, option_col, option)
        col_letter = xl_col_to_name(option_col)
        review_sheet.data_validation(target_row, 5, target_row, 5, {
            "validate": "list",
            "source": f"='Candidate_Options'!${col_letter}$2:${col_letter}${len(options) + 1}",
            "input_title": "Select final PN",
            "input_message": "Select the correct PN or N/A.",
            "error_title": "Invalid PN",
            "error_message": "Choose a PN from the approved candidate list.",
        })

    if len(review_df):
        review_sheet.autofilter(5, 0, last_data_row, len(headers) - 1)
        review_sheet.conditional_format(first_data_row, 6, last_data_row, 6, {
            "type": "text", "criteria": "containing", "value": "PENDING", "format": pending_fmt,
        })
        review_sheet.conditional_format(first_data_row, 6, last_data_row, 6, {
            "type": "text", "criteria": "containing", "value": "RESOLVED", "format": resolved_fmt,
        })
    else:
        review_sheet.merge_range("A7:N9", "No uncertain selections were found. The workbook is ready to generate the DEFE template.", resolved_fmt)

    data_sheet = workbook.add_worksheet("Selections_Data")
    data_sheet.hide()
    data_headers = [
        "ReviewID", "VIN", "HarnessFamily", "EnginePN", "MatchStatus",
        "RequiredCount", "MatchedCount", "MissingCount", "ExtraCount",
        "RequiredSalesCodes", "MatchedSalesCodes", "Giveaway",
        "MissingSalesCodes", "ExtraSalesCodes", "Score",
    ]
    for col, header in enumerate(data_headers):
        data_sheet.write(0, col, header)
    for row_idx, (_, selection) in enumerate(selections_df.iterrows(), start=1):
        vin = _review_text(selection.get("VIN"))
        family = _review_text(selection.get("HarnessFamily"))
        values = [
            f"{vin}|{family}", vin, family,
            _review_text(selection.get("SelectedHarnessPN")),
            _review_text(selection.get("MatchStatus")),
            selection.get("RequiredCount"), selection.get("MatchedCount"),
            selection.get("MissingCount"), selection.get("ExtraCount"),
            _review_text(selection.get("RequiredSalesCodes")),
            _review_text(selection.get("MatchedSalesCodes")),
            _review_text(selection.get("Giveaway")),
            _review_text(selection.get("MissingSalesCodes")),
            _review_text(selection.get("ExtraSalesCodes")),
            selection.get("Score"),
        ]
        for col, value in enumerate(values):
            if value is None or (isinstance(value, float) and pd.isna(value)):
                value = ""
            data_sheet.write(row_idx, col, value)

    config_sheet = workbook.add_worksheet("Config")
    config_sheet.hide()
    config_sheet.write_row("A1", ["Key", "Value"])
    config_sheet.write_row("A2", ["ReviewFirstRow", excel_first_row])
    config_sheet.write_row("A3", ["ReviewLastRow", excel_last_row if len(review_df) else excel_first_row - 1])
    config_sheet.write_row("A4", ["SelectionLastRow", len(selections_df) + 1])
    config_sheet.write_row("A5", ["OutputFile", defe_output_name])

    _copy_template_sheet_to_xlsxwriter(workbook, template_path)
    workbook.close()
    return output_path
def create_formatted_output(template_path: str, my: str, program: str, out_dir: str,
                           selections_df: pd.DataFrame, vin_matrix_df: pd.DataFrame):
    """
    Copy Template.xlsx with all formatting preserved,
    populate with VINs and harness selections,
    rename to '{MY_last_2_digits}_{Program}_VBOM_Template_for_DEFE.xlsx'
    e.g., 27_RU_VBOM_Template_for_DEFE.xlsx (where MY=2027)
    """
    from openpyxl import load_workbook
    
    try:
        # Extract last 2 digits of MY
        my_short = my[-2:] if len(my) >= 2 else my
        
        # Load template
        wb_new = load_workbook(template_path)
        ws_new = wb_new.active
        
        # Clear data rows (keep header formatting in rows 1-2)
        # Clear from row 3 onwards
        for r in range(3, ws_new.max_row + 1):
            for c in range(1, ws_new.max_column + 1):
                cell = ws_new.cell(row=r, column=c)
                cell.value = None
        
        # Populate VINs in row 2, starting from column C (col 3)
        vin_order = list(vin_matrix_df["VIN"])
        start_col = 3  # Column C
        for idx, vin in enumerate(vin_order):
            ws_new.cell(row=2, column=start_col + idx, value=vin)
        
        # Clear any leftover VINs beyond current list
        for c in range(start_col + len(vin_order), ws_new.max_column + 1):
            ws_new.cell(row=2, column=c, value=None)
        
        # Build mapping: family -> PN -> set(VINs)
        family_map = {}
        for _, row in selections_df.iterrows():
            family = str(row["HarnessFamily"]).strip()
            pn = str(row["SelectedHarnessPN"]).strip()
            vin = row["VIN"]
            if pn.upper() == NOT_APPLICABLE_PN or str(row.get("MatchStatus", "")).strip().upper() == NOT_APPLICABLE_STATUS:
                continue
            family_map.setdefault(family, {}).setdefault(pn, set()).add(vin)
        
        # Build ordered entries
        entries = []
        for family in sorted(family_map.keys()):
            pn_map = family_map[family]
            for pn in sorted(pn_map.keys()):
                entries.append((pn, f"WIRING ASSY -{family}", pn_map[pn]))
        
        # Write rows
        for i, (pn, harness_name, vin_set) in enumerate(entries):
            r = 3 + i
            ws_new.cell(row=r, column=1, value=pn)
            ws_new.cell(row=r, column=2, value=harness_name)
            for j, vin in enumerate(vin_order):
                ws_new.cell(row=r, column=start_col + j, value=("x" if vin in vin_set else ""))
        
        # Clear any remaining old rows beyond the new entries
        for r in range(3 + len(entries), ws_new.max_row + 1):
            ws_new.cell(row=r, column=1, value=None)
            ws_new.cell(row=r, column=2, value=None)
            for j in range(len(vin_order)):
                ws_new.cell(row=r, column=start_col + j, value=None)
        
        # Save with new name
        output_name = f"{my_short}_{program}_VBOM_Template_for_DEFE.xlsx"
        output_path = os.path.join(out_dir, output_name)
        wb_new.save(output_path)
        
        return output_path
    except Exception as e:
        print(f"ERROR: Could not create formatted output file: {e}")
        import traceback
        traceback.print_exc()
        return None
