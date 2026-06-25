from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path
from typing import Iterable

import pandas as pd


def _resolve_vbom_root() -> Path:
    return Path(__file__).resolve().parent.parent / "VBOMxRISKMATRIX 2"


def _load_vbom_module():
    vbom_root = _resolve_vbom_root()
    if str(vbom_root) not in sys.path:
        sys.path.append(str(vbom_root))
    import main_app as vbom_main_app

    return vbom_main_app


def _write_uploaded_file(uploaded_file, destination_path: Path) -> Path:
    destination_path.write_bytes(uploaded_file.getvalue())
    return destination_path


def run_vbom_workflow(
    *,
    my: str,
    program: str,
    source_type: str,
    input_upload,
    complexity_uploads: Iterable,
    output_dir: Path | None = None,
) -> dict:
    vbom_main_app = _load_vbom_module()

    target_dir = output_dir or Path(tempfile.mkdtemp(prefix="vbom_streamlit_", dir=str(Path.cwd())))
    target_dir.mkdir(parents=True, exist_ok=True)

    input_name = getattr(input_upload, "name", "input_file") or "input_file"
    input_path = target_dir / Path(input_name).name
    _write_uploaded_file(input_upload, input_path)

    complexity_paths = []
    for uploaded_file in complexity_uploads:
        file_name = getattr(uploaded_file, "name", "harness_complexity.xlsx") or "harness_complexity.xlsx"
        destination_path = target_dir / Path(file_name).name
        _write_uploaded_file(uploaded_file, destination_path)
        complexity_paths.append(destination_path)

    if not complexity_paths:
        raise ValueError("At least one harness complexity file is required.")

    source_mode = "buildspec" if source_type.lower() == "buildspec" else "doall"
    vin_matrix_df, vin_codes_sorted = vbom_main_app.build_vin_matrix(str(input_path), source_type=source_mode)

    per_file_master = []
    per_file_families = []
    all_complexity_codes = set()

    for complexity_path in complexity_paths:
        df_comp, header_codes, pn_rows = vbom_main_app.read_complexity_sheet(str(complexity_path))
        family = vbom_main_app.try_get_harness_family(str(complexity_path))
        per_file_master.append((str(complexity_path), df_comp))
        per_file_families.append(
            {
                "family": family,
                "header_codes": set(header_codes),
                "pns": pn_rows,
            }
        )
        all_complexity_codes.update(header_codes)

    vin_code_set = set(vin_codes_sorted)
    complexity_code_set = set(all_complexity_codes)
    diff_df = vbom_main_app.build_salescode_diff(vin_code_set, complexity_code_set)
    excluded_codes = sorted(list(vin_code_set - complexity_code_set))
    excluded_df = pd.DataFrame({"SalesCode_Not_In_Any_Harness": excluded_codes})

    family_stats_df, global_code_df = vbom_main_app.build_salescode_statistics(per_file_families)
    selected_codes_by_family = {
        fam["family"]: set(fam["header_codes"]) for fam in per_file_families
    }
    families_for_matching = vbom_main_app.filter_per_file_families(per_file_families, selected_codes_by_family)

    selections_df, all_candidates_df, final_bom_df = vbom_main_app.build_outputs(vin_matrix_df, families_for_matching)

    master_path = target_dir / vbom_main_app.MASTER_FILE_NAME
    vin_matrix_path = target_dir / vbom_main_app.VIN_MATRIX_FILE
    selections_path = target_dir / vbom_main_app.SELECTIONS_FILE

    used_sheetnames = set()
    with pd.ExcelWriter(master_path, engine="openpyxl") as writer:
        for complexity_path, df_comp in per_file_master:
            sheet_name = vbom_main_app.safe_sheetname(Path(complexity_path).stem, used_sheetnames)
            df_comp.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    vin_matrix_df.to_excel(vin_matrix_path, index=False, engine="openpyxl")
    vbom_main_app.write_df_to_excel_append(str(vin_matrix_path), "SalesCode_Diff", diff_df)

    with pd.ExcelWriter(selections_path, engine="openpyxl") as writer:
        selections_df.to_excel(writer, sheet_name="Selections", index=False)
        all_candidates_df.to_excel(writer, sheet_name="AllCandidates", index=False)
        excluded_df.to_excel(writer, sheet_name="Excluded_SalesCodes", index=False)
        final_bom_df.to_excel(writer, sheet_name="Final_BOM_By_VIN", index=False)
        if not family_stats_df.empty:
            family_stats_df.to_excel(writer, sheet_name="Family_Code_Stats", index=False)
            global_code_df.to_excel(writer, sheet_name="Global_Code_Overview", index=False)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(selections_path)
        ws = wb["AllCandidates"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        highlight_fill = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")
        header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        isbest_col = header_map.get("IsBest")
        if isbest_col is not None:
            for row_num in range(2, ws.max_row + 1):
                val = ws.cell(row=row_num, column=isbest_col).value
                is_true = (
                    (isinstance(val, bool) and val)
                    or (isinstance(val, (int, float)) and val == 1)
                    or (isinstance(val, str) and val.strip().upper() == "TRUE")
                )
                if is_true:
                    for col_num in range(1, ws.max_column + 1):
                        ws.cell(row=row_num, column=col_num).fill = highlight_fill
        wb.save(selections_path)
    except Exception:
        pass

    template_path = _resolve_vbom_root() / vbom_main_app.TEMPLATE_SOURCE_FILE
    formatted_template_path = None
    if template_path.exists():
        formatted_template_path = vbom_main_app.create_formatted_output(
            str(template_path),
            my,
            program,
            str(target_dir),
            selections_df,
            vin_matrix_df,
        )

    return {
        "output_dir": target_dir,
        "master_path": master_path,
        "vin_matrix_path": vin_matrix_path,
        "selections_path": selections_path,
        "formatted_template_path": formatted_template_path,
    }
