import os
import re
import pandas as pd
import tkinter as tk
from tkinter import Tk, filedialog, messagebox
from tkinter import ttk

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_TK_DND = True
except Exception:
    DND_FILES = None
    TkinterDnD = None
    HAS_TK_DND = False

# ========================
# Configurable parameters
# ========================
VIN_COL = "VIN"
SALES_CODES_COL = "Sales Code ( 3 Char)"   # matches your file

# Your requested rules for VIN/spec parsing
INCLUDE_NUMERIC_3CHAR = False   # if False, drop '191', '590', etc
ONLY_ALPHA_3CHAR = False         # if True, keep only A–Z triplets

# Optional extras:
SPLIT_SUFFIX_X3 = False         # True => RLX3 -> RLX
SAVE_OUTPUT = True

# Output names
MASTER_FILE_NAME = "Master_Combined_Harness_Complexity.xlsx"
VIN_MATRIX_FILE = "VIN_Salescode_matrix_.xlsx"
SELECTIONS_FILE = "VIN_to_Harness_Selection.xlsx"  # final multi-tab output
TEMPLATE_SOURCE_FILE = "Template.xlsx"

# Complexity header parsing: keep alpha-numeric, but numeric-only codes are dropped.
COMPLEXITY_ALLOW_ALPHANUMERIC = True   # regex [A-Z0-9]{3} then filters out 3-digit codes

# =========
# File pickers
# =========
def pick_file(title="Select VIN/spec file"):
    root = Tk(); root.withdraw(); root.attributes('-topmost', True)
    filetypes = [("Excel files", "*.xlsx *.xls *.xlsm"), ("CSV files", "*.csv"), ("All files", "*.*")]
    path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    root.destroy()
    return path

def pick_multiple_files(title="Select one or more Harness Complexity files"):
    root = Tk(); root.withdraw(); root.attributes('-topmost', True)
    filetypes = [("Excel files", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")]
    paths = filedialog.askopenfilenames(title=title, filetypes=filetypes)
    root.destroy()
    return list(paths)


def _parse_drop_files(raw: str):
    """Parse tkdnd payload into local file paths (supports macOS file:// URIs)."""
    from urllib.parse import unquote, urlparse

    if not raw:
        return []

    tokens = re.findall(r'\{[^}]+\}|[^\s]+', raw)
    out = []
    for token in tokens:
        path = token[1:-1] if token.startswith("{") and token.endswith("}") else token
        path = path.strip().strip('"').strip("'")
        if not path:
            continue

        # Finder drops may come as file:// URLs; decode and normalize them.
        if path.startswith("file://"):
            parsed = urlparse(path)
            path = unquote(parsed.path)
            if os.name == "nt" and path.startswith("/") and len(path) > 2 and path[2] == ":":
                path = path[1:]

        path = os.path.normpath(path)
        out.append(path)

    return out


class RunSetupDialog:
    """Startup window for MY/Program and input file selection."""

    BRAND_NAVY = "#11314f"
    BRAND_ORANGE = "#d9851f"
    BRAND_WHITE = "#ffffff"
    BRAND_MIST = "#eef3f8"
    BRAND_TEXT = "#173552"

    def __init__(self):
        self.result = None
        self.dnd_ready = False
        if HAS_TK_DND and TkinterDnD is not None:
            try:
                self.root = TkinterDnD.Tk()
                self.dnd_ready = True
            except Exception:
                # Continue without drag-and-drop if tkdnd runtime assets are unavailable.
                self.root = Tk()
        else:
            self.root = Tk()
        self.root.title("VBOM Generator")
        self.root.geometry("1100x920")
        self.root.minsize(1000, 840)
        self.root.configure(bg=self.BRAND_NAVY)
        self.root.protocol("WM_DELETE_WINDOW", self._on_cancel)

        self.my_var = tk.StringVar()
        self.program_var = tk.StringVar()
        self.doall_var = tk.StringVar()
        self.buildspec_var = tk.StringVar()
        self.doall_confirmed = False
        self.buildspec_confirmed = False
        self.active_spec_source = None
        self.harness_confirmed = False
        self.harness_files = []
        self.logo_img = None
        self.start_btn_top = None
        self.last_selected_folder = os.getcwd()

        self._build_ui()
        self.root.mainloop()

    def _build_ui(self):
        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure("Card.TLabelframe", background=self.BRAND_WHITE)
        style.configure("Card.TLabelframe.Label", background=self.BRAND_WHITE, foreground=self.BRAND_TEXT, font=("Segoe UI", 13, "bold"))
        style.configure("Card.TFrame", background=self.BRAND_WHITE)
        style.configure("Brand.TButton", font=("Segoe UI", 10, "bold"), padding=8)

        outer = tk.Frame(self.root, bg=self.BRAND_NAVY, padx=16, pady=16)
        outer.pack(fill="both", expand=True)

        hero = tk.Frame(outer, bg=self.BRAND_NAVY)
        hero.pack(fill="x", pady=(0, 14))
        hero_top = tk.Frame(hero, bg=self.BRAND_NAVY)
        hero_top.pack(fill="x")
        self._render_logo(hero_top)
        self.start_btn_top = tk.Button(
            hero_top,
            text="Start Analysis",
            command=self._on_start,
            bg=self.BRAND_ORANGE,
            fg="#ffffff",
            activebackground="#bf7417",
            activeforeground="#ffffff",
            relief="flat",
            padx=14,
            pady=6,
            state="disabled"
        )
        self.start_btn_top.pack(side="right", anchor="ne")
        tk.Label(
            hero,
            text="VBOM TOOL",
            bg=self.BRAND_NAVY,
            fg=self.BRAND_WHITE,
            font=("Segoe UI", 22, "bold")
        ).pack(anchor="w", pady=(8, 2))
        tk.Label(
            hero,
            text="Enter MY/Program, upload files, confirm each section, then start analysis.",
            bg=self.BRAND_NAVY,
            fg="#d9e4ef",
            font=("Segoe UI", 11)
        ).pack(anchor="w")

        body = tk.Frame(outer, bg=self.BRAND_MIST, bd=0, relief="flat")
        body.pack(fill="both", expand=True)

        form = ttk.LabelFrame(body, text="Program Information", padding=12, style="Card.TLabelframe")
        form.pack(fill="x")
        ttk.Label(form, text="Model Year (MY):").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=4)
        ttk.Entry(form, textvariable=self.my_var, width=18).grid(row=0, column=1, sticky="w", pady=4, padx=(0, 14))
        ttk.Label(form, text="Program:").grid(row=0, column=2, sticky="w", padx=(16, 6), pady=4)
        ttk.Entry(form, textvariable=self.program_var, width=20).grid(row=0, column=3, sticky="w", pady=4)
        ttk.Label(form, text="Example: MY=27, Program=RU", foreground="#4a6178").grid(row=1, column=0, columnspan=4, sticky="w", pady=(2, 0))

        spec_sections = tk.Frame(body, bg=self.BRAND_MIST)
        spec_sections.pack(fill="x", pady=(12, 0))

        doall_box = ttk.LabelFrame(spec_sections, text="DoAll (columns: VIN, Salescodes)", padding=12, style="Card.TLabelframe")
        doall_box.pack(side="left", fill="both", expand=True, padx=(0, 6))
        doall_row = ttk.Frame(doall_box, style="Card.TFrame")
        doall_row.pack(fill="x")
        self.doall_entry = ttk.Entry(doall_row, textvariable=self.doall_var)
        self.doall_entry.pack(side="left", fill="x", expand=True)
        self.doall_browse_btn = ttk.Button(doall_row, text="Browse", command=self._browse_doall)
        self.doall_browse_btn.pack(side="left", padx=(8, 0))
        self.doall_confirm_btn = ttk.Button(doall_row, text="Confirm DoAll", command=self._confirm_doall)
        self.doall_confirm_btn.pack(side="left", padx=(8, 0))

        self.doall_drop = tk.Label(
            doall_box,
            text="Drag and drop DoAll file here (.xlsx/.xls/.xlsm/.csv)",
            bg="#ffffff", fg=self.BRAND_TEXT, relief="ridge", bd=1, padx=10, pady=16
        )
        self.doall_drop.pack(fill="x", pady=(8, 0))
        self.doall_status_lbl = tk.Label(doall_box, text="Pending", fg="#a65000", bg=self.BRAND_WHITE, anchor="w")
        self.doall_status_lbl.pack(fill="x", pady=(6, 0))

        buildspec_box = ttk.LabelFrame(spec_sections, text="BuildSpec File", padding=12, style="Card.TLabelframe")
        buildspec_box.pack(side="left", fill="both", expand=True, padx=(6, 0))
        buildspec_row = ttk.Frame(buildspec_box, style="Card.TFrame")
        buildspec_row.pack(fill="x")
        self.buildspec_entry = ttk.Entry(buildspec_row, textvariable=self.buildspec_var)
        self.buildspec_entry.pack(side="left", fill="x", expand=True)
        self.buildspec_browse_btn = ttk.Button(buildspec_row, text="Browse", command=self._browse_buildspec)
        self.buildspec_browse_btn.pack(side="left", padx=(8, 0))
        self.buildspec_confirm_btn = ttk.Button(buildspec_row, text="Confirm BuildSpec", command=self._confirm_buildspec)
        self.buildspec_confirm_btn.pack(side="left", padx=(8, 0))

        self.buildspec_drop = tk.Label(
            buildspec_box,
            text="Drag and drop BuildSpec file here (.xlsx/.xls/.xlsm)",
            bg="#ffffff", fg=self.BRAND_TEXT, relief="ridge", bd=1, padx=10, pady=16
        )
        self.buildspec_drop.pack(fill="x", pady=(8, 0))
        self.buildspec_status_lbl = tk.Label(buildspec_box, text="Pending", fg="#a65000", bg=self.BRAND_WHITE, anchor="w")
        self.buildspec_status_lbl.pack(fill="x", pady=(6, 0))

        harness_box = ttk.LabelFrame(body, text="Harness Complexity Files (multiple)", padding=12, style="Card.TLabelframe")
        harness_box.pack(fill="both", expand=True, pady=(12, 0))

        bar = ttk.Frame(harness_box, style="Card.TFrame")
        bar.pack(fill="x")
        ttk.Button(bar, text="Add Files", command=self._browse_harness).pack(side="left")
        ttk.Button(bar, text="Remove Selected", command=self._remove_selected_harness).pack(side="left", padx=6)
        ttk.Button(bar, text="Clear", command=self._clear_harness).pack(side="left")
        ttk.Button(bar, text="Confirm Harness Files", command=self._confirm_harness).pack(side="left", padx=8)
        self.harness_count_lbl = ttk.Label(bar, text="0 file(s)")
        self.harness_count_lbl.pack(side="right")

        list_frame = ttk.Frame(harness_box, style="Card.TFrame")
        list_frame.pack(fill="both", expand=True, pady=(8, 0))
        self.harness_list = tk.Listbox(
            list_frame,
            selectmode="extended",
            bg="#ffffff",
            fg=self.BRAND_TEXT,
            highlightbackground="#c7d4e2",
            relief="flat"
        )
        ysb = ttk.Scrollbar(list_frame, orient="vertical", command=self.harness_list.yview)
        self.harness_list.configure(yscrollcommand=ysb.set)
        self.harness_list.pack(side="left", fill="both", expand=True)
        ysb.pack(side="right", fill="y")

        self.harness_drop = tk.Label(
            harness_box,
            text="Drag and drop one or many Harness Complexity files here (.xlsx/.xls/.xlsm)",
            bg="#ffffff", fg=self.BRAND_TEXT, relief="ridge", bd=1, padx=10, pady=14
        )
        self.harness_drop.pack(fill="x", pady=(8, 0))
        self.harness_status_lbl = tk.Label(harness_box, text="Pending", fg="#a65000", bg=self.BRAND_WHITE, anchor="w")
        self.harness_status_lbl.pack(fill="x", pady=(6, 0))

        if self.dnd_ready and DND_FILES is not None:
            self._enable_dnd()
        else:
            self.doall_drop.configure(text="Drag-and-drop unavailable (install tkinterdnd2). Use Browse.", fg="#8a4f00")
            self.buildspec_drop.configure(text="Drag-and-drop unavailable (install tkinterdnd2). Use Browse.", fg="#8a4f00")
            self.harness_drop.configure(text="Drag-and-drop unavailable (install tkinterdnd2). Use Add Files.", fg="#8a4f00")

        actions = tk.Frame(outer, bg=self.BRAND_NAVY)
        actions.pack(fill="x", pady=(12, 0))
        self.readiness_lbl = tk.Label(
            actions,
            text="Waiting for confirmations",
            bg=self.BRAND_NAVY,
            fg="#d9e4ef",
            font=("Segoe UI", 10)
        )
        self.readiness_lbl.pack(side="left")
        tk.Button(
            actions,
            text="Cancel",
            command=self._on_cancel,
            bg="#8c3b2a",
            fg="#ffffff",
            activebackground="#762d1e",
            activeforeground="#ffffff",
            relief="flat",
            padx=14,
            pady=8
        ).pack(side="right")
        self.start_btn = tk.Button(
            actions,
            text="Start Analysis",
            command=self._on_start,
            bg=self.BRAND_ORANGE,
            fg="#ffffff",
            activebackground="#bf7417",
            activeforeground="#ffffff",
            relief="flat",
            padx=18,
            pady=8,
            state="disabled"
        )
        self.start_btn.pack(side="right", padx=8)

        self.my_var.trace_add("write", lambda *_: self._refresh_start_state())
        self.program_var.trace_add("write", lambda *_: self._refresh_start_state())
        self.doall_var.trace_add("write", lambda *_: self._on_doall_changed())
        self.buildspec_var.trace_add("write", lambda *_: self._on_buildspec_changed())
        self._refresh_spec_source_state()
        self._refresh_start_state()

    def _render_logo(self, parent):
        logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
        if not os.path.exists(logo_path):
            return
        try:
            # Works on Tk versions that support SVG directly.
            raw = tk.PhotoImage(file=logo_path)
            self.logo_img = raw.subsample(2, 2)
            tk.Label(parent, image=self.logo_img, bg=self.BRAND_NAVY).pack(anchor="w")
            return
        except Exception:
            pass

        try:
            import io
            from cairosvg import svg2png
            from PIL import Image, ImageTk

            png_bytes = svg2png(url=logo_path, output_width=130)
            pil = Image.open(io.BytesIO(png_bytes))
            self.logo_img = ImageTk.PhotoImage(pil)
            tk.Label(parent, image=self.logo_img, bg=self.BRAND_NAVY).pack(anchor="w")
        except Exception:
            tk.Label(parent, text="VERSIGENT", bg=self.BRAND_NAVY, fg=self.BRAND_WHITE, font=("Segoe UI", 18, "bold")).pack(anchor="w")

    def _enable_dnd(self):
        self.doall_drop.drop_target_register(DND_FILES)
        self.doall_drop.dnd_bind('<<Drop>>', self._on_drop_doall)
        self.doall_drop.dnd_bind('<<DragEnter>>', lambda _e: self.doall_drop.configure(bg="#fff3e1"))
        self.doall_drop.dnd_bind('<<DragLeave>>', lambda _e: self.doall_drop.configure(bg="#ffffff"))
        self.buildspec_drop.drop_target_register(DND_FILES)
        self.buildspec_drop.dnd_bind('<<Drop>>', self._on_drop_buildspec)
        self.buildspec_drop.dnd_bind('<<DragEnter>>', lambda _e: self.buildspec_drop.configure(bg="#fff3e1"))
        self.buildspec_drop.dnd_bind('<<DragLeave>>', lambda _e: self.buildspec_drop.configure(bg="#ffffff"))
        self.harness_drop.drop_target_register(DND_FILES)
        self.harness_drop.dnd_bind('<<Drop>>', self._on_drop_harness)
        self.harness_drop.dnd_bind('<<DragEnter>>', lambda _e: self.harness_drop.configure(bg="#fff3e1"))
        self.harness_drop.dnd_bind('<<DragLeave>>', lambda _e: self.harness_drop.configure(bg="#ffffff"))

    def _browse_doall(self):
        path = filedialog.askopenfilename(
            title="Select DoAll file",
            filetypes=[("DoAll", "*.xlsx *.xls *.xlsm *.csv"), ("All files", "*.*")],
        )
        if path:
            self._select_input_source("doall", path)
            self.last_selected_folder = os.path.dirname(path)
            self.doall_drop.configure(bg="#ffffff")

    def _browse_buildspec(self):
        path = filedialog.askopenfilename(
            title="Select BuildSpec file",
            filetypes=[("BuildSpec", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")],
        )
        if path:
            self._select_input_source("buildspec", path)
            self.last_selected_folder = os.path.dirname(path)
            self.buildspec_drop.configure(bg="#ffffff")

    def _browse_harness(self):
        paths = filedialog.askopenfilenames(
            title="Select one or more Harness Complexity files",
            filetypes=[("Excel files", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")],
        )
        if paths:
            self.last_selected_folder = os.path.dirname(paths[-1])
        self._add_harness_files(list(paths))

    def _on_drop_doall(self, event):
        files = _parse_drop_files(event.data)
        if not files:
            return
        if self.active_spec_source == "buildspec":
            self.doall_status_lbl.configure(text="Unavailable while BuildSpec is selected", fg="#8a4f00")
            return
        allowed = {".xlsx", ".xls", ".xlsm", ".csv"}
        chosen = None
        for p in files:
            if os.path.splitext(p)[1].lower() in allowed:
                chosen = p
                break
        if not chosen:
            self.doall_status_lbl.configure(text="No valid DoAll file dropped", fg="#b52a2a")
            return
        self._select_input_source("doall", chosen)
        self.last_selected_folder = os.path.dirname(chosen)
        self.doall_drop.configure(bg="#ffffff")

    def _on_drop_buildspec(self, event):
        files = _parse_drop_files(event.data)
        if not files:
            return
        if self.active_spec_source == "doall":
            self.buildspec_status_lbl.configure(text="Unavailable while DoAll is selected", fg="#8a4f00")
            return
        allowed = {".xlsx", ".xls", ".xlsm"}
        chosen = None
        for p in files:
            if os.path.splitext(p)[1].lower() in allowed:
                chosen = p
                break
        if not chosen:
            self.buildspec_status_lbl.configure(text="No valid BuildSpec file dropped", fg="#b52a2a")
            return
        self._select_input_source("buildspec", chosen)
        self.last_selected_folder = os.path.dirname(chosen)
        self.buildspec_drop.configure(bg="#ffffff")

    def _on_drop_harness(self, event):
        files = _parse_drop_files(event.data)
        self.harness_drop.configure(bg="#ffffff")
        self._add_harness_files(files)

    def _add_harness_files(self, files):
        allowed = {".xlsx", ".xls", ".xlsm"}
        added = False
        invalid = 0
        last_added_folder = None
        for p in files:
            ext = os.path.splitext(p)[1].lower()
            if ext not in allowed:
                invalid += 1
                continue
            norm = os.path.normpath(p)
            if norm not in self.harness_files:
                self.harness_files.append(norm)
                last_added_folder = os.path.dirname(norm)
                added = True
        if added:
            if last_added_folder:
                self.last_selected_folder = last_added_folder
            self.harness_confirmed = False
            self._refresh_harness_list()
        if invalid:
            self.harness_status_lbl.configure(text=f"Ignored {invalid} non-Excel file(s)", fg="#b56f00")

    def _refresh_harness_list(self):
        self.harness_list.delete(0, tk.END)
        for p in self.harness_files:
            self.harness_list.insert(tk.END, p)
        self.harness_count_lbl.configure(text=f"{len(self.harness_files)} file(s)")
        if self.harness_files:
            self.harness_status_lbl.configure(text=f"{len(self.harness_files)} file(s) loaded. Click Confirm Harness Files.", fg="#4a6178")
        else:
            self.harness_status_lbl.configure(text="Pending", fg="#a65000")
        self._refresh_start_state()

    def _remove_selected_harness(self):
        sel = list(self.harness_list.curselection())
        if not sel:
            return
        for idx in reversed(sel):
            if 0 <= idx < len(self.harness_files):
                self.harness_files.pop(idx)
        self.harness_confirmed = False
        self._refresh_harness_list()

    def _clear_harness(self):
        self.harness_files = []
        self.harness_confirmed = False
        self._refresh_harness_list()

    def _on_cancel(self):
        self.result = None
        self.root.destroy()

    def _select_input_source(self, source: str, path: str):
        if source == "doall":
            self.doall_var.set(path)
            self.doall_confirmed = False
            self.doall_status_lbl.configure(text="File selected. Click Confirm DoAll.", fg="#4a6178")
            self.buildspec_var.set("")
            self.buildspec_confirmed = False
            self.buildspec_status_lbl.configure(text="Unavailable while DoAll is selected", fg="#8a4f00")
            self.active_spec_source = "doall"
        else:
            self.buildspec_var.set(path)
            self.buildspec_confirmed = False
            self.buildspec_status_lbl.configure(text="File selected. Click Confirm BuildSpec.", fg="#4a6178")
            self.doall_var.set("")
            self.doall_confirmed = False
            self.doall_status_lbl.configure(text="Unavailable while BuildSpec is selected", fg="#8a4f00")
            self.active_spec_source = "buildspec"
        self._refresh_spec_source_state()
        self._refresh_start_state()

    def _refresh_spec_source_state(self):
        doall_available = self.active_spec_source in (None, "doall")
        buildspec_available = self.active_spec_source in (None, "buildspec")

        self.doall_entry.configure(state=("normal" if doall_available else "disabled"))
        self.doall_browse_btn.configure(state=("normal" if doall_available else "disabled"))
        self.doall_confirm_btn.configure(state=("normal" if doall_available else "disabled"))

        self.buildspec_entry.configure(state=("normal" if buildspec_available else "disabled"))
        self.buildspec_browse_btn.configure(state=("normal" if buildspec_available else "disabled"))
        self.buildspec_confirm_btn.configure(state=("normal" if buildspec_available else "disabled"))

        if doall_available and not self.doall_var.get().strip() and self.active_spec_source is None:
            self.doall_status_lbl.configure(text="Pending", fg="#a65000")
        if buildspec_available and not self.buildspec_var.get().strip() and self.active_spec_source is None:
            self.buildspec_status_lbl.configure(text="Pending", fg="#a65000")

    def _on_doall_changed(self):
        if self.doall_var.get().strip():
            if self.active_spec_source != "doall":
                self.active_spec_source = "doall"
                self.buildspec_var.set("")
                self.buildspec_confirmed = False
                self.buildspec_status_lbl.configure(text="Unavailable while DoAll is selected", fg="#8a4f00")
            self.doall_confirmed = False
            self.doall_status_lbl.configure(text="File selected. Click Confirm DoAll.", fg="#4a6178")
        elif self.active_spec_source == "doall":
            self.active_spec_source = None
            self.doall_confirmed = False
            self.doall_status_lbl.configure(text="Pending", fg="#a65000")
        self._refresh_spec_source_state()
        self._refresh_start_state()

    def _on_buildspec_changed(self):
        if self.buildspec_var.get().strip():
            if self.active_spec_source != "buildspec":
                self.active_spec_source = "buildspec"
                self.doall_var.set("")
                self.doall_confirmed = False
                self.doall_status_lbl.configure(text="Unavailable while BuildSpec is selected", fg="#8a4f00")
            self.buildspec_confirmed = False
            self.buildspec_status_lbl.configure(text="File selected. Click Confirm BuildSpec.", fg="#4a6178")
        elif self.active_spec_source == "buildspec":
            self.active_spec_source = None
            self.buildspec_confirmed = False
            self.buildspec_status_lbl.configure(text="Pending", fg="#a65000")
        self._refresh_spec_source_state()
        self._refresh_start_state()

    def _confirm_doall(self):
        path = self.doall_var.get().strip()
        if not path:
            messagebox.showerror("Missing file", "Please select a DoAll file first.", parent=self.root)
            return
        if not os.path.isfile(path):
            messagebox.showerror("Invalid file", "DoAll file does not exist.", parent=self.root)
            return
        self.doall_confirmed = True
        self.active_spec_source = "doall"
        self.doall_status_lbl.configure(text="DoAll confirmed", fg="#1f7a1f")
        self.buildspec_status_lbl.configure(text="Unavailable while DoAll is selected", fg="#8a4f00")
        self._refresh_spec_source_state()
        self._refresh_start_state()

    def _confirm_buildspec(self):
        path = self.buildspec_var.get().strip()
        if not path:
            messagebox.showerror("Missing file", "Please select a BuildSpec file first.", parent=self.root)
            return
        if not os.path.isfile(path):
            messagebox.showerror("Invalid file", "BuildSpec file does not exist.", parent=self.root)
            return
        self.buildspec_confirmed = True
        self.active_spec_source = "buildspec"
        self.buildspec_status_lbl.configure(text="BuildSpec confirmed", fg="#1f7a1f")
        self.doall_status_lbl.configure(text="Unavailable while BuildSpec is selected", fg="#8a4f00")
        self._refresh_spec_source_state()
        self._refresh_start_state()

    def _confirm_harness(self):
        if not self.harness_files:
            messagebox.showerror("Missing files", "Please add at least one Harness file first.", parent=self.root)
            return
        missing = [p for p in self.harness_files if not os.path.isfile(p)]
        if missing:
            messagebox.showerror("Invalid files", "Some selected Harness files no longer exist.", parent=self.root)
            return
        self.harness_confirmed = True
        self.harness_status_lbl.configure(text=f"Harness files confirmed ({len(self.harness_files)})", fg="#1f7a1f")
        self._refresh_start_state()

    def _refresh_start_state(self):
        missing = []
        if not self.my_var.get().strip():
            missing.append("Model Year")
        if not self.program_var.get().strip():
            missing.append("Program")
        if self.active_spec_source == "doall":
            if not self.doall_confirmed:
                missing.append("DoAll confirmation")
        elif self.active_spec_source == "buildspec":
            if not self.buildspec_confirmed:
                missing.append("BuildSpec confirmation")
        else:
            missing.append("DoAll or BuildSpec confirmation")
        if not self.harness_confirmed:
            missing.append("Harness Files confirmation")

        ready = len(missing) == 0
        new_state = ("normal" if ready else "disabled")
        self.start_btn.configure(state=new_state)
        if self.start_btn_top is not None:
            self.start_btn_top.configure(state=new_state)
        if ready:
            self.readiness_lbl.configure(text="Ready to run")
        else:
            self.readiness_lbl.configure(text=f"Missing: {', '.join(missing)}")

    def _on_start(self):
        my = self.my_var.get().strip()
        program = self.program_var.get().strip()
        doall_path = self.doall_var.get().strip()
        buildspec_path = self.buildspec_var.get().strip()

        if not my or not program:
            messagebox.showerror("Missing data", "Please enter both Model Year and Program.", parent=self.root)
            return

        input_type = self.active_spec_source
        if input_type == "doall":
            input_path = doall_path
            if not input_path:
                messagebox.showerror("Missing file", "Please select a DoAll file.", parent=self.root)
                return
            if not os.path.isfile(input_path):
                messagebox.showerror("Invalid file", "DoAll file does not exist.", parent=self.root)
                return
            if not self.doall_confirmed:
                messagebox.showerror("Not confirmed", "Please confirm DoAll before starting.", parent=self.root)
                return
        elif input_type == "buildspec":
            input_path = buildspec_path
            if not input_path:
                messagebox.showerror("Missing file", "Please select a BuildSpec file.", parent=self.root)
                return
            if not os.path.isfile(input_path):
                messagebox.showerror("Invalid file", "BuildSpec file does not exist.", parent=self.root)
                return
            if not self.buildspec_confirmed:
                messagebox.showerror("Not confirmed", "Please confirm BuildSpec before starting.", parent=self.root)
                return
        else:
            messagebox.showerror("Missing file", "Please select and confirm either DoAll or BuildSpec.", parent=self.root)
            return

        if not self.harness_files:
            messagebox.showerror("Missing files", "Please add at least one Harness Complexity file.", parent=self.root)
            return
        if not self.harness_confirmed:
            messagebox.showerror("Not confirmed", "Please confirm Harness Files before starting.", parent=self.root)
            return

        self.result = {
            "my": my,
            "program": program,
            "input_type": input_type,
            "input_file": input_path,
            "harness_files": list(self.harness_files),
            "output_dir": self.last_selected_folder or os.path.dirname(input_path),
        }
        self.root.destroy()


def collect_run_setup_inputs():
    """Open startup setup window and return run inputs or None if cancelled."""
    dlg = RunSetupDialog()
    return dlg.result

# =========
# IO helpers
# =========
def load_dataframe(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return pd.read_csv(path)
    elif ext in (".xlsx", ".xls", ".xlsm"):
        engine = "openpyxl" if ext in (".xlsx", ".xlsm") else "xlrd"
        return pd.read_excel(path, engine=engine)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

def write_df_to_excel_append(path: str, sheet_name: str, df: pd.DataFrame):
    """Append/overwrite a sheet in an xlsx file (engine=openpyxl)."""
    from openpyxl import load_workbook
    if not os.path.exists(path):
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
        wb.save(path)
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def safe_sheetname(base: str, used: set):
    """Return a safe/unique Excel sheet name (<=31 chars)."""
    name = os.path.splitext(os.path.basename(base))[0]
    name = name[:31] if len(name) > 31 else name
    candidate = name; i = 1
    while candidate in used:
        suffix = f"_{i}"
        candidate = (name[: (31 - len(suffix))] + suffix) if len(name) + len(suffix) > 31 else name + suffix
        i += 1
    used.add(candidate)
    return candidate

# ====================
# Sales code parsing
# ====================
def _is_nullish(x):
    if x is None:
        return True
    try:
        return pd.isna(x)
    except Exception:
        return False

def parse_sales_codes(cell):
    """
    Normalize one cell from 'Sales Code ( 3 Char)' into a list of 3-char codes.
    Respects INCLUDE_NUMERIC_3CHAR / ONLY_ALPHA_3CHAR.
    """
    if isinstance(cell, (list, tuple, set)):
        items = []
        for it in cell:
            items.extend(parse_sales_codes(str(it)))
        # de-dup in-order
        seen, dedup = set(), []
        for c in items:
            if c not in seen:
                seen.add(c); dedup.append(c)
        return dedup

    if _is_nullish(cell):
        return []

    text = str(cell)
    tokens = []
    for raw in text.split():
        t = re.sub(r'^[^A-Za-z0-9]+|[^A-Za-z0-9]+$', '', raw)
        if not t:
            continue
        if SPLIT_SUFFIX_X3 and len(t) == 4 and t.upper().endswith("X3"):
            t = t[:3]
        t_up = t.upper()
        if len(t_up) != 3:
            continue
        if ONLY_ALPHA_3CHAR and not re.fullmatch(r'[A-Z]{3}', t_up):
            continue
        if not INCLUDE_NUMERIC_3CHAR and t_up.isdigit():
            continue
        if not ONLY_ALPHA_3CHAR and not re.fullmatch(r'[A-Z0-9]{3}', t_up):
            continue
        tokens.append(t_up)

    seen, dedup = set(), []
    for c in tokens:
        if c not in seen:
            seen.add(c); dedup.append(c)
    return dedup

def derive_sorted_unique_codes(series: pd.Series):
    code_set = set()
    for cell in series:
        items = cell if isinstance(cell, (list, tuple, set)) else parse_sales_codes(cell)
        for c in items:
            code_set.add(c)
    return sorted(code_set)

# =========================================
# Read Complexity and extract header/PN map
# =========================================
def read_complexity_sheet(path: str):
    """
    Returns:
      df_complexity: raw Complexity sheet (header=None)
      header_codes: sorted unique 3-char codes found in the FIRST ROW
            rows: list of (pn, set_of_codes, set_of_giveaway_codes)
                        where:
                            set_of_codes          => cells marked 'X'
                            set_of_giveaway_codes => cells marked 'G'
    """
    ext = os.path.splitext(path)[1].lower()
    engine = "openpyxl" if ext in (".xlsx", ".xlsm") else "xlrd"
    df = pd.read_excel(path, sheet_name="Complexity", header=None, engine=engine)

    # find header codes from row 0
    header_values = df.iloc[0].tolist()
    code_regex = r'[A-Z0-9]{3}' if COMPLEXITY_ALLOW_ALPHANUMERIC else r'[A-Z]{3}'
    colidx_to_code = {}
    for j, v in enumerate(header_values):
        if _is_nullish(v):
            continue
        txt = str(v).upper()
        m = re.search(code_regex, txt)
        if m:
            code = m.group(0)
            # Drop numeric-only 3-char codes like 108, 110.
            if code.isdigit():
                continue
            colidx_to_code[j] = code
    header_codes = sorted(set(colidx_to_code.values()))

    # pick PNs and their 'X' / 'G' codes
    pn_pattern = re.compile(r'^\d{8}[A-Z]{2}$')  # e.g., 68720520AA
    rows = []
    for i in range(1, len(df)):
        a0 = df.iat[i, 0]
        if _is_nullish(a0):
            continue
        pn = str(a0).strip()
        if not pn_pattern.match(pn):
            continue
        pn_codes = set()
        pn_giveaway_codes = set()
        row_vals = df.iloc[i].tolist()
        for j, val in enumerate(row_vals):
            if j in colidx_to_code:
                val_up = (val if isinstance(val, str) else ("" if _is_nullish(val) else str(val))).strip().upper()
                if val_up == "X":
                    pn_codes.add(colidx_to_code[j])
                elif val_up == "G":
                    pn_giveaway_codes.add(colidx_to_code[j])
        rows.append((pn, pn_codes, pn_giveaway_codes))

    return df, header_codes, rows

def try_get_harness_family(path: str):
    """
    Try to get HarnessFamily from 'Harness PN' tab near the 'Harness:' label,
    fallback to the filename segment.
    """
    base = os.path.basename(path)
    m = re.match(r'^Harness_Complexity_[^_]+_[^_]+_([^_]+)_', base)
    fallback = m.group(1) if m else os.path.splitext(base)[0]
    try:
        ext = os.path.splitext(path)[1].lower()
        engine = "openpyxl" if ext in (".xlsx", ".xlsm") else "xlrd"
        df_hp = pd.read_excel(path, sheet_name="Harness PN", header=None, engine=engine)
        for i in range(min(200, df_hp.shape[0])):
            for j in range(min(20, df_hp.shape[1])):
                val = df_hp.iat[i, j]
                if isinstance(val, str) and val.strip().lower() == "harness:":
                    if j + 1 < df_hp.shape[1]:
                        v = df_hp.iat[i, j+1]
                        if isinstance(v, str) and v.strip():
                            return v.strip().upper()
                    if i + 1 < df_hp.shape[0]:
                        v = df_hp.iat[i+1, j]
                        if isinstance(v, str) and v.strip():
                            return v.strip().upper()
        return fallback.upper()
    except Exception:
        return fallback.upper()

# ==============================
# VIN matrix & SalesCode Diff
# ==============================
def _build_vin_matrix_from_doall(vin_file_path: str):
    df = load_dataframe(vin_file_path)
    missing = [c for c in (VIN_COL, SALES_CODES_COL) if c not in df.columns]
    if missing:
        raise KeyError(f"Missing required column(s): {missing}. Columns found: {list(df.columns)}")
    df = df.copy()
    df["_parsed_codes"] = df[SALES_CODES_COL].apply(parse_sales_codes)
    sorted_unique_codes = derive_sorted_unique_codes(df["_parsed_codes"])

    rows = []
    for _, row in df.iterrows():
        row_data = {VIN_COL: row[VIN_COL]}
        present = set(row["_parsed_codes"])
        for code in sorted_unique_codes:
            row_data[code] = '|' if code in present else ''
        rows.append(row_data)
    matrix_df = pd.DataFrame(rows)
    return matrix_df, sorted_unique_codes


def _is_buildspec_marked(val) -> bool:
    if _is_nullish(val):
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val != 0
    s = str(val).strip().upper()
    if s in ("", "0", "N", "NO", "FALSE", "-"):
        return False
    return True


def _build_vin_matrix_from_buildspec(path: str):
    ext = os.path.splitext(path)[1].lower()
    engine = "openpyxl" if ext in (".xlsx", ".xlsm") else "xlrd"
    df = pd.read_excel(path, header=None, engine=engine)

    # User-defined BuildSpec layout:
    # - Row 7 contains VIN values
    # - Row 6 contains MVON values
    # - Column A contains SalesCodes
    # Prefer VIN when present, else use MVON.
    vin_row_idx = 6
    mvon_row_idx = 5
    code_col_idx = 0
    data_start_row = 7

    if df.shape[0] <= vin_row_idx:
        raise ValueError("BuildSpec does not contain required VIN/MVON rows (rows 6 and 7).")
    if df.shape[1] <= 1:
        raise ValueError("BuildSpec does not contain VIN/MVON data columns.")

    vin_values = df.iloc[vin_row_idx].tolist() if df.shape[0] > vin_row_idx else []
    mvon_values = df.iloc[mvon_row_idx].tolist() if df.shape[0] > mvon_row_idx else []

    vin_to_codes = {}

    for col in range(1, df.shape[1]):
        vin_raw = vin_values[col] if col < len(vin_values) else None
        mvon_raw = mvon_values[col] if col < len(mvon_values) else None

        vin_text = "" if _is_nullish(vin_raw) else str(vin_raw).strip()
        mvon_text = "" if _is_nullish(mvon_raw) else str(mvon_raw).strip()
        vin_id = vin_text if vin_text else mvon_text
        if not vin_id:
            continue
        vin_id_upper = vin_id.strip().upper()
        if vin_id_upper in {"VIN", "MVON", "SALES CODE", "SALESCODE", "SALESCODES"}:
            continue

        code_set = vin_to_codes.setdefault(vin_id, set())

        for r in range(data_start_row, df.shape[0]):
            code_cell = df.iat[r, code_col_idx] if code_col_idx < df.shape[1] else None
            parsed_codes = parse_sales_codes(code_cell)
            if not parsed_codes:
                continue
            marker = df.iat[r, col]
            if _is_buildspec_marked(marker):
                for c in parsed_codes:
                    code_set.add(c)

    if not vin_to_codes:
        raise ValueError("No VIN/MVON and SalesCode assignments were found in BuildSpec.")

    all_codes = sorted({code for codes in vin_to_codes.values() for code in codes})
    rows = []
    for vin_id in sorted(vin_to_codes.keys()):
        present = vin_to_codes[vin_id]
        row_data = {VIN_COL: vin_id}
        for code in all_codes:
            row_data[code] = "|" if code in present else ""
        rows.append(row_data)

    return pd.DataFrame(rows), all_codes


def build_vin_matrix(vin_file_path: str, source_type: str = "doall"):
    if source_type == "buildspec":
        return _build_vin_matrix_from_buildspec(vin_file_path)
    return _build_vin_matrix_from_doall(vin_file_path)

def build_salescode_diff(vin_codes_set, complexity_codes_set):
    vin_not_in_complex = sorted(vin_codes_set - complexity_codes_set)
    complex_not_in_vin = sorted(complexity_codes_set - vin_codes_set)
    max_len = max(len(vin_not_in_complex), len(complex_not_in_vin))
    vin_col = vin_not_in_complex + [""] * (max_len - len(vin_not_in_complex))
    comp_col = complex_not_in_vin + [""] * (max_len - len(complex_not_in_vin))
    return pd.DataFrame({
        "Salescode_not_in_Complexity": vin_col,
        "Salescode_not_in_SPEC": comp_col
    })
# ==============================
# Sales Code Statistics
# ==============================
def build_salescode_statistics(per_file_complexity: list):
    """
    For each harness family compute, per sales code:
      - how many PNs contain it
      - coverage % (PNsWithCode / TotalPNs * 100)
    - status: STANDARD (100%) | OPTIONAL (>0 and <100%) | UNUSED (0%)

        Also produces a global overview across all families, including a single
        classification bucket per SalesCode:
            - STANDARD_IN_ALL_HARNESSES
            - STANDARD_IN_SOME_HARNESSES
            - OPTIONAL_IN_ALL_HARNESSES
            - OTHER

    Returns:
        family_stats_df  – one row per (HarnessFamily, SalesCode)
        global_df        – one row per SalesCode summarising cross-family presence
                   and final classification bucket
        (no pivot returned)
    """
    family_stats_rows = []
    for fam in per_file_complexity:
        family_name = fam["family"]
        pns = fam["pns"]          # list of (pn, set_of_codes, set_of_giveaway_codes)
        total_pns = len(pns)
        if total_pns == 0:
            continue

        # count how many PNs include each code.
        # G (giveaway) means the code is present on that PN too.
        code_counts = {}
        for _pn, codes, _giveaway_codes in pns:
            effective_codes = set(codes) | set(_giveaway_codes)
            for code in effective_codes:
                code_counts[code] = code_counts.get(code, 0) + 1

        # also register header codes that appear in 0 PNs
        for code in fam["header_codes"]:
            if code not in code_counts:
                code_counts[code] = 0

        for code in sorted(code_counts):
            count = code_counts[code]
            pct = round(count / total_pns * 100, 1)
            if pct == 100.0:
                status = "STANDARD"
            elif pct == 0.0:
                status = "UNUSED"
            else:
                status = "OPTIONAL"
            family_stats_rows.append({
                "HarnessFamily": family_name,
                "SalesCode": code,
                "PNsWithCode": count,
                "TotalPNs": total_pns,
                "Coverage_Pct": pct,
                "Status": status,
            })

    family_stats_df = pd.DataFrame(family_stats_rows, columns=[
        "HarnessFamily", "SalesCode", "PNsWithCode", "TotalPNs", "Coverage_Pct", "Status"
    ])

    if family_stats_df.empty:
        return family_stats_df, pd.DataFrame()

    all_families = [fam["family"] for fam in per_file_complexity]
    total_families = len(all_families)
    all_codes = sorted(family_stats_df["SalesCode"].unique())

    global_rows = []
    for code in all_codes:
        sub = family_stats_df[family_stats_df["SalesCode"] == code].copy()
        status_map = dict(zip(sub["HarnessFamily"], sub["Status"]))
        statuses = [status_map.get(fam, "UNUSED") for fam in all_families]

        standard_cnt = sum(1 for s in statuses if s == "STANDARD")
        optional_cnt = sum(1 for s in statuses if s == "OPTIONAL")
        unused_cnt = sum(1 for s in statuses if s == "UNUSED")
        fams_present = [fam for fam in all_families if status_map.get(fam, "UNUSED") != "UNUSED"]

        if standard_cnt == total_families:
            classification = "STANDARD_IN_ALL_HARNESSES"
        elif 0 < standard_cnt < total_families:
            classification = "STANDARD_IN_SOME_HARNESSES"
        elif optional_cnt == total_families:
            classification = "OPTIONAL_IN_ALL_HARNESSES"
        else:
            classification = "OTHER"

        global_rows.append({
            "SalesCode": code,
            "FamiliesPresent": len(fams_present),
            "TotalFamilies": total_families,
            "FamilyCoverage_Pct": round(len(fams_present) / total_families * 100, 1),
            "StandardInFamilies": standard_cnt,
            "OptionalInFamilies": optional_cnt,
            "UnusedInFamilies": unused_cnt,
            "Classification": classification,
            "FamilyList": ", ".join(sorted(fams_present)),
        })

    global_df = pd.DataFrame(global_rows, columns=[
        "SalesCode", "FamiliesPresent", "TotalFamilies",
        "FamilyCoverage_Pct", "StandardInFamilies", "OptionalInFamilies",
        "UnusedInFamilies", "Classification", "FamilyList"
    ])

    return family_stats_df, global_df


def _format_stats_sheets(wb_path: str):
    """Apply conditional color formatting to the three SalesCode statistics sheets."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    STATUS_FILLS = {
        "STANDARD": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "OPTIONAL": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "UNUSED":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    wb = load_workbook(wb_path)

    # Family_Code_Stats: color every row by Status
    if "Family_Code_Stats" in wb.sheetnames:
        ws = wb["Family_Code_Stats"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        hdr = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        status_col = hdr.get("Status")
        if status_col:
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=status_col).value
                fill = STATUS_FILLS.get(str(val).upper() if val else "", None)
                if fill:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill

    # Global_Code_Overview: freeze + autofilter only
    if "Global_Code_Overview" in wb.sheetnames:
        ws = wb["Global_Code_Overview"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Highlight the requested single-view categories.
        from openpyxl.styles import PatternFill
        cls_fills = {
            "STANDARD_IN_ALL_HARNESSES": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "STANDARD_IN_SOME_HARNESSES": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "OPTIONAL_IN_ALL_HARNESSES": PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid"),
        }
        hdr = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        cls_col = hdr.get("Classification")
        if cls_col:
            for r in range(2, ws.max_row + 1):
                cls = ws.cell(row=r, column=cls_col).value
                fill = cls_fills.get(str(cls).upper() if cls else "")
                if fill:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill

    wb.save(wb_path)
# ====================================
# Build Selections + AllCandidates
# ====================================
def compute_status_score(req_cnt, matched, extra, missing):
    if missing == 0 and extra == 0:
        status = "EXACT"
    elif missing == 0 and extra > 0:
        status = "OVERBUILT"
    else:
        status = "INCOMPLETE"
    score = matched - extra - 100 * missing
    return status, score

def _variant_base_family(name: str):
    """Normalize family name to group Pacifica/Voyager variants."""
    if not isinstance(name, str):
        return str(name)
    s = name.upper()
    # Remove PACIFICA and VOYAGER variants (handles both word boundaries and underscores)
    s = re.sub(r'[_\s]*(PACIFICA|VOYAGER)[_\s]*', ' ', s)
    # Normalize whitespace
    s = re.sub(r'[\s_]+', '_', s).strip().rstrip('_')
    return s

def build_outputs(vin_matrix_df: pd.DataFrame,
                  per_file_complexity: list):
    """
    per_file_complexity = list of dicts: {
       'family': str,
       'header_codes': set(...),
         'pns': list of (pn, set_of_codes, set_of_giveaway_codes)
    }
    Returns: selections_df, all_candidates_df, final_bom_df
    """
    families = per_file_complexity
    selections_rows = []
    all_candidates_rows = []

    code_columns = [c for c in vin_matrix_df.columns if c != VIN_COL]

    for _, vrow in vin_matrix_df.iterrows():
        vin = vrow[VIN_COL]
        vin_codes = {c for c in code_columns if str(vrow[c]) == '|'}

        # group families by Pacifica/Voyager base to prevent double selection
        grouped = {}
        for fam in families:
            base = _variant_base_family(fam["family"])
            grouped.setdefault(base, []).append(fam)

        for _base, fam_list in grouped.items():
            group_best = None  # (key, family_name, best_pn, best_info)

            for fam in fam_list:
                family_name = fam["family"]
                fam_codes = fam["header_codes"]
                vin_required_for_family = vin_codes & fam_codes

                # evaluate all PNs for this family
                best_key = None
                best_pn = None
                best_info = None  # (status, req_cnt, matched_cnt, missing_cnt, extra_cnt, missing_list, extra_list, score)

                pn_evals = []
                for pn, pn_codes, pn_giveaway_codes in fam["pns"]:
                    # Treat both X and G as present for matching semantics.
                    effective_pn_codes = set(pn_codes) | set(pn_giveaway_codes)

                    matched_codes = vin_required_for_family & effective_pn_codes
                    missing_codes = vin_required_for_family - effective_pn_codes
                    extra_codes = effective_pn_codes - vin_required_for_family

                    matched = len(matched_codes)
                    missing = len(missing_codes)
                    extra = len(extra_codes)
                    status, score = compute_status_score(len(vin_required_for_family), matched, extra, missing)

                    # sorting key for "best"
                    key = (score, matched, -missing, -extra, -len(extra_codes), pn)
                    pn_evals.append((pn, score, status, matched, missing, extra,
                                     ",".join(sorted(missing_codes)),
                                     ",".join(sorted(extra_codes)),
                                     len(vin_required_for_family),
                                     ",".join(sorted(pn_giveaway_codes))))

                    if (best_key is None) or (key > best_key):
                        best_key = key
                        best_pn = pn
                        best_info = (status, len(vin_required_for_family), matched, missing, extra,
                                     ",".join(sorted(missing_codes)),
                                     ",".join(sorted(extra_codes)),
                                     score,
                                     ",".join(sorted(pn_giveaway_codes)))

                # add candidates (and mark IsBest later if they win the group)
                for pn, score, _status, _matched, _missing, _extra, miss_list, extra_list, _req, giveaway_list in pn_evals:
                    all_candidates_rows.append({
                        "VIN": vin,
                        "HarnessFamily": family_name,
                        "PN": pn,
                        "Score": score,
                        "IsBest": False,
                        "Giveaway": giveaway_list if giveaway_list else None,
                        "MissingSalesCodes": miss_list if miss_list else None,
                        "ExtraSalesCodes": extra_list if extra_list else None
                    })

                if best_info is not None:
                    if (group_best is None) or (best_key > group_best[0]):
                        group_best = (best_key, family_name, best_pn, best_info)

            if group_best is None:
                continue

            # mark best in AllCandidates for the selected family/pn
            _, family_name, best_pn, best_info = group_best
            for row in all_candidates_rows:
                if row["VIN"] == vin and row["HarnessFamily"] == family_name and row["PN"] == best_pn:
                    row["IsBest"] = True

            # add selection (only one per group)
            status, req_cnt, matched_cnt, missing_cnt, extra_cnt, miss_list, extra_list, score, giveaway_list = best_info
            selections_rows.append({
                "VIN": vin,
                "HarnessFamily": family_name,
                "SelectedHarnessPN": best_pn,
                "MatchStatus": status,
                "RequiredCount": req_cnt,
                "MatchedCount": matched_cnt,
                "MissingCount": missing_cnt,
                "ExtraCount": extra_cnt,
                "Giveaway": giveaway_list if giveaway_list else None,
                "MissingSalesCodes": miss_list if miss_list else None,
                "ExtraSalesCodes": extra_list if extra_list else None,
                "Score": score
            })

    selections_df = pd.DataFrame(selections_rows, columns=[
        "VIN","HarnessFamily","SelectedHarnessPN","MatchStatus",
        "RequiredCount","MatchedCount","MissingCount","ExtraCount",
        "Giveaway","MissingSalesCodes","ExtraSalesCodes","Score"
    ])

    all_candidates_df = pd.DataFrame(all_candidates_rows, columns=[
        "VIN","HarnessFamily","PN","Score","IsBest","Giveaway","MissingSalesCodes","ExtraSalesCodes"
    ])

    # Final_BOM_By_VIN: pivot selections to VIN x family => PN
    bom = selections_df.pivot(index="VIN", columns="HarnessFamily", values="SelectedHarnessPN")
    bom = bom.reset_index()
    # stable column order: VIN first, then families sorted
    ordered_cols = ["VIN"] + sorted([c for c in bom.columns if c != "VIN"])
    bom = bom[ordered_cols]
    final_bom_df = bom

    return selections_df, all_candidates_df, final_bom_df


def find_same_score_ties(all_candidates_df: pd.DataFrame):
    if all_candidates_df.empty:
        return []
    ties = []
    grouped = all_candidates_df.groupby(["VIN", "HarnessFamily"], dropna=False)
    for (vin, family), grp in grouped:
        if grp.empty:
            continue
        max_score = grp["Score"].max()
        top = grp[grp["Score"] == max_score].copy()
        if len(top) > 1:
            ties.append((vin, family, top.reset_index(drop=True)))
    return ties


def _split_codes_cell(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return []
    s = str(val).strip()
    return [x.strip() for x in s.split(",") if x.strip()]


def apply_tie_break_overrides(selections_df: pd.DataFrame, all_candidates_df: pd.DataFrame, overrides: dict):
    if not overrides:
        return selections_df, all_candidates_df

    selections = selections_df.copy()
    candidates = all_candidates_df.copy()

    for (vin, family), chosen_pn in overrides.items():
        mask_group = (candidates["VIN"] == vin) & (candidates["HarnessFamily"] == family)
        if not mask_group.any():
            continue
        candidates.loc[mask_group, "IsBest"] = False
        mask_chosen = mask_group & (candidates["PN"] == chosen_pn)
        if not mask_chosen.any():
            continue
        candidates.loc[mask_chosen, "IsBest"] = True

        chosen_row = candidates.loc[mask_chosen].iloc[0]
        mask_sel = (selections["VIN"] == vin) & (selections["HarnessFamily"] == family)
        if not mask_sel.any():
            continue

        required = int(selections.loc[mask_sel, "RequiredCount"].iloc[0])
        missing_codes = _split_codes_cell(chosen_row.get("MissingSalesCodes"))
        extra_codes = _split_codes_cell(chosen_row.get("ExtraSalesCodes"))
        missing_cnt = len(missing_codes)
        extra_cnt = len(extra_codes)
        matched_cnt = max(required - missing_cnt, 0)
        if missing_cnt == 0 and extra_cnt == 0:
            status = "EXACT"
        elif missing_cnt == 0:
            status = "OVERBUILT"
        else:
            status = "INCOMPLETE"

        selections.loc[mask_sel, "SelectedHarnessPN"] = chosen_pn
        selections.loc[mask_sel, "MatchStatus"] = status
        selections.loc[mask_sel, "MatchedCount"] = matched_cnt
        selections.loc[mask_sel, "MissingCount"] = missing_cnt
        selections.loc[mask_sel, "ExtraCount"] = extra_cnt
        selections.loc[mask_sel, "Giveaway"] = chosen_row.get("Giveaway")
        selections.loc[mask_sel, "MissingSalesCodes"] = chosen_row.get("MissingSalesCodes")
        selections.loc[mask_sel, "ExtraSalesCodes"] = chosen_row.get("ExtraSalesCodes")
        selections.loc[mask_sel, "Score"] = chosen_row.get("Score")

    return selections, candidates


class HarnessTieRevisionDialog:
    def __init__(self, parent, ties):
        self.result = None
        self._parent = parent
        self._ties = ties
        self._choices = {}
        self._options_by_key = {}
        self._check_vars_by_key = {}
        self._build_window()

    def _build_window(self):
        win = tk.Toplevel(self._parent)
        self._win = win
        win.title("Harness PN Tie Revision")
        win.geometry("1380x800")
        win.minsize(980, 620)
        win.attributes("-topmost", True)
        win.grab_set()

        # Explicit colors — override macOS dark mode
        BG       = "#F0F2F5"
        BOX_BG   = "#FFFFFF"
        BAR_BG   = "#2C3E50"
        BAR_FG   = "#FFFFFF"
        LABEL_FG = "#1A1A1A"
        SUB_FG   = "#555555"
        SCORE_FG = "#1A6FBF"
        HDR_BG   = "#E8EEF5"
        HDR_FG   = "#1F2D3D"
        ROW_ALT  = "#EBF3FB"
        BOLD_F   = ("Segoe UI", 10, "bold")
        NORM_F   = ("Segoe UI", 9)

        win.configure(bg=BG)

        tk.Label(win,
             text="Review tied top-score cases grouped by Harness Family. Select exactly one Harness PN per VIN.",
                 font=BOLD_F, bg=BG, fg=LABEL_FG,
                 ).pack(fill="x", padx=12, pady=(12, 4))

        outer = tk.Frame(win, bg=BG)
        outer.pack(fill="both", expand=True, padx=10, pady=(0, 8))

        canvas = tk.Canvas(outer, borderwidth=0, highlightthickness=0, bg=BG)
        ysb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=BG)
        inner.bind("<Configure>", lambda _e: canvas.configure(scrollregion=canvas.bbox("all")))
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(win_id, width=e.width))
        canvas.configure(yscrollcommand=ysb.set)
        ysb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        ties_by_family = {}
        for vin, family, cand_df in self._ties:
            ties_by_family.setdefault(family, []).append((vin, cand_df))

        for family_idx, family in enumerate(sorted(ties_by_family.keys()), start=1):
            family_items = ties_by_family[family]
            card = tk.Frame(inner, bg=BOX_BG, relief="solid", bd=1)
            card.pack(fill="x", padx=6, pady=6)

            title_bar = tk.Frame(card, bg=BAR_BG)
            title_bar.pack(fill="x")
            tk.Label(title_bar,
                     text=f"  Family Group {family_idx}:  {family}  |  VIN ties: {len(family_items)}  ",
                     font=BOLD_F, bg=BAR_BG, fg=BAR_FG, anchor="w",
                     ).pack(side="left", padx=4, pady=5)

            content = tk.Frame(card, bg=BOX_BG, padx=10, pady=8)
            content.pack(fill="x")

            for vin_idx, (vin, cand_df) in enumerate(sorted(family_items, key=lambda x: str(x[0])), start=1):
                group_key = (vin, family)
                top_score = cand_df["Score"].max()

                vin_box = tk.Frame(content, bg=BOX_BG, relief="groove", bd=1)
                vin_box.pack(fill="x", pady=(0, 10))

                vin_hdr = tk.Frame(vin_box, bg=HDR_BG)
                vin_hdr.pack(fill="x")
                tk.Label(
                    vin_hdr,
                    text=f" VIN {vin_idx}: {vin}   |   Top score: {top_score}",
                    font=BOLD_F,
                    bg=HDR_BG,
                    fg=HDR_FG,
                    anchor="w",
                ).pack(fill="x", padx=6, pady=4)

                row_hdr = tk.Frame(vin_box, bg=BOX_BG)
                row_hdr.pack(fill="x", padx=6, pady=(6, 2))
                tk.Label(row_hdr, text="Pick", width=5, anchor="w", bg=BOX_BG, fg=LABEL_FG, font=BOLD_F).grid(row=0, column=0, sticky="w")
                tk.Label(row_hdr, text="PN", width=16, anchor="w", bg=BOX_BG, fg=LABEL_FG, font=BOLD_F).grid(row=0, column=1, sticky="w")
                tk.Label(row_hdr, text="Score", width=8, anchor="w", bg=BOX_BG, fg=LABEL_FG, font=BOLD_F).grid(row=0, column=2, sticky="w")
                tk.Label(row_hdr, text="Giveaway", width=20, anchor="w", bg=BOX_BG, fg=LABEL_FG, font=BOLD_F).grid(row=0, column=3, sticky="w")
                tk.Label(row_hdr, text="MissingSalesCodes", width=32, anchor="w", bg=BOX_BG, fg=LABEL_FG, font=BOLD_F).grid(row=0, column=4, sticky="w")
                tk.Label(row_hdr, text="ExtraSalesCodes", width=32, anchor="w", bg=BOX_BG, fg=LABEL_FG, font=BOLD_F).grid(row=0, column=5, sticky="w")

                current_best = cand_df[cand_df["IsBest"] == True]
                default_pn = str(current_best.iloc[0]["PN"]) if not current_best.empty else str(cand_df.iloc[0]["PN"])
                options = [str(x) for x in cand_df["PN"].tolist()]

                choice_var = tk.StringVar(value=default_pn)
                self._choices[group_key] = choice_var
                self._options_by_key[group_key] = options
                self._check_vars_by_key[group_key] = {}

                for row_i, (_, r) in enumerate(cand_df.iterrows()):
                    pn = str(r.get("PN", ""))
                    is_default = pn == default_pn
                    check_var = tk.BooleanVar(value=is_default)
                    self._check_vars_by_key[group_key][pn] = check_var

                    row_bg = BOX_BG if row_i % 2 == 0 else ROW_ALT
                    row = tk.Frame(vin_box, bg=row_bg)
                    row.pack(fill="x", padx=6, pady=1)

                    cb = tk.Checkbutton(
                        row,
                        variable=check_var,
                        bg=row_bg,
                        activebackground=row_bg,
                        command=lambda k=group_key, p=pn: self._on_candidate_toggle(k, p),
                    )
                    cb.grid(row=0, column=0, sticky="w")
                    tk.Label(row, text=pn, width=16, anchor="w", bg=row_bg, fg=LABEL_FG, font=NORM_F).grid(row=0, column=1, sticky="w")
                    tk.Label(row, text=str(r.get("Score", "")), width=8, anchor="w", bg=row_bg, fg=LABEL_FG, font=NORM_F).grid(row=0, column=2, sticky="w")
                    tk.Label(row, text=(r.get("Giveaway") or ""), width=20, anchor="w", bg=row_bg, fg=LABEL_FG, font=NORM_F).grid(row=0, column=3, sticky="w")
                    tk.Label(row, text=(r.get("MissingSalesCodes") or ""), width=32, anchor="w", bg=row_bg, fg=LABEL_FG, font=NORM_F).grid(row=0, column=4, sticky="w")
                    tk.Label(row, text=(r.get("ExtraSalesCodes") or ""), width=32, anchor="w", bg=row_bg, fg=LABEL_FG, font=NORM_F).grid(row=0, column=5, sticky="w")

        bar = tk.Frame(win, bg=BG)
        bar.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(bar, text="Auto Select First Occurrence",
                   command=self._auto_select_first).pack(side="left", padx=4)
        ttk.Button(bar, text="Cancel", command=self._on_cancel).pack(side="right", padx=4)
        ttk.Button(bar, text="Apply Selections", command=self._on_apply).pack(side="right", padx=4)

        win.wait_window()

    def _on_candidate_toggle(self, key, chosen_pn):
        vars_by_pn = self._check_vars_by_key.get(key, {})
        chosen_var = vars_by_pn.get(chosen_pn)
        if chosen_var is None:
            return

        if chosen_var.get():
            for pn, var in vars_by_pn.items():
                if pn != chosen_pn:
                    var.set(False)
            self._choices[key].set(chosen_pn)
            return

        # Enforce one selection minimum: do not allow all unchecked.
        if not any(var.get() for var in vars_by_pn.values()):
            chosen_var.set(True)
            self._choices[key].set(chosen_pn)

    def _auto_select_first(self):
        for key, opts in self._options_by_key.items():
            if not opts:
                continue
            first = opts[0]
            self._choices[key].set(first)
            for pn, var in self._check_vars_by_key.get(key, {}).items():
                var.set(pn == first)

    def _on_cancel(self):
        self.result = None
        self._win.destroy()

    def _on_apply(self):
        self.result = {k: v.get() for k, v in self._choices.items()}
        self._win.destroy()


# ======================================================
# Interactive SalesCode Review & Selection Dialog
# ======================================================
class SalesCodeReviewDialog:


    _STATUS_BG = {
        "STANDARD": "#C6EFCE",
        "OPTIONAL": "#FFEB9C",
        "UNUSED":    "#FFC7CE",
    }
    _STATUS_FG = {
        "STANDARD": "#276221",
        "OPTIONAL": "#7D6608",
        "UNUSED":    "#9C0006",
    }

    def __init__(self, parent,
                 family_stats_df: "pd.DataFrame",
                 global_code_df:  "pd.DataFrame"):
        self.result = None
        self._family_stats = family_stats_df
        self._parent        = parent

        self._families = sorted(self._family_stats["HarnessFamily"].unique().tolist())
        self._current_family = tk.StringVar(value=self._families[0] if self._families else "")
        self._profile_loaded = False

        # Build family -> status -> sorted codes map.
        self._family_status_codes = {}
        self._family_code_vars = {}
        self._family_reviewed = {}
        for fam in self._families:
            sub = self._family_stats[self._family_stats["HarnessFamily"] == fam]
            std_codes = sorted(sub[sub["Status"] == "STANDARD"]["SalesCode"].unique().tolist())
            opt_codes = sorted(sub[sub["Status"] == "OPTIONAL"]["SalesCode"].unique().tolist())
            self._family_status_codes[fam] = {
                "STANDARD": std_codes,
                "OPTIONAL": opt_codes,
            }
            all_codes = sorted(set(std_codes + opt_codes))
            self._family_code_vars[fam] = {c: tk.BooleanVar(value=True) for c in all_codes}
            self._family_reviewed[fam] = False

        self._status_lbl = None
        self._continue_btn_bottom = None
        self._continue_btn_tab = None
        self._export_btn = None
        self._progress_tree = None
        self._codes_inner = None

        self._build_window()

    # ------------------------------------------------------------------ build
    def _build_window(self):
        win = tk.Toplevel(self._parent)
        self._win = win
        win.title("SalesCode Analysis & Selection")
        win.geometry("1300x820")
        win.minsize(960, 640)
        win.resizable(True, True)
        win.attributes('-topmost', True)
        win.grab_set()

        # style tweaks
        style = ttk.Style(win)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("Heading.TLabel", font=("Segoe UI", 10, "bold"))

        nb = ttk.Notebook(win)
        nb.pack(fill="both", expand=True, padx=6, pady=(6, 0))

        t1 = ttk.Frame(nb); nb.add(t1, text="  Per-Family Analysis  ")
        t2 = ttk.Frame(nb); nb.add(t2, text="  Code Selection  ")

        self._build_family_tab(t1)
        self._build_selection_tab(t2)

        # bottom bar
        bar = ttk.Frame(win, relief="groove")
        bar.pack(fill="x", padx=6, pady=6)

        self._status_lbl = ttk.Label(bar, text="", font=("Segoe UI", 9))
        self._status_lbl.pack(side="left", padx=10)
        self._refresh_status_label()

        ttk.Button(bar, text="Cancel",
                   command=win.destroy).pack(side="right", padx=6, pady=3)
        self._continue_btn_bottom = ttk.Button(bar, text="✓  Continue",
                               command=self._on_proceed)
        self._continue_btn_bottom.pack(side="right", padx=2, pady=3)

        self._refresh_selection_readiness()
        win.wait_window()

    # --------------------------------------------------- Per-Family tab
    def _build_family_tab(self, parent):
        top = ttk.Frame(parent)
        top.pack(fill="x", padx=8, pady=6)

        ttk.Label(top, text="Harness Family:").pack(side="left")
        families = ["(All)"] + sorted(self._family_stats["HarnessFamily"].unique())
        self._fam_filter_var = tk.StringVar(value="(All)")
        cb = ttk.Combobox(top, textvariable=self._fam_filter_var,
                          values=families, width=42, state="readonly")
        cb.pack(side="left", padx=6)
        cb.bind("<<ComboboxSelected>>", lambda _e: self._refresh_family_tree())

        ttk.Label(top, text="Status:").pack(side="left", padx=(14, 4))
        self._fam_status_var = tk.StringVar(value="(All)")
        sf = ttk.Combobox(top, textvariable=self._fam_status_var,
                          values=["(All)", "STANDARD", "OPTIONAL", "UNUSED"],
                          width=14, state="readonly")
        sf.pack(side="left")
        sf.bind("<<ComboboxSelected>>", lambda _e: self._refresh_family_tree())

        # legend
        leg = ttk.Frame(parent)
        leg.pack(fill="x", padx=8)
        for status, bg in self._STATUS_BG.items():
            fr = tk.Frame(leg, bg=bg, width=14, height=14)
            fr.pack(side="left", padx=(0, 2))
            ttk.Label(leg, text=status, foreground=self._STATUS_FG[status],
                      font=("Segoe UI", 8, "bold")).pack(side="left", padx=(0, 12))

        cols = ("HarnessFamily", "SalesCode", "PNsWithCode", "TotalPNs", "Coverage_Pct", "Status")
        tree = self._make_tree(parent, cols, {
            "HarnessFamily": 220, "SalesCode": 90, "PNsWithCode": 110,
            "TotalPNs": 90, "Coverage_Pct": 110, "Status": 110,
        })
        self._family_tree = tree
        for s, bg in self._STATUS_BG.items():
            tree.tag_configure(s, background=bg, foreground=self._STATUS_FG[s])
        self._refresh_family_tree()

    def _refresh_family_tree(self):
        tree = self._family_tree
        tree.delete(*tree.get_children())
        df = self._family_stats
        fsel = self._fam_filter_var.get()
        ssel = self._fam_status_var.get()
        if fsel != "(All)":
            df = df[df["HarnessFamily"] == fsel]
        if ssel != "(All)":
            df = df[df["Status"] == ssel]
        for _, row in df.iterrows():
            st = row["Status"]
            tree.insert("", "end", values=(
                row["HarnessFamily"], row["SalesCode"],
                row["PNsWithCode"], row["TotalPNs"],
                f"{row['Coverage_Pct']:.1f}%", st,
            ), tags=(st,))

    # --------------------------------------------------- Code Selection tab
    def _build_selection_tab(self, parent):
        top = ttk.Frame(parent)
        top.pack(fill="x", padx=8, pady=6)

        ttk.Label(top, text="Review each harness family box and mark it reviewed.",
                  font=("Segoe UI", 10, "bold")).pack(side="left")

        right_actions = ttk.Frame(top)
        right_actions.pack(side="right")
        ttk.Button(right_actions, text="Mark All Reviewed", command=self._mark_all_reviewed).pack(side="left", padx=2)
        ttk.Button(right_actions, text="Load Selection Profile", command=self._load_selection_profile).pack(side="left", padx=2)
        self._export_btn = ttk.Button(right_actions, text="Export Selection Profile", command=self._export_selection_profile)
        self._export_btn.pack(side="left", padx=2)
        self._continue_btn_tab = ttk.Button(right_actions, text="Continue", command=self._on_proceed)
        self._continue_btn_tab.pack(side="left", padx=2)

        self._selection_summary_lbl = ttk.Label(parent, text="", font=("Segoe UI", 9))
        self._selection_summary_lbl.pack(fill="x", padx=8)

        outer = ttk.Frame(parent)
        outer.pack(fill="both", expand=True, padx=8, pady=4)

        self._codes_canvas = tk.Canvas(outer, borderwidth=0, highlightthickness=0, bg="#f8f8f8")
        ysb = ttk.Scrollbar(outer, orient="vertical", command=self._codes_canvas.yview)
        self._codes_inner = ttk.Frame(self._codes_canvas)
        self._codes_inner.bind(
            "<Configure>",
            lambda _e: self._codes_canvas.configure(scrollregion=self._codes_canvas.bbox("all"))
        )

        # Keep the inner frame stretched to the full visible canvas width.
        self._codes_window = self._codes_canvas.create_window((0, 0), window=self._codes_inner, anchor="nw")
        self._codes_canvas.bind(
            "<Configure>",
            lambda e: self._codes_canvas.itemconfigure(self._codes_window, width=e.width)
        )
        self._codes_canvas.configure(yscrollcommand=ysb.set)
        ysb.pack(side="right", fill="y")
        self._codes_canvas.pack(side="left", fill="both", expand=True)

        self._render_all_family_boxes()
        self._refresh_selection_readiness()

    def _render_all_family_boxes(self):
        if self._codes_inner is None:
            return
        for w in self._codes_inner.winfo_children():
            w.destroy()

        for fam in self._families:
            reviewed = self._family_reviewed.get(fam, False)
            mark = "✅" if reviewed else "⬜"
            fam_box = tk.LabelFrame(self._codes_inner, text=f"{mark} {fam}", bg="#ffffff", padx=8, pady=6)
            fam_box.pack(fill="x", padx=8, pady=6)

            # Explicit heading inside the card so family name is always obvious.
            tk.Label(
                fam_box,
                text=f"Harness Family: {fam}",
                bg="#ffffff",
                fg="#11314f",
                font=("Segoe UI", 10, "bold")
            ).pack(anchor="w", pady=(0, 4))

            head = ttk.Frame(fam_box)
            head.pack(fill="x", pady=(0, 4))
            ttk.Button(head, text="Select all", command=lambda f=fam: self._set_family_checks(f, True)).pack(side="left", padx=2)
            ttk.Button(head, text="Deselect all", command=lambda f=fam: self._set_family_checks(f, False)).pack(side="left", padx=2)
            ttk.Button(head, text="Mark Reviewed", command=lambda f=fam: self._mark_family_reviewed(f)).pack(side="left", padx=8)

            vars_map = self._family_code_vars.get(fam, {})
            std_codes = self._family_status_codes[fam]["STANDARD"]
            opt_codes = self._family_status_codes[fam]["OPTIONAL"]

            self._render_family_code_group(fam_box, fam, "STANDARD SalesCodes", std_codes, vars_map, "#C6EFCE")
            self._render_family_code_group(fam_box, fam, "OPTIONAL SalesCodes", opt_codes, vars_map, "#FFEB9C")

        self._update_selection_summary()

    def _render_family_code_group(self, parent, family_name, title, codes, vars_map, bg):
        box = tk.LabelFrame(parent, text=title, bg=bg, padx=8, pady=6)
        box.pack(fill="x", padx=4, pady=4)
        if not codes:
            tk.Label(box, text="(none)", bg=bg).pack(anchor="w")
            return
        cols = 6
        for i, code in enumerate(codes):
            r, c = divmod(i, cols)
            tk.Checkbutton(
                box,
                text=code,
                variable=vars_map[code],
                bg=bg,
                activebackground=bg,
                command=lambda f=family_name: self._on_family_code_toggle(f),
            ).grid(row=r, column=c, sticky="w", padx=4, pady=2)

    def _on_family_code_toggle(self, family_name: str):
        self._family_reviewed[family_name] = False
        self._profile_loaded = False
        self._render_all_family_boxes()
        self._refresh_status_label()
        self._refresh_selection_readiness()

    def _set_family_checks(self, family_name: str, value: bool):
        for var in self._family_code_vars.get(family_name, {}).values():
            var.set(value)
        self._family_reviewed[family_name] = False
        self._profile_loaded = False
        self._render_all_family_boxes()
        self._refresh_status_label()
        self._refresh_selection_readiness()

    def _mark_family_reviewed(self, family_name: str):
        self._family_reviewed[family_name] = True
        self._render_all_family_boxes()
        self._refresh_status_label()
        self._refresh_selection_readiness()

    def _mark_all_reviewed(self):
        for fam in self._families:
            self._family_reviewed[fam] = True
        self._render_all_family_boxes()
        self._refresh_status_label()
        self._refresh_selection_readiness()

    def _update_selection_summary(self):
        reviewed = sum(1 for fam in self._families if self._family_reviewed.get(fam, False))
        total_codes = sum(len(self._family_code_vars.get(fam, {})) for fam in self._families)
        selected_codes = sum(
            sum(1 for v in self._family_code_vars.get(fam, {}).values() if v.get())
            for fam in self._families
        )
        if hasattr(self, "_selection_summary_lbl") and self._selection_summary_lbl is not None:
            self._selection_summary_lbl.config(
                text=f"Harnesses reviewed: {reviewed}/{len(self._families)}   |   "
                     f"Selected: {selected_codes}/{total_codes}"
            )

    def _all_reviewed(self):
        return bool(self._families) and all(self._family_reviewed.get(f, False) for f in self._families)

    def _refresh_selection_readiness(self):
        ready = self._all_reviewed() or self._profile_loaded
        state = "normal" if ready else "disabled"
        if self._continue_btn_tab is not None:
            self._continue_btn_tab.configure(state=state)
        if self._continue_btn_bottom is not None:
            self._continue_btn_bottom.configure(state=state)
        if self._export_btn is not None:
            self._export_btn.configure(state=("normal" if self._all_reviewed() else "disabled"))

    def _build_profile_dataframe(self):
        rows = []
        for fam in self._families:
            vars_map = self._family_code_vars.get(fam, {})
            std_set = set(self._family_status_codes[fam]["STANDARD"])
            opt_set = set(self._family_status_codes[fam]["OPTIONAL"])
            for code in sorted(vars_map.keys()):
                if code in std_set:
                    status = "STANDARD"
                elif code in opt_set:
                    status = "OPTIONAL"
                else:
                    status = "UNKNOWN"
                rows.append({
                    "HarnessFamily": fam,
                    "SalesCode": code,
                    "Status": status,
                    "Selected": bool(vars_map[code].get()),
                    "Reviewed": bool(self._family_reviewed.get(fam, False)),
                })
        return pd.DataFrame(rows)

    def _export_selection_profile(self):
        if not self._all_reviewed():
            messagebox.showinfo("Not Ready", "Review all harness families before exporting.", parent=self._win)
            return
        path = filedialog.asksaveasfilename(
            title="Export SalesCode Selection Profile",
            defaultextension=".xlsx",
            initialfile="SalesCode_Selection_Profile.xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            parent=self._win,
        )
        if not path:
            return
        df = self._build_profile_dataframe()
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Family_Code_Selection", index=False)
        messagebox.showinfo("Exported", f"Selection profile saved:\n{path}", parent=self._win)

    def _load_selection_profile(self):
        path = filedialog.askopenfilename(
            title="Load SalesCode Selection Profile",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")],
            parent=self._win,
        )
        if not path:
            return
        try:
            df = pd.read_excel(path, sheet_name="Family_Code_Selection", engine="openpyxl")
        except Exception:
            df = pd.read_excel(path, engine="openpyxl")

        required = {"HarnessFamily", "SalesCode", "Selected"}
        if not required.issubset(set(df.columns)):
            messagebox.showerror("Invalid File", "Profile must include columns: HarnessFamily, SalesCode, Selected", parent=self._win)
            return

        # Reset current selections first.
        for fam in self._families:
            for var in self._family_code_vars.get(fam, {}).values():
                var.set(False)

        for _, row in df.iterrows():
            fam = str(row.get("HarnessFamily", "")).strip()
            code = str(row.get("SalesCode", "")).strip().upper()
            sel = row.get("Selected", False)
            if fam in self._family_code_vars and code in self._family_code_vars[fam]:
                is_selected = bool(sel)
                if isinstance(sel, str):
                    is_selected = sel.strip().upper() in ("TRUE", "1", "YES", "Y")
                self._family_code_vars[fam][code].set(is_selected)

        for fam in self._families:
            self._family_reviewed[fam] = True

        self._profile_loaded = True
        self._render_all_family_boxes()
        self._refresh_status_label()
        self._refresh_selection_readiness()
        messagebox.showinfo("Loaded", "Selection profile loaded successfully.", parent=self._win)

    def _refresh_status_label(self):
        total = 0
        selected = 0
        for fam in self._families:
            vars_map = self._family_code_vars.get(fam, {})
            total += len(vars_map)
            selected += sum(1 for v in vars_map.values() if v.get())
        reviewed = sum(1 for fam in self._families if self._family_reviewed.get(fam, False))
        if self._status_lbl is None:
            return
        self._status_lbl.config(
            text=f"Reviewed harnesses: {reviewed}/{len(self._families)}   |   "
                 f"Selected codes: {selected}/{total}   |   Excluded: {total - selected}")

    # --------------------------------------------------- Treeview helper
    def _make_tree(self, parent, cols, col_widths=None):
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True, padx=6, pady=4)
        vsb = ttk.Scrollbar(frame, orient="vertical")
        hsb = ttk.Scrollbar(frame, orient="horizontal")
        tree = ttk.Treeview(frame, columns=cols, show="headings",
                            yscrollcommand=vsb.set, xscrollcommand=hsb.set,
                            selectmode="browse")
        vsb.configure(command=tree.yview)
        hsb.configure(command=tree.xview)
        for col in cols:
            w = (col_widths or {}).get(col, 100)
            tree.heading(col, text=col,
                         command=lambda c=col: self._sort_tree(tree, c, False))
            tree.column(col, width=w, minwidth=50, anchor="center")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        return tree

    def _sort_tree(self, tree, col, reverse):
        data = [(tree.set(k, col), k) for k in tree.get_children("")]
        try:
            data.sort(key=lambda t: float(t[0].rstrip('%')), reverse=reverse)
        except ValueError:
            data.sort(key=lambda t: t[0].lower(), reverse=reverse)
        for idx, (_, k) in enumerate(data):
            tree.move(k, "", idx)
        tree.heading(col, command=lambda: self._sort_tree(tree, col, not reverse))

    # --------------------------------------------------- proceed / cancel
    def _on_proceed(self):
        if not (self._all_reviewed() or self._profile_loaded):
            messagebox.showinfo("Not Ready", "Review all harness families or load a profile before continuing.", parent=self._win)
            return
        self.result = {
            fam: {code for code, var in self._family_code_vars.get(fam, {}).items() if var.get()}
            for fam in self._families
        }
        self._win.destroy()


# ======================================================
# Filter families by user-selected codes
# ======================================================
def filter_per_file_families(per_file_families: list, selected_codes_by_family: dict) -> list:
    """
    Return a copy of per_file_families where each family's header_codes and
    per-PN code sets are restricted to that family's selected SalesCodes.
    """
    filtered = []
    for fam in per_file_families:
        family_name = fam["family"]
        chosen = selected_codes_by_family.get(family_name)
        if chosen is None:
            chosen = set(fam["header_codes"])
        else:
            chosen = set(chosen)
        filtered.append({
            "family": family_name,
            "header_codes": fam["header_codes"] & chosen,
            "pns": [(pn, codes & chosen, giveaway_codes & chosen) for pn, codes, giveaway_codes in fam["pns"]],
        })
    return filtered

# ==============================
# Populate formatted Template.xlsx
# ==============================
def ask_my_and_program():
    """Prompt user for MY (model year) and Program (e.g., RU)."""
    from tkinter import simpledialog
    root = Tk(); root.withdraw(); root.attributes('-topmost', True)
    
    my = simpledialog.askstring(
        "Model Year",
        "Enter Model Year (MY):\n(e.g., 27)",
        parent=root
    )
    if not my:
        root.destroy()
        return None, None
    
    program = simpledialog.askstring(
        "Program",
        "Enter Program:\n(e.g., RU)",
        parent=root
    )
    root.destroy()
    return my.strip() if my else None, program.strip() if program else None

def ask_save_folder():
    """Prompt user to select folder to save the populated template."""
    root = Tk(); root.withdraw(); root.attributes('-topmost', True)
    folder = filedialog.askdirectory(
        title="Select folder to save the populated template"
    )
    root.destroy()
    return folder


def ask_output_folder(default_dir: str = ""):
    """Prompt user to select where all output files should be saved."""
    root = Tk(); root.withdraw(); root.attributes('-topmost', True)
    folder = filedialog.askdirectory(
        title="Select folder to save output files",
        initialdir=(default_dir or os.getcwd())
    )
    root.destroy()
    return folder

def create_formatted_output(template_path: str, my: str, program: str, out_dir: str,
                           selections_df: pd.DataFrame, vin_matrix_df: pd.DataFrame):
    """
    Copy Template.xlsx with all formatting preserved,
    populate with VINs and harness selections,
    rename to '{MY_last_2_digits}_{Program}_VBOM_Template_for_DEFE.xlsx'
    e.g., 27_RU_VBOM_Template_for_DEFE.xlsx (where MY=2027)
    """
    from openpyxl import load_workbook
    
    try:
        # Extract last 2 digits of MY
        my_short = my[-2:] if len(my) >= 2 else my
        
        # Load template
        wb_new = load_workbook(template_path)
        ws_new = wb_new.active
        
        # Clear data rows (keep header formatting in rows 1-2)
        # Clear from row 3 onwards
        for r in range(3, ws_new.max_row + 1):
            for c in range(1, ws_new.max_column + 1):
                cell = ws_new.cell(row=r, column=c)
                cell.value = None
        
        # Populate VINs in row 2, starting from column C (col 3)
        vin_order = list(vin_matrix_df["VIN"])
        start_col = 3  # Column C
        for idx, vin in enumerate(vin_order):
            ws_new.cell(row=2, column=start_col + idx, value=vin)
        
        # Clear any leftover VINs beyond current list
        for c in range(start_col + len(vin_order), ws_new.max_column + 1):
            ws_new.cell(row=2, column=c, value=None)
        
        # Build mapping: family -> PN -> set(VINs)
        family_map = {}
        for _, row in selections_df.iterrows():
            family = str(row["HarnessFamily"]).strip()
            pn = str(row["SelectedHarnessPN"]).strip()
            vin = row["VIN"]
            family_map.setdefault(family, {}).setdefault(pn, set()).add(vin)
        
        # Build ordered entries
        entries = []
        for family in sorted(family_map.keys()):
            pn_map = family_map[family]
            for pn in sorted(pn_map.keys()):
                entries.append((pn, f"WIRING ASSY -{family}", pn_map[pn]))
        
        # Write rows
        for i, (pn, harness_name, vin_set) in enumerate(entries):
            r = 3 + i
            ws_new.cell(row=r, column=1, value=pn)
            ws_new.cell(row=r, column=2, value=harness_name)
            for j, vin in enumerate(vin_order):
                ws_new.cell(row=r, column=start_col + j, value=("x" if vin in vin_set else ""))
        
        # Clear any remaining old rows beyond the new entries
        for r in range(3 + len(entries), ws_new.max_row + 1):
            ws_new.cell(row=r, column=1, value=None)
            ws_new.cell(row=r, column=2, value=None)
            for j in range(len(vin_order)):
                ws_new.cell(row=r, column=start_col + j, value=None)
        
        # Save with new name
        output_name = f"{my_short}_{program}_VBOM_Template_for_DEFE.xlsx"
        output_path = os.path.join(out_dir, output_name)
        wb_new.save(output_path)
        
        return output_path
    except Exception as e:
        print(f"ERROR: Could not create formatted output file: {e}")
        import traceback
        traceback.print_exc()
        return None

# =========
# Main app
# =========
def main():
    # Startup setup window (MY + Program + input files)
    setup = collect_run_setup_inputs()
    if not setup:
        try:
            root = Tk(); root.withdraw()
            messagebox.showinfo("Canceled", "Setup canceled. Exiting.")
            root.destroy()
        except Exception:
            pass
        return

    my = setup["my"]
    program = setup["program"]
    input_type = setup.get("input_type", "doall")
    path = setup.get("input_file") or setup.get("build_spec")
    complexity_files = setup["harness_files"]
    out_dir = setup.get("output_dir") or os.path.dirname(path)

    print("\n" + "="*80)
    if input_type == "buildspec":
        print("STEP 1/5: LOAD BUILDSPEC FILE")
    else:
        print("STEP 1/5: LOAD DOALL FILE")
    print("="*80)
    print(f"✓ Loaded: {os.path.basename(path)}")

    # 2) VIN matrix
    print("\nProcessing VIN specifications...")
    vin_matrix_df, vin_codes_sorted = build_vin_matrix(path, source_type=input_type)
    print(f"✓ Found {len(vin_matrix_df)} VINs with {len(vin_codes_sorted)} unique sales codes")

    # 4) Harness files were selected in startup setup window
    print("\n" + "="*80)
    print("STEP 2/5: LOAD HARNESS COMPLEXITY FILES")
    print("="*80)
    if not complexity_files:
        try:
            root = Tk(); root.withdraw()
            messagebox.showinfo("No files selected", "No Complexity files selected. Done with VIN/spec only.")
            root.destroy()
        except Exception:
            pass
        return

    # 5) Read each Complexity and build data for master + outputs
    print("\nProcessing complexity files...")
    per_file_master = []  # (file, df_complexity)
    per_file_families = []  # dicts for selections/candidates
    all_complexity_codes = set()

    for f in complexity_files:
        try:
            df_comp, header_codes, pn_rows = read_complexity_sheet(f)
            family = try_get_harness_family(f)
            per_file_master.append((f, df_comp))
            per_file_families.append({
                "family": family,
                "header_codes": set(header_codes),
                "pns": pn_rows
            })
            all_complexity_codes.update(header_codes)
            print(f"✓ {os.path.basename(f)} | {family} | {len(header_codes)} codes | {len(pn_rows)} PNs")
        except Exception as e:
            print(f"✗ SKIPPED '{os.path.basename(f)}': {e}")

    if not per_file_master:
        print("No valid Complexity sheets read. Exiting.")
        return

    # 6) Prepare all output datasets (files are written after final output folder selection)
    print("\n" + "="*80)
    print("STEP 3/5: PROCESS DATA")
    print("="*80)
    print("✓ Data prepared for output files")

    # 7) SalesCode_Diff dataset (sheet is written after final output folder selection)
    vin_code_set = set(vin_codes_sorted)
    complexity_code_set = set(all_complexity_codes)
    diff_df = build_salescode_diff(vin_code_set, complexity_code_set)
    print(f"✓ SalesCode_Diff prepared")

    # Excluded_SalesCodes
    excluded_codes = sorted(list(vin_code_set - complexity_code_set))
    excluded_df = pd.DataFrame({"SalesCode_Not_In_Any_Harness": excluded_codes})

    # SalesCode statistics — computed BEFORE user review so the dialog can show them
    print("\nComputing sales code statistics...")
    family_stats_df, global_code_df = build_salescode_statistics(per_file_families)
    if not family_stats_df.empty:
        standard_pairs = int((family_stats_df["Status"] == "STANDARD").sum())
        unused_pairs    = int((family_stats_df["Status"] == "UNUSED").sum())
        print(f"✓ Stats: {standard_pairs} standard code/family pairs, "
              f"{unused_pairs} unused code/family pairs")

    # Show interactive review & code-selection dialog
    print("\nOpening SalesCode Analysis & Selection window...")
    _rev_root = Tk(); _rev_root.withdraw(); _rev_root.attributes('-topmost', True)
    if not family_stats_df.empty:
        _dialog = SalesCodeReviewDialog(_rev_root, family_stats_df,
                                        global_code_df)
        selected_codes_by_family = _dialog.result
    else:
        selected_codes_by_family = {
            fam["family"]: set(fam["header_codes"])
            for fam in per_file_families
        }
    _rev_root.destroy()

    if selected_codes_by_family is None:
        print("User cancelled the review. Exiting.")
        try:
            root = Tk(); root.withdraw()
            messagebox.showinfo("Cancelled", "Process cancelled at code-review step.")
            root.destroy()
        except Exception:
            pass
        return

    families_for_matching = filter_per_file_families(per_file_families, selected_codes_by_family)
    selected_union = set()
    for codes in selected_codes_by_family.values():
        selected_union.update(codes)
    excluded_by_user = all_complexity_codes - selected_union
    if excluded_by_user:
        print(f"✓ User excluded {len(excluded_by_user)} sales code(s) from matching: "
              f"{', '.join(sorted(excluded_by_user))}")

    # 8) Build Selections, AllCandidates, Final_BOM
    print("\nMatching harnesses to VINs...")
    selections_df, all_candidates_df, final_bom_df = build_outputs(
        vin_matrix_df, families_for_matching)
    print(f"✓ Generated selections for {len(selections_df)} VIN-Family pairs")

    # 8b) Tie-revision dialog — show when 2+ PNs share the top score
    ties = find_same_score_ties(all_candidates_df)
    if ties:
        print(f"✓ Found {len(ties)} tie case(s) — opening Harness PN Tie Revision window...")
        _tie_root = Tk()
        _tie_root.withdraw()
        _tie_root.attributes("-topmost", True)
        tie_dlg = HarnessTieRevisionDialog(_tie_root, ties)
        overrides = tie_dlg.result
        _tie_root.destroy()

        if overrides is None:
            print("User cancelled at tie revision. Exiting.")
            try:
                root = Tk(); root.withdraw()
                messagebox.showinfo("Cancelled", "Process cancelled at tie-revision step.")
                root.destroy()
            except Exception:
                pass
            return

        selections_df, all_candidates_df = apply_tie_break_overrides(
            selections_df, all_candidates_df, overrides)
        # Rebuild BOM from updated selections
        bom = selections_df.pivot(index="VIN", columns="HarnessFamily",
                                   values="SelectedHarnessPN").reset_index()
        final_bom_df = bom[["VIN"] + sorted([c for c in bom.columns if c != "VIN"])]
        print(f"✓ Tie overrides applied")
    else:
        print("✓ No ties detected")

    # Final output location prompt (after processing, before writing files)
    chosen_out_dir = ask_output_folder(out_dir)
    if chosen_out_dir:
        out_dir = chosen_out_dir
        print(f"✓ Output folder selected: {out_dir}")
    else:
        print(f"✓ Output folder not changed (using): {out_dir}")

    # Resolve final output paths only after user picks destination.
    master_path = os.path.join(out_dir, MASTER_FILE_NAME)
    vin_out_path = os.path.join(out_dir, VIN_MATRIX_FILE)
    selections_out = os.path.join(out_dir, SELECTIONS_FILE)

    # Write Master workbook in final selected folder.
    used_sheetnames = set()
    with pd.ExcelWriter(master_path, engine="openpyxl") as writer:
        for f, df_comp in per_file_master:
            sheet = safe_sheetname(os.path.basename(f), used_sheetnames)
            df_comp.to_excel(writer, sheet_name=sheet, index=False, header=False)
    print(f"✓ Created: {MASTER_FILE_NAME}")

    # Write VIN matrix and append SalesCode_Diff in final selected folder.
    vin_matrix_df.to_excel(vin_out_path, index=False, engine="openpyxl")
    write_df_to_excel_append(vin_out_path, "SalesCode_Diff", diff_df)
    print(f"✓ Created: {VIN_MATRIX_FILE} (with SalesCode_Diff)")

    # 9) Save VIN_to_Harness_Selection.xlsx
    print("\n" + "="*80)
    print("STEP 4/5: GENERATE OUTPUT FILES")
    print("="*80)
    with pd.ExcelWriter(selections_out, engine="openpyxl") as writer:
        selections_df.to_excel(writer, sheet_name="Selections", index=False)
        all_candidates_df.to_excel(writer, sheet_name="AllCandidates", index=False)
        excluded_df.to_excel(writer, sheet_name="Excluded_SalesCodes", index=False)
        final_bom_df.to_excel(writer, sheet_name="Final_BOM_By_VIN", index=False)
        if not family_stats_df.empty:
            family_stats_df.to_excel(writer, sheet_name="Family_Code_Stats", index=False)
            global_code_df.to_excel(writer, sheet_name="Global_Code_Overview", index=False)
    print(f"✓ Created: {SELECTIONS_FILE}")

    # Format AllCandidates: highlight ONLY rows where IsBest == TRUE in soft orange
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        wb = load_workbook(selections_out)
        ws = wb["AllCandidates"]

        # Freeze top row & autofilter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Soft orange (peach) highlight
        highlight_fill = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")

        # Find the IsBest column index by header text
        header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        isbest_col = header_map.get("IsBest", None)

        if isbest_col is not None:
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=isbest_col).value
                # Accept True, 'TRUE', 1 as true-ish
                is_true = (
                    (isinstance(val, bool) and val) or
                    (isinstance(val, (int, float)) and val == 1) or
                    (isinstance(val, str) and val.strip().upper() == "TRUE")
                )
                if is_true:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = highlight_fill

        wb.save(selections_out)
        print(f"✓ Highlighted IsBest rows in AllCandidates sheet")
    except Exception as e:
        print(f"✓ Formatting applied (note: {e})")

    # Format SalesCode statistics sheets
    try:
        _format_stats_sheets(selections_out)
        print(f"✓ Formatted SalesCode statistics sheets")
    except Exception as e:
        print(f"✓ Statistics sheets created (note: {e})")

    # 9.5) Save formatted template in the same output folder used for all files
    print("\n" + "="*80)
    print("STEP 5/5: SAVE FORMATTED TEMPLATE")
    print("="*80)
    save_folder = out_dir
    formatted_template_path = None

    # Create and populate Template.xlsx
    template_src_path = os.path.join(os.path.dirname(__file__), TEMPLATE_SOURCE_FILE)
    if not os.path.exists(template_src_path):
        template_src_path = os.path.join(out_dir, TEMPLATE_SOURCE_FILE)

    if os.path.exists(template_src_path):
        try:
            formatted_template_path = create_formatted_output(
                template_src_path,
                my,
                program,
                save_folder,
                selections_df,
                vin_matrix_df
            )
            if formatted_template_path:
                print(f"✓ Formatted template created: {os.path.basename(formatted_template_path)}")
        except Exception as e:
            print(f"✗ Could not create formatted template: {e}")
            import traceback
            traceback.print_exc()
    else:
        print(f"⚠ Template.xlsx not found at: {template_src_path}")

    # Final completion dialog
    print("\n" + "="*80)
    print("✅ PROCESS COMPLETE!")
    print("="*80)
    
    try:
        root = Tk()
        root.withdraw()
        
        template_info = ""
        if formatted_template_path:
            template_info += f"  ✅ Formatted Template:\n      {os.path.basename(formatted_template_path)}\n"
        if not template_info:
            template_info = "  ⚠ Templates: Not saved\n"
        
        completion_msg = f"""✅ PROCESS COMPLETE!

📊 SUMMARY:
  • VINs Processed: {len(vin_matrix_df)}
  • Harness Families: {len(per_file_families)}
  • Selections Created: {len(selections_df)}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📁 OUTPUT LOCATIONS:

Main Output Folder:
  {out_dir}

Data Files:
  • {MASTER_FILE_NAME}
  • {SELECTIONS_FILE}
  • {VIN_MATRIX_FILE}

{template_info}
All files are ready for use!"""
        
        messagebox.showinfo("✅ Success", completion_msg)
        root.destroy()
    except Exception as e:
        print(f"Completion message: {e}")


if __name__ == "__main__":
    main()